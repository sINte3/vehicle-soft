"""
Yuklama Hisobot -- Engine Hours Workload Report
Generates web data and Excel export for equipment workload analysis.
Matches the approved sample: Yuklama_Hisoboti_Namuna_Aprel_2025.xlsx
"""

import os
from datetime import date
from sqlalchemy import text, func

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from models import db, Organization, Equipment


# ---------------------------------------------------------------------------
# Style constants (matching sample file)
# ---------------------------------------------------------------------------
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE   = Font(name='Times New Roman', size=14, bold=True)
FONT_HEADER  = Font(name='Times New Roman', size=11, bold=True)
FONT_DATA    = Font(name='Times New Roman', size=10)
FONT_NOTE    = Font(name='Times New Roman', size=9, italic=True)

FILL_YELLOW  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
FILL_RED     = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_R = Alignment(horizontal='right',  vertical='center')

PCT_FMT  = '0.0%'
NUM_FMT  = '#,##0.0'
INT_FMT  = '#,##0'


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def _get_mapped_equipment_ids():
    """Return set of equipment_ids that have a Wialon mapping entry."""
    try:
        rows = db.session.execute(
            text('SELECT DISTINCT equipment_id FROM vialon_mappings WHERE equipment_id IS NOT NULL')
        ).fetchall()
        return {r[0] for r in rows}
    except Exception:
        return None


def _get_hours_map(d_from, d_to):
    """Return {equipment_id: total_engine_hours} for the date range."""
    try:
        rows = db.session.execute(
            text('''
                SELECT equipment_id, SUM(engine_hours)
                FROM engine_hours_records
                WHERE work_date >= :df AND work_date <= :dt
                GROUP BY equipment_id
            '''),
            {'df': d_from, 'dt': d_to}
        ).fetchall()
        return {r[0]: float(r[1]) for r in rows if r[1] is not None}
    except Exception:
        return {}


def get_workload_data(d_from, d_to, org_ids=None):
    """
    Returns (orgs_data, calendar_days, norm_hours) where orgs_data is:
    [
        {
          'org': Organization,
          'rows': [
              {
                'eq': Equipment,
                'fact': float,       # moto-hours for period
                'norm': float,       # calendar_days * 8
                'exec_pct': float,   # fact / norm  (0..1)
                'work_days': int,    # round(fact / 8)
                'idle_days': int,    # calendar_days - work_days
                'calendar_days': int,
                'note': str,
              }
          ]
        }
    ]
    Only organisations that have at least one Wialon-mapped equipment are included.
    Within each org equipment is sorted by fact DESC.
    """
    calendar_days = (d_to - d_from).days + 1
    norm_hours    = calendar_days * 8

    mapped_ids = _get_mapped_equipment_ids()   # None means table missing
    hours_map  = _get_hours_map(d_from, d_to)

    org_q = Organization.query.order_by(Organization.sort_order)
    if org_ids:
        org_q = org_q.filter(Organization.id.in_(org_ids))
    orgs = org_q.all()

    result = []
    for org in orgs:
        eq_q = (Equipment.query
                .filter_by(organization_id=org.id, is_active=True)
                .order_by(Equipment.name))

        # Filter to Wialon-mapped equipment only (if mapping table exists)
        if mapped_ids is not None:
            equipment = [e for e in eq_q.all() if e.id in mapped_ids]
        else:
            # Fallback: show equipment that has any engine hours data ever
            eq_with_hours = set(hours_map.keys())
            equipment = [e for e in eq_q.all() if e.id in eq_with_hours]

        if not equipment:
            continue

        rows = []
        for eq in equipment:
            fact      = hours_map.get(eq.id, 0.0)
            norm      = float(norm_hours)
            exec_pct  = (fact / norm) if norm > 0 else 0.0
            work_days = round(fact / 8) if fact > 0 else 0
            idle_days = calendar_days - int(work_days)

            rows.append({
                'eq':           eq,
                'fact':         fact,
                'norm':         norm,
                'exec_pct':     exec_pct,
                'work_days':    int(work_days),
                'idle_days':    idle_days,
                'calendar_days': calendar_days,
                'note':         '',
            })

        # Sort by fact DESC
        rows.sort(key=lambda r: r['fact'], reverse=True)

        result.append({'org': org, 'rows': rows})

    return result, calendar_days, norm_hours


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------
def _sc(cell, font=None, fill=None, align=None, border=BORDER, num_fmt=None):
    """Style a cell."""
    if font:    cell.font      = font
    if fill:    cell.fill      = fill
    if align:   cell.alignment = align
    if border is not None: cell.border = border
    if num_fmt: cell.number_format = num_fmt


def _period_label(d_from, d_to):
    """Human-readable period label for subtitle."""
    if d_from == d_to:
        return d_from.strftime('%d.%m.%Y')
    return f'{d_from.strftime("%d.%m.%Y")} \u2013 {d_to.strftime("%d.%m.%Y")}'


def generate_workload_excel(d_from, d_to, output_dir, org_ids=None):
    """
    Generate Yuklama Hisobot Excel file.
    Returns absolute filepath of created file.
    """
    orgs_data, calendar_days, norm_hours = get_workload_data(d_from, d_to, org_ids)

    wb = Workbook()
    ws = wb.active

    # Sheet name
    if d_from.month == d_to.month and d_from.year == d_to.year:
        from calendar import month_name as _mn
        # Use Uzbek month names
        uz_months = [
            '', '\u042f\u043d\u0432\u0430\u0440', '\u0424\u0435\u0432\u0440\u0430\u043b',
            '\u041c\u0430\u0440\u0442', '\u0410\u043f\u0440\u0435\u043b', '\u041c\u0430\u0439',
            '\u0418\u044e\u043d', '\u0418\u044e\u043b', '\u0410\u0432\u0433\u0443\u0441\u0442',
            '\u0421\u0435\u043d\u0442\u044f\u0431\u0440', '\u041e\u043a\u0442\u044f\u0431\u0440',
            '\u041d\u043e\u044f\u0431\u0440', '\u0414\u0435\u043a\u0430\u0431\u0440',
        ]
        ws.title = '{} {}'.format(uz_months[d_from.month], d_from.year)
    else:
        ws.title = f'{d_from.strftime("%d.%m")}-{d_to.strftime("%d.%m.%Y")}'

    # ── Row 1: Title ───────────────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = (u'"\u0411\u0443\u0445\u043e\u0440\u043e \u0410\u0433\u0440\u043e\u043a\u043b\u0430\u0441\u0442\u0435\u0440" \u041c\u0427\u0416 '
               u'\u0442\u0438\u0437\u0438\u043c\u0438\u0434\u0430\u0433\u0438 \u0442\u0435\u0445\u043d\u0438\u043a\u0430\u043b\u0430\u0440\u043d\u0438\u043d\u0433 '
               u'\u0438\u0448 \u044e\u043a\u043b\u0430\u043c\u0430\u0441\u0438 \u0442\u045c\u0493\u0440\u0438\u0441\u0438\u0434\u0430 '
               u'\u041c\u0410\u042a\u041b\u0423\u041c\u041e\u0422')
    _sc(c, font=FONT_TITLE, align=ALIGN_C, border=None)
    ws.row_dimensions[1].height = 40

    # ── Row 2: Period subtitle ─────────────────────────────────────────────
    ws.merge_cells('A2:L2')
    period_str = _period_label(d_from, d_to)
    c = ws['A2']
    c.value = (u'\u0414\u0430\u0432\u0440: {}  |  '
               u'\u041a\u0430\u043b\u0435\u043d\u0434\u0430\u0440 \u043a\u0443\u043d\u043b\u0430\u0440: {}  |  '
               u'\u041d\u043e\u0440\u043c\u0430: {} \u043c\u043e\u0442\u043e\u0441\u043e\u0430\u0442').format(
               period_str, calendar_days, int(norm_hours))
    _sc(c, font=Font(name='Times New Roman', size=11), align=ALIGN_C, border=None)
    ws.row_dimensions[2].height = 22

    # ── Rows 3-4: Headers ──────────────────────────────────────────────────
    # Merged header cells (row 3 spans to row 4 unless noted)
    merges_r3 = [
        ('A3', 'A4'), ('B3', 'B4'), ('C3', 'C4'), ('D3', 'D4'),
        ('G3', 'G4'), ('J3', 'J4'), ('K3', 'K4'), ('L3', 'L4'),
    ]
    for top, bot in merges_r3:
        ws.merge_cells(f'{top}:{bot}')

    # Group headers
    ws.merge_cells('E3:F3')   # MOTOSOOTLAR
    ws.merge_cells('H3:I3')   # KUN TAHLILI

    headers_r3 = {
        'A3': u'\u2116',
        'B3': u'\u0422\u0430\u0448\u043a\u0438\u043b\u043e\u0442 \u043d\u043e\u043c\u0438',
        'C3': u'\u0422\u0435\u0445\u043d\u0438\u043a\u0430 \u0440\u0443\u0441\u0443\u043c\u0438',
        'D3': u'\u0414\u0430\u0432\u043b\u0430\u0442 \u0440\u0430\u049b\u0430\u043c\u0438',
        'E3': u'\u041c\u041e\u0422\u041e\u0421\u041e\u0410\u0422\u041b\u0410\u0420',
        'G3': u'\u0411\u0430\u0436\u0430\u0440\u0438\u0448,\n%',
        'H3': u'\u041a\u0423\u041d \u0422\u0410\u04b2\u041b\u0418\u041b\u0418',
        'J3': u'\u041d\u043e\u0440\u043c\u0430\n\u043a\u0443\u043d\u043b\u0430\u0440\u0438',
        'K3': u'\u042e\u043a\u043b\u0430\u043c\u0430,\n%',
        'L3': u'\u0418\u0437\u043e\u04b3',
    }
    headers_r4 = {
        'E4': u'\u041d\u043e\u0440\u043c\u0430\n(\u0440\u0435\u0436\u0430)',
        'F4': u'\u0424\u0430\u043a\u0442',
        'H4': u'\u0418\u0448\n\u043a\u0443\u043d\u043b\u0430\u0440\u0438',
        'I4': u'\u0418\u0448\u043b\u0430\u043c\u0430\u0433\u0430\u043d\n\u043a\u0443\u043d\u043b\u0430\u0440',
    }

    for addr, val in headers_r3.items():
        c = ws[addr]
        c.value = val
        _sc(c, font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_C)

    for addr, val in headers_r4.items():
        c = ws[addr]
        c.value = val
        _sc(c, font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_C)

    # Style empty merged cells in header area
    for r in (3, 4):
        for col in range(1, 13):
            c = ws.cell(row=r, column=col)
            if not c.value:
                _sc(c, fill=FILL_YELLOW, align=ALIGN_C, font=FONT_HEADER)

    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 32

    # ── Data rows ──────────────────────────────────────────────────────────
    row = 5
    global_num = 1

    for od in orgs_data:
        org = od['org']

        for r in od['rows']:
            eq         = r['eq']
            fact       = r['fact']
            norm       = r['norm']
            exec_pct   = r['exec_pct']   # 0..1
            work_days  = r['work_days']
            idle_days  = r['idle_days']
            cal_days   = r['calendar_days']
            note       = r.get('note', '')

            is_zero = (fact == 0)
            fill = FILL_RED if is_zero else None

            values = [
                (1,  global_num),
                (2,  org.name),
                (3,  eq.name),
                (4,  eq.plate or ''),
                (5,  norm),
                (6,  fact),
                (7,  exec_pct),
                (8,  work_days),
                (9,  idle_days),
                (10, cal_days),
                (11, exec_pct),
                (12, note),
            ]
            for col, val in values:
                c = ws.cell(row=row, column=col, value=val)
                align = ALIGN_L if col in (2, 3, 12) else ALIGN_C

                if col in (7, 11):       # percent columns
                    _sc(c, font=FONT_DATA, fill=fill, align=ALIGN_C,
                        border=BORDER, num_fmt=PCT_FMT)
                elif col in (5, 6):      # hours columns
                    _sc(c, font=FONT_DATA, fill=fill, align=ALIGN_C,
                        border=BORDER, num_fmt=NUM_FMT)
                elif col in (8, 9, 10):  # day columns (integers)
                    _sc(c, font=FONT_DATA, fill=fill, align=ALIGN_C,
                        border=BORDER, num_fmt=INT_FMT)
                else:
                    _sc(c, font=FONT_DATA, fill=fill, align=align, border=BORDER)

            ws.row_dimensions[row].height = 20
            global_num += 1
            row += 1

        # Empty separator row between organisations
        for col in range(1, 13):
            ws.cell(row=row, column=col).value = None
        ws.row_dimensions[row].height = 6
        row += 1

    # ── Notes section ─────────────────────────────────────────────────────
    row += 1  # blank line
    note_lines = [
        (u'\u0418\u0437\u043e\u04b3 (\u0444\u043e\u0440\u043c\u0443\u043b\u0430\u043b\u0430\u0440):', True),
        (u'* \u041d\u043e\u0440\u043c\u0430 (\u0440\u0435\u0436\u0430) = \u041a\u0430\u043b\u0435\u043d\u0434\u0430\u0440 \u043a\u0443\u043d\u043b\u0430\u0440 \u00d7 8 \u043c\u043e\u0442\u043e\u0441\u043e\u0430\u0442  ({} \u00d7 8 = {})'.format(calendar_days, int(norm_hours)), False),
        (u'* \u0411\u0430\u0436\u0430\u0440\u0438\u0448 % = \u0424\u0430\u043a\u0442 \u00f7 \u041d\u043e\u0440\u043c\u0430 \u00d7 100%', False),
        (u'* \u0418\u0448 \u043a\u0443\u043d\u043b\u0430\u0440\u0438 (\u04b3\u0438\u0441\u043e\u0431\u0438\u0439) = \u0424\u0430\u043a\u0442 \u043c\u043e\u0442\u043e\u0441\u043e\u0430\u0442 \u00f7 8  (\u044f\u0445\u043b\u0438\u0442\u043b\u0430\u0448)', False),
        (u'* \u0418\u0448\u043b\u0430\u043c\u0430\u0433\u0430\u043d \u043a\u0443\u043d\u043b\u0430\u0440 = {} \u2212 \u0418\u0448 \u043a\u0443\u043d\u043b\u0430\u0440\u0438 (\u04b3\u0438\u0441\u043e\u0431\u0438\u0439)'.format(calendar_days), False),
        (u'* \u0422\u0435\u0445\u043d\u0438\u043a\u0430 \u0442\u0430\u0440\u0442\u0438\u0431\u0438: \u0444\u0430\u043a\u0442\u0438\u043a \u043c\u043e\u0442\u043e\u0441\u043e\u0430\u0442 \u0431\u045c\u0439\u0438\u0447\u0430 \u043a\u0430\u043c\u0430\u0439\u0438\u0448 \u0442\u0430\u0440\u0442\u0438\u0431\u0438\u0434\u0430 (\u043a\u045c\u043f\u0434\u0430\u043d \u043e\u0437\u0433\u0430)', False),
        (u'', False),
        (u'\u049a\u0438\u0437\u0438\u043b \u0444\u043e\u043d:  \u0412\u0438\u0430\u043b\u043e\u043d \u043c\u0430\u044a\u043b\u0443\u043c\u043e\u0442\u0438 \u0439\u045c\u049b (0 \u043c\u043e\u0442\u043e\u0441\u043e\u0430\u0442) \u2014 \u0442\u0435\u043a\u0448\u0438\u0440\u0438\u0448 \u043a\u0435\u0440\u0430\u043a', False),
    ]
    for text_val, is_header in note_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        c = ws.cell(row=row, column=1, value=text_val)
        _sc(c, font=Font(name='Times New Roman', size=9, bold=is_header),
            align=ALIGN_L, border=None)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Column widths (matching sample) ────────────────────────────────────
    col_widths = {
        'A': 5, 'B': 22, 'C': 22, 'D': 14,
        'E': 14, 'F': 12, 'G': 12,
        'H': 13, 'I': 14, 'J': 12, 'K': 11, 'L': 18,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Page setup ─────────────────────────────────────────────────────────
    ws.page_setup.orientation  = 'landscape'
    ws.page_setup.paperSize    = 9   # A4
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '3:4'
    ws.print_area = f'A1:L{row - 1}'

    # ── Save ───────────────────────────────────────────────────────────────
    if d_from == d_to:
        fname = f'Yuklama_{d_from.strftime("%d_%m_%Y")}.xlsx'
    else:
        fname = f'Yuklama_{d_from.strftime("%d_%m_%Y")}_{d_to.strftime("%d_%m_%Y")}.xlsx'

    os.makedirs(output_dir, exist_ok=True)
    fpath = os.path.join(output_dir, fname)
    wb.save(fpath)
    return fpath
