# -*- coding: utf-8 -*-
"""
excel_daily_activity.py - Kunlik ish bilan band hisoboti.
Two-table daily activity report on one sheet.
Table 1: 8 agriculture categories (27 cols A-AA).
Table 2: yuk_transport by brand (cols A-K).
"""

import os
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import db, Organization, Equipment, DailyRecord
from excel_export import (
    FONT_TITLE, FONT_HEADER, FONT_HEADER_SM, FONT_DATA,
    FONT_TOTAL, FILL_YELLOW, FILL_GREEN, FILL_BLUE,
    ALIGN_CENTER, ALIGN_LEFT, BORDER, set_print, style_cell, NUM_FMT,
    translate_workbook_to_ru, polish_report_workbook,
)

AGR_CATS = ['yukori', 'mtz', 'qatnov', 'mini', 'combine', 'special', 'motorcycle', 'passenger']

AGR_LABELS = [
    u'Юқори\nунумли',
    u'Чопиқ\nтрактор',
    u'Қатнов\nтрактор',
    u'Мини\nтрактор',
    u'Комбайн',
    u'Махсус\nтехника',
    u'Мотоцикл',
    u'Йўловчи\nташиш',
]

YUK_BRANDS = ['HOWO', 'MAN', 'Kamaz', 'Hyundai', 'Isuzu', 'Zoomlion']


def load_activity_data(report_date, org_ids=None):
    """
    Returns (orgs_list, dict keyed by org_id).
    Each value: org, agr_total, agr_working, yuk_total, yuk_total_all, yuk_working.
    Only orgs with at least one piece of equipment are included in the dict.
    """
    org_q = Organization.query.order_by(Organization.sort_order)
    if org_ids:
        org_q = org_q.filter(Organization.id.in_(org_ids))
    orgs = org_q.all()

    result = {}
    for org in orgs:
        agr_total = {}
        for cat in AGR_CATS:
            agr_total[cat] = Equipment.query.filter_by(
                organization_id=org.id, category=cat, is_active=True).count()

        agr_working = {cat: 0 for cat in AGR_CATS}
        working_recs = (DailyRecord.query
                        .join(Equipment)
                        .filter(Equipment.organization_id == org.id,
                                Equipment.category.in_(AGR_CATS),
                                DailyRecord.work_date == report_date,
                                DailyRecord.status == 'working')
                        .with_entities(Equipment.category,
                                       func.count(func.distinct(Equipment.id)))
                        .group_by(Equipment.category)
                        .all())
        for cat, cnt in working_recs:
            agr_working[cat] = cnt

        yuk_eqs = Equipment.query.filter_by(
            organization_id=org.id, category='yuk_transport', is_active=True).all()
        yuk_total_all = len(yuk_eqs)
        yuk_total = {brand: 0 for brand in YUK_BRANDS}
        for eq in yuk_eqs:
            for brand in YUK_BRANDS:
                # [REASON]: eq_type stores the brand/model string; case-insensitive match
                if brand.lower() in (eq.eq_type or '').lower():
                    yuk_total[brand] += 1
                    break

        yuk_working_cnt = (DailyRecord.query
                           .join(Equipment)
                           .filter(Equipment.organization_id == org.id,
                                   Equipment.category == 'yuk_transport',
                                   DailyRecord.work_date == report_date,
                                   DailyRecord.status == 'working')
                           .with_entities(func.count(func.distinct(Equipment.id)))
                           .scalar() or 0)

        if sum(agr_total.values()) > 0 or yuk_total_all > 0:
            result[org.id] = {
                'org': org,
                'agr_total': agr_total,
                'agr_working': agr_working,
                'yuk_total': yuk_total,
                'yuk_total_all': yuk_total_all,
                'yuk_working': yuk_working_cnt,
            }

    return orgs, result


def _sc(ws, row, col, value=None, font=None, fill=None, align=ALIGN_CENTER, border=BORDER):
    """Write value and apply style to a single cell."""
    cell = ws.cell(row, col)
    if value is not None:
        cell.value = value
    style_cell(cell, font=font, fill=fill, align=align, border=border)
    return cell


def generate_daily_activity(report_date, output_dir, org_ids=None, lang='uz'):
    """Generate two-table daily activity Excel report. Returns saved file path."""
    orgs, data = load_activity_data(report_date, org_ids)
    date_str = report_date.strftime('%d.%m.%Y')

    wb = Workbook()
    ws = wb.active
    ws.title = report_date.strftime('%d.%m.%Y')

    set_print(ws)

    # Column widths: B=22, C-K=12 (satisfies both Table 1 and Table 2 needs), L-AA=9
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    for col_idx in range(3, 12):   # C through K
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    for col_idx in range(12, 28):  # L through AA
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    # ═══════════════════════════════════════════════════════════════
    # TABLE 1: Agricultural machinery — 27 columns (A through AA)
    # ═══════════════════════════════════════════════════════════════

    # Row 1 — title (no border per spec)
    ws.merge_cells('A1:AA1')
    c = ws['A1']
    c.value = (u'"Бухоро Агрокластер" '
               u'МЧЖ таркибидаги корхоналарда '
               u'мавжуд қишлоқ хўжалиги '
               u'техникалари кунлик '
               u'харакати бўйича МАЪЛУМОТ')
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 65

    # Row 2 — date (right side of sheet)
    ws.merge_cells('T2:AA2')
    c = ws['T2']
    c.value = u'Сана: ' + date_str
    style_cell(c, font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)
    ws.row_dimensions[2].height = 20

    # Rows 3-4 — group and category headers
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 45

    ws.merge_cells('A3:A4')
    style_cell(ws['A3'], font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)
    ws['A3'] = '№'

    ws.merge_cells('B3:B4')
    ws['B3'] = u'Корхона номи'
    style_cell(ws['B3'], font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('C3:C4')
    ws['C3'] = u'Жами\nқиш.хўж-ги\nтехника'
    style_cell(ws['C3'], font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('D3:K3')
    ws['D3'] = u'ШУНДАН'
    style_cell(ws['D3'], font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('L3:S3')
    ws['L3'] = u'ИШ БИЛАН БАНД'
    style_cell(ws['L3'], font=FONT_HEADER, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('T3:AA3')
    ws['T3'] = u'ИШГА ЧИҚМАГАН'
    style_cell(ws['T3'], font=FONT_HEADER, fill=FILL_BLUE, align=ALIGN_CENTER, border=BORDER)

    # Row 4 sub-headers: 3 groups × 8 categories
    for i, label in enumerate(AGR_LABELS):
        # SHUNDAN group cols D-K (4-11)
        c = ws.cell(4, 4 + i, label)
        style_cell(c, font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)
        # ISH BILAN BAND group cols L-S (12-19)
        c = ws.cell(4, 12 + i, label)
        style_cell(c, font=FONT_HEADER_SM, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)
        # ISHGA CHIQMAGAN group cols T-AA (20-27)
        c = ws.cell(4, 20 + i, label)
        style_cell(c, font=FONT_HEADER_SM, fill=FILL_BLUE, align=ALIGN_CENTER, border=BORDER)

    # A4, B4, C4 are covered by row-3 merges; style to avoid blank border artifacts
    for col in [1, 2, 3]:
        style_cell(ws.cell(4, col), font=FONT_HEADER_SM, fill=FILL_YELLOW,
                   align=ALIGN_CENTER, border=BORDER)

    # Data rows
    row = 5
    num = 1
    grand_total = 0
    grand_agr_total = {cat: 0 for cat in AGR_CATS}
    grand_agr_working = {cat: 0 for cat in AGR_CATS}
    grand_agr_idle = {cat: 0 for cat in AGR_CATS}

    for org in orgs:
        if org.id not in data:
            continue
        d = data[org.id]
        total_agr = sum(d['agr_total'][cat] for cat in AGR_CATS)
        if total_agr == 0:
            continue

        ws.row_dimensions[row].height = 22
        ws.cell(row, 1, num)
        style_cell(ws.cell(row, 1), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)

        ws.cell(row, 2, d['org'].name)
        style_cell(ws.cell(row, 2), font=FONT_DATA, align=ALIGN_LEFT, border=BORDER)

        ws.cell(row, 3, total_agr or None)
        style_cell(ws.cell(row, 3), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)

        grand_total += total_agr

        for i, cat in enumerate(AGR_CATS):
            tot = d['agr_total'][cat]
            wrk = d['agr_working'][cat]
            # [REASON]: idle computed in Python, not formula, for reliable Excel rendering
            idl = tot - wrk

            ws.cell(row, 4 + i, tot or None)
            style_cell(ws.cell(row, 4 + i), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)

            ws.cell(row, 12 + i, wrk or None)
            style_cell(ws.cell(row, 12 + i), font=FONT_DATA, fill=FILL_GREEN,
                       align=ALIGN_CENTER, border=BORDER)

            ws.cell(row, 20 + i, idl or None)
            style_cell(ws.cell(row, 20 + i), font=FONT_DATA, fill=FILL_BLUE,
                       align=ALIGN_CENTER, border=BORDER)

            grand_agr_total[cat] += tot
            grand_agr_working[cat] += wrk
            grand_agr_idle[cat] += idl

        num += 1
        row += 1

    # Totals row
    ws.row_dimensions[row].height = 28
    ws.merge_cells('A{}:B{}'.format(row, row))
    ws.cell(row, 1, u'Жами :')
    style_cell(ws.cell(row, 1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT, border=BORDER)

    ws.cell(row, 3, grand_total or None)
    style_cell(ws.cell(row, 3), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    total_working_agr = 0
    total_idle_agr = 0
    for i, cat in enumerate(AGR_CATS):
        ws.cell(row, 4 + i, grand_agr_total[cat] or None)
        style_cell(ws.cell(row, 4 + i), font=FONT_TOTAL, fill=FILL_GREEN,
                   align=ALIGN_CENTER, border=BORDER)

        ws.cell(row, 12 + i, grand_agr_working[cat] or None)
        style_cell(ws.cell(row, 12 + i), font=FONT_TOTAL, fill=FILL_GREEN,
                   align=ALIGN_CENTER, border=BORDER)

        ws.cell(row, 20 + i, grand_agr_idle[cat] or None)
        style_cell(ws.cell(row, 20 + i), font=FONT_TOTAL, fill=FILL_GREEN,
                   align=ALIGN_CENTER, border=BORDER)

        total_working_agr += grand_agr_working[cat]
        total_idle_agr += grand_agr_idle[cat]

    row += 1

    # Summary row: total working and total idle counts
    ws.row_dimensions[row].height = 22
    ws.merge_cells('A{}:L{}'.format(row, row))
    ws.cell(row, 1, u'ИШ БИЛАН БАНД: {}'.format(total_working_agr))
    style_cell(ws.cell(row, 1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('M{}:AA{}'.format(row, row))
    ws.cell(row, 13, u'ИШГА ЧИҚМАГАН: {}'.format(total_idle_agr))
    style_cell(ws.cell(row, 13), font=FONT_TOTAL, fill=FILL_BLUE, align=ALIGN_CENTER, border=BORDER)

    row += 5  # gap between tables

    # ═══════════════════════════════════════════════════════════════
    # TABLE 2: Cargo transport — yuk_transport by brand (cols A-K)
    # ═══════════════════════════════════════════════════════════════

    ws.merge_cells('A{}:AA{}'.format(row, row))
    c = ws.cell(row, 1)
    c.value = (u'"Бухоро Агрокластер" '
               u'МЧЖ таркибидаги корхоналарда '
               u'мавжуд юк транспортлари кунлик '
               u'харакати бўйича МАЪЛУМОТ')
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 65
    row += 1

    ws.merge_cells('T{}:AA{}'.format(row, row))
    c = ws.cell(row, 20)  # col T = 20
    c.value = u'Сана: ' + date_str
    style_cell(c, font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)
    ws.row_dimensions[row].height = 20
    row += 1

    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 45

    ws.merge_cells('A{}:A{}'.format(row, row + 1))
    ws.cell(row, 1, '№')
    style_cell(ws.cell(row, 1), font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('B{}:B{}'.format(row, row + 1))
    ws.cell(row, 2, u'Корхона номи')
    style_cell(ws.cell(row, 2), font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('C{}:C{}'.format(row, row + 1))
    ws.cell(row, 3, u'Жами\nтехника')
    style_cell(ws.cell(row, 3), font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER, border=BORDER)

    for i, brand in enumerate(YUK_BRANDS):
        col = 4 + i
        ws.merge_cells('{r}{s}:{r}{e}'.format(
            r=get_column_letter(col), s=row, e=row + 1))
        ws.cell(row, col, brand)
        style_cell(ws.cell(row, col), font=FONT_HEADER_SM, fill=FILL_YELLOW,
                   align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('J{}:J{}'.format(row, row + 1))
    ws.cell(row, 10, u'ИШ БИЛАН\nБАНД')
    style_cell(ws.cell(row, 10), font=FONT_HEADER, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    ws.merge_cells('K{}:K{}'.format(row, row + 1))
    ws.cell(row, 11, u'ИШГА\nЧИҚМАГАН')
    style_cell(ws.cell(row, 11), font=FONT_HEADER, fill=FILL_BLUE, align=ALIGN_CENTER, border=BORDER)

    # Style second header row (covered by merges, but border needs applying)
    for col in [1, 2, 3]:
        style_cell(ws.cell(row + 1, col), font=FONT_HEADER_SM, fill=FILL_YELLOW,
                   align=ALIGN_CENTER, border=BORDER)
    for i in range(len(YUK_BRANDS)):
        style_cell(ws.cell(row + 1, 4 + i), font=FONT_HEADER_SM, fill=FILL_YELLOW,
                   align=ALIGN_CENTER, border=BORDER)
    style_cell(ws.cell(row + 1, 10), font=FONT_HEADER_SM, fill=FILL_GREEN,
               align=ALIGN_CENTER, border=BORDER)
    style_cell(ws.cell(row + 1, 11), font=FONT_HEADER_SM, fill=FILL_BLUE,
               align=ALIGN_CENTER, border=BORDER)

    row += 2

    num = 1
    grand_yuk_total = 0
    grand_yuk_brands = {brand: 0 for brand in YUK_BRANDS}
    grand_yuk_working = 0

    for org in orgs:
        if org.id not in data:
            continue
        d = data[org.id]
        if d['yuk_total_all'] == 0:
            continue

        ws.row_dimensions[row].height = 22
        ws.cell(row, 1, num)
        style_cell(ws.cell(row, 1), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)

        ws.cell(row, 2, d['org'].name)
        style_cell(ws.cell(row, 2), font=FONT_DATA, align=ALIGN_LEFT, border=BORDER)

        ws.cell(row, 3, d['yuk_total_all'] or None)
        style_cell(ws.cell(row, 3), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)

        for i, brand in enumerate(YUK_BRANDS):
            val = d['yuk_total'][brand]
            ws.cell(row, 4 + i, val or None)
            style_cell(ws.cell(row, 4 + i), font=FONT_DATA, align=ALIGN_CENTER, border=BORDER)
            grand_yuk_brands[brand] += val

        ws.cell(row, 10, d['yuk_working'] or None)
        style_cell(ws.cell(row, 10), font=FONT_DATA, fill=FILL_GREEN,
                   align=ALIGN_CENTER, border=BORDER)

        idle_yuk = d['yuk_total_all'] - d['yuk_working']
        ws.cell(row, 11, idle_yuk or None)
        style_cell(ws.cell(row, 11), font=FONT_DATA, fill=FILL_BLUE,
                   align=ALIGN_CENTER, border=BORDER)

        grand_yuk_total += d['yuk_total_all']
        grand_yuk_working += d['yuk_working']
        num += 1
        row += 1

    # Totals row Table 2
    ws.row_dimensions[row].height = 28
    ws.merge_cells('A{}:B{}'.format(row, row))
    ws.cell(row, 1, u'Жами :')
    style_cell(ws.cell(row, 1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT, border=BORDER)

    ws.cell(row, 3, grand_yuk_total or None)
    style_cell(ws.cell(row, 3), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    for i, brand in enumerate(YUK_BRANDS):
        ws.cell(row, 4 + i, grand_yuk_brands[brand] or None)
        style_cell(ws.cell(row, 4 + i), font=FONT_TOTAL, fill=FILL_GREEN,
                   align=ALIGN_CENTER, border=BORDER)

    ws.cell(row, 10, grand_yuk_working or None)
    style_cell(ws.cell(row, 10), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER, border=BORDER)

    grand_yuk_idle = grand_yuk_total - grand_yuk_working
    ws.cell(row, 11, grand_yuk_idle or None)
    style_cell(ws.cell(row, 11), font=FONT_TOTAL, fill=FILL_BLUE, align=ALIGN_CENTER, border=BORDER)

    ws.print_area = 'A1:AA{}'.format(row)

    if lang == 'ru':
        translate_workbook_to_ru(wb)
    polish_report_workbook(wb)

    prefix = 'Dnevnaya_zanyatost' if lang == 'ru' else 'Kunlik'
    fname = '{}_{}.xlsx'.format(prefix, report_date.strftime('%d_%m_%Y'))
    fpath = os.path.join(output_dir, fname)
    wb.save(fpath)
    return fpath
