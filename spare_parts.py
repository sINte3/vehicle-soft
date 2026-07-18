import bisect
import io
import math
import os
import re
import uuid
from difflib import SequenceMatcher

from flask import (Blueprint, render_template, request, redirect, url_for, flash,
                   abort, g, jsonify, current_app, send_from_directory, send_file)
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from sqlalchemy import and_, exists, func, update
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import joinedload, selectinload

from models import (db, SparePart, SparePartCategory, SparePartPriceAudit,
                    SparePartAttachment, SparePartInventory,
                    SparePartInventoryMovement, SparePartRequest,
                    SparePartRequestItem, SparePartReservation,
                    SparePartSku, SparePartStatusHistory,
                    SparePartUnit, SparePartWarehouse, SparePartWriteOffAct,
                    SparePartWriteOffActItem, Organization, Equipment,
                    EquipmentModel, SparePartCompatibility,
                    SparePartMaintenanceNorm, EngineHoursRecord, User,
                    module_required)
from sec003a_ext import log_audit, diff_dict

# [REASON]: BOT003 — Best-effort Telegram notifications. Import is guarded so that
# missing module does not crash spare parts. The actual enqueue functions catch all
# exceptions internally and never raise.
try:
    from bot003_notifications import (
        enqueue_spare_request_submitted_best_effort,
        enqueue_spare_request_status_best_effort,
    )
    _BOT003_AVAILABLE = True
except ImportError:
    _BOT003_AVAILABLE = False

spare_parts_bp = Blueprint('spare_parts', __name__, url_prefix='/spare-parts')

STATUS_LABELS = {
    'draft':     {'uz': 'Чорнов',       'ru': 'Черновик'},
    'submitted': {'uz': 'Юборилган',    'ru': 'Отправлено'},
    # [REASON]: SPARE-STAGE1 — application-level status for the price
    # reject/return workflow; no schema change (status column is free VARCHAR).
    # Flow: draft -> submitted -> (returned_for_revision -> submitted) -> approved | rejected
    'returned_for_revision': {'uz': 'Қайта ишлашга қайтарилган', 'ru': 'На доработке'},
    'approved':  {'uz': 'Тасдиқланган', 'ru': 'Утверждено'},
    'rejected':  {'uz': 'Рад этилган',  'ru': 'Отклонено'},
    # [REASON]: SPARE-STAGE2 — new terminal state, reachable ONLY from
    # 'approved' (issue_request enforces the transition): the approved goods
    # were physically handed over and a write-off act was generated.
    'issued':    {'uz': 'Берилган',     'ru': 'Выдано'},
}
STATUS_COLORS = {
    'draft':     'var(--text2)',
    'submitted': 'var(--info)',
    'returned_for_revision': 'var(--warn)',
    'approved':  'var(--accent)',
    'rejected':  'var(--danger)',
    'issued':    'var(--success)',
}

PRICE_STATUS_LABELS = {
    'pending':   {'uz': 'Нарх кутилмоқда',      'ru': 'Цена не подтверждена'},
    'confirmed': {'uz': 'Нарх тасдиқланган',    'ru': 'Цена подтверждена'},
    'rejected':  {'uz': 'Нарх рад этилган',     'ru': 'Цена отклонена'},
    'returned':  {'uz': 'Нарх қайтарилган',     'ru': 'Цена возвращена'},
}
PRICE_STATUS_COLORS = {
    'pending':   'var(--text2)',
    'confirmed': 'var(--accent)',
    'rejected':  'var(--danger)',
    'returned':  'var(--warn)',
}

# ─── SPARE-STAGE2: warning rule 3/4/6 thresholds ─────────────────────────────
# [REASON]: SPARE-STAGE2 — rules 3/4/6 are warn-only (never block), same
# principle as the SPARE-STAGE1 repeat engine (rules 1-2). The values below
# are tunable starting points agreed with the project owner, not validated
# business constants.

# Tunable starting value: rule 3 warns when the same equipment + same catalog
# part appears on MORE THAN this many approved requests in the trailing window.
RULE3_MAX_OCCURRENCES_90D = 3
# Tunable starting value: rule 3 trailing lookback window in days (mirrors the
# 90-day window already used by the rules 1-2 repeat engine).
RULE3_WINDOW_DAYS = 90
# Tunable starting value: rule 4 warns when current-month approved+confirmed
# spend exceeds the trailing-months average by more than this multiplier.
RULE4_COST_ANOMALY_MULTIPLIER = 2.0
# Tunable starting value: rule 4 baseline length in calendar months (the
# months immediately before the current one).
RULE4_BASELINE_MONTHS = 3
# Tunable starting value: rule 6 warns when the same equipment + part was on a
# request rejected within this many trailing days.
RULE6_REJECTED_LOOKBACK_DAYS = 30

# ─── SPARE-STAGE2: fuzzy catalog search ──────────────────────────────────────
# Tunable starting value: minimum difflib ratio for a fuzzy-only match
# (same default cutoff difflib.get_close_matches uses).
FUZZY_MATCH_MIN_RATIO = 0.6
# Tunable starting value: maximum results returned by api_catalog_search.
FUZZY_SEARCH_MAX_RESULTS = 10

# ─── SPARE-STAGE2: inventory movement types ──────────────────────────────────
INV_MOVEMENT_TYPES = ('receipt', 'issue', 'adjustment', 'write_off')
INV_MOVEMENT_LABELS = {
    'receipt':    {'uz': 'Кирим',         'ru': 'Приход'},
    'issue':      {'uz': 'Бериш',         'ru': 'Выдача'},
    'adjustment': {'uz': 'Тузатиш',       'ru': 'Корректировка'},
    'write_off':  {'uz': 'Ҳисобдан чиқариш', 'ru': 'Списание'},
}

# Photo/video upload rules (SPARE-STAGE1 + MOBILE-UPLOAD-001). Extension AND
# content signature are both checked -- Content-Type alone is never trusted.
# [REASON]: MOBILE-UPLOAD-001 — field mechanics use phone cameras; iPhones
# default to HEIC, and Android/action-cam footage is commonly MP4/WEBM/
# MKV/AVI, so JPG/PNG-only was too narrow for real field use.
ALLOWED_PHOTO_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.heic', '.heif'}
ALLOWED_VIDEO_EXTENSIONS = {'.mp4', '.mov', '.webm', '.avi', '.mkv'}
ALLOWED_ATTACHMENT_EXTENSIONS = ALLOWED_PHOTO_EXTENSIONS | ALLOWED_VIDEO_EXTENSIONS

# [REASON]: No Pillow/ffmpeg in requirements.txt (no new dependencies without
# approval), so file content is verified with minimal magic-byte checks per
# container format instead (see _validate_photo). ISO-BMFF (`ftyp` at offset
# 4) covers MP4/MOV/HEIC/HEIF -- deliberately NOT distinguishing exact brand
# codes between them (real complexity for negligible anti-spoofing benefit
# here); a valid `ftyp` box is sufficient proof the file is a real media
# container and not, e.g., a renamed .txt/.html (the actual failure mode
# this project hit once in QA, caught by the JPG/PNG check -- the new
# formats keep that same bar).

MAX_ATTACHMENT_SIZE_MB = 50
MAX_ATTACHMENT_SIZE = MAX_ATTACHMENT_SIZE_MB * 1024 * 1024
# [REASON]: MOBILE-UPLOAD-001 — photos and videos combined, per request line
# item (both the create flow and the detail-page add-more flow enforce this
# against the same running count).
MAX_ATTACHMENTS_PER_ITEM = 5

# [REASON]: MOBILE-UPLOAD-001 — attachment_file() no longer relies on
# mimetypes.guess_type() (unreliable for WEBP/HEIC across Windows Python
# installs); Content-Type is looked up explicitly per stored extension.
ATTACHMENT_CONTENT_TYPES = {
    '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
    '.png': 'image/png',
    '.webp': 'image/webp',
    '.heic': 'image/heic', '.heif': 'image/heif',
    '.mp4': 'video/mp4',
    '.mov': 'video/quicktime',
    '.webm': 'video/webm',
    '.avi': 'video/x-msvideo',
    '.mkv': 'video/x-matroska',
}


def _spare_lang():
    lang = getattr(g, 'lang', None)
    if not lang and getattr(current_user, 'is_authenticated', False):
        lang = getattr(current_user, 'language', None)
    return 'ru' if lang == 'ru' else 'uz'


def _spare_t(uz_text, ru_text):
    return ru_text if _spare_lang() == 'ru' else uz_text


def _spare_format_errors(errors, title_uz=None, title_ru=None):
    unique = []
    seen = set()
    for err in errors or []:
        text = str(err).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        unique.append(text)
    if not unique:
        return ''
    title = _spare_t(title_uz or 'Заявка сақланмади. Хатоларни тузатинг:',
                     title_ru or 'Заявка не сохранена. Исправьте ошибки:')
    return title + '\n' + '\n'.join('- ' + item for item in unique)


def _spare_flash_errors(errors, title_uz=None, title_ru=None):
    msg = _spare_format_errors(errors, title_uz, title_ru)
    if msg:
        flash(msg, 'warning')



def _add_status_history(request_id, old_status, new_status, comment="", changed_by=None):
    """Insert a SparePartStatusHistory row for a status transition.

    Called before db.session.commit() in status-changing routes.
    Does not flush or commit -- caller is responsible for the final commit.
    """
    history = SparePartStatusHistory(
        request_id=request_id,
        old_status=old_status,
        new_status=new_status,
        comment=comment or "",
        changed_by=changed_by,
    )
    db.session.add(history)


def _date_iso(value):
    if not value:
        return None
    try:
        return value.isoformat()
    except Exception:
        return str(value)


def _spare_user_org_ids():
    return current_user.get_org_ids() if not current_user.is_admin else None


def _spare_check_org_access(org_id):
    if not org_id or not current_user.can_access_org(org_id):
        abort(403)


def _spare_check_request_access(req):
    if not current_user.is_admin and not current_user.can_access_org(req.organization_id):
        abort(403)


def _parse_spare_date(value):
    try:
        if not value:
            raise ValueError()
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        raise ValueError('request_date')


def _active_units():
    """Active unit-directory rows for pickers, ordered like other references.

    [REASON]: SP-F-024 — strict managed directory (owner decision 2026-07-14).
    An EMPTY result (fresh install before migrate_spare_parts_units.py runs)
    makes every caller fall back to the legacy free-text behavior, so the app
    keeps working pre-migration and in dev environments.
    """
    try:
        return (SparePartUnit.query.filter_by(is_active=True)
                .order_by(SparePartUnit.sort_order, SparePartUnit.id).all())
    except OperationalError as exc:
        # [REASON]: RE-SP-009 — deploy-order safety net, NOT normal operation:
        # if code that queries spare_part_units runs before
        # migrate_spare_parts_units.py (misordered deploy), or a rollback
        # drops the table before the code stops querying it, the request
        # form/catalog routes must degrade to the same legacy free-text
        # fallback an empty directory already takes instead of 500-ing.
        # Deliberately narrow: ONLY the SQLite "no such table" condition is
        # swallowed; any other database error still surfaces.
        if 'no such table' in str(exc.orig or exc).lower():
            db.session.rollback()
            current_app.logger.warning(
                'spare_part_units table missing — falling back to free-text '
                'units (deploy/rollback transition window?)')
            return []
        raise


def _unit_labels_map(lang=None):
    """Unit code -> localized display label ({'dona': 'шт', ...}) for `lang`
    (defaults to the current UI language). Cached per request+language in g.

    [REASON]: RE-SP-010 — output documents (PDF act, Excel report, detail
    tables) showed the raw stored code (e.g. `dona`) instead of the
    directory's localized word. This map covers ALL directory rows, active or
    not: a deactivated unit must still render readably on historical rows.
    Reuses the RE-SP-009 "no such table" deploy-window guard so rendering
    degrades to raw codes instead of 500-ing pre-migration.
    """
    lang = lang or _spare_lang()
    cache = getattr(g, '_spare_unit_labels', None)
    if cache is None:
        cache = {}
        g._spare_unit_labels = cache
    if lang in cache:
        return cache[lang]
    try:
        rows = SparePartUnit.query.all()
    except OperationalError as exc:
        if 'no such table' in str(exc.orig or exc).lower():
            db.session.rollback()
            rows = []
        else:
            raise
    labels = {u.code: (u.name_ru if lang == 'ru' else u.name_uz) for u in rows}
    cache[lang] = labels
    return labels


def _unit_label(code, lang=None):
    """Localized display word for a stored unit code.

    [REASON]: RE-SP-010 — display-time translation ONLY (stored values are
    never rewritten). An unknown/legacy code (free-text era value, or an
    empty directory pre-migration) falls back to the raw code unchanged —
    never raises, never blanks a value out.
    """
    code = (code or '').strip()
    if not code:
        return code
    return _unit_labels_map(lang).get(code) or code


@spare_parts_bp.app_template_filter('spare_unit_label')
def _spare_unit_label_filter(code):
    # Template-side entry point for _unit_label (detail/act/catalog tables).
    return _unit_label(code)


def _unit_options(unit_rows):
    """JSON-ready picker options: code + language-correct label + lowercase
    alt spellings (code|name_ru|name_uz) for matching legacy catalog values
    client-side."""
    lang = _spare_lang()
    return [{
        'code': u.code,
        'label': (u.name_ru if lang == 'ru' else u.name_uz),
        'alt': '|'.join(x.lower() for x in (u.code, u.name_ru, u.name_uz)),
    } for u in unit_rows]


def _parse_spare_positive_qty(value):
    try:
        result = float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        raise ValueError('quantity')
    # [REASON]: SP-F-016 — float() happily parses 'nan'/'inf', and NaN slips
    # past every comparison (NaN <= 0 is False), so finiteness is an explicit
    # extra condition on the existing check, not a new error category.
    if not math.isfinite(result) or result <= 0:
        raise ValueError('quantity')
    return result


def _request_snapshot(req):
    if not req:
        return None
    return {
        'id': getattr(req, 'id', None),
        'request_date': _date_iso(getattr(req, 'request_date', None)),
        'organization_id': getattr(req, 'organization_id', None),
        'equipment_id': getattr(req, 'equipment_id', None),
        'status': getattr(req, 'status', ''),
        'note': getattr(req, 'note', ''),
        'created_by': getattr(req, 'created_by', None),
        'reviewed_by': getattr(req, 'reviewed_by', None),
        'reviewed_at': _date_iso(getattr(req, 'reviewed_at', None)),
        'review_comment': getattr(req, 'review_comment', ''),
        'items_count': len(getattr(req, 'items', []) or []),
    }


def _item_snapshot(item):
    if not item:
        return None
    return {
        'id': getattr(item, 'id', None),
        'request_id': getattr(item, 'request_id', None),
        'spare_part_id': getattr(item, 'spare_part_id', None),
        'sku_id': getattr(item, 'sku_id', None),
        'name': getattr(item, 'name', ''),
        'part_number': getattr(item, 'part_number', ''),
        'quantity': getattr(item, 'quantity', None),
        'unit': getattr(item, 'unit', ''),
        'note': getattr(item, 'note', ''),
        'price': getattr(item, 'price', None),
        'price_status': getattr(item, 'price_status', ''),
    }


def _catalog_snapshot(part):
    if not part:
        return None
    return {
        'id': getattr(part, 'id', None),
        'name': getattr(part, 'name', ''),
        'name_uz': getattr(part, 'name_uz', None),
        'part_number': getattr(part, 'part_number', ''),
        'unit': getattr(part, 'unit', ''),
        'category': getattr(part, 'category', ''),
        'category_id': getattr(part, 'category_id', None),
        'status': getattr(part, 'status', ''),
        'merged_into_id': getattr(part, 'merged_into_id', None),
        'source_request_item_id': getattr(part, 'source_request_item_id', None),
        'created_at': _date_iso(getattr(part, 'created_at', None)),
    }


def _audit_spare(action, entity_type='', entity_id=None, entity_label='', before=None,
                 after=None, changes=None, description=''):
    log_audit(
        db,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_label=entity_label,
        module='spare_parts',
        before=before,
        after=after,
        changes=changes,
        description=description,
    )


def _deny_spare(attempted, entity_type='', entity_id=None, entity_label=''):
    """Record a permission-denied attempt on a financially significant spare
    parts action, then abort(403).

    [REASON]: CYCLE-2-3 Part 8 — forward-going audit contract completion:
    denials on money-relevant actions (viewing acts, issuing stock,
    approving/rejecting requests, price actions, catalog mutations) now
    leave a forensic trail. Deliberately NOT applied to every possible 403
    (org-scope checks, harmless read screens) — only where an owner would
    investigate who tried. Commits its own row because the request ends in
    abort(403) and no later caller commit will run; a logging failure never
    masks the 403 itself.
    """
    try:
        log_audit(
            db,
            action='spare_parts_access_denied',
            entity_type=entity_type,
            entity_id=entity_id,
            entity_label=entity_label,
            module='spare_parts',
            status='denied',
            description='Permission denied: ' + attempted,
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('audit write failed for denied %s', attempted)
    abort(403)


# ─── SPARE-STAGE1: price audit / repeat-order / photo-gate helpers ────────────

def write_price_audit(item_id, old_price, new_price, action, user_id):
    """Append a SparePartPriceAudit row for a price action.

    [REASON]: SPARE-STAGE1 — plain function with explicit arguments only (no
    current_user / request access) so the future Telegram bot process (BOT002B)
    can call it outside a Flask request context. Does not flush or commit --
    the caller is responsible for the final commit.
    """
    db.session.add(SparePartPriceAudit(
        item_id=item_id,
        old_price=old_price,
        new_price=new_price,
        action=action,
        changed_by=user_id,
    ))


def _update_sku_price_stats(sku_id):
    """Recompute a SKU's informational last_price / avg_price from its
    confirmed-price history.

    [REASON]: SPARE-STAGE2 business rule — one direction only: the price
    CONFIRM workflow writes SKU stats; SKU stats never set or bypass a
    request price (SparePartPriceAudit + price_status stay the single source
    of truth). avg_price is a simple running average over every 'confirm'
    audit row of items referencing this SKU. Does not commit — the caller's
    transaction covers it.
    """
    if not sku_id:
        return
    sku = SparePartSku.query.get(sku_id)
    if sku is None:
        return
    rows = (db.session.query(SparePartPriceAudit.new_price)
            .join(SparePartRequestItem,
                  SparePartPriceAudit.item_id == SparePartRequestItem.id)
            .filter(SparePartRequestItem.sku_id == sku_id,
                    SparePartPriceAudit.action == 'confirm',
                    SparePartPriceAudit.new_price.isnot(None))
            .order_by(SparePartPriceAudit.changed_at.asc(),
                      SparePartPriceAudit.id.asc())
            .all())
    prices = [float(r[0]) for r in rows]
    if not prices:
        return
    sku.last_price = prices[-1]
    sku.avg_price = round(sum(prices) / len(prices), 2)


def _check_repeat_orders(equipment_id, spare_part_id, exclude_request_id=None,
                         as_of_date=None, eligible_statuses=None):
    """Repeat-order warning engine.

    Returns {'severity': 'red'|'yellow'|None, 'days_since': int|None, 'history': [...]}.

    [REASON]: SPARE-STAGE1 — same equipment + same catalog part requested again
    within 90 days is surfaced to the operator/approver. <=7 days is 'red',
    <=30 days 'yellow', <=90 days history only. Rejected requests are excluded.
    This engine WARNS only -- it never blocks saving or submitting.

    [REASON]: SP-F-009 — hybrid contract (owner decision 2026-07-14): the live
    form/detail call sites pass neither new parameter, keeping today's broad
    behavior byte-identical (all non-rejected statuses, anchored to today);
    the REPORT passes both, making its repeat table strict and reproducible:
      - as_of_date anchors the window and days_since to the examined line's
        own request date instead of "today", so a report re-pulled later
        shows the same numbers;
      - eligible_statuses, when given, restricts matches to those statuses
        AND to requests strictly BEFORE as_of_date (a historical request must
        never be flagged because of a LATER one).
    """
    empty = {'severity': None, 'days_since': None, 'history': []}
    if not equipment_id or not spare_part_id:
        return empty
    as_of_date = as_of_date or date.today()
    window_start = as_of_date - timedelta(days=90)
    q = (db.session.query(SparePartRequestItem, SparePartRequest)
         .join(SparePartRequest,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .filter(SparePartRequest.equipment_id == equipment_id,
                 SparePartRequestItem.spare_part_id == spare_part_id,
                 SparePartRequest.request_date >= window_start))
    if eligible_statuses is not None:
        q = q.filter(SparePartRequest.status.in_(eligible_statuses),
                     SparePartRequest.request_date < as_of_date)
    else:
        q = q.filter(SparePartRequest.status != 'rejected')
    if exclude_request_id:
        q = q.filter(SparePartRequest.id != exclude_request_id)
    rows = q.order_by(SparePartRequest.request_date.desc(),
                      SparePartRequest.id.desc()).all()
    if not rows:
        return empty
    history = [{
        'request_id': req.id,
        'request_date': _date_iso(req.request_date),
        'status': req.status,
        'quantity': item.quantity,
        'unit': item.unit or '',
    } for item, req in rows]
    days_since = (as_of_date - rows[0][1].request_date).days
    if days_since <= 7:
        severity = 'red'
    elif days_since <= 30:
        severity = 'yellow'
    else:
        severity = None
    return {'severity': severity, 'days_since': days_since, 'history': history}


# ─── SPARE-STAGE2: warning rules 3, 4, 6 ─────────────────────────────────────
# Additive rule set alongside the SPARE-STAGE1 repeat engine above; the
# _check_repeat_orders function (rules 1-2) is deliberately not modified.

def _fmt_sum_text(value):
    """1234567.0 -> '1 234 567' (same style as the fmt_sum Jinja filter in app.py)."""
    return '{:,.0f}'.format(float(value or 0)).replace(',', ' ')


def _fmt_qty_text(value):
    """Trim trailing zeros: 3.0 -> '3', 2.5 -> '2.5' (same as spare_parts_pdf)."""
    value = float(value or 0)
    if value == int(value):
        return str(int(value))
    return ('{:.3f}'.format(value)).rstrip('0').rstrip('.')


def _month_start_back(d, months_back):
    """First day of the month `months_back` calendar months before d's month."""
    total = d.year * 12 + (d.month - 1) - months_back
    return date(total // 12, total % 12 + 1, 1)


def _rule3_result(request_dates):
    """Build the rule-3 warning dict from matched request dates (newest first).

    [REASON]: CYCLE-2-3 Part 9 — shared by the per-pair query path and the
    batched path so both produce identical dicts by construction.
    """
    if len(request_dates) <= RULE3_MAX_OCCURRENCES_90D:
        return None
    shown = [d.strftime('%d.%m.%Y') for d in request_dates[:8]]
    dates_text = ', '.join(shown) + ('…' if len(request_dates) > 8 else '')
    return {
        'rule': 3,
        'severity': 'yellow',
        'count': len(request_dates),
        'dates': [d.isoformat() for d in request_dates],
        'message': _spare_t(
            'Охирги {} кунда бу эҳтиёт қисм ушбу техника учун {} марта тасдиқланган: {}',
            'За последние {} дней эта деталь для этой техники уже утверждалась {} раз: {}'
        ).format(RULE3_WINDOW_DAYS, len(request_dates), dates_text),
    }


def _check_rule3_frequency(equipment_id, spare_part_id, exclude_request_id=None):
    """Rule 3 — frequency: same equipment + same canonical part on more than
    RULE3_MAX_OCCURRENCES_90D approved requests in the trailing window.

    [REASON]: SPARE-STAGE2 — authorized requests only (unlike rules 1-2, which
    count everything non-rejected): the frequency signal is about parts the
    business actually paid for repeatedly, per the task spec. Yellow only.

    [REASON]: SPARE-STAGE2-QA-FIX3 — count both 'approved' and 'issued'.
    'issued' is a Stage-2 terminal state reachable ONLY from 'approved' (goods
    physically handed over + write-off act exists), so it is a stronger signal
    of real authorized spend, not a reason to drop the request from the count.
    """
    window_start = date.today() - timedelta(days=RULE3_WINDOW_DAYS)
    q = (db.session.query(SparePartRequest.id, SparePartRequest.request_date)
         .join(SparePartRequestItem,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .filter(SparePartRequest.equipment_id == equipment_id,
                 SparePartRequestItem.spare_part_id == spare_part_id,
                 SparePartRequest.request_date >= window_start,
                 SparePartRequest.status.in_(('approved', 'issued'))))
    if exclude_request_id:
        q = q.filter(SparePartRequest.id != exclude_request_id)
    # distinct: the same part listed twice on one request counts once.
    # [REASON]: CYCLE-2-3 Part 9 — the id.desc() tiebreaker makes same-date
    # ordering deterministic (previously unspecified by SQLite), so the
    # per-pair and batched paths cannot disagree on cosmetic date order.
    rows = q.distinct().order_by(SparePartRequest.request_date.desc(),
                                 SparePartRequest.id.desc()).all()
    return _rule3_result([r.request_date for r in rows])


def _rule4_result(part, cost_rows, today):
    """Build the rule-4 warning dict from that category's (price, quantity,
    request_date) rows in the baseline-to-current window.

    [REASON]: CYCLE-2-3 Part 9 — shared by the per-pair query path and the
    batched path so both produce identical dicts by construction. Contains
    the entire original aggregation, unchanged.
    """
    month_start = today.replace(day=1)
    current_total = 0.0
    prior_total = 0.0
    prior_months_with_data = set()
    for price, quantity, req_date in cost_rows:
        cost = float(price or 0) * float(quantity or 0)
        # [REASON]: SP-F-016 defense-in-depth — these are already-stored
        # values, not raw input, but one corrupted NaN row must not poison
        # the whole category baseline.
        if not math.isfinite(cost):
            continue
        if req_date >= month_start:
            current_total += cost
        else:
            prior_total += cost
            prior_months_with_data.add((req_date.year, req_date.month))
    if not prior_months_with_data or prior_total <= 0:
        return None
    baseline_avg = prior_total / len(prior_months_with_data)
    if current_total <= RULE4_COST_ANOMALY_MULTIPLIER * baseline_avg:
        return None
    category = part.category_ref
    cat_name = ((category.name_ru if _spare_lang() == 'ru' else category.name_uz)
                if category else '')
    return {
        'rule': 4,
        'severity': 'yellow',
        'current_month_total': round(current_total, 2),
        'baseline_avg': round(baseline_avg, 2),
        'message': _spare_t(
            '«{cat}» категорияси: шу ойдаги харажат {cur} сўм — олдинги {months} ой ўртачасидан ({avg} сўм) {mult} баробардан кўп',
            'Категория «{cat}»: расходы в этом месяце {cur} сум — более чем в {mult} раза выше среднего за прошлые {months} мес. ({avg} сум)'
        ).format(cat=cat_name,
                 cur=_fmt_sum_text(current_total),
                 avg=_fmt_sum_text(baseline_avg),
                 months=RULE4_BASELINE_MONTHS,
                 mult=RULE4_COST_ANOMALY_MULTIPLIER),
    }


def _check_rule4_cost_anomaly(equipment_id, spare_part_id, exclude_request_id=None):
    """Rule 4 — category cost anomaly: current calendar month approved+confirmed
    spend for this equipment in this item's category vs the average of the
    RULE4_BASELINE_MONTHS months before the current one.

    [REASON]: SPARE-STAGE2 — spend counts only authorized money (approved or
    issued request + confirmed price, same rule as the SPARE-REPORTS cost
    tables). No warning without at least one prior month of spend — a first-ever
    purchase in a category has nothing to compare against.

    [REASON]: SP-F-023 — the baseline average divides by the number of prior
    months that actually had spend (len(prior_months_with_data)), not by the
    fixed RULE4_BASELINE_MONTHS window size: with real spend in only one of
    the three prior months, dividing by three artificially diluted the
    baseline to a third and made the anomaly trigger fire far too easily.

    [REASON]: SPARE-STAGE2-QA-FIX3 — 'issued' requests count exactly like
    'approved' ones (issued is reachable only from approved and means the money
    was actually spent), so both statuses feed the cost baseline.
    """
    part = SparePart.query.get(spare_part_id)
    if part is None or part.category_id is None:
        return None
    today = date.today()
    next_month_start = _month_start_back(today, -1)
    baseline_start = _month_start_back(today, RULE4_BASELINE_MONTHS)
    q = (db.session.query(SparePartRequestItem.price,
                          SparePartRequestItem.quantity,
                          SparePartRequest.request_date)
         .join(SparePartRequest,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .join(SparePart, SparePartRequestItem.spare_part_id == SparePart.id)
         .filter(SparePartRequest.equipment_id == equipment_id,
                 SparePart.category_id == part.category_id,
                 SparePartRequest.status.in_(('approved', 'issued')),
                 SparePartRequestItem.price_status == 'confirmed',
                 SparePartRequest.request_date >= baseline_start,
                 SparePartRequest.request_date < next_month_start))
    if exclude_request_id:
        q = q.filter(SparePartRequest.id != exclude_request_id)
    return _rule4_result(part, q.all(), today)


def _rule6_result(rejected):
    """Build the rule-6 warning dict from the most recent matching rejection.

    [REASON]: CYCLE-2-3 Part 9 — shared by the per-pair query path and the
    batched path so both produce identical dicts by construction.
    """
    if rejected is None:
        return None
    rejected_on = (rejected.reviewed_at.date() if rejected.reviewed_at
                   else rejected.request_date)
    comment = (rejected.review_comment or '').strip()
    comment_text = comment if comment else _spare_t('изоҳсиз', 'без комментария')
    return {
        'rule': 6,
        'severity': 'yellow',
        'request_id': rejected.id,
        'rejected_date': rejected_on.isoformat(),
        'comment': comment,
        'message': _spare_t(
            'Бу эҳтиёт қисм бўйича #{} сўров {} куни рад этилган. Изоҳ: {}',
            'Похожая заявка #{} на эту деталь была отклонена {}. Комментарий: {}'
        ).format(rejected.id, rejected_on.strftime('%d.%m.%Y'), comment_text),
    }


def _rule6_cutoffs(today=None):
    """(cutoff_dt, cutoff_date) pair for rule 6 — one definition for both paths."""
    cutoff_date = (today or date.today()) - timedelta(days=RULE6_REJECTED_LOOKBACK_DAYS)
    # [REASON]: SP-F-010 — both comparison paths use the same calendar-day
    # granularity: cutoff_dt is midnight of the cutoff day, so a rejection at
    # ANY time on that day matches, exactly like the whole-day request_date
    # fallback. Previously reviewed_at was compared against an exact UTC
    # timestamp while the fallback used whole days, so the boundary day
    # behaved differently depending on which field was populated.
    cutoff_dt = datetime.combine(cutoff_date, datetime.min.time())
    return cutoff_dt, cutoff_date


def _check_rule6_recent_rejection(equipment_id, spare_part_id, exclude_request_id=None):
    """Rule 6 — recently rejected, resubmitted: same equipment + same part was
    on a request rejected within the trailing RULE6_REJECTED_LOOKBACK_DAYS.

    The warning carries the rejection date and the reviewer's comment so the
    approver sees WHY it was rejected last time.
    """
    cutoff_dt, cutoff_date = _rule6_cutoffs()
    q = (db.session.query(SparePartRequest)
         .join(SparePartRequestItem,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .filter(SparePartRequest.equipment_id == equipment_id,
                 SparePartRequestItem.spare_part_id == spare_part_id,
                 SparePartRequest.status == 'rejected',
                 # [REASON]: reviewed_at is when the rejection actually happened;
                 # request_date is only the fallback for legacy rows without it.
                 db.or_(SparePartRequest.reviewed_at >= cutoff_dt,
                        db.and_(SparePartRequest.reviewed_at.is_(None),
                                SparePartRequest.request_date >= cutoff_date))))
    if exclude_request_id:
        q = q.filter(SparePartRequest.id != exclude_request_id)
    rejected = q.order_by(SparePartRequest.reviewed_at.desc(),
                          SparePartRequest.id.desc()).first()
    return _rule6_result(rejected)


def _model_display_name(model):
    """Model label honouring the viewer's language (name_uz when set + uz)."""
    if model is None:
        return ''
    if _spare_lang() == 'uz' and (model.name_uz or '').strip():
        return model.name_uz
    return model.name or ''


def _part_display_name(part, fallback=''):
    """Catalog part label honouring the viewer's language.

    [REASON]: CYCLE-2-3 Part 7 — the Uzbek interface shows name_uz when the
    owner has filled it in, and falls back to the canonical (Russian) `name`
    otherwise. This fallback is the intended long-term behavior, not a bug:
    translations arrive over time through the catalog UI and a missing one
    must never blank a part name. Same shape as _model_display_name above.
    """
    if part is None:
        return fallback
    if _spare_lang() == 'uz' and (getattr(part, 'name_uz', '') or '').strip():
        return part.name_uz
    return part.name or fallback


def _snapshot_display_name(snapshot_name, part, lang=None):
    """Display override for a STORED snapshot name (request/act item rows).

    [REASON]: CYCLE-2-3 Part 7 — only the Uzbek interface with a non-empty
    catalog alias overrides the snapshot; in every other case the snapshot
    is returned byte-identical, so Russian output and unlinked free-text
    items are completely unchanged. Stored values are never rewritten.
    """
    lang = lang or _spare_lang()
    if (lang == 'uz' and part is not None
            and (getattr(part, 'name_uz', '') or '').strip()):
        return part.name_uz
    return snapshot_name


def _check_rule5_incompatibility(equipment_id, spare_part_id, exclude_request_id=None):
    """Rule 5 — incompatibility: the part is NOT in the compatibility list of
    the request's equipment model.

    [REASON]: SPARE-STAGE3 — severity RED (like rules 1-2, unlike the yellow
    rules 3/4/6): requesting a part that is known-incompatible with the machine
    is a hard mistake, not a soft frequency/cost signal. Still warn-only — it
    never blocks submission or approval, same principle as every other rule.

    Fires ONLY when the part has at least one SparePartCompatibility row AND the
    equipment's model_id is not among them. Stays completely SILENT when:
      - the part has zero compatibility rows (compatibility "not yet defined" —
        the matrix starts empty on deploy, absence is never "incompatible"), or
      - the equipment's model_id is NULL (defensive; shouldn't happen after the
        Stage-3 migration, but never crash or false-positive if it does).

    exclude_request_id is accepted for signature parity with the other rules; it
    has no effect here (rule 5 depends on the catalog/equipment, not on prior
    requests).
    """
    if not equipment_id or not spare_part_id:
        return None
    compatible_ids = [
        row[0] for row in
        db.session.query(SparePartCompatibility.equipment_model_id)
        .filter(SparePartCompatibility.spare_part_id == spare_part_id).all()
    ]
    if not compatible_ids:
        # Compatibility not yet defined for this part -> silent.
        return None
    eq = Equipment.query.get(equipment_id)
    model_id = eq.model_id if eq else None
    if not model_id:
        # Equipment has no canonical model -> can't judge; silent (defensive).
        return None
    if model_id in compatible_ids:
        return None
    model = EquipmentModel.query.get(model_id)
    return _rule5_result(model_id, model)


def _rule5_result(model_id, model):
    """Build the rule-5 warning dict for a known-incompatible model.

    [REASON]: CYCLE-2-3 Part 9 — shared by the per-pair query path and the
    batched path so both produce identical dicts by construction.
    """
    model_name = _model_display_name(model) or ('#%s' % model_id)
    return {
        'rule': 5,
        'severity': 'red',
        'equipment_model_id': model_id,
        'message': _spare_t(
            'Бу деталь «{model}» модели учун мос деталлар рўйхатига кирмайди.',
            'Эта деталь не входит в список совместимых для модели «{model}».'
        ).format(model=model_name),
    }


def _check_extra_warnings(equipment_id, spare_part_id, exclude_request_id=None):
    """SPARE-STAGE2/3 aggregator for rules 3/4/5/6.

    Returns a list of warning dicts (possibly empty). Purely additive to the
    rules 1-2 engine: callers keep using _check_repeat_orders unchanged and
    attach this list alongside it. Each dict carries its own 'severity', so the
    red rule 5 renders red and the yellow rules 3/4/6 render yellow.
    """
    if not equipment_id or not spare_part_id:
        return []
    warnings = []
    # [REASON]: SPARE-STAGE3 — rule 5 (red incompatibility) listed first so the
    # hardest signal renders at the top of the warnings block.
    for check in (_check_rule5_incompatibility,
                  _check_rule3_frequency,
                  _check_rule4_cost_anomaly,
                  _check_rule6_recent_rejection):
        result = check(equipment_id, spare_part_id, exclude_request_id)
        if result:
            warnings.append(result)
    return warnings


# ─── CYCLE-2-3 Part 9: batched warning engines ────────────────────────────────
# [REASON]: the detail page ran _check_repeat_orders + 4 rule queries PER
# ITEM (N+1). The batched versions below run a FIXED number of queries for
# any item count and feed the exact same result-builder functions
# (_rule3_result/_rule4_result/_rule5_result/_rule6_result) as the per-pair
# engines, so their output is identical by construction; equivalence against
# the per-item originals is additionally asserted by tests and the
# scripts/spare_parts_nplus1_benchmark.py synthetic-volume run.

def _check_repeat_orders_batch(equipment_id, spare_part_ids, exclude_request_id=None):
    """Rules 1-2 for MANY parts of one request in a single query.

    Returns {spare_part_id: result-dict} with exactly the dict shape of
    _check_repeat_orders (today-anchored, non-rejected — the live detail
    contract; the report's strict/reproducible variant has its own batch in
    _reports_repeat_rows). Parts with no matches map to the empty result.
    """
    empty = {'severity': None, 'days_since': None, 'history': []}
    results = {pid: dict(empty, history=[]) for pid in spare_part_ids}
    if not equipment_id or not spare_part_ids:
        return results
    as_of_date = date.today()
    window_start = as_of_date - timedelta(days=90)
    q = (db.session.query(SparePartRequestItem, SparePartRequest)
         .join(SparePartRequest,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .filter(SparePartRequest.equipment_id == equipment_id,
                 SparePartRequestItem.spare_part_id.in_(set(spare_part_ids)),
                 SparePartRequest.request_date >= window_start,
                 SparePartRequest.status != 'rejected'))
    if exclude_request_id:
        q = q.filter(SparePartRequest.id != exclude_request_id)
    rows = q.order_by(SparePartRequest.request_date.desc(),
                      SparePartRequest.id.desc()).all()
    grouped = {}
    for item, req in rows:
        grouped.setdefault(item.spare_part_id, []).append((item, req))
    for pid, pair_rows in grouped.items():
        history = [{
            'request_id': req.id,
            'request_date': _date_iso(req.request_date),
            'status': req.status,
            'quantity': item.quantity,
            'unit': item.unit or '',
        } for item, req in pair_rows]
        days_since = (as_of_date - pair_rows[0][1].request_date).days
        if days_since <= 7:
            severity = 'red'
        elif days_since <= 30:
            severity = 'yellow'
        else:
            severity = None
        results[pid] = {'severity': severity, 'days_since': days_since,
                        'history': history}
    return results


def _check_extra_warnings_batch(equipment_id, spare_part_ids, exclude_request_id=None):
    """Rules 3/4/5/6 for MANY parts of one request in a fixed number of queries.

    Returns {spare_part_id: [warning dicts]} equal to calling
    _check_extra_warnings once per part (same per-part rule order 5,3,4,6).
    """
    spare_part_ids = [pid for pid in dict.fromkeys(spare_part_ids) if pid]
    if not equipment_id or not spare_part_ids:
        return {pid: [] for pid in spare_part_ids}
    today = date.today()
    parts = {p.id: p for p in
             SparePart.query.options(joinedload(SparePart.category_ref))
             .filter(SparePart.id.in_(spare_part_ids)).all()}

    # Rule 5 — compatibility rows for every part at once; equipment/model
    # resolved once (they are the same for every item of the request).
    compat = {}
    for pid, model_id in (db.session.query(
            SparePartCompatibility.spare_part_id,
            SparePartCompatibility.equipment_model_id)
            .filter(SparePartCompatibility.spare_part_id.in_(spare_part_ids)).all()):
        compat.setdefault(pid, set()).add(model_id)
    eq = Equipment.query.get(equipment_id)
    eq_model_id = eq.model_id if eq else None
    eq_model = EquipmentModel.query.get(eq_model_id) if eq_model_id else None

    # Rule 3 — distinct (request, date) matches for every part at once.
    r3_window_start = today - timedelta(days=RULE3_WINDOW_DAYS)
    r3_q = (db.session.query(SparePartRequest.id,
                             SparePartRequest.request_date,
                             SparePartRequestItem.spare_part_id)
            .join(SparePartRequestItem,
                  SparePartRequestItem.request_id == SparePartRequest.id)
            .filter(SparePartRequest.equipment_id == equipment_id,
                    SparePartRequestItem.spare_part_id.in_(spare_part_ids),
                    SparePartRequest.request_date >= r3_window_start,
                    SparePartRequest.status.in_(('approved', 'issued'))))
    if exclude_request_id:
        r3_q = r3_q.filter(SparePartRequest.id != exclude_request_id)
    r3_rows = {}
    for rid_, rdate, pid in r3_q.distinct().all():
        r3_rows.setdefault(pid, []).append((rdate, rid_))
    for pid in r3_rows:
        r3_rows[pid].sort(key=lambda t: (t[0], t[1]), reverse=True)

    # Rule 4 — one cost query covering every category the items belong to.
    cat_ids = {p.category_id for p in parts.values() if p.category_id is not None}
    r4_rows = {}
    if cat_ids:
        next_month_start = _month_start_back(today, -1)
        baseline_start = _month_start_back(today, RULE4_BASELINE_MONTHS)
        r4_q = (db.session.query(SparePartRequestItem.price,
                                 SparePartRequestItem.quantity,
                                 SparePartRequest.request_date,
                                 SparePart.category_id)
                .join(SparePartRequest,
                      SparePartRequestItem.request_id == SparePartRequest.id)
                .join(SparePart, SparePartRequestItem.spare_part_id == SparePart.id)
                .filter(SparePartRequest.equipment_id == equipment_id,
                        SparePart.category_id.in_(cat_ids),
                        SparePartRequest.status.in_(('approved', 'issued')),
                        SparePartRequestItem.price_status == 'confirmed',
                        SparePartRequest.request_date >= baseline_start,
                        SparePartRequest.request_date < next_month_start))
        if exclude_request_id:
            r4_q = r4_q.filter(SparePartRequest.id != exclude_request_id)
        for price, quantity, req_date, cat_id in r4_q.all():
            r4_rows.setdefault(cat_id, []).append((price, quantity, req_date))

    # Rule 6 — all matching rejections for every part at once; newest per part.
    cutoff_dt, cutoff_date = _rule6_cutoffs(today)
    r6_q = (db.session.query(SparePartRequest, SparePartRequestItem.spare_part_id)
            .join(SparePartRequestItem,
                  SparePartRequestItem.request_id == SparePartRequest.id)
            .filter(SparePartRequest.equipment_id == equipment_id,
                    SparePartRequestItem.spare_part_id.in_(spare_part_ids),
                    SparePartRequest.status == 'rejected',
                    db.or_(SparePartRequest.reviewed_at >= cutoff_dt,
                           db.and_(SparePartRequest.reviewed_at.is_(None),
                                   SparePartRequest.request_date >= cutoff_date))))
    if exclude_request_id:
        r6_q = r6_q.filter(SparePartRequest.id != exclude_request_id)
    r6_latest = {}
    for req, pid in r6_q.all():
        # Same pick as ORDER BY reviewed_at DESC, id DESC LIMIT 1 (SQLite
        # sorts NULL reviewed_at last in DESC order).
        key = (req.reviewed_at is not None, req.reviewed_at or datetime.min, req.id)
        if pid not in r6_latest or key > r6_latest[pid][0]:
            r6_latest[pid] = (key, req)

    results = {}
    for pid in spare_part_ids:
        part = parts.get(pid)
        warnings = []
        # Rule 5 — same silence conditions as _check_rule5_incompatibility.
        compatible_ids = compat.get(pid)
        if compatible_ids and eq_model_id and eq_model_id not in compatible_ids:
            warnings.append(_rule5_result(eq_model_id, eq_model))
        r3 = _rule3_result([d for d, _rid in r3_rows.get(pid, [])])
        if r3:
            warnings.append(r3)
        if part is not None and part.category_id is not None:
            r4 = _rule4_result(part, r4_rows.get(part.category_id, []), today)
            if r4:
                warnings.append(r4)
        r6 = _rule6_result(r6_latest[pid][1] if pid in r6_latest else None)
        if r6:
            warnings.append(r6)
        results[pid] = warnings
    return results


def _photo_required(item):
    """True when the item needs at least one photo before submit.

    [REASON]: SPARE-STAGE1 — categories with kind='unit' require photo proof;
    an item whose catalog part has no category yet (including fresh
    pending_review candidates) is treated as requiring a photo (safe default).
    Only kind='consumable' makes photos optional.
    """
    part = item.spare_part
    if part is None or part.category_id is None:
        return True
    cat = part.category_ref
    return (cat.kind if cat else 'unit') != 'consumable'


def _items_missing_photos(items):
    """Return bilingual row labels for items that require a photo but have none."""
    missing = []
    for idx, item in enumerate(items, start=1):
        if not _photo_required(item):
            continue
        if not (item.attachments or []):
            missing.append('{}. {}'.format(idx, item.name))
    return missing


# ─── SPARE-STAGE1: photo upload infrastructure ────────────────────────────────

def _upload_dir():
    """Absolute path of the spare parts upload folder (created on demand).

    UPLOAD_FOLDER is a plain path relative to the app folder so the Telegram
    bot process can resolve the same location without Flask context.
    """
    folder = current_app.config.get('UPLOAD_FOLDER') or 'instance/uploads/spare_parts'
    if not os.path.isabs(folder):
        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder)
    os.makedirs(folder, exist_ok=True)
    return folder


def _validate_photo(fs):
    """Validate an uploaded photo or video attachment.

    Returns (ext, None) on success, or (None, bilingual error) on failure.

    [REASON]: MOBILE-UPLOAD-001 — extends the SPARE-STAGE1 JPG/PNG-only
    check to the formats mechanics' phones actually produce (WEBP/HEIC
    photos, MP4/MOV/WEBM/AVI/MKV videos). Extension whitelist, size limit
    AND content magic-byte check are all required; the client-supplied
    Content-Type is never trusted.
    """
    filename = fs.filename or ''
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_ATTACHMENT_EXTENSIONS:
        return None, _spare_t(
            '«{}»: рухсат этилган форматлар — JPG, PNG, WEBP, HEIC/HEIF, MP4, MOV, WEBM, AVI, MKV',
            '«{}»: допустимые форматы — JPG, PNG, WEBP, HEIC/HEIF, MP4, MOV, WEBM, AVI, MKV').format(filename)

    fs.stream.seek(0, os.SEEK_END)
    size = fs.stream.tell()
    fs.stream.seek(0)
    if size > MAX_ATTACHMENT_SIZE:
        return None, _spare_t(
            '«{}»: файл ҳажми {} МБдан ошмаслиги керак',
            '«{}»: размер файла не должен превышать {} МБ').format(filename, MAX_ATTACHMENT_SIZE_MB)

    # [REASON]: WEBP/AVI (RIFF) signatures need bytes up to offset 12; read a
    # 16-byte header once and slice it for every check below.
    head = fs.stream.read(16)
    fs.stream.seek(0)

    if ext in ('.jpg', '.jpeg') and head.startswith(b'\xff\xd8\xff'):
        return ext, None
    if ext == '.png' and head.startswith(b'\x89PNG\r\n\x1a\n'):
        return ext, None
    if ext == '.webp' and head[0:4] == b'RIFF' and head[8:12] == b'WEBP':
        return ext, None
    if ext == '.avi' and head[0:4] == b'RIFF' and head[8:12] == b'AVI ':
        return ext, None
    if ext in ('.mp4', '.mov', '.heic', '.heif') and head[4:8] == b'ftyp':
        return ext, None
    if ext in ('.webm', '.mkv') and head[0:4] == b'\x1a\x45\xdf\xa3':
        return ext, None

    return None, _spare_t(
        '«{}»: файл мазмуни кўрсатилган кенгайтмага мос эмас',
        '«{}»: содержимое файла не соответствует указанному расширению').format(filename)


def _remove_stored_files(paths):
    """Best-effort removal of files written during a failed save.

    [REASON]: SP-F-003 — same try/except-OSError style as the PDF cleanup in
    issue_request: cleanup must never mask the original failure.
    """
    for full_path in paths or []:
        try:
            os.remove(full_path)
        except OSError:
            pass


def _store_item_photo(fs, item_id, user_id):
    """Validate and persist one uploaded photo for an item.

    Returns (error, full_path): a bilingual error string (full_path None), or
    (None, absolute path of the written file) on success. Adds the
    SparePartAttachment row without committing (caller commits).

    [REASON]: SPARE-STAGE1 — the on-disk name is built ONLY from item_id and a
    random suffix; original_filename is stored as metadata, never used for the
    path (path traversal risk), and the naming scheme is reproducible by the
    future bot process.
    """
    ext, err = _validate_photo(fs)
    if err:
        return err, None
    fname = '{}_{}{}'.format(item_id, uuid.uuid4().hex[:8], ext)
    full_path = os.path.join(_upload_dir(), fname)
    # [REASON]: SP-F-003 — if anything fails after bytes hit the disk (partial
    # save on a full disk, a failing stat, a session error), the file must not
    # survive as an orphan no row will ever reference.
    try:
        fs.save(full_path)
        db.session.add(SparePartAttachment(
            item_id=item_id,
            file_path=fname,
            original_filename=(fs.filename or '')[:300],
            file_size=os.path.getsize(full_path),
            uploaded_by=user_id,
        ))
    except Exception:
        _remove_stored_files([full_path])
        raise
    return None, full_path


def _store_attachments_for_item(files, item_id, user_id, existing_count):
    """Validate and store files for one item, up to MAX_ATTACHMENTS_PER_ITEM total.

    Returns (stored_count, errors, written_paths). Files beyond the per-item
    cap (existing attachments + files already stored earlier in this same
    call) are rejected with a clear bilingual error instead of being silently
    dropped or truncated mid-file -- the caller's already-stored files are
    untouched.

    [REASON]: SP-F-003 — written_paths lets the caller delete every file this
    call produced if the surrounding request later fails before commit; on an
    exception INSIDE this loop the already-written files of this call are
    removed here before re-raising.

    [REASON]: MOBILE-UPLOAD-001 — shared by the create-request flow (always
    existing_count=0) and the detail-page add-more flow (existing_count is
    the item's current SparePartAttachment count), so both enforce the same
    5-photos-and-videos-combined-per-item cap the same way.
    """
    errors = []
    stored = 0
    written_paths = []
    try:
        for fs in files:
            if existing_count + stored >= MAX_ATTACHMENTS_PER_ITEM:
                errors.append(_spare_t(
                    '«{}»: позицияга энг кўпи билан {} та файл бириктириш мумкин',
                    '«{}»: к позиции можно прикрепить не более {} файлов'
                ).format(fs.filename or '', MAX_ATTACHMENTS_PER_ITEM))
                continue
            err, full_path = _store_item_photo(fs, item_id, user_id)
            if err:
                errors.append(err)
            else:
                stored += 1
                written_paths.append(full_path)
    except Exception:
        _remove_stored_files(written_paths)
        raise
    return stored, errors, written_paths


def _store_submitted_item_photos(created_items):
    """Intake per-row photo/video files posted with the request form.

    created_items is a list of (item, prepared) pairs; each row's files arrive
    as item_photo_{form_index}. Invalid files are skipped and reported — if
    that leaves a mandatory photo/video missing, the submit gate keeps the
    request as a draft with its own explicit message.

    Returns the list of absolute paths written, so the caller can remove them
    if the request fails after this point but before commit (SP-F-003); an
    exception inside this loop removes THIS call's earlier items' files
    before re-raising.
    """
    errors = []
    written_paths = []
    try:
        for item, prepared in created_items:
            files = [fs for fs in
                     request.files.getlist('item_photo_{}'.format(prepared['form_index']))
                     if fs and fs.filename]
            _, item_errors, item_paths = _store_attachments_for_item(
                files, item.id, current_user.id, existing_count=0)
            errors.extend(item_errors)
            written_paths.extend(item_paths)
    except Exception:
        _remove_stored_files(written_paths)
        raise
    if errors:
        _spare_flash_errors(errors,
                            title_uz='Баъзи фотолар қабул қилинмади:',
                            title_ru='Некоторые фото не приняты:')
    return written_paths


def _approval_blockers(spr):
    """Bilingual error list explaining why a submitted request cannot be approved.

    [REASON]: SPARE-STAGE1 approval gate — a request is approvable only when
    every item points to a reviewed catalog entry (no pending_review) and every
    item price is confirmed.
    """
    errors = []
    for idx, item in enumerate(spr.items, start=1):
        part = item.spare_part
        if part is not None and part.status == 'pending_review':
            errors.append(_spare_t(
                '{}-позиция «{}»: янги каталог ёзуви ҳали текширилмаган (каталог менежери тасдиқлаши керак)',
                '{}. позиция «{}»: новая запись каталога ещё не проверена (нужно решение менеджера каталога)'
            ).format(idx, item.name))
        if item.price_status != 'confirmed':
            errors.append(_spare_t(
                '{}-позиция «{}»: нарх тасдиқланмаган',
                '{}. позиция «{}»: цена не подтверждена'
            ).format(idx, item.name))
    return errors


@spare_parts_bp.route('/')
@module_required('spare_parts')
def index():
    lang = getattr(g, 'lang', 'uz')
    status_filter = request.args.get('status', '')
    org_id = request.args.get('org_id', type=int)
    date_from_s = request.args.get('date_from', '')
    date_to_s = request.args.get('date_to', '')

    # PERF-SPARE-001B: eager-load spare parts index relations.
    # [REASON]: Avoid N+1 SELECTs for req.items, req.equipment, req.organization and req.creator.
    q = (SparePartRequest.query
         .options(
             joinedload(SparePartRequest.organization),
             joinedload(SparePartRequest.equipment),
             joinedload(SparePartRequest.creator),
             selectinload(SparePartRequest.items),
         ))
    user_org_ids = _spare_user_org_ids()
    if user_org_ids is not None:
        q = q.filter(SparePartRequest.organization_id.in_(user_org_ids))
    if status_filter:
        q = q.filter(SparePartRequest.status == status_filter)
    if org_id:
        if not current_user.can_access_org(org_id):
            abort(403)
        q = q.filter(SparePartRequest.organization_id == org_id)
    if date_from_s:
        try:
            q = q.filter(SparePartRequest.request_date >= datetime.strptime(date_from_s, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_s:
        try:
            q = q.filter(SparePartRequest.request_date <= datetime.strptime(date_to_s, '%Y-%m-%d').date())
        except ValueError:
            pass

    requests_list = q.order_by(SparePartRequest.created_at.desc()).all()

    counts = {s: 0 for s in STATUS_LABELS}
    count_base = SparePartRequest.query
    if user_org_ids:
        count_base = count_base.filter(SparePartRequest.organization_id.in_(user_org_ids))

    count_rows = (count_base
                  .with_entities(SparePartRequest.status,
                                 func.count(SparePartRequest.id))
                  .group_by(SparePartRequest.status)
                  .all())

    counts['all'] = 0
    for row_status, row_count in count_rows:
        row_count = int(row_count or 0)
        counts['all'] += row_count
        if row_status in counts:
            counts[row_status] = row_count

    if user_org_ids is None:
        organizations = Organization.query.order_by(Organization.sort_order).all()
    else:
        organizations = (Organization.query
                         .filter(Organization.id.in_(user_org_ids))
                         .order_by(Organization.sort_order).all())
    return render_template('spare_parts_list.html',
                           requests_list=requests_list,
                           status_filter=status_filter,
                           org_id=org_id,
                           date_from_s=date_from_s,
                           date_to_s=date_to_s,
                           counts=counts,
                           organizations=organizations,
                           status_labels=STATUS_LABELS,
                           status_colors=STATUS_COLORS,
                           lang=lang)


@spare_parts_bp.route('/desk')
@module_required('spare_parts')
def desk():
    """SP-DESK-001: role-aware operator workspace («Рабочий стол»).

    Read-only home page for the module: per-permission action queues with
    live counts. Built ONLY on existing statuses/permissions — no schema
    change, no new permission codes, and index() stays the requests list.
    """
    org_ids = _spare_user_org_ids()          # None => admin/all orgs
    uid = current_user.id

    R = SparePartRequest
    I = SparePartRequestItem
    P = SparePart

    def _scoped(q):
        # [REASON]: SEC003F org scoping — non-admin queue counts must match
        # what the user can actually open through index()/detail().
        return q if org_ids is None else q.filter(R.organization_id.in_(org_ids))

    # [REASON]: SP-DESK-001 — aggregate/EXISTS only (PERF-SPARE-001 rule: no
    # per-request Python loops on this hot page). A submitted request needs
    # pricing while ANY item is unconfirmed; it is ready for approval only
    # when fully priced AND no item points at a pending_review catalog part.
    unpriced = exists().where(and_(I.request_id == R.id,
                                   I.price_status != 'confirmed'))
    pending_part = exists().where(and_(I.request_id == R.id,
                                       I.spare_part_id == P.id,
                                       P.status == 'pending_review'))

    # [REASON]: SP-RESERVE-003 — «Готовы к выдаче» must stop lying: a SKU
    # item is covered iff an ACTIVE reservation holds its full quantity, and
    # only requests with no under-covered SKU item are physically issuable
    # right now; the rest of the approved queue waits for stock. Correlated
    # EXISTS only (PERF-SPARE-001: no Python loop over requests), and the two
    # tiles partition the approved queue exactly.
    # Coverage compares the reservation row against its OWN snapshot, not the
    # live item: requested_quantity is written with the same round(..., 3) as
    # quantity, so the comparison is self-consistent, whereas item.quantity is
    # the raw stored float — a >3-decimal quantity (possible via non-browser
    # callers; _parse_spare_positive_qty does not round) would otherwise
    # compare as under-covered forever. The snapshot cannot drift from the
    # item it describes: request items are immutable after creation (there is
    # no edit route — save_request only creates).
    RES = SparePartReservation
    item_covered = exists().where(and_(RES.request_item_id == I.id,
                                       RES.status == 'active',
                                       RES.quantity >= RES.requested_quantity))
    under_covered = exists().where(and_(I.request_id == R.id,
                                        I.sku_id.isnot(None),
                                        ~item_covered))

    submitted = _scoped(R.query.filter(R.status == 'submitted'))
    approved = _scoped(R.query.filter(R.status == 'approved'))

    counts = {
        # «Требуют действия». awaiting_price + awaiting_approval need NOT
        # equal total submitted: a fully priced request whose item still sits
        # in the classification queue counts in neither — it is blocked on
        # catalog work, surfaced by the needs_classification tile instead.
        'awaiting_price':       submitted.filter(unpriced).count(),
        'awaiting_approval':    submitted.filter(~unpriced).filter(~pending_part).count(),
        # Catalog parts are global reference data, deliberately NOT org-scoped.
        'needs_classification': P.query.filter(P.status == 'pending_review').count(),
        'ready_to_issue':       approved.filter(~under_covered).count(),
        'awaiting_stock':       approved.filter(under_covered).count(),
        'due_maintenance':      len(_maintenance_due_rows(org_ids=org_ids)),
        # «Мои заявки»
        'my_drafts':            _scoped(R.query.filter(R.status == 'draft',
                                                       R.created_by == uid)).count(),
        'my_returned':          _scoped(R.query.filter(R.status == 'returned_for_revision',
                                                       R.created_by == uid)).count(),
    }

    return render_template('spare_parts_desk.html',
                           counts=counts,
                           lang=_spare_lang())


@spare_parts_bp.route('/new')
@module_required('spare_parts')
def new_request():
    if not current_user.can_edit:
        flash(_spare_t('Сизда ҳуқуқ йўқ', 'У вас нет прав'), 'warning')
        return redirect(url_for('spare_parts.index'))
    user_org_ids = _spare_user_org_ids()
    if user_org_ids is None:
        organizations = Organization.query.order_by(Organization.sort_order).all()
    else:
        organizations = (Organization.query
                         .filter(Organization.id.in_(user_org_ids))
                         .order_by(Organization.sort_order).all())
    # [REASON]: SPARE-STAGE1 — equipment is now loaded per organization via
    # AJAX (api_equipment_by_org), so the full equipment list is no longer
    # rendered into the form. Catalog parts feed the datalist part picker.
    catalog_parts = (SparePart.query
                     .options(joinedload(SparePart.category_ref))
                     .filter(SparePart.status == 'active',
                             db.or_(SparePart.is_active == True,  # noqa: E712
                                    SparePart.is_active.is_(None)))
                     .order_by(SparePart.name)
                     .all())
    return render_template('spare_part_form.html',
                           req=None,
                           today=date.today().isoformat(),
                           organizations=organizations,
                           catalog_parts=catalog_parts,
                           # [REASON]: SP-F-024 — unit picker options; empty
                           # list keeps the legacy free-text input.
                           unit_options=_unit_options(_active_units()),
                           lang=_spare_lang())


@spare_parts_bp.route('/save', methods=['POST'])
@module_required('spare_parts')
def save_request():
    if not current_user.can_edit:
        abort(403)
    action = request.form.get('action', 'draft')
    if action not in {'draft', 'submit'}:
        abort(400)
    status = 'submitted' if action == 'submit' else 'draft'

    req_date_s = request.form.get('request_date', date.today().isoformat())
    try:
        req_date = _parse_spare_date(req_date_s)
    except ValueError:
        flash(_spare_t('Сана нотўғри', 'Некорректная дата'), 'warning')
        return redirect(url_for('spare_parts.new_request'))

    org_id = request.form.get('organization_id', type=int)
    eq_id = request.form.get('equipment_id', type=int) or None
    note = request.form.get('note', '').strip()

    if not org_id:
        flash(_spare_t('Ташкилотни танланг', 'Выберите организацию'), 'warning')
        return redirect(url_for('spare_parts.new_request'))
    _spare_check_org_access(org_id)
    if eq_id:
        eq = Equipment.query.get_or_404(eq_id)
        if eq.organization_id != org_id or not eq.is_active or not current_user.can_access_org(eq.organization_id):
            abort(403)

    names = request.form.getlist('item_name')
    parts = request.form.getlist('item_part_number')
    qtys = request.form.getlist('item_quantity')
    units = request.form.getlist('item_unit')
    notes = request.form.getlist('item_note')
    # [REASON]: SPARE-STAGE1 — hidden per-row field filled by the catalog part
    # picker; empty value means free-text entry ("not in catalog" flow).
    sp_ids = request.form.getlist('item_spare_part_id')
    # SPARE-STAGE2 — optional per-row SKU choice (select is always submitted,
    # empty value = no SKU, which keeps the exact Stage 1 item behaviour).
    sku_ids = request.form.getlist('item_sku_id')

    # [REASON]: SP-F-024 — strict unit directory: submitted values must be an
    # active SparePartUnit.code. An empty directory (fresh install before
    # migrate_spare_parts_units.py) keeps the legacy free-text acceptance so
    # nothing breaks pre-migration.
    valid_unit_codes = {u.code for u in _active_units()}

    prepared_items = []
    validation_errors = []
    for i, raw_name in enumerate(names):
        row_no = i + 1
        name = raw_name.strip()
        qty_raw = qtys[i].strip() if i < len(qtys) else ''
        part_number = parts[i].strip() if i < len(parts) else ''
        unit = units[i].strip() if i < len(units) else 'dona'
        item_note = notes[i].strip() if i < len(notes) else ''
        sp_raw = sp_ids[i].strip() if i < len(sp_ids) else ''

        if not name and not qty_raw and not part_number and not item_note:
            continue
        if not name:
            validation_errors.append(_spare_t('{}. қатор: запчасть номини киритинг', '{} строка: введите название запчасти').format(row_no))
            continue
        try:
            qty = _parse_spare_positive_qty(qty_raw)
        except ValueError:
            validation_errors.append(_spare_t('{}. қатор: миқдор мусбат бўлиши керак', '{} строка: количество должно быть больше нуля').format(row_no))
            continue
        # [REASON]: SP-F-024 — reject unknown units explicitly, never coerce
        # silently to a default.
        if valid_unit_codes and (unit or 'dona') not in valid_unit_codes:
            validation_errors.append(_spare_t(
                '{}. қатор: ўлчов бирлигини рўйхатдан танланг',
                '{} строка: выберите единицу измерения из справочника').format(row_no))
            continue

        # Resolve the catalog link. A stale/merged/inactive id silently falls
        # back to the free-text flow instead of failing the whole request.
        spare_part = None
        if sp_raw:
            try:
                sp_id = int(sp_raw)
            except ValueError:
                sp_id = None
            if sp_id:
                cand = SparePart.query.get(sp_id)
                if cand and cand.status == 'merged' and cand.merged_into_id:
                    cand = SparePart.query.get(cand.merged_into_id)
                if cand and cand.status in ('active', 'pending_review') and cand.is_active != False:  # noqa: E712
                    spare_part = cand

        # [REASON]: SPARE-STAGE2 — resolve the optional SKU with the same
        # silent-fallback rule as the catalog link above: a stale, inactive or
        # wrong-part SKU id degrades to "no SKU" instead of failing the save.
        sku = None
        sku_raw = sku_ids[i].strip() if i < len(sku_ids) else ''
        if sku_raw and spare_part is not None:
            try:
                sku_id = int(sku_raw)
            except ValueError:
                sku_id = None
            if sku_id:
                sku_cand = SparePartSku.query.get(sku_id)
                if (sku_cand and sku_cand.spare_part_id == spare_part.id
                        and sku_cand.is_active != False):  # noqa: E712
                    sku = sku_cand

        prepared_items.append({
            'name': name,
            'part_number': part_number,
            'quantity': qty,
            'unit': unit or 'dona',
            'note': item_note,
            'spare_part': spare_part,
            'sku': sku,
            'form_index': i,
        })

    if validation_errors:
        _spare_flash_errors(validation_errors)
        return redirect(url_for('spare_parts.new_request'))

    if not prepared_items:
        _spare_flash_errors([_spare_t('Камида битта позиция киритинг', 'Добавьте хотя бы одну позицию')])
        return redirect(url_for('spare_parts.new_request'))

    spr = SparePartRequest(
        request_date=req_date,
        organization_id=org_id,
        equipment_id=eq_id,
        status='draft',
        note=note,
        created_by=current_user.id,
    )
    db.session.add(spr)
    db.session.flush()

    created_items = []
    for prepared in prepared_items:
        item = SparePartRequestItem(
            request_id=spr.id,
            spare_part_id=prepared['spare_part'].id if prepared['spare_part'] else None,
            sku_id=prepared['sku'].id if prepared['sku'] else None,
            name=prepared['name'],
            part_number=prepared['part_number'],
            quantity=prepared['quantity'],
            unit=prepared['unit'],
            note=prepared['note'],
        )
        db.session.add(item)
        created_items.append((item, prepared))

    db.session.flush()

    # [REASON]: SPARE-STAGE1 "not in catalog" flow — a free-text item creates a
    # pending_review catalog candidate linked back to this item. The request
    # can be saved/submitted with the pending link but cannot be approved until
    # a catalog manager approves or merges the candidate (_approval_blockers).
    for item, prepared in created_items:
        if prepared['spare_part'] is not None:
            continue
        candidate = SparePart(
            name=item.name,
            part_number=item.part_number,
            unit=item.unit,
            status='pending_review',
            created_by=current_user.id,
            source_request_item_id=item.id,
        )
        db.session.add(candidate)
        db.session.flush()
        item.spare_part_id = candidate.id
        _audit_spare(
            'spare_part_catalog_created',
            entity_type='spare_part_catalog',
            entity_id=candidate.id,
            entity_label=candidate.name,
            after=_catalog_snapshot(candidate),
            description='Pending-review catalog candidate auto-created from request item #{}'.format(item.id)
        )

    # [REASON]: SP-F-003 — everything between the first byte written to disk
    # and the commit is one failure domain: if the transaction dies anywhere
    # in this block, every attachment file written for it is removed again
    # (best-effort), so no orphan files survive a failed save. The exception
    # then propagates exactly as before — behavior is unchanged apart from
    # the cleanup.
    written_paths = []
    try:
        written_paths = _store_submitted_item_photos(created_items)
        db.session.flush()

        # [REASON]: SPARE-STAGE1 mandatory-photo rule — items in 'unit' (or not yet
        # categorized) categories need at least one photo before submit. To avoid
        # losing the operator's typed rows, a failed submit is kept as a draft with
        # a clear bilingual error; photos can be added on the detail page.
        if status == 'submitted':
            missing = _items_missing_photos([item for item, _ in created_items])
            if missing:
                status = 'draft'
                _spare_flash_errors(
                    missing,
                    title_uz='Сўров юборилмади, чорнов сифатида сақланди. Қуйидаги позицияларга камида битта фото керак:',
                    title_ru='Заявка не отправлена и сохранена как черновик. Этим позициям нужно хотя бы одно фото:',
                )

        if status == 'submitted':
            spr.status = 'submitted'

        _audit_spare(
            'spare_part_request_created',
            entity_type='spare_part_request',
            entity_id=spr.id,
            entity_label='Request #{}'.format(spr.id),
            after=_request_snapshot(spr),
            description='Spare part request created'
        )
        for item, _prepared in created_items:
            _audit_spare(
                'spare_part_item_created',
                entity_type='spare_part_request_item',
                entity_id=item.id,
                entity_label=item.name,
                after=_item_snapshot(item),
                description='Spare part request item created'
            )
        # [REASON]: FIX003A — Record status history before commit.
        if status == 'submitted':
            _add_status_history(spr.id, None, 'submitted', changed_by=current_user.id)
        db.session.commit()
    except Exception:
        db.session.rollback()
        _remove_stored_files(written_paths)
        raise
    if status == 'submitted':
        flash(_spare_t('Сўров юборилди', 'Заявка отправлена'), 'success')
        # [REASON]: BOT003 — Post-commit best-effort notification, never raises.
        if _BOT003_AVAILABLE:
            enqueue_spare_request_submitted_best_effort(spr.id)
    elif status == 'draft':
        flash(_spare_t('Чорнов сақланди', 'Черновик сохранён'), 'info')
    return redirect(url_for('spare_parts.detail', rid=spr.id))


@spare_parts_bp.route('/<int:rid>')
@module_required('spare_parts')
def detail(rid):
    lang = getattr(g, 'lang', 'uz')
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    # SPARE-STAGE1 action visibility flags for the template.
    can_price = (current_user.has_module_access('spare_parts_price_confirm')
                 and spr.status == 'submitted')
    can_approve = (current_user.has_module_access('spare_parts_approve')
                   and spr.status == 'submitted')
    # [REASON]: SPARE-STAGE2 — issuing is a separate permission from approval
    # (warehouse handover vs purchase decision); only approved requests offer it.
    can_issue = (current_user.has_module_access('spare_parts_issue')
                 and spr.status == 'approved')
    issue_ctx = _issue_context(spr) if can_issue else None
    # Acts are visible on the request once issued (permanent, no un-issue).
    acts = (SparePartWriteOffAct.query.filter_by(request_id=spr.id)
            .order_by(SparePartWriteOffAct.id).all())
    can_edit_photos = (current_user.can_edit
                       and spr.created_by == current_user.id
                       and spr.status in ('draft', 'returned_for_revision'))
    # [REASON]: RE-SP-001 (4b) — a DB row whose file is gone from disk must be
    # a VISIBLE bilingual "file unavailable" status, not a silent broken
    # image. Display-only: the row is never altered or deleted here (issued
    # requests are immutable evidence; repair happens only through the
    # owner-approved reconciliation manifest process).
    upload_dir = _upload_dir()
    missing_attachment_ids = set()
    for item in spr.items:
        for att in (item.attachments or []):
            fname = os.path.basename(att.file_path or '')
            if not fname or not os.path.isfile(os.path.join(upload_dir, fname)):
                missing_attachment_ids.add(att.id)
    # [REASON]: SPARE-STAGE2 — the warn-only rules must be visible to the
    # reviewer at approval time, not only to the operator while typing the
    # request. Computed per item while the request is under review; rules 1-2
    # come from the untouched SPARE-STAGE1 engine, 3/4/6 from the new checks.
    item_warnings = {}
    if spr.status == 'submitted' and (can_price or can_approve) and spr.equipment_id:
        # [REASON]: CYCLE-2-3 Part 9 — batched: a fixed number of queries for
        # the whole request instead of ~5 queries per item. Output equality
        # with the per-item engines is covered by tests.
        part_ids = [item.spare_part_id for item in spr.items if item.spare_part_id]
        base_map = _check_repeat_orders_batch(spr.equipment_id, part_ids,
                                              exclude_request_id=spr.id)
        extra_map = _check_extra_warnings_batch(spr.equipment_id, part_ids,
                                                exclude_request_id=spr.id)
        for item in spr.items:
            if not item.spare_part_id:
                continue
            base = base_map.get(item.spare_part_id) or {
                'severity': None, 'days_since': None, 'history': []}
            extra = extra_map.get(item.spare_part_id, [])
            if base['severity'] or extra:
                item_warnings[item.id] = {'base': base, 'extra': extra}

    # ── SP-DETAIL-002: next-action resolver (read-only; every action still POSTs
    #    to its existing route or scrolls to its existing card — this only decides
    #    which single action is surfaced as primary for THIS user right now). ──
    is_owner = (current_user.id == spr.created_by)
    any_unpriced = any(i.price_status != 'confirmed' for i in spr.items)
    next_action = None
    waiting_for = None
    if spr.status in ('draft', 'returned_for_revision') and is_owner:
        # One-click POST in the block (matches the existing submit button's gate).
        next_action = {
            'kind': 'submit',
            'route': url_for('spare_parts.submit_request', rid=spr.id),
            'label': (_spare_t('Тузатиб, қайта юбориш', 'Исправить и отправить')
                      if spr.status == 'returned_for_revision'
                      else _spare_t('Кўриб чиқишга юбориш', 'Отправить на рассмотрение')),
        }
    elif spr.status == 'submitted' and can_price and any_unpriced:
        next_action = {
            'kind': 'scroll', 'target': 'sp-action-price',
            'label': _spare_t('Нарх қўйиш/тасдиқлаш', 'Проставить/подтвердить цену'),
        }
    elif spr.status == 'submitted' and can_approve and not _approval_blockers(spr):
        next_action = {
            'kind': 'scroll', 'target': 'sp-action-approve',
            'label': _spare_t('Тасдиқлаш', 'Утвердить'),
            'secondary': _spare_t('Рад этиш', 'Отклонить'),
        }
    elif spr.status == 'approved' and can_issue:
        next_action = {
            'kind': 'scroll', 'target': 'sp-action-issue',
            'label': _spare_t('Омбордан бериш', 'Выдать со склада'),
        }

    if next_action is None and spr.status not in ('rejected', 'issued'):
        if spr.status == 'draft':
            waiting_for = _spare_t('Оператор черновикни юборади', 'Оператор отправит заявку')
        elif spr.status == 'returned_for_revision':
            waiting_for = _spare_t('Оператор тузатиб юборади', 'Оператор исправит и отправит')
        elif spr.status == 'submitted':
            if any_unpriced:
                waiting_for = _spare_t('Нарх тасдиқланиши', 'Подтверждение цены')
            elif _approval_blockers(spr):
                waiting_for = _spare_t('Каталог текшируви', 'Проверка каталога')
            else:
                waiting_for = _spare_t('Администратор қарори', 'Решение администратора')
        elif spr.status == 'approved':
            waiting_for = _spare_t('Омбордан бериш', 'Выдача со склада')

    # ── SP-DETAIL-002: status stepper (draft→submitted→approved→issued; returned
    #    and rejected are side states). Pure presentation data. ──
    _step_defs = [
        ('draft',     {'uz': 'Чорнов',       'ru': 'Черновик'}),
        ('submitted', {'uz': 'Юборилган',    'ru': 'Отправлено'}),
        ('approved',  {'uz': 'Тасдиқланган', 'ru': 'Утверждено'}),
        ('issued',    {'uz': 'Берилган',     'ru': 'Выдано'}),
    ]
    if spr.status == 'returned_for_revision':
        _reached = 'submitted'; _side = 'returned'
    elif spr.status == 'rejected':
        _reached = 'submitted'; _side = 'rejected'
    else:
        _reached = spr.status; _side = None
    _order = [k for k, _ in _step_defs]
    _cur_idx = _order.index(_reached) if _reached in _order else 0
    steps = []
    for i, (key, label) in enumerate(_step_defs):
        state = 'done' if i < _cur_idx else ('current' if i == _cur_idx else 'future')
        steps.append({'key': key, 'label': label, 'state': state})
    stepper = {'steps': steps, 'side': _side}

    # ── SP-DETAIL-002: unified event timeline (created + status history + price
    #    audit + attachments), oldest → newest. No schema change; actors resolved
    #    here so the template stays presentation-only. ──
    def _uname(u):
        return (u.full_name or u.username) if u else _spare_t('номаълум', 'неизвестно')
    hist = (SparePartStatusHistory.query
            .filter_by(request_id=spr.id)
            .order_by(SparePartStatusHistory.changed_at.asc(),
                      SparePartStatusHistory.id.asc()).all())
    _hist_ids = {h.changed_by for h in hist if h.changed_by}
    _hist_users = ({u.id: u for u in User.query.filter(User.id.in_(_hist_ids)).all()}
                   if _hist_ids else {})
    price_events = (SparePartPriceAudit.query
                    .join(SparePartRequestItem,
                          SparePartPriceAudit.item_id == SparePartRequestItem.id)
                    .filter(SparePartRequestItem.request_id == spr.id)
                    .order_by(SparePartPriceAudit.changed_at.asc(),
                              SparePartPriceAudit.id.asc()).all())
    timeline = []
    timeline.append({'ts': spr.created_at, 'type': 'created', 'actor': _uname(spr.creator)})
    for h in hist:
        timeline.append({'ts': h.changed_at, 'type': 'status',
                         'actor': _uname(_hist_users.get(h.changed_by)),
                         'old_status': h.old_status, 'new_status': h.new_status,
                         'comment': h.comment or ''})
    for a in price_events:
        timeline.append({'ts': a.changed_at, 'type': 'price', 'actor': _uname(a.user),
                         'price_action': a.action,
                         'old_price': a.old_price, 'new_price': a.new_price,
                         'item_name': (a.item.name if a.item else '')})
    for item in spr.items:
        for att in (item.attachments or []):
            timeline.append({'ts': att.uploaded_at, 'type': 'attachment',
                             'actor': _uname(att.uploader),
                             'filename': (att.original_filename or ''),
                             'item_name': item.name})
    timeline.sort(key=lambda e: e['ts'] or spr.created_at)
    return render_template('spare_part_detail.html',
                           req=spr,
                           status_labels=STATUS_LABELS,
                           status_colors=STATUS_COLORS,
                           price_status_labels=PRICE_STATUS_LABELS,
                           price_status_colors=PRICE_STATUS_COLORS,
                           can_price=can_price,
                           can_approve=can_approve,
                           can_edit_photos=can_edit_photos,
                           can_issue=can_issue,
                           issue_ctx=issue_ctx,
                           acts=acts,
                           item_warnings=item_warnings,
                           missing_attachment_ids=missing_attachment_ids,
                           lang=lang,
                           next_action=next_action,
                           waiting_for=waiting_for,
                           stepper=stepper,
                           timeline=timeline)


@spare_parts_bp.route('/<int:rid>/submit', methods=['POST'])
@module_required('spare_parts')
def submit_request(rid):
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    # [REASON]: SPARE-STAGE1 — the price reject/return workflow sends a request
    # back as 'returned_for_revision'; the owner fixes it and resubmits from
    # the same button, so both statuses are submittable.
    if spr.status not in ('draft', 'returned_for_revision') or spr.created_by != current_user.id:
        abort(403)
    # [REASON]: SPARE-STAGE1 mandatory-photo rule at submit time.
    missing = _items_missing_photos(spr.items)
    if missing:
        _spare_flash_errors(
            missing,
            title_uz='Сўров юборилмади. Қуйидаги позицияларга камида битта фото керак:',
            title_ru='Заявка не отправлена. Этим позициям нужно хотя бы одно фото:',
        )
        return redirect(url_for('spare_parts.detail', rid=rid))
    before = _request_snapshot(spr)
    old_status = spr.status
    spr.status = 'submitted'
    after = _request_snapshot(spr)
    _audit_spare(
        'spare_part_request_status_changed',
        entity_type='spare_part_request',
        entity_id=spr.id,
        entity_label='Request #{}'.format(spr.id),
        before=before,
        after=after,
        changes={'status': {'before': old_status, 'after': spr.status}},
        description='Spare part request submitted'
    )
    # [REASON]: FIX003A — Record status history before commit.
    _add_status_history(spr.id, old_status, 'submitted', changed_by=current_user.id)
    db.session.commit()
    flash(_spare_t('Сўров юборилди', 'Заявка отправлена'), 'success')
    # [REASON]: BOT003 — Post-commit best-effort notification, never raises.
    if _BOT003_AVAILABLE:
        enqueue_spare_request_submitted_best_effort(spr.id)
    return redirect(url_for('spare_parts.detail', rid=rid))


@spare_parts_bp.route('/<int:rid>/approve', methods=['POST'])
@module_required('spare_parts')
def approve_request(rid):
    # [REASON]: SPARE-STAGE1 — final approval now requires the explicit
    # 'spare_parts_approve' permission; has_module_access() keeps the admin
    # bypass, so admins retain access and nobody else gains it automatically.
    if not current_user.has_module_access('spare_parts_approve'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('approve request #{}'.format(rid),
                    entity_type='spare_part_request', entity_id=rid,
                    entity_label='Request #{}'.format(rid))
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    if spr.status != 'submitted':
        abort(400)
    # [REASON]: SPARE-STAGE1 approval gate — pending_review catalog candidates
    # and unconfirmed prices block approval with an itemised bilingual error.
    blockers = _approval_blockers(spr)
    if blockers:
        _spare_flash_errors(
            blockers,
            title_uz='Сўровни тасдиқлаб бўлмайди:',
            title_ru='Заявку нельзя утвердить:',
        )
        return redirect(url_for('spare_parts.detail', rid=rid))
    before = _request_snapshot(spr)
    old_status = spr.status
    spr.status = 'approved'
    spr.reviewed_by = current_user.id
    spr.reviewed_at = datetime.utcnow()
    spr.review_comment = request.form.get('review_comment', '').strip()
    after = _request_snapshot(spr)
    _audit_spare(
        'spare_part_request_status_changed',
        entity_type='spare_part_request',
        entity_id=spr.id,
        entity_label='Request #{}'.format(spr.id),
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Spare part request approved; status {} -> {}'.format(old_status, spr.status)
    )
    # [REASON]: FIX003A — Record status history before commit.
    _add_status_history(spr.id, old_status, 'approved', comment=spr.review_comment or '', changed_by=current_user.id)
    # [REASON]: SP-RESERVE-003 — reserve stock in the SAME transaction as the
    # status change: either the request becomes approved together with its
    # reservation rows, or neither happens. Soft by design — approval is never
    # blocked by stock (see _create_reservations_for_request); shortages only
    # produce the post-commit warning below.
    shortages = _create_reservations_for_request(spr, current_user.id)
    sku_items = [i for i in spr.items if i.sku_id]
    total_needed = round(sum(round(float(i.quantity or 0), 3)
                             for i in sku_items), 3)
    total_short = round(sum(s['short'] for s in (shortages or [])), 3)
    reservation_wh = SparePartWarehouse.query.filter_by(
        organization_id=spr.organization_id).first()
    _audit_spare(
        'spare_part_reservations_created',
        entity_type='spare_part_request',
        entity_id=spr.id,
        entity_label='Request #{}'.format(spr.id),
        after={'warehouse_id': reservation_wh.id if reservation_wh else None,
               'rows': len(sku_items) if shortages is not None else 0,
               'total_reserved': (round(total_needed - total_short, 3)
                                  if shortages is not None else 0),
               'total_short': total_short},
        description='Reservations created on approval of request #{}'.format(spr.id)
    )
    db.session.commit()
    flash(_spare_t('Сўров тасдиқланди', 'Заявка утверждена'), 'success')
    # [REASON]: SP-RESERVE-003 — non-blocking notices about the reservation
    # outcome, shown alongside (never instead of) the success flash.
    if shortages is None and sku_items:
        flash(_spare_t('Захира яратилмади: ташкилотда эҳтиёт қисмлар омбори йўқ.',
                       'Резерв не создан: у организации нет склада запчастей.'),
              'info')
    elif shortages:
        _spare_flash_errors(
            [_spare_t('«{}»: керак {}, захираланган {}, етишмайди {}',
                      '«{}»: нужно {}, зарезервировано {}, не хватает {}'
                      ).format(s['item'].name,
                               _fmt_qty_text(s['needed']),
                               _fmt_qty_text(s['reserved']),
                               _fmt_qty_text(s['short']))
             for s in shortages],
            title_uz='Сўров тасдиқланди, лекин омбор ҳаммасига етмади:',
            title_ru='Заявка утверждена, но склада хватило не на всё:')
    # [REASON]: BOT003 — Post-commit best-effort notification, never raises.
    if _BOT003_AVAILABLE:
        enqueue_spare_request_status_best_effort(spr.id, 'spare_request_approved')
    return redirect(url_for('spare_parts.detail', rid=rid))


@spare_parts_bp.route('/<int:rid>/reject', methods=['POST'])
@module_required('spare_parts')
def reject_request(rid):
    # [REASON]: SPARE-STAGE1 — same permission as approval (see approve_request).
    if not current_user.has_module_access('spare_parts_approve'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('reject request #{}'.format(rid),
                    entity_type='spare_part_request', entity_id=rid,
                    entity_label='Request #{}'.format(rid))
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    if spr.status != 'submitted':
        abort(400)
    before = _request_snapshot(spr)
    old_status = spr.status
    spr.status = 'rejected'
    spr.reviewed_by = current_user.id
    spr.reviewed_at = datetime.utcnow()
    spr.review_comment = request.form.get('review_comment', '').strip()
    after = _request_snapshot(spr)
    _audit_spare(
        'spare_part_request_status_changed',
        entity_type='spare_part_request',
        entity_id=spr.id,
        entity_label='Request #{}'.format(spr.id),
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Spare part request rejected; status {} -> {}'.format(old_status, spr.status)
    )
    # [REASON]: FIX003A — Record status history before commit.
    _add_status_history(spr.id, old_status, 'rejected', comment=spr.review_comment or '', changed_by=current_user.id)
    db.session.commit()
    flash(_spare_t('Сўров рад этилди', 'Заявка отклонена'), 'warning')
    # [REASON]: BOT003 — Post-commit best-effort notification, never raises.
    if _BOT003_AVAILABLE:
        enqueue_spare_request_status_best_effort(spr.id, 'spare_request_rejected')
    return redirect(url_for('spare_parts.detail', rid=rid))


# ─── SPARE-STAGE1: price workflow ─────────────────────────────────────────────

def _price_action_target(rid, item_id):
    """Shared guard for price actions: access, permission, status, ownership."""
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    # [REASON]: SPARE-STAGE1 — price actions require the explicit
    # 'spare_parts_price_confirm' permission (admin passes automatically).
    if not current_user.has_module_access('spare_parts_price_confirm'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('price action on request #{} item #{}'.format(rid, item_id),
                    entity_type='spare_part_request_item', entity_id=item_id)
    # Prices are set/confirmed while the request is under review.
    if spr.status != 'submitted':
        abort(400)
    item = SparePartRequestItem.query.get_or_404(item_id)
    if item.request_id != spr.id:
        abort(404)
    return spr, item


@spare_parts_bp.route('/<int:rid>/items/<int:item_id>/price/set', methods=['POST'])
@module_required('spare_parts')
def item_price_set(rid, item_id):
    spr, item = _price_action_target(rid, item_id)
    raw = (request.form.get('price', '') or '').strip()
    try:
        price = float(raw.replace(',', '.').replace(' ', ''))
        # [REASON]: SP-F-016 — reject 'nan'/'inf' the same way as a malformed
        # number (NaN would pass the < 0 check and poison SKU price stats).
        if not math.isfinite(price) or price < 0:
            raise ValueError()
    except (TypeError, ValueError):
        _spare_flash_errors([_spare_t('Нарх мусбат сон бўлиши керак',
                                      'Цена должна быть неотрицательным числом')],
                            title_uz='Нарх сақланмади:', title_ru='Цена не сохранена:')
        return redirect(url_for('spare_parts.detail', rid=rid))
    old_price = item.price
    item.price = price
    # Setting a new value restarts the confirm cycle.
    item.price_status = 'pending'
    item.price_set_by = current_user.id
    item.price_set_at = datetime.utcnow()
    write_price_audit(item.id, old_price, price, 'set', current_user.id)
    _audit_spare(
        'spare_part_item_price_set',
        entity_type='spare_part_request_item',
        entity_id=item.id,
        entity_label=item.name,
        changes={'price': {'before': old_price, 'after': price}},
        description='Item price set on request #{}'.format(spr.id)
    )
    db.session.commit()
    flash(_spare_t('Нарх сақланди', 'Цена сохранена'), 'success')
    return redirect(url_for('spare_parts.detail', rid=rid))


@spare_parts_bp.route('/<int:rid>/items/<int:item_id>/price/confirm', methods=['POST'])
@module_required('spare_parts')
def item_price_confirm(rid, item_id):
    spr, item = _price_action_target(rid, item_id)
    if item.price is None:
        _spare_flash_errors([_spare_t('Аввал нархни киритинг', 'Сначала укажите цену')],
                            title_uz='Нарх тасдиқланмади:', title_ru='Цена не подтверждена:')
        return redirect(url_for('spare_parts.detail', rid=rid))
    # [REASON]: SP-F-007 — confirm is idempotent: a second confirm (stale tab,
    # double click) is a no-op, otherwise it would write a duplicate
    # price_audit row and double-count this price in the SKU's average via
    # _update_sku_price_stats. Only a 'pending' price can be confirmed;
    # rejected/returned prices leave the submitted state anyway.
    if item.price_status != 'pending':
        flash(_spare_t('Нарх аллақачон тасдиқланган', 'Цена уже подтверждена'), 'info')
        return redirect(url_for('spare_parts.detail', rid=rid))
    item.price_status = 'confirmed'
    write_price_audit(item.id, item.price, item.price, 'confirm', current_user.id)
    # [REASON]: SPARE-STAGE2 — confirm is the only event that feeds SKU price
    # stats (one-way write; see _update_sku_price_stats). Runs before commit so
    # the stats land in the same transaction as the confirmation itself.
    if item.sku_id:
        _update_sku_price_stats(item.sku_id)
    _audit_spare(
        'spare_part_item_price_confirmed',
        entity_type='spare_part_request_item',
        entity_id=item.id,
        entity_label=item.name,
        changes={'price_status': {'before': 'pending', 'after': 'confirmed'}},
        description='Item price confirmed on request #{}'.format(spr.id)
    )
    db.session.commit()
    flash(_spare_t('Нарх тасдиқланди', 'Цена подтверждена'), 'success')
    return redirect(url_for('spare_parts.detail', rid=rid))


@spare_parts_bp.route('/<int:rid>/items/<int:item_id>/price/reject', methods=['POST'])
@module_required('spare_parts')
def item_price_reject(rid, item_id):
    spr, item = _price_action_target(rid, item_id)
    mode = request.form.get('mode', 'reject')
    if mode not in ('reject', 'return'):
        abort(400)
    comment = (request.form.get('comment', '') or '').strip()
    old_price_status = item.price_status
    item.price_status = 'rejected' if mode == 'reject' else 'returned'
    write_price_audit(item.id, item.price, item.price, mode, current_user.id)
    # [REASON]: SPARE-STAGE1 — a price rejection/return sends the whole request
    # back to the operator; recorded in BOTH SparePartPriceAudit (above) and
    # SparePartStatusHistory (below), per the approved workflow.
    old_status = spr.status
    spr.status = 'returned_for_revision'
    _add_status_history(spr.id, old_status, 'returned_for_revision',
                        comment=comment or 'Item #{} price {}'.format(item.id, item.price_status),
                        changed_by=current_user.id)
    _audit_spare(
        'spare_part_request_status_changed',
        entity_type='spare_part_request',
        entity_id=spr.id,
        entity_label='Request #{}'.format(spr.id),
        changes={'status': {'before': old_status, 'after': spr.status},
                 'item_price_status': {'before': old_price_status, 'after': item.price_status}},
        description='Request returned for revision: item #{} price {}'.format(item.id, item.price_status)
    )
    db.session.commit()
    flash(_spare_t('Сўров қайта ишлашга қайтарилди', 'Заявка возвращена на доработку'), 'warning')
    return redirect(url_for('spare_parts.detail', rid=rid))


# ─── SPARE-STAGE1: photo routes ───────────────────────────────────────────────

@spare_parts_bp.route('/<int:rid>/items/<int:item_id>/photo', methods=['POST'])
@module_required('spare_parts')
def item_photo_upload(rid, item_id):
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    item = SparePartRequestItem.query.get_or_404(item_id)
    if item.request_id != spr.id:
        abort(404)
    is_owner = current_user.can_edit and spr.created_by == current_user.id
    if not (is_owner or current_user.is_admin):
        abort(403)
    # Photos are attached while the operator can still change the request.
    if spr.status not in ('draft', 'returned_for_revision'):
        abort(400)
    # [REASON]: MOBILE-UPLOAD-001 — freshly counted (not item.attachments, which
    # may be a stale relationship cache) so the 5-per-item cap is enforced
    # against what is actually stored right now.
    existing_count = SparePartAttachment.query.filter_by(item_id=item.id).count()
    files = [fs for fs in request.files.getlist('photo') if fs and fs.filename]
    # [REASON]: SP-F-003 — same write-then-commit failure domain as
    # save_request: this route shares the identical gap (files written, commit
    # later, no surrounding try/except), so a failure between the two removes
    # every file written by this call before propagating.
    written_paths = []
    stored = 0
    errors = []
    try:
        stored, errors, written_paths = _store_attachments_for_item(
            files, item.id, current_user.id, existing_count)
        if stored:
            _audit_spare(
                'spare_part_item_photo_uploaded',
                entity_type='spare_part_request_item',
                entity_id=item.id,
                entity_label=item.name,
                description='{} photo(s) uploaded for item on request #{}'.format(stored, spr.id)
            )
            db.session.commit()
    except Exception:
        db.session.rollback()
        _remove_stored_files(written_paths)
        raise
    if errors:
        _spare_flash_errors(errors,
                            title_uz='Баъзи фотолар қабул қилинмади:',
                            title_ru='Некоторые фото не приняты:')
    if stored:
        flash(_spare_t('Фото юкланди', 'Фото загружено'), 'success')
    elif not errors:
        flash(_spare_t('Файл танланмаган', 'Файл не выбран'), 'warning')
    return redirect(url_for('spare_parts.detail', rid=rid))


@spare_parts_bp.route('/attachments/<int:att_id>')
@module_required('spare_parts')
def attachment_file(att_id):
    att = SparePartAttachment.query.get_or_404(att_id)
    item = att.item
    spr = item.request if item else None
    if spr is None:
        abort(404)
    # [REASON]: SPARE-STAGE1 — uploads are NOT served as raw static files; the
    # same org-scoped access check as detail(rid) applies to every download.
    _spare_check_request_access(spr)
    fname = os.path.basename(att.file_path or '')
    if not fname:
        abort(404)
    # [REASON]: MOBILE-UPLOAD-001 — explicit Content-Type per stored
    # extension so videos render/scrub correctly in <video> and images
    # don't fall back to a generic download prompt; conditional=True (Flask
    # default) gives Range-request support for free, which video seeking
    # needs.
    ext = os.path.splitext(fname)[1].lower()
    mimetype = ATTACHMENT_CONTENT_TYPES.get(ext, 'application/octet-stream')
    return send_from_directory(_upload_dir(), fname, mimetype=mimetype)


# ─── SPARE-STAGE1: JSON endpoints ─────────────────────────────────────────────

@spare_parts_bp.route('/api/equipment-by-org')
@module_required('spare_parts')
def api_equipment_by_org():
    # [REASON]: SPARE-STAGE1 — mirrors work_orders.api_equipment_by_org (that
    # file is outside this task's scope fence and returns no eq_type, which the
    # spare parts form must show as read-only reference text). Same org-access
    # rule and active-equipment filter.
    org_id = request.args.get('org_id', type=int)
    if not org_id or not current_user.can_access_org(org_id):
        return jsonify([])
    eqs = (Equipment.query
           .filter_by(organization_id=org_id, is_active=True)
           .order_by(Equipment.name, Equipment.plate).all())
    return jsonify([
        {
            'id': e.id,
            'name': e.name,
            'plate': e.plate or '',
            'eq_type': e.eq_type or '',
            'category': e.category_display,
        }
        for e in eqs
    ])


@spare_parts_bp.route('/api/repeat-check')
@module_required('spare_parts')
def api_repeat_check():
    equipment_id = request.args.get('equipment_id', type=int)
    spare_part_id = request.args.get('spare_part_id', type=int)
    exclude_request_id = request.args.get('exclude_request_id', type=int)
    empty = {'severity': None, 'days_since': None, 'history': [], 'extra_warnings': []}
    if not equipment_id or not spare_part_id:
        return jsonify(empty)
    eq = Equipment.query.get(equipment_id)
    if not eq or not current_user.can_access_org(eq.organization_id):
        return jsonify(empty)
    result = _check_repeat_orders(equipment_id, spare_part_id, exclude_request_id)
    # [REASON]: SPARE-STAGE2 — rules 3/4/6 ride along under an additive key;
    # the original rules 1-2 payload stays byte-identical for existing callers.
    result['extra_warnings'] = _check_extra_warnings(
        equipment_id, spare_part_id, exclude_request_id)
    return jsonify(result)


@spare_parts_bp.route('/api/price-suggestions')
@module_required('spare_parts')
def api_price_suggestions():
    # [REASON]: SPARE-STAGE1 — last 3 confirmed prices for a catalog part,
    # shown as a hint when setting a price. Exact part match only, no fuzzy.
    spare_part_id = request.args.get('spare_part_id', type=int)
    if not spare_part_id:
        return jsonify([])
    rows = (SparePartPriceAudit.query
            .join(SparePartRequestItem,
                  SparePartPriceAudit.item_id == SparePartRequestItem.id)
            .filter(SparePartRequestItem.spare_part_id == spare_part_id,
                    SparePartPriceAudit.action == 'confirm')
            .order_by(SparePartPriceAudit.changed_at.desc())
            .limit(3)
            .all())
    return jsonify([
        {'price': a.new_price,
         'date': a.changed_at.date().isoformat() if a.changed_at else None}
        for a in rows
    ])


# ─── SPARE-STAGE2: fuzzy catalog search ───────────────────────────────────────

def _norm_search_text(value):
    """Normalize text for catalog search comparison.

    [REASON]: SPARE-STAGE2 — same normalization approach as app.py's
    _task_ref_001d_norm (casefold + collapse whitespace), plus punctuation
    folded to spaces, so «Фильтр, масляный.» equals «фильтр масляный» before
    any fuzzy scoring; difflib then only has to absorb genuine typos.
    """
    value = (value or '').casefold()
    value = re.sub(r'[^\w\s]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def _fuzzy_score(query_norm, name_norm):
    """Best difflib ratio between direct and token-sorted comparison.

    [REASON]: the token-sorted pass makes word-order swaps («масляный фильтр»
    vs «фильтр масляный») score near 1.0 instead of ~0.5, without any
    third-party fuzzy library.
    """
    direct = SequenceMatcher(None, query_norm, name_norm).ratio()
    query_sorted = ' '.join(sorted(query_norm.split()))
    name_sorted = ' '.join(sorted(name_norm.split()))
    if query_sorted == query_norm and name_sorted == name_norm:
        return direct
    return max(direct, SequenceMatcher(None, query_sorted, name_sorted).ratio())


def _search_catalog_parts(query, limit=FUZZY_SEARCH_MAX_RESULTS):
    """Ranked catalog search over selectable parts: exact > substring > fuzzy.

    Matches against the part name only (single `name` column in Stage 1/2).
    Returns a list of (part, match_kind, score) tuples. Stdlib difflib only —
    at this catalog's scale (a few hundred rows) scoring the full candidate
    list per request is fast enough, no search dependency needed.

    [REASON]: exact and substring tiers always rank above fuzzy-only hits, so
    a correctly typed name can never be outranked by a fuzzy competitor.
    """
    query_norm = _norm_search_text(query)
    if not query_norm:
        return []
    parts = (SparePart.query
             .options(joinedload(SparePart.category_ref))
             .filter(SparePart.status == 'active',
                     db.or_(SparePart.is_active == True,  # noqa: E712
                            SparePart.is_active.is_(None)))
             .order_by(SparePart.name)
             .all())
    exact, substr, fuzzy = [], [], []
    for part in parts:
        # [REASON]: CYCLE-2-3 Part 7 — the query is matched against BOTH the
        # canonical name and the optional Uzbek alias (best score wins), so
        # a part found by its Uzbek alias still resolves to the same catalog
        # entry. Parts without an alias behave exactly as before.
        norms = [n for n in (_norm_search_text(part.name),
                             _norm_search_text(getattr(part, 'name_uz', '') or ''))
                 if n]
        if not norms:
            continue
        if any(n == query_norm for n in norms):
            exact.append((part, 'exact', 1.0))
            continue
        score = max(_fuzzy_score(query_norm, n) for n in norms)
        if any(query_norm in n for n in norms):
            substr.append((part, 'substring', score))
        elif score >= FUZZY_MATCH_MIN_RATIO:
            fuzzy.append((part, 'fuzzy', score))
    # Python sort is stable: ties inside a tier keep alphabetical order.
    substr.sort(key=lambda t: -t[2])
    fuzzy.sort(key=lambda t: -t[2])
    return (exact + substr + fuzzy)[:limit]


@spare_parts_bp.route('/api/catalog-search')
@module_required('spare_parts')
def api_catalog_search():
    """Typo-tolerant catalog picker search (SPARE-STAGE2, Task 2).

    Used by the request form's part picker and the catalog screen's search
    box. Same candidate set as the form's datalist: active, non-merged parts.
    """
    q = (request.args.get('q', '') or '').strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify([
        {
            # [REASON]: CYCLE-2-3 Part 7 — suggestions show the viewer's
            # language (alias in uz when set); clicking one re-runs the
            # form's exact-match flow against the same display value.
            'id': part.id,
            'name': _part_display_name(part),
            'part_number': part.part_number or '',
            'unit': part.unit or 'dona',
            'kind': (part.category_ref.kind if part.category_ref else ''),
            'match': match_kind,
            'score': round(score, 3),
        }
        for part, match_kind, score in _search_catalog_parts(q)
    ])


# ─── SPARE-STAGE2: SKU catalog ────────────────────────────────────────────────

def _sku_snapshot(sku):
    if not sku:
        return None
    return {
        'id': getattr(sku, 'id', None),
        'spare_part_id': getattr(sku, 'spare_part_id', None),
        'brand': getattr(sku, 'brand', ''),
        'article_number': getattr(sku, 'article_number', ''),
        'supplier': getattr(sku, 'supplier', ''),
        'last_price': getattr(sku, 'last_price', None),
        'avg_price': getattr(sku, 'avg_price', None),
        'is_active': getattr(sku, 'is_active', None),
    }


@spare_parts_bp.route('/api/skus-by-part')
@module_required('spare_parts')
def api_skus_by_part():
    """Active SKUs of one canonical part, for the request form's SKU picker.

    last_price is included as a SUGGESTION only — the price confirm workflow
    still applies to whatever ends up on the request item.
    """
    spare_part_id = request.args.get('spare_part_id', type=int)
    if not spare_part_id:
        return jsonify([])
    skus = (SparePartSku.query
            .filter(SparePartSku.spare_part_id == spare_part_id,
                    db.or_(SparePartSku.is_active == True,  # noqa: E712
                           SparePartSku.is_active.is_(None)))
            .order_by(SparePartSku.brand, SparePartSku.article_number,
                      SparePartSku.id)
            .all())
    return jsonify([
        {
            'id': s.id,
            'brand': s.brand or '',
            'article_number': s.article_number or '',
            'supplier': s.supplier or '',
            'label': s.label,
            'last_price': s.last_price,
            'avg_price': s.avg_price,
        }
        for s in skus
    ])


@spare_parts_bp.route('/skus')
@module_required('spare_parts')
def skus():
    # [REASON]: SPARE-STAGE2 — SKU management is part of catalog stewardship,
    # so it sits behind the existing spare_parts_catalog_manage permission
    # (no new permission for this screen, per the task decision).
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        abort(403)
    sku_rows = (SparePartSku.query
                .options(joinedload(SparePartSku.spare_part))
                .order_by(SparePartSku.spare_part_id, SparePartSku.brand,
                          SparePartSku.id)
                .all())
    catalog_parts = (SparePart.query
                     .filter(SparePart.status == 'active',
                             db.or_(SparePart.is_active == True,  # noqa: E712
                                    SparePart.is_active.is_(None)))
                     .order_by(SparePart.name)
                     .all())
    return render_template('spare_parts_skus.html',
                           sku_rows=sku_rows,
                           catalog_parts=catalog_parts,
                           lang=_spare_lang())


def _sku_normalized_clash(spare_part_id, brand, article_number, supplier,
                          exclude_id=None):
    """True when an ACTIVE SKU with the same normalized identity exists.

    [REASON]: SP-F-008 — owner decision: exact normalized duplicates are
    forbidden. This pre-check mirrors the partial unique index
    uq_spare_part_skus_normalized (same lower(trim(...)) on BOTH sides so
    SQLite's ASCII-only lower() semantics match the index exactly, same
    active-only scope) to show a friendly message for the common case; the
    index remains the real enforcement for the race case (RE-SP-002).
    """
    clash_q = (SparePartSku.query
               .filter(SparePartSku.spare_part_id == spare_part_id,
                       SparePartSku.is_active == True,  # noqa: E712 — mirrors the index WHERE is_active = 1
                       func.lower(func.trim(SparePartSku.brand)) == func.lower(brand),
                       func.lower(func.trim(SparePartSku.article_number)) == func.lower(article_number),
                       func.lower(func.trim(SparePartSku.supplier)) == func.lower(supplier)))
    if exclude_id:
        clash_q = clash_q.filter(SparePartSku.id != exclude_id)
    return clash_q.first() is not None


def _sku_duplicate_response(part_id, race=False):
    """Shared friendly-duplicate response for BOTH the pre-check path and the
    IntegrityError race path, so the two converge on identical user-visible
    behavior (RE-SP-002)."""
    if race:
        current_app.logger.warning(
            'SKU save rejected by uq_spare_part_skus_normalized (race), part=%s',
            part_id)
    _spare_flash_errors(
        [_spare_t('Бу деталь учун бундай SKU аллақачон мавжуд',
                  'Такой SKU уже существует для этой детали')],
        title_uz='SKU сақланмади:', title_ru='SKU не сохранён:')
    return redirect(url_for('spare_parts.skus'))


@spare_parts_bp.route('/skus/save', methods=['POST'])
@module_required('spare_parts')
def sku_save():
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('save SKU', entity_type='spare_part_sku')
    sku_id = request.form.get('id', type=int)
    spare_part_id = request.form.get('spare_part_id', type=int)
    brand = (request.form.get('brand', '') or '').strip()
    article_number = (request.form.get('article_number', '') or '').strip()
    supplier = (request.form.get('supplier', '') or '').strip()

    part = SparePart.query.get(spare_part_id) if spare_part_id else None
    errors = []
    if part is None or part.status != 'active' or part.is_active == False:  # noqa: E712
        errors.append(_spare_t('Каталогдан фаол эҳтиёт қисмни танланг',
                               'Выберите активную запчасть из каталога'))
    # A SKU with no brand, no article and no supplier identifies nothing.
    if not (brand or article_number or supplier):
        errors.append(_spare_t(
            'Камида битта майдонни тўлдиринг: бренд, артикул ёки таъминотчи',
            'Заполните хотя бы одно поле: бренд, артикул или поставщик'))
    if errors:
        _spare_flash_errors(errors, title_uz='SKU сақланмади:',
                            title_ru='SKU не сохранён:')
        return redirect(url_for('spare_parts.skus'))

    # Friendly pre-check for the common duplicate case (see
    # _sku_normalized_clash); the IntegrityError handler below is the safety
    # net for the race, not a replacement (RE-SP-002).
    will_be_active = (request.form.get('is_active') is not None) if sku_id else True
    if will_be_active and _sku_normalized_clash(part.id, brand, article_number,
                                                supplier, exclude_id=sku_id):
        return _sku_duplicate_response(part.id, race=False)

    # [REASON]: RE-SP-002 — the WHOLE insert/update including flush() must sit
    # inside one IntegrityError boundary. flush() is where SQLite actually
    # executes the INSERT/UPDATE, so under a true race (two saves pass the
    # pre-check before either flushes) uq_spare_part_skus_normalized raises at
    # flush(), not at commit(); a try around commit() alone lets that escape
    # as an HTTP 500 with a dirty session.
    try:
        created = False
        before = None
        if sku_id:
            sku = SparePartSku.query.get_or_404(sku_id)
            before = _sku_snapshot(sku)
            sku.spare_part_id = part.id
            sku.brand = brand
            sku.article_number = article_number
            sku.supplier = supplier
            sku.is_active = request.form.get('is_active') is not None
        else:
            sku = SparePartSku(spare_part_id=part.id, brand=brand,
                               article_number=article_number, supplier=supplier,
                               created_by=current_user.id)
            db.session.add(sku)
            created = True
        db.session.flush()
        after = _sku_snapshot(sku)
        _audit_spare(
            'spare_part_sku_created' if created else 'spare_part_sku_updated',
            entity_type='spare_part_sku',
            entity_id=sku.id,
            entity_label='{} — {}'.format(part.name, sku.label),
            before=before,
            after=after,
            changes=diff_dict(before, after),
            description='Spare part SKU saved'
        )
        db.session.commit()
    except IntegrityError:
        # [REASON]: SP-F-008/RE-SP-002 — the race safety net: rollback() clears
        # the half-flushed state so the session serves the next request
        # normally, and the user sees the exact same friendly duplicate
        # message as the pre-check path.
        db.session.rollback()
        return _sku_duplicate_response(part.id, race=True)
    flash(_spare_t('SKU сақланди', 'SKU сохранён'), 'success')
    return redirect(url_for('spare_parts.skus'))


# ─── SPARE-STAGE2: warehouses + inventory + movements ─────────────────────────

def _apply_inventory_movement(warehouse_id, sku_id, movement_type, quantity,
                              reference_type='manual', reference_id=None,
                              note='', user_id=None):
    """Apply one signed stock change to a (warehouse, SKU) pair WITH its audit row.

    [REASON]: SPARE-STAGE2 core invariant — spare_part_inventory.quantity is
    never written anywhere except through this function, which records the
    matching movement row (with a balance_after snapshot) in the same
    transaction. A reviewer can sum a pair's movement quantities and always
    get the pair's current quantity. Does not flush/commit beyond the flush
    needed for a lazily created inventory row — the caller owns the
    transaction, so a multi-item action (Task 5 issue) stays all-or-nothing.

    Sign rules per type: receipt > 0; issue/write_off < 0; adjustment any
    non-zero.

    [REASON]: SP-F-001 — the resulting balance must never go negative, and the
    guard must be ATOMIC: a single conditional UPDATE ... WHERE quantity +
    delta >= 0 (with RETURNING for the confirmed balance) instead of a
    read-then-write pair, so two concurrent issues can never both pass a
    Python-side check and drive the balance below zero. Deliberately NOT a
    CHECK constraint — SQLite cannot add one to an existing table without a
    full rebuild, unnecessary risk on the live WAL database. RETURNING is
    available: SQLite >= 3.35 (production Python 3.14 bundles far newer).

    Raises ValueError with a short reason code on invalid input;
    'insufficient_stock' means the movement would make the balance negative.
    """
    if movement_type not in INV_MOVEMENT_TYPES:
        raise ValueError('movement_type')
    try:
        quantity = float(quantity)
    except (TypeError, ValueError):
        raise ValueError('quantity')
    # [REASON]: SP-F-016 — NaN/inf would corrupt the running balance; same
    # error code as a malformed number.
    if not math.isfinite(quantity):
        raise ValueError('quantity')
    if quantity == 0:
        raise ValueError('zero_quantity')
    if movement_type == 'receipt' and quantity < 0:
        raise ValueError('receipt_negative')
    if movement_type in ('issue', 'write_off') and quantity > 0:
        raise ValueError('issue_positive')

    inv = SparePartInventory.query.filter_by(
        warehouse_id=warehouse_id, sku_id=sku_id).first()
    if inv is None:
        # [REASON]: inventory rows are created lazily on first movement — no
        # zero-quantity seeding of every SKU × warehouse combination.
        # [REASON]: SP-F-001 — a brand-new pair starts at 0, so a negative
        # first movement can never be satisfied; reject before inserting.
        if quantity < 0:
            raise ValueError('insufficient_stock')
        sku = SparePartSku.query.get(sku_id)
        unit = (sku.spare_part.unit if sku and sku.spare_part else '') or 'dona'
        new_qty = round(quantity, 3)
        inv = SparePartInventory(warehouse_id=warehouse_id, sku_id=sku_id,
                                 quantity=new_qty, unit=unit)
        db.session.add(inv)
        db.session.flush()
    else:
        inv_table = SparePartInventory.__table__
        result = db.session.execute(
            update(inv_table)
            .where(inv_table.c.id == inv.id,
                   inv_table.c.quantity + quantity >= 0)
            .values(quantity=func.round(inv_table.c.quantity + quantity, 3),
                    updated_at=datetime.utcnow())
            .returning(inv_table.c.quantity)
        )
        row = result.first()
        if row is None:
            # Zero rows updated: the WHERE guard rejected the movement.
            raise ValueError('insufficient_stock')
        new_qty = float(row[0])
        # The Core UPDATE bypassed the ORM — expire the in-session object so
        # any later read in this transaction sees the confirmed balance.
        db.session.expire(inv)
    movement = SparePartInventoryMovement(
        warehouse_id=warehouse_id,
        sku_id=sku_id,
        movement_type=movement_type,
        quantity=quantity,
        balance_after=new_qty,
        reference_type=reference_type,
        reference_id=reference_id,
        note=note or '',
        created_by=user_id,
    )
    db.session.add(movement)
    return movement


# ─── SP-RESERVE-003: reservations ─────────────────────────────────────────────

RESERVATION_STATUSES = ('active', 'consumed', 'released')


def _reserved_totals(warehouse_id, sku_ids, exclude_request_id=None):
    """Sum of ACTIVE reserved quantity per sku_id on one warehouse.

    [REASON]: SP-RESERVE-003 — one GROUP BY query, never a per-SKU loop;
    exclude_request_id removes that request's own claims so its own reserve
    never blocks its own issue.
    """
    sku_ids = [s for s in (sku_ids or []) if s]
    if not sku_ids:
        return {}
    q = (db.session.query(SparePartReservation.sku_id,
                          func.sum(SparePartReservation.quantity))
         .filter(SparePartReservation.warehouse_id == warehouse_id,
                 SparePartReservation.sku_id.in_(sku_ids),
                 SparePartReservation.status == 'active'))
    if exclude_request_id is not None:
        q = q.filter(SparePartReservation.request_id != exclude_request_id)
    return {sku_id: round(float(total or 0), 3)
            for sku_id, total in q.group_by(SparePartReservation.sku_id).all()}


def _availability_map(warehouse_id, sku_ids, exclude_request_id=None):
    """{sku_id: {'on_hand', 'reserved', 'available'}} for one warehouse.

    [REASON]: SP-RESERVE-003 — available = on-hand minus active reservations
    (of OTHER requests when exclude_request_id is given). Derived on read,
    never stored; on-hand itself stays untouched by reservations (a
    reservation is not a stock movement). Available may legitimately be
    negative: a manual negative adjustment wins over a standing reserve and
    is displayed, not auto-corrected. Two queries total.
    """
    sku_ids = [s for s in (sku_ids or []) if s]
    if not sku_ids:
        return {}
    reserved = _reserved_totals(warehouse_id, sku_ids,
                                exclude_request_id=exclude_request_id)
    inv_rows = (SparePartInventory.query
                .filter(SparePartInventory.warehouse_id == warehouse_id,
                        SparePartInventory.sku_id.in_(sku_ids))
                .all())
    on_hand = {r.sku_id: float(r.quantity or 0) for r in inv_rows}
    out = {}
    for sku_id in sku_ids:
        oh = on_hand.get(sku_id, 0.0)
        rv = reserved.get(sku_id, 0.0)
        out[sku_id] = {'on_hand': oh, 'reserved': rv,
                       'available': round(oh - rv, 3)}
    return out


def _create_reservations_for_request(spr, user_id):
    """Create one reservation row per SKU item of a freshly approved request.

    Returns None when the organization has no warehouse (nothing created —
    the caller distinguishes "no warehouse" from "no shortages"), else the
    list of shortage dicts {'item', 'needed', 'reserved', 'short'}.

    [REASON]: SP-RESERVE-003 — reservation is SOFT: approval is never blocked
    by stock ("approved" means the purchase is authorized; the part often is
    not in stock yet, and that is the normal case). Each item gets
    min(needed, remaining available) with a running per-SKU decrement so two
    items pointing at the same SKU can never both be granted the full stock;
    a row is written even when nothing could be reserved (quantity 0) so the
    shortage itself is recorded at decision time. Advisory only: the atomic
    guard inside _apply_inventory_movement remains the real enforcement at
    issue time. Two near-simultaneous approvals could in theory over-reserve;
    this single-Waitress low-concurrency app already accepts the same race
    class for act numbering, so deliberately no locking/retries here — the
    reservation is computed inside the same transaction as the status change
    (caller owns the commit; this helper only session.add()s).
    """
    warehouse = SparePartWarehouse.query.filter_by(
        organization_id=spr.organization_id).first()
    if warehouse is None:
        return None
    # [REASON]: items with sku_id NULL are outside the mechanism entirely —
    # no reservation row, no availability check (their explicit
    # confirm_no_sku gate at issue time is unchanged).
    sku_items = sorted((i for i in spr.items if i.sku_id), key=lambda i: i.id)
    if not sku_items:
        return []
    sku_ids = {i.sku_id for i in sku_items}
    availability = _availability_map(warehouse.id, sku_ids,
                                     exclude_request_id=spr.id)
    # Running grantable balance per SKU; a negative available grants nothing.
    remaining = {sku_id: max(0.0, availability[sku_id]['available'])
                 for sku_id in sku_ids}
    shortages = []
    for item in sku_items:
        needed = round(float(item.quantity or 0), 3)
        granted = round(max(0.0, min(needed, remaining.get(item.sku_id, 0.0))), 3)
        remaining[item.sku_id] = remaining.get(item.sku_id, 0.0) - granted
        db.session.add(SparePartReservation(
            request_id=spr.id,
            request_item_id=item.id,
            warehouse_id=warehouse.id,
            sku_id=item.sku_id,
            quantity=granted,
            requested_quantity=needed,
            status='active',
            created_by=user_id,
        ))
        short = round(needed - granted, 3)
        if short > 0:
            shortages.append({'item': item, 'needed': needed,
                              'reserved': granted, 'short': short})
    return shortages


def _consume_reservations_for_request(request_id, user_id):
    """Flip every ACTIVE reservation of a request to 'consumed'.

    [REASON]: SP-RESERVE-003 — called inside the issue transaction, after the
    stock movements succeed: the claim is fulfilled the moment the stock
    physically leaves. No commit here — a failed issue rolls the flip back
    together with everything else.
    """
    rows = (SparePartReservation.query
            .filter_by(request_id=request_id, status='active')
            .all())
    now = datetime.utcnow()
    for row in rows:
        row.status = 'consumed'
        row.closed_at = now
        row.closed_by = user_id
    return rows


def _release_reservations(reservation_ids, user_id, note=''):
    """Flip the given ACTIVE reservations to 'released'; returns affected rows.

    [REASON]: SP-RESERVE-003 — the manual escape hatch for an
    approved-but-abandoned request: 'approved' has no other exit (reject and
    the price routes only accept 'submitted'), so freeing the claim is a
    warehouse-screen action recording who/when/why. Rows not currently
    active are ignored, never rewritten. No commit — caller owns the
    transaction.
    """
    if not reservation_ids:
        return []
    rows = (SparePartReservation.query
            .filter(SparePartReservation.id.in_(reservation_ids),
                    SparePartReservation.status == 'active')
            .all())
    now = datetime.utcnow()
    for row in rows:
        row.status = 'released'
        row.closed_at = now
        row.closed_by = user_id
        row.close_note = (note or '')[:300]
    return rows


def _spare_orgs_for_user():
    """Organizations the current user may see, ordered like the rest of the module."""
    user_org_ids = _spare_user_org_ids()
    q = Organization.query.order_by(Organization.sort_order)
    if user_org_ids is not None:
        q = q.filter(Organization.id.in_(user_org_ids))
    return q.all()


@spare_parts_bp.route('/inventory')
@module_required('spare_parts')
def inventory():
    # [REASON]: SPARE-STAGE2 — new permission primitive for warehouse work,
    # same has_module_access()+admin-bypass pattern as the other five
    # spare-parts permissions. Seeded by migrate_spare_parts_stage2.py.
    if not current_user.has_module_access('spare_parts_inventory_manage'):
        abort(403)
    organizations = _spare_orgs_for_user()
    org_id = request.args.get('org_id', type=int)
    if org_id and not current_user.can_access_org(org_id):
        abort(403)
    if not org_id and organizations:
        org_id = organizations[0].id

    warehouse = (SparePartWarehouse.query.filter_by(organization_id=org_id).first()
                 if org_id else None)
    inventory_rows = []
    movements = []
    if warehouse:
        inventory_rows = (SparePartInventory.query
                          .options(joinedload(SparePartInventory.sku)
                                   .joinedload(SparePartSku.spare_part))
                          .filter_by(warehouse_id=warehouse.id)
                          .all())
        inventory_rows.sort(key=lambda r: ((r.sku.spare_part.name if r.sku and r.sku.spare_part else ''),
                                           (r.sku.label if r.sku else '')))
        movements = (SparePartInventoryMovement.query
                     .options(joinedload(SparePartInventoryMovement.sku)
                              .joinedload(SparePartSku.spare_part),
                              joinedload(SparePartInventoryMovement.creator))
                     .filter_by(warehouse_id=warehouse.id)
                     .order_by(SparePartInventoryMovement.id.desc())
                     .limit(30)
                     .all())
    # [REASON]: SPARE-STAGE2-QA-FIX2 — for 'request_item' movements the stored
    # reference_id is the SparePartRequestItem PK, NOT the request number. Resolve
    # each item PK to its parent request number so the journal's "Основание" column
    # shows the real request an auditor can look up (e.g. «Заявка #10»), instead of
    # the item's internal id which looks like a non-existent request.
    movement_request_no = {}
    request_item_ids = [m.reference_id for m in movements
                        if m.reference_type == 'request_item' and m.reference_id]
    if request_item_ids:
        item_to_req = dict(
            db.session.query(SparePartRequestItem.id,
                             SparePartRequestItem.request_id)
            .filter(SparePartRequestItem.id.in_(request_item_ids)).all())
        for m in movements:
            if m.reference_type == 'request_item' and m.reference_id in item_to_req:
                movement_request_no[m.id] = item_to_req[m.reference_id]
    # [REASON]: SP-RESERVE-003 — the balances table shows three numbers
    # (on-hand / reserved / available = on-hand − reserved); reserved totals
    # come from ONE grouped query over the warehouse's active reservations,
    # never a per-row loop. The active reservation list below the balances
    # is the manual-release surface ('approved' has no other exit).
    reserved_map = {}
    reservations = []
    if warehouse:
        reserved_rows = (db.session.query(SparePartReservation.sku_id,
                                          func.sum(SparePartReservation.quantity))
                         .filter(SparePartReservation.warehouse_id == warehouse.id,
                                 SparePartReservation.status == 'active')
                         .group_by(SparePartReservation.sku_id)
                         .all())
        reserved_map = {sku_id: round(float(total or 0), 3)
                        for sku_id, total in reserved_rows}
        reservations = (SparePartReservation.query
                        .options(joinedload(SparePartReservation.request),
                                 joinedload(SparePartReservation.item),
                                 joinedload(SparePartReservation.sku)
                                 .joinedload(SparePartSku.spare_part),
                                 joinedload(SparePartReservation.creator))
                        .filter(SparePartReservation.warehouse_id == warehouse.id,
                                SparePartReservation.status == 'active')
                        .order_by(SparePartReservation.created_at.desc(),
                                  SparePartReservation.id.desc())
                        .all())
    # SKU picker for the manual movement form: all active SKUs, labelled with
    # their canonical part name.
    sku_options = (SparePartSku.query
                   .options(joinedload(SparePartSku.spare_part))
                   .filter(db.or_(SparePartSku.is_active == True,  # noqa: E712
                                  SparePartSku.is_active.is_(None)))
                   .all())
    sku_options.sort(key=lambda s: ((s.spare_part.name if s.spare_part else ''), s.label))
    return render_template('spare_parts_inventory.html',
                           organizations=organizations,
                           org_id=org_id,
                           warehouse=warehouse,
                           inventory_rows=inventory_rows,
                           reserved_map=reserved_map,
                           reservations=reservations,
                           movements=movements,
                           movement_request_no=movement_request_no,
                           sku_options=sku_options,
                           movement_labels=INV_MOVEMENT_LABELS,
                           lang=_spare_lang())


@spare_parts_bp.route('/inventory/warehouse', methods=['POST'])
@module_required('spare_parts')
def inventory_warehouse_create():
    if not current_user.has_module_access('spare_parts_inventory_manage'):
        abort(403)
    org_id = request.form.get('organization_id', type=int)
    name = (request.form.get('name', '') or '').strip()
    if not org_id or not current_user.can_access_org(org_id):
        abort(403)
    org = Organization.query.get_or_404(org_id)
    if not name:
        name = org.name
    # [REASON]: the UNIQUE(organization_id) constraint is the real guard; this
    # pre-check only exists to show a friendly message instead of a 500.
    if SparePartWarehouse.query.filter_by(organization_id=org_id).first():
        _spare_flash_errors(
            [_spare_t('Бу ташкилот учун омбор аллақачон мавжуд',
                      'Склад для этой организации уже существует')],
            title_uz='Омбор яратилмади:', title_ru='Склад не создан:')
        return redirect(url_for('spare_parts.inventory', org_id=org_id))
    warehouse = SparePartWarehouse(organization_id=org_id, name=name[:200])
    db.session.add(warehouse)
    db.session.flush()
    _audit_spare(
        'spare_part_warehouse_created',
        entity_type='spare_part_warehouse',
        entity_id=warehouse.id,
        entity_label=warehouse.name,
        after={'id': warehouse.id, 'organization_id': org_id, 'name': warehouse.name},
        description='Spare part warehouse created'
    )
    db.session.commit()
    flash(_spare_t('Омбор яратилди', 'Склад создан'), 'success')
    return redirect(url_for('spare_parts.inventory', org_id=org_id))


@spare_parts_bp.route('/inventory/movement', methods=['POST'])
@module_required('spare_parts')
def inventory_movement_create():
    """Manual receipt / adjustment entry."""
    if not current_user.has_module_access('spare_parts_inventory_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('create inventory movement',
                    entity_type='spare_part_inventory_movement')
    warehouse_id = request.form.get('warehouse_id', type=int)
    warehouse = SparePartWarehouse.query.get_or_404(warehouse_id)
    if not current_user.can_access_org(warehouse.organization_id):
        abort(403)
    movement_type = (request.form.get('movement_type', '') or '').strip()
    # Manual entry allows receipts and adjustments only; issues and write-offs
    # come from the request issue workflow (Task 5), never typed by hand.
    if movement_type not in ('receipt', 'adjustment'):
        abort(400)
    sku_id = request.form.get('sku_id', type=int)
    sku = SparePartSku.query.get(sku_id) if sku_id else None
    note = (request.form.get('note', '') or '').strip()
    raw_qty = (request.form.get('quantity', '') or '').strip().replace(',', '.')

    errors = []
    if sku is None or sku.is_active == False:  # noqa: E712
        errors.append(_spare_t('SKU танланг', 'Выберите SKU'))
    quantity = None
    try:
        parsed = float(raw_qty)
        # [REASON]: SP-F-016 — 'nan'/'inf' rejected with the same message as
        # a malformed number.
        if not math.isfinite(parsed):
            raise ValueError('quantity')
        quantity = parsed
    except (TypeError, ValueError):
        errors.append(_spare_t('Миқдор сон бўлиши керак', 'Количество должно быть числом'))
    if quantity is not None:
        if quantity == 0:
            errors.append(_spare_t('Миқдор нолдан фарқли бўлиши керак',
                                   'Количество не может быть нулём'))
        elif movement_type == 'receipt' and quantity < 0:
            errors.append(_spare_t('Кирим миқдори мусбат бўлиши керак',
                                   'Количество прихода должно быть положительным'))
    if errors:
        _spare_flash_errors(errors, title_uz='Ҳаракат сақланмади:',
                            title_ru='Движение не сохранено:')
        return redirect(url_for('spare_parts.inventory',
                                org_id=warehouse.organization_id))

    # [REASON]: SP-F-001 pre-flight — a negative manual adjustment may not
    # drive the balance below zero; friendly message instead of relying on the
    # atomic guard's exception for the ordinary (non-race) case.
    if movement_type == 'adjustment' and quantity < 0:
        inv = SparePartInventory.query.filter_by(
            warehouse_id=warehouse.id, sku_id=sku.id).first()
        current_qty = float(inv.quantity or 0) if inv else 0.0
        if current_qty + quantity < 0:
            _spare_flash_errors(
                [_spare_t('Омборда етарли қолдиқ йўқ (жорий қолдиқ: {})',
                          'На складе недостаточно остатка (текущий остаток: {})'
                          ).format(_fmt_qty_text(current_qty))],
                title_uz='Ҳаракат сақланмади:', title_ru='Движение не сохранено:')
            return redirect(url_for('spare_parts.inventory',
                                    org_id=warehouse.organization_id))

    try:
        movement = _apply_inventory_movement(
            warehouse.id, sku.id, movement_type, quantity,
            reference_type='manual', note=note, user_id=current_user.id)
    except ValueError as exc:
        # [REASON]: SP-F-001 — race with another manual entry between the
        # pre-flight above and the atomic UPDATE guard; a clean bilingual
        # message instead of an unhandled 500.
        db.session.rollback()
        if str(exc) == 'insufficient_stock':
            msg = _spare_t('Қолдиқ ўзгарди — саҳифани янгилаб, қайта уриниб кўринг',
                           'Остаток изменился — обновите страницу и повторите')
        else:
            msg = _spare_t('Ҳаракат маълумотлари нотўғри',
                           'Некорректные данные движения')
        _spare_flash_errors([msg], title_uz='Ҳаракат сақланмади:',
                            title_ru='Движение не сохранено:')
        return redirect(url_for('spare_parts.inventory',
                                org_id=warehouse.organization_id))
    _audit_spare(
        'spare_part_inventory_movement_created',
        entity_type='spare_part_inventory_movement',
        entity_id=None,
        entity_label='{} {} {}'.format(movement_type, quantity, sku.label),
        after={'warehouse_id': warehouse.id, 'sku_id': sku.id,
               'movement_type': movement_type, 'quantity': quantity,
               'balance_after': movement.balance_after},
        description='Manual inventory movement'
    )
    db.session.commit()
    flash(_spare_t('Ҳаракат сақланди. Янги қолдиқ: {}',
                   'Движение сохранено. Новый остаток: {}').format(movement.balance_after),
          'success')
    return redirect(url_for('spare_parts.inventory',
                            org_id=warehouse.organization_id))


@spare_parts_bp.route('/inventory/reservations/<int:res_id>/release', methods=['POST'])
@module_required('spare_parts')
def inventory_reservation_release(res_id):
    """Manually release one active reservation from the warehouse screen.

    [REASON]: SP-RESERVE-003 — the only way to free the claim of an
    approved-but-abandoned request: 'approved' has no other exit (reject and
    the price routes only accept 'submitted'). Does NOT change the request's
    status — only the claim is dropped, recording who/when/why.
    """
    # [REASON]: same permission primitive as the rest of the warehouse screen
    # (no new permission codes), with the CYCLE-2-3 Part 8 denial trail.
    if not current_user.has_module_access('spare_parts_inventory_manage'):
        _deny_spare('release reservation #{}'.format(res_id),
                    entity_type='spare_part_reservation', entity_id=res_id,
                    entity_label='Reservation #{}'.format(res_id))
    res = SparePartReservation.query.get_or_404(res_id)
    if res.status != 'active':
        abort(400)
    warehouse = res.warehouse
    _spare_check_org_access(warehouse.organization_id if warehouse else None)
    _release_reservations([res.id], current_user.id,
                          note=request.form.get('note', ''))
    _audit_spare(
        'spare_part_reservation_released',
        entity_type='spare_part_reservation',
        entity_id=res.id,
        entity_label='Reservation #{} (request #{})'.format(res.id, res.request_id),
        after={'request_id': res.request_id,
               'request_item_id': res.request_item_id,
               'warehouse_id': res.warehouse_id,
               'sku_id': res.sku_id,
               'quantity': res.quantity,
               'requested_quantity': res.requested_quantity,
               'note': res.close_note},
        description='Reservation released from warehouse screen'
    )
    db.session.commit()
    flash(_spare_t('Захира бўшатилди. Сўров ҳолати ўзгармади.',
                   'Резерв снят. Статус заявки не изменился.'), 'success')
    return redirect(url_for('spare_parts.inventory',
                            org_id=warehouse.organization_id))


# ─── SPARE-STAGE2: issue + write-off acts ─────────────────────────────────────

def _acts_dir():
    """Absolute path of the write-off act PDF folder (created on demand).

    Sibling of the photo UPLOAD_FOLDER, same resolution rule: a plain path
    relative to the app folder so non-Flask processes can find the files.
    """
    folder = os.path.join('instance', 'uploads', 'spare_parts_acts')
    root = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(root, folder)
    os.makedirs(folder, exist_ok=True)
    return folder


def _generate_act_number(year):
    """Next SPW-{year}-{00001} act number.

    [REASON]: SPARE-STAGE2 — exact reuse of the MAX+1-inside-a-transaction
    technique proven by work_orders._generate_wo_number: act_number is UNIQUE,
    so a rare concurrent race raises IntegrityError on commit instead of
    silently duplicating a number; acceptable for this low-concurrency app.
    """
    prefix = 'SPW-{}-'.format(year)
    last = db.session.query(func.max(SparePartWriteOffAct.act_number)).filter(
        SparePartWriteOffAct.act_number.like(prefix + '%')
    ).scalar()
    seq = 1
    if last:
        try:
            seq = int(last.split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return '{}{:05d}'.format(prefix, seq)


def _issue_context(spr):
    """Data the detail page needs to offer the issue action.

    Returns dict with: sku_items / no_sku_items, the organization's warehouse
    (or None), current stock per SKU item so the issuer sees what will be
    deducted (and what would go negative) BEFORE confirming, plus
    SP-RESERVE-003 keys: 'availability' ({sku_id: {'on_hand', 'reserved',
    'available', 'own_reserved'}}, where 'reserved'/'available' count OTHER
    requests' active reservations only) and 'needed_by_sku' (this request's
    quantity aggregated per SKU across items).
    """
    sku_items = [i for i in spr.items if i.sku_id]
    no_sku_items = [i for i in spr.items if not i.sku_id]
    warehouse = SparePartWarehouse.query.filter_by(
        organization_id=spr.organization_id).first()
    stock = {}
    if warehouse and sku_items:
        rows = (SparePartInventory.query
                .filter(SparePartInventory.warehouse_id == warehouse.id,
                        SparePartInventory.sku_id.in_([i.sku_id for i in sku_items]))
                .all())
        stock = {r.sku_id: r.quantity for r in rows}
    # [REASON]: SP-RESERVE-003 — per-SKU aggregation (two items pointing at
    # the same SKU are checked and displayed against their combined need),
    # in first-appearance order so pre-flight messages stay deterministic.
    needed_by_sku = {}
    for item in sku_items:
        needed_by_sku[item.sku_id] = round(
            needed_by_sku.get(item.sku_id, 0.0) + float(item.quantity or 0), 3)
    availability = {}
    if warehouse and sku_items:
        # [REASON]: SP-RESERVE-003 — exclude_request_id makes 'reserved' and
        # 'available' count OTHER requests only, so a request's own standing
        # reserve never blocks its own issue; 'own_reserved' is surfaced
        # separately for display and the under-reserved hint.
        availability = _availability_map(warehouse.id, list(needed_by_sku),
                                         exclude_request_id=spr.id)
        own_rows = (db.session.query(SparePartReservation.sku_id,
                                     func.sum(SparePartReservation.quantity))
                    .filter(SparePartReservation.request_id == spr.id,
                            SparePartReservation.warehouse_id == warehouse.id,
                            SparePartReservation.status == 'active')
                    .group_by(SparePartReservation.sku_id).all())
        own = {sku_id: round(float(total or 0), 3) for sku_id, total in own_rows}
        for sku_id, info in availability.items():
            info['own_reserved'] = own.get(sku_id, 0.0)
    return {
        'sku_items': sku_items,
        'no_sku_items': no_sku_items,
        'warehouse': warehouse,
        'stock': stock,
        'availability': availability,
        'needed_by_sku': needed_by_sku,
    }


@spare_parts_bp.route('/<int:rid>/issue', methods=['POST'])
@module_required('spare_parts')
def issue_request(rid):
    """approved -> issued: deduct SKU stock, create the write-off act + PDF.

    All-or-nothing: every deduction, the act row, its item snapshots and the
    PDF render happen in ONE transaction — any failure rolls the whole issue
    back and the request stays 'approved'.
    """
    # [REASON]: SPARE-STAGE2 — separate permission from spare_parts_approve,
    # matching the real-world split between "approve the purchase" and
    # "physically hand it over from the warehouse".
    if not current_user.has_module_access('spare_parts_issue'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('issue request #{}'.format(rid),
                    entity_type='spare_part_request', entity_id=rid,
                    entity_label='Request #{}'.format(rid))
    spr = SparePartRequest.query.get_or_404(rid)
    _spare_check_request_access(spr)
    # [REASON]: 'issued' is reachable ONLY from 'approved' — never from any
    # other status, and never twice (no un-issue; acts are permanent).
    if spr.status != 'approved':
        abort(400)

    ctx = _issue_context(spr)
    errors = []
    if ctx['sku_items'] and ctx['warehouse'] is None:
        errors.append(_spare_t(
            'Ташкилотда эҳтиёт қисмлар омбори йўқ. Аввал «Омбор» экранида омбор яратинг.',
            'У организации нет склада запчастей. Сначала создайте склад на экране «Склад».'))
    # [REASON]: SPARE-STAGE2 — untracked (no-SKU) items must be issued
    # CONSCIOUSLY: the form shows which items skip stock deduction and the
    # server independently requires the explicit confirmation flag, so the
    # rule holds even if the client-side UI is bypassed.
    if ctx['no_sku_items'] and request.form.get('confirm_no_sku') != '1':
        errors.append(_spare_t(
            'SKUсиз позициялар омбордан ҳисобдан чиқарилмайди — беришдан олдин буни тасдиқлаш шарт',
            'Позиции без SKU не будут списаны со склада — перед выдачей это нужно явно подтвердить'))
    # [REASON]: SP-F-001 pre-flight — a friendly shortage message BEFORE any
    # transaction work. SP-RESERVE-003: aggregated per SKU (two items on the
    # same SKU are checked against their combined need) and against
    # AVAILABILITY = on-hand minus OTHER requests' active reservations, so a
    # part already promised to another approved request cannot be issued past
    # it — while this request's own reserve never blocks its own issue. The
    # atomic guard inside _apply_inventory_movement stays the real
    # enforcement for the concurrent case below.
    if ctx['sku_items'] and ctx['warehouse'] is not None:
        for sku_id, needed in ctx['needed_by_sku'].items():
            info = ctx['availability'].get(sku_id) or {
                'on_hand': 0.0, 'reserved': 0.0, 'available': 0.0,
                'own_reserved': 0.0}
            available = float(info['available'])
            if available - needed < 0:
                item_name = next((i.name for i in ctx['sku_items']
                                  if i.sku_id == sku_id), '')
                msg = _spare_t(
                    '«{}»: омборда етарли қолдиқ йўқ (керак: {}, мавжуд: {})',
                    '«{}»: на складе недостаточно остатка (нужно: {}, доступно: {})'
                ).format(item_name, _fmt_qty_text(needed), _fmt_qty_text(available))
                # [REASON]: SP-RESERVE-003 — when the request itself is
                # under-reserved, say so explicitly: the issuer must see the
                # part was never fully promised to THIS request, as opposed
                # to stock having drained afterwards.
                if float(info['own_reserved']) + 0.001 < needed:
                    msg += _spare_t(
                        ' — сўров захираси кераклидан кам (захираланган: {})',
                        ' — резерв заявки меньше требуемого (зарезервировано: {})'
                    ).format(_fmt_qty_text(info['own_reserved']))
                errors.append(msg)
    if errors:
        _spare_flash_errors(errors, title_uz='Бериб бўлмади:',
                            title_ru='Выдача не выполнена:')
        return redirect(url_for('spare_parts.detail', rid=rid))

    warehouse = ctx['warehouse']
    act_pdf_full_path = None

    def _rollback_failed_issue():
        # Shared failure path: roll everything back and remove the orphan PDF
        # if it was written before the transaction died.
        db.session.rollback()
        if act_pdf_full_path and os.path.exists(act_pdf_full_path):
            try:
                os.remove(act_pdf_full_path)
            except OSError:
                pass

    try:
        # 1. Inventory movements for every SKU item (negative issue quantities).
        for item in ctx['sku_items']:
            _apply_inventory_movement(
                warehouse.id, item.sku_id, 'issue', -float(item.quantity or 0),
                reference_type='request_item', reference_id=item.id,
                note='Request #{}'.format(spr.id), user_id=current_user.id)

        # [REASON]: SP-RESERVE-003 — the claim is fulfilled the moment the
        # stock physically leaves: consume this request's active reservations
        # INSIDE the same transaction, so a failed issue (rollback below)
        # leaves them active and nothing is partially applied.
        _consume_reservations_for_request(spr.id, current_user.id)

        # 2. The act with MAX+1 numbering and snapshotted items.
        issued_date = date.today()
        act = SparePartWriteOffAct(
            act_number=_generate_act_number(issued_date.year),
            request_id=spr.id,
            organization_id=spr.organization_id,
            warehouse_id=warehouse.id if warehouse else None,
            issued_date=issued_date,
            issued_by=current_user.id,
        )
        db.session.add(act)
        db.session.flush()
        for item in spr.items:
            total = None
            if item.price is not None:
                total = round(float(item.price) * float(item.quantity or 0), 2)
            db.session.add(SparePartWriteOffActItem(
                act_id=act.id,
                request_item_id=item.id,
                name=item.name,
                sku_label=(item.sku.label if item.sku else ''),
                quantity=item.quantity,
                unit=item.unit or 'dona',
                price=item.price,
                total=total,
            ))
        db.session.flush()

        # 3. Status transition + history + audit.
        old_status = spr.status
        spr.status = 'issued'
        _add_status_history(spr.id, old_status, 'issued',
                            comment='Act {}'.format(act.act_number),
                            changed_by=current_user.id)
        _audit_spare(
            'spare_part_request_status_changed',
            entity_type='spare_part_request',
            entity_id=spr.id,
            entity_label='Request #{}'.format(spr.id),
            changes={'status': {'before': old_status, 'after': 'issued'},
                     'act_number': {'before': None, 'after': act.act_number}},
            description='Request issued; write-off act {} created'.format(act.act_number)
        )
        # [REASON]: SP-F-017 — the act is its own auditable entity (a numbered
        # accounting document), so its creation gets a DISTINCT audit event in
        # the same transaction, not just an act_number embedded in the request
        # status change: an auditor filtering by entity/action must find every
        # act without parsing another event's changes payload.
        _audit_spare(
            'spare_part_write_off_act_created',
            entity_type='spare_part_write_off_act',
            entity_id=act.id,
            entity_label=act.act_number,
            after={'id': act.id, 'act_number': act.act_number,
                   'request_id': spr.id, 'organization_id': spr.organization_id,
                   'warehouse_id': act.warehouse_id,
                   'issued_date': _date_iso(act.issued_date),
                   'items_count': len(spr.items)},
            description='Write-off act {} created for request #{}'.format(
                act.act_number, spr.id)
        )

        # 4. Render the PDF inside the same transaction: a request must never
        # end up 'issued' without its printable act.
        from spare_parts_pdf import generate_write_off_act_pdf
        fname = '{}_{}.pdf'.format(act.id, act.act_number)
        act_pdf_full_path = os.path.join(_acts_dir(), fname)
        # [REASON]: RE-SP-010 — localized unit words in the printed act.
        generate_write_off_act_pdf(act, act_pdf_full_path, lang=_spare_lang(),
                                   unit_labels=_unit_labels_map())
        act.pdf_path = fname

        db.session.commit()
    except ValueError as exc:
        # [REASON]: SP-F-001 — the true-concurrency path: another transaction
        # consumed the stock between the pre-flight above and the atomic
        # UPDATE guard. Distinct message so the issuer knows to reload, not
        # to report a system error. Any other ValueError code is unexpected
        # here and gets the generic treatment.
        _rollback_failed_issue()
        if str(exc) == 'insufficient_stock':
            current_app.logger.warning(
                'Issue rejected for request %s: stock changed concurrently', rid)
            _spare_flash_errors(
                [_spare_t('Қолдиқ ўзгарди — саҳифани янгилаб, қайта уриниб кўринг',
                          'Остаток изменился — обновите страницу и повторите')],
                title_uz='Бериб бўлмади:', title_ru='Выдача не выполнена:')
        else:
            current_app.logger.exception('Issue action failed for request %s', rid)
            _spare_flash_errors(
                [_spare_t('Ички хатолик — ҳеч нарса ўзгартирилмади',
                          'Внутренняя ошибка — ничего не изменено')],
                title_uz='Бериб бўлмади:', title_ru='Выдача не выполнена:')
        return redirect(url_for('spare_parts.detail', rid=rid))
    except Exception:
        _rollback_failed_issue()
        current_app.logger.exception('Issue action failed for request %s', rid)
        _spare_flash_errors(
            [_spare_t('Ички хатолик — ҳеч нарса ўзгартирилмади',
                      'Внутренняя ошибка — ничего не изменено')],
            title_uz='Бериб бўлмади:', title_ru='Выдача не выполнена:')
        return redirect(url_for('spare_parts.detail', rid=rid))

    flash(_spare_t('Сўров берилди. Далолатнома: {}',
                   'Заявка выдана. Акт: {}').format(act.act_number), 'success')
    # [REASON]: deliberately NO bot003 notification here — the bot's event
    # allowlist (out of this task's scope fence) has no issued event type, so
    # a call would only produce "Unknown event_type" log noise on every issue.
    return redirect(url_for('spare_parts.act_detail', act_id=act.id))


@spare_parts_bp.route('/acts/<int:act_id>')
@module_required('spare_parts')
def act_detail(act_id):
    # [REASON]: CYCLE-2-3-HOTFIX F-002 — same eager-loaded chain as act_pdf
    # (see comment there); this page renders the same per-line item list.
    act = (SparePartWriteOffAct.query
           .options(selectinload(SparePartWriteOffAct.items)
                    .joinedload(SparePartWriteOffActItem.request_item)
                    .joinedload(SparePartRequestItem.spare_part))
           .filter(SparePartWriteOffAct.id == act_id)
           .first_or_404())
    # Same org-scoped access rule as the request it belongs to.
    spr = act.request
    if spr is None:
        abort(404)
    _spare_check_request_access(spr)
    # [REASON]: SP-F-002 — acts carry prices/quantities and were readable by
    # any base spare_parts user; viewing them now needs the explicit
    # spare_parts_acts permission (admin bypass via has_module_access, and
    # migrate_spare_parts_acts_permission.py auto-grants it to existing
    # issue/approve holders so no current workflow user loses access).
    if not current_user.has_module_access('spare_parts_acts'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('view act #{}'.format(act_id),
                    entity_type='spare_part_write_off_act', entity_id=act_id,
                    entity_label=act.act_number)
    return render_template('spare_part_act.html',
                           act=act,
                           req=spr,
                           lang=_spare_lang())


@spare_parts_bp.route('/acts/<int:act_id>/pdf')
@module_required('spare_parts')
def act_pdf(act_id):
    # [REASON]: CYCLE-2-3-HOTFIX F-002 — since Part 7 the PDF renderer walks
    # item.request_item.spare_part per act line to resolve the Uzbek alias
    # (name_uz); a bare get_or_404 made that N+1 (one query per line). Eager-
    # load the whole rendering chain in a fixed number of queries. Rendered
    # output is unchanged — only the query pattern.
    act = (SparePartWriteOffAct.query
           .options(selectinload(SparePartWriteOffAct.items)
                    .joinedload(SparePartWriteOffActItem.request_item)
                    .joinedload(SparePartRequestItem.spare_part))
           .filter(SparePartWriteOffAct.id == act_id)
           .first_or_404())
    spr = act.request
    if spr is None:
        abort(404)
    # [REASON]: act PDFs are NOT served as raw static files; the same
    # org-scoped access check as the act/request detail applies to downloads.
    _spare_check_request_access(spr)
    # [REASON]: SP-F-002 — same explicit permission as act_detail (this is the
    # same document as a file).
    if not current_user.has_module_access('spare_parts_acts'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('download act PDF #{}'.format(act_id),
                    entity_type='spare_part_write_off_act', entity_id=act_id,
                    entity_label=act.act_number)
    # [REASON]: SPARE-STAGE2-QA-FIX1 — the act RECORD (number, items, prices,
    # organization, permanence/no-un-issue) is frozen at issue time and must
    # never change. Only the PRINTED rendering is regenerated fresh here, in the
    # language the viewer currently wants — mirroring how the Excel report export
    # already follows the current UI language instead of the language at some past
    # event. An optional ?lang=ru|uz overrides the UI default so an act issued in
    # one language can be printed in the other without touching the global toggle.
    # The archival copy written at issue time to act.pdf_path is left untouched.
    lang = request.args.get('lang')
    if lang not in ('ru', 'uz'):
        lang = _spare_lang()
    from spare_parts_pdf import generate_write_off_act_pdf
    buffer = io.BytesIO()
    # [REASON]: RE-SP-010 — unit labels follow the same explicit-lang rule as
    # the rest of this regenerated rendering (?lang= override included).
    generate_write_off_act_pdf(act, buffer, lang=lang,
                               unit_labels=_unit_labels_map(lang))
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     download_name='{}.pdf'.format(act.act_number))


# ─── CYCLE-2-3 Part 5: acts index ─────────────────────────────────────────────

@spare_parts_bp.route('/acts')
@module_required('spare_parts')
def acts_index():
    """Standalone browsable list of write-off acts.

    [REASON]: CYCLE-2-3 Part 5 — until now acts were reachable ONLY from
    their request's detail page; an operator with the spare_parts_acts
    permission had no way to browse issue history without hunting through
    individual requests. Purely additive: gated exactly like act_detail/
    act_pdf (base module + explicit spare_parts_acts permission + org
    scoping), no existing route changes.
    """
    if not current_user.has_module_access('spare_parts_acts'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('browse acts index', entity_type='spare_part_write_off_act')

    org_id = request.args.get('org_id', type=int)
    date_from_s = request.args.get('date_from', '')
    date_to_s = request.args.get('date_to', '')

    q = (SparePartWriteOffAct.query
         .options(joinedload(SparePartWriteOffAct.organization),
                  joinedload(SparePartWriteOffAct.issuer),
                  joinedload(SparePartWriteOffAct.request)
                  .joinedload(SparePartRequest.equipment),
                  selectinload(SparePartWriteOffAct.items)))
    # Same org scoping rule as the request list: non-admins see only the
    # organizations assigned to them.
    user_org_ids = _spare_user_org_ids()
    if user_org_ids is not None:
        q = q.filter(SparePartWriteOffAct.organization_id.in_(user_org_ids))
    if org_id:
        if not current_user.can_access_org(org_id):
            abort(403)
        q = q.filter(SparePartWriteOffAct.organization_id == org_id)
    # Lenient date filters, same style as index(): a malformed value is
    # ignored rather than failing the page.
    if date_from_s:
        try:
            q = q.filter(SparePartWriteOffAct.issued_date
                         >= datetime.strptime(date_from_s, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_s:
        try:
            q = q.filter(SparePartWriteOffAct.issued_date
                         <= datetime.strptime(date_to_s, '%Y-%m-%d').date())
        except ValueError:
            pass
    acts = q.order_by(SparePartWriteOffAct.issued_date.desc(),
                      SparePartWriteOffAct.id.desc()).all()

    # Per-act grand total over snapshotted item totals (None = no confirmed
    # price at issue time), same summation rule as spare_part_act.html.
    act_totals = {
        act.id: sum(float(i.total) for i in act.items if i.total is not None)
        for act in acts
    }

    if user_org_ids is None:
        organizations = Organization.query.order_by(Organization.sort_order).all()
    else:
        organizations = (Organization.query
                         .filter(Organization.id.in_(user_org_ids))
                         .order_by(Organization.sort_order).all())
    return render_template('spare_parts_acts.html',
                           acts=acts,
                           act_totals=act_totals,
                           org_id=org_id,
                           date_from_s=date_from_s,
                           date_to_s=date_to_s,
                           organizations=organizations,
                           lang=_spare_lang())


# ─── Catalog ──────────────────────────────────────────────────────────────────

@spare_parts_bp.route('/catalog')
@module_required('spare_parts')
def catalog():
    # [REASON]: SP-F-015 — the GET page splits read from manage: any base
    # spare_parts user may BROWSE the canonical catalog (they already see the
    # same names through the request-form picker), while the pending-review
    # queue, add/edit forms and category management stay behind
    # spare_parts_catalog_manage (the POST routes below keep their own
    # independent permission checks — only the read path changes).
    can_manage = current_user.has_module_access('spare_parts_catalog_manage')
    parts = (SparePart.query
             .options(joinedload(SparePart.category_ref))
             .order_by(SparePart.name).all())
    pending_parts = []
    if can_manage:
        # No reason to query data the viewer cannot act on.
        pending_parts = (SparePart.query
                         .options(joinedload(SparePart.creator),
                                  joinedload(SparePart.source_item))
                         .filter(SparePart.status == 'pending_review')
                         .order_by(SparePart.created_at)
                         .all())
    merge_targets = [p for p in parts
                     if p.status == 'active' and p.is_active != False]  # noqa: E712
    categories = (SparePartCategory.query
                  .order_by(SparePartCategory.sort_order, SparePartCategory.id)
                  .all())
    return render_template('spare_parts_catalog.html',
                           parts=parts,
                           pending_parts=pending_parts,
                           merge_targets=merge_targets,
                           categories=categories,
                           can_manage=can_manage,
                           # [REASON]: SP-F-024 — unit picker for the manage
                           # form; empty list keeps the legacy free-text input.
                           units=_active_units(),
                           lang=_spare_lang())


@spare_parts_bp.route('/catalog/save', methods=['POST'])
@module_required('spare_parts')
def catalog_save():
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('save catalog part', entity_type='spare_part_catalog')
    pid = request.form.get('id', type=int)
    name = request.form.get('name', '').strip()
    # [REASON]: CYCLE-2-3 Part 7 — optional Uzbek alias; empty means "not
    # translated yet" and is stored as NULL so displays fall back to name.
    name_uz = (request.form.get('name_uz', '') or '').strip() or None
    part_number = request.form.get('part_number', '').strip()
    unit = request.form.get('unit', 'dona').strip()
    # [REASON]: SPARE-STAGE1 — the form now sends category_id (managed
    # categories); the deprecated free-text `category` column is left untouched.
    category_id = request.form.get('category_id', type=int) or None
    if category_id and not SparePartCategory.query.get(category_id):
        category_id = None

    if not name:
        flash(_spare_t('Номини киритинг', 'Введите название'), 'warning')
        return redirect(url_for('spare_parts.catalog'))

    # [REASON]: SP-F-024 — same strict unit-directory rule as save_request
    # (empty directory = legacy free-text fallback, no silent coercion).
    valid_unit_codes = {u.code for u in _active_units()}
    if valid_unit_codes and (unit or 'dona') not in valid_unit_codes:
        _spare_flash_errors(
            [_spare_t('Ўлчов бирлигини рўйхатдан танланг',
                      'Выберите единицу измерения из справочника')],
            title_uz='Сақланмади:', title_ru='Не сохранено:')
        return redirect(url_for('spare_parts.catalog'))

    created = False
    before = None
    if pid:
        part = SparePart.query.get_or_404(pid)
        before = _catalog_snapshot(part)
        part.name = name
        part.name_uz = name_uz
        part.part_number = part_number
        part.unit = unit
        part.category_id = category_id
    else:
        part = SparePart(name=name, name_uz=name_uz, part_number=part_number,
                         unit=unit, category_id=category_id,
                         created_by=current_user.id)
        db.session.add(part)
        created = True
    db.session.flush()
    after = _catalog_snapshot(part)
    _audit_spare(
        'spare_part_catalog_created' if created else 'spare_part_catalog_updated',
        entity_type='spare_part_catalog',
        entity_id=part.id,
        entity_label=part.name,
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Spare part catalog saved'
    )
    db.session.commit()
    flash(_spare_t('Сақланди', 'Сохранено'), 'success')
    return redirect(url_for('spare_parts.catalog'))


@spare_parts_bp.route('/catalog/<int:pid>/name-uz', methods=['POST'])
@module_required('spare_parts')
def catalog_name_uz_save(pid):
    """Inline per-row edit of a part's Uzbek alias (CYCLE-2-3 Part 7).

    [REASON]: the owner fills translations in over time; a dedicated
    single-field route lets the catalog table offer one text input per row
    without round-tripping every other field through the main edit form.
    Same spare_parts_catalog_manage gate as every other catalog mutation.
    """
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('save catalog part Uzbek alias #{}'.format(pid),
                    entity_type='spare_part_catalog', entity_id=pid)
    part = SparePart.query.get_or_404(pid)
    before = _catalog_snapshot(part)
    part.name_uz = (request.form.get('name_uz', '') or '').strip() or None
    after = _catalog_snapshot(part)
    _audit_spare(
        'spare_part_catalog_updated',
        entity_type='spare_part_catalog',
        entity_id=part.id,
        entity_label=part.name,
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Spare part Uzbek alias saved'
    )
    db.session.commit()
    flash(_spare_t('Сақланди', 'Сохранено'), 'success')
    return redirect(url_for('spare_parts.catalog'))


@spare_parts_bp.route('/catalog/categories/save', methods=['POST'])
@module_required('spare_parts')
def catalog_category_save():
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('save catalog category', entity_type='spare_part_category')
    cid = request.form.get('id', type=int)
    name_ru = (request.form.get('name_ru', '') or '').strip()
    name_uz = (request.form.get('name_uz', '') or '').strip()
    kind = (request.form.get('kind', 'unit') or 'unit').strip()
    sort_order = request.form.get('sort_order', type=int) or 0
    if kind not in ('unit', 'consumable'):
        abort(400)
    if not name_ru or not name_uz:
        _spare_flash_errors(
            [_spare_t('Категория номини иккала тилда киритинг',
                      'Введите название категории на обоих языках')],
            title_uz='Категория сақланмади:', title_ru='Категория не сохранена:')
        return redirect(url_for('spare_parts.catalog'))

    created = False
    if cid:
        cat = SparePartCategory.query.get_or_404(cid)
        cat.name_ru = name_ru
        cat.name_uz = name_uz
        cat.kind = kind
        cat.sort_order = sort_order
        cat.is_active = request.form.get('is_active') is not None
    else:
        cat = SparePartCategory(name_ru=name_ru, name_uz=name_uz, kind=kind,
                                sort_order=sort_order, created_by=current_user.id)
        db.session.add(cat)
        created = True
    db.session.flush()
    _audit_spare(
        'spare_part_category_created' if created else 'spare_part_category_updated',
        entity_type='spare_part_category',
        entity_id=cat.id,
        entity_label=cat.name_ru,
        description='Spare part category saved (kind={})'.format(cat.kind)
    )
    db.session.commit()
    flash(_spare_t('Категория сақланди', 'Категория сохранена'), 'success')
    return redirect(url_for('spare_parts.catalog'))


@spare_parts_bp.route('/catalog/review/<int:pid>/approve', methods=['POST'])
@module_required('spare_parts')
def catalog_review_approve(pid):
    """Approve a pending_review candidate as a new canonical catalog entry."""
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('approve catalog candidate #{}'.format(pid),
                    entity_type='spare_part_catalog', entity_id=pid)
    part = SparePart.query.get_or_404(pid)
    if part.status != 'pending_review':
        abort(400)
    category_id = request.form.get('category_id', type=int)
    category = SparePartCategory.query.get(category_id) if category_id else None
    # [REASON]: SPARE-STAGE1 — approving as new requires a category so the
    # photo rule (unit/consumable) and analytics have a defined bucket.
    if not category:
        _spare_flash_errors(
            [_spare_t('Категорияни танланг', 'Выберите категорию')],
            title_uz='Ёзув тасдиқланмади:', title_ru='Запись не подтверждена:')
        return redirect(url_for('spare_parts.catalog'))
    before = _catalog_snapshot(part)
    part.category_id = category.id
    part.status = 'active'
    after = _catalog_snapshot(part)
    _audit_spare(
        'spare_part_catalog_review_approved',
        entity_type='spare_part_catalog',
        entity_id=part.id,
        entity_label=part.name,
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Pending-review candidate approved as new catalog entry'
    )
    db.session.commit()
    flash(_spare_t('Каталог ёзуви тасдиқланди', 'Запись каталога подтверждена'), 'success')
    return redirect(url_for('spare_parts.catalog'))


@spare_parts_bp.route('/catalog/review/<int:pid>/merge', methods=['POST'])
@module_required('spare_parts')
def catalog_review_merge(pid):
    """Merge a pending_review candidate into an existing canonical entry."""
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        # CYCLE-2-3 Part 8: forward-going denial trail (see _deny_spare).
        _deny_spare('merge catalog candidate #{}'.format(pid),
                    entity_type='spare_part_catalog', entity_id=pid)
    part = SparePart.query.get_or_404(pid)
    if part.status != 'pending_review':
        abort(400)
    target_id = request.form.get('target_id', type=int)
    target = SparePart.query.get(target_id) if target_id else None
    if not target or target.id == part.id or target.status != 'active':
        _spare_flash_errors(
            [_spare_t('Бирлаштириш учун фаол каталог ёзувини танланг',
                      'Выберите активную запись каталога для объединения')],
            title_uz='Бирлаштирилмади:', title_ru='Объединение не выполнено:')
        return redirect(url_for('spare_parts.catalog'))
    before = _catalog_snapshot(part)
    # [REASON]: SPARE-STAGE1 — repoint every request item that referenced the
    # duplicate candidate to the canonical entry so history and the repeat
    # engine count them against one part. Bulk update runs before the part is
    # modified so no in-session state needs re-synchronisation.
    repointed = (SparePartRequestItem.query
                 .filter(SparePartRequestItem.spare_part_id == part.id)
                 .update({'spare_part_id': target.id}, synchronize_session=False))
    part.status = 'merged'
    part.merged_into_id = target.id
    after = _catalog_snapshot(part)
    _audit_spare(
        'spare_part_catalog_review_merged',
        entity_type='spare_part_catalog',
        entity_id=part.id,
        entity_label=part.name,
        before=before,
        after=after,
        changes=diff_dict(before, after),
        description='Candidate merged into catalog entry #{} ({} items repointed)'.format(
            target.id, repointed)
    )
    db.session.commit()
    flash(_spare_t('Ёзув бирлаштирилди', 'Записи объединены'), 'success')
    return redirect(url_for('spare_parts.catalog'))


# ─── SPARE-STAGE3: equipment-model reference management ───────────────────────

def _require_catalog_manage():
    # [REASON]: SPARE-STAGE3 — all Stage-3 management screens (equipment models,
    # compatibility, maintenance norms) reuse the Stage-1 catalog permission.
    if not current_user.has_module_access('spare_parts_catalog_manage'):
        abort(403)


@spare_parts_bp.route('/equipment-models')
@module_required('spare_parts')
def equipment_models():
    """Manage the canonical equipment-model reference (Part 1b).

    Lists every model with how many Equipment rows point at it and how many
    compatibility rows reference it. This is where the owner does the one-time
    reconciliation of the ~336 migrated models (rename typos, merge duplicates)
    at their own pace — it never needs to be "finished" for the app to work.
    """
    _require_catalog_manage()
    models = (EquipmentModel.query
              .order_by(EquipmentModel.is_active.desc(), EquipmentModel.name)
              .all())
    # Bulk counters (avoid N+1): equipment per model, compatibility per model.
    eq_counts = dict(
        db.session.query(Equipment.model_id, func.count(Equipment.id))
        .filter(Equipment.model_id.isnot(None))
        .group_by(Equipment.model_id).all())
    compat_counts = dict(
        db.session.query(SparePartCompatibility.equipment_model_id,
                         func.count(SparePartCompatibility.id))
        .group_by(SparePartCompatibility.equipment_model_id).all())
    active_models = [m for m in models if m.is_active]
    return render_template('spare_parts_equipment_models.html',
                           models=models,
                           active_models=active_models,
                           eq_counts=eq_counts,
                           compat_counts=compat_counts,
                           lang=_spare_lang())


@spare_parts_bp.route('/equipment-models/save', methods=['POST'])
@module_required('spare_parts')
def equipment_model_save():
    """Create a model or edit its name/name_uz/manufacturer/active flag.

    Renaming keeps eq_type text in sync on every Equipment row pointing at the
    model, so legacy eq_type readers (daily_entry.html etc.) stay correct.
    """
    _require_catalog_manage()
    mid = request.form.get('id', type=int)
    name = (request.form.get('name', '') or '').strip()
    name_uz = (request.form.get('name_uz', '') or '').strip()
    manufacturer = (request.form.get('manufacturer', '') or '').strip()
    is_active = request.form.get('is_active') is not None
    if not name:
        flash(_spare_t('Модель номини киритинг', 'Введите название модели'), 'warning')
        return redirect(url_for('spare_parts.equipment_models'))
    # name is UNIQUE at the DB level — pre-check for a friendly message.
    clash = (EquipmentModel.query
             .filter(func.lower(EquipmentModel.name) == name.lower())
             .filter(EquipmentModel.id != (mid or 0)).first())
    if clash:
        flash(_spare_t('Бундай номли модель мавжуд',
                       'Модель с таким названием уже существует'), 'warning')
        return redirect(url_for('spare_parts.equipment_models'))

    created = False
    if mid:
        model = EquipmentModel.query.get_or_404(mid)
        old_name = model.name
        model.name = name
        model.name_uz = name_uz or None
        model.manufacturer = manufacturer or None
        model.is_active = is_active
        synced = 0
        if old_name != name:
            # [REASON]: SPARE-STAGE3 — keep eq_type in sync with the model name
            # so any code still reading eq_type directly sees the canonical text.
            synced = (Equipment.query.filter_by(model_id=model.id)
                      .update({'eq_type': name}, synchronize_session=False))
        _audit_spare(
            'spare_part_equipment_model_updated',
            entity_type='equipment_model', entity_id=model.id,
            entity_label=model.name,
            description='Model renamed {!r}->{!r}, {} equipment eq_type synced'.format(
                old_name, name, synced) if old_name != name else 'Model updated')
    else:
        model = EquipmentModel(name=name, name_uz=name_uz or None,
                               manufacturer=manufacturer or None, is_active=is_active)
        db.session.add(model)
        db.session.flush()
        created = True
        _audit_spare(
            'spare_part_equipment_model_created',
            entity_type='equipment_model', entity_id=model.id,
            entity_label=model.name, description='Model created')
    db.session.commit()
    flash(_spare_t('Сақланди', 'Сохранено'), 'success')
    return redirect(url_for('spare_parts.equipment_models'))


@spare_parts_bp.route('/equipment-models/merge', methods=['POST'])
@module_required('spare_parts')
def equipment_model_merge():
    """Merge one model into a survivor (Part 1b).

    Reassigns every Equipment.model_id and SparePartCompatibility /
    SparePartMaintenanceNorm equipment_model_id from the merged-away model to
    the survivor, syncs eq_type text on affected Equipment to the survivor's
    name, then deactivates (never hard-deletes) the merged model so its
    migrated_from_eq_type audit trail survives.
    """
    _require_catalog_manage()
    survivor_id = request.form.get('survivor_id', type=int)
    merged_id = request.form.get('merged_id', type=int)
    survivor = EquipmentModel.query.get(survivor_id) if survivor_id else None
    merged = EquipmentModel.query.get(merged_id) if merged_id else None
    if not survivor or not merged or survivor.id == merged.id:
        flash(_spare_t('Бирлаштириш учун иккита ҳар хил моделни танланг',
                       'Выберите две разные модели для объединения'), 'warning')
        return redirect(url_for('spare_parts.equipment_models'))

    # Compatibility: repoint, but first drop rows that would collide with a
    # survivor row for the same part (UNIQUE(spare_part_id, equipment_model_id)).
    survivor_parts = [r[0] for r in db.session.query(
        SparePartCompatibility.spare_part_id)
        .filter(SparePartCompatibility.equipment_model_id == survivor.id).all()]
    dropped_dupes = 0
    if survivor_parts:
        dropped_dupes = (SparePartCompatibility.query
                         .filter(SparePartCompatibility.equipment_model_id == merged.id,
                                 SparePartCompatibility.spare_part_id.in_(survivor_parts))
                         .delete(synchronize_session=False))
    compat_moved = (SparePartCompatibility.query
                    .filter_by(equipment_model_id=merged.id)
                    .update({'equipment_model_id': survivor.id}, synchronize_session=False))
    norms_moved = (SparePartMaintenanceNorm.query
                   .filter_by(equipment_model_id=merged.id)
                   .update({'equipment_model_id': survivor.id}, synchronize_session=False))
    eq_moved = (Equipment.query.filter_by(model_id=merged.id)
                .update({'model_id': survivor.id, 'eq_type': survivor.name},
                        synchronize_session=False))
    merged.is_active = False
    _audit_spare(
        'spare_part_equipment_model_merged',
        entity_type='equipment_model', entity_id=merged.id,
        entity_label=merged.name,
        description=('Merged {!r}(#{}) into {!r}(#{}): {} equipment, {} compat '
                     '({} dup dropped), {} norms repointed').format(
            merged.name, merged.id, survivor.name, survivor.id,
            eq_moved, compat_moved, dropped_dupes, norms_moved))
    db.session.commit()
    flash(_spare_t('Моделлар бирлаштирилди', 'Модели объединены'), 'success')
    return redirect(url_for('spare_parts.equipment_models'))


# ─── SPARE-STAGE3: part <-> model compatibility matrix (Part 2) ───────────────

@spare_parts_bp.route('/catalog/<int:pid>/compatibility')
@module_required('spare_parts')
def catalog_compatibility(pid):
    """Edit which equipment models a catalog part is compatible with."""
    _require_catalog_manage()
    part = SparePart.query.get_or_404(pid)
    selected_ids = {r[0] for r in db.session.query(
        SparePartCompatibility.equipment_model_id)
        .filter(SparePartCompatibility.spare_part_id == pid).all()}
    # [REASON]: SPARE-STAGE3 — show active models PLUS any already-linked model
    # even if it became inactive, so saving the form never silently drops a
    # compatibility row that had no visible checkbox to represent it.
    models = (EquipmentModel.query
              .filter(db.or_(EquipmentModel.is_active.is_(True),
                             EquipmentModel.id.in_(selected_ids)))
              .order_by(EquipmentModel.name).all())
    return render_template('spare_part_compatibility.html',
                           part=part, models=models,
                           selected_ids=selected_ids, lang=_spare_lang())


@spare_parts_bp.route('/catalog/<int:pid>/compatibility/save', methods=['POST'])
@module_required('spare_parts')
def catalog_compatibility_save(pid):
    """Reconcile SparePartCompatibility rows for a part from the checkbox set."""
    _require_catalog_manage()
    part = SparePart.query.get_or_404(pid)
    chosen = set(request.form.getlist('model_ids', type=int))
    # Only allow real, active models to be added.
    if chosen:
        valid = {m.id for m in EquipmentModel.query
                 .filter(EquipmentModel.id.in_(chosen)).all()}
        chosen &= valid
    existing = {r.equipment_model_id: r for r in
                SparePartCompatibility.query.filter_by(spare_part_id=pid).all()}
    to_add = chosen - set(existing.keys())
    to_remove = set(existing.keys()) - chosen
    for model_id in to_add:
        db.session.add(SparePartCompatibility(
            spare_part_id=pid, equipment_model_id=model_id,
            created_by=current_user.id))
    for model_id in to_remove:
        db.session.delete(existing[model_id])
    _audit_spare(
        'spare_part_compatibility_updated',
        entity_type='spare_part_catalog', entity_id=part.id,
        entity_label=part.name,
        description='Compatibility set: +{} -{} (now {} models)'.format(
            len(to_add), len(to_remove), len(chosen)))
    db.session.commit()
    flash(_spare_t('Мослик сақланди', 'Совместимость сохранена'), 'success')
    return redirect(url_for('spare_parts.catalog_compatibility', pid=pid))


# ─── SPARE-STAGE3: maintenance-interval norms + passive notification (Part 4) ──

def _current_engine_hours(equipment_id):
    """Cumulative engine-hours for a machine = SUM of the per-day Wialon records.

    [REASON]: SPARE-STAGE3 — EngineHoursRecord.engine_hours is a PER-DAY value
    (hours the engine ran that work_date), NOT a lifetime odometer. Cumulative
    hours are therefore the running sum of those daily rows. See the Part-4
    note in docs for why this is the only engine-hours signal available.
    """
    total = (db.session.query(func.coalesce(func.sum(EngineHoursRecord.engine_hours), 0.0))
             .filter(EngineHoursRecord.equipment_id == equipment_id).scalar())
    return float(total or 0.0)


def _last_replacement_date(equipment_id, spare_part_id):
    """Most recent date part `spare_part_id` was physically issued for machine
    `equipment_id` (its write-off act's issued_date), or None if never.

    [REASON]: SPARE-STAGE3 — there is NO engine-hours snapshot captured at issue
    time anywhere in the Stage-1/2 flow, so "last replacement" is anchored to
    the write-off act's calendar issued_date (the physical handover). Engine-
    hours accumulated since are then summed from the per-day records strictly
    after that date. When no issue exists, we DON'T know when the part was last
    replaced, so the caller stays silent — the feature warms up as parts get
    issued through the module, never guessing on missing data.
    """
    row = (db.session.query(func.max(SparePartWriteOffAct.issued_date))
           .join(SparePartWriteOffActItem,
                 SparePartWriteOffActItem.act_id == SparePartWriteOffAct.id)
           .join(SparePartRequestItem,
                 SparePartWriteOffActItem.request_item_id == SparePartRequestItem.id)
           .join(SparePartRequest,
                 SparePartWriteOffAct.request_id == SparePartRequest.id)
           .filter(SparePartRequest.equipment_id == equipment_id,
                   SparePartRequestItem.spare_part_id == spare_part_id)
           .scalar())
    return row


def _hours_since_replacement(equipment_id, since_date):
    """Engine-hours accumulated strictly AFTER since_date for this machine."""
    total = (db.session.query(func.coalesce(func.sum(EngineHoursRecord.engine_hours), 0.0))
             .filter(EngineHoursRecord.equipment_id == equipment_id,
                     EngineHoursRecord.work_date > since_date).scalar())
    return float(total or 0.0)


def _maintenance_due_rows(org_ids=None):
    """Compute which (equipment, part) pairs have crossed their interval.

    Passive only: returns a list of due dicts. A pair contributes a row ONLY
    when it has a real replacement anchor (an issued write-off act) AND the
    engine-hours accumulated since that anchor reach the norm's interval_hours.
    No anchor -> silent (we don't fabricate a replacement time). This can only
    ever under-alert (miss) if Wialon data is incomplete; it never false-fires.
    """
    due = []
    norms = (SparePartMaintenanceNorm.query
             .filter_by(is_active=True)
             .options(joinedload(SparePartMaintenanceNorm.spare_part),
                      joinedload(SparePartMaintenanceNorm.equipment_model))
             .all())
    norms = [n for n in norms if n.interval_hours and n.interval_hours > 0]
    if not norms:
        return due
    part_ids = {n.spare_part_id for n in norms}

    # [REASON]: CYCLE-2-3 Part 9 — batched: previously every norm ran its own
    # anchor-set query and every candidate machine two more queries (last
    # replacement date + hours since), i.e. 1 + N_norms + 2×N_pairs queries.
    # Now: ONE grouped query yields every (equipment, part) pair's newest
    # issue date (the anchor), and ONE more fetches the per-day engine-hour
    # totals of the anchored machines after the oldest anchor; per-pair
    # "hours strictly after the anchor" is then a suffix-sum lookup in
    # Python. The anchor-only/never-issued silence rule is preserved because
    # pairs without an act never appear in the anchor query. Output equality
    # with the per-pair helpers is asserted by tests.
    anchor_rows = (db.session.query(
            SparePartRequest.equipment_id,
            SparePartRequestItem.spare_part_id,
            func.max(SparePartWriteOffAct.issued_date))
        .join(SparePartWriteOffAct,
              SparePartWriteOffAct.request_id == SparePartRequest.id)
        .join(SparePartWriteOffActItem,
              SparePartWriteOffActItem.act_id == SparePartWriteOffAct.id)
        .join(SparePartRequestItem,
              SparePartWriteOffActItem.request_item_id == SparePartRequestItem.id)
        .filter(SparePartRequestItem.spare_part_id.in_(part_ids),
                SparePartRequest.equipment_id.isnot(None))
        .group_by(SparePartRequest.equipment_id,
                  SparePartRequestItem.spare_part_id)
        .all())
    # (equipment_id, part_id) -> anchor date
    pair_anchor = {(eq_id, part_id): anchor
                   for eq_id, part_id, anchor in anchor_rows}
    if not pair_anchor:
        return due
    anchored_eq_ids = {eq_id for eq_id, _p in pair_anchor}
    min_anchor = min(pair_anchor.values())
    day_rows = (db.session.query(EngineHoursRecord.equipment_id,
                                 EngineHoursRecord.work_date,
                                 func.sum(EngineHoursRecord.engine_hours))
                .filter(EngineHoursRecord.equipment_id.in_(anchored_eq_ids),
                        EngineHoursRecord.work_date > min_anchor)
                .group_by(EngineHoursRecord.equipment_id,
                          EngineHoursRecord.work_date)
                .all())
    per_eq_days = {}
    for eq_id, work_date, day_hours in day_rows:
        per_eq_days.setdefault(eq_id, []).append((work_date, float(day_hours or 0.0)))
    # Per equipment: dates ascending + suffix sums, so "hours after anchor"
    # is one bisect instead of one SUM query per (equipment, part) pair.
    per_eq_index = {}
    for eq_id, days in per_eq_days.items():
        days.sort()
        dates = [d for d, _h in days]
        suffix = [0.0] * (len(days) + 1)
        for i in range(len(days) - 1, -1, -1):
            suffix[i] = suffix[i + 1] + days[i][1]
        per_eq_index[eq_id] = (dates, suffix)

    def _hours_after(eq_id, anchor):
        dates_suffix = per_eq_index.get(eq_id)
        if not dates_suffix:
            return 0.0
        dates, suffix = dates_suffix
        return suffix[bisect.bisect_right(dates, anchor)]

    pair_data = {key: (anchor, _hours_after(key[0], anchor))
                 for key, anchor in pair_anchor.items()}

    # One equipment fetch for every anchored machine; the per-norm
    # active/model/org filters are applied per pair below, exactly as the
    # per-norm SQL filters did. joinedload keeps eq.model from lazy-loading
    # per row in the template.
    eq_map = {e.id: e for e in
              Equipment.query.options(joinedload(Equipment.model))
              .filter(Equipment.id.in_({eq_id for eq_id, _p in pair_data}))
              .all()}

    for norm in norms:
        # Ascending id order mirrors the original per-norm query's row order
        # so equal-overdue ties sort identically after the stable sort below.
        for eq_id in sorted(eq_id for (eq_id, part_id) in pair_data
                            if part_id == norm.spare_part_id):
            eq = eq_map.get(eq_id)
            if eq is None or eq.is_active is not True:
                continue
            if norm.equipment_model_id and eq.model_id != norm.equipment_model_id:
                continue
            if org_ids is not None and eq.organization_id not in org_ids:
                continue
            anchor, hours_since = pair_data[(eq_id, norm.spare_part_id)]
            if hours_since < norm.interval_hours:
                continue
            due.append({
                'equipment': eq,
                'part': norm.spare_part,
                'norm': norm,
                'model': eq.model,
                'last_replaced': anchor,
                'hours_since': round(hours_since, 1),
                'interval_hours': norm.interval_hours,
                'overdue_by': round(hours_since - norm.interval_hours, 1),
            })
    # Most overdue first.
    due.sort(key=lambda d: d['overdue_by'], reverse=True)
    return due


@spare_parts_bp.route('/maintenance-norms')
@module_required('spare_parts')
def maintenance_norms():
    """CRUD list for engine-hours replacement intervals."""
    _require_catalog_manage()
    norms = (SparePartMaintenanceNorm.query
             .options(joinedload(SparePartMaintenanceNorm.spare_part),
                      joinedload(SparePartMaintenanceNorm.equipment_model))
             .order_by(SparePartMaintenanceNorm.is_active.desc(),
                       SparePartMaintenanceNorm.id.desc())
             .all())
    parts = (SparePart.query.filter_by(status='active')
             .order_by(SparePart.name).all())
    models = (EquipmentModel.query.filter_by(is_active=True)
              .order_by(EquipmentModel.name).all())
    return render_template('spare_parts_maintenance_norms.html',
                           norms=norms, parts=parts, models=models,
                           lang=_spare_lang())


@spare_parts_bp.route('/maintenance-norms/save', methods=['POST'])
@module_required('spare_parts')
def maintenance_norm_save():
    _require_catalog_manage()
    nid = request.form.get('id', type=int)
    spare_part_id = request.form.get('spare_part_id', type=int)
    equipment_model_id = request.form.get('equipment_model_id', type=int) or None
    is_active = request.form.get('is_active') is not None
    try:
        interval_hours = float((request.form.get('interval_hours', '') or '').replace(',', '.'))
        # [REASON]: SP-F-016 — NaN passes the <= 0 gate below (NaN <= 0 is
        # False) and inf makes the norm unreachable garbage; both fold into
        # the existing "must be > 0" rejection.
        if not math.isfinite(interval_hours):
            interval_hours = 0
    except ValueError:
        interval_hours = 0
    if not spare_part_id or not SparePart.query.get(spare_part_id):
        flash(_spare_t('Деталью танланг', 'Выберите деталь'), 'warning')
        return redirect(url_for('spare_parts.maintenance_norms'))
    if interval_hours <= 0:
        flash(_spare_t('Интервал (моточас) 0 дан катта бўлиши керак',
                       'Интервал (моточасы) должен быть больше 0'), 'warning')
        return redirect(url_for('spare_parts.maintenance_norms'))
    if equipment_model_id and not EquipmentModel.query.get(equipment_model_id):
        equipment_model_id = None
    if nid:
        norm = SparePartMaintenanceNorm.query.get_or_404(nid)
        norm.spare_part_id = spare_part_id
        norm.equipment_model_id = equipment_model_id
        norm.interval_hours = interval_hours
        norm.is_active = is_active
        action = 'spare_part_maintenance_norm_updated'
    else:
        norm = SparePartMaintenanceNorm(
            spare_part_id=spare_part_id, equipment_model_id=equipment_model_id,
            interval_hours=interval_hours, is_active=is_active,
            created_by=current_user.id)
        db.session.add(norm)
        action = 'spare_part_maintenance_norm_created'
    db.session.flush()
    _audit_spare(action, entity_type='spare_part_maintenance_norm',
                 entity_id=norm.id, entity_label=str(norm.spare_part_id),
                 description='Norm interval={}h model={}'.format(
                     interval_hours, equipment_model_id or 'ALL'))
    db.session.commit()
    flash(_spare_t('Сақланди', 'Сохранено'), 'success')
    return redirect(url_for('spare_parts.maintenance_norms'))


@spare_parts_bp.route('/maintenance-norms/<int:nid>/delete', methods=['POST'])
@module_required('spare_parts')
def maintenance_norm_delete(nid):
    _require_catalog_manage()
    norm = SparePartMaintenanceNorm.query.get_or_404(nid)
    _audit_spare('spare_part_maintenance_norm_deleted',
                 entity_type='spare_part_maintenance_norm', entity_id=norm.id,
                 entity_label=str(norm.spare_part_id), description='Norm deleted')
    db.session.delete(norm)
    db.session.commit()
    flash(_spare_t('Ўчирилди', 'Удалено'), 'success')
    return redirect(url_for('spare_parts.maintenance_norms'))


@spare_parts_bp.route('/maintenance')
@module_required('spare_parts')
def maintenance_due():
    """Passive 'upcoming/overdue maintenance' list (Part 4).

    [REASON]: SPARE-STAGE3 — deliberately notification-only. It never creates a
    draft request; the owner decided passive notification for this stage.

    [REASON]: SP-F-011 — read-only due list, visible to any base spare_parts
    user scoped to their organizations (the manual's documented split):
    deliberately NO _require_catalog_manage() here, unlike the norm CRUD
    routes above, which keep it — viewing "what is due" is operations,
    editing the norms is catalog stewardship.
    """
    org_ids = None if current_user.is_admin else current_user.get_org_ids()
    rows = _maintenance_due_rows(org_ids=org_ids)
    return render_template('spare_parts_maintenance_due.html',
                           rows=rows, lang=_spare_lang())


# ─── SPARE-REPORTS: costs & repeat-warning analytics ──────────────────────────

def _reports_parse_date(value, default):
    """Lenient date parser for report query-string filters.

    [REASON]: SPARE-REPORTS — mirrors fuel_routes._parse_report_date: a bad
    query-string value falls back to the default period instead of failing
    the whole page (unlike _parse_spare_date, which raises for form input).
    """
    try:
        return datetime.strptime(value or '', '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return default


def _reports_filters():
    """Read and validate the filter set shared by reports and reports_export."""
    today = date.today()
    default_start = today.replace(day=1)
    d_from = _reports_parse_date(request.args.get('date_from'), default_start)
    d_to = _reports_parse_date(request.args.get('date_to'), today)
    if d_from > d_to:
        d_from, d_to = d_to, d_from

    org_id = request.args.get('org_id', type=int) or None
    if org_id and not current_user.can_access_org(org_id):
        abort(403)

    equipment_id = request.args.get('equipment_id', type=int) or None
    if equipment_id:
        eq = Equipment.query.get(equipment_id)
        if eq is None:
            equipment_id = None
        elif not current_user.can_access_org(eq.organization_id):
            abort(403)
        elif org_id and eq.organization_id != org_id:
            # Stale equipment kept from a previously selected organization —
            # the organization filter wins, the equipment filter is dropped.
            equipment_id = None

    category_id = request.args.get('category_id', type=int) or None
    if category_id and not SparePartCategory.query.get(category_id):
        category_id = None

    return d_from, d_to, org_id, equipment_id, category_id


def _reports_data(d_from, d_to, org_id=None, equipment_id=None, category_id=None):
    """Collect the five spare parts report tables for one filter set.

    [REASON]: SPARE-REPORTS business rule — cost tables count only money
    actually authorized: SparePartRequest.status in ('approved', 'issued') AND
    SparePartRequestItem.price_status == 'confirmed'. Draft/pending amounts
    never enter cost totals. Line cost = price * quantity.

    [REASON]: SPARE-STAGE2-QA-FIX3 — 'issued' is a Stage-2 terminal state
    reachable ONLY from 'approved' (goods handed over + write-off act exists);
    it is strictly a stronger signal of real spend than 'approved' alone, so it
    must be counted in every cost table and the top-20 — excluding it was an
    unintended Stage-1-era assumption that Stage 2 broke.
    """
    user_org_ids = _spare_user_org_ids()

    q = (db.session.query(SparePartRequestItem, SparePartRequest, SparePart)
         .join(SparePartRequest,
               SparePartRequestItem.request_id == SparePartRequest.id)
         .outerjoin(SparePart,
                    SparePartRequestItem.spare_part_id == SparePart.id)
         .filter(SparePartRequest.request_date >= d_from,
                 SparePartRequest.request_date <= d_to))
    if user_org_ids is not None:
        q = q.filter(SparePartRequest.organization_id.in_(user_org_ids))
    if org_id:
        q = q.filter(SparePartRequest.organization_id == org_id)
    if equipment_id:
        q = q.filter(SparePartRequest.equipment_id == equipment_id)
    if category_id:
        # NULL-category items (free-text/no catalog link) drop out here by
        # SQL NULL semantics, which is correct for an explicit category filter.
        q = q.filter(SparePart.category_id == category_id)
    lines = q.order_by(SparePartRequest.request_date,
                       SparePartRequest.id,
                       SparePartRequestItem.id).all()

    org_map = {o.id: o for o in Organization.query.all()}
    cat_map = {c.id: c for c in SparePartCategory.query.all()}
    eq_ids = {req.equipment_id for _item, req, _part in lines if req.equipment_id}
    eq_map = ({e.id: e for e in Equipment.query.filter(Equipment.id.in_(eq_ids)).all()}
              if eq_ids else {})

    def _cost(item):
        return float(item.price or 0) * float(item.quantity or 0)

    cost_lines = [(item, req, part) for item, req, part in lines
                  if req.status in ('approved', 'issued')
                  and item.price_status == 'confirmed']

    # Table 1: costs by equipment. Grouped by (organization_id, equipment_id)
    # so the NULL-equipment bucket stays split per organization and every row
    # can show its organization; real equipment belongs to exactly one
    # organization, so this equals grouping by equipment_id for it.
    eq_totals = {}
    for item, req, _part in cost_lines:
        key = (req.organization_id, req.equipment_id)
        row = eq_totals.setdefault(key, {
            'organization': org_map.get(req.organization_id),
            'equipment': eq_map.get(req.equipment_id),
            'total': 0.0, 'lines': 0,
        })
        row['total'] += _cost(item)
        row['lines'] += 1
    by_equipment = sorted(eq_totals.values(),
                          key=lambda r: r['total'], reverse=True)

    # Table 2: costs by organization.
    org_totals = {}
    for item, req, _part in cost_lines:
        row = org_totals.setdefault(req.organization_id, {
            'organization': org_map.get(req.organization_id),
            'total': 0.0, 'lines': 0,
        })
        row['total'] += _cost(item)
        row['lines'] += 1
    by_organization = sorted(org_totals.values(),
                             key=lambda r: r['total'], reverse=True)

    # Table 3: costs by catalog category. Items without a category (no catalog
    # link, or a part not yet categorized) get their own None bucket — they are
    # shown as «без категории» / «категориясиз», never silently dropped.
    cat_totals = {}
    for item, _req, part in cost_lines:
        cat_id = part.category_id if part is not None else None
        row = cat_totals.setdefault(cat_id, {
            'category': cat_map.get(cat_id),
            'total': 0.0, 'lines': 0,
        })
        row['total'] += _cost(item)
        row['lines'] += 1
    by_category = sorted(cat_totals.values(),
                         key=lambda r: r['total'], reverse=True)

    grand_total = sum(_cost(item) for item, _req, _part in cost_lines)

    # Table 4: repeat-order warnings triggered in the period.
    # [REASON]: SP-F-009/SP-F-023 — hybrid contract (owner decision
    # 2026-07-14): unlike the live form/detail (broad, today-anchored,
    # unchanged), the REPORT's repeat table is strict and reproducible:
    # examined lines and their matches are approved/issued only, and each
    # line's days_since is computed against that line's own request date
    # (as_of_date), never against "today" — so re-pulling a past period
    # later shows identical numbers. Deliberately NOT reusing cost_lines'
    # confirmed-price filter: repeat detection is about equipment/part
    # recurrence, not cost.
    # [REASON]: CYCLE-2-3 Part 9 — batched: ONE candidate query for the whole
    # report instead of one _check_repeat_orders query per examined line.
    # The per-line window/status/exclusion semantics (SP-F-009 hybrid
    # contract) are re-applied in Python over the shared candidate set, so
    # the resulting rows are identical to the per-line queries; equality is
    # asserted by tests against the original per-line implementation.
    examined = [(item, req, part) for item, req, part in lines
                if req.status in ('approved', 'issued')
                and req.equipment_id and item.spare_part_id]
    candidates = {}
    if examined:
        min_window_start = min(req.request_date for _i, req, _p in examined) \
            - timedelta(days=90)
        pair_eq_ids = {req.equipment_id for _i, req, _p in examined}
        pair_part_ids = {item.spare_part_id for item, _r, _p in examined}
        cand_rows = (db.session.query(SparePartRequest.equipment_id,
                                      SparePartRequestItem.spare_part_id,
                                      SparePartRequest.id,
                                      SparePartRequest.request_date)
                     .join(SparePartRequestItem,
                           SparePartRequestItem.request_id == SparePartRequest.id)
                     .filter(SparePartRequest.equipment_id.in_(pair_eq_ids),
                             SparePartRequestItem.spare_part_id.in_(pair_part_ids),
                             SparePartRequest.status.in_(('approved', 'issued')),
                             SparePartRequest.request_date >= min_window_start)
                     .distinct().all())
        for eq_id, part_id, cand_req_id, cand_date in cand_rows:
            candidates.setdefault((eq_id, part_id), []).append(
                (cand_date, cand_req_id))
        for pair_rows in candidates.values():
            pair_rows.sort(reverse=True)   # newest first, same as ORDER BY

    repeat_rows = []
    for item, req, part in examined:
        as_of = req.request_date
        window_start = as_of - timedelta(days=90)
        # Exact per-line re-application of the strict report contract:
        # eligible statuses only (already in SQL), window anchored to the
        # line's own date, strictly-earlier requests, self excluded.
        matches = [(d, rid_) for d, rid_ in candidates.get(
                       (req.equipment_id, item.spare_part_id), [])
                   if d >= window_start and d < as_of and rid_ != req.id]
        if not matches:
            continue
        days_since = (as_of - matches[0][0]).days
        if days_since <= 7:
            severity = 'red'
        elif days_since <= 30:
            severity = 'yellow'
        else:
            severity = None
        if severity in ('red', 'yellow'):
            repeat_rows.append({
                'request_id': req.id,
                'request_date': req.request_date,
                'equipment': eq_map.get(req.equipment_id),
                'organization': org_map.get(req.organization_id),
                # [REASON]: CYCLE-2-3 Part 7 — Uzbek alias override at
                # display time only; RU output stays the stored snapshot.
                'part_name': _snapshot_display_name(item.name, part),
                'severity': severity,
                'days_since': days_since,
            })
    repeat_rows.sort(key=lambda r: (0 if r['severity'] == 'red' else 1,
                                    r['days_since'] if r['days_since'] is not None else 999))

    # Table 5: top-20 most expensive confirmed approved line items.
    top_lines = sorted(cost_lines, key=lambda t: _cost(t[0]), reverse=True)[:20]
    top_items = [{
        'request_id': req.id,
        'request_date': req.request_date,
        'organization': org_map.get(req.organization_id),
        'equipment': eq_map.get(req.equipment_id),
        # [REASON]: CYCLE-2-3 Part 7 — Uzbek alias override at display time
        # only; RU output stays the stored snapshot.
        'name': _snapshot_display_name(item.name, part),
        'quantity': item.quantity,
        # [REASON]: RE-SP-010 — display-time localization; the stored
        # item.unit snapshot is untouched. Unknown codes pass through raw.
        'unit': _unit_label(item.unit or ''),
        'price': item.price,
        'total': _cost(item),
    } for item, req, part in top_lines]

    return {
        'd_from': d_from,
        'd_to': d_to,
        'by_equipment': by_equipment,
        'by_organization': by_organization,
        'by_category': by_category,
        'grand_total': grand_total,
        'cost_lines_count': len(cost_lines),
        'repeat_rows': repeat_rows,
        'top_items': top_items,
    }


def _xlsx_safe(value):
    """Neutralize spreadsheet formula injection in a user-controlled string.

    [REASON]: SP-F-004 — Excel/LibreOffice treat a cell starting with = + - @
    as a formula, so a part named '=1+1' (or worse, '=HYPERLINK(...)') typed
    by any operator would execute when the exported report is opened. The
    standard defense is a leading apostrophe: Excel shows the text verbatim
    and never evaluates it. Applied ONLY to user-controlled strings; fixed
    labels, ids, dates and numeric columns are left untouched.
    """
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped and stripped[0] in ('=', '+', '-', '@'):
            return "'" + value
    return value


def _spare_reports_workbook(data, lang='uz'):
    """Build the 5-sheet Excel workbook for the spare parts reports.

    Follows the styling conventions of fuel_routes._fuel_report_workbook
    (bold filled header row, frozen header, borders, auto column widths).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def L(ru, uz):
        return ru if lang == 'ru' else uz

    # [REASON]: SPARE-REPORTS — costs are сум amounts, so the money format is
    # excel_export.py's NUM_FMT ('#,##0', no decimals — same style as fmt_sum
    # in templates), not the fuel report's '#,##0.00' liters format.
    money_fmt = '#,##0'

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill('solid', fgColor='D9EAD3')
    header_font = Font(bold=True)
    total_font = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_table(ws, money_cols=()):
        ws.freeze_panes = 'A2'
        ws.sheet_view.showGridLines = False
        max_col = ws.max_column
        max_row = ws.max_row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                if cell.row > 1 and cell.column in money_cols \
                        and isinstance(cell.value, (int, float)):
                    cell.number_format = money_fmt
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            width = 10
            for cell in ws[letter]:
                val = '' if cell.value is None else str(cell.value)
                width = max(width, min(len(val) + 2, 38))
            ws.column_dimensions[letter].width = width

    # [REASON]: SP-F-004 — equipment/organization/part/unit names are operator
    # -typed reference data and must be formula-neutralized in every sheet;
    # the fixed '—'-prefixed fallbacks pass through _xlsx_safe unchanged.
    def eq_name(eq):
        return _xlsx_safe(eq.name) if eq else L('— без техники —', '— техникасиз —')

    def eq_plate(eq):
        return _xlsx_safe((eq.plate or '')) if eq else ''

    def org_name(org):
        return _xlsx_safe((org.short_name or org.name)) if org else '—'

    severity_labels = {
        'red': L('Красный (≤7 дней)', 'Қизил (≤7 кун)'),
        'yellow': L('Жёлтый (≤30 дней)', 'Сариқ (≤30 кун)'),
    }

    # Sheet 1: costs by equipment.
    ws = wb.create_sheet(L('Затраты по технике', 'Техника бўйича харажатлар'))
    ws.append([L('Техника', 'Техника'), L('Гос. номер', 'Давлат рақами'),
               L('Организация', 'Ташкилот'), L('Позиций', 'Позициялар'),
               L('Сумма, сум', 'Сумма, сўм')])
    for r in data['by_equipment']:
        ws.append([eq_name(r['equipment']), eq_plate(r['equipment']),
                   org_name(r['organization']), r['lines'], r['total']])
    ws.append([L('ИТОГО', 'ЖАМИ'), '', '', data['cost_lines_count'],
               data['grand_total']])
    for cell in ws[ws.max_row]:
        cell.font = total_font
    style_table(ws, money_cols=(5,))

    # Sheet 2: costs by organization.
    ws = wb.create_sheet(L('Затраты по организациям', 'Ташкилотлар бўйича харажатлар'))
    ws.append([L('Организация', 'Ташкилот'), L('Позиций', 'Позициялар'),
               L('Сумма, сум', 'Сумма, сўм')])
    for r in data['by_organization']:
        ws.append([org_name(r['organization']), r['lines'], r['total']])
    ws.append([L('ИТОГО', 'ЖАМИ'), data['cost_lines_count'], data['grand_total']])
    for cell in ws[ws.max_row]:
        cell.font = total_font
    style_table(ws, money_cols=(3,))

    # Sheet 3: costs by catalog category.
    ws = wb.create_sheet(L('Затраты по категориям', 'Категориялар бўйича харажатлар'))
    ws.append([L('Категория', 'Категория'), L('Позиций', 'Позициялар'),
               L('Сумма, сум', 'Сумма, сўм')])
    for r in data['by_category']:
        cat = r['category']
        label = (_xlsx_safe(cat.name_ru if lang == 'ru' else cat.name_uz) if cat
                 else L('— без категории —', '— категориясиз —'))
        ws.append([label, r['lines'], r['total']])
    ws.append([L('ИТОГО', 'ЖАМИ'), data['cost_lines_count'], data['grand_total']])
    for cell in ws[ws.max_row]:
        cell.font = total_font
    style_table(ws, money_cols=(3,))

    # Sheet 4: repeat-order warnings.
    # [REASON]: MOBILE-UPLOAD-001/i18n — Excel worksheet names are capped at
    # 31 characters and cannot contain : \ / ? * [ ]. The on-screen card
    # title for this table also carries a "(red <=7d, yellow <=30d)"
    # parenthetical, which alone pushes the RU/UZ pair to 49/47 characters.
    # That detail is already shown per-row via the Уровень/Даража column
    # (severity_labels below), so the sheet name reuses only the base
    # RU/UZ title without the parenthetical -- still the exact label pair
    # used on screen, just without the redundant, over-length suffix.
    ws = wb.create_sheet(L('Повторные заказы', 'Такрорий сўровлар'))
    ws.append([L('Заявка', 'Сўров'), L('Дата', 'Сана'),
               L('Организация', 'Ташкилот'), L('Техника', 'Техника'),
               L('Запчасть', 'Эҳтиёт қисм'), L('Уровень', 'Даража'),
               L('Дней с прошлого заказа', 'Олдинги сўровдан кунлар')])
    for r in data['repeat_rows']:
        ws.append(['#{}'.format(r['request_id']),
                   r['request_date'].strftime('%d.%m.%Y'),
                   org_name(r['organization']), eq_name(r['equipment']),
                   _xlsx_safe(r['part_name']),
                   severity_labels.get(r['severity'], r['severity']),
                   r['days_since']])
    style_table(ws)

    # Sheet 5: top-20 most expensive line items.
    ws = wb.create_sheet(L('Топ-20 самых дорогих позиций', 'Топ-20 энг қиммат позициялар'))
    ws.append([L('Заявка', 'Сўров'), L('Дата', 'Сана'),
               L('Организация', 'Ташкилот'), L('Техника', 'Техника'),
               L('Запчасть', 'Эҳтиёт қисм'), L('Кол-во', 'Миқдор'),
               L('Ед. изм.', 'Ўлчов бирлиги'), L('Цена, сум', 'Нарх, сўм'),
               L('Сумма, сум', 'Сумма, сўм')])
    for r in data['top_items']:
        ws.append(['#{}'.format(r['request_id']),
                   r['request_date'].strftime('%d.%m.%Y'),
                   org_name(r['organization']), eq_name(r['equipment']),
                   _xlsx_safe(r['name']), r['quantity'], _xlsx_safe(r['unit']),
                   r['price'], r['total']])
    style_table(ws, money_cols=(8, 9))

    return wb


@spare_parts_bp.route('/reports')
@module_required('spare_parts')
def reports():
    # [REASON]: SPARE-REPORTS — the whole screen sits behind the explicit
    # spare_parts_reports permission, checked the same way as the three
    # SPARE-STAGE1 permissions (has_module_access keeps the admin bypass).
    if not current_user.has_module_access('spare_parts_reports'):
        abort(403)
    d_from, d_to, org_id, equipment_id, category_id = _reports_filters()
    data = _reports_data(d_from, d_to, org_id=org_id,
                         equipment_id=equipment_id, category_id=category_id)

    user_org_ids = _spare_user_org_ids()
    if user_org_ids is None:
        organizations = Organization.query.order_by(Organization.sort_order).all()
    else:
        organizations = (Organization.query
                         .filter(Organization.id.in_(user_org_ids))
                         .order_by(Organization.sort_order).all())
    categories = (SparePartCategory.query
                  .order_by(SparePartCategory.sort_order, SparePartCategory.id)
                  .all())
    # Initial options for the equipment filter; refreshed on organization
    # change via the same AJAX endpoint the request form uses
    # (api_equipment_by_org — active equipment of the selected organization).
    if org_id:
        equipment_options = (Equipment.query
                             .filter_by(organization_id=org_id, is_active=True)
                             .order_by(Equipment.name, Equipment.plate).all())
    else:
        equipment_options = []
    return render_template('spare_parts_reports.html',
                           data=data,
                           date_from_s=d_from.isoformat(),
                           date_to_s=d_to.isoformat(),
                           org_id=org_id,
                           equipment_id=equipment_id,
                           category_id=category_id,
                           organizations=organizations,
                           categories=categories,
                           equipment_options=equipment_options,
                           lang=_spare_lang())


@spare_parts_bp.route('/reports/export')
@module_required('spare_parts')
def reports_export():
    # Same gate as the reports screen (this is the same data, as a file).
    if not current_user.has_module_access('spare_parts_reports'):
        abort(403)
    d_from, d_to, org_id, equipment_id, category_id = _reports_filters()
    data = _reports_data(d_from, d_to, org_id=org_id,
                         equipment_id=equipment_id, category_id=category_id)
    wb = _spare_reports_workbook(data, lang=_spare_lang())
    # Local imports mirror fuel_routes.balance_report_export — keeps this
    # change additive (module-level flask import line untouched).
    from io import BytesIO
    from flask import send_file
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    prefix = ('Spare_parts_report' if _spare_lang() == 'ru'
              else 'Ehtiyot_qismlar_hisoboti')
    fname = '{}_{}_{}.xlsx'.format(prefix, d_from.strftime('%d_%m_%Y'),
                                   d_to.strftime('%d_%m_%Y'))
    return send_file(
        buffer,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
