"""
Excel report generator v4 — fully dynamic per-category structure.
Sheet 1: ОБШИЙ (dynamic columns based on selected categories)
Sheet 2: ОБШИЙ КАМЧИЛИК
Sheets 3..N+2: СВОД [category] — one per selected category
Sheets N+3..2N+2: Свод ичи [category] — one per selected category
"""

import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.worksheet.properties import PageSetupProperties

from models import db, Organization, Equipment, DailyRecord, Deficiency, CATEGORIES


# ─── Style constants (matching operator's report) ────────────────────
THIN = Side(style='thin', color='000000')
MEDIUM = Side(style='medium', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_M = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FONT_TITLE = Font(name='Times New Roman', size=24, bold=True)
FONT_SUBTITLE = Font(name='Times New Roman', size=18, bold=True)
FONT_HEADER = Font(name='Times New Roman', size=16, bold=True)
FONT_HEADER_SM = Font(name='Times New Roman', size=14, bold=True)
FONT_DATA = Font(name='Times New Roman', size=20, bold=True)
FONT_DATA_SM = Font(name='Times New Roman', size=14)
FONT_TOTAL = Font(name='Times New Roman', size=18, bold=True)

# Fills
FILL_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
FILL_GREEN = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
FILL_BLUE = PatternFill(start_color='DAE3F3', end_color='DAE3F3', fill_type='solid')
FILL_GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

NUM_FMT = '#,##0'


def set_print(ws, orientation='landscape', fit_width=1, fit_height=0):
    """Setup landscape A4 with fit-to-width."""
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)
    ws.print_options.horizontalCentered = True


def style_cell(cell, font=None, fill=None, align=None, border=BORDER, num_fmt=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt


# ─── Data Loading ─────────────────────────────────────────────────────
def load_data(d_from, d_to, org_ids=None):
    """Load all org/records for report date range. Returns dict."""
    q = Organization.query.order_by(Organization.sort_order)
    if org_ids:
        q = q.filter(Organization.id.in_(org_ids))
    orgs = q.all()

    ALL_CATS = ['yukori', 'mtz', 'qatnov', 'mini', 'combine',
                'special', 'yuk_transport', 'motorcycle', 'passenger']

    data = {}

    for org in orgs:
        records = (DailyRecord.query
                   .filter(DailyRecord.work_date >= d_from,
                           DailyRecord.work_date <= d_to)
                   .join(Equipment)
                   .filter(Equipment.organization_id == org.id)
                   .order_by(Equipment.category, Equipment.name, DailyRecord.line_index)
                   .all())

        cats = {cat: [] for cat in ALL_CATS}
        for r in records:
            cat = r.equipment.category
            if cat in cats:
                cats[cat].append(r)

        sums = {}
        for cat in ALL_CATS:
            recs = cats[cat]
            sums[cat] = {
                'cash':     sum(r.amount_cash or 0 for r in recs),
                'transfer': sum(r.amount_transfer or 0 for r in recs),
                'internal': sum(r.amount_internal or 0 for r in recs),
                'other':    sum(r.amount_other or 0 for r in recs),
            }

        data[org.id] = {
            'org': org,
            'records': records,
            'cats': cats,
            'sums': sums,
        }

    return orgs, data


# ─── Sheet 1: ОБШИЙ (dynamic per selected categories) ─────────────────
def build_obshiy(wb, orgs, data, rd, nd, selected_cats):
    from openpyxl.utils import get_column_letter

    ws = wb.active
    ws.title = f'ОБШИЙ {rd.strftime("%d.%m")}'

    N = len(selected_cats)
    total_cols = 3 + 4 * N  # A, B, C + 4 payment groups x N cats

    PAY_LABELS = ['Накд', 'Пул кўчириш', 'Тизим корхонасида ишлади', 'Бошқа']
    PAY_KEYS   = ['cash', 'transfer', 'internal', 'other']

    # ── Row 1: Title ──────────────────────────────────────────────────
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    c = ws['A1']
    c.value = ('"Бухоро Агрокластер" МЧЖ тизимидаги барча корхоналарда мавжуд '
               'техникалар иштирокида амалга оширилган\nишлар суммаси ва тўлов '
               'турлари тугрисида МАЪЛУМОТИ')
    style_cell(c, font=FONT_TITLE, align=ALIGN_CENTER, border=None)
    ws.row_dimensions[1].height = 82.5

    # ── Row 2: Date ───────────────────────────────────────────────────
    date_start_col = 3 + 3 * N + 1  # start of last payment group
    date_end_col   = total_cols
    ws.merge_cells(
        start_row=2, start_column=date_start_col,
        end_row=2,   end_column=date_end_col
    )
    c = ws.cell(row=2, column=date_start_col)
    c.value = f'{rd.strftime("%d.%m")}-{nd.strftime("%d.%m.%Y")}'
    style_cell(c, font=FONT_HEADER, align=ALIGN_CENTER, border=None)
    ws.row_dimensions[2].height = 21.75

    # ── Row 3: Fixed headers + payment group headers ──────────────────
    ws.merge_cells('A3:A4')
    ws['A3'].value = '№'
    ws.merge_cells('B3:B4')
    ws['B3'].value = 'Ташкилот номи'
    ws.merge_cells('C3:C4')
    ws['C3'].value = 'Корхоналар кесимида иш хажми суммаси миқдори'

    for pay_idx, pay_label in enumerate(PAY_LABELS):
        start_col = 4 + pay_idx * N
        end_col   = start_col + N - 1
        ws.merge_cells(
            start_row=3, start_column=start_col,
            end_row=3,   end_column=end_col
        )
        c = ws.cell(row=3, column=start_col, value=pay_label)
        style_cell(c, font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER)
        for col in range(start_col + 1, end_col + 1):
            style_cell(ws.cell(row=3, column=col),
                       fill=FILL_YELLOW, font=FONT_HEADER, align=ALIGN_CENTER)

    # ── Row 4: Category name headers (repeated per payment type) ──────
    for pay_idx in range(4):
        for cat_idx, cat in enumerate(selected_cats):
            col = 4 + pay_idx * N + cat_idx
            cat_label = CATEGORIES.get(cat, cat)
            c = ws.cell(row=4, column=col, value=cat_label)
            style_cell(c, font=FONT_HEADER_SM, fill=FILL_YELLOW, align=ALIGN_CENTER)

    # Style fixed header cells rows 3-4
    for hrow in range(3, 5):
        for col in range(1, 4):
            cell = ws.cell(row=hrow, column=col)
            style_cell(cell, font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER)

    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 54

    # ── Data rows ────────────────────────────────────────────────────
    row = 5
    num = 1
    col_sums = {col: 0 for col in range(4, total_cols + 1)}

    for org in orgs:
        d = data.get(org.id)
        if not d:
            continue

        total_org = sum(
            d['sums'].get(cat, {}).get(pay, 0)
            for cat in selected_cats
            for pay in PAY_KEYS
        )

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=org.name)
        ws.cell(row=row, column=3, value=total_org if total_org else 0)

        style_cell(ws.cell(row=row, column=1), font=FONT_DATA, align=ALIGN_CENTER)
        style_cell(ws.cell(row=row, column=2), font=FONT_DATA, align=ALIGN_LEFT)
        style_cell(ws.cell(row=row, column=3), font=FONT_DATA, align=ALIGN_CENTER,
                   num_fmt=NUM_FMT)

        for pay_idx, pay_key in enumerate(PAY_KEYS):
            for cat_idx, cat in enumerate(selected_cats):
                col = 4 + pay_idx * N + cat_idx
                val = d['sums'].get(cat, {}).get(pay_key, 0)
                c = ws.cell(row=row, column=col, value=val if val else None)
                style_cell(c, font=FONT_DATA, align=ALIGN_CENTER, num_fmt=NUM_FMT)
                col_sums[col] += val

        ws.row_dimensions[row].height = 38
        num += 1
        row += 1

    # ── Сумма row ────────────────────────────────────────────────────
    ws.cell(row=row, column=2, value='Сумма:')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER)
    style_cell(ws.cell(row=row, column=2), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT)
    style_cell(ws.cell(row=row, column=3), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER)
    for col in range(4, total_cols + 1):
        c = ws.cell(row=row, column=col, value=col_sums[col])
        style_cell(c, font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    ws.row_dimensions[row].height = 30
    row += 1

    # ── ЖАМИ сумма row (payment type totals) ─────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1,
            value='ЖАМИ сумма:\n(Нақд/Пул кучириш/Тизим/Бошқа)')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT)
    style_cell(ws.cell(row=row, column=2), fill=FILL_GREEN, font=FONT_TOTAL)
    style_cell(ws.cell(row=row, column=3), fill=FILL_GREEN, font=FONT_TOTAL)

    pay_totals = []
    for pay_idx in range(4):
        start_col = 4 + pay_idx * N
        end_col   = start_col + N - 1
        pay_total = sum(col_sums[c] for c in range(start_col, end_col + 1))
        pay_totals.append(pay_total)
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row,   end_column=end_col)
        ws.cell(row=row, column=start_col, value=pay_total)
        style_cell(ws.cell(row=row, column=start_col), font=FONT_TOTAL,
                   fill=FILL_YELLOW, align=ALIGN_CENTER, num_fmt=NUM_FMT)
        for c in range(start_col + 1, end_col + 1):
            style_cell(ws.cell(row=row, column=c), fill=FILL_YELLOW, font=FONT_TOTAL)
    ws.row_dimensions[row].height = 50
    row += 1

    # ── ХАММАСИ row ──────────────────────────────────────────────────
    grand = sum(pay_totals)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1,
            value='ХАММАСИ \n(Нақд/Пул кучириш/Тизим/Бошқа):')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT)
    style_cell(ws.cell(row=row, column=2), fill=FILL_GREEN, font=FONT_TOTAL)
    style_cell(ws.cell(row=row, column=3), fill=FILL_GREEN, font=FONT_TOTAL)
    ws.merge_cells(start_row=row, start_column=4,
                   end_row=row,   end_column=total_cols)
    ws.cell(row=row, column=4, value=grand)
    style_cell(ws.cell(row=row, column=4), font=FONT_TITLE,
               fill=FILL_YELLOW, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    for c in range(5, total_cols + 1):
        style_cell(ws.cell(row=row, column=c), fill=FILL_YELLOW)
    ws.row_dimensions[row].height = 50

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    data_col_width = max(14, min(18, 80 // max(N, 1)))
    for col_idx in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = data_col_width

    set_print(ws, fit_height=1)
    ws.print_title_rows = '3:4'
    ws.print_area = f'A1:{last_col_letter}{row}'


# ─── Sheet 2: КАМЧИЛИК ───────────────────────────────────────────────
def build_kamchilik(wb, rd, deficiencies):
    ws = wb.create_sheet('ОБШИЙ КАМЧИЛИК')

    # Header row
    ws['A1'].value = 'NN'
    ws.merge_cells('B1:I1')
    ws['B1'].value = 'Аникланган камчиликлар:'

    style_cell(ws['A1'], font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_CENTER)
    for c in range(2, 10):
        style_cell(ws.cell(row=1, column=c), font=FONT_HEADER, fill=FILL_YELLOW, align=ALIGN_LEFT)

    ws.row_dimensions[1].height = 35

    # Deficiencies
    row = 2
    for i, d in enumerate(deficiencies, 1):
        ws.cell(row=row, column=1, value=i)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws.cell(row=row, column=2, value=d.text)

        style_cell(ws.cell(row=row, column=1), font=FONT_HEADER_SM, align=ALIGN_CENTER)
        for c in range(2, 10):
            style_cell(ws.cell(row=row, column=c), font=FONT_HEADER_SM, align=ALIGN_LEFT)

        text_len = len(d.text or '')
        ws.row_dimensions[row].height = max(35, min(150, text_len // 3))
        row += 1

    if not deficiencies:
        ws.cell(row=2, column=1, value=1)
        ws.merge_cells('B2:I2')
        ws.cell(row=2, column=2, value='(Камчиликлар киритилмаган)')
        for c in range(1, 10):
            style_cell(ws.cell(row=2, column=c), font=FONT_HEADER_SM, align=ALIGN_LEFT)
        row = 3

    ws.column_dimensions['A'].width = 8
    for col in 'BCDEFGHI':
        ws.column_dimensions[col].width = 25

    set_print(ws)
    ws.print_area = f'A1:I{row-1}'


# ─── СВОД (summary by single category) ───────────────────────────────
def build_svod(wb, orgs, data, rd, nd, category, sheet_title, cat_label):
    ws = wb.create_sheet(sheet_title)

    ws.merge_cells('A1:E1')
    ws['A1'].value = f'Бажарилган ишлар суммаси МАЪЛУМОТИ\n({cat_label})'
    style_cell(ws['A1'],
               font=Font(name='Times New Roman', size=22, bold=True, italic=False),
               align=ALIGN_CENTER, border=None)
    ws.row_dimensions[1].height = 65

    ws.merge_cells('C2:E2')
    ws['C2'].value = f'{rd.strftime("%d/%m/%Y")} - {nd.strftime("%d/%m/%Y")} (12:00 дан - 12:00 гача)'
    style_cell(ws['C2'], font=FONT_HEADER, align=ALIGN_CENTER, border=None)
    ws.row_dimensions[2].height = 25

    ws.merge_cells('A3:A4')
    ws['A3'].value = 'N'
    ws.merge_cells('B3:B4')
    ws['B3'].value = 'Ташкилот номи'
    ws.merge_cells('C3:F3')
    ws['C3'].value = 'БАЖАРИЛГАН ИШ СУММАСИ'
    ws['C4'].value = 'Накд'
    ws['D4'].value = 'Пул кўчириш'
    ws['E4'].value = 'Тизим корхонасида ишлади'
    ws['F4'].value = 'Бошқа'

    for r in range(3, 5):
        for c in range(1, 7):
            style_cell(ws.cell(row=r, column=c), font=FONT_HEADER,
                       fill=FILL_YELLOW, align=ALIGN_CENTER)
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 35

    row = 5
    num = 1
    sum_c, sum_t, sum_i, sum_o = 0, 0, 0, 0

    for org in orgs:
        d = data.get(org.id)
        if not d:
            continue

        eq_count = Equipment.query.filter_by(
            organization_id=org.id, category=category).count()
        if eq_count == 0:
            continue

        cash     = d['sums'].get(category, {}).get('cash', 0)
        transfer = d['sums'].get(category, {}).get('transfer', 0)
        internal = d['sums'].get(category, {}).get('internal', 0)
        other    = d['sums'].get(category, {}).get('other', 0)

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=org.name)
        ws.cell(row=row, column=3, value=cash if cash else 0)
        ws.cell(row=row, column=4, value=transfer if transfer else 0)
        ws.cell(row=row, column=5, value=internal if internal else 0)
        ws.cell(row=row, column=6, value=other if other else 0)

        style_cell(ws.cell(row=row, column=1), font=FONT_DATA, align=ALIGN_CENTER)
        style_cell(ws.cell(row=row, column=2), font=FONT_DATA, align=ALIGN_LEFT)
        for c in range(3, 7):
            style_cell(ws.cell(row=row, column=c), font=FONT_DATA,
                       align=ALIGN_CENTER, num_fmt=NUM_FMT)

        sum_c += cash
        sum_t += transfer
        sum_i += internal
        sum_o += other
        ws.row_dimensions[row].height = 40
        num += 1
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value='ЖАМИ СУММА:')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_CENTER)
    style_cell(ws.cell(row=row, column=2), font=FONT_TOTAL, fill=FILL_GREEN)
    ws.cell(row=row, column=3, value=sum_c)
    ws.cell(row=row, column=4, value=sum_t)
    ws.cell(row=row, column=5, value=sum_i)
    ws.cell(row=row, column=6, value=sum_o)
    for c in range(3, 7):
        style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL,
                   fill=FILL_GREEN, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    ws.row_dimensions[row].height = 40
    row += 1

    grand = sum_c + sum_t + sum_i + sum_o
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value='ХАММАСИ \n(Накд, Пул кучириш, Тизим, Бошқа):')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT)
    style_cell(ws.cell(row=row, column=2), font=FONT_TOTAL, fill=FILL_GREEN)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.cell(row=row, column=3, value=grand)
    style_cell(ws.cell(row=row, column=3), font=FONT_TITLE,
               fill=FILL_YELLOW, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    for c in range(4, 7):
        style_cell(ws.cell(row=row, column=c), fill=FILL_YELLOW)
    ws.row_dimensions[row].height = 55

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 22

    set_print(ws)
    ws.print_title_rows = '4:4'
    ws.print_area = f'A1:F{row}'


# ─── Свод ичи (detail by single category) ─────────────────────────────
def build_svod_ichi(wb, orgs, data, rd, nd, category, sheet_title, cat_label):
    ws = wb.create_sheet(sheet_title)

    label_map = {
        'yukori':        'юқори унумли ва махсус техникаларнинг',
        'mtz':           'чопиқ тракторларнинг',
        'qatnov':        'қатнов тракторларнинг',
        'mini':          'мини тракторларнинг',
        'combine':       'комбайнларнинг',
        'special':       'махсус техникаларнинг',
        'yuk_transport': 'юк ташувчи техникаларнинг',
        'motorcycle':    'мотоцикл техникаларнинг',
        'passenger':     'йўловчи ташиш техникаларнинг',
    }
    label = label_map.get(category, cat_label + 'нинг')

    ws.merge_cells('A1:M1')
    ws['A1'].value = (f'"Бухоро Агрокластер" МЧЖ тизимдаги ташкилотларга тегишли '
                      f'барча {label} иш билан бандлиги тўғрисида МАЪЛУМОТ')
    style_cell(ws['A1'], font=FONT_TITLE, align=ALIGN_CENTER, border=None)
    ws.row_dimensions[1].height = 50

    ws.merge_cells('I2:M2')
    ws['I2'].value = f'{rd.strftime("%d/%m/%Y")} - {nd.strftime("%d/%m/%Y")} (12:00 дан - 12:00 гача)'
    style_cell(ws['I2'], font=FONT_HEADER, align=ALIGN_CENTER, border=None)
    ws.row_dimensions[2].height = 25

    headers_r3 = [
        ('A', 'N'),
        ('B', 'Сана'),
        ('C', 'Ташкилот номи'),
        ('D', 'Техника русуми ва \nдавлат рақам белгиси'),
        ('E', 'Иш тури'),
        ('F', 'Буюртмачи \nноми '),
        ('G', 'Ўлчов бирлиги'),
        ('H', 'Микдори'),
        ('I', 'Нархи, сум'),
    ]
    for col, val in headers_r3:
        ws.merge_cells(f'{col}3:{col}4')
        ws[f'{col}3'].value = val

    ws.merge_cells('J3:M3')
    ws['J3'].value = 'БАЖАРИЛГАН ИШ СУММАСИ'
    ws['J4'].value = 'Накд'
    ws['K4'].value = 'Пул кўчириш'
    ws['L4'].value = 'Тизим корхонасида ишлади'
    ws['M4'].value = 'Бошқа'

    for r in range(3, 5):
        for c in range(1, 14):
            style_cell(ws.cell(row=r, column=c), font=FONT_HEADER_SM,
                       fill=FILL_YELLOW, align=ALIGN_CENTER)
    ws.row_dimensions[3].height = 35
    ws.row_dimensions[4].height = 35

    row = 5
    grand_c, grand_t, grand_i, grand_o = 0, 0, 0, 0

    for org in orgs:
        d = data.get(org.id)
        if not d:
            continue

        recs = data[org.id]['cats'].get(category, [])
        eq_in_cat = (Equipment.query
                     .filter_by(organization_id=org.id,
                                category=category,
                                is_active=True)
                     .order_by(Equipment.name).all())

        if not eq_in_cat and not recs:
            continue

        eq_records = {}
        for r in recs:
            eq_records.setdefault(r.equipment_id, []).append(r)
        for eq_id in eq_records:
            eq_records[eq_id].sort(key=lambda r: (r.work_date, r.line_index))

        num = 1
        org_c = org_t = org_i = org_o = 0
        first_in_org = True

        for eq in eq_in_cat:
            eq_recs = eq_records.get(eq.id, [])

            if not eq_recs:
                ws.cell(row=row, column=1, value=num)
                if first_in_org:
                    ws.cell(row=row, column=3, value=org.name)
                    first_in_org = False
                ws.cell(row=row, column=4, value=eq.full_name)
                ws.cell(row=row, column=5, value='Вақтинча бўш')

                for c in range(1, 14):
                    cell = ws.cell(row=row, column=c)
                    fill = FILL_YELLOW if c == 5 else None
                    style_cell(cell, font=FONT_DATA_SM, fill=fill,
                               align=ALIGN_LEFT if c in (3, 4, 5, 6) else ALIGN_CENTER)
                ws.row_dimensions[row].height = 30
                num += 1
                row += 1
            else:
                first_line = True
                for rec in eq_recs:
                    if first_line:
                        ws.cell(row=row, column=1, value=num)
                        if first_in_org:
                            ws.cell(row=row, column=3, value=org.name)
                            first_in_org = False
                        ws.cell(row=row, column=4, value=eq.full_name)

                    ws.cell(row=row, column=2, value=rec.work_date.strftime('%d.%m.%Y'))

                    is_idle = rec.status != 'working'
                    if rec.status == 'working':
                        ws.cell(row=row, column=5, value=rec.work_type)
                        ws.cell(row=row, column=6, value=rec.customer)
                        ws.cell(row=row, column=7, value=rec.unit)
                        ws.cell(row=row, column=8, value=rec.quantity)
                        ws.cell(row=row, column=9, value=rec.price)
                        if rec.amount_cash:
                            ws.cell(row=row, column=10, value=rec.amount_cash)
                        if rec.amount_transfer:
                            ws.cell(row=row, column=11, value=rec.amount_transfer)
                        if rec.amount_internal:
                            ws.cell(row=row, column=12, value=rec.amount_internal)
                        if rec.amount_other:
                            ws.cell(row=row, column=13, value=rec.amount_other)
                        org_c += rec.amount_cash or 0
                        org_t += rec.amount_transfer or 0
                        org_i += rec.amount_internal or 0
                        org_o += rec.amount_other or 0
                    else:
                        ws.cell(row=row, column=5,
                                value=rec.idle_reason or 'Вақтинча бўш')

                    for c in range(1, 14):
                        cell = ws.cell(row=row, column=c)
                        fill = FILL_YELLOW if (is_idle and c == 5) else None
                        style_cell(cell, font=FONT_DATA_SM, fill=fill,
                                   align=ALIGN_LEFT if c in (3, 4, 5, 6) else ALIGN_CENTER,
                                   num_fmt=NUM_FMT if c in (8, 9, 10, 11, 12, 13) else None)
                    ws.row_dimensions[row].height = 30
                    first_line = False
                    row += 1
                num += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=f'Жами {org.name}:')
        style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL,
                   fill=FILL_GREEN, align=ALIGN_LEFT)
        for c in range(2, 10):
            style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL, fill=FILL_GREEN)
        ws.cell(row=row, column=10, value=org_c)
        ws.cell(row=row, column=11, value=org_t)
        ws.cell(row=row, column=12, value=org_i)
        ws.cell(row=row, column=13, value=org_o)
        for c in range(10, 14):
            style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL,
                       fill=FILL_GREEN, align=ALIGN_CENTER, num_fmt=NUM_FMT)
        ws.row_dimensions[row].height = 35

        grand_c += org_c
        grand_t += org_t
        grand_i += org_i
        grand_o += org_o
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value='ЖАМИ СУММА:')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_YELLOW, align=ALIGN_CENTER)
    for c in range(2, 10):
        style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL, fill=FILL_YELLOW)
    ws.cell(row=row, column=10, value=grand_c)
    ws.cell(row=row, column=11, value=grand_t)
    ws.cell(row=row, column=12, value=grand_i)
    ws.cell(row=row, column=13, value=grand_o)
    for c in range(10, 14):
        style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL,
                   fill=FILL_YELLOW, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    ws.row_dimensions[row].height = 40
    row += 1

    grand = grand_c + grand_t + grand_i + grand_o
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1,
            value='ХАММАСИ (Накд, Пул кучириш, Тизим, Бошқа):')
    style_cell(ws.cell(row=row, column=1), font=FONT_TOTAL, fill=FILL_GREEN, align=ALIGN_LEFT)
    for c in range(2, 10):
        style_cell(ws.cell(row=row, column=c), font=FONT_TOTAL, fill=FILL_GREEN)
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    ws.cell(row=row, column=10, value=grand)
    style_cell(ws.cell(row=row, column=10), font=FONT_TITLE,
               fill=FILL_YELLOW, align=ALIGN_CENTER, num_fmt=NUM_FMT)
    for c in range(11, 14):
        style_cell(ws.cell(row=row, column=c), fill=FILL_YELLOW)
    ws.row_dimensions[row].height = 50

    widths = {
        'A': 5, 'B': 12, 'C': 22, 'D': 25, 'E': 22, 'F': 28,
        'G': 12, 'H': 11, 'I': 15, 'J': 18, 'K': 18, 'L': 20, 'M': 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    set_print(ws)
    ws.print_title_rows = '3:4'
    ws.print_area = f'A1:M{row}'


# ─── Main Generator ──────────────────────────────────────────────────
def generate_report(d_from, d_to, output_dir, org_ids=None, cat_filter=None):
    """
    cat_filter: list of category codes e.g. ['mtz', 'yukori', 'qatnov']
                None means all 9 categories
    """
    ALL_CATS = ['yukori', 'mtz', 'qatnov', 'mini', 'combine',
                'special', 'yuk_transport', 'motorcycle', 'passenger']

    if cat_filter:
        selected_cats = [c for c in ALL_CATS if c in cat_filter]
    else:
        selected_cats = ALL_CATS

    orgs, data = load_data(d_from, d_to, org_ids)
    rd, nd = d_from, d_to

    q = (Deficiency.query
         .filter(Deficiency.work_date >= d_from, Deficiency.work_date <= d_to)
         .order_by(Deficiency.work_date, Deficiency.sort_order, Deficiency.id))
    if org_ids:
        q = q.filter(
            (Deficiency.organization_id.in_(org_ids)) | (Deficiency.organization_id == None)
        )
    deficiencies = q.all()

    wb = Workbook()

    if d_from == d_to:
        dl = rd.strftime('%d.%m')
    else:
        dl = f'{rd.strftime("%d.%m")}-{nd.strftime("%d.%m")}'

    # Sheet 1: ОБШИЙ (dynamic columns)
    build_obshiy(wb, orgs, data, rd, nd, selected_cats)

    # Sheet 2: КАМЧИЛИК
    build_kamchilik(wb, rd, deficiencies)

    prefix_svod = f'СВОД {dl} '
    prefix_ichi = f'Ичи {dl} '
    max_svod = 31 - len(prefix_svod)  # remaining chars for category label
    max_ichi = 31 - len(prefix_ichi)

    for cat in selected_cats:
        cat_label = CATEGORIES.get(cat, cat)
        svod_title = f'{prefix_svod}{cat_label}'[:31]
        ichi_title = f'{prefix_ichi}{cat_label}'[:31]
        build_svod(wb, orgs, data, rd, nd, cat, svod_title, cat_label)
        build_svod_ichi(wb, orgs, data, rd, nd, cat, ichi_title, cat_label)

    if d_from == d_to:
        filename = f'Hisobot_{rd.strftime("%d_%m_%Y")}.xlsx'
    else:
        filename = f'Hisobot_{rd.strftime("%d_%m_%Y")}_{nd.strftime("%d_%m_%Y")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
