# -*- coding: utf-8 -*-
"""DRONE-RECEIVED-BLANK-IS-ZERO-001 -- «Получено» stops having an unknown.

THE OWNER'S DECISION, 2026-08-07, not a fact derived from the data: in the
dispatcher sheets an empty «олинмаган пуллар» cell means NOTHING WAS
COLLECTED, not «nobody wrote it down». DRONE-ZERO-VS-UNKNOWN-001 had treated
it like a blank «Сумма» and printed a dash, which produced live rows saying
two things at once:

    Бухоро Сервис Агрокластер    3    55 429 140    —    55 429 140

«collected: unknown; uncollected: all of it.» So the dash leaves «Получено».

What does NOT change, and is asserted here to make sure it did not:

  * «Сумма» keeps its dash and its «сумма не указана: N» counter.
  * «Не получено» keeps its dash, still keyed on no_amount.
  * The three-bucket split on the debts page.
  * Every piece of arithmetic, including the column SUM -- an empty cell
    already counts as zero inside SUM, so only the statement changes.

What these assertions can and cannot prove:

  * The negative controls load drones.py out of git at THIS task's base,
    de97c8d, and run it against the same fixture. Where git cannot produce
    that commit they skip loudly; a skipped control is not a passed one.
  * Nothing here was run against staging. The reviewer's blank counts (81 on
    «Долги — заказчики», 94 on «По заказчикам», column sum 1 904 558 438.98)
    were measured on staging workbooks and cannot be reproduced from this
    branch. The counts asserted below are this fixture's own.

Run:
  python -m unittest tests.test_drone_received_blank_is_zero -v
"""
import io
import re
import unittest

from tests.harness import app, create_admin, login
from tests.test_drone_zero_vs_unknown import (
    ALL_NULL, CUT_MONEY, CUT_SHEETS, DEBT_MONEY, DEBT_SHEETS, MIXED,
    NONE_NULL, NO_RECEIVED, PARTIAL, TOUCHED_TEMPLATES, UNKNOWN_LABEL_RU,
    UNSTATED_LABEL_RU, XLSX_ROUTES, by_label, cut_cells, pre_commit_drones,
    read_template, report_data, seed, set_language, text_of)

# The base this task was written against: the merge of PR #48.
BASE_COMMIT = 'de97c8d'


def money_cells(row):
    import drones
    return drones._drone_work_money_cells(row)


def every_row(data):
    """Every dict a workbook writer will hand to _drone_work_money_cells()."""
    rows = []
    for name in ('by_customer', 'by_operator', 'by_subdivision', 'by_month',
                 'by_payment'):
        cut = data[name]
        rows.extend(cut['rows'])
        rows.extend(cut['services'])
        rows.append(cut['total'])
    for name in ('debt_by_customer', 'debt_by_operator'):
        debt = data[name]
        rows.extend(debt['rows'])
        for bucket in (debt['unknown'], debt['rest']):
            if bucket is not None:
                rows.append(bucket)
        rows.append(debt['total'])
    return rows


class PreCommitMixin(object):
    """This task's base, not the previous task's."""

    def pre_commit(self):
        module = pre_commit_drones(BASE_COMMIT)
        if module is None:
            self.skipTest('git cannot produce %s:drones.py -- the negative '
                          'control did NOT run' % BASE_COMMIT)
        return module


# ---------------------------------------------------------------------------
# A  commit 1 -- drones.py
# ---------------------------------------------------------------------------

class TestA1TheMiddleCellIsPlain(unittest.TestCase):
    """A-1: an all-NULL-received group gets 0, not None."""

    @classmethod
    def setUpClass(cls):
        seed()

    def setUp(self):
        self.rows = by_label(report_data()['by_customer'])

    def test_a1_all_null_received_and_all_null_amount(self):
        """ALL_NULL: nothing priced, nothing collected.

        Sum unknown, debt unknown, collected ZERO -- the three columns are
        now three different rules and this row shows all three.
        """
        amount, received, outstanding = money_cells(self.rows[ALL_NULL])
        self.assertIsNone(amount)
        self.assertEqual(0, received)
        self.assertIsNone(outstanding)

    def test_a1_all_null_received_with_amounts_known(self):
        """NO_RECEIVED: 13 000 000 charged, nothing collected, all of it due."""
        row = self.rows[NO_RECEIVED]
        self.assertEqual(2, row['no_received'])
        self.assertEqual(row['jobs'], row['no_received'])
        amount, received, outstanding = money_cells(row)
        self.assertEqual(13000000.0, amount)
        self.assertEqual(0, received)
        self.assertEqual(13000000.0, outstanding)

    def test_a1_the_service_row_too(self):
        amount, received, outstanding = money_cells(
            self.rows[UNSTATED_LABEL_RU])
        self.assertIsNone(amount)
        self.assertEqual(0, received)
        self.assertIsNone(outstanding)

    def test_a1_a_group_that_did_collect_is_untouched(self):
        amount, received, outstanding = money_cells(self.rows[NONE_NULL])
        self.assertEqual(7000000.0, amount)
        self.assertEqual(7000000.0, received)
        self.assertEqual(0.0, outstanding)

    def test_a1_a_partial_group_is_untouched(self):
        amount, received, outstanding = money_cells(self.rows[MIXED])
        self.assertEqual(5000000.0, amount)
        self.assertEqual(3000000.0, received)
        self.assertEqual(2000000.0, outstanding)

    def test_a1_the_returned_value_is_the_rows_own_received(self):
        """Plain pass-through: no rounding, no coalescing, no rule."""
        for label, row in self.rows.items():
            with self.subTest(group=label):
                self.assertEqual(row['received'], money_cells(row)[1])


class TestA2TheCounterSurvives(unittest.TestCase):
    """A-2: no_received is still computed and still correct.

    Coverage is deliberately NOT duplicated here. TestA1Counters and
    TestA2CountersAddUp in tests/test_drone_zero_vs_unknown.py already assert
    no_received on ordinary rows, on both service rows, on the operator
    partition, on `rest`, on every cut total, and that each total equals the
    sum over its groups. Those tests are untouched by this task and still
    pass; this class only pins the two things they cannot know: that the
    counter is still there AFTER the column stopped using it, and that
    nothing renders it.
    """

    @classmethod
    def setUpClass(cls):
        seed()

    def test_a2_every_row_still_carries_the_counter(self):
        for row in every_row(report_data()):
            with self.subTest(label=row.get('label')):
                self.assertIn('no_received', row)
                self.assertIsInstance(row['no_received'], int)

    def test_a2_the_counter_is_still_correct(self):
        rows = by_label(report_data()['by_customer'])
        self.assertEqual(2, rows[ALL_NULL]['no_received'])
        self.assertEqual(0, rows[MIXED]['no_received'])
        self.assertEqual(0, rows[NONE_NULL]['no_received'])
        self.assertEqual(2, rows[NO_RECEIVED]['no_received'])
        self.assertEqual(6, report_data()['by_customer']['total']
                         ['no_received'])

    def test_a2_the_deliberate_non_use_is_written_down(self):
        """A counter computed and never shown needs a reason IN THE CODE.

        Without it the next reader deletes it as dead weight, or restores the
        dash believing they are fixing an oversight. Both were named as the
        risk when the decision was taken.
        """
        import inspect
        import drones
        for func in (drones._drone_work_cut, drones._drone_work_bucket_row,
                     drones._drone_work_money_cells):
            with self.subTest(function=func.__name__):
                source = inspect.getsource(func)
                self.assertIn('DRONE-RECEIVED-BLANK-IS-ZERO-001', source)
                self.assertIn('2026-08-07', source)


class TestA3TheMiddleCellIsNeverNone(unittest.TestCase):
    """A-3: the check that fails if somebody reinstates the blank."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_a3_no_row_anywhere_produces_a_blank_received(self):
        rows = every_row(report_data())
        self.assertGreater(len(rows), 20, 'the sweep read almost nothing')
        for row in rows:
            with self.subTest(label=row.get('label')):
                self.assertIsNotNone(money_cells(row)[1])

    def test_a3_including_the_rows_whose_other_columns_are_blank(self):
        """The case that matters: blank neighbours, numeric middle."""
        blanks = [row for row in every_row(report_data())
                  if money_cells(row)[0] is None]
        self.assertTrue(blanks, 'no unpriced row in the fixture at all')
        for row in blanks:
            with self.subTest(label=row.get('label')):
                amount, received, outstanding = money_cells(row)
                self.assertIsNone(amount)
                self.assertIsNone(outstanding)
                self.assertIsNotNone(received)
                self.assertEqual(0, received)

    def test_a3_the_check_can_fail(self):
        """Negative control for the control: a planted blank is caught."""
        import drones
        planted = {'label': 'planted', 'jobs': 2, 'area': 0.0,
                   'amount': 0.0, 'received': 0.0, 'outstanding': 0.0,
                   'no_amount': 2, 'no_received': 2}
        self.assertEqual(0.0, drones._drone_work_money_cells(planted)[1])
        # what the pre-commit rule would have produced for the same dict
        self.assertIsNone(
            drones._drone_money_or_blank(planted, 'received', 'no_received'),
            'the old rule no longer blanks -- then A-3 proves nothing')


class TestA4NegativeControl(PreCommitMixin, unittest.TestCase):
    """A-4: at de97c8d the same fixture returns None in the middle."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_a4_pre_commit_blanks_the_received_cell(self):
        module = self.pre_commit()
        row = by_label(report_data()['by_customer'])[ALL_NULL]
        before = module._drone_work_money_cells(row)
        self.assertIsNone(before[1],
                          'the control is measuring nothing: de97c8d already '
                          'returns a number here')
        self.assertIsNone(before[0])
        self.assertIsNone(before[2])

    def test_a4_pre_commit_blanks_it_for_no_received_too(self):
        module = self.pre_commit()
        row = by_label(report_data()['by_customer'])[NO_RECEIVED]
        before = module._drone_work_money_cells(row)
        self.assertEqual(13000000.0, before[0])
        self.assertIsNone(before[1])
        self.assertEqual(13000000.0, before[2])

    def test_a4_and_this_commit_does_not(self):
        rows = by_label(report_data()['by_customer'])
        for label in (ALL_NULL, NO_RECEIVED):
            with self.subTest(group=label):
                self.assertEqual(0, money_cells(rows[label])[1])

    def test_a4_the_control_is_reading_the_other_module(self):
        import drones
        module = self.pre_commit()
        self.assertIsNot(module, drones)
        self.assertNotEqual(module.__file__, drones.__file__)


# ---------------------------------------------------------------------------
# B  commit 2 -- the two screens
# ---------------------------------------------------------------------------

REPORTS_TEMPLATE = 'templates/drones/works_reports.html'
DEBTS_TEMPLATE = 'templates/drones/works_debts.html'

# (page, table title RU, table title UZ, column index of «Сумма»)
PAGES = (
    ('/drones/works/reports', 'По заказчикам', 'Буюртмачилар бўйича', 3),
    ('/drones/works/debts', 'Долги — по заказчикам',
     'Қарзлар — буюртмачилар бўйича', 2),
)

# Every money_cell(...) CALL in the module, with its note argument if any.
# The negative lookbehind skips the macro's own declaration in the partial --
# «{% macro money_cell(value, jobs, missing, note='') %}» is not a call site,
# and counting it reported note='' as an undefined key on the first run.
_CALL_RE = re.compile(r'(?<!macro )money_cell\(\s*(?P<args>[^()]*?)\s*\)')
# The note dictionary inside the partial: 'lang': {'key': 'text', ...}
_NOTE_KEY_RE = re.compile(r"'([a-z_]+)'\s*:\s*'[^']+'")


def note_keys():
    """The keys the partial's dictionary actually defines, per language."""
    source = read_template(PARTIAL)
    body = source.split('{% macro money_cell', 1)[1]
    keys = {}
    for lang in ('ru', 'uz'):
        block = body.split("'%s': {" % lang, 1)[1].split('}', 1)[0]
        keys[lang] = set(_NOTE_KEY_RE.findall("{" + block + "}"))
    return keys


def call_sites():
    """[(template, line, note-argument-or-None)] over the whole module."""
    from tests.test_drone_zero_vs_unknown import (_JINJA_COMMENT_RE,
                                                  drones_templates)
    found = []
    for path in drones_templates():
        source = _JINJA_COMMENT_RE.sub(
            lambda m: re.sub(r'[^\n]', ' ', m.group(0)), read_template(path))
        for match in _CALL_RE.finditer(source):
            args = [a.strip() for a in match.group('args').split(',')]
            note = None
            if len(args) >= 4:
                note = args[3].strip('\'"')
            found.append((path, source[:match.start()].count('\n') + 1, note))
    return found


class ScreenMixin(object):
    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin(cls.__name__.lower()[:24])
        set_language(cls.admin, 'ru')

    def page(self, url, lang='ru'):
        set_language(self.admin, lang)
        client = app.test_client()
        login(client, self.admin)
        response = client.get(url)
        self.assertEqual(200, response.status_code, url)
        return response.data.decode('utf-8')

    def money_columns(self, url, title, first, lang='ru'):
        body = self.page(url, lang=lang)
        columns = [cut_cells(body, title, first + offset)
                   for offset in (0, 1, 2)]
        return {label: tuple(column[label] for column in columns)
                for label in columns[0]}


class TestB1ReceivedShowsZero(ScreenMixin, unittest.TestCase):
    """B-1: an all-NULL-received group renders 0, neighbours unaffected."""

    def test_b1_received_is_zero_on_both_pages(self):
        for url, title, _uz, first in PAGES:
            with self.subTest(page=url):
                cells = self.money_columns(url, title, first)
                for label in cells:
                    if 'Долг неизвестен' in label or label == ALL_NULL:
                        _amount, received, _out = cells[label]
                        self.assertEqual('0', text_of(received))

    def test_b1_the_neighbours_keep_their_dashes(self):
        """«Сумма» and «Не получено» are untouched by this task."""
        cells = self.money_columns('/drones/works/reports',
                                   'По заказчикам', 3)
        amount, received, outstanding = cells[ALL_NULL]
        self.assertEqual('—', text_of(amount))
        self.assertEqual('0', text_of(received))
        self.assertEqual('—', text_of(outstanding))

    def test_b1_the_unknown_bucket_reads_zero_between_two_dashes(self):
        """Deliberate and coherent: nothing collected, debt unknown.

        The task names this explicitly as the intended result rather than
        something to «fix».
        """
        cells = self.money_columns('/drones/works/debts',
                                   'Долги — по заказчикам', 2)
        label = [k for k in cells if 'Долг неизвестен' in k][0]
        amount, received, outstanding = cells[label]
        self.assertEqual('—', text_of(amount))
        self.assertEqual('0', text_of(received))
        self.assertEqual('—', text_of(outstanding))

    def test_b1_a_group_that_collected_is_unchanged(self):
        cells = self.money_columns('/drones/works/reports',
                                   'По заказчикам', 3)
        self.assertIn('7\xa0000\xa0000', cells[NONE_NULL][1])
        self.assertIn('3\xa0000\xa0000', cells[MIXED][1])

    def test_b1_the_amount_note_still_renders(self):
        """The «Сумма» counter must survive: only «Получено» changed."""
        cells = self.money_columns('/drones/works/reports',
                                   'По заказчикам', 3)
        self.assertIn('сумма не указана', text_of(cells[MIXED][0]))


class TestB2TheReceivedNoteIsGoneFromEveryPage(ScreenMixin,
                                               unittest.TestCase):
    """B-2: «получено не указано» appears nowhere, in either language."""

    GONE = ('получено не указано', 'кирим қилингани кўрсатилмаган')

    def test_b2_no_page_renders_the_received_wording(self):
        for url, _ru, _uz, _first in PAGES:
            for lang in ('ru', 'uz'):
                with self.subTest(page=url, lang=lang):
                    body = self.page(url, lang=lang)
                    for phrase in self.GONE:
                        self.assertNotIn(phrase, body)

    def test_b2_the_reports_page_still_renders_the_amount_wording(self):
        """Or the check above would pass on a page that lost both."""
        body = self.page('/drones/works/reports', lang='ru')
        self.assertIn('сумма не указана', body)
        body = self.page('/drones/works/reports', lang='uz')
        self.assertIn('сумма кўрсатилмаган', body)


class TestB3NoteKeysExist(unittest.TestCase):
    """B-3: every call site passes a key the dictionary defines, or none.

    [REASON]: the dictionary lookup returns '' for an unknown key. That is
    what makes a missing `with context` visible, and it is also what makes a
    stale key silent -- a caller passing a deleted key renders «: 3» with no
    text. Deleting the received wording is only safe with this check beside
    it.
    """

    def test_b3_the_received_keys_are_gone(self):
        source = read_template(PARTIAL)
        self.assertNotIn("'received':", source)
        self.assertNotIn('получено не указано', source)
        self.assertNotIn('кирим қилингани кўрсатилмаган', source)

    def test_b3_the_dictionary_still_defines_amount_in_both_languages(self):
        keys = note_keys()
        self.assertEqual({'amount'}, keys['ru'])
        self.assertEqual({'amount'}, keys['uz'])

    def test_b3_every_call_site_passes_a_known_key_or_nothing(self):
        keys = note_keys()
        known = keys['ru'] & keys['uz']
        sites = call_sites()
        self.assertGreater(len(sites), 15, 'the scan found almost no calls')
        offenders = [(path, line, note) for path, line, note in sites
                     if note is not None and note not in known]
        self.assertEqual([], offenders, '\n'.join(
            '%s:%d passes note %r, which the dictionary does not define'
            % (p, l, n) for p, l, n in offenders))

    def test_b3_a_language_that_defines_a_key_the_other_does_not_is_caught(
            self):
        """Both languages must define the same keys, or one renders empty."""
        keys = note_keys()
        self.assertEqual(keys['ru'], keys['uz'])

    def test_b3_the_check_can_fail(self):
        """Negative control, in-process: a planted key is rejected.

        The real plant-and-revert against the template is in the PR body;
        this keeps the control permanently.
        """
        known = note_keys()['ru'] & note_keys()['uz']
        planted = [(REPORTS_TEMPLATE, 999, 'received')]
        offenders = [s for s in planted if s[2] not in known]
        self.assertEqual(1, len(offenders),
                         'a stale key is no longer detected')

    def test_b3_no_call_site_passes_received_any_more(self):
        self.assertEqual([], [(p, l) for p, l, n in call_sites()
                              if n == 'received'])

    def test_b3_the_received_call_sites_can_never_dash_or_note(self):
        """missing = 0 is what makes the call unconditional. Pin it."""
        sites = 0
        for path in (REPORTS_TEMPLATE, DEBTS_TEMPLATE):
            for line in read_template(path).split('\n'):
                if '.received,' not in line or 'money_cell' not in line:
                    continue
                sites += 1
                self.assertRegex(
                    line, r'money_cell\([A-Za-z_.]+\.received, '
                          r'[A-Za-z_.]+\.jobs, 0\)', line.strip())
        self.assertEqual(7, sites, 'expected 7 received cells, saw %d' % sites)


class TestB4NegativeControl(ScreenMixin, unittest.TestCase):
    """B-4: at de97c8d the same fixture renders a dash on both pages."""

    def test_b4_the_pre_commit_markup_renders_a_dash(self):
        """The exact call the two templates carried at de97c8d, rendered
        through the REAL macro against the REAL row."""
        row = by_label(report_data()['by_customer'])[ALL_NULL]
        pre = ("{% from 'drones/_money_cell.html' import money_cell "
               "with context %}"
               "{{ money_cell(r.received, r.jobs, r.no_received) }}")
        from flask import g
        with app.test_request_context('/drones/works/reports'):
            g.lang = 'ru'
            rendered = app.jinja_env.from_string(pre).render(r=row, lang='ru')
        self.assertEqual('—', text_of(rendered))

    def test_b4_the_pre_commit_call_really_is_what_shipped(self):
        """Quoting a call proves nothing unless it is THE call."""
        import subprocess
        for path in (REPORTS_TEMPLATE, DEBTS_TEMPLATE):
            source = subprocess.check_output(
                ['git', 'show', '%s:%s' % (BASE_COMMIT, path)],
                cwd=__import__('tests.test_drone_zero_vs_unknown',
                               fromlist=['REPO_ROOT']).REPO_ROOT
            ).decode('utf-8')
            with self.subTest(template=path):
                self.assertIn("no_received, 'received')", source)

    def test_b4_and_the_current_pages_do_not(self):
        for url, title, _uz, first in PAGES:
            with self.subTest(page=url):
                cells = self.money_columns(url, title, first)
                label = ([k for k in cells if 'Долг неизвестен' in k]
                         or [ALL_NULL])[0]
                self.assertEqual('0', text_of(cells[label][1]))


class TestB5BothLanguages(ScreenMixin, unittest.TestCase):
    """B-5: RU and UZ, both pages, no empty cell in the three columns."""

    def test_b5_no_money_cell_renders_empty(self):
        for url, title_ru, title_uz, first in PAGES:
            for lang, title in (('ru', title_ru), ('uz', title_uz)):
                with self.subTest(page=url, lang=lang):
                    cells = self.money_columns(url, title, first, lang=lang)
                    self.assertTrue(cells, 'the table was not found at all')
                    for label, three in cells.items():
                        for index, cell in enumerate(three):
                            self.assertTrue(
                                text_of(cell),
                                'empty cell: %s / %s / column %d'
                                % (url, label, first + index))

    def test_b5_every_note_that_renders_still_has_wording(self):
        from tests.test_drone_zero_vs_unknown import note_text
        for url, title_ru, title_uz, first in PAGES:
            for lang, title in (('ru', title_ru), ('uz', title_uz)):
                with self.subTest(page=url, lang=lang):
                    for label, cells in self.money_columns(
                            url, title, first, lang=lang).items():
                        for cell in cells:
                            if 'vs-muted' not in cell:
                                continue
                            self.assertTrue(note_text(cell),
                                            'empty wording: %s / %s'
                                            % (url, label))


class TestB6CssScanStillClean(unittest.TestCase):
    """B-6: F-1 still reports zero findings over the touched templates."""

    def test_b6_no_unreachable_class(self):
        from tests.test_drone_zero_vs_unknown import (css_last_compounds,
                                                      unreachable_classes)
        compounds = css_last_compounds()
        offenders = []
        for path in TOUCHED_TEMPLATES:
            for finding in unreachable_classes(path, compounds):
                offenders.append((path,) + finding)
        self.assertEqual([], offenders)


# ---------------------------------------------------------------------------
# C  the workbooks -- §4, what the reviewer will re-check after the merge
# ---------------------------------------------------------------------------

class WorkbookMixin(object):
    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin(cls.__name__.lower()[:24])
        set_language(cls.admin, 'ru')

    def workbook(self, url, module=None):
        """The live workbook, or the one de97c8d's writers would produce."""
        from openpyxl import load_workbook
        if module is None:
            client = app.test_client()
            login(client, self.admin)
            response = client.get(url)
            self.assertEqual(200, response.status_code, url)
            return load_workbook(io.BytesIO(response.data))
        return self.pre_commit_workbook(module)

    def pre_commit_workbook(self, module):
        """de97c8d's sheet writers over the SAME rows, into a real file."""
        from flask import g
        from openpyxl import Workbook, load_workbook
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = 'ru'
            data = module._drone_works_report_data(
                module._drone_work_conditions(
                    module._drone_works_filters_from_args(MultiDict())))
            st = module._drone_xlsx_styler()
            wb = Workbook()
            wb.remove(wb.active)
            module._drone_work_cut_sheet(wb, st, 'По заказчикам', 'Заказчик',
                                         data['by_customer'])
            module._drone_work_debt_sheet(wb, st, 'Долги — заказчики',
                                          'Заказчик',
                                          data['debt_by_customer'])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return load_workbook(buffer)

    def column(self, ws, index):
        """Every body cell of one column, the total row excluded."""
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return [row[index] for row in rows[:-1]]


class TestC1ReceivedHasNoBlanks(WorkbookMixin, PreCommitMixin,
                                unittest.TestCase):
    """C-1: no None in «Получено»; «Сумма» and «Не получено» still have some.

    The counts below are THIS FIXTURE'S, measured here. The reviewer's
    staging figures -- 81 blanks on «Долги — заказчики», 94 on «По
    заказчикам», column sum 1 904 558 438.98 -- were measured on workbooks
    exported from staging at de97c8d and cannot be reproduced from this
    branch.
    """

    def blank_counts(self, wb):
        """{sheet: (blank amounts, blank received, blank outstanding)}."""
        counts = {}
        for title in CUT_SHEETS + DEBT_SHEETS:
            if title not in wb.sheetnames:
                continue
            money = CUT_MONEY if title in CUT_SHEETS else DEBT_MONEY
            counts[title] = tuple(
                sum(1 for v in self.column(wb[title], index) if v is None)
                for index in money)
        return counts

    def test_c1_no_received_cell_is_blank_in_either_workbook(self):
        for url in XLSX_ROUTES:
            wb = self.workbook(url)
            counts = self.blank_counts(wb)
            self.assertTrue(counts, 'no sheet was read at all')
            for title, (_amount, received, _out) in counts.items():
                with self.subTest(url=url, sheet=title):
                    self.assertEqual(0, received)

    def test_c1_the_neighbours_still_have_their_blanks(self):
        """If this moves, the change is too wide."""
        wb = self.workbook(XLSX_ROUTES[0])
        counts = self.blank_counts(wb)
        # «По заказчикам»: ALL_NULL and «Заказчик не указан» are unpriced
        self.assertEqual((2, 0, 2), counts['По заказчикам'])
        # both debt sheets: the collapsed «Долг неизвестен» line
        self.assertEqual((1, 0, 1), counts['Долги — заказчики'])
        self.assertEqual((1, 0, 1), counts['Долги — операторы'])

    def test_c1_the_before_and_after_counts_per_sheet(self):
        """The table the PR body quotes, computed rather than asserted from
        memory. Received blanks go to zero; the other two do not move."""
        module = self.pre_commit()
        before = self.blank_counts(self.pre_commit_workbook(module))
        after = self.blank_counts(self.workbook(XLSX_ROUTES[0]))
        for title in ('По заказчикам', 'Долги — заказчики'):
            with self.subTest(sheet=title):
                self.assertGreater(before[title][1], 0,
                                   'nothing was blank before -- the control '
                                   'is measuring nothing')
                self.assertEqual(0, after[title][1])
                self.assertEqual(before[title][0], after[title][0])
                self.assertEqual(before[title][2], after[title][2])


class TestC2TheColumnSumDoesNotMove(WorkbookMixin, PreCommitMixin,
                                    unittest.TestCase):
    """C-2: an empty cell already counts as zero inside SUM.

    So the arithmetic cannot change and only the statement does. Computed on
    both trees over the same fixture and compared exactly, not approximately:
    a difference of one tiyin is a defect, not a rounding artefact.
    """

    def received_sum(self, wb, title, index):
        return sum(v for v in self.column(wb[title], index) if v is not None)

    def test_c2_the_received_column_sums_identically(self):
        module = self.pre_commit()
        before_wb = self.pre_commit_workbook(module)
        after_wb = self.workbook(XLSX_ROUTES[0])
        for title, index in (('По заказчикам', CUT_MONEY[1]),
                             ('Долги — заказчики', DEBT_MONEY[1])):
            with self.subTest(sheet=title):
                before = self.received_sum(before_wb, title, index)
                after = self.received_sum(after_wb, title, index)
                self.assertEqual(before, after)
                self.assertEqual(10000000.0, after,
                                 'the fixture collected 10 000 000 in total')

    def test_c2_the_total_row_is_unchanged_too(self):
        module = self.pre_commit()
        before_wb = self.pre_commit_workbook(module)
        after_wb = self.workbook(XLSX_ROUTES[0])
        for title, index in (('По заказчикам', CUT_MONEY[1]),
                             ('Долги — заказчики', DEBT_MONEY[1])):
            with self.subTest(sheet=title):
                before = list(before_wb[title].iter_rows(
                    values_only=True))[-1][index]
                after = list(after_wb[title].iter_rows(
                    values_only=True))[-1][index]
                self.assertEqual(before, after)
                self.assertEqual(10000000.0, after)

    def test_c2_the_body_still_adds_up_to_the_total(self):
        """A sum that matched by both columns being wrong would pass above."""
        wb = self.workbook(XLSX_ROUTES[0])
        for title, index in (('По заказчикам', CUT_MONEY[1]),
                             ('Долги — заказчики', DEBT_MONEY[1])):
            with self.subTest(sheet=title):
                total = list(wb[title].iter_rows(values_only=True))[-1][index]
                self.assertAlmostEqual(
                    total, self.received_sum(wb, title, index), places=2)

    def test_c2_the_amount_column_sum_is_unchanged_as_well(self):
        """Nothing outside «Получено» may move."""
        module = self.pre_commit()
        before_wb = self.pre_commit_workbook(module)
        after_wb = self.workbook(XLSX_ROUTES[0])
        for index in (CUT_MONEY[0], CUT_MONEY[2]):
            with self.subTest(column=index):
                self.assertEqual(
                    self.received_sum(before_wb, 'По заказчикам', index),
                    self.received_sum(after_wb, 'По заказчикам', index))


if __name__ == '__main__':
    unittest.main(verbosity=2)
