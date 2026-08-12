# -*- coding: utf-8 -*-
"""DRONE-FLIGHTS-MISSING-001: гектары, летавшие без строки в книгах.

Синтетическая база, где каждое правило посажено нарочно:

  * недостача дня и машины -- вылеты минус работы, привязанные к машине
    через оператора (10 − 6 = 4);
  * граница привязки: работа Усмона 13.04 уходит на №4, 14.04 -- на №8
    (история апрельской сверки 2026-08-03);
  * два оператора на машине в один день -- названы оба; работа оператора,
    привязанного в этот день к ДВУМ машинам, не вычитается ни из одной;
  * два периода одного человека на одной машине -- один кандидат, не
    ложная неоднозначность;
  * многодневная работа лежит целиком в дне начала; внутренний день помечен
    «внутри диапазона», а не объявлен покрытым;
  * работа без даты / с нераспознанным оператором / без привязки на дату --
    в листе «Работы без машины» с причиной, из машин не вычитается;
  * вылет 2026-03-31 22:30 UTC -- это 1 апреля по UTC+5;
  * нераспознанный ник -- строка недостачи целиком, без оператора.

Run:
  python -m unittest tests.test_drone_flights_missing_from_books_001 -v
"""
import contextlib
import hashlib
import io
import os
import sqlite3
import sys
import tempfile
import unittest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO_ROOT, 'tools'))

import drone_flights_missing_from_books as tool  # noqa: E402

BOOK_APRIL = 'Апрель книга.xlsx'
BOOK_MAY = 'Май книга.xlsx'

OP_NURALI = 'Нурали Қодиров'
OP_BEHZOD = 'Беҳзод Имомов'
OP_USMON = 'Усмон Анваров'
OP_SHOHRUH = 'Шохрух Хамроев'


def build_db(path):
    con = sqlite3.connect(path)
    con.execute('CREATE TABLE drone_units ('
                'id INTEGER PRIMARY KEY, number INTEGER)')
    con.execute('CREATE TABLE drone_flights ('
                'id INTEGER PRIMARY KEY, drone_unit_id INTEGER, '
                'nickname_raw TEXT, started_at TEXT, area_ha REAL, '
                'region TEXT, location_text TEXT)')
    con.execute('CREATE TABLE drone_operators ('
                'id INTEGER PRIMARY KEY, full_name TEXT)')
    con.execute('CREATE TABLE drone_operator_assignments ('
                'id INTEGER PRIMARY KEY, operator_id INTEGER, '
                'drone_unit_id INTEGER, date_from TEXT, date_to TEXT)')
    con.execute('CREATE TABLE drone_works ('
                'id INTEGER PRIMARY KEY, period_month TEXT, '
                'work_date_from TEXT, work_date_to TEXT, date_raw TEXT, '
                'customer_raw TEXT, area_ha REAL, subdivision_name TEXT, '
                'operator_raw TEXT, drone_operator_id INTEGER, '
                'source_file TEXT, source_sheet TEXT, source_row INTEGER)')
    con.executemany(
        'INSERT INTO drone_units (id, number) VALUES (?, ?)',
        [(1, 2), (2, 10), (3, 4), (4, 8)])
    con.executemany(
        'INSERT INTO drone_operators (id, full_name) VALUES (?, ?)',
        [(1, OP_NURALI), (2, OP_BEHZOD), (3, OP_USMON), (4, OP_SHOHRUH)])
    con.executemany(
        'INSERT INTO drone_operator_assignments '
        '(operator_id, drone_unit_id, date_from, date_to) '
        'VALUES (?, ?, ?, ?)',
        [
            (1, 1, '2026-03-01', None),           # Нурали на №2, открыт
            # Второй период того же человека на той же машине: один
            # кандидат, а не ложная неоднозначность.
            (1, 1, '2026-04-01', '2026-04-05'),
            (2, 2, '2026-04-01', '2026-04-30'),   # Беҳзод на №10
            # Перекрытие: 10.04 Беҳзод одновременно и на №2 -- у №2 два
            # оператора, у работы Беҳзода две машины.
            (2, 1, '2026-04-10', '2026-04-10'),
            (3, 3, '2026-04-01', '2026-04-13'),   # Усмон: №4 до 13.04...
            (3, 4, '2026-04-14', None),           # ...№8 с 14.04
        ])
    con.executemany(
        'INSERT INTO drone_flights '
        '(drone_unit_id, nickname_raw, started_at, area_ha, region, '
        ' location_text) VALUES (?, ?, ?, ?, ?, ?)',
        [
            (1, None, '2026-04-10 06:00:00.000000', 6.0,
             'Бухоро', 'Гиждуван, массив А'),
            (1, None, '2026-04-10 07:00:00.000000', 4.0,
             'Бухоро', 'Гиждуван, массив Б'),
            (2, None, '2026-04-10 05:00:00.000000', 3.0, 'Когон', None),
            # 22:30 UTC = 03:30 следующего дня по UTC+5: первое апреля.
            (1, None, '2026-03-31 22:30:00.000000', 2.0, None, None),
            (1, None, '2026-04-02 05:00:00.000000', 5.0, None, None),
            (3, None, '2026-04-13 05:00:00.000000', 5.0, None, None),
            (4, None, '2026-04-14 05:00:00.000000', 8.0, None, None),
            (None, '8 GardenU', '2026-05-05 05:00:00.000000', 1.2,
             'Гарден', None),
            # Целиком вне окна: февраль и июль по локальному времени.
            (1, None, '2026-02-20 10:00:00.000000', 9.9, None, None),
            (1, None, '2026-07-01 02:00:00.000000', 9.9, None, None),
        ])
    con.executemany(
        'INSERT INTO drone_works '
        '(id, period_month, work_date_from, work_date_to, date_raw, '
        ' customer_raw, area_ha, subdivision_name, operator_raw, '
        ' drone_operator_id, source_file, source_sheet, source_row) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        [
            # Привязывается к №2: покрыта часть дня 10.04 (10 − 6 = 4).
            (1, '2026-04', '2026-04-10', '2026-04-10', '10.04.2026',
             'Тонг фх', 6.0, None, 'Нурали', 1, BOOK_APRIL, 'свод ичи', 4),
            # У Беҳзода 10.04 ДВЕ машины (№10 и №2): не вычитается нигде.
            (2, '2026-04', '2026-04-10', '2026-04-10', '10.04.2026',
             'Спорный фх', 2.0, None, 'Беҳзод', 2, BOOK_APRIL,
             'свод ичи', 5),
            # Многодневная: целиком в дне начала 01.04 (книги 9 > вылеты 2).
            (3, '2026-04', '2026-04-01', '2026-04-03', '1-3.04.2026',
             'Баркамол фх', 9.0, None, 'Нурали', 1, BOOK_APRIL,
             'свод ичи', 6),
            # Оператор не распознан.
            (4, '2026-04', '2026-04-05', '2026-04-05', '05.04.2026',
             'Номаълум фх', 7.0, None, 'Ким-дир', None, BOOK_APRIL,
             'свод ичи', 7),
            # У Шохруха нет ни одной привязки.
            (5, '2026-04', '2026-04-06', '2026-04-06', '06.04.2026',
             'Шохрух мижози', 3.3, None, 'Шохрух', 4, BOOK_APRIL,
             'свод ичи', 8),
            # Без даты: только месяц книги.
            (6, '2026-05', None, None, None,
             'Оазис', 7.5, None, 'Нурали', 1, BOOK_MAY, 'свод ичи', 9),
            # Граница привязки Усмона: 13.04 -> №4 (покрывает день целиком),
            # 14.04 -> №8 (8 − 3 = 5).
            (7, '2026-04', '2026-04-13', '2026-04-13', '13.04.2026',
             'Учун фх', 5.0, None, 'Усмон', 3, BOOK_APRIL, 'свод ичи', 10),
            (8, '2026-04', '2026-04-14', '2026-04-14', '14.04.2026',
             'Кейин фх', 3.0, None, 'Усмон', 3, BOOK_APRIL, 'свод ичи', 11),
            # Вне окна: июльская дата в июньской книге и февральская книга.
            (9, '2026-06', '2026-07-02', '2026-07-02', '02.07.2026',
             'Июл фх', 9.9, None, None, None, 'Июнь.xlsx', 'свод ичи', 3),
            (10, '2026-02', None, None, None,
             'Феврал фх', 9.9, None, None, None, 'Февраль.xlsx',
             'свод ичи', 3),
        ])
    con.commit()
    con.close()


class MissingFromBooksTests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix='fl_missing_')
        cls.db = os.path.join(cls.tmp, 'transport.db')
        cls.out = os.path.join(cls.tmp, 'report.xlsx')
        build_db(cls.db)
        with open(cls.db, 'rb') as fh:
            cls.db_sha_before = hashlib.sha256(fh.read()).hexdigest()
        cls.stdout = io.StringIO()
        with contextlib.redirect_stdout(cls.stdout):
            cls.code = tool.main(['--db', cls.db, '--out', cls.out])
        from openpyxl import load_workbook
        cls.wb = load_workbook(cls.out)

    # -- строки листов --------------------------------------------------

    def day_rows(self):
        ws = self.wb['Нет в книгах по дням']
        return {(row[1], row[2]): row
                for row in ws.iter_rows(min_row=3, values_only=True)
                if row[1]}

    def month_row(self, month, label):
        ws = self.wb['По месяцам и машинам']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == month and row[1] == label:
                return row
        self.fail('row (%s, %s) not found in the monthly sheet'
                  % (month, label))

    def summary_row(self, month):
        ws = self.wb['Сводка']
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] == month:
                return row
        self.fail('month %s not found in the summary' % month)

    def unmapped_rows(self):
        ws = self.wb['Работы без машины']
        return {row[1]: row
                for row in ws.iter_rows(min_row=3, values_only=True)
                if row[1] is not None}

    # -- каркас -----------------------------------------------------------

    def test_exit_code_zero_and_database_untouched(self):
        self.assertEqual(self.code, 0)
        with open(self.db, 'rb') as fh:
            self.assertEqual(hashlib.sha256(fh.read()).hexdigest(),
                             self.db_sha_before,
                             'read-only инструмент изменил файл базы')

    def test_console_is_ascii_only(self):
        text = self.stdout.getvalue()
        self.assertTrue(text.strip())
        self.assertTrue(all(ord(ch) < 128 for ch in text),
                        'в консоли не-ASCII: %r' % text[:120])

    def test_offset_is_shared_with_the_daily_tool_and_the_application(self):
        """Смещение не дублируется: одна константа на оба инструмента."""
        import drone_flights_vs_books_daily as daily
        self.assertIs(tool.UTC_OFFSET_MINUTES, daily.UTC_OFFSET_MINUTES)
        with open(os.path.join(REPO_ROOT, 'drones.py'),
                  encoding='utf-8') as fh:
            source = fh.read()
        self.assertIn('DRONE_DISPLAY_UTC_OFFSET = timedelta(hours=5)', source)
        self.assertEqual(tool.UTC_OFFSET_MINUTES, 300)

    # -- ядро: недостача по дням и машинам ---------------------------------

    def test_day_sheet_contains_exactly_the_uncovered_rows(self):
        rows = self.day_rows()
        self.assertEqual(
            set(rows), {
                ('2026-04-02', '№2'),
                ('2026-04-10', '№2'),
                ('2026-04-10', '№10'),
                ('2026-04-14', '№8'),
                ('2026-05-05', 'ник: 8 GardenU'),
            })

    def test_partial_day_missing_is_flights_minus_mapped_books(self):
        row = self.day_rows()[('2026-04-10', '№2')]
        self.assertAlmostEqual(row[4], 10.0)   # га по вылетам
        self.assertEqual(row[5], 2)            # вылетов
        self.assertAlmostEqual(row[6], 6.0)    # га в книгах
        self.assertAlmostEqual(row[7], 4.0)    # нет в книгах
        self.assertIn('Тонг фх (6.00)', row[8])

    def test_overlap_day_names_both_operators(self):
        row = self.day_rows()[('2026-04-10', '№2')]
        self.assertEqual(row[3], '%s; %s' % (OP_BEHZOD, OP_NURALI))

    def test_two_periods_of_one_man_are_one_candidate(self):
        row = self.day_rows()[('2026-04-02', '№2')]
        self.assertEqual(row[3], OP_NURALI,
                         'два периода одного человека дали дубль имени')

    def test_ambiguous_work_is_subtracted_from_no_machine(self):
        # Работа Беҳзода (2 га) 10.04 могла бы уйти на №10 или №2 --
        # не ушла ни туда, ни туда.
        self.assertAlmostEqual(self.day_rows()[('2026-04-10', '№10')][7], 3.0)
        self.assertAlmostEqual(self.day_rows()[('2026-04-10', '№2')][7], 4.0)
        row = self.unmapped_rows()[2]
        self.assertEqual(row[0], tool.REASON_MANY_UNITS)

    def test_assignment_boundary_moves_the_work_between_machines(self):
        # 13.04: работа Усмона покрывает №4 целиком -- строки недостачи нет.
        self.assertNotIn(('2026-04-13', '№4'), self.day_rows())
        # 14.04: та же пара оператор-дата на день позже -- уже №8, 8 − 3 = 5.
        row = self.day_rows()[('2026-04-14', '№8')]
        self.assertEqual(row[3], OP_USMON)
        self.assertAlmostEqual(row[7], 5.0)

    def test_multiday_start_day_is_covered_inner_day_is_annotated(self):
        rows = self.day_rows()
        # 01.04: книги (9) больше вылетов (2 -- пограничный вылет 22:30 UTC),
        # строки недостачи нет: это и проверка границы дня UTC+5.
        self.assertNotIn(('2026-04-01', '№2'), rows)
        inner = rows[('2026-04-02', '№2')]
        self.assertAlmostEqual(inner[7], 5.0)
        self.assertIn('внутри диапазона работы ID 3', inner[11])
        self.assertIn(BOOK_APRIL, inner[11])
        self.assertIn('2026-04-01..2026-04-03', inner[11])

    def test_unrecognized_nickname_is_fully_missing_without_operator(self):
        row = self.day_rows()[('2026-05-05', 'ник: 8 GardenU')]
        self.assertAlmostEqual(row[7], 1.2)
        # Пустая строка в xlsx читается как None: клетка честно пуста.
        self.assertIn(row[3], (None, ''))
        self.assertEqual(row[9], 'Гарден')

    def test_places_come_from_dji(self):
        row = self.day_rows()[('2026-04-10', '№2')]
        self.assertEqual(row[9], 'Бухоро')
        self.assertIn('Гиждуван, массив А', row[10])
        self.assertIn('Гиждуван, массив Б', row[10])

    # -- работы без машины ---------------------------------------------------

    def test_unmapped_works_carry_their_reasons(self):
        rows = self.unmapped_rows()
        self.assertEqual(set(rows), {2, 4, 5, 6})
        self.assertEqual(rows[4][0], tool.REASON_NO_OPERATOR)
        self.assertEqual(rows[5][0], tool.REASON_NO_COVER)
        self.assertEqual(rows[6][0], tool.REASON_NO_DATE)
        self.assertEqual(rows[6][2], '2026-05')
        self.assertEqual((rows[5][9], rows[5][11]), (BOOK_APRIL, 8))

    # -- месяц х машина и сводка ------------------------------------------

    def test_monthly_machine_rows_and_buckets_cross_foot(self):
        n2 = self.month_row('2026-04', '№2')
        self.assertEqual(n2[3], 4)                 # вылетов
        self.assertAlmostEqual(n2[4], 17.0)        # га по вылетам
        self.assertEqual(n2[5], 2)                 # работ привязано
        self.assertAlmostEqual(n2[6], 15.0)        # га привязано
        self.assertAlmostEqual(n2[7], 2.0)         # нетто
        self.assertAlmostEqual(n2[8], 9.0)         # недостача по дням
        n4 = self.month_row('2026-04', '№4')
        self.assertAlmostEqual(n4[7], 0.0)
        self.assertAlmostEqual(n4[8], 0.0)
        bucket = self.month_row(
            '2026-04', '(не привязано: %s)' % tool.REASON_MANY_UNITS)
        self.assertEqual(bucket[5], 1)
        self.assertAlmostEqual(bucket[6], 2.0)
        total = self.month_row('2026-04', 'ИТОГО МЕСЯЦА')
        self.assertEqual(total[3], 7)
        self.assertAlmostEqual(total[4], 33.0)
        self.assertAlmostEqual(total[6], 35.3)
        self.assertAlmostEqual(total[7], -2.3)
        self.assertAlmostEqual(total[8], 17.0)

    def test_summary_reconciles_and_matches_the_daily_report_sign(self):
        april = self.summary_row('2026-04')
        self.assertAlmostEqual(april[1], 33.0)
        self.assertAlmostEqual(april[2], 35.3)
        self.assertAlmostEqual(april[3], 23.0)     # привязано
        self.assertAlmostEqual(april[4], 12.3)     # не привязано
        self.assertAlmostEqual(april[2], april[3] + april[4])
        self.assertAlmostEqual(april[5], -2.3)     # нетто = выл − вед
        self.assertAlmostEqual(april[6], 17.0)
        may = self.summary_row('2026-05')
        self.assertAlmostEqual(may[1], 1.2)
        self.assertAlmostEqual(may[2], 7.5)
        self.assertAlmostEqual(may[6], 1.2)
        march = self.summary_row('2026-03')
        self.assertEqual(march[7], tool.NOTE_NO_FLIGHTS)

    def test_window_excludes_by_the_screen_month_rule(self):
        ws = self.wb['Работы без машины']
        ids = [row[1] for row in ws.iter_rows(min_row=3, values_only=True)]
        self.assertNotIn(9, ids, 'июльская дата в июньской книге -- вне окна')
        self.assertNotIn(10, ids, 'февральская книга -- вне окна')
        june = self.summary_row('2026-06')
        self.assertAlmostEqual(june[1], 0.0)
        self.assertAlmostEqual(june[2], 0.0)

    # -- отказы ------------------------------------------------------------

    def test_missing_database_exits_two(self):
        code = tool.main(['--db', os.path.join(self.tmp, 'absent.db'),
                          '--out', os.path.join(self.tmp, 'x.xlsx')])
        self.assertEqual(code, 2)

    def test_bad_month_arguments_exit_one(self):
        self.assertEqual(tool.main(['--db', self.db, '--month-from', '2026',
                                    '--out', self.out]), 1)
        self.assertEqual(tool.main(['--db', self.db,
                                    '--month-from', '2026-05',
                                    '--month-to', '2026-03',
                                    '--out', self.out]), 1)


if __name__ == '__main__':
    unittest.main()
