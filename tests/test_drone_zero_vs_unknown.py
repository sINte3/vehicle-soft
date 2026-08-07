# -*- coding: utf-8 -*-
"""DRONE-ZERO-VS-UNKNOWN-001 -- «ноль» и «не записано» перестают быть одним
пикселем в трёх колонках и в третьей корзине долгов.

  A  the counters              _drone_work_cut() also counts missing received
  B  the shared macro          one money_cell(), two templates, `with context`
  C  the two screens           three columns obey the rule of §2
  D  the third bucket          «долг неизвестен» leaves «долга нет»
  E  the workbooks             an unknown figure is an EMPTY cell, never 0

What these assertions can and cannot prove:

  * Every negative control here re-runs the REAL pre-commit code, loaded out
    of git at BASE_COMMIT as a second module object, against the SAME fixture
    rows -- not a paraphrase of it. Where git cannot produce that commit the
    control skips loudly instead of passing quietly; a skipped control is not
    a passed one and the PR body says so.
  * B-2 is the `with context` check. It is written so that deleting
    `with context` from either import makes it fail, and that was verified by
    deleting it once -- see the PR body for the raw output.
  * Nothing here was run against production. The production figures quoted in
    the task (904 jobs, 852 with a received figure, 52 without) are the task's
    numbers; this branch has no access to that database.

Run:
  python -m unittest tests.test_drone_zero_vs_unknown -v
"""
import importlib.util
import io
import os
import re
import subprocess
import sys
import tempfile
import unittest

from tests.harness import app, reset_db, create_admin, login
from models import db, DroneCustomer, DroneCustomerAlias, DroneOperator, \
    DroneWork, User

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# The merge base this task was written against, named in the task itself.
# The negative controls load drones.py OUT OF THIS COMMIT and run it against
# the same fixture, so «the pre-commit tree renders 0» is measured and not
# asserted from memory.
BASE_COMMIT = '116e15a'


# ── the fixture ──────────────────────────────────────────────────────────────
# Six customer groups, chosen so that every case the three columns and the
# three buckets have to tell apart is present exactly once.
#
#   group          jobs  amount           received        no_amount no_received
#   ALL_NULL         2   NULL, NULL       NULL, NULL          2         2
#   MIXED            2   NULL, 5 000 000  1 000 000,          1         0
#                                          2 000 000
#   NONE_NULL        2   3 000 000,       3 000 000,          0         0
#                        4 000 000        4 000 000
#   NO_RECEIVED      2   6 000 000,       NULL, NULL          0         2
#                        7 000 000
#   UNSTATED  (svc) 1    NULL             NULL                1         1
#   UNRESOLVED(svc) 1    1 000 000        NULL                0         1
ALL_NULL = 'Бухоро Агрокластер Заминлари МЧЖ'
MIXED = 'Дўстлик ФХ'
NONE_NULL = 'Пешку АМТ'
NO_RECEIVED = 'Миробод АМТ'
UNSTATED_LABEL_RU = 'Заказчик не указан'
UNRESOLVED_LABEL_RU = 'Заказчик не определён'
UNSTATED_LABEL_UZ = 'Буюртмачи кўрсатилмаган'
UNRESOLVED_LABEL_UZ = 'Буюртмачи аниқланмаган'

# (amount, received, operator) per job, in group order.
#
# [REASON]: the operator is assigned PER JOB, not per group, and deliberately
# cuts across the customer grouping. With one operator per customer the two
# cuts are the same partition under two names, and three of the things this
# task has to get right are never exercised: the operator debt cut had no
# «unknown» bucket at all, no «settled» bucket, and no group with
# 0 < no_received < jobs -- so the «получено не указано» NOTE (as opposed to
# the dash) rendered nowhere in it. The assignment below gives the operator
# cut all three buckets and that partial group:
#
#   op1  MIXED x2 + NO_RECEIVED's second job   3 jobs  no_received 1  -> note
#   op2  NO_RECEIVED's first job + UNRESOLVED  2 jobs  owing
#   op3  ALL_NULL x2 + UNSTATED                3 jobs  unknown
#   op4  NONE_NULL x2                          2 jobs  settled
GROUPS = (
    (ALL_NULL, ((None, None, 'op3'), (None, None, 'op3'))),
    (MIXED, ((None, 1000000.0, 'op1'), (5000000.0, 2000000.0, 'op1'))),
    (NONE_NULL, ((3000000.0, 3000000.0, 'op4'),
                 (4000000.0, 4000000.0, 'op4'))),
    (NO_RECEIVED, ((6000000.0, None, 'op2'), (7000000.0, None, 'op1'))),
)

# What the operator cut must look like, so a change to the assignment above
# that quietly collapses a bucket is caught rather than absorbed.
#           jobs amount     received  no_amount no_received
OPERATORS = {
    'op1': (3, 12000000.0, 3000000.0, 1, 1),
    'op2': (2, 7000000.0, 0.0, 0, 2),
    'op3': (3, 0.0, 0.0, 3, 3),
    'op4': (2, 7000000.0, 7000000.0, 0, 0),
}
OPERATOR_NAMES = {
    'op1': 'Файзуллаев Фурқат',
    'op2': 'Хамроев Шохрух',
    'op3': 'Холмуродов Шахзод',
    'op4': 'Болтаев Шахзод',
}

EXPECTED = {
    #             jobs amount     received   no_amount no_received
    ALL_NULL: (2, 0.0, 0.0, 2, 2),
    MIXED: (2, 5000000.0, 3000000.0, 1, 0),
    NONE_NULL: (2, 7000000.0, 7000000.0, 0, 0),
    NO_RECEIVED: (2, 13000000.0, 0.0, 0, 2),
    UNSTATED_LABEL_RU: (1, 0.0, 0.0, 1, 1),
    UNRESOLVED_LABEL_RU: (1, 1000000.0, 0.0, 0, 1),
}

TOTAL_JOBS = 10
TOTAL_AMOUNT = 26000000.0
TOTAL_RECEIVED = 10000000.0
TOTAL_OUTSTANDING = 16000000.0
TOTAL_NO_AMOUNT = 4
TOTAL_NO_RECEIVED = 6


def seed():
    """Build the fixture in the disposable test database."""
    reset_db()
    with app.app_context():
        operators = {}
        for key in sorted(OPERATOR_NAMES):
            row = DroneOperator(full_name=OPERATOR_NAMES[key],
                                subdivision_name='Гарден')
            db.session.add(row)
            db.session.flush()
            operators[key] = row.id

        def add(amount, received, customer_id, customer_raw, operator):
            db.session.add(DroneWork(
                period_month='2026-04',
                drone_customer_id=customer_id,
                customer_raw=customer_raw,
                drone_operator_id=operators[operator],
                operator_raw=operator,
                area_ha=10.0, amount=amount, received_amount=received,
                payment_type='cash', subdivision_name='Гарден',
                source_file='fixture.xlsx', source_sheet='свод ичи',
                import_batch='zero-vs-unknown'))

        for name, jobs in GROUPS:
            customer = DroneCustomer(name=name)
            db.session.add(customer)
            db.session.flush()
            db.session.add(DroneCustomerAlias(
                raw_name=name, normalized_key=name.casefold(),
                drone_customer_id=customer.id, is_active=True))
            for amount, received, operator in jobs:
                add(amount, received, customer.id, name, operator)

        # [REASON]: the two service rows are DIFFERENT facts and the cut keeps
        # them apart -- «имени в ведомости не было» (the holding's own land,
        # customer_raw = '') versus «написание есть, алиаса нет». Both have to
        # be in the fixture, because both are ordinary groups as far as the
        # three columns and the three buckets are concerned, and a fixture
        # that only exercised named customers would never render them.
        add(None, None, None, '', 'op3')
        add(1000000.0, None, None, 'Хўжалик номаълум', 'op2')
        db.session.commit()


def set_language(user_id, lang):
    with app.app_context():
        user = db.session.get(User, user_id)
        user.language = lang
        db.session.commit()


def report_data(lang='ru'):
    """The structure both screens and both workbooks are built from."""
    import drones
    from flask import g
    from werkzeug.datastructures import MultiDict
    with app.test_request_context('/drones/works/reports'):
        g.lang = lang
        return drones._drone_works_report_data(
            drones._drone_work_conditions(
                drones._drone_works_filters_from_args(MultiDict())))


def by_label(cut):
    """{label: row} over the ordinary rows AND the service rows of a cut."""
    return {r['label']: r for r in cut['rows'] + list(cut['services'])}


_CELL_RE = re.compile(r'<td[^>]*>(.*?)</td>', re.S)


def cut_cells(body, title, column):
    """{row label: one <td> as rendered} for one table of a rendered page.

    column is the 0-based index inside the row. The report cuts run
    label, jobs, area, amount, received, outstanding, share; the debt tables
    run label, jobs, amount, received, outstanding.
    """
    after = body.split(title, 1)[1]
    tbody = after.split('<tbody>')[1].split('</tbody>')[0]
    cells = {}
    for row in tbody.split('<tr'):
        tds = _CELL_RE.findall(row)
        if len(tds) <= column:
            continue
        label = text_of(tds[0])
        cells[label] = tds[column]
    return cells


# ── the pre-commit tree, loaded out of git ───────────────────────────────────
_PRE_COMMIT_CACHE = {}


def pre_commit_drones(commit=None):
    """drones.py as of `commit`, imported as a SEPARATE module object.

    [REASON]: a negative control that quotes the old code into the test file
    proves only that the quote differs. This loads the real pre-commit module
    and runs it against the same rows in the same session, so «the pre-commit
    tree renders 0 here» is a measurement.

    The module builds its own Blueprint('drones') object at import time; it is
    never registered on the app, so the live blueprint is untouched. models.db
    is shared, which is exactly what makes the comparison meaningful.

    `commit` defaults to BASE_COMMIT, the base of DRONE-ZERO-VS-UNKNOWN-001.
    DRONE-RECEIVED-BLANK-IS-ZERO-001 passes its own base instead: each task's
    control has to run against the tree IT changed, or it measures somebody
    else's commit. One loader, one cache keyed by commit, two callers.

    Returns None when git cannot produce the blob (a shallow clone, a tarball
    export). Callers skip in that case -- loudly, never silently passing.
    """
    commit = commit or BASE_COMMIT
    if commit in _PRE_COMMIT_CACHE:
        return _PRE_COMMIT_CACHE[commit]
    _PRE_COMMIT_CACHE[commit] = None
    try:
        source = subprocess.check_output(
            ['git', 'show', '%s:drones.py' % commit],
            cwd=REPO_ROOT, stderr=subprocess.DEVNULL)
    except (OSError, subprocess.CalledProcessError):
        return None
    handle, path = tempfile.mkstemp(prefix='drones_pre_', suffix='.py')
    with os.fdopen(handle, 'wb') as fh:
        fh.write(source)
    name = 'drones_at_%s' % commit
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    _PRE_COMMIT_CACHE[commit] = module
    return module


class PreCommitMixin(object):
    def pre_commit(self):
        module = pre_commit_drones()
        if module is None:
            self.skipTest('git cannot produce %s:drones.py -- the negative '
                          'control did NOT run' % BASE_COMMIT)
        return module

    def pre_commit_report_data(self, lang='ru'):
        module = self.pre_commit()
        from flask import g
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = lang
            return module._drone_works_report_data(
                module._drone_work_conditions(
                    module._drone_works_filters_from_args(MultiDict())))


# ---------------------------------------------------------------------------
# A  commit 1 -- the counters
# ---------------------------------------------------------------------------

class TestA1Counters(unittest.TestCase):
    """A-1: no_amount and no_received on every row, service row and total."""

    @classmethod
    def setUpClass(cls):
        seed()

    def setUp(self):
        self.cut = report_data()['by_customer']
        self.rows = by_label(self.cut)

    def test_a1_every_group_carries_both_counters(self):
        for label, (jobs, amount, received, no_amount, no_received) in \
                EXPECTED.items():
            with self.subTest(group=label):
                row = self.rows[label]
                self.assertEqual(jobs, row['jobs'])
                self.assertAlmostEqual(amount, row['amount'], places=2)
                self.assertAlmostEqual(received, row['received'], places=2)
                self.assertEqual(no_amount, row['no_amount'])
                self.assertEqual(no_received, row['no_received'])

    def test_a1_the_four_cases_really_are_four_different_cases(self):
        """A fixture where two cases coincide proves nothing about either."""
        shapes = {(self.rows[label]['no_amount'] == self.rows[label]['jobs'],
                   self.rows[label]['no_received'] == self.rows[label]['jobs'],
                   bool(self.rows[label]['no_amount']))
                  for label in (ALL_NULL, MIXED, NONE_NULL, NO_RECEIVED)}
        self.assertEqual(4, len(shapes), shapes)

    def test_a1_the_service_rows_carry_them_too(self):
        services = {r['label']: r for r in self.cut['services']}
        self.assertEqual({UNSTATED_LABEL_RU, UNRESOLVED_LABEL_RU},
                         set(services))
        self.assertEqual(1, services[UNSTATED_LABEL_RU]['no_received'])
        self.assertEqual(1, services[UNSTATED_LABEL_RU]['no_amount'])
        self.assertEqual(1, services[UNRESOLVED_LABEL_RU]['no_received'])
        self.assertEqual(0, services[UNRESOLVED_LABEL_RU]['no_amount'])

    def test_a1_the_total_carries_them(self):
        total = self.cut['total']
        self.assertEqual(TOTAL_JOBS, total['jobs'])
        self.assertEqual(TOTAL_NO_AMOUNT, total['no_amount'])
        self.assertEqual(TOTAL_NO_RECEIVED, total['no_received'])

    def test_a1_the_operator_cut_is_the_partition_the_fixture_claims(self):
        """The operator dimension cuts ACROSS the customer one on purpose.

        If this drifts back to one-operator-per-customer, the operator debt
        cut silently loses its «unknown» and «settled» buckets and its only
        partially-missing «Получено» group, and D-1, D-2 and C-1 stop
        exercising them without any of them failing.
        """
        rows = by_label(report_data()['by_operator'])
        for key, (jobs, amount, received, no_amount, no_received) in \
                OPERATORS.items():
            with self.subTest(operator=key):
                row = rows[OPERATOR_NAMES[key]]
                self.assertEqual(jobs, row['jobs'])
                self.assertAlmostEqual(amount, row['amount'], places=2)
                self.assertAlmostEqual(received, row['received'], places=2)
                self.assertEqual(no_amount, row['no_amount'])
                self.assertEqual(no_received, row['no_received'])


class TestA2CountersAddUp(unittest.TestCase):
    """A-2: a counter that does not add up is measuring something else."""

    CUTS = ('by_customer', 'by_operator', 'by_subdivision', 'by_month',
            'by_payment')

    @classmethod
    def setUpClass(cls):
        seed()

    def setUp(self):
        self.data = report_data()

    def test_a2_each_total_equals_the_sum_over_its_groups(self):
        for name in self.CUTS:
            with self.subTest(cut=name):
                cut = self.data[name]
                everything = cut['rows'] + list(cut['services'])
                self.assertEqual(cut['total']['no_amount'],
                                 sum(r['no_amount'] for r in everything))
                self.assertEqual(cut['total']['no_received'],
                                 sum(r['no_received'] for r in everything))

    def test_a2_every_cut_agrees_on_the_two_counts(self):
        counts = {name: (self.data[name]['total']['no_amount'],
                         self.data[name]['total']['no_received'])
                  for name in self.CUTS}
        self.assertEqual(
            {name: (TOTAL_NO_AMOUNT, TOTAL_NO_RECEIVED) for name in counts},
            counts)

    def test_a2_the_money_totals_are_untouched_by_the_new_column(self):
        """Adding a counter must not disturb the arithmetic it sits beside."""
        for name in self.CUTS:
            with self.subTest(cut=name):
                total = self.data[name]['total']
                self.assertAlmostEqual(TOTAL_AMOUNT, total['amount'], places=2)
                self.assertAlmostEqual(TOTAL_RECEIVED, total['received'],
                                       places=2)
                self.assertAlmostEqual(TOTAL_OUTSTANDING,
                                       total['outstanding'], places=2)
                self.assertTrue(self.data[name]['reconciled'], name)

    def test_a2_rest_carries_both_counters(self):
        """`rest` is assembled by naming its keys, so a counter reaches it
        only if somebody named it. no_amount never was."""
        rest = self.data['debt_by_customer']['rest']
        self.assertIsNotNone(rest)
        self.assertIn('no_amount', rest)
        self.assertIn('no_received', rest)


class TestA3NegativeControl(PreCommitMixin, unittest.TestCase):
    """A-3: the same assertions fail against the real pre-commit module."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_a3_pre_commit_rows_have_no_no_received(self):
        cut = self.pre_commit_report_data()['by_customer']
        row = by_label(cut)[ALL_NULL]
        self.assertEqual(2, row['no_amount'], 'the fixture is the same one')
        with self.assertRaises(KeyError):
            row['no_received']

    def test_a3_pre_commit_total_has_no_no_received(self):
        cut = self.pre_commit_report_data()['by_customer']
        with self.assertRaises(KeyError):
            cut['total']['no_received']

    def test_a3_pre_commit_rest_drops_the_counter(self):
        rest = self.pre_commit_report_data()['debt_by_customer']['rest']
        self.assertIsNotNone(rest, 'the settled bucket exists pre-commit')
        self.assertNotIn('no_amount', rest)
        self.assertNotIn('no_received', rest)

    def test_a3_the_control_is_reading_the_other_module(self):
        """Guard against the control silently importing the live module."""
        import drones
        module = self.pre_commit()
        self.assertIsNot(module, drones)
        self.assertNotEqual(module.__file__, drones.__file__)


# ---------------------------------------------------------------------------
# B  commit 2 -- one macro, shared by two templates
# ---------------------------------------------------------------------------

PARTIAL = 'templates/drones/_money_cell.html'
REPORTS_TEMPLATE = 'templates/drones/works_reports.html'
DEBTS_TEMPLATE = 'templates/drones/works_debts.html'
MACRO_TEMPLATES = (REPORTS_TEMPLATE, DEBTS_TEMPLATE)

IMPORT_LINE = ("{% from 'drones/_money_cell.html' import money_cell "
               "with context %}")


def read_template(relative_path):
    with io.open(os.path.join(REPO_ROOT, relative_path), encoding='utf-8') as f:
        return f.read()


def render_macro(call, lang='ru', with_context=True):
    """Render one money_cell() call through the REAL application Jinja env.

    with_context=False reproduces exactly the mistake §4 of the task warns
    about, so the check that catches it can be shown to catch it.
    """
    from flask import g
    source = ("{%% set is_ru = (lang == 'ru') %%}"
              "{%% from 'drones/_money_cell.html' import money_cell%s %%}"
              "%s") % (' with context' if with_context else '', call)
    with app.test_request_context('/drones/works/reports'):
        g.lang = lang
        return app.jinja_env.from_string(source).render(lang=lang)


# [REASON]: NBSP is NOT whitespace here. str.split() eats U+00A0, so the
# obvious ' '.join(text.split()) turns «5\xa0000\xa0000» into «5 000 000» and
# every assertion that looks for the grouped number then matches the ungrouped
# one -- a check that gives the same answer for right and wrong code. The
# grouping separator IS the vs_num filter's output and has to survive.
_WS_RE = re.compile(r'[^\S\xa0]+')
_NOTE_RE = re.compile(r'<div class="vs-muted"[^>]*>(.*?)</div>', re.S)


def text_of(html):
    return _WS_RE.sub(' ', re.sub(r'<[^>]+>', ' ', html)).strip()


def note_text(rendered):
    """The muted note's WORDING alone: «сумма не указана: 1» -> the words.

    Returns '' when there is no note at all and when the note rendered with
    an empty wording -- the two ways this can be broken, kept apart by the
    caller asserting which one it expects.
    """
    match = _NOTE_RE.search(rendered)
    if match is None:
        return ''
    return _WS_RE.sub(' ', match.group(1)).rsplit(':', 1)[0].strip()


class TestB1TheFourCombinations(unittest.TestCase):
    """B-1: the partial rendered directly, one case per row of §2's table."""

    def test_b1_missing_equals_jobs_is_a_dash(self):
        out = render_macro("{{ money_cell(0, 2, 2, 'amount') }}")
        self.assertEqual('—', text_of(out))
        self.assertNotIn('0', text_of(out))

    def test_b1_partial_with_a_note_shows_the_value_and_the_count(self):
        out = render_macro("{{ money_cell(5000000, 2, 1, 'amount') }}")
        self.assertIn('5\xa0000\xa0000', out)
        self.assertIn('сумма не указана', text_of(out))
        self.assertTrue(text_of(out).endswith('1'), text_of(out))

    def test_b1_partial_without_a_note_shows_the_value_only(self):
        """«Не получено» keys off no_amount and gets no note of its own."""
        out = render_macro("{{ money_cell(5000000, 2, 1) }}")
        self.assertIn('5\xa0000\xa0000', out)
        self.assertNotIn('vs-muted', out)
        self.assertEqual('5\xa0000\xa0000', text_of(out))

    def test_b1_missing_zero_shows_the_value_only(self):
        out = render_macro("{{ money_cell(7000000, 2, 0, 'amount') }}")
        self.assertIn('7\xa0000\xa0000', out)
        self.assertNotIn('vs-muted', out)

    def test_b1_the_four_renderings_are_four(self):
        rendered = {text_of(render_macro(call)) for call in (
            "{{ money_cell(0, 2, 2, 'amount') }}",
            "{{ money_cell(5000000, 2, 1, 'amount') }}",
            "{{ money_cell(5000000, 2, 1) }}",
            "{{ money_cell(7000000, 2, 0, 'amount') }}")}
        self.assertEqual(4, len(rendered), rendered)

    def test_b1_a_partial_group_is_never_a_dash(self):
        """Two of six missing is a lower bound with a count, not a dash."""
        out = render_macro("{{ money_cell(12000000, 6, 2, 'amount') }}")
        self.assertNotEqual('—', text_of(out))
        self.assertIn('12\xa0000\xa0000', out)
        self.assertTrue(text_of(out).endswith('2'), text_of(out))

    def test_b1_zero_jobs_is_not_a_dash(self):
        """jobs == 0 and missing == 0 are equal but mean «nothing here»."""
        self.assertEqual('0', text_of(render_macro(
            "{{ money_cell(0, 0, 0, 'amount') }}")))


class TestB2WithContextIsLoadBearing(unittest.TestCase):
    """B-2: the note is non-empty in both languages, and empty without
    `with context`.

    [REASON]: the note text is selected by a DICTIONARY LOOKUP on `lang`
    rather than by `if is_ru`. With a ternary, a macro imported without
    context evaluates `lang == 'ru'` to False and silently renders the Uzbek
    branch -- non-empty, plausible, and wrong in Russian. The lookup makes the
    same mistake produce an empty string, which is what these assertions can
    see. Verified by deleting `with context` once; the raw output is in the
    PR body.
    """

    NOTE_CALL = "{{ money_cell(5000000, 2, 1, '%s') }}"

    def test_b2_the_note_is_non_empty_in_both_languages(self):
        # NARROWED by DRONE-RECEIVED-BLANK-IS-ZERO-001: the loop used to run
        # ('amount', 'received'). «Получено» has no note any more and its two
        # strings were deleted from the dictionary, so asking for them here
        # would assert that a deliberately removed key still resolves.
        for lang in ('ru', 'uz'):
            for note in ('amount',):
                with self.subTest(lang=lang, note=note):
                    wording = note_text(
                        render_macro(self.NOTE_CALL % note, lang=lang))
                    self.assertTrue(wording,
                                    'empty note for %s/%s' % (lang, note))
                    self.assertNotIn(':', wording)

    def test_b2_the_two_notes_are_two_distinct_strings(self):
        """Was four: two keys x two languages. One key remains."""
        seen = {note_text(render_macro(self.NOTE_CALL % note, lang=lang))
                for lang in ('ru', 'uz') for note in ('amount',)}
        self.assertEqual(2, len(seen), seen)

    def test_b2_without_with_context_the_note_is_empty(self):
        """The negative control for the control: this is the failure mode."""
        for lang in ('ru', 'uz'):
            with self.subTest(lang=lang):
                rendered = render_macro(self.NOTE_CALL % 'amount', lang=lang,
                                        with_context=False)
                self.assertIn('vs-muted', rendered,
                              'the note div is still emitted -- only its '
                              'wording is lost')
                self.assertEqual('', note_text(rendered),
                                 'the check cannot tell the two cases apart')
                # the value itself still renders -- only the wording is lost,
                # which is why nothing else on the page would look wrong
                self.assertIn('5\xa0000\xa0000', rendered)

    def test_b2_the_note_is_non_empty_rendered_through_the_real_page(self):
        """Not the macro in isolation -- the page, through the real route.

        The MIXED group is the only one with 0 < no_amount < jobs, so its
        «Сумма» cell is the one carrying a note. Commit 3 extends this to the
        debts page and to «Получено» (C-4).
        """
        seed()
        admin = create_admin('b2_admin')
        for lang, title in (('ru', 'По заказчикам'),
                            ('uz', 'Буюртмачилар бўйича')):
            with self.subTest(lang=lang):
                set_language(admin, lang)
                client = app.test_client()
                login(client, admin)
                body = client.get(
                    '/drones/works/reports').data.decode('utf-8')
                cell = cut_cells(body, title, 3)[MIXED]
                self.assertIn('vs-muted', cell, 'no note rendered at all')
                self.assertTrue(note_text(cell),
                                'the note rendered with an EMPTY wording -- '
                                'this is what a missing `with context` looks '
                                'like on the page')

    def test_b2_both_templates_import_with_context(self):
        for path in MACRO_TEMPLATES:
            with self.subTest(template=path):
                source = read_template(path)
                # assertTrue, not assertIn: assertIn would dump the whole
                # template into the failure message.
                self.assertTrue(IMPORT_LINE in source,
                                '%s does not import money_cell WITH CONTEXT'
                                % path)
                self.assertEqual(
                    1, source.count("import money_cell"),
                    'exactly one import of the macro per template')

    def test_b2_the_note_wording_lives_only_in_the_partial(self):
        """One owner for the wording, or the two spellings come back."""
        for path in MACRO_TEMPLATES:
            with self.subTest(template=path):
                self.assertNotIn('сумма не указана', read_template(path))
                self.assertNotIn('сумма кўрсатилмаган', read_template(path))

    def test_b2_the_partial_selects_by_lookup_not_by_a_ternary(self):
        """The property B-2 rests on, asserted where it can be seen.

        A ternary on is_ru would make test_b2_without_with_context_the_note
        _is_empty pass by accident in Uzbek and fail to protect Russian.
        """
        source = read_template(PARTIAL)
        macro_body = source.split('{% macro money_cell', 1)[1]
        self.assertIn(".get(lang, {})", macro_body)
        self.assertNotIn('if is_ru', macro_body)


class TestB3TheOldMacroIsGone(unittest.TestCase):
    """B-3: works_reports.html no longer defines its own amount_cell."""

    DECL = re.compile(r'\{%-?\s*macro\s+([A-Za-z_][A-Za-z0-9_]*)')

    def test_b3_no_template_declares_amount_cell(self):
        offenders = []
        for root, _dirs, files in os.walk(os.path.join(REPO_ROOT,
                                                       'templates')):
            for name in files:
                if not name.endswith('.html'):
                    continue
                path = os.path.join(root, name)
                with io.open(path, encoding='utf-8') as fh:
                    for macro in self.DECL.findall(fh.read()):
                        if macro == 'amount_cell':
                            offenders.append(os.path.relpath(path, REPO_ROOT))
        self.assertEqual([], offenders)

    def test_b3_nothing_calls_amount_cell_any_more(self):
        for path in MACRO_TEMPLATES:
            with self.subTest(template=path):
                self.assertNotIn('amount_cell(', read_template(path))

    def test_b3_money_cell_is_declared_exactly_once_in_the_repo(self):
        declarations = []
        for root, _dirs, files in os.walk(os.path.join(REPO_ROOT,
                                                       'templates')):
            for name in files:
                if not name.endswith('.html'):
                    continue
                path = os.path.join(root, name)
                with io.open(path, encoding='utf-8') as fh:
                    if 'money_cell' in self.DECL.findall(fh.read()):
                        declarations.append(os.path.relpath(path, REPO_ROOT))
        self.assertEqual([PARTIAL], declarations)


# ---------------------------------------------------------------------------
# C  commit 3 -- the two screens
# ---------------------------------------------------------------------------

# (page, table title RU, table title UZ, column index of «Сумма»)
# Report cuts:  label jobs area amount received outstanding share
# Debt tables:  label jobs      amount received outstanding
PAGES = (
    ('/drones/works/reports', 'По заказчикам', 'Буюртмачилар бўйича', 3),
    ('/drones/works/debts', 'Долги — по заказчикам',
     'Қарзлар — буюртмачилар бўйича', 2),
)


class ScreenMixin(object):
    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin(cls.__name__.lower()[:24])
        set_language(cls.admin, 'ru')

    def page(self, url, lang='ru'):
        set_language(self.admin, lang)
        client = app.test_client()
        login(client, self.admin)
        response = client.get(url)
        self.assertEqual(200, response.status_code, url)
        return response.data.decode('utf-8')

    def money_columns(self, url, title, first, lang='ru'):
        """{label: (amount cell, received cell, outstanding cell)}."""
        body = self.page(url, lang=lang)
        columns = [cut_cells(body, title, first + offset)
                   for offset in (0, 1, 2)]
        return {label: tuple(column[label] for column in columns)
                for label in columns[0]}


class TestC1TheRuleOnBothScreens(ScreenMixin, unittest.TestCase):
    """C-1: the rule of §2, in all three columns, on both pages.

    ORDERING, stated rather than adapted to: §5's C-1 asks for the all-NULL
    group to be asserted on BOTH pages at this commit. On the debts page it
    is not a row yet -- it has outstanding == 0, so _drone_work_debt_cut()
    folds it into «Долга нет (остальные)», which is the defect commit 4
    removes. Asserting it here would mean writing a test that only passes two
    commits later and leaving this tree red. So the all-NULL group is checked
    on the reports page here and on the debts page in D-1, and
    test_c1_the_debts_page_still_hides_it_at_this_commit records exactly where
    it currently sits -- that assertion INVERTS in commit 4 and is rewritten
    there.
    """

    def test_c1_amount_and_outstanding_are_dashes_received_follows_its_own(
            self):
        cells = self.money_columns('/drones/works/reports', 'По заказчикам', 3)
        amount, received, outstanding = cells[ALL_NULL]
        self.assertEqual('—', text_of(amount))
        self.assertEqual('—', text_of(outstanding),
                         'a debt computed from an unknown amount is unknown, '
                         'not zero')
        # INVERTED by DRONE-RECEIVED-BLANK-IS-ZERO-001. This line read
        # `self.assertEqual('—', text_of(received))` under «ALL_NULL also has
        # every received NULL, so «Получено» is a dash here by its OWN
        # counter». The owner decided on 2026-08-07 that a blank «олинмаган
        # пуллар» means nothing was collected, so the column has no dash at
        # all now. The two neighbours above are untouched and still asserted.
        self.assertEqual('0', text_of(received))

    def test_c1_the_debts_page_no_longer_hides_it(self):
        """INVERTED by commit 4 rather than deleted.

        At commit 3 this assertion read the other way round: the all-NULL
        group was not a row of the debts table at all and «Долга нет» carried
        it. The «сумма не указана: 3» note on that folded row was commit 1's
        counter reaching the screen -- the folded row at least admitted three
        of its five jobs had no price. Commit 4 takes those three jobs out of
        it, so the note is gone and the row is honest without one.
        """
        cells = self.money_columns('/drones/works/debts',
                                   'Долги — по заказчикам', 2)
        rest = [label for label in cells if 'Долга нет' in label]
        self.assertEqual(1, len(rest), cells.keys())
        amount, _received, outstanding = cells[rest[0]]
        # Nothing unpriced is left inside it, so no note and a real zero.
        self.assertEqual('', note_text(amount))
        self.assertEqual('0', text_of(outstanding))

    def test_c1_received_has_no_dash_rule_of_its_own_any_more(self):
        """NO_RECEIVED: amounts known, received all NULL.

        REVERSED by DRONE-RECEIVED-BLANK-IS-ZERO-001. This test was called
        test_c1_received_keys_off_its_own_counter_not_off_no_amount and
        asserted `self.assertEqual('—', text_of(received))` -- «Получено» is
        a dash. It keys off nothing now: a blank means nothing was collected,
        so the row reads 13 000 000 charged, 0 collected, 13 000 000 due.
        «Сумма» and «Не получено» are unchanged and still asserted.
        """
        for url, title, _uz, first in PAGES:
            with self.subTest(page=url):
                amount, received, outstanding = \
                    self.money_columns(url, title, first)[NO_RECEIVED]
                self.assertIn('13\xa0000\xa0000', amount)
                self.assertEqual('0', text_of(received))
                self.assertIn('13\xa0000\xa0000', outstanding)

    def test_c1_outstanding_keys_off_no_amount_not_off_its_own(self):
        """UNSTATED: one job, no amount, no received.

        Proves «Не получено» reads no_amount: its own value is 0.0 and a
        column keyed off «is it zero» would print 0 here. On the reports page
        for the same reason as above -- the debts page folds it until
        commit 4.
        """
        cells = self.money_columns('/drones/works/reports', 'По заказчикам', 3)
        label = [k for k in cells if UNSTATED_LABEL_RU in k][0]
        amount, received, outstanding = cells[label]
        self.assertEqual('—', text_of(amount))
        # was '—'; see DRONE-RECEIVED-BLANK-IS-ZERO-001 above
        self.assertEqual('0', text_of(received))
        self.assertEqual('—', text_of(outstanding))

    def test_c1_a_partial_group_shows_its_sum_and_the_count(self):
        for url, title, _uz, first in PAGES:
            with self.subTest(page=url):
                amount, received, outstanding = \
                    self.money_columns(url, title, first)[MIXED]
                self.assertIn('5\xa0000\xa0000', amount)
                self.assertEqual('сумма не указана', note_text(amount))
                # 0 of 2 missing -> a plain sum, no note
                self.assertIn('3\xa0000\xa0000', received)
                self.assertEqual('', note_text(received))
                # «Не получено» never carries a note of its own
                self.assertIn('2\xa0000\xa0000', outstanding)
                self.assertEqual('', note_text(outstanding))

    def test_c1_a_fully_known_group_is_three_plain_numbers(self):
        for url, title, _uz, first in PAGES:
            with self.subTest(page=url):
                cells = self.money_columns(url, title, first)
                if NONE_NULL not in cells:
                    # settled groups live under «Долга нет» on the debts page
                    continue
                for cell in cells[NONE_NULL]:
                    self.assertNotIn('vs-muted', cell)
                    self.assertNotEqual('—', text_of(cell))

    def test_c1_the_received_note_renders_nowhere_now(self):
        """INVERTED by DRONE-RECEIVED-BLANK-IS-ZERO-001.

        This asserted the opposite: op1 is the group with
        0 < no_received < jobs, and the note «получено не указано» had to
        render on it. «Получено» has no note any more, so the same cut is
        asserted to carry none -- and the amount note on the same page is
        asserted to survive, or this would pass on a page that lost both.
        """
        cells = self.money_columns('/drones/works/reports',
                                   'По операторам', 3)
        self.assertTrue(cells, 'the operator cut was not found')
        self.assertEqual([], [label for label, c in cells.items()
                              if note_text(c[1])])
        amount_notes = [label for label, c in cells.items()
                        if note_text(c[0])]
        self.assertTrue(amount_notes,
                        'no «Сумма» note either -- the check proves nothing')


class TestC2NegativeControl(ScreenMixin, PreCommitMixin, unittest.TestCase):
    """C-2: the pre-commit tree prints 0 in «Не получено» on both pages."""

    def test_c2_the_group_really_does_have_zero_outstanding(self):
        """The data has not changed -- only what the screens say about it."""
        row = by_label(self.pre_commit_report_data()['by_customer'])[ALL_NULL]
        self.assertEqual(0.0, row['amount'])
        self.assertEqual(0.0, row['received'])
        self.assertEqual(0.0, row['outstanding'])
        self.assertEqual(2, row['jobs'])

    def test_c2_pre_commit_markup_prints_that_zero_as_zero(self):
        """The exact expressions the two templates carried at 116e15a,
        rendered through the REAL vs_num filter."""
        row = by_label(self.pre_commit_report_data()['by_customer'])[ALL_NULL]
        for column in ('received', 'outstanding'):
            with self.subTest(column=column):
                pre_fix = "{{ '%%.0f'|format(cell.%s)|vs_num }}" % column
                with app.app_context():
                    rendered = app.jinja_env.from_string(pre_fix).render(
                        cell=row)
                self.assertEqual('0', rendered.strip())

    def test_c2_the_pre_commit_expression_really_is_what_shipped(self):
        """Quoting an expression proves nothing unless it is THE expression."""
        for path, names in ((REPORTS_TEMPLATE,
                             ('r.received', 'r.outstanding')),
                            (DEBTS_TEMPLATE, ('r.amount', 'r.received',
                                              'r.outstanding'))):
            source = subprocess.check_output(
                ['git', 'show', '%s:%s' % (BASE_COMMIT, path)],
                cwd=REPO_ROOT).decode('utf-8')
            for name in names:
                with self.subTest(template=path, cell=name):
                    self.assertIn("{{ '%%.0f'|format(%s)|vs_num }}" % name,
                                  source)

    def test_c2_and_the_current_reports_page_does_not(self):
        cells = self.money_columns('/drones/works/reports', 'По заказчикам', 3)
        self.assertEqual('—', text_of(cells[ALL_NULL][2]))
        self.assertNotIn('0', text_of(cells[ALL_NULL][2]))

    def test_c2_and_the_current_debts_page_does_not(self):
        """The debts page never gives an unpriced group a row of its OWN --
        before commit 4 it was folded into «Долга нет», after it into the
        «Долг неизвестен» line. So the cell to compare against the pre-commit
        0 is that line's, and it is a dash."""
        cells = self.money_columns('/drones/works/debts',
                                   'Долги — по заказчикам', 2)
        self.assertNotIn(ALL_NULL, cells)
        unknown = [c for label, c in cells.items()
                   if 'Долг неизвестен' in label]
        self.assertEqual(1, len(unknown), cells.keys())
        self.assertEqual('—', text_of(unknown[0][2]))
        self.assertNotIn('0', text_of(unknown[0][2]))


class TestC3DivBalance(unittest.TestCase):
    """C-3: report the counts, do not assert they are equal.

    The macro legitimately adds a muted <div> per partial group, and moving
    the amount cell out removed one; asserting equality would only be true by
    coincidence. tools/check_templates.py is what enforces BALANCE (open ==
    close) per file, and it is a blocking CI check.
    """

    OPEN_RE = re.compile(r'<div(?=[\s>/])', re.I)
    CLOSE_RE = re.compile(r'</div\s*>', re.I)
    COMMENT_RE = re.compile(r'\{#.*?#\}', re.S)

    def counts(self, source):
        stripped = self.COMMENT_RE.sub('', source)
        return (len(self.OPEN_RE.findall(stripped)),
                len(self.CLOSE_RE.findall(stripped)))

    def test_c3_both_templates_are_balanced_now(self):
        for path in MACRO_TEMPLATES + (PARTIAL,):
            with self.subTest(template=path):
                opened, closed = self.counts(read_template(path))
                self.assertEqual(opened, closed, path)

    def test_c3_report_the_before_and_after(self):
        """Recorded, not asserted equal. The numbers go in the PR body."""
        reported = {}
        for path in MACRO_TEMPLATES:
            before = subprocess.check_output(
                ['git', 'show', '%s:%s' % (BASE_COMMIT, path)],
                cwd=REPO_ROOT).decode('utf-8')
            reported[path] = (self.counts(before),
                              self.counts(read_template(path)))
        self.assertEqual(
            {REPORTS_TEMPLATE: ((41, 41), (40, 40)),
             DEBTS_TEMPLATE: ((39, 39), (39, 39))},
            reported)


class TestC4BothLanguages(ScreenMixin, unittest.TestCase):
    """C-4: RU and UZ, both pages, no empty label in the three columns."""

    def test_c4_no_money_cell_renders_empty_in_either_language(self):
        for url, title_ru, title_uz, first in PAGES:
            for lang, title in (('ru', title_ru), ('uz', title_uz)):
                with self.subTest(page=url, lang=lang):
                    cells = self.money_columns(url, title, first, lang=lang)
                    self.assertTrue(cells, 'the table was not found at all')
                    for label, three in cells.items():
                        for index, cell in enumerate(three):
                            self.assertTrue(
                                text_of(cell),
                                'empty cell: %s / %s / column %d'
                                % (url, label, first + index))

    def test_c4_every_note_that_renders_has_wording(self):
        for url, title_ru, title_uz, first in PAGES:
            for lang, title in (('ru', title_ru), ('uz', title_uz)):
                with self.subTest(page=url, lang=lang):
                    for label, cells in self.money_columns(
                            url, title, first, lang=lang).items():
                        for cell in cells:
                            if 'vs-muted' not in cell:
                                continue
                            self.assertTrue(note_text(cell),
                                            'empty wording: %s / %s'
                                            % (url, label))

    def test_c4_the_uzbek_pages_really_are_uzbek(self):
        """Guards against the language switch silently not switching."""
        for url, _ru, title_uz, _first in PAGES:
            with self.subTest(page=url):
                self.assertIn(title_uz, self.page(url, lang='uz'))


# ---------------------------------------------------------------------------
# D  commit 4 -- the third bucket on the debts page
# ---------------------------------------------------------------------------

UNKNOWN_LABEL_RU = 'Долг неизвестен — сумма не записана'
UNKNOWN_LABEL_UZ = 'Қарз номаълум — сумма ёзилмаган'
NO_DEBT_LABEL_RU = 'Долга нет (остальные)'

DEBT_CUTS = ('debt_by_customer', 'debt_by_operator')


class TestD1EveryGroupLandsInOneBucket(unittest.TestCase):
    """D-1: three kinds, three buckets, nothing in two and nothing in none."""

    @classmethod
    def setUpClass(cls):
        seed()

    def setUp(self):
        self.data = report_data()
        self.debt = self.data['debt_by_customer']

    def buckets(self):
        return {'owing': list(self.debt['rows']),
                'unknown': ([self.debt['unknown']]
                            if self.debt['unknown'] else []),
                'rest': [self.debt['rest']] if self.debt['rest'] else []}

    def test_d1_the_three_buckets_all_exist_on_this_fixture(self):
        """A fixture missing a bucket cannot prove anything about it."""
        self.assertTrue(self.debt['rows'])
        self.assertIsNotNone(self.debt['unknown'])
        self.assertIsNotNone(self.debt['rest'])

    def test_d1_owing_holds_exactly_the_groups_that_owe(self):
        self.assertEqual(
            {MIXED, NO_RECEIVED, UNRESOLVED_LABEL_RU},
            {r['label'] for r in self.debt['rows']})

    def test_d1_unknown_holds_exactly_the_unpriced_groups(self):
        """ALL_NULL (2 jobs) + UNSTATED (1 job), by construction of the
        fixture -- the two groups where every job's amount is NULL."""
        unknown = self.debt['unknown']
        self.assertEqual(UNKNOWN_LABEL_RU, unknown['label'])
        self.assertEqual(3, unknown['jobs'])
        self.assertEqual(3, unknown['no_amount'])
        self.assertEqual(unknown['jobs'], unknown['no_amount'])
        self.assertAlmostEqual(0.0, unknown['amount'], places=2)

    def test_d1_settled_holds_only_the_genuinely_settled(self):
        rest = self.debt['rest']
        self.assertEqual(NO_DEBT_LABEL_RU, rest['label'])
        self.assertEqual(2, rest['jobs'])
        self.assertEqual(0, rest['no_amount'],
                         'an unpriced group in «Долга нет» is the defect')
        self.assertAlmostEqual(7000000.0, rest['amount'], places=2)

    def test_d1_no_group_is_in_two_buckets_or_in_none(self):
        for name in DEBT_CUTS:
            with self.subTest(cut=name):
                debt = self.data[name]
                source = name.replace('debt_', '')
                cut = self.data[source]
                everything = cut['rows'] + list(cut['services'])
                buckets = [debt['rows'],
                           [debt['unknown']] if debt['unknown'] else [],
                           [debt['rest']] if debt['rest'] else []]
                placed = sum(len(b) for b in buckets)
                collapsed = sum(1 for b in buckets[1:] for _ in b)
                # every ordinary group is its own row in `owing`; the two
                # service lines each stand for >= 1 group, so compare JOBS
                self.assertEqual(
                    sum(r['jobs'] for r in everything),
                    sum(r['jobs'] for b in buckets for r in b))
                self.assertGreaterEqual(placed, 1)
                self.assertLessEqual(collapsed, 2)

    def test_d1_unknown_is_tested_before_settled(self):
        """The ordering §6 calls load-bearing, asserted where it decides.

        Every group in `unknown` has outstanding <= 0.005, so with the two
        tests in the other order every one of them would land in `rest`.
        """
        unknown = self.debt['unknown']
        self.assertLessEqual(unknown['outstanding'], 0.005)
        self.assertNotIn(UNKNOWN_LABEL_RU, [self.debt['rest']['label']])


class TestD2BucketsSumToTheCutTotal(unittest.TestCase):
    """D-2: assert the reconciliation; do not reason about it."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_d2_the_three_buckets_equal_the_total(self):
        data = report_data()
        for name in DEBT_CUTS:
            with self.subTest(cut=name):
                debt = data[name]
                rows = list(debt['rows'])
                for bucket in (debt['unknown'], debt['rest']):
                    if bucket is not None:
                        rows.append(bucket)
                self.assertEqual(debt['total']['jobs'],
                                 sum(r['jobs'] for r in rows))
                for key in ('area', 'amount', 'received'):
                    self.assertAlmostEqual(
                        debt['total'][key], sum(r[key] for r in rows),
                        delta=0.01, msg='%s / %s' % (name, key))

    def test_d2_the_counters_reconcile_too(self):
        data = report_data()
        for name in DEBT_CUTS:
            with self.subTest(cut=name):
                debt = data[name]
                rows = list(debt['rows'])
                for bucket in (debt['unknown'], debt['rest']):
                    if bucket is not None:
                        rows.append(bucket)
                for key in ('no_amount', 'no_received'):
                    self.assertEqual(debt['total'][key],
                                     sum(r[key] for r in rows),
                                     '%s / %s' % (name, key))

    def test_d2_three_buckets_cover_what_two_covered(self):
        """owing + unknown + settled == owing + settled, as it was."""
        after = report_data()
        module = pre_commit_drones()
        if module is None:
            self.skipTest('git cannot produce %s:drones.py' % BASE_COMMIT)
        from flask import g
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/debts'):
            g.lang = 'ru'
            before = module._drone_works_report_data(
                module._drone_work_conditions(
                    module._drone_works_filters_from_args(MultiDict())))
        for name in DEBT_CUTS:
            with self.subTest(cut=name):
                old = before[name]
                old_rows = list(old['rows']) + (
                    [old['rest']] if old['rest'] else [])
                new = after[name]
                new_rows = list(new['rows']) + [
                    b for b in (new['unknown'], new['rest']) if b]
                self.assertEqual(sum(r['jobs'] for r in old_rows),
                                 sum(r['jobs'] for r in new_rows))
                for key in ('amount', 'received', 'outstanding'):
                    self.assertAlmostEqual(sum(r[key] for r in old_rows),
                                           sum(r[key] for r in new_rows),
                                           delta=0.01,
                                           msg='%s / %s' % (name, key))


class TestD3ReconciledIsUnchanged(PreCommitMixin, unittest.TestCase):
    """D-3: reconciled stays True wherever it was True before."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_d3_reconciled_survives_the_split(self):
        before = self.pre_commit_report_data()
        after = report_data()
        for name in DEBT_CUTS + ('by_customer', 'by_operator',
                                 'by_subdivision', 'by_month', 'by_payment'):
            with self.subTest(cut=name):
                if before[name]['reconciled']:
                    self.assertTrue(after[name]['reconciled'], name)

    def test_d3_it_was_true_before_so_the_check_is_not_vacuous(self):
        before = self.pre_commit_report_data()
        self.assertTrue(all(before[name]['reconciled'] for name in DEBT_CUTS))


class TestD4NegativeControl(PreCommitMixin, unittest.TestCase):
    """D-4: pre-commit, an unpriced group is INSIDE «Долга нет»."""

    @classmethod
    def setUpClass(cls):
        seed()

    def test_d4_pre_commit_puts_the_unpriced_jobs_in_the_no_debt_row(self):
        """The assertion that proves it, quoted in the commit message:

            self.assertEqual(5, rest['jobs'])

        Five jobs, not two: ALL_NULL's two and UNSTATED's one are folded in
        beside NONE_NULL's two. After commit 4 the same row holds 2.
        """
        rest = self.pre_commit_report_data()['debt_by_customer']['rest']
        self.assertEqual(NO_DEBT_LABEL_RU, rest['label'])
        self.assertEqual(5, rest['jobs'])
        self.assertAlmostEqual(7000000.0, rest['amount'], places=2)

    def test_d4_pre_commit_has_no_third_bucket_at_all(self):
        debt = self.pre_commit_report_data()['debt_by_customer']
        self.assertNotIn('unknown', debt)

    def test_d4_after_the_commit_the_same_row_holds_two_jobs(self):
        rest = report_data()['debt_by_customer']['rest']
        self.assertEqual(2, rest['jobs'])
        self.assertEqual(0, rest['no_amount'])

    def test_d4_the_three_unpriced_jobs_moved_and_were_not_lost(self):
        before = self.pre_commit_report_data()['debt_by_customer']
        after = report_data()['debt_by_customer']
        self.assertEqual(before['rest']['jobs'],
                         after['rest']['jobs'] + after['unknown']['jobs'])


class TestD5TheLabelIsCyrillicUzbek(unittest.TestCase):
    """D-5: the new label in both languages, Uzbek Cyrillic by code point."""

    # Latin is allowed only in product and brand names -- none appear in this
    # label, so the allowance is empty here on purpose.
    LATIN = re.compile(r'[A-Za-z]')

    def label(self, lang):
        import drones
        from flask import g
        with app.test_request_context('/drones/works/debts'):
            g.lang = lang
            return drones._drone_t('Қарз номаълум — сумма ёзилмаган',
                                   'Долг неизвестен — сумма не записана')

    def test_d5_both_languages_exist_and_differ(self):
        self.assertEqual(UNKNOWN_LABEL_RU, self.label('ru'))
        self.assertEqual(UNKNOWN_LABEL_UZ, self.label('uz'))
        self.assertNotEqual(self.label('ru'), self.label('uz'))

    def test_d5_the_uzbek_label_carries_no_latin_letter(self):
        self.assertIsNone(self.LATIN.search(UNKNOWN_LABEL_UZ),
                          UNKNOWN_LABEL_UZ)

    def test_d5_it_is_cyrillic_by_code_point_not_by_eye(self):
        """«Қ» is U+049A. A Latin K with a tail pasted on is not it."""
        letters = [ch for ch in UNKNOWN_LABEL_UZ if ch.isalpha()]
        self.assertTrue(letters)
        for ch in letters:
            self.assertTrue(0x0400 <= ord(ch) <= 0x04FF,
                            '%r is U+%04X, outside Cyrillic'
                            % (ch, ord(ch)))
        self.assertIn('Қ', UNKNOWN_LABEL_UZ)

    def test_d5_the_scan_fires_on_a_planted_latin_string(self):
        """Negative control: a check that cannot fail is not a check."""
        planted = 'Qarz nomalum - summa yozilmagan'
        self.assertIsNotNone(self.LATIN.search(planted))
        with self.assertRaises(AssertionError):
            for ch in [c for c in planted if c.isalpha()]:
                self.assertTrue(0x0400 <= ord(ch) <= 0x04FF)

    def test_d5_the_label_reaches_both_rendered_pages(self):
        seed()
        admin = create_admin('d5_admin')
        for lang, expected in (('ru', UNKNOWN_LABEL_RU),
                               ('uz', UNKNOWN_LABEL_UZ)):
            with self.subTest(lang=lang):
                set_language(admin, lang)
                client = app.test_client()
                login(client, admin)
                body = client.get('/drones/works/debts').data.decode('utf-8')
                self.assertIn(expected, body)


class TestD6TheBucketOnTheRenderedPage(ScreenMixin, unittest.TestCase):
    """The row itself: three dashes, its jobs, and above «Долга нет»."""

    def test_d6_the_unknown_row_prints_two_dashes_and_a_zero(self):
        """Was three dashes. DRONE-RECEIVED-BLANK-IS-ZERO-001 makes the
        middle a 0, and that reading is coherent and intended: nothing was
        collected, and what was owed is unknown."""
        cells = self.money_columns('/drones/works/debts',
                                   'Долги — по заказчикам', 2)
        row = [c for label, c in cells.items() if UNKNOWN_LABEL_RU in label]
        self.assertEqual(1, len(row), cells.keys())
        amount, received, outstanding = row[0]
        self.assertEqual('—', text_of(amount))
        self.assertEqual('0', text_of(received))
        self.assertEqual('—', text_of(outstanding))

    def table(self):
        """The <tbody> of the customer debt table, alone.

        [REASON]: both labels also appear in the explanatory alert above the
        tables, and there «Долга нет» comes first. Comparing positions over
        the whole page measured the prose, not the table -- and it did fail
        that way before this was narrowed.
        """
        body = self.page('/drones/works/debts')
        after = body.split('Долги — по заказчикам', 1)[1]
        return after.split('<tbody>')[1].split('</tbody>')[0]

    def test_d6_it_sits_above_the_no_debt_row(self):
        tbody = self.table()
        self.assertLess(tbody.index(UNKNOWN_LABEL_RU),
                        tbody.index(NO_DEBT_LABEL_RU))

    def test_d6_its_jobs_are_shown(self):
        row = [r for r in self.table().split('<tr')
               if UNKNOWN_LABEL_RU in r][0]
        self.assertEqual('3', text_of(_CELL_RE.findall(row)[1]))

    def test_d6_the_rendered_row_carries_the_class_that_paints_a_row(self):
        """Not the template source -- the bytes the browser receives.

        `.vs-table tr.is-warning-row` is the only rule in design-system.css
        that paints a table row; `is-warning` alone paints nothing on a <tr>,
        so the row that exists to be noticed rendered like every other row.
        F-1 is the check that catches the class of defect; this asserts the
        one row.
        """
        row = [r for r in self.table().split('<tr')
               if UNKNOWN_LABEL_RU in r][0]
        self.assertIn('class="is-warning-row"', '<tr' + row)
        self.assertNotIn('class="is-warning"', '<tr' + row)

    def test_d6_the_page_explains_the_new_row(self):
        """A visible service row nobody explains is a support ticket."""
        for lang, needle in (('ru', 'Долг неизвестен'),
                             ('uz', 'Қарз номаълум')):
            with self.subTest(lang=lang):
                body = self.page('/drones/works/debts', lang=lang)
                self.assertIn(needle, body.split('<table', 1)[0]
                              + body.rsplit('</table>', 1)[-1])


# ---------------------------------------------------------------------------
# E  commit 5 -- Excel must not disagree with the screen
# ---------------------------------------------------------------------------

XLSX_ROUTES = ('/drones/works/reports.xlsx', '/drones/works/debt.xlsx')

# Both writers DO emit per-group rows. reports.xlsx writes five cut sheets
# through _drone_work_cut_sheet() and both debt sheets through
# _drone_work_debt_sheet(); debt.xlsx writes the two debt sheets. Nothing was
# skipped and no row was invented.
CUT_SHEETS = ('По заказчикам', 'По операторам', 'По подразделениям',
              'По месяцам', 'По типам оплаты')
DEBT_SHEETS = ('Долги — заказчики', 'Долги — операторы')
# 0-based index of (amount, received, outstanding) in each sheet kind.
CUT_MONEY = (3, 4, 5)
DEBT_MONEY = (2, 3, 4)


class WorkbookMixin(object):
    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin(cls.__name__.lower()[:24])
        set_language(cls.admin, 'ru')

    def workbook(self, url):
        from openpyxl import load_workbook
        client = app.test_client()
        login(client, self.admin)
        response = client.get(url)
        self.assertEqual(200, response.status_code, url)
        return load_workbook(io.BytesIO(response.data))

    def sheet_rows(self, ws):
        return {row[0]: row
                for row in ws.iter_rows(min_row=2, values_only=True)}


class TestE1UnknownCellsAreEmpty(WorkbookMixin, unittest.TestCase):
    """E-1: an unknown figure reads back as None -- not 0, not '', not '—'.

    NARROWED to «Сумма» and «Не получено» by
    DRONE-RECEIVED-BLANK-IS-ZERO-001. «Получено» is no longer a column that
    can be unknown: the owner decided on 2026-08-07 that a blank «олинмаган
    пуллар» cell means nothing was collected. So the received column is
    asserted to be a NUMBER here, with the same force the blank used to be
    asserted with -- inverted, not dropped. See
    tests/test_drone_received_blank_is_zero.py for that task's own tests.
    """

    # Column indices of (amount, outstanding) -- the two that can still be
    # unknown. received deliberately absent; it has its own assertions.
    CUT_UNKNOWNABLE = (CUT_MONEY[0], CUT_MONEY[2])
    DEBT_UNKNOWNABLE = (DEBT_MONEY[0], DEBT_MONEY[2])

    def unknown_rows(self, wb):
        """Every row of every sheet whose group has no priced job at all.

        On the cut sheets that is ALL_NULL and «Заказчик не указан»; on the
        debt sheets it is the collapsed «Долг неизвестен» line.
        """
        found = []
        for title in CUT_SHEETS:
            if title not in wb.sheetnames:
                continue
            for label, row in self.sheet_rows(wb[title]).items():
                if label in (ALL_NULL, UNSTATED_LABEL_RU):
                    found.append((title, label, row, self.CUT_UNKNOWNABLE))
        for title in DEBT_SHEETS:
            if title not in wb.sheetnames:
                continue
            for label, row in self.sheet_rows(wb[title]).items():
                if label == UNKNOWN_LABEL_RU:
                    found.append((title, label, row, self.DEBT_UNKNOWNABLE))
        return found

    def test_e1_the_unknown_cells_read_back_as_none(self):
        for url in XLSX_ROUTES:
            wb = self.workbook(url)
            rows = self.unknown_rows(wb)
            self.assertTrue(rows, 'no unknown row found in %s' % url)
            for title, label, row, columns in rows:
                for index in columns:
                    with self.subTest(url=url, sheet=title, label=label,
                                      column=index):
                        value = row[index]
                        self.assertIsNone(value)
                        # spelled out, because each of these is a way the
                        # old behaviour could come back wearing a new coat
                        self.assertNotEqual(0, value)
                        self.assertNotEqual('', value)
                        self.assertNotEqual('—', value)

    def test_e1_the_received_cell_of_an_unknown_row_is_a_number(self):
        """The half DRONE-RECEIVED-BLANK-IS-ZERO-001 inverted.

        Same rows, same workbooks, opposite expectation: «Получено» holds 0,
        not a blank. Asserted here rather than deleted so the reversal is
        visible where the old rule was written down.
        """
        for url in XLSX_ROUTES:
            wb = self.workbook(url)
            for title in CUT_SHEETS:
                if title not in wb.sheetnames:
                    continue
                for label, row in self.sheet_rows(wb[title]).items():
                    if label in (ALL_NULL, UNSTATED_LABEL_RU):
                        with self.subTest(url=url, sheet=title, label=label):
                            self.assertEqual(0, row[CUT_MONEY[1]])
            for title in DEBT_SHEETS:
                if title not in wb.sheetnames:
                    continue
                for label, row in self.sheet_rows(wb[title]).items():
                    if label == UNKNOWN_LABEL_RU:
                        with self.subTest(url=url, sheet=title, label=label):
                            self.assertEqual(0, row[DEBT_MONEY[1]])

    def test_e1_the_scan_saw_every_sheet_it_should_have(self):
        """A scan that found nothing would pass every assertion above.

        The counts are measured, not guessed. ALL_NULL and «Заказчик не
        указан» are CUSTOMER labels, so they appear on the customer cut sheet
        only -- the other four cuts group by operator, subdivision, month and
        payment type and label their rows accordingly. The «Долг неизвестен»
        line appears on both debt sheets in both workbooks, which is the
        thing the per-job operator assignment in the fixture buys.
        """
        reports = self.unknown_rows(self.workbook(XLSX_ROUTES[0]))
        debts = self.unknown_rows(self.workbook(XLSX_ROUTES[1]))
        self.assertEqual(
            [('По заказчикам', ALL_NULL), ('По заказчикам',
                                           UNSTATED_LABEL_RU),
             ('Долги — заказчики', UNKNOWN_LABEL_RU),
             ('Долги — операторы', UNKNOWN_LABEL_RU)],
            [(title, label) for title, label, _row, _cols in reports])
        self.assertEqual(
            [('Долги — заказчики', UNKNOWN_LABEL_RU),
             ('Долги — операторы', UNKNOWN_LABEL_RU)],
            [(title, label) for title, label, _row, _cols in debts])

    def test_e1_a_received_only_gap_blanks_nothing_now(self):
        """NO_RECEIVED: amounts known, every received NULL.

        REVERSED by DRONE-RECEIVED-BLANK-IS-ZERO-001. This test used to read
        `self.assertIsNone(row[CUT_MONEY[1]])` under «only «Получено» is
        blank». It is now the clearest single case of the new rule: nothing
        was collected, so the column says 0, and the debt is the whole amount.
        """
        wb = self.workbook(XLSX_ROUTES[0])
        row = self.sheet_rows(wb['По заказчикам'])[NO_RECEIVED]
        self.assertEqual(13000000, row[CUT_MONEY[0]])
        self.assertEqual(0, row[CUT_MONEY[1]])
        self.assertIsNotNone(row[CUT_MONEY[1]])
        self.assertEqual(13000000, row[CUT_MONEY[2]])

    def test_e1_a_partial_group_keeps_all_three_numbers(self):
        """MIXED: one of two jobs unpriced -- a lower bound, not a blank."""
        wb = self.workbook(XLSX_ROUTES[0])
        row = self.sheet_rows(wb['По заказчикам'])[MIXED]
        self.assertEqual(5000000, row[CUT_MONEY[0]])
        self.assertEqual(3000000, row[CUT_MONEY[1]])
        self.assertEqual(2000000, row[CUT_MONEY[2]])

    def test_e1_the_workbook_agrees_with_the_screen_cell_for_cell(self):
        """The property §7 is actually about, asserted directly.

        The received column dropped out of the pairing when
        DRONE-RECEIVED-BLANK-IS-ZERO-001 made it unconditional: `None` is the
        missing_key it would need, and «no rule» is not a rule to compare
        against. Its own agreement is asserted by
        test_e1_the_received_cell_is_never_blank_and_matches_the_row below.
        """
        data = report_data()
        wb = self.workbook(XLSX_ROUTES[0])
        rows = self.sheet_rows(wb['По заказчикам'])
        for row in data['by_customer']['rows'] + \
                list(data['by_customer']['services']):
            for index, missing_key in ((CUT_MONEY[0], 'no_amount'),
                                       (CUT_MONEY[2], 'no_amount')):
                with self.subTest(label=row['label'], column=index):
                    screen_is_dash = (row['jobs']
                                      and row[missing_key] == row['jobs'])
                    cell = rows[row['label']][index]
                    self.assertEqual(screen_is_dash, cell is None,
                                     'screen and workbook disagree')

    def test_e1_the_received_cell_is_never_blank_and_matches_the_row(self):
        """Received: no dash rule at all, and the figure still agrees."""
        data = report_data()
        rows = self.sheet_rows(self.workbook(XLSX_ROUTES[0])['По заказчикам'])
        for row in data['by_customer']['rows'] + \
                list(data['by_customer']['services']):
            with self.subTest(label=row['label']):
                cell = rows[row['label']][CUT_MONEY[1]]
                self.assertIsNotNone(cell)
                self.assertAlmostEqual(row['received'], cell, places=2)


class TestE2EveryOtherMoneyCellIsStillANumber(WorkbookMixin,
                                              unittest.TestCase):
    """E-2: numbers stay numbers, and no separator character travels."""

    def test_e2_no_money_cell_is_a_string(self):
        for url in XLSX_ROUTES:
            wb = self.workbook(url)
            offenders = []
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if not isinstance(cell.value, str):
                            continue
                        if '\xa0' in cell.value or cell.value.strip() == '—':
                            offenders.append((url, ws.title, cell.coordinate,
                                              cell.value))
            self.assertEqual([], offenders)

    def test_e2_every_known_money_cell_is_numeric(self):
        for url, sheets, columns in (
                (XLSX_ROUTES[0], CUT_SHEETS, CUT_MONEY),
                (XLSX_ROUTES[0], DEBT_SHEETS, DEBT_MONEY),
                (XLSX_ROUTES[1], DEBT_SHEETS, DEBT_MONEY)):
            wb = self.workbook(url)
            checked = 0
            for title in sheets:
                for label, row in self.sheet_rows(wb[title]).items():
                    for index in columns:
                        if row[index] is None:
                            continue
                        checked += 1
                        with self.subTest(url=url, sheet=title, label=label):
                            self.assertIsInstance(row[index], (int, float))
                            self.assertNotIsInstance(row[index], bool)
            self.assertGreater(checked, 10, '%s / %s' % (url, sheets[0]))

    def test_e2_the_totals_still_add_up_inside_the_sheet(self):
        """A blank must not have been achieved by dropping a number."""
        wb = self.workbook(XLSX_ROUTES[0])
        ws = wb['По заказчикам']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        body, total = rows[:-1], rows[-1]
        self.assertEqual('Итого', total[0])
        self.assertEqual(TOTAL_JOBS, total[1])
        self.assertEqual(TOTAL_JOBS, sum(r[1] for r in body))
        for index, expected in zip(CUT_MONEY, (TOTAL_AMOUNT, TOTAL_RECEIVED,
                                               TOTAL_OUTSTANDING)):
            with self.subTest(column=index):
                self.assertAlmostEqual(
                    expected, sum(r[index] or 0.0 for r in body), places=2)
                self.assertAlmostEqual(expected, total[index], places=2)

    def test_e2_the_sheet_layout_is_unchanged(self):
        """Blanking a cell must not move or rename a column."""
        wb = self.workbook(XLSX_ROUTES[0])
        self.assertEqual(
            ['Заказчик', 'Работ', 'Гектары', 'Сумма', 'Получено',
             'Не получено', 'Доля, %'],
            list(next(wb['По заказчикам'].iter_rows(max_row=1,
                                                    values_only=True))))
        self.assertEqual(
            ['Заказчик', 'Работ', 'Сумма', 'Получено', 'Не получено'],
            list(next(wb['Долги — заказчики'].iter_rows(max_row=1,
                                                        values_only=True))))


class TestE3NegativeControl(WorkbookMixin, unittest.TestCase):
    """E-3: pre-commit, the same cells read back as 0."""

    def pre_commit_cut_sheet_values(self):
        """The pre-commit writers, run over the same rows into a real
        workbook, and read back with openpyxl."""
        module = pre_commit_drones()
        if module is None:
            self.skipTest('git cannot produce %s:drones.py' % BASE_COMMIT)
        from flask import g
        from openpyxl import Workbook, load_workbook
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = 'ru'
            data = module._drone_works_report_data(
                module._drone_work_conditions(
                    module._drone_works_filters_from_args(MultiDict())))
            st = module._drone_xlsx_styler()
            wb = Workbook()
            wb.remove(wb.active)
            module._drone_work_cut_sheet(wb, st, 'По заказчикам', 'Заказчик',
                                         data['by_customer'])
            module._drone_work_debt_sheet(wb, st, 'Долги — заказчики',
                                          'Заказчик',
                                          data['debt_by_customer'])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return load_workbook(buffer)

    def test_e3_pre_commit_writes_zero_where_this_commit_writes_nothing(self):
        wb = self.pre_commit_cut_sheet_values()
        rows = self.sheet_rows(wb['По заказчикам'])
        for index in CUT_MONEY:
            with self.subTest(column=index):
                self.assertEqual(0, rows[ALL_NULL][index])
                self.assertIsNotNone(rows[ALL_NULL][index])

    def test_e3_pre_commit_no_received_column_is_zero_too(self):
        wb = self.pre_commit_cut_sheet_values()
        self.assertEqual(0, self.sheet_rows(wb['По заказчикам'])[
            NO_RECEIVED][CUT_MONEY[1]])

    def test_e3_and_this_commit_writes_none_for_the_same_cells(self):
        """NARROWED to the two columns that can still be unknown.

        It used to cover all three and to add
        `self.assertIsNone(rows[NO_RECEIVED][CUT_MONEY[1]])`.
        DRONE-RECEIVED-BLANK-IS-ZERO-001 made «Получено» unconditional, so
        those two lines moved to the assertion below rather than vanishing.
        """
        wb = self.workbook(XLSX_ROUTES[0])
        rows = self.sheet_rows(wb['По заказчикам'])
        for index in (CUT_MONEY[0], CUT_MONEY[2]):
            with self.subTest(column=index):
                self.assertIsNone(rows[ALL_NULL][index])
        self.assertEqual(0, rows[ALL_NULL][CUT_MONEY[1]])
        self.assertEqual(0, rows[NO_RECEIVED][CUT_MONEY[1]])

    def test_e3_pre_commit_debt_sheet_hides_the_unpriced_jobs(self):
        """The workbook's half of D-4: five jobs under «Долга нет»."""
        wb = self.pre_commit_cut_sheet_values()
        rows = self.sheet_rows(wb['Долги — заказчики'])
        self.assertIn(NO_DEBT_LABEL_RU, rows)
        self.assertNotIn(UNKNOWN_LABEL_RU, rows)
        self.assertEqual(5, rows[NO_DEBT_LABEL_RU][1])


# ---------------------------------------------------------------------------
# F  every CSS class this task's templates use is REACHABLE on the element
#    it is used on
# ---------------------------------------------------------------------------
#
# [REASON]: DRONE-UI-FIX-002 carried this check as its X-5 and the prompt for
# DRONE-ZERO-VS-UNKNOWN-001 dropped it, so a defect walked straight through to
# review: the third debt bucket shipped as <tr class="is-warning">, and
# `is-warning` does not style a table row. design-system.css defines
# `.vs-table tr.is-warning-row` (lines 646 and 649) for that, and defines
# `is-warning` only as a modifier of .vs-mini, .vs-stat-value, .vs-alert and
# .vs-report-tile. The row that exists to be noticed rendered like every other
# row, and nothing failed.
#
# «Is the class defined anywhere in the CSS» would NOT have caught it --
# is-warning IS defined, four times over. What catches it is asking whether
# the class is defined for THE ELEMENT IT IS USED ON: a rule reaches
# <tr class="X"> only if some selector's last compound carries .X, names
# either no tag or `tr`, and requires no class the element does not have.

CSS_PATH = os.path.join(REPO_ROOT, 'static', 'css', 'design-system.css')

_CSS_COMMENT_RE = re.compile(r'/\*.*?\*/', re.S)
_CSS_RULE_RE = re.compile(r'([^{}]+)\{', re.S)
_CSS_PSEUDO_RE = re.compile(r'::?[a-zA-Z-]+(\([^)]*\))?')
_CSS_ATTR_RE = re.compile(r'\[[^\]]*\]')
_CSS_COMBINATOR_RE = re.compile(r'\s+|>|\+|~')
_COMPOUND_RE = re.compile(r'^([a-zA-Z][a-zA-Z0-9]*)?((?:\.[A-Za-z0-9_-]+)*)$')

_JINJA_RE = re.compile(r'\{\{.*?\}\}|\{%.*?%\}', re.S)
_JINJA_COMMENT_RE = re.compile(r'\{#.*?#\}', re.S)
_TAG_RE = re.compile(r'<([a-zA-Z][a-zA-Z0-9]*)\b([^>]*)>', re.S)
_CLASS_ATTR_RE = re.compile(r'''\bclass\s*=\s*(?:"([^"]*)"|'([^']*)')''', re.S)
_PLAIN_CLASS_RE = re.compile(r'^[A-Za-z0-9_-]+$')


def css_last_compounds(path=CSS_PATH):
    """[(tag or None, frozenset(classes))] -- one per class-bearing selector.

    Only the LAST compound of each selector matters: it is what decides which
    element the rule paints. Ancestor requirements are deliberately ignored --
    checking them would need a DOM, and the defect class this catches does not
    need one.
    """
    with io.open(path, encoding='utf-8') as handle:
        source = _CSS_COMMENT_RE.sub('', handle.read())
    compounds = []
    for block in _CSS_RULE_RE.findall(source):
        block = block.strip()
        # @media / @supports / @keyframes wrappers are not selectors; the
        # rules nested inside them are matched separately by the same regex.
        if not block or block.startswith('@'):
            continue
        for selector in block.split(','):
            selector = _CSS_PSEUDO_RE.sub(
                '', _CSS_ATTR_RE.sub('', selector)).strip()
            if not selector:
                continue
            parts = [p for p in _CSS_COMBINATOR_RE.split(selector) if p]
            if not parts:
                continue
            match = _COMPOUND_RE.match(parts[-1])
            if match is None:
                continue
            classes = frozenset(c for c in match.group(2).split('.') if c)
            if classes:
                compounds.append((match.group(1), classes))
    return compounds


def template_class_usages(path):
    """[(line, tag, frozenset(classes))] for one template.

    Jinja comments are blanked rather than deleted so reported line numbers
    are the ones in the file. Jinja expressions inside a class attribute are
    blanked too: `class="vs-btn {{ x }}"` contributes vs-btn and nothing else,
    because what x expands to is not knowable here.
    """
    with io.open(os.path.join(REPO_ROOT, path), encoding='utf-8') as handle:
        source = handle.read()
    source = _JINJA_COMMENT_RE.sub(
        lambda m: re.sub(r'[^\n]', ' ', m.group(0)), source)
    usages = []
    for match in _TAG_RE.finditer(source):
        for attr in _CLASS_ATTR_RE.finditer(match.group(2)):
            raw = _JINJA_RE.sub(' ', attr.group(1) or attr.group(2) or '')
            classes = frozenset(c for c in raw.split()
                                if _PLAIN_CLASS_RE.match(c))
            if classes:
                usages.append((source[:match.start()].count('\n') + 1,
                               match.group(1).lower(), classes))
    return usages


def unreachable_classes(path, compounds=None):
    """[(line, tag, css_class, why)] for classes no rule can paint here."""
    compounds = css_last_compounds() if compounds is None else compounds
    defined = set()
    for _tag, classes in compounds:
        defined |= classes
    findings = []
    for line, tag, classes in template_class_usages(path):
        for css_class in sorted(classes):
            if css_class not in defined:
                findings.append((line, tag, css_class,
                                 'not defined in design-system.css at all'))
                continue
            if any(css_class in required
                   and (rule_tag is None or rule_tag == tag)
                   and required <= classes
                   for rule_tag, required in compounds):
                continue
            where = sorted({'%s.%s' % (rule_tag or '',
                                       '.'.join(sorted(required)))
                            for rule_tag, required in compounds
                            if css_class in required})
            findings.append((line, tag, css_class,
                             'defined, but only as %s -- none of those can '
                             'paint <%s> with these classes'
                             % (', '.join(where), tag)))
    return findings


# [REASON]: EMPTY, and the constant stays so that it is empty ON PURPOSE
# rather than by never having existed. It briefly held one entry --
# (works_reports.html, 'tr', 'is-warning'), the same mistake on the service
# rows of the five report cuts, dating from 9f018e1 and present at the merge
# base 116e15a. It was left out of the review fix as an already-accepted
# screen's appearance, and the owner overruled that: those service rows are
# «Заказчик не указан», «Оператор не определён», «Подразделение не указано» --
# the same «we cannot say» rows the third debt bucket exists for. Painting one
# and not the others makes two screens disagree about one idea.
#
# test_f1_the_exemption_list_is_exactly_this asserts the SCAN FINDS NOTHING,
# so adding an entry here is a deliberate act against a test that says the
# list must stay empty -- not a quiet place to put the next one.
KNOWN_UNREACHABLE = set()


DRONES_TEMPLATE_DIR = os.path.join(REPO_ROOT, 'templates', 'drones')

# Every template this task edited. works_assignment_hints.html joined the list
# when its «Машина не распознана» row was fixed.
TOUCHED_TEMPLATES = MACRO_TEMPLATES + (
    PARTIAL, 'templates/drones/works_assignment_hints.html')


def drones_templates():
    """Every template under templates/drones/, sorted, repo-relative.

    [REASON]: scanning only the templates a task touched WAS the defect. It is
    why <tr class="is-warning"> survived on works_assignment_hints.html --
    nothing had changed that file, so nothing looked at it. A check that only
    reads where the last edit happened misses the next one the same way: it
    measures recency, not correctness. The directory is enumerated rather than
    listed, so a new template is covered the day it appears rather than the
    day somebody remembers to add it.

    F-1 does NOT use this yet, and the reason is scope, not doubt. Running it
    over the whole module surfaces three more findings, all in customers.html
    and none of them this increment's:

        customers.html:136  <tr class="is-focus">      is-focus is in no CSS
        customers.html:164  <label class="vs-check">   vs-check is in no CSS
        customers.html:190  <label class="vs-check">   (.vs-check-card and
                                                        .vs-check-grid exist;
                                                        vs-check does not)

    Fixing them would repaint a screen this PR has no business repainting --
    a highlighted row and two checkbox labels that have never been styled.
    They are recorded in the PR body with file, line, tag and class. Widening
    F-1 is one line: TEMPLATES = drones_templates().
    """
    return tuple(sorted(
        'templates/drones/' + name
        for name in os.listdir(DRONES_TEMPLATE_DIR)
        if name.endswith('.html')))


class TestF1EveryClassIsReachable(unittest.TestCase):
    """F-1: the check DRONE-UI-FIX-002 had as X-5 and this task dropped.

    Scope is the templates this task touched. drones_templates() is the
    module-wide scan and its docstring records what it finds and why this PR
    does not act on it.
    """

    TEMPLATES = TOUCHED_TEMPLATES

    def findings(self):
        compounds = css_last_compounds()
        found = []
        for path in self.TEMPLATES:
            for line, tag, css_class, why in unreachable_classes(path,
                                                                 compounds):
                found.append((path, line, tag, css_class, why))
        return found

    def test_f1_no_template_uses_a_class_no_rule_can_paint(self):
        offenders = [(path, line, tag, css_class, why)
                     for path, line, tag, css_class, why in self.findings()
                     if (path, tag, css_class) not in KNOWN_UNREACHABLE]
        self.assertEqual([], offenders, '\n'.join(
            '%s:%d <%s class="%s"> -- %s' % (p, l, t, c, w)
            for p, l, t, c, w in offenders))

    def test_f1_the_exemption_list_is_exactly_this(self):
        """The list must stay EMPTY: the scan finds nothing to exempt.

        An allowlist nobody checks becomes a place to hide things. This
        asserts both directions at once -- no template has an unreachable
        class, AND the constant carries no entry papering over one. Adding an
        entry means editing this assertion too, which is the point.
        """
        seen = {(path, tag, css_class)
                for path, _line, tag, css_class, _why in self.findings()}
        self.assertEqual(set(), seen)
        self.assertEqual(set(), KNOWN_UNREACHABLE)

    def test_f1_the_unknown_bucket_row_is_the_class_that_paints_a_row(self):
        """The defect this check exists for, asserted on its own.

        Jinja comments are stripped first: the comment above the row QUOTES
        the broken markup to explain why it is broken, and a naive substring
        assertion matches the explanation instead of the row. It did, on the
        first run.
        """
        source = _JINJA_COMMENT_RE.sub('', read_template(DEBTS_TEMPLATE))
        self.assertTrue('<tr class="is-warning-row">' in source,
                        'the unknown-bucket row does not carry is-warning-row')
        self.assertFalse('<tr class="is-warning">' in source,
                         'a <tr> still carries the class that paints nothing')

    def test_f1_that_class_is_defined_for_a_table_row(self):
        """Read out of the real CSS, not asserted from the task text."""
        compounds = css_last_compounds()
        self.assertIn(('tr', frozenset({'is-warning-row'})), compounds)
        self.assertNotIn(('tr', frozenset({'is-warning'})), compounds)

    def test_f1_the_assignment_hints_row_is_covered_now(self):
        """It was fixed, so the file is touched, so F-1 scans it.

        The point of the entry: before this commit that template was outside
        every scope, which is exactly how its row survived.
        """
        path = 'templates/drones/works_assignment_hints.html'
        self.assertIn(path, self.TEMPLATES)
        source = _JINJA_COMMENT_RE.sub('', read_template(path))
        self.assertTrue('<tr class="is-warning-row">' in source,
                        'the «Машина не распознана» row lost the fix')
        self.assertFalse('<tr class="is-warning">' in source,
                         'a <tr> still carries the class that paints nothing')

    def test_f1_the_module_wide_gap_is_visible_in_code(self):
        """What F-1 does NOT scan, asserted rather than left to memory.

        drones_templates() is live and one line from being F-1's scope. This
        pins the gap so it cannot quietly close by accident or be forgotten:
        the module has templates this check does not read, and the three
        findings in them are named in drones_templates()' docstring and in
        the PR body.
        """
        unscanned = set(drones_templates()) - set(self.TEMPLATES)
        self.assertTrue(unscanned, 'nothing is unscanned -- widen F-1')
        self.assertIn('templates/drones/customers.html', unscanned)
        self.assertEqual(set(self.TEMPLATES) - set(drones_templates()), set(),
                         'F-1 scans a template that is not in the module')

    def test_f1_the_scan_actually_read_something(self):
        """A parser that found no classes would pass every assertion above."""
        compounds = css_last_compounds()
        self.assertGreater(len(compounds), 200)
        total = sum(len(template_class_usages(path))
                    for path in self.TEMPLATES)
        self.assertGreater(total, 40, 'the template scan read almost nothing')


class TestF2NegativeControl(unittest.TestCase):
    """F-2: plant a class that does not exist, prove the check catches it.

    Two plants, because they fail for two different reasons and only the
    second one is the defect that got through review:

      * a class defined NOWHERE                  -- the obvious case
      * a class defined, but not for this element -- what <tr class=
        "is-warning"> was, and what a plain «is it in the file» check misses

    The plants go through the same unreachable_classes() the live check uses,
    against a temporary file, so the control exercises the real detector.
    """

    def scan_source(self, source):
        handle, path = tempfile.mkstemp(prefix='class_probe_', suffix='.html',
                                        dir=REPO_ROOT)
        with os.fdopen(handle, 'w', encoding='utf-8') as fh:
            fh.write(source)
        try:
            return unreachable_classes(os.path.basename(path))
        finally:
            os.unlink(path)

    def test_f2_a_class_that_does_not_exist_is_caught(self):
        findings = self.scan_source(
            '<div class="vs-card vs-not-a-real-class-at-all"></div>')
        self.assertEqual(1, len(findings), findings)
        line, tag, css_class, why = findings[0]
        self.assertEqual(1, line)
        self.assertEqual('div', tag)
        self.assertEqual('vs-not-a-real-class-at-all', css_class)
        self.assertIn('not defined', why)

    def test_f2_a_class_defined_for_another_element_is_caught(self):
        """The exact shape of the defect, planted and caught."""
        findings = self.scan_source('<table class="vs-table">\n'
                                    '<tr class="is-warning"></tr>\n'
                                    '</table>')
        self.assertEqual(1, len(findings), findings)
        line, tag, css_class, why = findings[0]
        self.assertEqual(2, line)
        self.assertEqual('tr', tag)
        self.assertEqual('is-warning', css_class)
        self.assertIn('vs-alert', why)

    def test_f2_the_fixed_row_passes_the_same_scan(self):
        """The control must ACCEPT the fix, or it is measuring nothing."""
        self.assertEqual([], self.scan_source(
            '<table class="vs-table">\n'
            '<tr class="is-warning-row"><td>x</td></tr>\n'
            '</table>'))

    def test_f2_a_plain_is_it_in_the_file_check_would_have_missed_it(self):
        """Why the check is element-aware, demonstrated rather than claimed."""
        with io.open(CSS_PATH, encoding='utf-8') as handle:
            css = handle.read()
        self.assertIn('is-warning', css)
        findings = self.scan_source('<tr class="is-warning"></tr>')
        self.assertEqual(1, len(findings),
                         'the element-aware check catches what the substring '
                         'check cannot')

    def test_f2_jinja_expressions_do_not_become_class_names(self):
        """Guard against the parser inventing findings out of {{ ... }}."""
        self.assertEqual([], self.scan_source(
            '<div class="vs-card {{ extra_class }}">'
            '<span class="{{ \'vs-badge\' if x else \'\' }}"></span></div>'))


if __name__ == '__main__':
    unittest.main(verbosity=2)
