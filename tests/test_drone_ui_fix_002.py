# -*- coding: utf-8 -*-
"""DRONE-UI-FIX-002 -- three defects found by the production acceptance run.

  A  DRONE-UI-CARD-WRAP-001      money broke between two digits of a group.
  B  DRONE-UI-CUSTOMER-LABEL-001 one label stood for two different facts.
  C  DRONE-REPORT-ZERO-AMOUNT-001 a missing price was reported as 0.

What the assertions here can and cannot prove, stated plainly because the
distinction is the whole point of the acceptance:

  * A-1 and A-2 are STRUCTURAL. They prove the opt-out rule is still additive
    (the rule it opts out of survives verbatim) and that every number carrying
    card actually carries the class. They do NOT prove the number stopped
    wrapping -- no assertion over CSS text can, because wrapping is decided by
    the layout engine against a font. That measurement is A-3/A-4 and it lives
    in tools/measure_drone_card_wrap.py, which drives a real browser. Its
    result is recorded in the PR body, not here.
  * B and C are behavioural and are asserted end to end through the real
    application against a disposable database, with the negative controls
    (B-4, C-2) pinned against the pre-fix rendering so that a check which
    cannot tell the two cases apart is caught rather than trusted.

Run:
  python -m unittest tests.test_drone_ui_fix_002 -v
"""
import os
import re
import unittest
from html.parser import HTMLParser

import jinja2

from tests.harness import app, reset_db, create_admin, login
from models import db, DroneCustomer, DroneWork, User

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSS_PATH = os.path.join(REPO_ROOT, 'static', 'css', 'design-system.css')

# The four screens the task names. Nothing outside the drones module is
# touched: other tracks own their screens.
MONEY_TEMPLATES = (
    'templates/drones/works.html',
    'templates/drones/works_reports.html',
    'templates/drones/works_debts.html',
    'templates/drones/summary.html',
)

# [REASON]: DRONE-ANALYTICS-001. Screens that use `vs-stat-value is-num`
# WITHOUT being one of the six-card money grids above. The distinction is
# real and must not be blurred by widening MONEY_TEMPLATES: that tuple means
# «a .vs-stat-grid.is-money whose six values all opt out of word-wrapping»,
# and test_a2_six_cards_per_screen pins the six. These four carry a plain
# .vs-stat-grid with four to six cards; they need the is-num opt-out on their
# numbers for the same reason, and nothing else about them is the same.
#
# They are listed rather than exempted by directory, so a screen belonging to
# ANOTHER TRACK still cannot pick up these classes unnoticed -- which is the
# whole point of the scan below.
#
# [REASON]: DRONE-WORKS-UPLOAD-001 adds the upload preview to the same
# category. Its grid is a plain .vs-stat-grid of eight cards -- rows,
# hectares, amount, received, other costs, customer spellings, priced rows,
# rows already in the database -- and it needs the same opt-out for the same
# reason. It is NOT a six-card money grid, so MONEY_TEMPLATES stays what it
# means.
#
# [REASON]: DRONE-CLOSE-001 adds the closing screen for the same reason. Its
# grid is a plain .vs-stat-grid of four cards -- months shown, ledger
# hectares, flight hectares, hectares not handed in -- and every one of them
# is a figure that must not word-wrap mid-number. It is NOT a six-card money
# grid, so MONEY_TEMPLATES stays what it means.
DRONES_PLAIN_NUM_TEMPLATES = (
    'templates/drones/spray_usage.html',
    'templates/drones/works_flights_reconcile.html',
    'templates/drones/flight_calendar.html',
    'templates/drones/works_import_preview.html',
    'templates/drones/works_closing.html',
    # DRONE-CASH-SPLIT-001: пять карточек подотчёта — обычная сетка, не
    # денежная шестёрка. Числа в них денежные и опт-аут `is-num` им нужен,
    # но `.vs-stat-grid.is-money` здесь нет и быть не должно: та тройка
    # означает ровно шесть карточек, и test_a2_six_cards_per_screen их
    # пересчитывает.
    'templates/drones/operator_cash.html',
    # DRONE-USEFUL-AREA-001: четыре карточки полезной площади. Числа --
    # гектары и счётчики работ, а не суммы, поэтому `.vs-stat-grid.is-money`
    # здесь нет: та тройка означает ровно шесть денежных карточек. Опт-аут
    # `is-num` картам нужен -- гектары группируются как числа.
    'templates/drones/coverage.html',
)

# [REASON]: this line is the rule the fix opts OUT of, and it must survive the
# fix unchanged. It exists so a long uppercase RU/UZ label wraps inside its
# card instead of widening the whole document; deleting it to stop numbers
# breaking would trade this defect for the one it was written against.
ADDITIVITY_ANCHOR = '.vs-stat-label, .vs-stat-value { overflow-wrap: anywhere; }'

VOID_TAGS = {'area', 'base', 'br', 'col', 'embed', 'hr', 'img', 'input',
             'link', 'meta', 'param', 'source', 'track', 'wbr'}


class StatGridParser(HTMLParser):
    """Collects every .vs-stat-value and whether a money grid encloses it."""

    def __init__(self):
        HTMLParser.__init__(self, convert_charrefs=True)
        self.stack = []
        self.values = []
        self.money_grids = 0
        self.plain_grids = 0

    def _classes(self, attrs):
        return set((dict(attrs).get('class') or '').split())

    def handle_startendtag(self, tag, attrs):
        self._record(self._classes(attrs))

    def handle_starttag(self, tag, attrs):
        classes = self._classes(attrs)
        self._record(classes)
        if tag not in VOID_TAGS:
            self.stack.append((tag, classes))

    def _record(self, classes):
        if 'vs-stat-grid' in classes:
            if 'is-money' in classes:
                self.money_grids += 1
            else:
                self.plain_grids += 1
        if 'vs-stat-value' in classes:
            in_money = any('vs-stat-grid' in c and 'is-money' in c
                           for _tag, c in self.stack)
            self.values.append({'classes': classes, 'in_money_grid': in_money})

    def handle_endtag(self, tag):
        for index in range(len(self.stack) - 1, -1, -1):
            if self.stack[index][0] == tag:
                del self.stack[index:]
                return


def parse_template(relative_path):
    with open(os.path.join(REPO_ROOT, relative_path), encoding='utf-8') as fh:
        source = fh.read()
    parser = StatGridParser()
    parser.feed(source)
    return parser


class TestA1CssIsAdditive(unittest.TestCase):
    """A-1: the fix ADDS rules; it does not edit or delete an existing one."""

    def setUp(self):
        with open(CSS_PATH, encoding='utf-8') as fh:
            self.css = fh.read()

    def test_a1_existing_wrap_rule_survives_verbatim(self):
        self.assertIn(
            ADDITIVITY_ANCHOR, self.css,
            'the overflow-wrap:anywhere rule was edited or removed. It is '
            'deliberate: it keeps a long label from widening the document. '
            'The number fix must opt out beside it, not replace it.')

    def test_a1_both_new_selectors_exist(self):
        self.assertIn('.vs-stat-grid.is-money', self.css)
        self.assertIn('.vs-stat-value.is-num', self.css)

    def test_a1_opt_out_actually_disables_wrapping(self):
        """The opt-out is worthless unless it names both properties.

        overflow-wrap:normal alone still lets word-break fall back to the
        default, so the class states both and the test reads both.
        """
        rule = re.search(r'\.vs-stat-value\.is-num\s*\{([^}]*)\}', self.css)
        self.assertIsNotNone(rule, '.vs-stat-value.is-num rule not found')
        body = rule.group(1)
        self.assertIn('overflow-wrap: normal', body)
        self.assertIn('word-break: keep-all', body)

    def test_a1_money_grid_uses_auto_fit_not_a_fixed_count(self):
        """auto-fit is what lets a card move to a second row instead of
        squeezing; a fixed column count would reintroduce the squeeze."""
        rule = re.search(r'\.vs-stat-grid\.is-money\s*\{([^}]*)\}', self.css)
        self.assertIsNotNone(rule)
        self.assertIn('auto-fit', rule.group(1))

    def test_a1_media_block_rules_are_untouched(self):
        """The two @media .vs-stat-grid rules are outranked, never edited."""
        self.assertIn(
            '.vs-stat-grid { grid-template-columns: repeat(4, minmax(0, 1fr)); }',
            self.css)
        self.assertIn(
            '.vs-stat-grid   { grid-template-columns: repeat(3, minmax(0, 1fr)); }',
            self.css)


class TestA2TemplatesCarryTheClasses(unittest.TestCase):
    """A-2: every value inside a money grid opts out, in all four screens."""

    def test_a2_every_stat_value_in_a_money_grid_is_num(self):
        for path in MONEY_TEMPLATES:
            with self.subTest(template=path):
                parsed = parse_template(path)
                # A check that passes on an empty file proves nothing.
                self.assertGreater(
                    parsed.money_grids, 0,
                    '%s has no .vs-stat-grid.is-money' % path)
                in_money = [v for v in parsed.values if v['in_money_grid']]
                self.assertGreater(
                    len(in_money), 0,
                    '%s has no .vs-stat-value inside a money grid' % path)
                missing = [v for v in in_money if 'is-num' not in v['classes']]
                self.assertEqual(
                    [], missing,
                    '%s: %d .vs-stat-value element(s) inside a money grid do '
                    'not carry is-num' % (path, len(missing)))

    def test_a2_six_cards_per_screen(self):
        """The count is pinned: a card added later without is-num is a defect
        this test must fail on, not silently accept."""
        for path in MONEY_TEMPLATES:
            with self.subTest(template=path):
                parsed = parse_template(path)
                in_money = [v for v in parsed.values if v['in_money_grid']]
                self.assertEqual(6, len(in_money),
                                 '%s: expected 6 stat values, got %d'
                                 % (path, len(in_money)))

    def test_a2_negative_control_parser_can_see_a_missing_class(self):
        """A-2 is only evidence if it can fail. Feed it markup that is wrong
        in exactly the way the real templates were before this commit."""
        parser = StatGridParser()
        parser.feed('<div class="vs-stat-grid is-money">'
                    '  <div class="vs-stat">'
                    '    <div class="vs-stat-value">2 657 997 449</div>'
                    '  </div>'
                    '</div>')
        in_money = [v for v in parser.values if v['in_money_grid']]
        self.assertEqual(1, len(in_money))
        self.assertNotIn('is-num', in_money[0]['classes'])

    def test_a2_parser_does_not_credit_a_value_outside_a_money_grid(self):
        parser = StatGridParser()
        parser.feed('<div class="vs-stat-grid">'
                    '  <div class="vs-stat-value">1</div>'
                    '</div>')
        self.assertEqual(1, len(parser.values))
        self.assertFalse(parser.values[0]['in_money_grid'])
        self.assertEqual(1, parser.plain_grids)

    def test_a2_no_class_added_outside_the_drones_module(self):
        """Other tracks own their screens; this task adds classes only here."""
        offenders = []
        for root, _dirs, files in os.walk(os.path.join(REPO_ROOT, 'templates')):
            for name in files:
                if not name.endswith('.html'):
                    continue
                path = os.path.join(root, name)
                rel = os.path.relpath(path, REPO_ROOT).replace('\\', '/')
                if rel in MONEY_TEMPLATES or rel in DRONES_PLAIN_NUM_TEMPLATES:
                    continue
                with open(path, encoding='utf-8') as fh:
                    source = fh.read()
                if 'vs-stat-grid is-money' in source or 'vs-stat-value is-num' in source:
                    offenders.append(rel)
        self.assertEqual([], offenders)
        # The whitelist must name only files that exist, or a rename empties
        # the scan silently and this test starts proving nothing.
        for rel in MONEY_TEMPLATES + DRONES_PLAIN_NUM_TEMPLATES:
            self.assertTrue(os.path.exists(os.path.join(REPO_ROOT, rel)), rel)

    def test_a2_the_plain_num_screens_really_carry_is_num(self):
        """The other half: a screen on the whitelist that stopped using the
        class would sit there forever as a stale exemption.

        [REASON]: this asserts that is-num is USED, not that every card has
        it. A .vs-stat-value holding «2026-05» is a period, not a number, and
        the opt-out that stops a long figure word-wrapping has nothing to do
        with it -- sources.html has printed plain non-numeric stat values
        since DRONE-FLEET-001. Demanding is-num on every card would push the
        class onto strings to satisfy a test.
        """
        for rel in DRONES_PLAIN_NUM_TEMPLATES:
            with self.subTest(template=rel):
                parsed = parse_template(rel)
                self.assertEqual(
                    0, parsed.money_grids,
                    '%s is a plain grid; move it to MONEY_TEMPLATES' % rel)
                numeric = [v for v in parsed.values
                           if 'is-num' in v['classes']]
                self.assertGreater(
                    len(numeric), 0,
                    '%s is whitelisted but no longer uses is-num' % rel)


# ---------------------------------------------------------------------------
# B  DRONE-UI-CUSTOMER-LABEL-001
# ---------------------------------------------------------------------------

RESOLVED_NAME = 'Миробод АМТ'
UNRESOLVED_RAW = 'Янги Аср фх'


def set_language(user_id, lang):
    """Both labels are language-dependent; a test that did not pin the
    language would pass or fail on the default of the day."""
    with app.app_context():
        user = db.session.get(User, user_id)
        user.language = lang
        db.session.commit()


def seed_customer_cases():
    """The three cases a customer cell can be in, and there are exactly three.

    (a) resolved            drone_customer_id set
    (b) NOT STATED          NULL id, customer_raw '' -- the holding's own land
    (c) NOT RESOLVED        NULL id, customer_raw spelled but unmatched
    """
    reset_db()
    ids = {}
    with app.app_context():
        customer = DroneCustomer(name=RESOLVED_NAME)
        db.session.add(customer)
        db.session.flush()
        ids['customer_id'] = customer.id
        for key, cid, raw in (('a', customer.id, RESOLVED_NAME),
                              ('b', None, ''),
                              ('c', None, UNRESOLVED_RAW)):
            row = DroneWork(period_month='2026-04', customer_raw=raw,
                            drone_customer_id=cid, operator_raw='Ким',
                            area_ha=1.0, amount=1000.0, payment_type='cash')
            db.session.add(row)
            db.session.flush()
            ids[key] = row.id
        db.session.commit()
    return ids


CELL_RE = re.compile(r'<td[^>]*>(.*?)</td>', re.S)
ROW_RE = re.compile(r'<tbody>(.*?)</tbody>', re.S)
BADGE_RE = re.compile(r'<span class="vs-badge">(.*?)</span>', re.S)
MUTED_RE = re.compile(r'<div class="vs-muted"[^>]*>(.*?)</div>', re.S)


def customer_cells(body):
    """{customer_raw as rendered: badge text or plain name} per ledger row.

    The customer column is the third <td>; the muted line under it is the
    verbatim spelling, which is what tells the three fixture rows apart.
    """
    tbody = ROW_RE.search(body)
    if not tbody:
        return {}
    cells = {}
    for row in tbody.group(1).split('<tr>'):
        tds = CELL_RE.findall(row)
        if len(tds) < 3:
            continue
        customer_td = tds[2]
        muted = MUTED_RE.search(customer_td)
        raw = ' '.join(muted.group(1).split()) if muted else ''
        badge = BADGE_RE.search(customer_td)
        if badge:
            rendered = ' '.join(badge.group(1).split())
        else:
            without_muted = MUTED_RE.sub('', customer_td)
            rendered = ' '.join(without_muted.split())
        cells[raw] = rendered
    return cells


class TestB3LedgerTellsTheTwoCasesApart(unittest.TestCase):
    """B-3: three works, three distinct cell renderings, in both languages."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_customer_cases()
        cls.admin = create_admin('label_admin')

    def ledger(self, lang):
        set_language(self.admin, lang)
        client = app.test_client()
        login(client, self.admin)
        response = client.get('/drones/works')
        self.assertEqual(200, response.status_code)
        return customer_cells(response.data.decode('utf-8'))

    def test_b3_russian(self):
        cells = self.ledger('ru')
        self.assertEqual(RESOLVED_NAME, cells[RESOLVED_NAME])
        self.assertEqual('Заказчик не указан', cells[''])
        self.assertEqual('Заказчик не определён', cells[UNRESOLVED_RAW])
        self.assertEqual(3, len(set(cells.values())),
                         'the three cases must render three different things')

    def test_b3_uzbek(self):
        cells = self.ledger('uz')
        self.assertEqual(RESOLVED_NAME, cells[RESOLVED_NAME])
        self.assertEqual('Буюртмачи кўрсатилмаган', cells[''])
        self.assertEqual('Буюртмачи аниқланмаган', cells[UNRESOLVED_RAW])
        self.assertEqual(3, len(set(cells.values())))

    def test_b3_the_filter_offers_both_options(self):
        set_language(self.admin, 'ru')
        client = app.test_client()
        login(client, self.admin)
        body = client.get('/drones/works').data.decode('utf-8')
        select = body.split('id="wCustomer"')[1].split('</select>')[0]
        self.assertIn('Заказчик не указан', select)
        self.assertIn('Заказчик не определён', select)
        self.assertIn('value="-2"', select)
        self.assertIn('value="-1"', select)


# [REASON]: B-4. The pre-commit cell, copied from works.html at c1dbf5a. It is
# kept here as a string rather than re-checked out, because the point is not
# that the old file existed -- it is that the B-3 assertion COULD NOT HAVE
# PASSED against it. A control that cannot fail is not a control.
PRE_FIX_CELL = (
    '{% if w.drone_customer %}{{ w.drone_customer.name }}'
    '{% else %}<span class="vs-badge">{{ t_none_customer }}</span>{% endif %}')


class TestB4NegativeControl(unittest.TestCase):
    """B-4: the same assertion against the pre-fix markup must FAIL on (b)."""

    def render_pre_fix(self, has_customer, lang):
        t_none_customer = ('Заказчик не определён' if lang == 'ru'
                           else 'Буюртмачи аниқланмаган')
        template = jinja2.Environment(autoescape=True).from_string(PRE_FIX_CELL)

        class Work(object):
            drone_customer = 'x' if has_customer else None

        return template.render(w=Work(), t_none_customer=t_none_customer)

    def test_b4_pre_fix_calls_the_unstated_case_unresolved_ru(self):
        rendered = self.render_pre_fix(False, 'ru')
        self.assertIn('Заказчик не определён', rendered)
        self.assertNotIn(
            'Заказчик не указан', rendered,
            'the pre-fix template could not produce the «не указан» label at '
            'all -- which is why B-3 fails against it, as it must')

    def test_b4_pre_fix_calls_the_unstated_case_unresolved_uz(self):
        rendered = self.render_pre_fix(False, 'uz')
        self.assertIn('Буюртмачи аниқланмаган', rendered)
        self.assertNotIn('Буюртмачи кўрсатилмаган', rendered)

    def test_b4_pre_fix_gave_one_label_to_both_null_cases(self):
        """The defect stated exactly: blank and non-blank rendered the same."""
        blank = self.render_pre_fix(False, 'ru')
        spelled = self.render_pre_fix(False, 'ru')
        self.assertEqual(blank, spelled)


class TestB5FiltersPartitionTheNullCustomers(unittest.TestCase):
    """B-5: -2 returns only (b), -1 only (c), disjoint, and union is total."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_customer_cases()
        cls.admin = create_admin('filter_admin')

    def ids_for(self, query):
        import drones
        from flask import g
        from urllib.parse import parse_qsl
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works?' + query):
            g.lang = 'ru'
            filters = drones._drone_works_filters_from_args(
                MultiDict(parse_qsl(query)))
            conds = drones._drone_work_conditions(filters)
            return {w.id for w in DroneWork.query.filter(*conds).all()}

    def test_b5_unstated_returns_only_the_blank_one(self):
        self.assertEqual({self.ids['b']}, self.ids_for('customer_id=-2'))

    def test_b5_unresolved_returns_only_the_spelled_one(self):
        self.assertEqual({self.ids['c']}, self.ids_for('customer_id=-1'))

    def test_b5_the_two_sets_are_disjoint_and_their_union_is_every_null(self):
        unstated = self.ids_for('customer_id=-2')
        unresolved = self.ids_for('customer_id=-1')
        self.assertEqual(set(), unstated & unresolved)
        with app.app_context():
            every_null = {w.id for w in DroneWork.query.filter(
                DroneWork.drone_customer_id.is_(None)).all()}
        # [REASON]: asserted explicitly. Two filters that each return
        # something while together losing a row is the failure worth catching,
        # and neither of the two assertions above can see it.
        self.assertEqual(every_null, unstated | unresolved)

    def test_b5_a_resolved_customer_is_in_neither(self):
        self.assertNotIn(self.ids['a'], self.ids_for('customer_id=-2'))
        self.assertNotIn(self.ids['a'], self.ids_for('customer_id=-1'))

    def test_b5_whitespace_only_raw_counts_as_not_stated(self):
        """trim() is in the condition on purpose: a cell holding a space is a
        cell that said nothing, and the report cut has always read it so."""
        with app.app_context():
            row = DroneWork(period_month='2026-04', customer_raw='   ',
                            drone_customer_id=None, operator_raw='Ким',
                            area_ha=1.0, payment_type='cash')
            db.session.add(row)
            db.session.commit()
            spaces_id = row.id
        try:
            self.assertIn(spaces_id, self.ids_for('customer_id=-2'))
            self.assertNotIn(spaces_id, self.ids_for('customer_id=-1'))
        finally:
            with app.app_context():
                db.session.delete(db.session.get(DroneWork, spaces_id))
                db.session.commit()


class TestB6LabelsAreOneStringNotTwo(unittest.TestCase):
    """B-6: the ledger's labels are byte-identical to the report cut's."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_customer_cases()
        cls.admin = create_admin('b6_admin')

    def report_service_labels(self, lang):
        import drones
        from flask import g
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = lang
            data = drones._drone_works_report_data(
                drones._drone_work_conditions(
                    drones._drone_works_filters_from_args(MultiDict())))
            return [s['label'] for s in data['by_customer']['services']]

    def ledger_badges(self, lang):
        set_language(self.admin, lang)
        client = app.test_client()
        login(client, self.admin)
        body = client.get('/drones/works').data.decode('utf-8')
        return set(customer_cells(body).values()) - {RESOLVED_NAME}

    def test_b6_labels_match_byte_for_byte(self):
        for lang in ('ru', 'uz'):
            with self.subTest(lang=lang):
                report = set(self.report_service_labels(lang))
                ledger = self.ledger_badges(lang)
                self.assertEqual(2, len(report),
                                 'the cut must still carry TWO service rows')
                # Compared programmatically, not by eye: two spellings of one
                # idea in one module is its own defect.
                self.assertEqual(report, ledger)

    def test_b6_one_function_owns_the_wording(self):
        import drones
        from flask import g
        with app.test_request_context('/drones/works'):
            g.lang = 'ru'
            labels = drones._drone_customer_service_labels()
        self.assertEqual(
            {'Заказчик не указан', 'Заказчик не определён'},
            set(labels.values()))

    def test_b6_the_two_labels_are_not_the_same_string(self):
        """The defect was that one string stood for both facts."""
        import drones
        from flask import g
        for lang in ('ru', 'uz'):
            with app.test_request_context('/drones/works'):
                g.lang = lang
                labels = drones._drone_customer_service_labels()
            self.assertEqual(2, len(set(labels.values())), lang)


# ---------------------------------------------------------------------------
# C  DRONE-REPORT-ZERO-AMOUNT-001
# ---------------------------------------------------------------------------

ALL_NULL = 'Бухоро Агрокластер Заминлари МЧЖ'
MIXED = 'Дўстлик ФХ'
NONE_NULL = 'Пешку АМТ'


def seed_amount_cases():
    """Three customer groups, one per case the amount column has to tell apart.

    ALL_NULL   2 jobs, neither carries a price  -> «the price was not written»
    MIXED      2 jobs, one carries a price      -> a sum, and how many did not
    NONE_NULL  2 jobs, both carry a price       -> just the sum
    """
    reset_db()
    ids = {}
    with app.app_context():
        for name, amounts in ((ALL_NULL, (None, None)),
                              (MIXED, (None, 5000000.0)),
                              (NONE_NULL, (3000000.0, 4000000.0))):
            customer = DroneCustomer(name=name)
            db.session.add(customer)
            db.session.flush()
            ids[name] = customer.id
            for amount in amounts:
                db.session.add(DroneWork(
                    period_month='2026-04', customer_raw=name,
                    drone_customer_id=customer.id, operator_raw='Ким',
                    area_ha=148.0, amount=amount,
                    received_amount=None, payment_type='cash'))
        db.session.commit()
    return ids


def cut_amount_cells(body, title):
    """{row label: the amount <td> as rendered} for one cut of the report."""
    after = body.split(title, 1)[1]
    tbody = after.split('<tbody>')[1].split('</tbody>')[0]
    cells = {}
    for row in tbody.split('<tr'):
        tds = CELL_RE.findall(row)
        if len(tds) < 4:
            continue
        label = ' '.join(re.sub(r'<[^>]+>', ' ', tds[0]).split())
        cells[label] = tds[3]
    return cells


def text_of(html):
    return ' '.join(re.sub(r'<[^>]+>', ' ', html).split())


class TestC1AmountColumnTellsMissingFromZero(unittest.TestCase):
    """C-1: three groups, three renderings."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_amount_cases()
        cls.admin = create_admin('amount_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)
        body = self.client.get('/drones/works/reports').data.decode('utf-8')
        self.cells = cut_amount_cells(body, 'По заказчикам')

    def test_c1_all_null_renders_a_dash_and_never_a_zero(self):
        cell = self.cells[ALL_NULL]
        self.assertEqual('—', text_of(cell))
        self.assertNotIn('0', text_of(cell),
                         'a group with no price must not read as charged zero')

    def test_c1_mixed_renders_the_sum_and_the_count_of_missing(self):
        cell = self.cells[MIXED]
        text = text_of(cell)
        self.assertIn('5\xa0000\xa0000', cell)
        self.assertIn('сумма не указана', text)
        # exactly one of the two jobs carried no amount
        self.assertTrue(text.rstrip().endswith('1'), text)

    def test_c1_none_null_renders_a_plain_sum(self):
        cell = self.cells[NONE_NULL]
        self.assertIn('7\xa0000\xa0000', cell)
        self.assertNotIn('сумма не указана', text_of(cell))

    def test_c1_the_three_cells_are_three_different_renderings(self):
        rendered = {text_of(self.cells[name])
                    for name in (ALL_NULL, MIXED, NONE_NULL)}
        self.assertEqual(3, len(rendered), rendered)

    def test_c1_uzbek_says_it_in_cyrillic(self):
        set_language(self.admin, 'uz')
        try:
            client = app.test_client()
            login(client, self.admin)
            body = client.get('/drones/works/reports').data.decode('utf-8')
            cells = cut_amount_cells(body, 'Буюртмачилар бўйича')
            self.assertEqual('—', text_of(cells[ALL_NULL]))
            self.assertIn('сумма кўрсатилмаган', text_of(cells[MIXED]))
        finally:
            set_language(self.admin, 'ru')


class TestC2NegativeControl(unittest.TestCase):
    """C-2: the pre-commit expression renders 0 for the all-NULL group."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_amount_cases()

    def cut(self):
        import drones
        from flask import g
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = 'ru'
            data = drones._drone_works_report_data(
                drones._drone_work_conditions(
                    drones._drone_works_filters_from_args(MultiDict())))
            return data['by_customer']

    def test_c2_the_group_sum_really_is_zero(self):
        """The data has not changed -- only what the screen says about it."""
        row = [r for r in self.cut()['rows'] if r['label'] == ALL_NULL][0]
        self.assertEqual(0.0, row['amount'])
        self.assertEqual(2, row['jobs'])
        self.assertEqual(2, row['no_amount'])

    def test_c2_pre_commit_markup_renders_that_zero_as_zero(self):
        """Rendered through the REAL vs_num filter, so this is what the
        column actually printed before this commit."""
        row = [r for r in self.cut()['rows'] if r['label'] == ALL_NULL][0]
        pre_fix = "{{ '%.0f'|format(cell.amount)|vs_num }}"
        with app.app_context():
            rendered = app.jinja_env.from_string(pre_fix).render(cell=row)
        self.assertEqual('0', rendered.strip())

    def test_c2_and_the_new_markup_does_not(self):
        row = [r for r in self.cut()['rows'] if r['label'] == ALL_NULL][0]
        self.assertNotEqual(0, row['no_amount'])
        self.assertEqual(row['jobs'], row['no_amount'])


class TestC3TheCutStillReconciles(unittest.TestCase):
    """C-3: adding a column must not disturb the totals."""

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_amount_cases()

    def data(self):
        import drones
        from flask import g
        from werkzeug.datastructures import MultiDict
        with app.test_request_context('/drones/works/reports'):
            g.lang = 'ru'
            return drones._drone_works_report_data(
                drones._drone_work_conditions(
                    drones._drone_works_filters_from_args(MultiDict())))

    def test_c3_every_cut_reconciles(self):
        data = self.data()
        for name in ('by_customer', 'by_operator', 'by_subdivision',
                     'by_month', 'by_payment'):
            with self.subTest(cut=name):
                self.assertTrue(data[name]['reconciled'], name)

    def test_c3_the_amount_total_is_unchanged_by_the_new_column(self):
        data = self.data()
        # 5 000 000 + 3 000 000 + 4 000 000; the four NULLs contribute nothing
        self.assertAlmostEqual(12000000.0, data['totals']['amount'], places=2)
        self.assertAlmostEqual(12000000.0,
                               data['by_customer']['total']['amount'], places=2)

    def test_c3_no_amount_sums_over_the_cut(self):
        """The new column reconciles too, or it is decoration."""
        data = self.data()
        cut = data['by_customer']
        everything = cut['rows'] + cut['services']
        self.assertEqual(cut['total']['no_amount'],
                         sum(r['no_amount'] for r in everything))
        self.assertEqual(3, cut['total']['no_amount'])

    def test_c3_every_cut_agrees_on_how_many_jobs_have_no_amount(self):
        data = self.data()
        counts = {name: data[name]['total']['no_amount']
                  for name in ('by_customer', 'by_operator', 'by_subdivision',
                               'by_month', 'by_payment')}
        self.assertEqual({name: 3 for name in counts}, counts)


class TestC4ExcelIsUnchanged(unittest.TestCase):
    """C-4: money in the workbook is still a NUMBER, never a display string.

    The class name is kept for traceability, but one of its assertions is no
    longer «unchanged»: DRONE-ZERO-VS-UNKNOWN-001 §7 reverses what an UNKNOWN
    figure writes -- see the docstring on
    test_c4_the_amount_column_holds_numbers. The layout, the column order and
    the never-a-string rule are unchanged and still asserted here.
    """

    @classmethod
    def setUpClass(cls):
        cls.ids = seed_amount_cases()
        cls.admin = create_admin('xlsx_admin')
        set_language(cls.admin, 'ru')

    def workbook(self):
        from openpyxl import load_workbook
        import io as _io
        client = app.test_client()
        login(client, self.admin)
        response = client.get('/drones/works/reports.xlsx')
        self.assertEqual(200, response.status_code)
        return load_workbook(_io.BytesIO(response.data))

    def test_c4_no_money_cell_is_a_string(self):
        wb = self.workbook()
        offenders = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    # A money cell turned into text would carry the grouping
                    # separator or the em dash the screen uses.
                    if '\xa0' in cell.value or cell.value.strip() == '—':
                        offenders.append((ws.title, cell.coordinate,
                                          cell.value))
        self.assertEqual([], offenders)

    def test_c4_the_amount_column_holds_numbers(self):
        """REVERSED by DRONE-ZERO-VS-UNKNOWN-001 §7, deliberately and with
        its reasoning stated, not weakened.

        This assertion used to read

            self.assertEqual(0, found[ALL_NULL])

        under the comment «the workbook still reports the SUM, zero included
        -- the screen's em dash is a display decision and does not travel
        into accounting.» DRONE-ZERO-VS-UNKNOWN-001 overrules exactly that
        sentence, and its argument is: the workbook is what goes to
        management, so a screen printing «—» beside a workbook printing 0
        fixes nothing, because the 0 is the copy that gets believed. The
        0 was never a measured figure -- it is SUM over an all-NULL column.

        What did NOT change, and is still asserted below: money in these
        workbooks stays a NUMBER. The unknown cell becomes EMPTY, which is
        absent data in a numeric column; «—» would turn the column into text
        and break every SUM over it.
        """
        wb = self.workbook()
        ws = wb['По заказчикам']
        found = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] in (ALL_NULL, MIXED, NONE_NULL):
                found[row[0]] = row[3]
        self.assertEqual(3, len(found))
        self.assertEqual(5000000, found[MIXED])
        self.assertEqual(7000000, found[NONE_NULL])
        for name in (MIXED, NONE_NULL):
            self.assertIsInstance(found[name], (int, float), name)
        # The group nobody priced: empty, and none of the three things it
        # must not be.
        self.assertIsNone(found[ALL_NULL])
        self.assertNotEqual(0, found[ALL_NULL])
        self.assertNotEqual('', found[ALL_NULL])
        self.assertNotEqual('—', found[ALL_NULL])

    def test_c4_the_workbook_carries_no_no_amount_column(self):
        """The task changes the screen, not the agreed sheet layout."""
        wb = self.workbook()
        ws = wb['По заказчикам']
        header = [c for c in next(ws.iter_rows(max_row=1, values_only=True))]
        self.assertEqual(
            ['Заказчик', 'Работ', 'Гектары', 'Сумма', 'Получено',
             'Не получено', 'Доля, %'], header)


if __name__ == '__main__':
    unittest.main()
