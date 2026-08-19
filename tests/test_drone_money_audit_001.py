# -*- coding: utf-8 -*-
"""DRONE-MONEY-001: деньги против гектаров, проверка на синтетике.

Инструмент только читает и обязан уметь ОТЛИЧИТЬ сошедшуюся строку от
несошедшейся. Поэтому каждое тождество проверяется дважды: на данных, где
оно выполнено, и на данных, где оно нарушено ровно в одной строке.

Отдельно проверяется то, из-за чего отчёт может незаметно соврать:
  * «0 расхождений» при нуле проверенных строк -- отчёт обязан называть,
    сколько строк он ВООБЩЕ смог проверить;
  * строка без ставки не должна тихо считаться сошедшейся;
  * «топшириши керак» (долг оператора) не складывается с приходом;
  * ни один гектар и ни один сум не теряется при группировке;
  * ни одного пишущего слова в запросах -- проверяется разбором ast, а не
    доверием к глазам.

Run:
  python -m unittest tests.test_drone_money_audit_001 -v
"""
import ast
import os
import sqlite3
import sys
import tempfile
import unittest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)
sys.path.insert(0, os.path.join(REPO_ROOT, 'tools'))

import drone_money_audit as tool  # noqa: E402

SCHEMA = """
CREATE TABLE drone_operators (id INTEGER PRIMARY KEY, full_name TEXT);
CREATE TABLE drone_units (id INTEGER PRIMARY KEY, number INTEGER);
CREATE TABLE drone_operator_assignments (id INTEGER PRIMARY KEY,
  operator_id INTEGER, drone_unit_id INTEGER, date_from TEXT, date_to TEXT,
  note TEXT);
CREATE TABLE drone_flights (id INTEGER PRIMARY KEY, drone_unit_id INTEGER,
  started_at TEXT, area_ha REAL);
CREATE TABLE drone_works (id INTEGER PRIMARY KEY, period_month TEXT,
  work_date_from TEXT, work_date_to TEXT, date_raw TEXT,
  drone_operator_id INTEGER, customer_raw TEXT, area_ha NUMERIC,
  price_per_ha NUMERIC, amount NUMERIC, other_costs NUMERIC,
  received_amount NUMERIC, received_kind TEXT, payment_type TEXT,
  source_sheet TEXT, source_row INTEGER);
"""

# (оператор, заказчик, га, ставка, сумма, расходы, приход, вид)
WORKS = (
    (1, 'Розия Бекова фх', 34.5, 200000, 6900000, 250000, 6650000, 'received'),
    (1, 'Зухро Агро фх', 15.0, 200000, 3000000, 0, 3000000, 'received'),
    (1, 'Гарден Агрокластер', 309.0, 75040, 23187360, 0, None, None),
    (2, 'Шамси Зиё фх', 9.0, 200000, 1800000, 0, 1800000, 'received'),
    (2, 'Достонбек фх', 2.0, 200000, 400000, 0, 400000, 'operator_due'),
    # «Пул олинмаган»: работа сделана, деньги не собраны. Это НЕ сломанное
    # тождество -- диспетчер ведёт такие строки колонкой «Олинмаган сумма».
    (2, 'Учон фх', 21.4, 200000, 4280000, 0, 0, 'received'),
)


def build_db(path, works=WORKS, flights=((1, '2025-09-15', 40.0),
                                         (2, '2025-09-20', 10.0))):
    conn = sqlite3.connect(path)
    conn.executescript(SCHEMA)
    conn.execute("INSERT INTO drone_operators (id, full_name) VALUES "
                 "(1, 'Холмуродов Шахзод'), (2, 'Анваров Усмон')")
    conn.execute('INSERT INTO drone_units (id, number) VALUES (1, 3), (2, 4)')
    conn.execute("INSERT INTO drone_operator_assignments "
                 "(operator_id, drone_unit_id, date_from, date_to) VALUES "
                 "(1, 1, '2025-09-01', '2025-09-30'), "
                 "(2, 2, '2025-09-01', '2025-09-30')")
    for idx, (unit, day, area) in enumerate(flights, 1):
        conn.execute('INSERT INTO drone_flights (id, drone_unit_id, '
                     'started_at, area_ha) VALUES (?, ?, ?, ?)',
                     (idx, unit, day + ' 06:00:00', area))
    for idx, (op, customer, ha, rate, amount, costs, recv, kind) in \
            enumerate(works, 1):
        conn.execute(
            'INSERT INTO drone_works (id, period_month, work_date_from, '
            'work_date_to, date_raw, drone_operator_id, customer_raw, '
            'area_ha, price_per_ha, amount, other_costs, received_amount, '
            'received_kind, payment_type, source_sheet, source_row) '
            "VALUES (?, '2025-09', '2025-09-10', '2025-09-10', '10.09.2025', "
            "?, ?, ?, ?, ?, ?, ?, ?, 'cash', 'свод ичи', ?)",
            (idx, op, customer, ha, rate, amount, costs, recv, kind, idx))
    conn.commit()
    conn.close()


def run(db, month='2025-09'):
    conn = tool.connect_ro(db)
    try:
        works = tool.load_works(conn, month)
        telemetry = tool.load_telemetry(conn, month)
    finally:
        conn.close()
    return tool.audit(works), telemetry


class MoneyAuditTest(unittest.TestCase):

    def setUp(self):
        self.dir = tempfile.mkdtemp(prefix='money_')
        self.db = os.path.join(self.dir, 'transport.db')
        build_db(self.db)

    # Тождество 1 выполнено -- и отчёт это ВИДИТ, а не молчит.
    def test_clean_books_pass_and_report_how_many_were_checked(self):
        result, _telemetry = run(self.db)
        self.assertEqual([], result['bad_amount'])
        self.assertEqual(6, result['checked_amount'])

    # Тождество 1 нарушено ровно в одной строке.
    def test_a_wrong_amount_is_found(self):
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET amount = 7000000 WHERE id = 1')
        conn.commit()
        conn.close()
        result, _telemetry = run(self.db)
        self.assertEqual(1, len(result['bad_amount']))
        self.assertEqual(1, result['bad_amount'][0][0])
        self.assertAlmostEqual(6900000.0, result['bad_amount'][0][-1], 2)
        self.assertEqual(6, result['checked_amount'])

    # Тождество 2 выполнено, и считается только по «кирим».
    def test_received_identity_holds_and_skips_operator_due(self):
        result, _telemetry = run(self.db)
        self.assertEqual([], result['bad_received'])
        # Проверены три строки с kind='received'; долг, пустой приход и
        # «олинмаган» -- нет.
        self.assertEqual(3, result['checked_received'])

    # Тождество 2 нарушено.
    def test_a_wrong_received_is_found(self):
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET received_amount = 6000000 '
                     'WHERE id = 1')
        conn.commit()
        conn.close()
        result, _telemetry = run(self.db)
        self.assertEqual(1, len(result['bad_received']))
        self.assertAlmostEqual(6650000.0, result['bad_received'][0][-1], 2)

    # Долг оператора не складывается с приходом.
    def test_uncollected_money_is_its_own_bucket_not_a_discrepancy(self):
        """«Пул олинмаган» -- не ошибка, а неполученные деньги.

        [REASON]: боевой прогон 2026-08-19 дал четыре «расхождения», и все
        четыре оказались неполученными деньгами -- они сходятся с колонкой
        «Олинмаган сумма» подписанного свода до сума. Считать их
        расхождением значит прятать настоящие ошибки за ложными.
        """
        result, _telemetry = run(self.db)
        self.assertEqual([], result['bad_received'])
        self.assertEqual(1, len(result['unpaid']))
        self.assertEqual('Учон фх', result['unpaid'][0][2])
        self.assertAlmostEqual(4280000.0, result['unpaid'][0][-1], 2)
        self.assertAlmostEqual(
            4280000.0, result['per_operator']['Анваров Усмон']['unpaid'], 2)

    def test_a_partial_payment_is_still_a_discrepancy(self):
        """Отрицательный контроль: ноль -- «олинмаган», а недоплата -- ошибка.

        [REASON]: если бы корзина ловила ЛЮБОЙ приход меньше должного, она
        бы проглотила настоящие расхождения вместе с неполученными деньгами.
        """
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET received_amount = 1000000 '
                     'WHERE customer_raw = ?', ('Учон фх',))
        conn.commit()
        conn.close()
        result, _telemetry = run(self.db)
        self.assertEqual([], result['unpaid'])
        self.assertEqual(1, len(result['bad_received']))

    def test_operator_due_is_not_counted_as_received(self):
        result, _telemetry = run(self.db)
        anvarov = result['per_operator']['Анваров Усмон']
        self.assertAlmostEqual(1800000.0, anvarov['received'], 2)
        self.assertAlmostEqual(400000.0, anvarov['operator_due'], 2)

    # Строка без ставки не считается сошедшейся, а называется поимённо.
    def test_a_row_without_a_rate_is_named_not_silently_passed(self):
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET price_per_ha = NULL WHERE id = 2')
        conn.commit()
        conn.close()
        result, _telemetry = run(self.db)
        self.assertEqual(5, result['checked_amount'])
        self.assertEqual([], result['bad_amount'])
        reasons = [row[-1] for row in result['uncheckable'] if row[0] == 2]
        self.assertIn('нет ставки', reasons)

    # «0 расхождений» при нуле проверенных строк не выглядит как успех.
    def test_zero_checked_is_visible_in_the_summary(self):
        """[REASON]: отчёт «расходится 0» на пустой проверке -- самообман."""
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET price_per_ha = NULL')
        conn.commit()
        conn.close()
        result, telemetry = run(self.db)
        summary = dict(tool.summary_rows('2025-09', result, telemetry))
        self.assertEqual(0, summary['  Т1: строк, где было чем проверить'])
        self.assertEqual(0, summary['  Т1: расходится'])
        self.assertEqual(0, summary['  Т1: из них сошлось'])
        # Каждая из пяти строк попала в «проверить нечем» по обоим тождествам
        # (нет ставки) либо по одному -- ни одна не пропала молча.
        self.assertGreaterEqual(summary['Проверить нечем, случаев'], 6)

    # Тождество 3: ничего не потеряно.
    def test_nothing_is_lost_while_grouping(self):
        result, _telemetry = run(self.db)
        self.assertEqual([], tool.conservation(result))
        self.assertAlmostEqual(390.9, result['rows_total']['ha'], 2)
        self.assertAlmostEqual(
            390.9, sum(a['ha'] for a in result['per_operator'].values()), 2)

    # И отрицательный контроль на само тождество 3.
    def test_conservation_can_actually_fail(self):
        result, _telemetry = run(self.db)
        result['rows_total']['ha'] += 1.0
        problems = tool.conservation(result)
        self.assertTrue(any('hectares lost' in p for p in problems), problems)

    # Разница в гектарах переводится в деньги по СВОЕЙ ставке оператора.
    def test_the_gap_is_priced_at_the_operators_own_average_rate(self):
        result, telemetry = run(self.db)
        col = {name: i for i, (name, _w) in
               enumerate(tool.OPERATOR_COLUMNS)}
        rows = {r[0]: r for r in tool.operator_rows(result, telemetry)}
        kholm = rows['Холмуродов Шахзод']
        book_ha = kholm[col['Га книги']]
        tel_ha = kholm[col['Га телеметрии']]
        rate = kholm[col['Своя средняя ставка, сум/га']]
        gap_money = kholm[col['Разница в сумах']]
        self.assertAlmostEqual(358.5, book_ha, 2)
        self.assertAlmostEqual(40.0, tel_ha, 2)
        # 33 087 360 / 358.5 -- своя средняя, а не 200 000.
        self.assertNotEqual(200000, rate)
        # Столбцы обязаны сходиться между собой: перемножив показанное,
        # читатель получает ровно третий столбец.
        self.assertAlmostEqual(round((tel_ha - book_ha) * rate, 2),
                               gap_money, 2)

    # Спорное покрытие называется, а не растворяется в чьих-то гектарах.
    def test_a_day_covered_twice_goes_to_the_ambiguous_bucket(self):
        conn = sqlite3.connect(self.db)
        conn.execute("INSERT INTO drone_operator_assignments "
                     "(operator_id, drone_unit_id, date_from, date_to) "
                     "VALUES (2, 1, '2025-09-01', '2025-09-30')")
        conn.commit()
        conn.close()
        _result, telemetry = run(self.db)
        by_operator, _unassigned, ambiguous, total = telemetry
        self.assertAlmostEqual(40.0, ambiguous, 2)
        self.assertAlmostEqual(50.0, total, 2)
        self.assertNotIn('Холмуродов Шахзод', by_operator)

    # База открывается только на чтение -- запись обязана падать.
    def test_the_database_is_opened_read_only(self):
        conn = tool.connect_ro(self.db)
        try:
            with self.assertRaises(sqlite3.OperationalError):
                conn.execute('DELETE FROM drone_works')
        finally:
            conn.close()

    # И ни одного пишущего слова в самих запросах.
    def test_no_write_statements_anywhere_in_the_source(self):
        """[REASON]: read-only соединение защищает от ошибки, а не от замысла.

        Если кто-то потом откроет базу обычным connect(), запрос уже будет
        написан. Проверяются все строковые литералы файла.
        """
        path = os.path.join(REPO_ROOT, 'tools', 'drone_money_audit.py')
        with open(path, encoding='utf-8') as handle:
            tree = ast.parse(handle.read())
        banned = ('insert ', 'update ', 'delete ', 'drop ', 'alter ',
                  'create table', 'replace into')
        for node in ast.walk(tree):
            if isinstance(node, ast.Constant) and isinstance(node.value, str):
                lowered = node.value.lower()
                for word in banned:
                    self.assertNotIn(
                        word, lowered,
                        'пишущее слово %r в литерале: %r' % (word,
                                                             node.value[:60]))

    def test_the_operator_sheet_header_matches_the_rows(self):
        """Заголовок и строки берутся из одной константы, а не из двух мест."""
        result, telemetry = run(self.db)
        for row in tool.operator_rows(result, telemetry):
            self.assertEqual(len(tool.OPERATOR_COLUMNS), len(row))
        out = os.path.join(self.dir, 'cols.xlsx')
        tool.write_xlsx(out, '2025-09', result, telemetry)
        import openpyxl
        wb = openpyxl.load_workbook(out)
        header = next(wb['По операторам'].iter_rows(min_row=1, max_row=1,
                                                    values_only=True))
        self.assertEqual([n for n, _w in tool.OPERATOR_COLUMNS], list(header))
        wb.close()

    def test_xlsx_is_written_and_readable(self):
        result, telemetry = run(self.db)
        out = os.path.join(self.dir, 'money.xlsx')
        tool.write_xlsx(out, '2025-09', result, telemetry)
        import openpyxl
        wb = openpyxl.load_workbook(out)
        self.assertEqual(['Сводка', 'По операторам', 'Сумма не сходится',
                          'Приход не сходится', 'Проверить нечем',
                          'Не собрано (олинмаган)', 'Без оператора',
                          'Строки без оператора'],
                         wb.sheetnames)
        summary = {r[0]: r[1] for r in
                   wb['Сводка'].iter_rows(min_row=2, values_only=True)}
        self.assertEqual(6, summary['Работ в книгах, строк'])
        self.assertEqual('ДА', summary['Тождество 3: ничего не потеряно'])
        wb.close()


    # Консоль обязана быть ЧИТАЕМОЙ, а не «?????».
    def test_every_console_label_is_ascii(self):
        """[REASON]: прогон на production 2026-08-19 напечатал «?????»

        вместо каждой строки сводки -- русские подписи шли через
        errors='replace'. Отчёт, который нельзя прочитать, отчётом не
        является. Проверяется КАЖДАЯ подпись, которую сводка может выдать,
        а не выборка.
        """
        result, telemetry = run(self.db)
        for label, _value in tool.summary_rows('2025-09', result, telemetry):
            if not label:
                continue
            line = tool.console(label)
            self.assertFalse(line.startswith('?? '),
                             'подпись без ASCII-перевода: %r' % label)
            line.encode('ascii')

    def test_the_printed_report_is_ascii_end_to_end(self):
        """Проверять надо НАПЕЧАТАННОЕ, а не функцию перевода подписи.

        [REASON]: на production 2026-08-19 функция была правильной, а main()
        печатал русские подписи мимо неё, и владелец получил «?????» на
        каждой строке. Тест, который зовёт console() сам, этого не видит --
        поэтому здесь запускается весь main() и разбирается его вывод.
        """
        import contextlib
        import io as _io
        buffer = _io.StringIO()
        out = os.path.join(self.dir, 'money.xlsx')
        argv = sys.argv
        sys.argv = ['drone_money_audit.py', '--db', self.db,
                    '--month', '2025-09', '--out', out]
        try:
            with contextlib.redirect_stdout(buffer):
                code = tool.main()
        finally:
            sys.argv = argv
        self.assertEqual(0, code)
        printed = buffer.getvalue()
        printed.encode('ascii')          # весь вывод обязан быть ASCII
        self.assertNotIn('??', printed)  # и ничего непереведённого в нём нет
        self.assertIn('Identity 1: amount = hectares x rate', printed)
        self.assertIn('Book hectares', printed)
        self.assertIn('Identity 3: nothing was lost               YES',
                      printed)

    def test_an_unknown_label_is_visible_not_silently_mangled(self):
        """Отрицательный контроль: неизвестная подпись обязана быть заметна."""
        self.assertTrue(tool.console('Небывалая строка').startswith('?? '))

    # Строки без оператора группируются по листу-источнику.
    def test_orphan_rows_are_grouped_by_the_sheet_they_came_from(self):
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET drone_operator_id = NULL '
                     'WHERE id IN (1, 2)')
        conn.execute("UPDATE drone_works SET source_sheet = 'свод ичи (Усмон)'"
                     ' WHERE id = 4')
        conn.execute('UPDATE drone_works SET drone_operator_id = NULL '
                     'WHERE id = 4')
        conn.commit()
        conn.close()
        result, _telemetry = run(self.db)
        self.assertEqual(3, result['orphan_rows'])
        self.assertAlmostEqual(58.5, result['orphan_ha'], 2)
        groups = result['orphans']
        self.assertEqual({('свод ичи',), ('свод ичи (Усмон)',)}, set(groups))
        self.assertEqual(2, groups[('свод ичи',)]['rows'])
        self.assertAlmostEqual(49.5, groups[('свод ичи',)]['ha'], 2)
        self.assertEqual((1, 2), (groups[('свод ичи',)]['first'],
                                  groups[('свод ичи',)]['last']))
        self.assertEqual(3, len(result['orphan_detail']))

    def test_the_orphan_block_reaches_the_console(self):
        """Блок происхождения печатается -- иначе владелец его не увидит.

        [REASON]: 1130.60 га сентября не привязаны ни к кому, и починить это
        может только человек по своей книге. Если блок не напечатан, отчёт
        назвал проблему, но не дал по ней работать.
        """
        import contextlib
        import io as _io
        conn = sqlite3.connect(self.db)
        conn.execute('UPDATE drone_works SET drone_operator_id = NULL '
                     'WHERE id IN (1, 2)')
        conn.commit()
        conn.close()
        buffer = _io.StringIO()
        argv = sys.argv
        sys.argv = ['drone_money_audit.py', '--db', self.db, '--month',
                    '2025-09', '--out', os.path.join(self.dir, 'm.xlsx')]
        try:
            with contextlib.redirect_stdout(buffer):
                tool.main()
        finally:
            sys.argv = argv
        printed = buffer.getvalue()
        printed.encode('ascii')
        self.assertIn('ROWS WITH NO OPERATOR', printed)
        self.assertIn('49.50', printed)          # га двух осиротевших строк
        self.assertIn('Rows with no operator                      2', printed)

    def test_no_orphans_means_no_group(self):
        """Отрицательный контроль: на чистой базе группа пуста, а не выдумана."""
        result, _telemetry = run(self.db)
        self.assertEqual(0, result['orphan_rows'])
        self.assertEqual({}, result['orphans'])
        self.assertEqual([], result['orphan_detail'])


if __name__ == '__main__':
    unittest.main()
