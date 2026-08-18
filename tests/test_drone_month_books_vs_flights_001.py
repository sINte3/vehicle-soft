# -*- coding: utf-8 -*-
"""DRONE-SEPT2025-RECON-001: разбор «книга против телеметрии» на синтетике.

К каждой проверке -- отрицательный контроль, потому что проверка, дающая
одинаковый результат при верной и неверной карте, проверкой не является.
Ровно так и был найден дефект в самом инструменте: первая редакция считала
итогом месяца ВСЕ строки книги, поэтому отмена переноса справки, датированной
чужим месяцем, не меняла итог -- и проверка не различала два случая.

  1. Дата работы, а не книга, решает месяц. Строка, датированная другим
     месяцем, из итога месяца ИСКЛЮЧАЕТСЯ и называется отдельно.
     Отрицательный контроль: та же строка, объявленная в redated_rows,
     возвращается в месяц и меняет итог оператора.
  2. «Без даты» и «датирована другим месяцем» -- разные корзины. Первая
     входит в итог месяца, вторая нет; смешение даёт правдоподобный,
     но неверный итог.
  3. Диапазоны, перечисления и текстовая дата разбираются: «06-07.09.2025»
     -> два дня, «22,25.09.2025» -> два дня, «19-30.10.2025» -> двенадцать,
     «2025-09-26» -> один. Последнее -- дефект, найденный этим тестом:
     дату, набранную текстом, первая редакция отправляла в «без даты».
     Отрицательный контроль: «зимой» и «06-07» без месяца остаются без даты,
     а не получают выдуманное число.
  4. Оператор берётся из ИМЕНИ ЛИСТА, а не из колонки оператора: колонка в
     книгах врёт (сентябрь-2025, лист Жураева, 132.70 га подписаны чужим
     именем).
  5. Имя листа приводится к справочному картой алиасов; имя, которого в карте
     нет, -- отказ кодом 1 с названием и гектарами, а НЕ молчаливый ноль.
  6. Карта назначений обязана покрыть каждый вылет: непокрытый -> код 3.
  7. Гектары оператора считаются по окну машины, а не по машине целиком:
     одна машина, разделённая датой между двумя людьми, даёт им разные суммы.
  8. Каталога книг или выгрузки нет -> код 2, отчёт НЕ создан.
 11. Ни одна ПОДПИСЬ не уходит в файл формулой. openpyxl решает тип клетки по
     первому символу, поэтому подпись «= Отчёт по работам (расчёт)» попадала в
     файл формулой и открывалась ошибкой -- дефект, найденный проверкой формул
     после того, как LibreOffice в контейнере пересчитать отказался. Плюс книга
     несёт fullCalcOnLoad: без него итоги читаются пустыми до первого открытия.
 10. Лист «Летал-записи нет» ставит рядом строки книг БЕЗ ДАТЫ и дни, где
     записи нет: без этого он выставлял бы записанную работу незаписанной
     (у Нурали в сентябре 324.30 га справки без даты против 322.54 га таких
     дней). Дат инструмент при этом не проставляет.
 12. Читаются ТОЛЬКО расшифровки по операторам. Сводные листы, годовой журнал
     оператора и список неполученных денег пересказывают те же работы: на
     полном архиве сентября-2025 чтение всех листов давало 8 336.03 га против
     настоящих 5 617.03. Пропущенные листы ПЕЧАТАЮТСЯ, чтобы лист не исчез
     молча.
 13. Одно короткое имя листа -- разные люди в разных книгах: «свод ичи
     (Шохрух)» в книге Когона это Хамроев, в книге Гардена -- Файзуллаев.
     Общий алиас по короткому имени увёл бы 412.60 га не тому человеку
     (поймано контролем по своду). Привязка листа к человеку -- явная,
     файлом и листом.
 15. Тип таблицы -- наличные или справка -- определяется её СТРУКТУРОЙ, а не
     меткой над ней: наличная несёт колонки расхода и прихода, справка нет.
     Метка врёт: в книге Агрокластера слова «Накд» над наличной таблицей нет
     вовсе, и все 826.80 га Сайфулло уходили в справку молча. Отрицательный
     контроль: та же таблица без колонок расхода и прихода читается справкой.
 16. Дата и деньги в соседних колонках не путаются: «Кирим қилинган сана» --
     дата, «Кирим қилинган» -- полученные деньги, отличаются одним словом.
     Баланс «не сдано» = сумма наличными − расходы − сдано, считается по
     строкам, а не берётся из итога книги.
 14. Свод самих диспетчеров читается как КОНТРОЛЬ и сверяется с ПОЛНЫМ
     разбором книги, включая строки с датой другого месяца: иначе контроль
     показывает ложное расхождение ровно на них. Отрицательный контроль:
     подменённое число в своде даёт видимую разницу.
  9. Мост к отчёту приложения сходится в ноль и называет каждое расхождение
     поимённо: строку, которой в отчёте нет; строку, где числа разошлись;
     строку, которой нет в книгах. Отрицательный контроль: если отчёт
     подменить, невязка перестаёт быть нулевой -- то есть мост считает, а
     не подтверждает сам себя.

Run:
  python -m unittest tests.test_drone_month_books_vs_flights_001 -v
"""
import datetime
import json
import os
import subprocess
import sys
import tempfile
import unittest

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOOL = os.path.join(REPO_ROOT, 'tools', 'drone_month_books_vs_flights.py')
sys.path.insert(0, os.path.join(REPO_ROOT, 'tools'))

CASH_HEADER = ['№', 'ФХ номи', 'Майдон (га)', 'Хизмат кўрсатиш суммаси',
               'Жами сумма', 'Бошка харажатлар', 'Кирим қилинган',
               'Кирим қилинган сана', 'Дрон бошқарувчи оператор', '*Изоҳ']
TRANSFER_HEADER = ['№', 'ФХ номи', 'Мадон (га)', 'Хизмат кўрсатиш санаси',
                   'Хизмат кўрсатиш суммаси', 'Жами сумма']


def write_book(path, sheets):
    """sheets: {лист: {'cash': [(фх, га, дата, оператор_в_колонке)],
                       'transfer': [(фх, га, дата)]}}"""
    book = openpyxl.Workbook()
    book.remove(book.active)
    for title, tables in sheets.items():
        sheet = book.create_sheet(title[:31])
        row = 1
        sheet.cell(row=row, column=1, value='Маълумот')
        row += 1
        if tables.get('cash'):
            sheet.cell(row=row, column=1, value='Накд')
            row += 1
            for col, name in enumerate(CASH_HEADER, 1):
                sheet.cell(row=row, column=col, value=name)
            row += 1
            for idx, (farm, area, date, who) in enumerate(tables['cash'], 1):
                sheet.cell(row=row, column=1, value=idx)
                sheet.cell(row=row, column=2, value=farm)
                sheet.cell(row=row, column=3, value=area)
                sheet.cell(row=row, column=8, value=date)
                sheet.cell(row=row, column=9, value=who)
                row += 1
            sheet.cell(row=row, column=2, value='Жами сумма:')
            row += 2
        if tables.get('transfer'):
            sheet.cell(row=row, column=1, value='Справка')
            row += 1
            for col, name in enumerate(TRANSFER_HEADER, 1):
                sheet.cell(row=row, column=col, value=name)
            row += 1
            for idx, (farm, area, date) in enumerate(tables['transfer'], 1):
                sheet.cell(row=row, column=1, value=idx)
                sheet.cell(row=row, column=2, value=farm)
                sheet.cell(row=row, column=3, value=area)
                sheet.cell(row=row, column=4, value=date)
                row += 1
            sheet.cell(row=row, column=1, value='Жами:')
    book.save(path)


def write_report(path, total, customers, operators=()):
    """Отчёт по работам приложения: «Сводка», «По заказчикам», «По операторам»."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Сводка'
    sheet.cell(row=1, column=1, value='Показатель')
    sheet.cell(row=1, column=2, value='Значение')
    sheet.cell(row=2, column=1, value='Гектаров')
    sheet.cell(row=2, column=2, value=total)
    for title, data in (('По заказчикам', customers),
                        ('По операторам', operators)):
        sheet2 = book.create_sheet(title)
        for col, name in enumerate(['Имя', 'Работ', 'Гектары'], 1):
            sheet2.cell(row=1, column=col, value=name)
        for idx, (name, area) in enumerate(data, 2):
            sheet2.cell(row=idx, column=1, value=name)
            sheet2.cell(row=idx, column=2, value=1)
            sheet2.cell(row=idx, column=3, value=area)
    book.save(path)


def write_flights(path, rows):
    """rows: [(дата-время, машина, ник, адрес, га)]"""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Вылеты'
    for col, name in enumerate(['Дата и время (UTC+5)', 'Машина (№)',
                                'Ник (DJI)', 'Область', 'Адрес', 'Гектары'], 1):
        sheet.cell(row=1, column=col, value=name)
    for idx, (when, machine, nick, addr, area) in enumerate(rows, 2):
        sheet.cell(row=idx, column=1, value=when)
        sheet.cell(row=idx, column=2, value=machine)
        sheet.cell(row=idx, column=3, value=nick)
        sheet.cell(row=idx, column=4, value='Бухарская область')
        sheet.cell(row=idx, column=5, value=addr)
        sheet.cell(row=idx, column=6, value=area)
    book.save(path)


class MonthReconTest(unittest.TestCase):

    def setUp(self):
        self.dir = tempfile.mkdtemp(prefix='recon_')
        self.books = os.path.join(self.dir, 'books')
        os.makedirs(self.books)
        write_book(os.path.join(self.books, 'Когон ПТЗ Дрон маълумот.xlsx'), {
            'свод ичи Ибодуллаев Хасан': {
                # Колонка оператора врёт: подписана чужим именем.
                'cash': [('Ферма А', 40.0,
                          datetime.datetime(2025, 9, 20), 'Қодиров Нурали'),
                         ('Ферма Б', 20.0, '06-07.09.2025', 'Қодиров Нурали'),
                         ('Ферма В', 10.0, '22,25.09.2025', None),
                         ('Ферма Г', 4.0, '2025-09-26', None)],
                'transfer': [('Справка октября', 100.0, '19-30.10.2025'),
                             ('Справка без даты', 7.0, None),
                             ('Справка зимой', 3.0, 'зимой')],
            },
        })
        self.flights = os.path.join(self.dir, 'flights.xlsx')
        write_flights(self.flights, [
            ('2025-09-06 10:00:00', 11, 'Kogon№8', 'Kogon District', 9.0),
            ('2025-09-07 10:00:00', 11, 'Kogon№8', 'Kogon District', 9.0),
            ('2025-09-20 10:00:00', 11, 'Kogon№8', 'Kogon District', 40.0),
            ('2025-09-22 10:00:00', 11, 'Kogon№8', 'Kogon District', 5.0),
            ('2025-09-25 10:00:00', 11, 'Kogon№8', 'Kogon District', 5.0),
            ('2025-09-26 10:00:00', 11, 'Kogon№8', 'Kogon District', 6.0),
            # Август -- вне месяца, в отчёт попасть не должен.
            ('2025-08-31 10:00:00', 11, 'Kogon№8', 'Kogon District', 500.0),
        ])
        self.out = os.path.join(self.dir, 'out.xlsx')
        # Отчёт приложения: «Ферма Г» в него не попала, «Ферма А» урезана
        # до 30, зато есть «Чужая ферма» 5.0, которой нет в книгах.
        # 84 (месяц) + 100 (октябрьская строка) - 4 - 10 + 5 = 175.
        self.report = os.path.join(self.dir, 'report.xlsx')
        write_report(self.report, 175.0,
                     [('Ферма А', 30.0), ('Ферма Б', 20.0),
                      ('Ферма В', 10.0), ('Справка октября', 100.0),
                      ('Справка без даты', 7.0), ('Справка зимой', 3.0),
                      ('Чужая ферма', 5.0)],
                     [('Ибодуллаев Хасанбой', 175.0)])

    def spec(self, **over):
        base = {
            'month': '2025-09',
            'operator_aliases': {'ибодуллаев хасан': 'Ибодуллаев Хасанбой'},
            'assignments': [{'machine': 11, 'from': '2025-09-01',
                             'to': '2025-09-30',
                             'operator': 'Ибодуллаев Хасанбой', 'why': 'тест'}],
        }
        base.update(over)
        path = os.path.join(self.dir, 'spec_%d.json' % len(over))
        with open(path, 'w', encoding='utf-8') as handle:
            json.dump(base, handle, ensure_ascii=False)
        return path

    def run_tool(self, spec_path, out=None, extra=()):
        args = [sys.executable, TOOL, '--books-dir', self.books,
                '--flights', self.flights, '--month', '2025-09',
                '--out', out or self.out]
        if spec_path:
            args += ['--assignments', spec_path]
        args += list(extra)
        return subprocess.run(args, capture_output=True, text=True)

    def operators(self, path):
        book = openpyxl.load_workbook(path, data_only=False)
        sheet = book['По операторам']
        out = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or row[0] == 'ИТОГО':
                continue
            out[row[0]] = {'book': row[2], 'undated': row[3], 'tele': row[4]}
        book.close()
        return out

    # 1 + 2. Месяц решает дата работы; «без даты» и «чужой месяц» -- разное.
    def test_other_month_row_is_excluded_from_the_month(self):
        result = self.run_tool(self.spec())
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        row = self.operators(self.out)['Ибодуллаев Хасанбой']
        # Наличные с датой 40 + 20 + 10 + 4 = 74, плюс 7 без даты и 3
        # нечитаемых = 84. Справка октября (100) в месяц не входит.
        self.assertAlmostEqual(84.0, row['book'], places=2)
        self.assertAlmostEqual(10.0, row['undated'], places=2)
        self.assertAlmostEqual(74.0, row['tele'], places=2)

    # 1, отрицательный контроль: объявленный перенос меняет итог.
    def test_redated_row_returns_into_the_month(self):
        spec = self.spec(redated_rows=[{
            'operator': 'Ибодуллаев Хасанбой', 'farm': 'Справка октября',
            'ha': 100.0, 'days': ['2025-09-28', '2025-09-29'],
        }])
        result = self.run_tool(spec)
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        row = self.operators(self.out)['Ибодуллаев Хасанбой']
        self.assertAlmostEqual(184.0, row['book'], places=2)

    # 3. Диапазоны и перечисления; отрицательный контроль -- мусор без даты.
    def test_date_ranges_and_lists_are_parsed_garbage_is_not(self):
        import drone_month_books_vs_flights as tool
        self.assertEqual(['2025-09-06', '2025-09-07'],
                         tool.parse_date_cell('06-07.09.2025', 2025, 9)[1])
        self.assertEqual(['2025-09-22', '2025-09-25'],
                         tool.parse_date_cell('22,25.09.2025', 2025, 9)[1])
        self.assertEqual(12, len(tool.parse_date_cell('19-30.10.2025',
                                                      2025, 9)[1]))
        self.assertEqual(['2025-09-26'],
                         tool.parse_date_cell('2025-09-26', 2025, 9)[1])
        self.assertEqual(['2025-09-26'],
                         tool.parse_date_cell('2025-09-26 07:15:00',
                                              2025, 9)[1])
        for garbage in ('зимой', '', None, '35.09.2025', '30-19.09.2025'):
            self.assertEqual([], tool.parse_date_cell(garbage, 2025, 9)[1],
                             'из %r нельзя выводить дату' % (garbage,))

    # 4. Оператор -- из имени листа, не из колонки.
    def test_operator_comes_from_the_sheet_name_not_the_column(self):
        rows = self.operators(self.out) if os.path.exists(self.out) else None
        if rows is None:
            self.run_tool(self.spec())
            rows = self.operators(self.out)
        self.assertIn('Ибодуллаев Хасанбой', rows)
        self.assertNotIn('Қодиров Нурали', rows)

    # 5. Непокрытое картой имя -- отказ с числом, не молчаливый ноль.
    def test_unknown_operator_name_fails_loudly(self):
        result = self.run_tool(self.spec(operator_aliases={}))
        self.assertEqual(1, result.returncode, result.stdout)
        self.assertIn('absent from the assignment map', result.stdout)
        self.assertIn('84.00', result.stdout)
        self.assertFalse(os.path.exists(self.out))

    # 6. Непокрытый вылет -- код 3.
    def test_uncovered_flight_fails_with_code_three(self):
        spec = self.spec(assignments=[{'machine': 11, 'from': '2025-09-01',
                                       'to': '2025-09-19',
                                       'operator': 'Ибодуллаев Хасанбой',
                                       'why': 'окно короче месяца'}])
        result = self.run_tool(spec)
        self.assertEqual(3, result.returncode, result.stdout)
        self.assertIn('not covered', result.stdout)

    # 7. Одна машина, разделённая датой, даёт двум людям разные суммы.
    def test_one_machine_split_by_date_splits_hectares(self):
        spec = self.spec(
            operator_aliases={'ибодуллаев хасан': 'Ибодуллаев Хасанбой'},
            assignments=[
                {'machine': 11, 'from': '2025-09-01', 'to': '2025-09-10',
                 'operator': 'Хамроев Шохрух', 'why': 'первые дни'},
                {'machine': 11, 'from': '2025-09-11', 'to': '2025-09-30',
                 'operator': 'Ибодуллаев Хасанбой', 'why': 'остаток'},
            ])
        result = self.run_tool(spec)
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        rows = self.operators(self.out)
        self.assertAlmostEqual(18.0, rows['Хамроев Шохрух']['tele'], places=2)
        self.assertAlmostEqual(56.0, rows['Ибодуллаев Хасанбой']['tele'],
                               places=2)

    # 8. Нет входных файлов -> код 2, отчёт не создан.
    def test_missing_inputs_give_code_two_and_no_report(self):
        args = [sys.executable, TOOL, '--books-dir',
                os.path.join(self.dir, 'nope'), '--flights', self.flights,
                '--month', '2025-09', '--out', self.out]
        result = subprocess.run(args, capture_output=True, text=True)
        self.assertEqual(2, result.returncode, result.stdout)
        self.assertFalse(os.path.exists(self.out))

    # 9. Мост сходится в ноль и называет расхождения поимённо.
    def test_bridge_to_the_application_report_closes(self):
        result = self.run_tool(self.spec(),
                               extra=['--works-report', self.report])
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        self.assertIn('residual +0.00', result.stdout)
        book = openpyxl.load_workbook(self.out)
        rows = [r for r in book['Мост к отчёту'].iter_rows(values_only=True)]
        book.close()
        text = ' | '.join(str(c) for r in rows for c in r if c is not None)
        # Имя показывается КАК В КНИГЕ, а не свёрнутым ключом.
        self.assertIn('Ферма Г', text)        # нет в отчёте
        self.assertIn('Ферма А', text)        # числа разошлись
        self.assertIn('Чужая ферма', text)    # нет в книгах
        self.assertNotIn('ферма г', text)
        # Все строки книг, включая октябрьскую: 84 + 100 = 184.
        self.assertAlmostEqual(184.0, rows[3][1], places=2)

    # 9, отрицательный контроль: подменённый отчёт ломает невязку.
    def test_bridge_residual_is_not_self_fulfilling(self):
        bad = os.path.join(self.dir, 'bad_report.xlsx')
        write_report(bad, 999.0, [('Ферма А', 30.0)])
        result = self.run_tool(self.spec(), extra=['--works-report', bad])
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        self.assertNotIn('residual +0.00', result.stdout)
        self.assertIn('residual', result.stdout)

    # 10. Строки без даты сопоставлены с днями без записи, а не забыты.
    def test_undated_rows_are_matched_against_days_without_a_record(self):
        result = self.run_tool(self.spec())
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        book = openpyxl.load_workbook(self.out)
        rows = list(book['Летал-записи нет'].iter_rows(values_only=True))
        book.close()
        head = next(i for i, r in enumerate(rows)
                    if r and r[0] == 'Сопоставление строк книг без даты')
        body = rows[head + 2]
        self.assertEqual('Ибодуллаев Хасанбой', body[0])
        # 7 без даты + 3 нечитаемых = 10 га книг; свободных дней телеметрии
        # 06 (9.0) и 07 (9.0) записаны, значит свободен только 22/25 -- нет:
        # 22 и 25 записаны строкой «22,25.09.2025». Свободных дней нет.
        self.assertAlmostEqual(10.0, body[1], places=2)
        self.assertAlmostEqual(0.0, body[2], places=2)

    # 11. Подписи не уходят формулами, книга пересчитывается при открытии.
    def test_no_label_is_written_as_a_formula(self):
        import drone_month_books_vs_flights as tool
        result = self.run_tool(self.spec(),
                               extra=['--works-report', self.report])
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        book = openpyxl.load_workbook(self.out)
        self.assertTrue(book.calculation.fullCalcOnLoad)
        bad = []
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != 'f':
                        continue
                    if not tool._FORMULA.fullmatch(str(cell.value)):
                        bad.append('%s!%s = %r'
                                   % (sheet.title, cell.coordinate, cell.value))
        book.close()
        self.assertEqual([], bad, 'подписи ушли формулами: %s' % bad)

    # 11, отрицательный контроль: гвардия ловит подпись со знака равенства.
    def test_label_starting_with_equals_is_refused(self):
        import drone_month_books_vs_flights as tool
        sheet = openpyxl.Workbook().active
        with self.assertRaises(ValueError):
            tool._put(sheet, 1, 1, '= Отчёт по работам (расчёт)')
        # Намеренная формула той же формы проходит.
        tool._put(sheet, 2, 1, '=SUM(B1:B9)')
        self.assertEqual('=SUM(B1:B9)', sheet.cell(row=2, column=1).value)

    # 12. Не-расшифровки пропускаются и называются.
    def test_only_breakdown_sheets_are_read(self):
        path = os.path.join(self.books, 'Когон ПТЗ Дрон маълумот.xlsx')
        book = openpyxl.load_workbook(path)
        # Годовой журнал того же человека: те же работы, другой лист.
        extra = book.create_sheet('Расул ака Дрон маълумот обём')
        for col, name in enumerate(CASH_HEADER, 1):
            extra.cell(row=1, column=col, value=name)
        extra.cell(row=2, column=1, value=1)
        extra.cell(row=2, column=2, value='Ферма А')
        extra.cell(row=2, column=3, value=999.0)
        extra.cell(row=2, column=8, value=datetime.datetime(2025, 9, 20))
        book.save(path)
        book.close()
        result = self.run_tool(self.spec())
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        # 999 га не попали в итог, а лист назван в выводе.
        self.assertAlmostEqual(
            84.0, self.operators(self.out)['Ибодуллаев Хасанбой']['book'],
            places=2)
        # Пропущенный лист назван в выводе; кириллица в консоли проекта
        # заменяется на «?», поэтому проверяется ASCII-часть строки.
        self.assertIn('skipped 1 sheet(s)', result.stdout)

    # 13. Одно короткое имя -- разные люди в разных книгах.
    def test_same_short_sheet_name_means_different_people(self):
        for name, area in (('Гарден Агрокластер Дрон маълумот.xlsx', 11.0),
                           ('Шофиркон ПТЗ Дрон маълумот.xlsx', 22.0)):
            write_book(os.path.join(self.books, name), {
                'свод ичи (Шохрух)': {
                    'cash': [('Ферма Ш', area,
                              datetime.datetime(2025, 9, 20), None)]},
            })
        spec = self.spec(sheet_operators=[
            {'file': 'Гарден Агрокластер Дрон маълумот.xlsx',
             'sheet': 'свод ичи (Шохрух)', 'operator': 'Файзуллаев Шохрух'},
            {'file': 'Шофиркон ПТЗ Дрон маълумот.xlsx',
             'sheet': 'свод ичи (Шохрух)', 'operator': 'Хамроев Шохрух'},
            {'file': 'Когон ПТЗ Дрон маълумот.xlsx',
             'sheet': 'свод ичи Ибодуллаев Хасан',
             'operator': 'Ибодуллаев Хасанбой'},
        ], assignments=[
            {'machine': 11, 'from': '2025-09-01', 'to': '2025-09-30',
             'operator': 'Ибодуллаев Хасанбой', 'why': 'тест'},
            {'machine': 21, 'from': '2025-09-01', 'to': '2025-09-30',
             'operator': 'Файзуллаев Шохрух', 'why': 'тест'},
            {'machine': 22, 'from': '2025-09-01', 'to': '2025-09-30',
             'operator': 'Хамроев Шохрух', 'why': 'тест'},
        ])
        result = self.run_tool(spec)
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        rows = self.operators(self.out)
        self.assertAlmostEqual(11.0, rows['Файзуллаев Шохрух']['book'], places=2)
        self.assertAlmostEqual(22.0, rows['Хамроев Шохрух']['book'], places=2)

    # 14. Свод диспетчеров как контроль; отрицательный контроль на него же.
    def test_control_sheet_compares_against_the_whole_book(self):
        import drone_month_books_vs_flights as tool
        path = os.path.join(self.books, 'Свод.xlsx')
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = 'ОБШИЙ СВОД '
        sheet.cell(row=1, column=3, value='Дрон бошқарувчи оператор')
        sheet.cell(row=2, column=3, value='Ибодуллаев Хасан')
        sheet.cell(row=2, column=4, value=110.0)   # справка
        sheet.cell(row=2, column=5, value=74.0)    # наличные
        sheet.cell(row=2, column=6, value=9000.0)  # жами сумма
        sheet.cell(row=2, column=7, value=8500.0)  # олинган
        sheet.cell(row=2, column=8, value=500.0)   # олинмаган
        sheet.cell(row=3, column=3, value='Жами:')
        book.save(path)
        book.close()
        control = tool.read_control(
            path, 'ОБШИЙ СВОД ',
            {'ибодуллаев хасан': 'Ибодуллаев Хасанбой'})
        entry = control['Ибодуллаев Хасанбой']
        # 184.00 -- ВСЕ строки книги, включая октябрьскую справку 100 га.
        self.assertAlmostEqual(
            184.0, entry['area_transfer'] + entry['area_cash'], places=2)
        self.assertEqual(['Ибодуллаев Хасан'], entry['spellings'])
        # Деньги свода читаются теми же смещениями.
        self.assertAlmostEqual(9000.0, entry['amount'], places=2)
        self.assertAlmostEqual(500.0, entry['not_received'], places=2)
        # Отрицательный контроль: подменённое число даёт другую сумму.
        book = openpyxl.load_workbook(path)
        book['ОБШИЙ СВОД '].cell(row=2, column=4, value=10.0)
        book.save(path)
        book.close()
        entry2 = tool.read_control(
            path, 'ОБШИЙ СВОД ',
            {'ибодуллаев хасан': 'Ибодуллаев Хасанбой'})['Ибодуллаев Хасанбой']
        self.assertAlmostEqual(
            84.0, entry2['area_transfer'] + entry2['area_cash'], places=2)

    # 14, отрицательный контроль: объявленный свод, которого нет -> код 2.
    def test_declared_control_sheet_missing_is_refused(self):
        spec = self.spec(control_sheet={'file': 'нет-такого.xlsx',
                                        'sheet': 'ОБШИЙ СВОД '})
        result = self.run_tool(spec)
        self.assertEqual(2, result.returncode, result.stdout)
        self.assertIn('control sheet file not found', result.stdout)

    # 15. Тип таблицы -- по структуре, а не по метке над ней.
    def test_table_kind_comes_from_its_columns_not_the_marker(self):
        import drone_month_books_vs_flights as tool
        path = os.path.join(self.dir, 'no_marker.xlsx')
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = 'свод ичи (Тест)'
        # Метки «Накд» нет; выше стоит слово «Справка» -- как в книге
        # Агрокластера, где на этом и терялись 826.80 га наличных.
        sheet.cell(row=1, column=1, value='Справка')
        for col, name in enumerate(CASH_HEADER, 1):
            sheet.cell(row=2, column=col, value=name)
        sheet.cell(row=3, column=1, value=1)
        sheet.cell(row=3, column=2, value='Ферма Н')
        sheet.cell(row=3, column=3, value=50.0)
        sheet.cell(row=3, column=5, value=1000.0)   # Жами сумма
        sheet.cell(row=3, column=6, value=100.0)    # Бошка харажатлар
        sheet.cell(row=3, column=7, value=900.0)    # Кирим қилинган
        sheet.cell(row=3, column=8, value=datetime.datetime(2025, 9, 20))
        book.save(path)
        book.close()
        rows = tool._read_sheet(path, openpyxl.load_workbook(path)['свод ичи (Тест)'],
                                2025, 9, {}, {})
        self.assertEqual(1, len(rows))
        self.assertEqual(tool.SECTION_CASH, rows[0]['section'])
        # Дата не утекла в деньги, деньги не утекли в дату.
        self.assertAlmostEqual(1000.0, rows[0]['amount'], places=2)
        self.assertAlmostEqual(100.0, rows[0]['expenses'], places=2)
        self.assertAlmostEqual(900.0, rows[0]['received'], places=2)
        self.assertEqual(['2025-09-20'], rows[0]['days'])

    # 15, отрицательный контроль: без колонок расхода и прихода -- справка.
    def test_table_without_money_columns_is_a_transfer_table(self):
        import drone_month_books_vs_flights as tool
        path = os.path.join(self.dir, 'transfer.xlsx')
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = 'свод ичи (Тест)'
        sheet.cell(row=1, column=1, value='Накд')   # метка врёт наоборот
        for col, name in enumerate(TRANSFER_HEADER, 1):
            sheet.cell(row=2, column=col, value=name)
        sheet.cell(row=3, column=1, value=1)
        sheet.cell(row=3, column=2, value='Ферма С')
        sheet.cell(row=3, column=3, value=20.0)
        sheet.cell(row=3, column=6, value=500.0)
        book.save(path)
        book.close()
        rows = tool._read_sheet(path, openpyxl.load_workbook(path)['свод ичи (Тест)'],
                                2025, 9, {}, {})
        self.assertEqual(tool.SECTION_TRANSFER, rows[0]['section'])

    # 16. Баланс: «не сдано» считается по строкам книги.
    def test_balance_is_computed_from_rows(self):
        result = self.run_tool(self.spec())
        self.assertEqual(0, result.returncode, result.stdout + result.stderr)
        book = openpyxl.load_workbook(self.out)
        rows = list(book['Баланс по операторам'].iter_rows(values_only=True))
        book.close()
        body = rows[2]
        self.assertEqual('Ибодуллаев Хасанбой', body[0])
        # В синтетике денег нет: суммы нулевые, но лист построен и считает.
        self.assertEqual('=E3-G3-H3', body[8])

    # Самоконтроль --expect-books-ha различает верное и неверное число.
    def test_expect_books_ha_discriminates(self):
        ok = self.run_tool(self.spec(), extra=['--expect-books-ha', '184.0'])
        self.assertEqual(0, ok.returncode, ok.stdout + ok.stderr)
        bad = self.run_tool(self.spec(), extra=['--expect-books-ha', '999.0'])
        self.assertEqual(1, bad.returncode, bad.stdout)
        self.assertIn('expected', bad.stdout)


if __name__ == '__main__':
    unittest.main()
