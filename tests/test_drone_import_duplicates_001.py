# -*- coding: utf-8 -*-
"""DRONE-IMPORT-DUP-001: поиск книг, загруженных дважды.

ТРИ ЛОВУШКИ, И ВСЕ ТРИ ПОЙМАНЫ НА PRODUCTION, А НЕ ЗДЕСЬ.

1. ЛИСТ-ТЁЗКА. «свод ичи (Шохрух)» есть и в книге Гардена (Файзуллаев
   Шохрух), и в книге Когона (Хамроев Шохрух). Первая редакция объявила
   гарденскую книгу лишней и предложила удалить 154.70 га живых работ.
   Отличает их ОПЕРАТОР.

2. СТОРОНА БЕЗ ОПЕРАТОРА. Молчание -- не согласие; такие группы UNKNOWN.

3. КНИГА ДРУГОГО МЕСЯЦА. Строки БЕЗ ДАТЫ стоят в месяце по period_month из
   формы загрузки. Книгу октября загрузили с периодом сентября -- её
   датированные строки ушли в октябрь, а недатированные остались в сентябре
   рядом с сентябрьской книгой того же человека, и это неотличимо от второй
   загрузки. 2026-08-20 по такому «дублю» с production удалили 15 работ на
   118.40 га; они оказались наличными блоками ОКТЯБРЬСКОЙ книги Сервиса.

ПОЧЕМУ ЭТОТ ФАЙЛ НЕ ПОЙМАЛ ТРЕТЬЮ. Прежняя фикстура давала дату КАЖДОЙ
строке, включая сторону-«остаток»: `work_date_from = month + '-10'`. Таких
строк в базе не было ни одной -- у настоящих 118.40 га даты нет вовсе.
Проверка на выдуманных данных одинаково зелена и при верном, и при неверном
коде. Теперь фикстура повторяет форму настоящих данных, и различие проверено
отрицательным контролем в обе стороны: недатированная сторона БЕЗ следа в
другом месяце остаётся неопознанной, а датированная -- кандидатом.

Run:
  python -m unittest tests.test_drone_import_duplicates_001 -v
"""
import ast
import os
import sqlite3
import sys
import tempfile
import unittest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)
sys.path.insert(0, os.path.join(REPO_ROOT, 'tools'))

import drone_import_duplicates as tool  # noqa: E402

SCHEMA = """
CREATE TABLE drone_operators (id INTEGER PRIMARY KEY, full_name TEXT);
CREATE TABLE drone_works (id INTEGER PRIMARY KEY AUTOINCREMENT,
  period_month TEXT NOT NULL, work_date_from TEXT, customer_raw TEXT,
  area_ha NUMERIC, amount NUMERIC, drone_operator_id INTEGER,
  source_file TEXT, source_sheet TEXT, source_row INTEGER);
"""

OPERATORS = ('Қудратов Мухриддин', 'Жумаев Фуркат', 'Имомов Беҳзод',
             'Қодиров Нурали', 'Файзуллаев Шохрух', 'Хамроев Шохрух',
             'Жураев Туйғун', 'Ибодуллаев Хасанбой')

# Сентябрьская книга Сервиса -- настоящая, с датами сентября.
SEPT = 'Сервис Дрон Маълумот (2).xlsx'
# ОКТЯБРЬСКАЯ книга Сервиса, загруженная с периодом «2025-09». Её
# датированные строки легли в октябрь по своим датам, недатированные остались
# в сентябре по period_month. Это и есть третья ловушка.
OCTOBER = 'Сервис Дрон Маълумот.xlsx'
# Настоящая двойная загрузка: обе стороны датированы своим месяцем.
TWICE = 'Шофиркон ПТЗ Дрон Октябрь.xlsx'
TWICE2 = 'Шофиркон ПТЗ Дрон Октябрь (2).xlsx'
# Двойная загрузка, у которой «остаток» недатирован, а следа в другом месяце
# нет: доказать «книга другого месяца» нечем, и удалять по ней нельзя тоже.
BLIND = 'Пешку ПТЗ Дрон Ноябрь.xlsx'
BLIND2 = 'Пешку ПТЗ Дрон Ноябрь (2).xlsx'

# (файл, лист, строк, га, period_month, оператор[, месяц дат])
# Седьмой элемент: 'YYYY-MM' -- даты этого месяца; None -- строки БЕЗ ДАТЫ;
# отсутствует -- даты своего period_month, как у обычной книги.
ROWS = (
    (SEPT, 'свод ичи (Мухриддин)', 11, 205.00, '2025-09',
     'Қудратов Мухриддин'),
    (OCTOBER, 'свод ичи (Мухриддин)', 6, 93.60, '2025-09',
     'Қудратов Мухриддин', None),
    (SEPT, 'свод ичи (Фурқат)', 37, 398.90, '2025-09', 'Жумаев Фуркат'),
    (OCTOBER, 'свод ичи (Фурқат)', 9, 24.80, '2025-09', 'Жумаев Фуркат',
     None),
    # Датированные строки той же октябрьской книги -- они и есть улика.
    (OCTOBER, 'свод ичи (Беҳзод)', 20, 191.70, '2025-09', 'Имомов Беҳзод',
     '2025-10'),
    ('Гарден Агрокластер.xlsx', 'свод ичи (Нурали)', 57, 884.50, '2025-09',
     'Қодиров Нурали'),
    # [REASON]: ПЕРВАЯ ловушка. «свод ичи (Шохрух)» есть и в книге Гардена,
    # и в книге Когона -- это книги ДВУХ РАЗНЫХ ЛЮДЕЙ.
    ('Гарден_Агрокластер_Дрон_маълумот.xlsx', 'свод ичи (Шохрух)', 14, 154.70,
     '2025-09', 'Файзуллаев Шохрух'),
    ('Когон ПТЗ Дрон маълумот (2).xlsx', 'свод ичи (Шохрух)', 36, 412.60,
     '2025-09', 'Хамроев Шохрух'),
    # [REASON]: ВТОРАЯ ловушка. Апрель 2026: «свод ичи (Шахзод)» приходит из
    # книги Ғиждувона (Холмуродов) и из книги Сервиса, где оператор не
    # привязан вовсе. Молчание -- не согласие.
    ('Дрон Ғиждувон ПТЗ Апрель.xlsx', 'свод ичи (Шахзод)', 21, 361.30,
     '2026-04', 'Холмуродов Шахзод'),
    ('Сервис Дрон маълумот АПРЕЛЬ.xlsx', 'свод ичи (Шахзод)', 28, 492.15,
     '2026-04', None),
    # ОТРИЦАТЕЛЬНЫЙ КОНТРОЛЬ к третьей ловушке: настоящая двойная загрузка,
    # обе стороны датированы своим месяцем. Гвардия обязана её пропустить,
    # иначе она не гвардия, а глушилка.
    (TWICE, 'свод ичи (Туйғун)', 8, 254.90, '2025-10', 'Жураев Туйғун'),
    (TWICE2, 'свод ичи (Туйғун)', 8, 254.90, '2025-10', 'Жураев Туйғун'),
    # ВТОРОЙ отрицательный контроль: сторона недатирована, но датированных
    # строк этого файла нет НИГДЕ. «Книга другого месяца» не доказана --
    # и DELETE всё равно не предлагается.
    (BLIND, 'свод ичи (Хасан)', 12, 200.00, '2025-11', 'Ибодуллаев Хасанбой'),
    (BLIND2, 'свод ичи (Хасан)', 5, 60.00, '2025-11', 'Ибодуллаев Хасанбой',
     None),
)

CLEAN = tuple(row for row in ROWS if row[0] not in (
    OCTOBER, TWICE2, BLIND2) and row[1] not in (
    'свод ичи (Шохрух)', 'свод ичи (Шахзод)'))


def build_db(path, rows=ROWS, manual=0):
    conn = sqlite3.connect(path)
    conn.executescript(SCHEMA)
    ids = {}
    for idx, name in enumerate(OPERATORS, 1):
        conn.execute('INSERT INTO drone_operators (id, full_name) '
                     'VALUES (?, ?)', (idx, name))
        ids[name] = idx
    row_id = 0
    for row in rows:
        source_file, sheet, count, hectares, month = row[:5]
        operator = row[5] if len(row) > 5 else None
        date_month = row[6] if len(row) > 6 else month
        per = hectares / count
        for number in range(count):
            row_id += 1
            conn.execute(
                'INSERT INTO drone_works (id, period_month, work_date_from, '
                'customer_raw, area_ha, amount, drone_operator_id, '
                'source_file, source_sheet, source_row) '
                'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (row_id, month,
                 (date_month + '-10') if date_month else None,
                 'ФХ %d' % row_id, per, per * 200000, ids.get(operator),
                 source_file, sheet, number + 4))
    # [REASON]: строка, набранная руками, не имеет ни файла, ни листа.
    # Она законно «ниоткуда»; если такие складывать в одну группу, отчёт
    # объявит двойной загрузкой всю ручную правку месяца.
    for number in range(manual):
        row_id += 1
        conn.execute(
            'INSERT INTO drone_works (id, period_month, work_date_from, '
            'customer_raw, area_ha, amount, source_file, source_sheet, '
            "source_row) VALUES (?, '2025-09', '2025-09-10', ?, 5.0, "
            '1000000, NULL, NULL, NULL)', (row_id, 'Ручная %d' % number))
    # [REASON]: и ещё одна ловушка -- файл, у которого имя листа пустое.
    # Вместе с ручными строками он даёт группу «месяц + пустой лист» из ДВУХ
    # источников, и отчёт без гвардии объявил бы это двойной загрузкой.
    for number in range(manual and 3):
        row_id += 1
        conn.execute(
            'INSERT INTO drone_works (id, period_month, work_date_from, '
            'customer_raw, area_ha, amount, source_file, source_sheet, '
            "source_row) VALUES (?, '2025-09', '2025-09-10', ?, 4.0, "
            "800000, 'Без листа.xlsx', '', ?)",
            (row_id, 'Безлистовая %d' % number, number + 1))
    conn.commit()
    conn.close()


def scan(db, month=None):
    conn = tool.connect_ro(db)
    try:
        rows = tool.load(conn, month)
        dated = tool.dated_months_by_file(conn)
    finally:
        conn.close()
    return rows, tool.find_duplicates(rows, dated)


class ImportDuplicatesTest(unittest.TestCase):

    def setUp(self):
        self.dir = tempfile.mkdtemp(prefix='impdup_')
        self.db = os.path.join(self.dir, 'transport.db')
        build_db(self.db)

    # ГЛАВНОЕ: лист-тёзка -- НЕ дубль, и удалять по нему нечего.
    def test_a_namesake_sheet_is_not_a_duplicate(self):
        """Два разных человека под одним именем листа.

        [REASON]: без различения по оператору отчёт предложил бы удалить
        154.70 га книги Файзуллаева Шохруха как «лишнюю загрузку» книги
        Хамроева Шохруха. Это и случилось на production 2026-08-20.
        """
        _rows, duplicates = scan(self.db)
        key = ('2025-09', 'свод ичи (Шохрух)')
        self.assertIn(key, duplicates)
        self.assertEqual(tool.NAMESAKE, duplicates[key]['kind'])
        self.assertEqual(
            {'Файзуллаев Шохрух', 'Хамроев Шохрух'},
            {f[4] for f in duplicates[key]['files']})
        # Ни в кандидатах, ни в SQL его быть не должно.
        self.assertNotIn(key, {(i[0], i[1]) for i in
                               tool.candidates(duplicates)})
        self.assertEqual([key[0]], [t[0] for t in tool.namesakes(duplicates)])

    # ВТОРАЯ ловушка: сторона без оператора -- не согласие.
    def test_a_side_without_an_operator_is_undecidable(self):
        """[REASON]: молчание не есть согласие.

        Первая редакция считала «нет оператора» за «оператор тот же» и на
        апреле 2026 предложила удалить 492.15 га как дубль. Такие группы
        обязаны уходить в UNKNOWN, и SQL по ним не печатается.
        """
        _rows, duplicates = scan(self.db)
        key = ('2026-04', 'свод ичи (Шахзод)')
        self.assertEqual(tool.UNKNOWN, duplicates[key]['kind'])
        self.assertEqual('operator', duplicates[key]['reason'])
        self.assertNotIn(key, {(i[0], i[1]) for i in
                               tool.candidates(duplicates)})

    # ТРЕТЬЯ ловушка -- та, что стоила 118.40 га живых данных.
    def test_another_months_book_is_not_a_double_upload(self):
        """Недатированные строки октябрьской книги, лежащие в сентябре.

        [REASON]: это ровно та группа, по которой 2026-08-20 на production
        выполнили DELETE. Обе стороны одного человека, обе в сентябре -- и
        отчёт называл их двойной загрузкой. Улика в том, что у стороны
        OCTOBER здесь нет НИ ОДНОЙ датированной строки, а датированные
        строки того же файла лежат в 2025-10.
        """
        _rows, duplicates = scan(self.db)
        for sheet, hectares in (('свод ичи (Мухриддин)', 93.60),
                                ('свод ичи (Фурқат)', 24.80)):
            key = ('2025-09', sheet)
            self.assertEqual(tool.OTHER_MONTH, duplicates[key]['kind'], sheet)
            self.assertEqual([(OCTOBER, ['2025-10'])],
                             duplicates[key]['strangers'], sheet)
            self.assertIn(hectares,
                          [round(f[2], 2) for f in duplicates[key]['files']])
            # Ни кандидатом, ни тёзкой -- и никакого SQL.
            self.assertNotIn(key, {(i[0], i[1]) for i in
                                   tool.candidates(duplicates)})
        self.assertEqual([('2025-09', 'свод ичи (Мухриддин)'),
                          ('2025-09', 'свод ичи (Фурқат)')],
                         [(m, sh) for m, sh, _f, _s
                          in tool.other_months(duplicates)])

    def test_another_months_book_gets_no_delete_at_all(self):
        """Даже закомментированного DELETE быть не должно.

        [REASON]: в прошлый раз выполнили именно закомментированный DELETE --
        его для того и раскомментировали. Строка, которой нет, не может быть
        раскомментирована.
        """
        self._run_main()
        sql = self._sql_text()
        self.assertIn('КНИГА ДРУГОГО МЕСЯЦА, НЕ ДУБЛЬ', sql)
        self.assertIn('свод ичи (Мухриддин)', sql)
        self.assertIn('датированные строки того же файла лежат в 2025-10',
                      sql)
        for line in sql.splitlines():
            if 'DELETE' in line:
                self.assertNotIn('Мухриддин', line)
                self.assertNotIn('Фурқат', line)

    def test_the_month_filter_still_sees_the_other_month(self):
        """--month 2025-09 обязан находить улику ЗА пределами сентября.

        [REASON]: улика -- октябрьские строки октябрьского файла, а фильтр
        по месяцу их отрезает. Если бы карта датированных месяцев считалась
        по той же выборке, прогон «--month 2025-09» снова предложил бы
        DELETE -- а именно так отчёт и запускают, помесячно.
        """
        _rows, september = scan(self.db, '2025-09')
        key = ('2025-09', 'свод ичи (Фурқат)')
        self.assertEqual(tool.OTHER_MONTH, september[key]['kind'])
        self.assertEqual([], tool.candidates(september))

    # ОТРИЦАТЕЛЬНЫЙ КОНТРОЛЬ: гвардия обязана различать, а не глушить.
    def test_a_real_double_upload_is_still_a_candidate(self):
        """Обе стороны датированы своим месяцем -- это настоящий дубль."""
        _rows, duplicates = scan(self.db)
        key = ('2025-10', 'свод ичи (Туйғун)')
        self.assertEqual(tool.CANDIDATE, duplicates[key]['kind'])
        self.assertEqual({TWICE, TWICE2},
                         {f[0] for f in duplicates[key]['files']})
        self.assertIn(key, {(i[0], i[1]) for i in tool.candidates(duplicates)})

    # ВТОРОЙ отрицательный контроль: половина признака -- не признак.
    def test_an_undated_side_without_a_trace_elsewhere_is_undecidable(self):
        """Недатированная сторона без следа в другом месяце -- UNKNOWN.

        [REASON]: доказать «книга другого месяца» нечем, но и предлагать
        DELETE нельзя: сторона стоит в этом месяце по period_month формы, а
        не по собственной дате. Ровно то же молчание, что и у стороны без
        оператора.
        """
        _rows, duplicates = scan(self.db)
        key = ('2025-11', 'свод ичи (Хасан)')
        self.assertEqual(tool.UNKNOWN, duplicates[key]['kind'])
        self.assertEqual('undated', duplicates[key]['reason'])
        self.assertNotIn(key, {(i[0], i[1]) for i in
                               tool.candidates(duplicates)})
        self.assertIn(key, {(m, sh) for m, sh, _f in tool.unknowns(
            duplicates)})

    def test_the_undecidable_groups_get_no_sql_at_all(self):
        self._run_main()
        sql = self._sql_text()
        self.assertIn('НЕ ОПОЗНАНО', sql)
        self.assertIn('нет НИ ОДНОЙ', sql)
        for sheet in ('свод ичи (Шахзод)', 'свод ичи (Хасан)'):
            self.assertIn(sheet, sql)
            for line in sql.splitlines():
                if sheet.split('(')[1] in line:
                    self.assertTrue(line.strip().startswith('--'), line)
        self.assertNotIn('2026-04', ''.join(
            line for line in sql.splitlines()
            if line.startswith('SELECT') or line.startswith('-- DELETE')))

    def test_both_files_are_offered_never_one_verdict(self):
        """Отчёт НЕ решает, какая загрузка лишняя -- это говорит свод.

        [REASON]: первая редакция считала полной загрузку с наибольшим
        числом строк. На апреле 2026 это назвало бы лишним файл с 361.30 га
        в пользу файла с 492.15, а на мае -- файл с 548.10 га в пользу файла
        с 226.51. Число строк такого не решает.
        """
        _rows, duplicates = scan(self.db)
        items = tool.candidates(duplicates)
        self.assertEqual(2, len(items))          # один лист x два файла
        self.assertEqual({TWICE, TWICE2}, {i[2] for i in items})

    def test_a_clean_month_is_silent(self):
        """Отрицательный контроль: отчёт, кричащий всегда, бесполезен."""
        os.remove(self.db)
        build_db(self.db, rows=CLEAN)
        _rows, duplicates = scan(self.db)
        self.assertEqual({}, duplicates)
        self.assertEqual([], tool.candidates(duplicates))

    def test_hand_typed_rows_are_not_a_double_upload(self):
        """[REASON]: у ручной строки нет ни файла, ни листа -- их много, и

        сложенные в одну группу они выглядели бы как загрузка из пустого
        файла. Тогда отчёт кричал бы на каждом месяце, где кто-то правил
        руками, и на него перестали бы смотреть.
        """
        os.remove(self.db)
        build_db(self.db, rows=CLEAN, manual=5)
        _rows, duplicates = scan(self.db)
        self.assertEqual({}, duplicates)

    def test_the_same_sheet_in_another_month_is_not_a_duplicate(self):
        os.remove(self.db)
        build_db(self.db, rows=(
            (SEPT, 'свод ичи (Мухриддин)', 11, 205.00, '2025-09',
             'Қудратов Мухриддин'),
            ('Сервис Октябрь.xlsx', 'свод ичи (Мухриддин)', 9, 180.0,
             '2025-10', 'Қудратов Мухриддин')))
        _rows, duplicates = scan(self.db)
        self.assertEqual({}, duplicates)

    def test_month_filter_narrows_the_scan(self):
        _rows, all_months = scan(self.db)
        self.assertEqual(6, len(all_months))
        _rows, september = scan(self.db, '2025-09')
        self.assertEqual(3, len(september))
        _rows, october = scan(self.db, '2025-10')
        self.assertEqual(1, len(october))
        _rows, april = scan(self.db, '2026-04')
        self.assertEqual(1, len(april))

    def test_the_printed_sql_targets_one_file_one_sheet_one_month(self):
        _rows, duplicates = scan(self.db)
        pairs = tool.sql_lines(tool.candidates(duplicates))
        self.assertEqual(2, len(pairs))
        for select, delete in pairs:
            self.assertTrue(select.startswith('SELECT'))
            self.assertTrue(delete.startswith('DELETE'))
            self.assertIn("period_month) = '2025-10'", delete)
            self.assertIn('source_sheet', delete)
            self.assertIn('source_file', delete)
            self.assertNotIn('Шохрух', delete)
            self.assertNotIn('Мухриддин', delete)

    def test_the_chosen_delete_removes_exactly_that_upload(self):
        """Печатаемый SQL проверяется ИСПОЛНЕНИЕМ, а не чтением глазами."""
        _rows, duplicates = scan(self.db)
        second = [i for i in tool.candidates(duplicates) if i[2] == TWICE2]
        pairs = tool.sql_lines(second)
        conn = sqlite3.connect(self.db)
        try:
            before = conn.execute('SELECT COUNT(*), ROUND(SUM(area_ha), 2) '
                                  'FROM drone_works').fetchone()
            for _select, delete in pairs:
                conn.execute(delete)
            conn.commit()
            after = conn.execute('SELECT COUNT(*), ROUND(SUM(area_ha), 2) '
                                 'FROM drone_works').fetchone()
            left = {r[0] for r in conn.execute(
                'SELECT DISTINCT source_file FROM drone_works')}
        finally:
            conn.close()
        self.assertEqual(before[0] - 8, after[0])
        self.assertAlmostEqual(before[1] - 254.90, after[1], 2)
        self.assertNotIn(TWICE2, left)
        self.assertIn(TWICE, left)
        # Ни книга Файзуллаева Шохруха, ни октябрьская книга Сервиса не
        # тронуты: по ним DELETE не печатался вовсе.
        self.assertIn('Гарден_Агрокластер_Дрон_маълумот.xlsx', left)
        self.assertIn(OCTOBER, left)

    def test_the_database_is_opened_read_only(self):
        conn = tool.connect_ro(self.db)
        try:
            with self.assertRaises(sqlite3.OperationalError):
                conn.execute('DELETE FROM drone_works')
        finally:
            conn.close()

    def test_the_tool_itself_never_writes(self):
        """[REASON]: read-only соединение защищает от ошибки, а не от замысла.

        Печатаемый DELETE -- это ТЕКСТ для владельца, он собирается в
        sql_lines и никогда не исполняется.
        """
        path = os.path.join(REPO_ROOT, 'tools', 'drone_import_duplicates.py')
        with open(path, encoding='utf-8') as handle:
            tree = ast.parse(handle.read())
        allowed = {'sql_lines'}
        banned = ('insert into', 'update ', 'delete from', 'drop ', 'alter ')
        for node in ast.walk(tree):
            if not isinstance(node, ast.FunctionDef) or node.name in allowed:
                continue
            for child in ast.walk(node):
                if isinstance(child, ast.Constant) and \
                        isinstance(child.value, str):
                    lowered = child.value.lower()
                    for word in banned:
                        self.assertNotIn(
                            word, lowered,
                            'пишущее слово %r в %s: %r'
                            % (word, node.name, child.value[:60]))

    def test_the_verdict_table_covers_every_kind(self):
        """[REASON]: вердикт жил в двух местах и в xlsx был двоичным -- всё,

        что не CANDIDATE, подписывалось «ТЁЗКИ, РАЗНЫЕ ЛЮДИ». UNKNOWN уже
        подписывался неверно; OTHER_MONTH подписался бы так же. Тот же класс
        дефекта, что и разъехавшийся порядок столбцов в drone_money_audit.
        """
        _rows, duplicates = scan(self.db)
        seen = set()
        for group in duplicates.values():
            for table in (tool.VERDICT_CONSOLE, tool.VERDICT_XLSX):
                label = tool.verdict(group, table)
                self.assertTrue(label)
                seen.add((group['kind'], group.get('reason')))
        self.assertEqual({(tool.CANDIDATE, None), (tool.NAMESAKE, None),
                          (tool.UNKNOWN, 'operator'),
                          (tool.UNKNOWN, 'undated'),
                          (tool.OTHER_MONTH, None)}, seen)
        self.assertNotEqual(tool.VERDICT_XLSX[tool.NAMESAKE],
                            tool.VERDICT_XLSX[tool.OTHER_MONTH])
        self.assertNotEqual(tool.VERDICT_XLSX[tool.NAMESAKE],
                            tool.VERDICT_XLSX[(tool.UNKNOWN, 'operator')])

    def test_find_duplicates_demands_the_evidence(self):
        """[REASON]: аргумент со значением по умолчанию молча вернул бы

        поведение, стоившее 118.40 га: забытая карта не отличалась бы от
        пустой базы, и группа снова стала бы кандидатом на удаление.
        """
        rows, _duplicates = scan(self.db)
        with self.assertRaises(TypeError):
            tool.find_duplicates(rows)

    def _sql_text(self):
        sql_path = os.path.join(self.dir, 'drone_import_duplicates.sql')
        with open(sql_path, encoding='utf-8') as handle:
            return handle.read()

    def _run_main(self):
        import contextlib
        import io as _io
        buffer = _io.StringIO()
        argv = sys.argv
        sys.argv = ['drone_import_duplicates.py', '--db', self.db]
        try:
            with contextlib.redirect_stdout(buffer):
                code = tool.main()
        finally:
            sys.argv = argv
        return code, buffer.getvalue()

    def test_exit_code_is_one_when_a_duplicate_is_found(self):
        code, printed = self._run_main()
        printed.encode('ascii')          # консоль -- только ASCII
        self.assertEqual(1, code)
        self.assertIn('NOTHING WAS DELETED', printed)
        self.assertIn('DOES NOT DECIDE', printed)
        self.assertIn('Namesake sheets     : 1', printed)
        self.assertIn('Undecidable         : 2', printed)
        self.assertIn("Another month's book: 2", printed)
        self.assertIn('no dated row here, dated rows in 2025-10', printed)

    def test_the_sql_file_marks_the_namesake_and_comments_every_delete(self):
        """Файл обязан и предупредить про тёзку, и не дать удалить сгоряча.

        [REASON]: DELETE выписывается ЗАКОММЕНТИРОВАННЫМ. Файл, который
        можно выполнить целиком одним махом, снёс бы обе стороны каждого
        кандидата -- то есть и верную загрузку тоже.
        """
        self._run_main()
        sql = self._sql_text()
        self.assertIn('ТЁЗКИ, НЕ ДУБЛЬ. Ничего не удалять.', sql)
        self.assertIn('свод ичи (Шохрух)', sql)
        # Ни одного DELETE по листу-тёзке.
        for line in sql.splitlines():
            if 'DELETE' in line:
                self.assertTrue(line.strip().startswith('--'), line)
                self.assertNotIn('Шохрух', line)
        self.assertEqual(2, sql.count('SELECT id,'))
        self.assertEqual(2, sql.count('-- DELETE FROM'))

    def test_the_header_does_not_promise_sql_that_is_not_there(self):
        """[REASON]: шапка обещала «ниже ДВА варианта, выполнить РОВНО ОДИН»

        всегда -- даже когда кандидатов нет и в файле одни предупреждения.
        Файл, зовущий выполнить то, чего в нём нет, толкает искать DELETE и
        раскомментировать первое похожее.
        """
        os.remove(self.db)
        build_db(self.db, rows=tuple(
            row for row in ROWS if row[0] not in (TWICE, TWICE2)))
        code, _printed = self._run_main()
        self.assertEqual(1, code)
        sql = self._sql_text()
        self.assertIn('ВЫПОЛНЯТЬ НЕЧЕГО', sql)
        self.assertNotIn('DELETE FROM', sql)
        self.assertNotIn('SELECT id,', sql)
        self.assertNotIn('выполнить надо РОВНО ОДИН', sql)

    def test_the_console_never_leaks_cyrillic(self):
        _code, printed = self._run_main()
        printed.encode('ascii')
        self.assertNotIn('DELETE FROM', printed)

    def test_exit_code_is_zero_on_a_clean_base(self):
        os.remove(self.db)
        build_db(self.db, rows=CLEAN)
        code, printed = self._run_main()
        self.assertEqual(0, code)
        self.assertIn('No book looks imported twice', printed)

    def test_missing_database_gives_code_two(self):
        argv = sys.argv
        sys.argv = ['drone_import_duplicates.py', '--db',
                    os.path.join(self.dir, 'nope.db')]
        try:
            self.assertEqual(2, tool.main())
        finally:
            sys.argv = argv

    def test_the_workbook_names_every_verdict_and_the_undated_count(self):
        out = os.path.join(self.dir, 'dup.xlsx')
        argv = sys.argv
        sys.argv = ['drone_import_duplicates.py', '--db', self.db,
                    '--out', out]
        try:
            import contextlib
            import io as _io
            with contextlib.redirect_stdout(_io.StringIO()):
                tool.main()
        finally:
            sys.argv = argv
        import openpyxl
        book = openpyxl.load_workbook(out)
        verdicts = {row[7] for row in
                    book['Двойные загрузки'].iter_rows(min_row=2,
                                                       values_only=True)}
        self.assertEqual({tool.VERDICT_XLSX[tool.CANDIDATE],
                          tool.VERDICT_XLSX[tool.NAMESAKE],
                          tool.VERDICT_XLSX[(tool.UNKNOWN, 'operator')],
                          tool.VERDICT_XLSX[(tool.UNKNOWN, 'undated')],
                          tool.VERDICT_XLSX[tool.OTHER_MONTH]}, verdicts)
        sources = {(row[2], row[7]) for row in
                   book['Все источники'].iter_rows(min_row=2,
                                                   values_only=True)}
        self.assertIn((OCTOBER, 6), sources)     # 6 строк без даты
        self.assertIn((SEPT, 0), sources)        # сентябрьская книга -- все с


if __name__ == '__main__':
    unittest.main()
