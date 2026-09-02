# -*- coding: utf-8 -*-
"""DRONE-WORKS-001: the ledger, the directory, the reports and the workbooks.

The load-bearing assertions here are the reconciliations. Every cut of the
report is built by grouping, and a group-by silently drops NULL -- which is
exactly the rows nobody has resolved yet. That is how the previous drone
system lost 2 036 flights and 2 082 hectares into an invisible bucket. So each
cut is checked to sum back to the same grand total, and each workbook is
re-read with openpyxl and compared cell by cell against the numbers the screen
computed. A screenshot proves nothing and neither does the route returning 200.

The fixture is deliberately small and hand-checkable: 10 jobs, 128.50 ha,
24 555 256 so'm, of which 12 500 000 received. One unresolved customer, one
customer missing entirely, one unresolved operator, one job with no date, one
job with no payment type, and one job whose own date falls outside the month
its book is filed under.

Run:
  python -m unittest tests.test_drone_works_screens -v
"""
import datetime
import io
import re
import unittest

from tests.harness import app, reset_db, create_admin, login, CSRF
from models import (db, DRONE_RECEIVED_KIND_OPERATOR_DUE,
                    DRONE_RECEIVED_KIND_RECEIVED, DroneCustomer,
                    DroneCustomerAlias, DroneFlight, DroneOperator,
                    DroneOperatorAssignment, DroneUnit, DroneWork,
                    Organization, User, ROLE_OPERATOR)


def set_language(user_id, lang):
    """The module is bilingual and the sheet titles follow the user.

    [REASON]: the workbook sheet names ARE the language -- «По заказчикам» in
    Russian and «Буюртмачилар бўйича» in Uzbek. A test that did not pin the
    language would pass or fail depending on the default of the day.
    """
    with app.app_context():
        user = db.session.get(User, user_id)
        user.language = lang
        db.session.commit()

# (period, date_from, date_to, operator, customer, area, price, amount,
#  received, payment, subdivision)
FIXTURE = [
    ('2026-04', datetime.date(2026, 4, 5), 'op1', 'c1', 12.5, 200000,
     2500000, 2500000, 'cash', 'Гарден'),
    ('2026-04', datetime.date(2026, 4, 6), 'op1', 'c1', 10.0, 200000,
     2000000, 0, 'cash', 'Гарден'),
    ('2026-04', datetime.date(2026, 4, 7), 'op2', 'c2', 20.0, 150000,
     3000000, 1000000, 'transfer', 'Когон'),
    ('2026-04', datetime.date(2026, 4, 8), 'op2', None, 5.0, None,
     700000, 0, 'transfer', 'Когон'),
    ('2026-03', datetime.date(2026, 3, 20), None, 'c2', 30.0, 300000,
     9000000, 9000000, 'transfer', 'Когон'),
    ('2026-03', datetime.date(2026, 3, 12), 'op1', 'c3', 8.0, 85633,
     685064, 0, 'internal', None),
    # A book filed as 2026-03 whose own date is in April: the dated row keeps
    # its own date and must land in April, never in March.
    ('2026-03', datetime.date(2026, 4, 25), 'op1', 'c3', 9.0, 200000,
     1800000, 0, 'cash', 'Пешку'),
    # No date at all: it has to stay in the report, in its own service row.
    ('2026-03', None, 'op2', 'c3', 6.0, 200000, 2015000, 0, 'cash', 'Пешку'),
    # DRONE-WORKS-IMPORT-FIX-001. The sheet said nothing about how this job
    # was paid -- roughly two fifths of the real corpus looks like this.
    ('2026-04', datetime.date(2026, 4, 9), 'op1', 'c1', 4.0, 200000,
     800000, 0, 'unknown', 'Гарден'),
    # The farm name was missing in the source: how the holding's own land is
    # written down. It must NOT share a service row with «не определён».
    ('2026-05', datetime.date(2026, 5, 4), 'op2', None, 24.0, 85633,
     2055192, 0, 'internal', 'Гарден'),
]
# The row with no customer at all is index 10 of the fixture; seed() gives it
# customer_raw = '' rather than a name.
EMPTY_CUSTOMER_INDEX = 10
# Fixture row 3 (20.0 ha, 3 000 000 amount, 1 000 000 «received») is the one
# whose figure came from «Оператор топшириши керак».
OPERATOR_DUE_INDEX = 3

TOTAL_JOBS = 10
TOTAL_AREA = 128.5
TOTAL_AMOUNT = 24555256.0
TOTAL_RECEIVED = 12500000.0
TOTAL_OUTSTANDING = TOTAL_AMOUNT - TOTAL_RECEIVED


def seed():
    """Build the fixture. Returns {'op1': id, ..., 'c1': id, ...}."""
    reset_db()
    ids = {}
    with app.app_context():
        for key, name, subdivision in (
                ('op1', 'Файзуллаев Фурқат', 'Гарден'),
                ('op2', 'Хамроев Шохрух', 'Когон')):
            row = DroneOperator(full_name=name, subdivision_name=subdivision)
            db.session.add(row)
            db.session.flush()
            ids[key] = row.id
        for key, name in (('c1', 'Миробод АМТ'), ('c2', 'Дўстлик ФХ'),
                          ('c3', 'Пешку АМТ')):
            row = DroneCustomer(name=name)
            db.session.add(row)
            db.session.flush()
            ids[key] = row.id
            db.session.add(DroneCustomerAlias(
                raw_name=name, normalized_key=name.casefold(),
                drone_customer_id=row.id, is_active=True))
        for index, (period, date_from, operator, customer, area, price,
                    amount, received, payment, subdivision) in enumerate(
                        FIXTURE, 1):
            db.session.add(DroneWork(
                period_month=period,
                work_date_from=date_from,
                work_date_to=date_from,
                drone_operator_id=ids.get(operator),
                operator_raw=operator or 'Ким',
                drone_customer_id=ids.get(customer),
                customer_raw=('' if index == EMPTY_CUSTOMER_INDEX
                              else 'Хозяйство %d' % index),
                area_ha=area, price_per_ha=price, amount=amount,
                received_amount=received,
                # [REASON]: one row carries «Оператор топшириши керак» -- money
                # the operator still OWES, sitting in received_amount. Without
                # one the debts page's memo never renders and the test that
                # checks it would pass on an empty string. On the 2026-08-05
                # staging import that is 164 rows out of 852.
                received_kind=(DRONE_RECEIVED_KIND_OPERATOR_DUE
                               if index == OPERATOR_DUE_INDEX
                               else (DRONE_RECEIVED_KIND_RECEIVED
                                     if received else None)),
                payment_type=payment,
                subdivision_name=subdivision,
                source_file='book%d.xlsx' % index, source_sheet='свод ичи',
                source_row=index, import_batch='fixture'))
        db.session.commit()
    return ids


def report_data(query='', lang='ru'):
    """The report structure the screen and the workbook are both built from.

    [REASON]: g.lang is set explicitly. Every service row label -- «Заказчик
    не определён», «Дата не указана» -- is language-dependent, and comparing a
    workbook produced for a Russian user against a structure computed with the
    module's Uzbek default would fail on the labels while the numbers agreed.
    """
    from flask import g
    import drones
    with app.test_request_context('/drones/works/reports?' + query):
        g.lang = lang
        filters = drones._drone_works_filters_from_args(
            _request_args(query))
        return drones._drone_works_report_data(
            drones._drone_work_conditions(filters)), filters


def _request_args(query):
    from werkzeug.datastructures import MultiDict
    from urllib.parse import parse_qsl
    return MultiDict(parse_qsl(query))


def sheet_rows(data, title):
    """{first cell: [rest of the row]} for one worksheet of a workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb[title]
    rows = list(ws.iter_rows(values_only=True))
    return rows


class DroneWorksScreenTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.ids = seed()
        cls.admin = create_admin('works_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    # ── The ledger ────────────────────────────────────────────────────────
    def test_the_ledger_opens_and_totals_the_whole_filtered_set(self):
        response = self.client.get('/drones/works')
        self.assertEqual(response.status_code, 200)
        body = response.data.decode('utf-8')
        self.assertIn('128.50', body)
        self.assertIn('Дата не указана', body)
        self.assertIn('Тип оплаты не указан', body)

    def test_the_period_filter_uses_the_jobs_own_date(self):
        """The April row of a March-labelled book belongs to April."""
        with app.app_context():
            import drones
            with app.test_request_context('/?period=2026-04'):
                filters = drones._drone_works_filters_from_args(
                    _request_args('period=2026-04'))
                conds = drones._drone_work_conditions(filters)
                april = drones._drone_works_totals(conds)
        self.assertEqual(april['jobs'], 6)
        self.assertAlmostEqual(april['area'], 60.5, places=2)

    def test_an_undated_job_falls_into_its_manifest_month(self):
        with app.app_context():
            import drones
            with app.test_request_context('/?period=2026-03'):
                filters = drones._drone_works_filters_from_args(
                    _request_args('period=2026-03'))
                march = drones._drone_works_totals(
                    drones._drone_work_conditions(filters))
        self.assertEqual(march['jobs'], 3)
        self.assertAlmostEqual(march['area'], 44.0, places=2)

    def test_the_two_month_buckets_add_up_to_everything(self):
        """Neither the undated job nor the out-of-month one is lost."""
        with app.app_context():
            import drones
            totals = drones._drone_works_totals([])
        self.assertEqual(totals['jobs'], TOTAL_JOBS)
        self.assertAlmostEqual(totals['area'], TOTAL_AREA, places=2)
        self.assertAlmostEqual(totals['amount'], TOTAL_AMOUNT, places=2)
        self.assertAlmostEqual(totals['received'], TOTAL_RECEIVED, places=2)

    def _customer_filter_totals(self, query):
        with app.app_context():
            import drones
            with app.test_request_context('/?' + query):
                filters = drones._drone_works_filters_from_args(
                    _request_args(query))
                return drones._drone_works_totals(
                    drones._drone_work_conditions(filters))

    def test_the_unresolved_filters_find_exactly_the_unresolved_rows(self):
        """DRONE-UI-CUSTOMER-LABEL-001 NARROWED customer_id=-1.

        It used to catch both kinds of NULL customer and this assertion read
        2 jobs / 29.0 ha. The two kinds are different facts -- «написание есть,
        алиаса нет» needs an alias created, «имени в ведомости не было» needs
        nothing -- so -1 now means only the first and -2 the second. The row
        counts below are the SAME rows, split; the union assertion at the end
        is what proves nothing fell out between the two options.
        """
        self.assertEqual(
            self.client.get('/drones/works?customer_id=-1').status_code, 200)
        self.assertEqual(
            self.client.get('/drones/works?customer_id=-2').status_code, 200)

        # Fixture row 4: no customer id, but «Хозяйство 4» IS spelled out.
        unresolved = self._customer_filter_totals('customer_id=-1')
        self.assertEqual(unresolved['jobs'], 1)
        self.assertAlmostEqual(unresolved['area'], 5.0, places=2)

        # Fixture row 10: customer_raw is '' -- the holding's own land.
        unstated = self._customer_filter_totals('customer_id=-2')
        self.assertEqual(unstated['jobs'], 1)
        self.assertAlmostEqual(unstated['area'], 24.0, places=2)

        # The pair still covers exactly what the single option used to.
        self.assertEqual(unresolved['jobs'] + unstated['jobs'], 2)
        self.assertAlmostEqual(unresolved['area'] + unstated['area'], 29.0,
                               places=2)

    def test_a_manual_job_needs_a_period_when_it_has_no_date(self):
        before = self._count()
        response = self.client.post('/drones/works/add', data={
            'csrf_token': CSRF, 'customer_raw': 'Без даты и периода',
            'area_ha': '3', 'payment_type': 'cash'})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self._count(), before)

    def test_a_manual_job_derives_its_period_from_the_date(self):
        response = self.client.post('/drones/works/add', data={
            'csrf_token': CSRF, 'customer_raw': 'С датой',
            'work_date_from': '2026-05-04', 'area_ha': '3,5',
            'payment_type': 'cash'})
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            row = DroneWork.query.filter_by(customer_raw='С датой').one()
            self.assertEqual(row.period_month, '2026-05')
            self.assertAlmostEqual(row.area_ha, 3.5, places=2)
            self.assertIsNone(row.price_per_ha)
            db.session.delete(row)
            db.session.commit()

    def test_a_blank_money_field_is_null_and_never_zero(self):
        self.client.post('/drones/works/add', data={
            'csrf_token': CSRF, 'customer_raw': 'Пустые деньги',
            'period_month': '2026-06', 'area_ha': '1',
            'payment_type': 'cash', 'price_per_ha': '', 'amount': '  '})
        with app.app_context():
            row = DroneWork.query.filter_by(customer_raw='Пустые деньги').one()
            self.assertIsNone(row.price_per_ha)
            self.assertIsNone(row.amount)
            db.session.delete(row)
            db.session.commit()

    def _count(self):
        with app.app_context():
            return DroneWork.query.count()

    # ── Reconciliation: the point of the whole report ─────────────────────
    def test_every_cut_reconciles_to_the_same_grand_total(self):
        data, _filters = report_data()
        for name in ('by_customer', 'by_operator', 'by_subdivision',
                     'by_month', 'by_payment'):
            cut = data[name]
            self.assertTrue(cut['reconciled'], name)
            self.assertEqual(cut['total']['jobs'], TOTAL_JOBS, name)
            self.assertAlmostEqual(cut['total']['area'], TOTAL_AREA,
                                   places=2, msg=name)
            self.assertAlmostEqual(cut['total']['amount'], TOTAL_AMOUNT,
                                   places=2, msg=name)
            self.assertAlmostEqual(cut['total']['received'], TOTAL_RECEIVED,
                                   places=2, msg=name)

    def test_the_reconciliation_check_can_actually_fail(self):
        """A control that never fails is not a control.

        [REASON]: `reconciled` is computed from the same query as the total,
        so it would read True on an empty table and on a broken one alike.
        Feeding it a deliberately wrong grand total proves it discriminates.
        """
        import drones
        with app.test_request_context('/'):
            wrong = drones._drone_works_totals([])
            wrong['area'] += 1.0
            cut = drones._drone_work_cut(
                [], DroneWork.drone_customer_id, wrong, ((None, 'x'),))
        self.assertFalse(cut['reconciled'])

    def test_the_unresolved_rows_are_visible_and_counted_in_the_total(self):
        data, _filters = report_data()
        customer = data['by_customer']
        labels = [s['label'] for s in customer['services']]
        self.assertEqual(labels,
                         ['Заказчик не указан', 'Заказчик не определён'])
        self.assertEqual(customer['total']['jobs'],
                         sum(r['jobs'] for r in customer['rows'])
                         + sum(s['jobs'] for s in customer['services']))

        operator = data['by_operator']
        self.assertEqual(len(operator['services']), 1)
        self.assertEqual(operator['services'][0]['jobs'], 1)
        self.assertAlmostEqual(operator['services'][0]['area'], 30.0,
                               places=2)

    def test_the_two_customer_service_rows_are_different_facts(self):
        """«не указан» is missing data; «не определён» is unmatched data."""
        data, _filters = report_data()
        services = {s['label']: s for s in data['by_customer']['services']}
        unstated = services['Заказчик не указан']
        self.assertEqual(unstated['jobs'], 1)
        self.assertAlmostEqual(unstated['area'], 24.0, places=2)
        unresolved = services['Заказчик не определён']
        self.assertEqual(unresolved['jobs'], 1)
        self.assertAlmostEqual(unresolved['area'], 5.0, places=2)

    def test_the_payment_cut_shows_the_unknown_bucket(self):
        data, _filters = report_data()
        cut = data['by_payment']
        self.assertTrue(cut['reconciled'])
        self.assertEqual([s['label'] for s in cut['services']],
                         ['Тип оплаты не указан'])
        self.assertEqual(cut['services'][0]['jobs'], 1)
        self.assertAlmostEqual(cut['services'][0]['area'], 4.0, places=2)
        self.assertEqual(cut['total']['jobs'], TOTAL_JOBS)

    def test_the_month_cut_uses_the_jobs_own_date_and_shows_the_undated_row(self):
        data, _filters = report_data()
        months = {r['key']: r for r in data['by_month']['rows']}
        self.assertEqual(set(months), {'2026-03', '2026-04', '2026-05'})
        # 12.5 + 10 + 20 + 5 + 9 + 4 -- including the April row of the March
        # book
        self.assertAlmostEqual(months['2026-04']['area'], 60.5, places=2)
        self.assertAlmostEqual(months['2026-03']['area'], 38.0, places=2)
        self.assertAlmostEqual(months['2026-05']['area'], 24.0, places=2)
        self.assertEqual(len(data['by_month']['services']), 1)
        service = data['by_month']['services'][0]
        self.assertEqual(service['jobs'], 1)
        self.assertAlmostEqual(service['area'], 6.0, places=2)
        # ... and it says which book's month the undated job belongs to
        self.assertEqual(service['periods'],
                         [{'period': '2026-03', 'jobs': 1, 'area': 6.0}])

    def test_the_subdivision_cut_shows_the_row_with_no_subdivision(self):
        data, _filters = report_data()
        self.assertEqual(len(data['by_subdivision']['services']), 1)
        self.assertAlmostEqual(
            data['by_subdivision']['services'][0]['area'], 8.0, places=2)

    def test_the_debt_view_sums_back_to_the_grand_total(self):
        data, _filters = report_data()
        for name in ('debt_by_customer', 'debt_by_operator'):
            debt = data[name]
            rows = list(debt['rows']) + ([debt['rest']] if debt['rest']
                                         else [])
            self.assertAlmostEqual(
                sum(r['outstanding'] for r in rows), TOTAL_OUTSTANDING,
                places=2, msg=name)
            self.assertEqual(sum(r['jobs'] for r in rows), TOTAL_JOBS, name)
            self.assertAlmostEqual(debt['total']['outstanding'],
                                   TOTAL_OUTSTANDING, places=2, msg=name)
        # every listed row really does owe something
        for row in data['debt_by_customer']['rows']:
            self.assertGreater(row['outstanding'], 0)

    def test_a_filter_narrows_every_cut_consistently(self):
        data, _filters = report_data('payment=transfer')
        self.assertEqual(data['totals']['jobs'], 3)
        self.assertAlmostEqual(data['totals']['area'], 55.0, places=2)
        for name in ('by_customer', 'by_operator', 'by_subdivision',
                     'by_month', 'by_payment'):
            self.assertTrue(data[name]['reconciled'], name)
            self.assertEqual(data[name]['total']['jobs'], 3, name)

    def test_the_reports_screen_opens(self):
        response = self.client.get('/drones/works/reports')
        self.assertEqual(response.status_code, 200)
        body = response.data.decode('utf-8')
        self.assertIn('Заказчик не определён', body)
        self.assertIn('Заказчик не указан', body)
        self.assertIn('Оператор не определён', body)
        self.assertIn('Дата не указана', body)
        self.assertIn('Подразделение не указано', body)
        self.assertIn('Тип оплаты не указан', body)

    # ── The workbooks, re-read with openpyxl ──────────────────────────────
    def test_the_report_workbook_matches_the_screen_cell_by_cell(self):
        response = self.client.get('/drones/works/reports.xlsx')
        self.assertEqual(response.status_code, 200)
        data, _filters = report_data()
        for title, cut_name in (('По заказчикам', 'by_customer'),
                                ('По операторам', 'by_operator'),
                                ('По подразделениям', 'by_subdivision'),
                                ('По месяцам', 'by_month'),
                                ('По типам оплаты', 'by_payment')):
            rows = sheet_rows(response.data, title)
            cut = data[cut_name]
            expected = len(cut['rows']) + len(cut['services']) + 2
            self.assertEqual(len(rows), expected, title)
            body = rows[1:-1]
            screen = cut['rows'] + list(cut['services'])
            for sheet_row, screen_row in zip(body, screen):
                self.assertEqual(sheet_row[0], screen_row['label'], title)
                self.assertEqual(sheet_row[1], screen_row['jobs'], title)
                self.assertAlmostEqual(sheet_row[2], screen_row['area'],
                                       places=2, msg=title)
                self.assertAlmostEqual(sheet_row[3], screen_row['amount'],
                                       places=2, msg=title)
                self.assertAlmostEqual(sheet_row[4], screen_row['received'],
                                       places=2, msg=title)
                self.assertAlmostEqual(sheet_row[5], screen_row['outstanding'],
                                       places=2, msg=title)
            total = rows[-1]
            self.assertEqual(total[1], TOTAL_JOBS, title)
            self.assertAlmostEqual(total[2], TOTAL_AREA, places=2, msg=title)
            self.assertAlmostEqual(total[3], TOTAL_AMOUNT, places=2,
                                   msg=title)
            self.assertAlmostEqual(total[4], TOTAL_RECEIVED, places=2,
                                   msg=title)

    def test_every_workbook_cut_reconciles_inside_the_file(self):
        """The file must stand on its own: sum the body, get the total row."""
        response = self.client.get('/drones/works/reports.xlsx')
        for title in ('По заказчикам', 'По операторам', 'По подразделениям',
                      'По месяцам', 'По типам оплаты'):
            rows = sheet_rows(response.data, title)
            body = rows[1:-1]
            total = rows[-1]
            self.assertAlmostEqual(sum(r[2] for r in body), total[2],
                                   places=2, msg=title)
            self.assertAlmostEqual(sum(r[3] for r in body), total[3],
                                   places=2, msg=title)
            self.assertEqual(sum(r[1] for r in body), total[1], title)

    def test_the_debt_workbook_matches_the_debt_view(self):
        response = self.client.get('/drones/works/debt.xlsx')
        self.assertEqual(response.status_code, 200)
        data, _filters = report_data()
        rows = sheet_rows(response.data, 'Долги — заказчики')
        body = rows[1:-1]
        screen = list(data['debt_by_customer']['rows'])
        if data['debt_by_customer']['rest']:
            screen.append(data['debt_by_customer']['rest'])
        self.assertEqual(len(body), len(screen))
        for sheet_row, screen_row in zip(body, screen):
            self.assertEqual(sheet_row[0], screen_row['label'])
            self.assertAlmostEqual(sheet_row[4], screen_row['outstanding'],
                                   places=2)
        self.assertAlmostEqual(rows[-1][4], TOTAL_OUTSTANDING, places=2)
        self.assertAlmostEqual(sum(r[4] for r in body), rows[-1][4], places=2)

    def test_the_flat_ledger_workbook_carries_every_row_and_its_provenance(self):
        response = self.client.get('/drones/works.xlsx')
        self.assertEqual(response.status_code, 200)
        rows = sheet_rows(response.data, 'Работы')
        self.assertEqual(len(rows), TOTAL_JOBS + 1)
        header = rows[0]
        self.assertIn('Файл-источник', header)
        self.assertIn('Заказчик — как написано', header)
        area_col = header.index('Гектары')
        self.assertAlmostEqual(sum(r[area_col] for r in rows[1:]), TOTAL_AREA,
                               places=2)
        # the unresolved rows arrive named, not blank
        customer_col = header.index('Заказчик')
        self.assertIn('Заказчик не определён',
                      {r[customer_col] for r in rows[1:]})

    def test_the_workbook_carries_the_filters_that_produced_it(self):
        response = self.client.get('/drones/works/reports.xlsx?period=2026-04')
        rows = sheet_rows(response.data, 'Сводка')
        values = {r[0]: r[1] for r in rows[1:]}
        self.assertEqual(values['Период'], '2026-04')
        self.assertEqual(values['Работ'], 6)
        self.assertAlmostEqual(values['Гектаров'], 60.5, places=2)

    def test_a_filtered_workbook_really_carries_fewer_rows(self):
        """The differential half: the filter has to reach the file."""
        everything = sheet_rows(
            self.client.get('/drones/works.xlsx').data, 'Работы')
        april = sheet_rows(
            self.client.get('/drones/works.xlsx?period=2026-04').data,
            'Работы')
        self.assertEqual(len(everything) - 1, TOTAL_JOBS)
        self.assertEqual(len(april) - 1, 6)


class DroneWorksPermissionTests(unittest.TestCase):
    """@module_required('drones') is enforced at the route, not at the link."""

    @classmethod
    def setUpClass(cls):
        seed()
        with app.app_context():
            user = User(username='no_drones', role=ROLE_OPERATOR,
                        full_name='No Drones')
            user.set_password('x')
            db.session.add(user)
            db.session.commit()
            cls.user_id = user.id

    def test_every_works_route_is_403_without_the_module(self):
        client = app.test_client()
        login(client, self.user_id)
        for url in ('/drones/works', '/drones/customers',
                    '/drones/works/reports', '/drones/works.xlsx',
                    '/drones/works/reports.xlsx', '/drones/works/debt.xlsx'):
            self.assertEqual(client.get(url).status_code, 403, url)

    def test_the_same_routes_are_200_for_an_admin(self):
        """The other half: a 403 everywhere is also what a broken route gives."""
        admin = create_admin('perm_admin')
        set_language(admin, 'ru')
        client = app.test_client()
        login(client, admin)
        for url in ('/drones/works', '/drones/customers',
                    '/drones/works/reports', '/drones/works.xlsx',
                    '/drones/works/reports.xlsx', '/drones/works/debt.xlsx'):
            self.assertEqual(client.get(url).status_code, 200, url)


class DroneWorksUzbekTests(unittest.TestCase):
    """Uzbek is CYRILLIC ONLY. Latin is allowed in product names alone."""

    ALLOWED_LATIN = ('Excel', 'DJI', 'xlsx', 'py', 'tools', 'import',
                     'drone', 'works', 'apply', 'RFID', 'PDF')

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('uz_admin')
        set_language(cls.admin, 'uz')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def test_the_screens_render_in_uzbek(self):
        for url in ('/drones/works', '/drones/customers',
                    '/drones/works/reports'):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)

    def test_the_service_row_labels_are_cyrillic(self):
        data, _filters = report_data(lang='uz')
        labels = []
        for name in ('by_customer', 'by_operator', 'by_subdivision',
                     'by_month', 'by_payment'):
            labels.extend(s['label'] for s in data[name]['services'])
        self.assertEqual(
            labels,
            ['Буюртмачи кўрсатилмаган', 'Буюртмачи аниқланмаган',
             'Оператор аниқланмаган', 'Бўлинма кўрсатилмаган',
             'Сана кўрсатилмаган', 'Тўлов тури кўрсатилмаган'])
        for label in labels:
            self.assertFalse(any('a' <= ch.lower() <= 'z' for ch in label),
                             label)

    def test_the_workbook_sheet_names_are_cyrillic(self):
        from openpyxl import load_workbook
        response = self.client.get('/drones/works/reports.xlsx')
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.data))
        self.assertEqual(
            wb.sheetnames,
            ['Жамланма', 'Буюртмачилар бўйича', 'Операторлар бўйича',
             'Бўлинмалар бўйича', 'Ойлар бўйича', 'Тўлов турлари бўйича',
             'Қарз — буюртмачилар', 'Қарз — операторлар'])
        for name in wb.sheetnames:
            self.assertFalse(any('a' <= ch.lower() <= 'z' for ch in name),
                             name)

    def test_the_uzbek_workbook_carries_the_same_numbers(self):
        """Different words, identical arithmetic."""
        response = self.client.get('/drones/works/reports.xlsx')
        rows = sheet_rows(response.data, 'Буюртмачилар бўйича')
        self.assertEqual(rows[-1][1], TOTAL_JOBS)
        self.assertAlmostEqual(rows[-1][2], TOTAL_AREA, places=2)
        self.assertAlmostEqual(rows[-1][3], TOTAL_AMOUNT, places=2)


class DroneCustomerKeyTests(unittest.TestCase):
    """The screen and the import must agree on what a spelling normalises to.

    [REASON]: they are two implementations of one rule -- drones.py cannot
    import from tools/ at request time. If they drift, the alias the screen
    creates stops matching the alias the import looks up and every re-import
    quietly creates a second customer. This test is the pin.
    """

    def test_the_screen_and_the_import_normalise_identically(self):
        import drones
        import sys
        import os
        tools = os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), 'tools')
        if tools not in sys.path:
            sys.path.insert(0, tools)
        import import_drone_works as imp
        for value in ('Миробод АМТ', '  Миробод   АМТ  ', 'МИРОБОД амт',
                      'Фукаро', 'Ғиждувон ПТЗ- ФХ', '', 'a  b\tc\nd'):
            self.assertEqual(drones._drone_customer_key(value),
                             imp.customer_key(value), repr(value))


class DroneAssignmentHintTests(unittest.TestCase):
    """The hint screen: read-only, and honest about what it cannot know.

    The fixture is built so the right answer is knowable by hand. In 2026-04
    the ledger gives Fayzullaev 35.5 ha (12.5 + 10.0 + the 9.0 ha row whose
    book is filed as March but whose own date is 25 April + the 4.0 ha row
    with no payment type) and Khamroev 25.0 ha; the flights give machine No 1
    22.4 ha, machine No 2 30.0 ha and 5.0 ha with no machine at all.

    So Fayzullaev's closest is No 2 at -15.5 % and Khamroev's is No 1 at
    -10.4 % -- NOT No 2 at +20.0 %, which is the point: the ranking is by the
    absolute relative difference, and a bigger machine is not a better match
    just because it is bigger.
    """

    @classmethod
    def setUpClass(cls):
        cls.ids = seed()
        cls.admin = create_admin('hint_admin')
        set_language(cls.admin, 'ru')
        with app.app_context():
            org = Organization(name='Agrocluster')
            db.session.add(org)
            db.session.flush()
            units = {}
            for number in (1, 2):
                unit = DroneUnit(number=number, organization_id=org.id)
                db.session.add(unit)
                db.session.flush()
                units[number] = unit.id
            cls.units = units
            # April flights, stored UTC. 19:30 UTC on 31 March is already
            # 1 April for the operators, so the fixture keeps well inside the
            # month rather than testing the boundary twice.
            flights = [
                (1, datetime.datetime(2026, 4, 5, 6, 0), 12.4),
                (1, datetime.datetime(2026, 4, 6, 6, 0), 10.0),
                (2, datetime.datetime(2026, 4, 7, 6, 0), 30.0),
                (None, datetime.datetime(2026, 4, 8, 6, 0), 5.0),
                # March, must not leak into the April comparison
                (1, datetime.datetime(2026, 3, 5, 6, 0), 500.0),
            ]
            for index, (number, started, area) in enumerate(flights, 1):
                db.session.add(DroneFlight(
                    dji_flight_id=index,
                    drone_unit_id=units.get(number),
                    nickname_raw='n%d' % index,
                    started_at=started, area_ha=area,
                    raw_json='{}'))
            db.session.commit()

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def hints(self, month='2026-04'):
        import drones
        from flask import g
        with app.test_request_context('/?month=' + month):
            g.lang = 'ru'
            return drones._drone_assignment_hints(month)

    def test_the_closest_machine_is_the_closest_by_relative_difference(self):
        data = self.hints()
        by_name = {r['name']: r for r in data['rows']}
        faiz = by_name['Файзуллаев Фурқат']
        self.assertAlmostEqual(faiz['area'], 35.5, places=2)
        self.assertEqual(faiz['best']['number'], 2)
        self.assertAlmostEqual(faiz['best']['area'], 30.0, places=2)
        self.assertAlmostEqual(faiz['best']['delta'], -15.4930, places=3)
        self.assertEqual(faiz['second']['number'], 1)
        self.assertAlmostEqual(faiz['second']['delta'], -36.9014, places=3)

        khamroev = by_name['Хамроев Шохрух']
        self.assertAlmostEqual(khamroev['area'], 25.0, places=2)
        # -10.4 beats +20.0 on absolute relative difference, which is the
        # whole ranking rule -- a bigger machine is not a better match.
        self.assertEqual(khamroev['best']['number'], 1)
        self.assertAlmostEqual(khamroev['best']['delta'], -10.4, places=2)
        self.assertEqual(khamroev['second']['number'], 2)
        self.assertAlmostEqual(khamroev['second']['delta'], 20.0, places=2)

    def test_the_month_boundary_holds(self):
        """March's 500 ha must not appear anywhere in the April comparison."""
        data = self.hints()
        self.assertAlmostEqual(data['machine_area'], 57.4, places=2)
        for machine in data['pool']:
            self.assertLess(machine['area'], 100.0)

    def test_flights_with_no_machine_are_shown_and_never_a_candidate(self):
        data = self.hints()
        self.assertEqual(data['unattributed']['flights'], 1)
        self.assertAlmostEqual(data['unattributed']['area'], 5.0, places=2)
        self.assertEqual({m['number'] for m in data['pool']}, {1, 2})

    def test_a_machine_already_assigned_to_that_operator_is_excluded(self):
        """The differential half: it is a candidate until the assignment exists."""
        before = self.hints()
        faiz_id = self.ids['op1']
        self.assertEqual(
            [r for r in before['rows']
             if r['operator_id'] == faiz_id][0]['best']['number'], 2)
        with app.app_context():
            row = DroneOperatorAssignment(
                operator_id=faiz_id, drone_unit_id=self.units[2],
                date_from=datetime.date(2026, 4, 1),
                date_to=datetime.date(2026, 4, 30))
            db.session.add(row)
            db.session.commit()
            assignment_id = row.id
        try:
            after = self.hints()
            faiz = [r for r in after['rows']
                    if r['operator_id'] == faiz_id][0]
            self.assertEqual(faiz['best']['number'], 1)
            self.assertIsNone(faiz['second'])
            self.assertEqual(faiz['already'], [2])
            # ... and the machine is still a candidate for the OTHER operator,
            # marked with who holds it, because two operators genuinely share
            # a machine in the source data.
            khamroev = [r for r in after['rows']
                        if r['operator_id'] == self.ids['op2']][0]
            candidates = [c for c in (khamroev['best'], khamroev['second'])
                          if c]
            taken = [c for c in candidates if c['number'] == 2]
            self.assertEqual(len(taken), 1)
            self.assertEqual(taken[0]['taken_by'], ['Файзуллаев Фурқат'])
        finally:
            with app.app_context():
                db.session.delete(db.session.get(DroneOperatorAssignment,
                                                 assignment_id))
                db.session.commit()

    def test_works_with_no_operator_are_reported_not_hidden(self):
        data = self.hints('2026-03')
        self.assertEqual(data['unresolved']['jobs'], 1)
        self.assertAlmostEqual(data['unresolved']['area'], 30.0, places=2)

    def test_the_screen_opens_and_says_loudly_that_it_only_suggests(self):
        response = self.client.get('/drones/works/assignment-hints')
        self.assertEqual(response.status_code, 200)
        body = response.data.decode('utf-8')
        self.assertIn('ЭТО ПОДСКАЗКА, А НЕ НАЗНАЧЕНИЕ', body)
        self.assertIn('довод', body.lower())

    def test_the_screen_writes_nothing(self):
        """Read-only means read-only: nothing exists that was not there."""
        with app.app_context():
            before = (DroneOperatorAssignment.query.count(),
                      DroneWork.query.count())
        self.client.get('/drones/works/assignment-hints?month=2026-04')
        self.client.get('/drones/works/assignment-hints?month=2026-03')
        with app.app_context():
            after = (DroneOperatorAssignment.query.count(),
                     DroneWork.query.count())
        self.assertEqual(before, after)

    def test_it_is_403_without_the_module(self):
        with app.app_context():
            user = User(username='no_drones_hints', role=ROLE_OPERATOR,
                        full_name='No Drones')
            user.set_password('x')
            db.session.add(user)
            db.session.commit()
            user_id = user.id
        client = app.test_client()
        login(client, user_id)
        self.assertEqual(
            client.get('/drones/works/assignment-hints').status_code, 403)


# ─── DRONE-REPORTS-HUB-001 ───────────────────────────────────────────────────

def ungroup(text):
    """Undo the display grouping so a rendered figure compares to its value.

    [REASON]: UI-NUMBER-FORMAT-001 groups thousands with U+00A0, and the
    whitespace normalisation these readers do turns that into a plain space --
    so a screen cell now reads «24 555 256» where the report structure holds
    24555256. The reconciliations below are about VALUES; grouping is display
    only, and no number changed. The separator is stripped only BETWEEN TWO
    DIGITS, so «Пешку АМТ» and «Хозяйство 0001» keep their spaces.

    This helper deliberately makes the reconciliations blind to grouping.
    Whether grouping happened, and whether it happened where it must not, is
    the job of DroneNumberFormatTests -- not of these.
    """
    import re
    return re.sub(r'(?<=\d)[\s\u00a0](?=\d)', '', text)


def debt_tables(body):
    """The two debt tables of a rendered page, cell by cell."""
    import re
    out = {}
    for title in ('Долги — по заказчикам', 'Долги — по операторам'):
        index = body.find(title)
        if index < 0:
            out[title] = None
            continue
        chunk = body[index:]
        table = chunk[chunk.find('<table'):chunk.find('</table>')]
        rows = []
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.S):
            cells = [' '.join(re.sub(r'<[^>]+>', '', c).split())
                     for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr,
                                         re.S)]
            if cells:
                rows.append(cells)
        out[title] = rows
    return out


class DroneReportsLauncherTests(unittest.TestCase):
    """/drones/reports -- ten tiles, no data, no filters, no export.

    Was five. DRONE-ANALYTICS-001 added four reports and DRONE-CLOSE-001 a
    tenth; the census below is updated rather than relaxed, because its whole
    purpose is that a tile cannot appear, move or change accent without
    somebody saying so here.
    """

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('hub_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def test_the_launcher_opens_with_twelve_tiles_in_order(self):
        import re
        response = self.client.get('/drones/reports')
        self.assertEqual(response.status_code, 200)
        body = response.data.decode('utf-8')
        tiles = re.findall(
            r'<a class="vs-card vs-report-tile ([\w-]+)" href="([^"]+)"', body)
        # DRONE-ANALYTICS-001: four tiles added. Five accents existed and all
        # five were in use, so three of the four REUSE an accent that says
        # what they are next to -- calendar beside the flight summary,
        # reconcile beside the works report, aging beside the debts -- and
        # only spray brings the sixth, is-purple, on tokens already defined.
        self.assertEqual(tiles, [
            ('is-primary', '/drones/summary'),
            ('is-info', '/drones/works/reports'),
            ('is-warning', '/drones/works/debts'),
            ('is-success', '/drones/works/assignment-hints'),
            ('is-primary', '/drones/reports/calendar'),
            ('is-info', '/drones/reports/reconcile'),
            ('is-warning', '/drones/works/debts/aging'),
            ('is-purple', '/drones/reports/spray'),
            # DRONE-CLOSE-001 reuses is-info, the accent `reconcile` carries:
            # the same comparison read by subdivision rather than by month
            # alone. A shared colour means shared data, and a new accent here
            # would claim a new data source that does not exist.
            ('is-info', '/drones/reports/closing'),
            # DRONE-CASH-SPLIT-001 reuses is-warning, the accent the debts
            # tile carries, and sits next to it: debts and the operator cash
            # sheet are two halves of one question, «who owes whom». The
            # customer owes the holding; the operator owes the till.
            ('is-warning', '/drones/reports/operator-cash'),
            # MEGA-3 reuses is-danger, the accent `sources` carries, and
            # stands beside it: both scream about the data itself -- a
            # machine gone silent there, a zero that cannot be a zero here.
            ('is-danger', '/drones/reports/health'),
            ('is-danger', '/drones/sources'),
        ])

    def test_every_tile_target_really_answers(self):
        """A launcher whose tiles 404 is worse than no launcher."""
        import re
        body = self.client.get('/drones/reports').data.decode('utf-8')
        for _accent, href in re.findall(
                r'<a class="vs-card vs-report-tile ([\w-]+)" href="([^"]+)"',
                body):
            self.assertEqual(self.client.get(href).status_code, 200, href)

    def test_the_launcher_carries_no_data_no_filter_no_export(self):
        body = self.client.get('/drones/reports').data.decode('utf-8')
        self.assertNotIn('<form', body.split('vs-report-tiles')[1])
        self.assertNotIn('.xlsx', body)
        self.assertNotIn('vs-stat-value', body)

    def test_the_icons_are_inline_svg_and_nothing_is_fetched(self):
        """UI-FONT-LOCAL-001: no CDN, no icon font, no remote asset."""
        body = self.client.get('/drones/reports').data.decode('utf-8')
        tiles = body.split('vs-report-tiles')[1]
        # One inline <svg> per tile, ten tiles. This is also X-6: a tile key
        # with no icon branch renders an EMPTY icon box and nothing fails, so
        # the count is the only thing that catches it.
        self.assertEqual(tiles.count('<svg'), 10)
        for marker in ('http://', 'https://', '<img', '@font-face', 'url('):
            self.assertNotIn(marker, tiles, marker)

    def test_it_is_403_without_the_module(self):
        with app.app_context():
            user = User(username='no_drones_hub', role=ROLE_OPERATOR,
                        full_name='No Drones')
            user.set_password('x')
            db.session.add(user)
            db.session.commit()
            user_id = user.id
        client = app.test_client()
        login(client, user_id)
        self.assertEqual(client.get('/drones/reports').status_code, 403)


class DroneNavStripTests(unittest.TestCase):
    """Отчёты in, Источники out -- and the route it left behind still works."""

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('nav_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    # [REASON]: ADAPT-1 свёл три реализации вторичной навигации в один
    # компонент (templates/_module_nav.html). Изменился ТОЛЬКО контейнер:
    # было <div class="vs-row vs-mb" style="..."> у дронов и два других вида
    # у АЗС и запчастей — стало <nav class="vs-pills vs-modulenav">.
    # Утверждения ниже не тронуты ни одним символом: те же семь вкладок, тот
    # же порядок, тот же список ссылок. Правка локатора — не ослабление
    # проверки; проверка, привязанная к атрибуту style, ловила вёрстку, а не
    # состав навигации.
    STRIP_OPEN = '<nav class="vs-pills vs-modulenav vs-mb"'

    def _strip(self, url='/drones/reports'):
        import re
        body = self.client.get(url).data.decode('utf-8')
        strip = body.split(self.STRIP_OPEN)[1]
        strip = strip[:strip.find('</nav>')]
        return re.findall(r'<a href="([^"]+)"[^>]*>([^<]+)</a>', strip)

    def test_the_strip_is_the_eight_tabs_in_order(self):
        # DRONE-USEFUL-AREA-001 добавил восьмую вкладку «Полезная площадь»
        # ПОСЛЕ «Работ»: показатель считается по вылетам, а не по ведомостям,
        # и стоять он должен рядом с производственной, а не денежной частью.
        # Прежние семь не тронуты ни порядком, ни написанием.
        self.assertEqual(
            [label.strip() for _href, label in self._strip()],
            ['Сводка', 'Вылеты', 'Машины', 'Операторы', 'Работы',
             'Полезная площадь', 'Заказчики', 'Отчёты'])

    def test_istochniki_left_the_strip_but_not_the_application(self):
        hrefs = [href for href, _label in self._strip()]
        self.assertNotIn('/drones/sources', hrefs)
        self.assertEqual(self.client.get('/drones/sources').status_code, 200)

    def test_svodka_is_in_the_strip_and_is_also_the_first_tile(self):
        """A launcher that omits the main report is confusing."""
        hrefs = [href for href, _label in self._strip()]
        self.assertIn('/drones/summary', hrefs)
        body = self.client.get('/drones/reports').data.decode('utf-8')
        first = body.split('vs-report-tiles')[1]
        self.assertLess(first.find('/drones/summary'),
                        first.find('/drones/works/reports'))

    def test_otchety_is_marked_active_on_the_pages_it_leads_to(self):
        # Требование не изменилось: на всех ПЯТИ страницах, входом в которые
        # служит лаунчер, полоса продолжает говорить «вы внутри отчётов».
        # Изменился признак активности: компонент полосы стал общим для трёх
        # модулей (ADAPT-1), и активное состояние в нём выражается
        # `vs-pill is-active`, а не `vs-btn vs-btn-primary`. Заодно
        # проверяется aria-current — то же самое, сказанное для скринридера;
        # раньше он этого не сообщал вовсе.
        for url in ('/drones/reports', '/drones/sources',
                    '/drones/works/reports', '/drones/works/debts',
                    '/drones/works/assignment-hints'):
            body = self.client.get(url).data.decode('utf-8')
            strip = body.split(self.STRIP_OPEN)[1]
            strip = strip[:strip.find('</nav>')]
            active = re.search(
                r'<a href="/drones/reports"\s+class="vs-pill is-active"\s+'
                r'aria-current="page"', strip)
            self.assertIsNotNone(active, url)


class DroneDebtsPageTests(unittest.TestCase):
    """The split that actually fixes the complaint.

    The load-bearing assertion is that the numbers did not move: the debts
    page renders the same _drone_works_report_data() output the reports page
    used to render, and both debt tables still sum to the same grand total as
    every other cut.
    """

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('debts_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def test_the_page_opens_and_carries_both_tables(self):
        response = self.client.get('/drones/works/debts')
        self.assertEqual(response.status_code, 200)
        tables = debt_tables(response.data.decode('utf-8'))
        self.assertIsNotNone(tables['Долги — по заказчикам'])
        self.assertIsNotNone(tables['Долги — по операторам'])

    def test_the_reports_page_no_longer_carries_them(self):
        """The differential half: the tables really left the old page."""
        body = self.client.get('/drones/works/reports').data.decode('utf-8')
        tables = debt_tables(body)
        self.assertIsNone(tables['Долги — по заказчикам'])
        self.assertIsNone(tables['Долги — по операторам'])
        self.assertNotIn('/drones/works/debt.xlsx', body)
        # ... but the page is not a dead end for somebody looking for them
        self.assertIn('/drones/works/debts', body)

    def test_the_excel_debt_button_moved_with_the_tables(self):
        body = self.client.get('/drones/works/debts').data.decode('utf-8')
        self.assertIn('/drones/works/debt.xlsx', body)
        self.assertEqual(
            self.client.get('/drones/works/debt.xlsx').status_code, 200)

    def test_every_cell_equals_what_the_report_structure_says(self):
        """Cell by cell against the same data the reports page computes."""
        data, _filters = report_data()
        tables = debt_tables(
            self.client.get('/drones/works/debts').data.decode('utf-8'))
        for title, key in (('Долги — по заказчикам', 'debt_by_customer'),
                           ('Долги — по операторам', 'debt_by_operator')):
            debt = data[key]
            screen = list(debt['rows']) + ([debt['rest']] if debt['rest']
                                           else [])
            rows = tables[title][1:-1]          # drop the header and total
            self.assertEqual(len(rows), len(screen), title)
            for rendered, computed in zip(rows, screen):
                self.assertEqual(rendered[0], computed['label'], title)
                self.assertEqual(ungroup(rendered[1]),
                                 str(computed['jobs']), title)
                self.assertEqual(ungroup(rendered[2]),
                                 '%.0f' % computed['amount'], title)
                self.assertEqual(ungroup(rendered[3]),
                                 '%.0f' % computed['received'], title)
                self.assertEqual(ungroup(rendered[4]),
                                 '%.0f' % computed['outstanding'], title)
            total = tables[title][-1]
            self.assertEqual(ungroup(total[1]), str(TOTAL_JOBS), title)
            self.assertEqual(ungroup(total[2]), '%.0f' % TOTAL_AMOUNT, title)
            self.assertEqual(ungroup(total[3]), '%.0f' % TOTAL_RECEIVED,
                             title)
            self.assertEqual(ungroup(total[4]), '%.0f' % TOTAL_OUTSTANDING,
                             title)

    def test_both_tables_reconcile_inside_the_page(self):
        """Sum the body, get the total row -- the page stands on its own."""
        tables = debt_tables(
            self.client.get('/drones/works/debts').data.decode('utf-8'))
        for title, rows in tables.items():
            body = rows[1:-1]
            total = rows[-1]
            for column in (1, 2, 3, 4):
                self.assertEqual(
                    sum(int(float(ungroup(r[column]))) for r in body),
                    int(float(ungroup(total[column]))),
                    '%s column %d' % (title, column))

    def test_the_no_debt_row_is_present_and_counted(self):
        tables = debt_tables(
            self.client.get('/drones/works/debts').data.decode('utf-8'))
        labels = [r[0] for r in tables['Долги — по операторам']]
        self.assertIn('Долга нет (остальные)', labels)

    def test_the_operator_due_memo_survives_the_move(self):
        """Without it the report shows money OWED as money collected."""
        body = self.client.get('/drones/works/debts').data.decode('utf-8')
        self.assertIn('Оператор топшириши керак', body)
        self.assertIn('ЕЩЁ ДОЛЖЕН', body)
        # Grouped since UI-NUMBER-FORMAT-001 -- the separator is U+00A0, and
        # asserting it literally here doubles as proof the filter reached the
        # memo, not only the tables.
        self.assertIn('<b>1\u00a0000\u00a0000</b>', body)

    def test_the_memo_can_be_absent_and_the_test_would_notice(self):
        """Negative control: no operator_due row, no memo.

        [REASON]: the memo is rendered inside {% if data.operator_due.jobs %}.
        A test that only ever saw the populated fixture could not tell the
        block from a hardcoded string.
        """
        with app.app_context():
            row = DroneWork.query.filter_by(
                received_kind=DRONE_RECEIVED_KIND_OPERATOR_DUE).one()
            row.received_kind = DRONE_RECEIVED_KIND_RECEIVED
            db.session.commit()
        try:
            body = self.client.get(
                '/drones/works/debts').data.decode('utf-8')
            self.assertNotIn('ЕЩЁ ДОЛЖЕН', body)
        finally:
            with app.app_context():
                row = DroneWork.query.filter_by(
                    source_row=OPERATOR_DUE_INDEX).one()
                row.received_kind = DRONE_RECEIVED_KIND_OPERATOR_DUE
                db.session.commit()

    def test_the_filters_reach_the_page(self):
        data, _filters = report_data('period=2026-04')
        tables = debt_tables(self.client.get(
            '/drones/works/debts?period=2026-04').data.decode('utf-8'))
        total = tables['Долги — по заказчикам'][-1]
        self.assertEqual(total[1], str(data['totals']['jobs']))
        self.assertNotEqual(total[1], str(TOTAL_JOBS))

    def test_it_is_403_without_the_module(self):
        with app.app_context():
            user = User(username='no_drones_debts', role=ROLE_OPERATOR,
                        full_name='No Drones')
            user.set_password('x')
            db.session.add(user)
            db.session.commit()
            user_id = user.id
        client = app.test_client()
        login(client, user_id)
        self.assertEqual(
            client.get('/drones/works/debts').status_code, 403)


class DroneReportsHubUzbekTests(unittest.TestCase):
    """Uzbek is CYRILLIC ONLY, verified by code point."""

    # Latin is allowed only in product and brand names.
    ALLOWED_LATIN = ('Excel', 'DJI')

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('hub_uz')
        set_language(cls.admin, 'uz')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    @staticmethod
    def latin_runs(text, allowed=()):
        """Every run of Latin letters left after the allowed words are cut."""
        import re
        for word in allowed:
            text = text.replace(word, ' ')
        return re.findall(r'[A-Za-z]+', text)

    def test_the_tile_strings_are_cyrillic(self):
        import drones
        offenders = []
        for tile in drones.DRONE_REPORT_TILES:
            for field in ('title_uz', 'subtitle_uz'):
                runs = self.latin_runs(tile[field], self.ALLOWED_LATIN)
                if runs:
                    offenders.append((tile['key'], field, runs))
        self.assertEqual(offenders, [])

    def test_the_scanner_fires_on_a_planted_latin_string(self):
        """Negative control. A scanner that never fires is not a scanner."""
        self.assertEqual(
            self.latin_runs('Ҳисоботлар', self.ALLOWED_LATIN), [])
        self.assertEqual(
            self.latin_runs('Ҳисоbотlар', self.ALLOWED_LATIN), ['b', 'l'])
        # ... and the allowance really allows
        self.assertEqual(
            self.latin_runs('Excel юклаб олиш', self.ALLOWED_LATIN), [])

    def test_the_uzbek_letters_really_are_cyrillic_code_points(self):
        """«Қ» must be U+049A, not a Latin K with a tail somebody pasted."""
        import drones
        titles = {t['key']: t['title_uz'] for t in drones.DRONE_REPORT_TILES}
        self.assertEqual(ord(titles['debts'][0]), 0x049A)      # Қ
        self.assertEqual(titles['debts'], 'Қарзлар')
        for name, char in (('Ҳ', 0x04B2), ('Ў', 0x040E), ('Ғ', 0x0492)):
            self.assertEqual(ord(name), char)

    def test_the_launcher_and_the_debts_page_render_in_uzbek(self):
        for url in ('/drones/reports', '/drones/works/debts'):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)

    def test_the_debts_page_hero_is_cyrillic_and_is_its_own(self):
        """After DRONE-DEBTS-DOUBLE-HEADER-001 there is one hero, and it is
        «Қарзлар» -- not the reports hero that used to sit above it."""
        body = self.client.get('/drones/works/debts').data.decode('utf-8')
        shape = page_shape(body)
        self.assertEqual(shape['hero_titles'], ['Қарзлар'])
        self.assertEqual(shape['forms'], ['/drones/works/debts'])
        self.assertEqual(shape['stat_grids'], 1)
        title = shape['hero_titles'][0]
        self.assertEqual(self.latin_runs(title, self.ALLOWED_LATIN), [])
        # «Қ» must be U+049A, not a Latin K somebody pasted a tail onto.
        self.assertEqual(ord(title[0]), 0x049A)

    def test_the_uzbek_summary_card_labels_are_cyrillic(self):
        import re
        body = self.client.get('/drones/works/debts').data.decode('utf-8')
        labels = [' '.join(l.split()) for l in re.findall(
            r'<div class="vs-stat-label">(.*?)</div>', body, re.S)]
        self.assertEqual(len(labels), 6)
        for label in labels:
            self.assertEqual(self.latin_runs(label, self.ALLOWED_LATIN), [],
                             label)
        # ... and the control still fires on the same scanner
        self.assertEqual(self.latin_runs('Кирим qилинган',
                                         self.ALLOWED_LATIN), ['q'])

    def test_the_uzbek_nav_strip_is_cyrillic(self):
        import re
        body = self.client.get('/drones/reports').data.decode('utf-8')
        strip = body.split('<nav class="vs-pills vs-modulenav vs-mb"')[1]
        strip = strip[:strip.find('</nav>')]
        labels = [label.strip() for _h, label
                  in re.findall(r'<a href="([^"]+)"[^>]*>([^<]+)</a>', strip)]
        self.assertEqual(labels[-1], 'Ҳисоботлар')
        for label in labels:
            self.assertEqual(self.latin_runs(label, self.ALLOWED_LATIN), [],
                             label)


# ─── DRONE-DEBTS-DOUBLE-HEADER-001 ───────────────────────────────────────────

def page_shape(body):
    """The structural counts a report page must satisfy.

    [REASON]: DRONE-REPORTS-HUB-001 proved the debt tables moved intact, cell
    by cell -- and nothing checked what came WITH them. works_debts.html was
    works_reports.html sliced too widely, so the page shipped two heroes, two
    filter forms and two rows of summary cards, and the FIRST form posted to
    the other page. An operator who set a period in the top form and pressed
    «Показать» was silently navigated somewhere else. A check on the numbers
    could never have caught that; this one looks at the shape.
    """
    import re
    return {
        'forms': re.findall(r'<form method="get" action="([^"]*)"', body),
        'hero_titles': [' '.join(t.split()) for t in re.findall(
            r'<div class="vs-hero-title">(.*?)</div>', body, re.S)],
        'stat_grids': len(re.findall(r'class="vs-stat-grid', body)),
    }


class DroneReportPageShapeTests(unittest.TestCase):
    """One form, one hero, one card row -- and the form stays on its page."""

    # (url, endpoint the single form must post to)
    FILTERED_PAGES = (
        ('/drones/works/reports', '/drones/works/reports'),
        ('/drones/works/debts', '/drones/works/debts'),
    )

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('shape_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def shape(self, url):
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200, url)
        return page_shape(response.data.decode('utf-8'))

    def test_each_filtered_page_has_exactly_one_form(self):
        for url, _endpoint in self.FILTERED_PAGES:
            self.assertEqual(len(self.shape(url)['forms']), 1, url)

    def test_each_form_posts_back_to_its_own_page(self):
        """The damaging half. A form that leaves is worse than no filter."""
        for url, endpoint in self.FILTERED_PAGES:
            self.assertEqual(self.shape(url)['forms'], [endpoint], url)

    def test_each_filtered_page_has_exactly_one_hero(self):
        for url, _endpoint in self.FILTERED_PAGES:
            self.assertEqual(len(self.shape(url)['hero_titles']), 1, url)

    def test_each_filtered_page_has_exactly_one_row_of_cards(self):
        for url, _endpoint in self.FILTERED_PAGES:
            self.assertEqual(self.shape(url)['stat_grids'], 1, url)

    def test_the_two_pages_share_no_hero_title(self):
        """Two pages titled the same thing are one page rendered twice."""
        reports = set(self.shape('/drones/works/reports')['hero_titles'])
        debts = set(self.shape('/drones/works/debts')['hero_titles'])
        self.assertEqual(reports & debts, set())

    def test_the_debts_page_is_titled_debts(self):
        self.assertEqual(self.shape('/drones/works/debts')['hero_titles'],
                         ['Долги'])

    def test_the_reports_page_is_titled_reports(self):
        self.assertEqual(self.shape('/drones/works/reports')['hero_titles'],
                         ['Отчёты по работам'])

    def test_the_launcher_has_one_hero_and_by_design_no_form_and_no_cards(self):
        """The same four assertions, adjusted for what a launcher IS.

        [REASON]: /drones/reports carries no filters and no data on purpose
        (DRONE-REPORTS-HUB-001), so «exactly one form» is the wrong assertion
        for it -- zero is the correct answer and the design intent. What
        transfers is the hero count, which is the check that would have caught
        the duplication.
        """
        shape = self.shape('/drones/reports')
        self.assertEqual(len(shape['hero_titles']), 1)
        self.assertEqual(shape['hero_titles'], ['Отчёты по дронам'])
        self.assertEqual(shape['forms'], [])
        self.assertEqual(shape['stat_grids'], 0)

    def test_no_report_page_repeats_a_hero(self):
        """Sweep: every drones page the launcher leads to."""
        for url in ('/drones/reports', '/drones/works/reports',
                    '/drones/works/debts', '/drones/works/assignment-hints',
                    '/drones/sources', '/drones/summary'):
            titles = self.shape(url)['hero_titles']
            self.assertEqual(len(titles), len(set(titles)), url)
            self.assertEqual(len(titles), 1, url)

    def test_no_page_carries_a_form_that_posts_somewhere_else(self):
        """Sweep: a GET form must never navigate off its own page."""
        for url in ('/drones/works/reports', '/drones/works/debts',
                    '/drones/works', '/drones/works/assignment-hints',
                    '/drones/summary'):
            for action in self.shape(url)['forms']:
                self.assertEqual(action, url, '%s -> %s' % (url, action))

    def test_the_shape_reader_would_notice_a_duplicate(self):
        """Negative control. A counter that never counts two is not a counter.

        [REASON]: every assertion above is an equality against 1. Fed markup
        that really does carry the duplication, the reader has to report 2 --
        otherwise «exactly one» is indistinguishable from «the regex never
        matched anything».
        """
        doubled = (
            '<div class="vs-hero-title">Отчёты по работам</div>'
            '<form method="get" action="/drones/works/reports">x</form>'
            '<div class="vs-stat-grid vs-mb">a</div>'
            '<div class="vs-hero-title">Долги</div>'
            '<form method="get" action="/drones/works/debts">y</form>'
            '<div class="vs-stat-grid vs-mb">b</div>')
        shape = page_shape(doubled)
        self.assertEqual(len(shape['forms']), 2)
        self.assertEqual(shape['forms'][0], '/drones/works/reports')
        self.assertEqual(len(shape['hero_titles']), 2)
        self.assertEqual(shape['stat_grids'], 2)
        # ... and on a single-section page it reports one of each
        single = doubled[doubled.index('<div class="vs-hero-title">Долги'):]
        shape = page_shape(single)
        self.assertEqual(len(shape['forms']), 1)
        self.assertEqual(len(shape['hero_titles']), 1)
        self.assertEqual(shape['stat_grids'], 1)


class DroneDebtsSummaryCardTests(unittest.TestCase):
    """The six figures on the debts page, before and against after.

    [REASON]: the fix deletes markup. These assertions exist so that deleting
    markup cannot quietly delete a number with it -- they read the rendered
    cards, not the structure that produces them.
    """

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('cards_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def cards(self, url):
        import re
        body = self.client.get(url).data.decode('utf-8')
        grid = body[body.index('vs-stat-grid'):]
        grid = grid[:grid.index('<div class="vs-help')
                    if '<div class="vs-help' in grid else len(grid)]
        labels = re.findall(r'<div class="vs-stat-label">(.*?)</div>', grid,
                            re.S)
        # [REASON]: DRONE-UI-CARD-WRAP-001 added `is-num` beside the class, so
        # the pattern admits FURTHER classes but still anchors on this one --
        # `vs-stat-value` followed by the quote or by a space and more classes.
        # Loosening it to a bare substring would let it match vs-stat-value-*
        # and read numbers this reader was never meant to see.
        values = re.findall(r'<div class="vs-stat-value(?: [^"]*)?">(.*?)</div>',
                            grid, re.S)
        return list(zip([' '.join(l.split()) for l in labels],
                        [' '.join(v.split()) for v in values]))

    def raw_cards(self, url):
        """The same six values WITHOUT whitespace normalisation.

        [REASON]: cards() collapses runs of whitespace, and U+00A0 counts as
        whitespace for str.split -- so it cannot tell a grouped figure from an
        ungrouped one. This reader keeps the separator byte for byte.
        """
        import re
        body = self.client.get(url).data.decode('utf-8')
        # Same anchoring as cards(), and for the same reason -- see there.
        return re.findall(r'<div class="vs-stat-value(?: [^"]*)?">(.*?)</div>',
                          body, re.S)

    def test_the_debts_page_shows_the_six_figures(self):
        # Grouped since UI-NUMBER-FORMAT-001; cards() has already turned the
        # U+00A0 into a plain space. The values are unchanged -- that is the
        # assertion right below this one.
        self.assertEqual(self.cards('/drones/works/debts'), [
            ('Работ', '10'),
            ('Гектары', '128.50'),
            ('Сумма', '24 555 256'),
            ('Получено', '12 500 000'),
            ('Не получено', '12 055 256'),
            ('Прочие расходы', '0'),
        ])

    def test_the_figures_match_the_report_structure(self):
        """No number changed value -- ungroup and the figures are identical."""
        data, _filters = report_data()
        totals = data['totals']
        cards = {label: ungroup(value)
                 for label, value in self.cards('/drones/works/debts')}
        self.assertEqual(cards['Работ'], str(totals['jobs']))
        self.assertEqual(cards['Сумма'], '%.0f' % totals['amount'])
        self.assertEqual(cards['Получено'], '%.0f' % totals['received'])
        self.assertEqual(cards['Не получено'],
                         '%.0f' % totals['outstanding'])

    def test_the_separator_really_is_the_non_breaking_space(self):
        """A plain space would let a browser break «2 491 / 814 505»."""
        values = [v.strip() for v in self.raw_cards('/drones/works/debts')]
        self.assertIn('24\u00a0555\u00a0256', values)
        self.assertIn('12\u00a0500\u00a0000', values)
        for value in values:
            self.assertNotIn(' ', value.replace('\u00a0', ''),
                             'plain space inside %r' % value)

    def test_the_short_figures_are_not_grouped(self):
        """10, 128.50 and 0 stay as they are: under five integer digits."""
        values = [v.strip() for v in self.raw_cards('/drones/works/debts')]
        self.assertIn('10', values)
        self.assertIn('128.50', values)
        self.assertIn('0', values)


# ─── DRONE-WORKS-PRICE-INTERNAL-001 ──────────────────────────────────────────

class DroneWorksPriceDefaultTests(unittest.TestCase):
    """The markup that lets «Цена за га» follow «Тип оплаты».

    HONESTY: this behaviour is CLIENT-SIDE. Everything below asserts that the
    page carries both tariffs, the three element ids the script binds to, and
    the script itself. It does NOT assert that clicking the select changes the
    field -- no browser runs in this suite. That half is proven by a human
    clicking it, and nobody has.
    """

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('price_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)
        self.body = self.client.get('/drones/works').data.decode('utf-8')

    def test_both_tariffs_are_in_the_markup_as_plain_integers(self):
        import re
        hint = self.body[self.body.index('id="wPriceHint"'):]
        hint = hint[:hint.index('>')]
        attrs = dict(re.findall(r'data-price-(\w+)="([^"]*)"', hint))
        self.assertEqual(attrs, {'cash': '200000', 'transfer': '200000',
                                 'internal': '85633'})
        for value in attrs.values():
            self.assertTrue(value.isdigit(), value)

    def test_the_tariffs_come_from_the_model_not_from_the_template(self):
        from models import DRONE_WORK_PRICE_SUGGESTIONS
        self.assertIn('data-price-internal="%d"'
                      % DRONE_WORK_PRICE_SUGGESTIONS['internal'], self.body)
        self.assertIn('data-price-cash="%d"'
                      % DRONE_WORK_PRICE_SUGGESTIONS['cash'], self.body)

    def test_the_three_ids_the_script_binds_to_all_exist(self):
        for element_id in ('wPriceHint', 'addWPayment', 'addWPrice'):
            self.assertIn('id="%s"' % element_id, self.body, element_id)

    def test_the_script_is_present_and_reads_the_data_attributes(self):
        self.assertIn('<script>', self.body)
        self.assertIn("getAttribute('data-price-internal')", self.body)
        self.assertIn("addEventListener('change'", self.body)

    def test_no_tojson_reaches_an_attribute(self):
        """The documented trap: tojson does not escape the double quote.

        [REASON]: CLAUDE.md records three forms already broken this way. The
        blocking checker tools/check_templates.py enforces it repository-wide;
        this asserts it for the one template that gained a script.
        """
        source = open('templates/drones/works.html', encoding='utf-8').read()
        import re
        self.assertEqual(re.findall(r'=\s*"[^"]*tojson[^"]*"', source), [])

    def test_the_internal_tariff_is_not_the_external_one(self):
        """The whole defect in one assertion: 85 633 is not 200 000."""
        from models import DRONE_WORK_PRICE_SUGGESTIONS
        self.assertNotEqual(DRONE_WORK_PRICE_SUGGESTIONS['internal'],
                            DRONE_WORK_PRICE_SUGGESTIONS['cash'])
        self.assertEqual(DRONE_WORK_PRICE_SUGGESTIONS['internal'], 85633)

    def test_the_form_still_prefills_the_external_rate(self):
        """The server-rendered default is unchanged: the script only reacts."""
        self.assertIn('id="addWPrice" class="vs-input" value="200000"',
                      self.body)


# ─── DRONE-CUSTOMERS-PAGE-SLOW-001 ───────────────────────────────────────────

def seed_many_customers(count=713, unresolved_jobs=10):
    """713 customers, the number the 2026-08-05 import actually created."""
    reset_db()
    with app.app_context():
        for index in range(1, count + 1):
            row = DroneCustomer(name='Хозяйство %04d' % index)
            db.session.add(row)
            db.session.flush()
            db.session.add(DroneCustomerAlias(
                raw_name='Хозяйство %04d' % index,
                normalized_key='хозяйство %04d' % index,
                drone_customer_id=row.id, is_active=True))
        # The one real spelling the task names, with its two internal spaces.
        special = DroneCustomer(name='Гурганжи Чорва   (Завкиддин)')
        db.session.add(special)
        db.session.flush()
        db.session.add(DroneCustomerAlias(
            raw_name='Гурганжи Чорва   (Завкиддин)',
            normalized_key='гурганжи чорва (завкиддин)',
            drone_customer_id=special.id, is_active=True))
        # An alias whose spelling shares nothing with its customer's name --
        # the case that proves search really looks at aliases.
        hidden = DroneCustomer(name='Пахтакор АМТ')
        db.session.add(hidden)
        db.session.flush()
        db.session.add(DroneCustomerAlias(
            raw_name='Фурқат фермер хўжалиги',
            normalized_key='фурқат фермер хўжалиги',
            drone_customer_id=hidden.id, is_active=True))
        for index in range(unresolved_jobs):
            db.session.add(DroneWork(
                period_month='2026-04', customer_raw='Неизвестное %d' % index,
                area_ha=23.1, payment_type='cash'))
        db.session.commit()
    return count + 2


class DroneCustomersPaginationTests(unittest.TestCase):
    """713 rows on one page made the browser QA agent give up four times."""

    PAGE_SIZE = 50

    @classmethod
    def setUpClass(cls):
        cls.total = seed_many_customers()
        cls.admin = create_admin('cust_page_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def rows_on(self, url):
        """How many customer rows the directory table renders."""
        import re
        body = self.client.get(url).data.decode('utf-8')
        table = body[body.index('Справочник'):]
        table = table[:table.index('</table>')] if '</table>' in table else table
        return len(re.findall(r'<tr(?: class="[^"]*")?>\s*\n?\s*<td style="font-weight:600;">',
                              table)), body

    def test_the_page_size_is_fifty(self):
        import drones
        self.assertEqual(drones.DRONE_CUSTOMERS_PAGE_SIZE, self.PAGE_SIZE)
        count, _body = self.rows_on('/drones/customers')
        self.assertEqual(count, self.PAGE_SIZE)

    def test_the_last_page_carries_the_remainder(self):
        pages = (self.total + self.PAGE_SIZE - 1) // self.PAGE_SIZE
        count, _body = self.rows_on('/drones/customers?page=%d' % pages)
        self.assertEqual(count, self.total - (pages - 1) * self.PAGE_SIZE)

    def test_a_page_beyond_the_end_clamps_instead_of_erroring(self):
        response = self.client.get('/drones/customers?page=9999')
        self.assertEqual(response.status_code, 200)
        response = self.client.get('/drones/customers?page=0')
        self.assertEqual(response.status_code, 200)
        response = self.client.get('/drones/customers?page=-3')
        self.assertEqual(response.status_code, 200)

    def test_the_total_stays_the_total_not_the_page_size(self):
        _count, body = self.rows_on('/drones/customers')
        self.assertIn('Всего: %d' % self.total, ' '.join(body.split()))

    def test_the_unresolved_banner_counts_across_all_customers(self):
        """[REASON]: a per-page count would read as «almost nothing to fix».

        The banner is how somebody notices there is work at all. It is
        asserted on the LAST page, where a page-scoped count would differ
        from a table-scoped one.
        """
        pages = (self.total + self.PAGE_SIZE - 1) // self.PAGE_SIZE
        for url in ('/drones/customers', '/drones/customers?page=%d' % pages,
                    '/drones/customers?q=Гурганжи'):
            body = ' '.join(
                self.client.get(url).data.decode('utf-8').split())
            self.assertIn('Работ без заказчика</span>: <b>10</b> (231.00 га)'
                          .replace('</span>', ''), body, url)

    def test_the_pickers_still_list_every_customer(self):
        """An alias must be movable to a customer on page 7."""
        import re
        body = self.client.get('/drones/customers').data.decode('utf-8')
        picker = body[body.index('id="addAliasCust"'):]
        picker = picker[:picker.index('</select>')]
        self.assertEqual(len(re.findall(r'<option value="\d+"', picker)),
                         self.total)

    def test_the_add_forms_are_still_there(self):
        body = self.client.get('/drones/customers').data.decode('utf-8')
        self.assertIn(url_for_test('drones.customers_add'), body)
        self.assertIn(url_for_test('drones.customer_alias_add'), body)


class DroneCustomersSearchTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.total = seed_many_customers()
        cls.admin = create_admin('cust_search_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    def body(self, url):
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200, url)
        return response.data.decode('utf-8')

    def test_search_collapses_the_internal_spaces(self):
        """«Гурганжи» must find «Гурганжи Чорва   (Завкиддин)»."""
        body = self.body('/drones/customers?q=Гурганжи')
        self.assertIn('Гурганжи Чорва', body)
        self.assertIn('найдено: 1', ' '.join(body.split()))

    def test_search_is_case_insensitive(self):
        body = ' '.join(self.body('/drones/customers?q=гурганжи').split())
        self.assertIn('найдено: 1', body)

    def test_search_folds_the_uzbek_letters(self):
        """«Фуркат» must find the alias spelled «Фурқат»."""
        body = ' '.join(self.body('/drones/customers?q=Фуркат').split())
        self.assertIn('найдено: 1', body)
        self.assertIn('Пахтакор АМТ', body)

    def test_search_matches_an_alias_whose_customer_name_differs(self):
        """The alias «Фурқат фермер хўжалиги» belongs to «Пахтакор АМТ»."""
        body = self.body('/drones/customers?q=хўжалиги')
        self.assertIn('Пахтакор АМТ', body)

    def test_search_narrows_the_result(self):
        """The differential half: the same page without q shows everything."""
        wide = ' '.join(self.body('/drones/customers').split())
        narrow = ' '.join(self.body('/drones/customers?q=Гурганжи').split())
        self.assertIn('Всего: %d' % self.total, wide)
        self.assertNotIn('найдено:', wide)
        self.assertIn('Всего: %d' % self.total, narrow)
        self.assertIn('найдено: 1', narrow)

    def test_a_search_that_finds_nothing_says_so(self):
        body = self.body('/drones/customers?q=такогонетнигде')
        self.assertIn('Ничего не найдено', body)

    def test_the_query_survives_paging(self):
        import re
        body = self.body('/drones/customers?q=Хозяйство')
        joined = ' '.join(body.split())
        self.assertIn('найдено: 713', joined)
        links = re.findall(r'href="(/drones/customers\?[^"]*page=2[^"]*)"',
                           body)
        self.assertTrue(links, 'no next-page link on a paged search')
        self.assertTrue(any('q=' in link for link in links), links)
        second = ' '.join(self.body(links[0].replace('&amp;', '&')).split())
        self.assertIn('найдено: 713', second)

    def test_the_page_survives_the_query(self):
        body = ' '.join(
            self.body('/drones/customers?q=Хозяйство&page=3').split())
        self.assertIn('страница 3 /', body)

    def test_the_search_fold_matches_the_import(self):
        """Pinned: search and import must fold the same, or search rots.

        [REASON]: drones.py cannot import from tools/ at request time, so the
        fold exists twice. This is the pin -- if one drifts, this fails
        instead of a farm quietly becoming unfindable.
        """
        import os
        import sys
        import drones
        tools = os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), 'tools')
        if tools not in sys.path:
            sys.path.insert(0, tools)
        import import_drone_works as imp
        for value in ('Гурганжи Чорва   (Завкиддин)', 'Фурқат', 'Фуркат',
                      'Беҳзод', 'Туйғун', 'ЎЗБЕКИСТОН', 'Ёрдамчи', '',
                      'a  b\tc\nd', 'Миробод АМТ'):
            self.assertEqual(drones._drone_search_key(value),
                             imp.uz_fold(value), repr(value))


# ─── UI-NUMBER-FORMAT-001 ────────────────────────────────────────────────

NBSP = '\u00a0'


class DroneNumberFormatTests(unittest.TestCase):
    """The grouping filter, value by value.

    The load-bearing assertion in this class is test_a_year_is_never_grouped.
    Everything else describes the feature; that one describes the defect this
    change is most likely to ship. A suite without it has not tested this.
    """

    def group(self, value):
        import drones
        return drones.drone_group_number(value)

    # ── What it must do ───────────────────────────────────────────────────

    def test_the_four_cases_the_task_names(self):
        self.assertEqual(self.group(2491814505),
                         '2' + NBSP + '491' + NBSP + '814' + NBSP + '505')
        self.assertEqual(self.group('15893.64'), '15' + NBSP + '893.64')
        self.assertEqual(self.group(906), '906')
        self.assertEqual(self.group(0), '0')
        self.assertIsNone(self.group(None))

    def test_the_separator_is_u00a0_by_code_point(self):
        """Not a plain space: a number must never wrap mid-figure."""
        out = self.group(2491814505)
        self.assertEqual([hex(ord(c)) for c in out if not c.isdigit()],
                         ['0xa0', '0xa0', '0xa0'])
        self.assertNotIn(' ', out)

    def test_the_decimal_separator_stays_the_point(self):
        """This task adds grouping. It does not restyle decimals."""
        self.assertEqual(self.group('2491814505.25'),
                         '2' + NBSP + '491' + NBSP + '814' + NBSP + '505.25')
        self.assertEqual(self.group('128.50'), '128.50')
        self.assertNotIn(',', self.group('15893.64'))

    def test_the_six_debt_figures_of_the_task_group_as_specified(self):
        """Section 4 of the task, verbatim, as the numbers the page reports."""
        self.assertEqual(
            [self.group(v) for v in (906, '15893.64', 2491814505,
                                     1904558439, 587256066, 145304400)],
            ['906',
             '15' + NBSP + '893.64',
             '2' + NBSP + '491' + NBSP + '814' + NBSP + '505',
             '1' + NBSP + '904' + NBSP + '558' + NBSP + '439',
             '587' + NBSP + '256' + NBSP + '066',
             '145' + NBSP + '304' + NBSP + '400'])

    def test_no_number_changes_value(self):
        """Strip the separators and the digits are the ones fed in."""
        for value in (0, 906, 2026, 15893, 2491814505, 1904558439,
                      587256066, 145304400, -2491814505):
            out = self.group(value)
            self.assertEqual(out.replace(NBSP, ''), str(value), out)

    def test_a_negative_number_keeps_its_sign_outside_the_grouping(self):
        self.assertEqual(self.group(-2491814505),
                         '-2' + NBSP + '491' + NBSP + '814' + NBSP + '505')

    def test_a_float_that_is_whole_does_not_grow_a_dot_zero(self):
        """report_data() hands the templates floats; 24555256.0 is money."""
        self.assertEqual(self.group(24555256.0),
                         '24' + NBSP + '555' + NBSP + '256')

    # ── THE NEGATIVE CONTROL ────────────────────────────────────────────

    def test_a_year_is_never_grouped(self):
        """2026 -> '2026', NOT '2 026'. The classic failure of this change.

        [REASON]: this is not a convention the filter follows, it is
        structural: DRONE_GROUP_MIN_DIGITS = 5 means a four-digit integer
        cannot be grouped by any call site, in any template, ever.
        """
        self.assertEqual(self.group(2026), '2026')
        self.assertEqual(self.group('2026'), '2026')
        for year in (1999, 2024, 2025, 2026, 2027, 2030):
            self.assertEqual(self.group(year), str(year))
            self.assertNotIn(NBSP, self.group(year))

    def test_the_naive_implementation_would_fail_that_control(self):
        """The control fires: prove it can tell right from wrong.

        [REASON]: CLAUDE.md -- a check that gives the same result for correct
        and incorrect code is not a check. This runs the obvious wrong
        implementation (format spec ',') past the same assertion and shows it
        is caught.
        """
        naive = lambda value: format(value, ',').replace(',', NBSP)
        self.assertEqual(naive(2026), '2' + NBSP + '026')     # the defect
        with self.assertRaises(AssertionError):
            self.assertEqual(naive(2026), '2026')

    def test_a_period_string_passes_through_untouched(self):
        """'2026-04' is a period, not a number."""
        for value in ('2026-04', '2026-03', '2026-05', '2026-04-05',
                      'За всё время'):
            self.assertEqual(self.group(value), value)

    def test_identifiers_pass_through_untouched(self):
        """Serials, nicknames, phones, IMEI, machine numbers.

        [REASON]: these read as digit strings but are names. Grouping one
        makes it un-searchable and un-copyable, and 1SZTJ4B001T0FC is not a
        quantity of anything.
        """
        for value in ('1SZTJ4B001T0FC', '+998901234567', 'DJI T50',
                      'Агро-3 (04)', '01 234 56 78', '3B7CH9004',
                      '86-1234-56', 'T50-2026-04'):
            self.assertEqual(self.group(value), value)

    def test_a_purely_numeric_identifier_string_is_the_known_limit(self):
        """Honest about what the filter cannot know.

        [REASON]: '998901234567' above survives because it carries a '+' or a
        space. A BARE 12-digit serial is indistinguishable from money to any
        filter, so the guarantee is at the CALL SITE: the filter is applied to
        money, area, litres and counts, and is not applied to identifier
        columns. test_no_identifier_column_carries_the_filter pins that.
        """
        self.assertIn(NBSP, self.group('998901234567'))

    def test_a_percentage_is_left_to_its_call_site(self):
        """«Доля, %» never reaches five integer digits anyway."""
        for value in ('12.5', '100.0', '0.0', '99.9'):
            self.assertEqual(self.group(value), value)

    def test_non_numbers_pass_through_unchanged_and_unwrapped(self):
        self.assertIsNone(self.group(None))
        self.assertEqual(self.group(''), '')
        self.assertEqual(self.group('—'), '—')
        self.assertIs(self.group(True), True)
        self.assertIs(self.group(False), False)
        self.assertEqual(self.group([1, 2]), [1, 2])

    def test_a_bool_is_not_treated_as_its_integer(self):
        """[REASON]: bool is a subclass of int; True would render '1'."""
        self.assertIsNot(self.group(True), 1)
        self.assertIs(self.group(True), True)

    def test_the_filter_is_published_app_wide_from_the_blueprint(self):
        """@bp.app_template_filter, so app.py is not touched."""
        import drones
        self.assertIs(app.jinja_env.filters['vs_num'],
                      drones.drone_group_number)
        with open('app.py', encoding='utf-8') as handle:
            source = handle.read()
        self.assertNotIn('vs_num', source)
        self.assertNotIn('drone_group_number', source)

    def test_it_works_as_a_jinja_filter_not_only_as_a_function(self):
        with app.app_context():
            rendered = app.jinja_env.from_string(
                '{{ a|vs_num }}|{{ b|vs_num }}|{{ c|vs_num }}').render(
                    a=2491814505, b=2026, c='2026-04')
        self.assertEqual(
            rendered,
            '2' + NBSP + '491' + NBSP + '814' + NBSP + '505|2026|2026-04')


class DroneNumberPlacementTests(unittest.TestCase):
    """WHERE the filter is applied -- the half a unit test usually misses."""

    TEMPLATES = ('_drones_nav.html', '_money_cell.html', 'coverage.html',
                 'customers.html',
                 'flight_calendar.html', 'health.html', 'list.html',
                 'operator_card.html', 'operator_cash.html',
                 'operators.html', 'reattach.html', 'reports.html',
                 'sources.html', 'spray_usage.html', 'summary.html',
                 'unit_card.html', 'units.html', 'works.html',
                 'works_assignment_hints.html', 'works_debts.html',
                 'works_debts_aging.html', 'works_flights_reconcile.html',
                 'works_closing.html',
                 'works_import.html', 'works_import_preview.html',
                 'works_reports.html')

    def source(self, name):
        with open('templates/drones/' + name, encoding='utf-8') as handle:
            return handle.read()

    def test_every_template_named_here_exists(self):
        """Guards the list against a rename silently emptying the scan."""
        import os
        on_disk = {n for n in os.listdir('templates/drones')
                   if n.endswith('.html')}
        self.assertEqual(set(self.TEMPLATES) - on_disk, set())
        self.assertEqual(on_disk - set(self.TEMPLATES), set())

    def test_no_number_filter_reaches_a_period_selects_options(self):
        """The template-level check the task asks for, by construction.

        [REASON]: a period is «2026-04» and a month select is where a
        careless grouping filter would land on a year. Every <select> whose
        name is a period-ish field is read out and its <option> markup is
        checked to carry no filter of any kind.
        """
        import re
        seen = 0
        for name in self.TEMPLATES:
            src = self.source(name)
            for match in re.finditer(
                    r'<select[^>]*name="(period|month|year|p)"[^>]*>(.*?)'
                    r'</select>', src, re.S):
                seen += 1
                block = match.group(2)
                self.assertNotIn('vs_num', block,
                                 '%s: %s select' % (name, match.group(1)))
                self.assertNotIn('|format', block,
                                 '%s: %s select' % (name, match.group(1)))
                for option in re.findall(r'<option[^>]*>(.*?)</option>',
                                         block, re.S):
                    self.assertNotIn('|', option,
                                     '%s: %s option %r'
                                     % (name, match.group(1), option))
        # The scan must have found the selects, or it proved nothing.
        # DRONE-ANALYTICS-001 brought two more: the period picker on the debt
        # aging page and the month picker on the flight calendar.
        # DRONE-CASH-SPLIT-001 brought the seventh -- the period picker of
        # the operator cash sheet. All are covered by the assertions above;
        # the count is raised so the scan still cannot pass by finding
        # nothing.
        self.assertEqual(seen, 7, 'expected 7 period selects, saw %d' % seen)

    def test_the_period_scan_would_notice_a_filter(self):
        """Negative control for the scan above: plant one, see it caught."""
        import re
        planted = ('<select name="period" id="x">'
                   '<option value="{{ p }}">{{ p|vs_num }}</option>'
                   '</select>')
        match = re.search(r'<select[^>]*name="(period)"[^>]*>(.*?)</select>',
                          planted, re.S)
        self.assertIsNotNone(match)
        with self.assertRaises(AssertionError):
            self.assertNotIn('vs_num', match.group(2))

    def test_no_identifier_column_carries_the_filter(self):
        """Serials, nicknames, phones and machine numbers stay unfiltered.

        [REASON]: the filter cannot tell a bare 12-digit serial from money --
        see test_a_purely_numeric_identifier_string_is_the_known_limit. The
        guarantee lives here, at the call site.
        """
        import re
        for name in self.TEMPLATES:
            for line in self.source(name).split('\n'):
                if 'vs_num' not in line:
                    continue
                lowered = line.lower()
                for word in ('serial', 'nickname', 'nick_', 'phone', 'imei',
                             'plate', 'firmware', 'reg_number', 'inn',
                             'apollo'):
                    self.assertNotIn(word, lowered,
                                     '%s: %s' % (name, line.strip()))

    def test_no_percentage_carries_the_filter(self):
        """«Доля, %» and Δ % stay exactly as they were."""
        import re
        for name in self.TEMPLATES:
            for line in self.source(name).split('\n'):
                if 'vs_num' not in line:
                    continue
                lowered = line.lower()
                for word in ('share', 'percent', 'delta'):
                    self.assertNotIn(word, lowered,
                                     '%s: %s' % (name, line.strip()))

    def test_the_filter_is_actually_applied_somewhere_on_every_screen(self):
        """A skipped template would make every assertion above vacuous."""
        # [REASON]: works_reports.html went from 18 call sites to 16, and no
        # number lost its grouping. DRONE-REPORT-ZERO-AMOUNT-001 replaced the
        # three hand-written amount cells -- ordinary row, service row, cut
        # total -- with one amount_cell() macro that applies the filter once.
        # This census counts CALL SITES; that the rendered figures are still
        # grouped is asserted against the rendered page by
        # DroneDebtsSummaryCardTests.test_the_separator_really_is_the_non_
        # breaking_space and by the cut-cell assertions in the C tests.
        #
        # DRONE-ZERO-VS-UNKNOWN-001 moved that macro out to _money_cell.html
        # and generalised it to all three money columns of both report
        # screens. Fifteen hand-written call sites -- three rows x (amount,
        # received, outstanding) on the debts page, three rows x (received,
        # outstanding) on the reports page -- collapsed into the ONE call site
        # inside money_cell(). works_reports.html 15 -> 9,
        # works_debts.html 15 -> 6, total 90 -> 75.
        #
        # [REASON]: the total is NOT an invariant and must not be read as one.
        # A shared macro applies the filter once for many cells, so a falling
        # total is the expected shape of this refactor and a census alone
        # cannot tell it apart from a number that lost its grouping. What
        # tells them apart is reading the RENDERED page: the cut-cell
        # assertions in tests/test_drone_zero_vs_unknown.py (C-1) look for
        # «13\xa0000\xa0000» in the rendered <td>, and
        # DroneDebtsSummaryCardTests.test_the_separator_really_is_the_non_
        # breaking_space does the same for the cards. This census only pins
        # WHERE the filter is written.
        #
        # DRONE-ANALYTICS-001 raised the census twice over, in two ways that
        # must not be confused with each other:
        #
        #   UI-NUMBER-FORMAT-002 applied the filter to counts of FLIGHTS and
        #   JOBS, which were printed raw everywhere. summary.html 26 -> 42,
        #   sources.html 4 -> 7, reattach.html 6 -> 12. Nothing else in those
        #   three templates changed. Counts of MACHINES (sources: silent,
        #   never), of DAYS (silent_days), machine numbers, ids and periods
        #   deliberately did NOT get the filter -- the task names flights and
        #   jobs, and a filter on a year or an id is the defect, not the fix.
        #
        #   Four new screens brought their own call sites: spray_usage 28,
        #   works_flights_reconcile 14, works_debts_aging 15, flight_calendar
        #   9. These are new numbers on new pages, not a change to old ones.
        #   works_flights_reconcile deliberately has NO filter on its two
        #   percentage cells and none on «Покрытие»: a percentage never gets
        #   the grouping filter, and that rule is asserted separately by
        #   test_no_percentage_carries_the_filter.
        #
        #   DRONE-WORKS-UPLOAD-001 brought the two upload screens. They print
        #   counts of JOBS and MONEY, which is exactly what UI-NUMBER-FORMAT-002
        #   named, so they carry the filter for the same reason: works_import
        #   4 (the four numeric journal columns), works_import_preview 30. The
        #   period on the form, the batch id, the sheet counts and the row
        #   numbers inside a rejected-row reference deliberately do NOT get it
        #   -- a period is «2026-08» and an id is an identifier, and the filter
        #   on either is the defect, not the fix.
        #
        #   DRONE-CLOSE-001 brought works_closing 24. Counts of WORKS and of
        #   FLIGHTS and hectares carry the filter, in the matrix cells, in the
        #   row and month totals, in the cards and in the «flew, no book»
        #   list. The COVERAGE PERCENTAGE deliberately does NOT -- a
        #   percentage never gets the grouping filter, the same rule
        #   works_flights_reconcile follows and
        #   test_no_percentage_carries_the_filter asserts -- and neither does
        #   the month string «2026-04», which is a period and not a number.
        #   MEGA-3 brought health 11: flights, hectares, counts and the sum
        #   carry the filter in all five sections; machine numbers («№ 5»)
        #   and months deliberately do NOT -- a machine number is an
        #   identifier and a month is a period.
        #
        #   DRONE-CASH-SPLIT-001 brought operator_cash 19 and raised
        #   works_debts from 6 to 18. Both are money and counts: collected,
        #   costs, handed in, balance, declared, and the same six figures per
        #   payment channel plus their footer. Operator NAMES carry nothing,
        #   because a name is not a number.
        #   DRONE-USEFUL-AREA-001 brought coverage.html 9: four cards
        #   (DJI area, calculated useful area, ready works, not-ready works),
        #   the works counter beside the table title, and per row the flight
        #   count, the flights-without-route note, the DJI area and the
        #   calculated one. The UNCERTAINTY column carries none: a percentage
        #   is never grouped in thousands, same rule as «Доля, %».
        expected = {
            '_money_cell.html': 1, 'coverage.html': 9,
            'customers.html': 2, 'flight_calendar.html': 9,
            'health.html': 11, 'list.html': 3,
            'operator_cash.html': 19,
            'reattach.html': 12, 'sources.html': 7, 'spray_usage.html': 28,
            'summary.html': 42, 'works.html': 9,
            'works_assignment_hints.html': 9, 'works_closing.html': 24,
            'works_debts.html': 18, 'works_debts_aging.html': 15,
            'works_flights_reconcile.html': 14, 'works_import.html': 4,
            'works_import_preview.html': 30, 'works_reports.html': 9,
        }
        actual = {name: self.source(name).count('|vs_num')
                  for name in self.TEMPLATES
                  if '|vs_num' in self.source(name)}
        self.assertEqual(actual, expected)
        self.assertEqual(sum(actual.values()), 275)


class DroneUiFixUzbekTests(unittest.TestCase):
    """Uzbek is CYRILLIC ONLY on the two screens this task changed.

    Latin is allowed only in product and brand names. The scan reads the
    UZBEK HALF of every «RU if is_ru else UZ» pair in the template source,
    which is exactly where a new string lands, and it is checked by code
    point -- «Қ» is U+049A, not a Latin K somebody pasted a tail onto.
    """

    # [REASON]: CLAUDE.md allows Latin only in product and brand names. A
    # command the operator has to TYPE is the same category and cannot be
    # transliterated -- «тоолс/импорт_дроне_воркс.пй» would not run. The
    # allowance is spelled out here so it stays two known strings rather than
    # a hole; test_the_scanner_fires_on_a_planted_latin_string shows the scan
    # still catches prose.
    ALLOWED_LATIN = ('Excel', 'DJI', 'tools/import_drone_works.py', '--apply')
    SCREENS = ('customers.html', 'works.html')

    @classmethod
    def setUpClass(cls):
        seed_many_customers()
        cls.admin = create_admin('uifix_uz')
        set_language(cls.admin, 'uz')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    @staticmethod
    def latin_runs(text, allowed=()):
        import re
        for word in allowed:
            text = text.replace(word, ' ')
        return re.findall(r'[A-Za-z]+', text)

    def uzbek_halves(self, name):
        """The UZ side of every «'RU' if is_ru else 'UZ'» in the template."""
        import re
        with open('templates/drones/' + name, encoding='utf-8') as handle:
            source = handle.read()
        return re.findall(r"if is_ru else '((?:[^'\\]|\\.)*)'", source)

    def test_the_scan_finds_the_strings_it_claims_to_scan(self):
        """A scan over an empty list proves nothing."""
        counts = {n: len(self.uzbek_halves(n)) for n in self.SCREENS}
        for name, count in counts.items():
            self.assertGreater(count, 10, '%s: %d' % (name, count))
        # The strings this task added must be among them.
        customers = self.uzbek_halves('customers.html')
        for expected in ('Топиш', 'Тозалаш', 'Саҳифа', 'Жами', 'топилди',
                         'саҳифа', 'Ҳеч нарса топилмади',
                         'Номи ва ёзилишлари бўйича қидирув'):
            self.assertIn(expected, customers, expected)

    def test_every_uzbek_string_on_both_screens_is_cyrillic(self):
        offenders = []
        for name in self.SCREENS:
            for text in self.uzbek_halves(name):
                runs = self.latin_runs(text, self.ALLOWED_LATIN)
                if runs:
                    offenders.append((name, text, runs))
        self.assertEqual(offenders, [])

    def test_the_scanner_fires_on_a_planted_latin_string(self):
        """The control. A scanner that never fires is not a scanner."""
        self.assertEqual(self.latin_runs('Саҳифа', self.ALLOWED_LATIN), [])
        self.assertEqual(self.latin_runs('Sahifa', self.ALLOWED_LATIN),
                         ['Sahifa'])
        self.assertEqual(self.latin_runs('Саҳifа', self.ALLOWED_LATIN),
                         ['if'])
        self.assertEqual(self.latin_runs('Excel юклаб олиш',
                                         self.ALLOWED_LATIN), [])

    def test_the_uzbek_letters_are_the_cyrillic_code_points(self):
        text = 'Номи ва ёзилишлари бўйича қидирув'
        self.assertIn(text, self.uzbek_halves('customers.html'))
        self.assertEqual(ord('ў'), 0x045E)
        self.assertEqual(ord('қ'), 0x049B)
        self.assertEqual(ord('ҳ'), 0x04B3)
        self.assertEqual(ord(text[text.index('ў')]), 0x045E)
        sahifa = 'Саҳифа'
        self.assertIn(sahifa, self.uzbek_halves('customers.html'))
        self.assertEqual(ord(sahifa[2]), 0x04B3)

    def test_both_screens_render_in_uzbek(self):
        for url in ('/drones/customers', '/drones/customers?q=Фуркат',
                    '/drones/customers?page=3', '/drones/works'):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)
            body = response.data.decode('utf-8')
            self.assertIn('Буюртмачилар', body, url)

    def test_the_uzbek_page_carries_the_search_and_the_pager(self):
        body = self.client.get('/drones/customers').data.decode('utf-8')
        self.assertIn('Номи ва ёзилишлари бўйича қидирув', body)
        self.assertIn('Топиш', body)
        self.assertIn('Саҳифа', body)
        self.assertIn('Жами', body)


class DroneExcelUntouchedTests(unittest.TestCase):
    """The exports must still hold NUMBERS, not formatted strings.

    [REASON]: the task forbids touching Excel at all. A cell holding
    '24 555 256' instead of 24555256 breaks every SUM the workbook does, and
    it breaks it silently -- the file still opens.
    """

    @classmethod
    def setUpClass(cls):
        seed()
        cls.admin = create_admin('xlsx_untouched_admin')
        set_language(cls.admin, 'ru')

    def setUp(self):
        self.client = app.test_client()
        login(self.client, self.admin)

    XLSX_ROUTES = ('/drones/summary.xlsx', '/drones/flights.xlsx',
                   '/drones/works.xlsx', '/drones/works/reports.xlsx',
                   '/drones/works/debt.xlsx')

    def test_no_exported_cell_carries_the_group_separator(self):
        from openpyxl import load_workbook
        checked = 0
        for url in self.XLSX_ROUTES:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)
            wb = load_workbook(io.BytesIO(response.data))
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        checked += 1
                        if isinstance(cell.value, str):
                            self.assertNotIn(
                                NBSP, cell.value,
                                '%s!%s!%s' % (url, ws.title,
                                              cell.coordinate))
        self.assertGreater(checked, 100, 'the scan read almost nothing')

    def test_the_money_cells_are_still_numeric(self):
        """Not «still present» -- still a NUMBER."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(
            self.client.get('/drones/works/debt.xlsx').data))
        found = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and \
                            not isinstance(cell.value, bool):
                        found.append(float(cell.value))
        self.assertIn(24555256.0, found)
        self.assertIn(12500000.0, found)
        self.assertIn(12055256.0, found)

    def test_no_xlsx_function_references_the_filter(self):
        """The filter is a Jinja filter; no export may reach for it."""
        import ast
        import inspect
        import drones
        tree = ast.parse(inspect.getsource(drones))
        names = []
        for node in ast.walk(tree):
            if isinstance(node, ast.FunctionDef) and \
                    node.name.endswith('_xlsx'):
                names.append(node.name)
                body = ast.dump(node)
                self.assertNotIn('vs_num', body, node.name)
                self.assertNotIn('drone_group_number', body, node.name)
                self.assertNotIn('DRONE_GROUP_SEPARATOR', body, node.name)
        # DRONE-ANALYTICS-001 added two exports; both are covered by the
        # assertions in the loop above, and the roster is updated so a third
        # one cannot appear without passing through this check.
        self.assertEqual(sorted(names),
                         ['flights_xlsx', 'spray_usage_xlsx', 'summary_xlsx',
                          'works_debt_xlsx', 'works_debts_aging_xlsx',
                          'works_reports_xlsx', 'works_xlsx'])


def url_for_test(endpoint):
    with app.test_request_context('/'):
        from flask import url_for
        return url_for(endpoint)


if __name__ == '__main__':
    unittest.main(verbosity=2)
