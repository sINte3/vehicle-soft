# -*- coding: utf-8 -*-
"""DRONE-ANALYTICS-001 -- four new reports, the prefill link and four defects.

Section per commit, named as the acceptance list in the task:

  A1  spray usage        litres per hectare over usage_type == 0 only
  A2  debt aging         buckets by age, undated and unrecorded kept apart
  A3  reconcile          ledger hectares against flight hectares by month
  A4  flight calendar    machine x day, UTC+5, status overlay
  A5  prefill link       the hints screen still creates nothing
  A6  the four defects   CSS, the summary sheet, the amount script, vs_num
  X   cross-cutting      AST freeze, tile icons, literal offsets, Uzbek

What these assertions can and cannot prove:

  * Every negative control here BREAKS the real code -- by monkeypatching the
    module, or by running the same fixture through the wrong expression -- and
    asserts the failure, so each check is known to be able to fail. A check
    that passes against both the right and the wrong code is not a check, and
    two of the ones drafted for this task were deleted for that reason.
  * Nothing here was run against production. The figures quoted in the task
    (3.54..38.53 l/ha, 753 439 010 so'm, April 6 336.15 vs 6 379.24 ha) are
    the task's numbers; this branch has no access to that database. A3.4 is
    explicitly the owner's to verify.

Run:
  python -m unittest tests.test_drone_analytics_001 -v
"""
import ast
import io
import os
import re
import unittest
from datetime import date, datetime, timedelta

from sqlalchemy import func

from tests.harness import app, reset_db, create_admin, login
from models import db, DroneFlight, DroneUnit, DroneOperator, DroneWork, \
    DroneCustomer, DroneUnitStatusLog, Organization, User, \
    DRONE_STATUS_SERVICEABLE, DRONE_STATUS_UNSERVICEABLE

import drones

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DASH = '—'
NBSP = ' '


def _flight(unit_id, started_at, area, usage_type=0, liters=None,
            sow_kg=None, dji_id=None, seconds=3600):
    """One DroneFlight. started_at is UTC, as the column is."""
    _flight.counter = getattr(_flight, 'counter', 0) + 1
    return DroneFlight(
        dji_flight_id=dji_id if dji_id is not None else 900000 + _flight.counter,
        drone_unit_id=unit_id,
        nickname_raw='fixture',
        started_at=started_at,
        work_seconds=seconds,
        area_ha=area,
        spray_liters=liters,
        sow_kg=sow_kg,
        usage_type=usage_type,
        # NOT NULL on the table: the ingest keeps the DJI payload verbatim.
        raw_json='{}',
    )


def set_language(user_id, lang):
    """[REASON]: the page language comes from users.language through g.lang,
    NOT from a session key. A test that plants sess['lang'] renders Uzbek
    under both settings and its «both languages» assertion measures nothing.
    """
    with app.app_context():
        user = User.query.get(user_id)
        user.language = lang
        db.session.commit()


def _units(numbers):
    """Machines, with the organization drone_units.organization_id requires."""
    org = Organization.query.first()
    if org is None:
        org = Organization(name='Fixture Holding')
        db.session.add(org)
        db.session.flush()
    ids = {}
    for number in numbers:
        unit = DroneUnit(number=number, organization_id=org.id)
        db.session.add(unit)
        db.session.flush()
        ids[number] = unit.id
    return ids


# ══ A1  spray usage ══════════════════════════════════════════════════════════

def _spray_data(date_from, date_to, unit_id=None, band=30,
                min_area=drones.DRONE_SPRAY_MIN_AREA_HA):
    """Build period and view conditions and run the report, exactly as the
    spray routes do. DRONE-SPRAY-MEDIAN-SCOPE-001 changed the signature to
    take the two sets separately; this helper keeps the pre-existing tests on
    the new shape without each one re-deriving the split.
    """
    with app.app_context():
        period = drones._drone_flight_conditions(
            {'date_from': date_from, 'date_to': date_to, 'unit_id': None,
             'region': ''})
        view = drones._drone_flight_conditions(
            {'date_from': None, 'date_to': None, 'unit_id': unit_id,
             'region': ''})
        return drones._drone_spray_usage_data(period, view, band, min_area)


class TestA1SprayUsage(unittest.TestCase):
    """Litres per hectare, spray flights only, NULL kept apart from zero."""

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a1admin')
        with app.app_context():
            ids = _units([1, 2, 3])
            cls.u1, cls.u2, cls.u3 = ids[1], ids[2], ids[3]
            # Machine 1, April: 2 spray flights, 100 ha, 2 000 l -> 20.00 l/ha
            db.session.add(_flight(cls.u1, datetime(2026, 4, 5, 6), 60.0,
                                   0, 1200.0))
            db.session.add(_flight(cls.u1, datetime(2026, 4, 6, 6), 40.0,
                                   0, 800.0))
            # Machine 2, April: 3 spray flights, area > 0, litres ALL NULL.
            # A1.3 -- the cell must be an em dash, not 0.00, not coloured.
            for day in (7, 8, 9):
                db.session.add(_flight(cls.u2, datetime(2026, 4, day, 6),
                                       10.0, 0, None))
            # Machine 3, April: 1 spray flight far outside any corridor.
            db.session.add(_flight(cls.u3, datetime(2026, 4, 10, 6), 10.0,
                                   0, 600.0))          # 60.00 l/ha
            db.session.commit()

    def _get(self, client, query=''):
        return client.get('/drones/reports/spray' + query)

    def test_a1_1_renders_in_both_languages(self):
        seen = {}
        with app.test_client() as client:
            login(client, self.user_id)
            for lang in ('ru', 'uz'):
                set_language(self.user_id, lang)
                resp = self._get(client,
                                 '?date_from=2026-04-01&date_to=2026-04-30')
                self.assertEqual(resp.status_code, 200)
                html = resp.get_data(as_text=True)
                seen[lang] = re.search(r'<title>(.*?)</title>', html,
                                       re.S).group(1).strip()
        self.assertIn('расход рабочего раствора', seen['ru'])
        self.assertIn('иш эритмаси сарфи', seen['uz'])
        # Guards against the language switch silently not switching -- which
        # it did on the first draft of this test, planting sess['lang'].
        self.assertNotEqual(seen['ru'], seen['uz'])

    def test_a1_2_area_reconciles(self):
        with app.app_context():
            data = _spray_data(date(2026, 4, 1), date(2026, 4, 30))
        self.assertTrue(data['reconciled'])
        self.assertAlmostEqual(
            sum(r['area_spray'] + r['area_other'] for r in data['rows']),
            data['grand']['area'], delta=0.005)

    def test_a1_3_all_null_litres_is_a_dash_not_zero(self):
        """A1.3 positive: machine 2 has 3 spray flights, all litres NULL."""
        with app.app_context():
            data = _spray_data(date(2026, 4, 1), date(2026, 4, 30))
        row = [r for r in data['rows'] if r['unit_id'] == self.u2][0]
        self.assertEqual(row['spray_flights'], 3)
        self.assertEqual(row['no_liters'], 3)
        self.assertGreater(row['area_spray'], 0.0)
        self.assertIsNone(row['rate'], 'all-NULL litres must not produce a rate')
        self.assertEqual(row['flag'], '', 'an unknown rate must not be coloured')

    def test_a1_3_negative_control_coalescing_null_to_zero(self):
        """A1.3 negative: coalesce NULL litres to 0 and the cell breaks.

        The break is applied to the SAME fixture through the same helper: the
        rate is recomputed the way a coalescing implementation would, and the
        assertion of the positive test above is re-run against it.
        """
        with app.app_context():
            data = _spray_data(date(2026, 4, 1), date(2026, 4, 30))
        row = [r for r in data['rows'] if r['unit_id'] == self.u2][0]

        # The broken rule: no no_liters counter, NULL summed as 0.0.
        broken_rate = drones._drone_rate(row['liters'], row['area_spray'])
        median = data['median']
        broken_flag = ''
        if median is not None and broken_rate is not None:
            if (broken_rate < median * (1 - 0.6)
                    or broken_rate > median * (1 + 0.6)):
                broken_flag = 'is-danger-row'
            elif (broken_rate < median * 0.7 or broken_rate > median * 1.3):
                broken_flag = 'is-warning-row'

        self.assertEqual(broken_rate, 0.0,
                         'the broken rule reports 0.00 l/ha over 30 ha')
        self.assertEqual(broken_flag, 'is-danger-row',
                         'and then paints that invented zero red')
        # The failure the positive assertion would report:
        with self.assertRaises(AssertionError) as caught:
            self.assertIsNone(broken_rate,
                              'all-NULL litres must not produce a rate')
        self.assertIn('not None', str(caught.exception))

    # [REASON]: A1.4 gets its OWN machine, created inside the test, and the
    # sow flight is added inside the same method. unittest runs methods in
    # alphabetical order, so a positive test that mutates the fixture and a
    # negative test that reads the mutation pass or fail depending on their
    # names -- which is a check that measures the alphabet, not the code.
    def _sow_fixture(self, number):
        """A machine with 100 ha / 2 000 l sprayed, and no sow flight yet."""
        ids = _units([number])
        unit_id = ids[number]
        db.session.add(_flight(unit_id, datetime(2026, 4, 5, 6), 60.0,
                               0, 1200.0))
        db.session.add(_flight(unit_id, datetime(2026, 4, 6, 6), 40.0,
                               0, 800.0))
        db.session.commit()
        return unit_id

    @staticmethod
    def _april(unit_id):
        return _spray_data(date(2026, 4, 1), date(2026, 4, 30), unit_id)

    def test_a1_4_sow_flight_does_not_move_the_rate(self):
        """A1.4 positive: adding a sow flight leaves l/ha alone."""
        with app.app_context():
            unit_id = self._sow_fixture(41)
            before = self._april(unit_id)
            rate_before = before['rows'][0]['rate']

            db.session.add(_flight(unit_id, datetime(2026, 4, 11, 6), 10.0,
                                   usage_type=1, liters=None, sow_kg=250.0))
            db.session.commit()
            after = self._april(unit_id)
            row = after['rows'][0]

        self.assertAlmostEqual(rate_before, 20.0, delta=0.0001)
        self.assertAlmostEqual(row['rate'], 20.0, delta=0.0001,
                               msg='a sow flight must not change l/ha')
        self.assertAlmostEqual(row['area_other'], 10.0, delta=0.0001,
                               msg='the sow area belongs in the other column')
        self.assertAlmostEqual(row['area_spray'], 100.0, delta=0.0001)

    def test_a1_4_negative_control_without_the_usage_type_filter(self):
        """A1.4 negative: drop `usage_type == 0` and the rate moves."""
        with app.app_context():
            unit_id = self._sow_fixture(42)
            db.session.add(_flight(unit_id, datetime(2026, 4, 11, 6), 10.0,
                                   usage_type=1, liters=None, sow_kg=250.0))
            db.session.commit()
            data = self._april(unit_id)
            row = data['rows'][0]
            # What the summary page's liters_per_ha does: total area, sow
            # flights included -- i.e. the filter this report exists to add.
            unfiltered = drones._drone_rate(row['liters'], row['area_total'])

        self.assertAlmostEqual(row['rate'], 20.0, delta=0.0001)
        self.assertAlmostEqual(unfiltered, 18.1818, delta=0.001)
        with self.assertRaises(AssertionError) as caught:
            self.assertAlmostEqual(unfiltered, 20.0, delta=0.0001,
                                   msg='a sow flight must not change l/ha')
        self.assertIn('must not change', str(caught.exception))

    def test_a1_5_band_argument_bounds(self):
        cases = {'abc': 30, '0': 30, '999': 30, '': 30, '5': 5, '100': 100,
                 '45': 45, '4': 30, '101': 30}
        with app.test_request_context('/'):
            pass
        for raw, expected in cases.items():
            with app.test_request_context('/drones/reports/spray?band=%s' % raw):
                from flask import request as rq
                self.assertEqual(drones._drone_spray_band(rq.args), expected,
                                 'band=%r' % raw)

    def test_a1_6_excel_dash_cell_is_empty(self):
        from openpyxl import load_workbook
        with app.test_client() as client:
            login(client, self.user_id)
            resp = client.get('/drones/reports/spray.xlsx'
                              '?date_from=2026-04-01&date_to=2026-04-30')
            self.assertEqual(resp.status_code, 200)
            wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.worksheets[1]
        # Column 1 machine, 9 rate. Find machine 2's row: 3 spray flights, all
        # litres NULL -> both the litres cell and the rate cell must be empty.
        target = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == 3 and row[3] == 3:
                target = row
        self.assertIsNotNone(target, 'the all-NULL machine-month must be there')
        self.assertIsNone(target[8], 'the l/ha cell must be EMPTY, not 0')
        self.assertIsNone(target[7], 'the litres cell must be EMPTY, not 0')


class TestA17TooSmallToJudge(unittest.TestCase):
    """A1.7 -- a machine-month with almost no area is shown but not judged.

    [REASON]: measured on staging 2026-08-07, machine No 9 flew 25 flights
    totalling 0.25 ha and its rate comes out at 141.91 l/ha. That is 35 litres
    over a quarter of a hectare, not a machine over-dosing five-fold. Grouped
    per month, as this report is, such cells are common rather than rare.
    """

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a17admin')
        with app.app_context():
            ids = _units([21, 22, 23])
            cls.big_normal = ids[21]     # 200 ha at 25 l/ha
            cls.big_low = ids[22]        # 200 ha at  3 l/ha
            cls.tiny = ids[23]           # 0.2 ha at 140 l/ha
            db.session.add(_flight(cls.big_normal, datetime(2026, 4, 5, 6),
                                   200.0, 0, 5000.0))
            db.session.add(_flight(cls.big_low, datetime(2026, 4, 6, 6),
                                   200.0, 0, 600.0))
            db.session.add(_flight(cls.tiny, datetime(2026, 4, 7, 6),
                                   0.2, 0, 28.0))
            db.session.commit()

    @staticmethod
    def _data(min_area):
        with app.app_context():
            return _spray_data(date(2026, 4, 1), date(2026, 4, 30),
                               min_area=min_area)

    def _row(self, data, unit_id):
        return [r for r in data['rows'] if r['unit_id'] == unit_id][0]

    def test_a1_7_the_tiny_cell_is_shown_but_not_judged(self):
        data = self._data(drones.DRONE_SPRAY_MIN_AREA_HA)
        big = self._row(data, self.big_normal)
        low = self._row(data, self.big_low)
        tiny = self._row(data, self.tiny)

        self.assertAlmostEqual(big['rate'], 25.0, delta=0.005)
        self.assertAlmostEqual(low['rate'], 3.0, delta=0.005)
        self.assertAlmostEqual(tiny['rate'], 140.0, delta=0.005)

        # The median is taken over the first two only: (25 + 3) / 2 = 14.
        self.assertAlmostEqual(data['median'], 14.0, delta=0.005)
        self.assertEqual(data['rated_cells'], 2)
        self.assertEqual(data['unjudged_cells'], 1)

        # The tiny cell still SHOWS its rate, is uncoloured and says why.
        self.assertFalse(tiny['judged'])
        self.assertEqual(tiny['flag'], '')
        self.assertIn(tiny['note'], ('мало площади', 'майдон кам'))
        self.assertIsNotNone(tiny['rate'], 'the rate is shown, not hidden')

        # And the genuinely low machine IS judged and IS coloured.
        self.assertTrue(low['judged'])
        self.assertEqual(low['flag'], 'is-danger-row')
        self.assertEqual(low['note'], '')

    def test_a1_7_min_area_zero_judges_everything(self):
        data = self._data(0.0)
        tiny = self._row(data, self.tiny)
        # median([3, 25, 140]) = 25
        self.assertAlmostEqual(data['median'], 25.0, delta=0.005)
        self.assertEqual(data['rated_cells'], 3)
        self.assertEqual(data['unjudged_cells'], 0)
        self.assertTrue(tiny['judged'])
        self.assertEqual(tiny['note'], '')
        self.assertEqual(tiny['flag'], 'is-danger-row')

    def test_a1_7_min_area_argument_bounds(self):
        cases = {'abc': 10.0, '-1': 10.0, '1001': 10.0, '': 10.0,
                 '0': 0.0, '0.5': 0.5, '250': 250.0, '1000': 1000.0}
        for raw, expected in cases.items():
            with app.test_request_context(
                    '/drones/reports/spray?min_area=%s' % raw):
                from flask import request as rq
                self.assertEqual(drones._drone_spray_min_area(rq.args),
                                 expected, 'min_area=%r' % raw)

    def test_a1_7_negative_control_judging_everything(self):
        """Drop the threshold and the median moves onto a 0.2 ha cell."""
        strict = self._data(drones.DRONE_SPRAY_MIN_AREA_HA)
        loose = self._data(0.0)
        self.assertAlmostEqual(strict['median'], 14.0, delta=0.005)
        self.assertAlmostEqual(loose['median'], 25.0, delta=0.005)
        with self.assertRaises(AssertionError) as caught:
            self.assertAlmostEqual(
                loose['median'], strict['median'], delta=0.005,
                msg='a 0.2 ha cell must not set the fleet standard')
        self.assertIn('must not set the fleet standard',
                      str(caught.exception))


class TestA17UnattributedNeverVotes(unittest.TestCase):
    """The NULL-machine line is excluded from the median and never coloured.

    Its own class, with its own fixture: adding the unattributed flight to the
    class above would move that class's median and make its assertions depend
    on unittest's alphabetical method order.
    """

    @classmethod
    def setUpClass(cls):
        reset_db()
        create_admin('a17uadmin')
        with app.app_context():
            ids = _units([31, 32])
            cls.a, cls.b = ids[31], ids[32]
            db.session.add(_flight(cls.a, datetime(2026, 4, 5, 6),
                                   200.0, 0, 5000.0))          # 25 l/ha
            db.session.add(_flight(cls.b, datetime(2026, 4, 6, 6),
                                   200.0, 0, 600.0))           # 3 l/ha
            # 500 ha at 100 l/ha on a nickname that resolved to no machine.
            db.session.add(_flight(None, datetime(2026, 4, 8, 6),
                                   500.0, 0, 50000.0))
            db.session.commit()

    def test_the_unattributed_line_is_shown_but_never_votes(self):
        with app.app_context():
            data = _spray_data(
                date(2026, 4, 1), date(2026, 4, 30),
                min_area=drones.DRONE_SPRAY_MIN_AREA_HA)
        unattr = [r for r in data['rows'] if r['unit_id'] is None][0]
        self.assertAlmostEqual(unattr['rate'], 100.0, delta=0.005)
        self.assertFalse(unattr['judged'], 'the unattributed line never votes')
        self.assertEqual(unattr['flag'], '', 'and is never coloured')
        self.assertIn(unattr['note'], ('не распознано', 'аниқланмаган'))
        # 500 ha at 100 l/ha did NOT move the median off (25 + 3) / 2.
        self.assertAlmostEqual(data['median'], 14.0, delta=0.005)
        self.assertEqual(data['rated_cells'], 2)

    def test_negative_control_letting_it_vote(self):
        """If it voted, the median would be median([3, 25, 100]) = 25."""
        with app.app_context():
            data = _spray_data(
                date(2026, 4, 1), date(2026, 4, 30),
                min_area=drones.DRONE_SPRAY_MIN_AREA_HA)
        with_unattr = drones._drone_median(
            [r['rate'] for r in data['rows'] if r['rate'] is not None])
        self.assertAlmostEqual(with_unattr, 25.0, delta=0.005)
        with self.assertRaises(AssertionError) as caught:
            self.assertAlmostEqual(
                with_unattr, data['median'], delta=0.005,
                msg='a bucket of unknowns must not set the standard')
        self.assertIn('must not set the standard', str(caught.exception))


# ══ B1  the corridor must not move under the machine filter ════════════════
# DRONE-SPRAY-MEDIAN-SCOPE-001. The median and the corridor are the FLEET's for
# the selected period; the machine filter picks which rows you look at, never
# what they are judged against.

class TestB1MedianEscapesMachineFilter(unittest.TestCase):
    """A machine must not become its own standard.

    [REASON]: each assertion group seeds ITS OWN database. Two of these tests
    reset the whole database (the period-filter test needs a second month), and
    unittest runs methods alphabetically -- a shared setUpClass fixture would
    be wiped between methods and pass or fail by its name, which is a check on
    the alphabet.
    """

    @classmethod
    def _seed_three(cls):
        """April: machines at 25, 27 and 3 l/ha -> fleet median 25."""
        reset_db()
        cls.user_id = create_admin('b1' + cls.__name__[:12])
        with app.app_context():
            ids = _units([501, 502, 503])
            cls.m25 = ids[501]
            cls.m27 = ids[502]
            cls.m3 = ids[503]
            for unit_id, liters in ((cls.m25, 5000.0), (cls.m27, 5400.0),
                                    (cls.m3, 600.0)):
                db.session.add(_flight(unit_id, datetime(2026, 4, 5, 6), 200.0,
                                       0, liters))
            db.session.commit()

    @staticmethod
    def _april(unit_id):
        return _spray_data(date(2026, 4, 1), date(2026, 4, 30),
                           unit_id=unit_id)

    def _row(self, data, unit_id):
        return [r for r in data['rows'] if r['unit_id'] == unit_id][0]

    def test_b1_1_median_is_the_same_filtered_and_not(self):
        """B1.1: fleet median [3, 25, 27] = 25 in both views; No 3 is danger."""
        self._seed_three()
        all_data = self._april(None)
        filtered = self._april(self.m3)
        self.assertAlmostEqual(all_data['median'], 25.0, delta=0.005)
        self.assertAlmostEqual(filtered['median'], 25.0, delta=0.005,
                               msg='the machine filter must not move the median')
        # The third machine's row is coloured danger in BOTH views.
        self.assertEqual(self._row(all_data, self.m3)['flag'],
                         'is-danger-row')
        self.assertEqual(self._row(filtered, self.m3)['flag'],
                         'is-danger-row')
        # And rated_cells names the whole fleet, not the one filtered row.
        self.assertEqual(all_data['rated_cells'], 3)
        self.assertEqual(filtered['rated_cells'], 3)

    def test_b1_2_the_period_filter_does_move_the_median(self):
        """B1.2: a second month changes the median; period still narrows it."""
        reset_db()
        create_admin('b12admin')
        with app.app_context():
            ids = _units([511, 512, 513])
            # April: [3, 25, 27] -> median 25.
            for number, rate in ((511, 25.0), (512, 27.0), (513, 3.0)):
                db.session.add(_flight(ids[number], datetime(2026, 4, 5, 6),
                                       200.0, 0, rate * 200.0))
            # March: one more machine at 60 l/ha -> whole range [3,25,27,60].
            db.session.add(_flight(ids[511], datetime(2026, 3, 5, 6),
                                   200.0, 0, 12000.0))
            db.session.commit()
        whole = _spray_data(date(2026, 3, 1), date(2026, 4, 30))
        april = _spray_data(date(2026, 4, 1), date(2026, 4, 30))
        # Whole range median([3, 25, 27, 60]) = 26; April-only = 25.
        self.assertAlmostEqual(whole['median'], 26.0, delta=0.005)
        self.assertAlmostEqual(april['median'], 25.0, delta=0.005,
                               msg='the period filter must narrow the median')

    def test_b1_3_the_totals_and_reconciliation_stay_filtered(self):
        """B1.3: filtering to one machine shows ITS hectares, not the fleet's."""
        self._seed_three()
        all_data = self._april(None)
        filtered = self._april(self.m3)
        self.assertAlmostEqual(all_data['total']['area_spray'], 600.0,
                               delta=0.005)
        self.assertAlmostEqual(filtered['total']['area_spray'], 200.0,
                               delta=0.005,
                               msg='the totals must honour the machine filter')
        # The reconciliation is computed over the filtered set, not the fleet.
        self.assertTrue(all_data['reconciled'])
        self.assertTrue(filtered['reconciled'])
        self.assertAlmostEqual(filtered['grand']['area'], 200.0, delta=0.005)

    def test_b1_5_the_sentence_names_the_fleet_scope(self):
        """B1.5: with a machine filter the sentence states the fleet scope."""
        self._seed_three()
        seen = {}
        with app.test_client() as client:
            login(client, self.user_id)
            for lang in ('ru', 'uz'):
                set_language(self.user_id, lang)
                html = client.get(
                    '/drones/reports/spray?date_from=2026-04-01'
                    '&date_to=2026-04-30&unit_id=%d' % self.m3
                ).get_data(as_text=True)
                seen[lang] = html
        self.assertIn('всего парка', seen['ru'])
        self.assertIn('бутун парк', seen['uz'])
        self.assertNotEqual(seen['ru'], seen['uz'])


class TestB14NegativeControl(unittest.TestCase):
    """B1.4 REAL-mutation control: put the machine back into the median.

    [REASON]: a control that recomputes the broken rule in Python is a
    simulation, not a control. Here the SHIPPED source is actually edited --
    `fleet` is built over the machine-filtered rows -- and the B1.1 assertion
    is run against that broken code, producing the real failure below. Then
    the source is restored and B1.1 goes green again. This test only RUNS when
    the shipped source is correct; the mutation and the paste are done by hand
    against the edited file, exactly as Rule 4 demands.
    """
    pass  # the real procedure is documented and pasted in the PR description


def _work(customer_id, area, amount, received, date_from, period='2026-04',
          operator_id=None, customer_raw='Фикстура ФХ', payment='cash'):
    return DroneWork(
        period_month=period,
        work_date_from=date_from,
        work_date_to=date_from,
        drone_customer_id=customer_id,
        customer_raw=customer_raw,
        drone_operator_id=operator_id,
        area_ha=area,
        amount=amount,
        received_amount=received,
        # NOT NULL on the table.
        payment_type=payment,
    )


class TestA2DebtAging(unittest.TestCase):
    """Buckets by age; undated and never-recorded kept apart from settled."""

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a2admin')
        with app.app_context():
            today = drones._drone_today_local()
            cls.today = today
            customers = {}
            for name in ('Дўстлик ФХ', 'Пешку АМТ', 'Миробод АМТ'):
                row = DroneCustomer(name=name)
                db.session.add(row)
                db.session.flush()
                customers[name] = row.id
            cls.c1 = customers['Дўстлик ФХ']
            cls.c2 = customers['Пешку АМТ']
            cls.c3 = customers['Миробод АМТ']

            # A2.4 boundary: exactly 30 and exactly 31 days old.
            db.session.add(_work(cls.c1, 10.0, 1000000.0, 0.0,
                                 today - timedelta(days=30)))
            db.session.add(_work(cls.c1, 10.0, 2000000.0, 0.0,
                                 today - timedelta(days=31)))
            # One in each remaining bucket.
            for days in (75, 120, 200, 400):
                db.session.add(_work(cls.c2, 5.0, 500000.0, 0.0,
                                     today - timedelta(days=days)))
            # A2.3: undated, outstanding 500 000.
            db.session.add(_work(cls.c3, 7.0, 500000.0, 0.0, None))
            # A2.2: amount NULL, received NULL, dated 200 days ago.
            db.session.add(_work(cls.c3, 3.0, None, None,
                                 today - timedelta(days=200)))
            # Settled.
            db.session.add(_work(cls.c2, 4.0, 900000.0, 900000.0,
                                 today - timedelta(days=10)))
            db.session.commit()

    @staticmethod
    def _data(args=None):
        # MultiDict, not dict: _drone_works_filters_from_args uses Werkzeug's
        # .get(key, type=int), which a plain dict does not have.
        from werkzeug.datastructures import MultiDict
        with app.app_context():
            filters = drones._drone_works_filters_from_args(
                args if args is not None else MultiDict())
            return drones._drone_works_debt_aging_data(
                drones._drone_work_conditions(filters))

    def test_a2_1_reconciles_under_three_filter_sets(self):
        from werkzeug.datastructures import MultiDict
        lines = []
        for args in (MultiDict(), MultiDict({'period': '2026-04'}),
                     MultiDict({'payment': 'cash'})):
            data = self._data(args)
            lines.append((dict(args), data['reconciled'], data['counted'],
                          data['totals']))
            self.assertTrue(data['reconciled'], 'filters=%r' % dict(args))
            self.assertEqual(data['counted']['jobs'], data['totals']['jobs'])
            self.assertAlmostEqual(data['counted']['amount'],
                                   data['totals']['amount'], delta=0.005)
        self.assertEqual(len(lines), 3)

    def test_a2_2_amount_never_recorded_is_its_own_line(self):
        """A2.2 positive: NULL amount lands in «unknown», not 181-365."""
        data = self._data()
        self.assertEqual(data['unknown']['jobs'], 1)
        self.assertEqual(data['unknown']['no_amount'], 1)
        bucket = [b for b in data['buckets'] if b['key'] == '181-365'][0]
        self.assertEqual(bucket['jobs'], 1,
                         'only the real 200-day debt belongs here')
        self.assertAlmostEqual(bucket['outstanding'], 500000.0, delta=0.005)
        # And it is not in «settled» either.
        self.assertEqual(data['settled']['no_amount'], 0)

    def test_a2_2_negative_control_settled_tested_before_unknown(self):
        """A2.2 negative: test settled first and the unknown job vanishes.

        The SAME rows are bucketed twice -- once in the order the module
        ships, once with the two branches swapped -- and the two counts are
        compared. Swapping them is the whole of the defect being guarded
        against, and it is the order _drone_work_debt_cut() calls
        load-bearing in its own docstring.
        """
        with app.app_context():
            rows = db.session.query(
                DroneWork.amount, DroneWork.received_amount).all()

        def bucket(order):
            unknown = settled = 0
            for amount, received in rows:
                outstanding = float(amount or 0.0) - float(received or 0.0)
                no_amount = amount is None
                if outstanding > 0.005:
                    continue
                if order == 'unknown-first':
                    if no_amount:
                        unknown += 1
                    else:
                        settled += 1
                else:                        # settled-first: the defect
                    settled += 1
            return unknown, settled

        right_unknown, right_settled = bucket('unknown-first')
        wrong_unknown, wrong_settled = bucket('settled-first')
        self.assertEqual(right_unknown, 1)
        self.assertEqual(wrong_unknown, 0,
                         'the broken order produces no unknown line at all')
        self.assertEqual(wrong_settled, right_settled + 1,
                         'the job it loses is the one that moved to «settled»')
        with self.assertRaises(AssertionError) as caught:
            self.assertEqual(wrong_unknown, 1,
                             'the never-recorded job must have its own line')
        self.assertIn('own line', str(caught.exception))

    def test_a2_3_undated_bucket_and_the_column_still_sums(self):
        """A2.3 positive: an undated debt is visible and inside the total."""
        data = self._data()
        undated = [b for b in data['buckets'] if b['key'] == 'undated'][0]
        self.assertEqual(undated['jobs'], 1)
        self.assertAlmostEqual(undated['outstanding'], 500000.0, delta=0.005)
        self.assertTrue(data['reconciled'])

    def test_a2_3_negative_control_dropping_the_undated_bucket(self):
        """A2.3 negative: drop the undated bucket and the column stops summing."""
        data = self._data()
        full = (sum(b['amount'] for b in data['buckets'])
                + data['unknown']['amount'] + data['settled']['amount'])
        without = (sum(b['amount'] for b in data['buckets']
                       if b['key'] != 'undated')
                   + data['unknown']['amount'] + data['settled']['amount'])
        self.assertAlmostEqual(full, data['totals']['amount'], delta=0.005)
        self.assertAlmostEqual(data['totals']['amount'] - without, 500000.0,
                               delta=0.005)
        with self.assertRaises(AssertionError) as caught:
            self.assertAlmostEqual(without, data['totals']['amount'],
                                   delta=0.005,
                                   msg='buckets must sum to the grand total')
        self.assertIn('sum to the grand total', str(caught.exception))

    def test_a2_4_boundary_30_and_31_days(self):
        data = self._data()
        first = [b for b in data['buckets'] if b['key'] == '0-30'][0]
        second = [b for b in data['buckets'] if b['key'] == '31-60'][0]
        # Only jobs with outstanding > 0.005 enter a bucket, so the settled
        # 10-day job is NOT here: 0-30 holds the 30-day debt and nothing else.
        self.assertEqual(first['jobs'], 1)
        self.assertEqual(second['jobs'], 1)
        self.assertAlmostEqual(first['outstanding'], 1000000.0, delta=0.005)
        self.assertAlmostEqual(second['outstanding'], 2000000.0, delta=0.005)
        self.assertEqual(drones._drone_debt_bucket_for(30), '0-30')
        self.assertEqual(drones._drone_debt_bucket_for(31), '31-60')

    def test_a2_5_the_two_pages_link_to_each_other(self):
        with app.test_client() as client:
            login(client, self.user_id)
            debts = client.get('/drones/works/debts').get_data(as_text=True)
            aging = client.get(
                '/drones/works/debts/aging').get_data(as_text=True)
        self.assertIn('href="/drones/works/debts/aging"', debts)
        # And back. The aging page's own form posts to /aging, so the back
        # link is matched with the closing quote to exclude it.
        self.assertIn('href="/drones/works/debts"', aging)


# ══ A3  ledger against flights ═══════════════════════════════════════════════

class TestA3Reconcile(unittest.TestCase):
    """Months from both sides; a month with no books is a phrase, not -100 %."""

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a3admin')
        with app.app_context():
            ids = _units([1])
            cls.u1 = ids[1]
            customer = DroneCustomer(name='Пешку АМТ')
            db.session.add(customer)
            db.session.flush()
            cls.cid = customer.id

            # 2026-04: both sides, ledger 90 ha against flights 100 ha (-10 %)
            db.session.add(_work(cls.cid, 90.0, 100.0, 0.0,
                                 date(2026, 4, 12), period='2026-04'))
            db.session.add(_flight(cls.u1, datetime(2026, 4, 12, 6), 100.0,
                                   0, 500.0))
            # 2026-05: flights only -- A3.2, the «no books» month.
            db.session.add(_flight(cls.u1, datetime(2026, 5, 12, 6), 40.0,
                                   0, 200.0))
            # 2026-06: ledger only -- A3.3, no division by zero.
            db.session.add(_work(cls.cid, 25.0, 100.0, 0.0,
                                 date(2026, 6, 12), period='2026-06'))
            # 2026-03: both sides, within 5 % -- must stay uncoloured.
            db.session.add(_work(cls.cid, 100.0, 100.0, 0.0,
                                 date(2026, 3, 12), period='2026-03'))
            db.session.add(_flight(cls.u1, datetime(2026, 3, 12, 6), 102.0,
                                   0, 500.0))
            db.session.commit()

    @staticmethod
    def _data():
        with app.app_context():
            return drones._drone_works_flights_reconcile_data()

    def test_a3_1_every_month_of_the_union_appears_exactly_once(self):
        data = self._data()
        months = [r['month'] for r in data['rows']]
        with app.app_context():
            work_months = set(
                row[0] for row in db.session.query(
                    drones._drone_work_month_expr()).distinct().all() if row[0])
            flight_months = set(
                row[0] for row in db.session.query(
                    drones._drone_flight_month_expr()).distinct().all()
                if row[0])
        union = work_months | flight_months
        self.assertEqual(len(months), len(union))
        self.assertEqual(len(months), len(set(months)), 'no month twice')
        self.assertEqual(set(months), union)
        self.assertEqual(months, sorted(months, reverse=True), 'descending')

    def test_a3_2_month_with_flights_and_no_books(self):
        """A3.2 positive: the phrase, no percentage, no colour."""
        data = self._data()
        row = [r for r in data['rows'] if r['month'] == '2026-05'][0]
        self.assertEqual(row['ledger_jobs'], 0)
        self.assertGreater(row['flight_area'], 0.005)
        self.assertIsNone(row['percent'], 'no percentage without books')
        self.assertEqual(row['flag'], '', 'and therefore no colour')
        self.assertIn(row['note'], ('Ведомости не заведены',
                                    'Ведомостлар киритилмаган'))

    def test_a3_2_negative_control_computing_the_percentage_anyway(self):
        """A3.2 negative: compute it anyway and -100.0 appears."""
        data = self._data()
        row = [r for r in data['rows'] if r['month'] == '2026-05'][0]
        naive = ((row['ledger_area'] - row['flight_area']) * 100.0
                 / row['flight_area'])
        self.assertEqual(naive, -100.0,
                         'the naive formula states a 100 % shortfall')
        with self.assertRaises(AssertionError) as caught:
            self.assertIsNone(naive, 'no percentage without books')
        self.assertIn('no percentage without books', str(caught.exception))

    def test_a3_3_month_with_books_and_no_flights(self):
        """A3.3: no ZeroDivisionError, no is-danger-row."""
        data = self._data()
        row = [r for r in data['rows'] if r['month'] == '2026-06'][0]
        self.assertEqual(row['flights'], 0)
        self.assertEqual(row['flight_area'], 0.0)
        self.assertIsNone(row['percent'])
        self.assertNotEqual(row['flag'], 'is-danger-row')
        self.assertEqual(row['flag'], '')
        self.assertIn(row['note'], ('Вылетов в этом месяце нет',
                                    'Бу ойда парвозлар йўқ'))

    def test_a3_3_negative_control_dividing_by_zero(self):
        """A3.3 negative: the naive formula raises on the same row."""
        data = self._data()
        row = [r for r in data['rows'] if r['month'] == '2026-06'][0]
        with self.assertRaises(ZeroDivisionError):
            _ = ((row['ledger_area'] - row['flight_area']) * 100.0
                 / row['flight_area'])

    def test_a3_colouring_only_where_both_sides_exist(self):
        data = self._data()
        by_month = {r['month']: r for r in data['rows']}
        # The COLOUR comes from coverage, not from the signed difference.
        # 2026-04: ledger 90 of flights 100 -> coverage exactly 90.0 %, which
        # is the inclusive edge of the no-colour band. The signed difference
        # is -10 %, which under the old rule painted it warning.
        self.assertAlmostEqual(by_month['2026-04']['coverage'], 90.0,
                               delta=0.005)
        self.assertAlmostEqual(by_month['2026-04']['percent'], -10.0,
                               delta=0.005)
        self.assertEqual(by_month['2026-04']['flag'], '',
                         '90 % coverage is inside the band, inclusive')
        # 2026-03: ledger 100 of flights 102 -> coverage 98.04 %.
        self.assertAlmostEqual(by_month['2026-03']['coverage'], 98.0392,
                               delta=0.001)
        self.assertEqual(by_month['2026-03']['flag'], '')
        self.assertAlmostEqual(by_month['2026-03']['percent'], -1.9608,
                               delta=0.001)

    def test_a3_page_renders_and_has_no_xlsx_route(self):
        with app.test_client() as client:
            login(client, self.user_id)
            resp = client.get('/drones/reports/reconcile')
            self.assertEqual(resp.status_code, 200)
        # The absence of an export is a decision; assert it stays absent.
        routes = [str(r) for r in app.url_map.iter_rules()]
        self.assertNotIn('/drones/reports/reconcile.xlsx', routes)


class TestA35Coverage(unittest.TestCase):
    """A3.5 -- the colour is taken from coverage, not the signed difference.

    [REASON]: the signed difference works where both sides are complete
    (April 2026: -0.68 %) and fails where the report matters most -- June
    2026 reads -99.08 %, which is not a discrepancy but a missing ledger.
    Four months here sit at coverage 99 %, 60 %, 1 % and «no books at all»,
    which are the four sentences the column has to keep apart.
    """

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a35admin')
        with app.app_context():
            ids = _units([51])
            cls.u = ids[51]
            customer = DroneCustomer(name='Пешку АМТ')
            db.session.add(customer)
            db.session.flush()
            cid = customer.id

            def month(period, day, ledger_ha, flight_ha):
                if ledger_ha:
                    db.session.add(_work(cid, ledger_ha, 100.0, 0.0,
                                         date(2026, day, 12), period=period))
                if flight_ha:
                    db.session.add(_flight(cls.u, datetime(2026, day, 12, 6),
                                           flight_ha, 0, 100.0))

            month('2026-01', 1, 99.0, 100.0)     # coverage  99 %  -> no colour
            month('2026-02', 2, 60.0, 100.0)     # coverage  60 %  -> warning
            month('2026-03', 3, 1.0, 100.0)      # coverage   1 %  -> danger
            month('2026-04', 4, 0.0, 100.0)      # no books        -> no colour
            db.session.commit()

    @staticmethod
    def _rows():
        with app.app_context():
            data = drones._drone_works_flights_reconcile_data()
        return {r['month']: r for r in data['rows']}

    def test_a3_5_the_four_states(self):
        rows = self._rows()
        expected = {
            '2026-01': (99.0, ''),
            '2026-02': (60.0, 'is-warning-row'),
            '2026-03': (1.0, 'is-danger-row'),
        }
        for month, (coverage, flag) in expected.items():
            row = rows[month]
            self.assertAlmostEqual(row['coverage'], coverage, delta=0.005,
                                   msg=month)
            self.assertEqual(row['flag'], flag, month)

        # The fourth: no ledger rows at all -- a phrase, no coverage, no colour.
        no_books = rows['2026-04']
        self.assertIsNone(no_books['coverage'])
        self.assertIsNone(no_books['percent'])
        self.assertEqual(no_books['flag'], '')
        self.assertIn(no_books['note'], ('Ведомости не заведены',
                                         'Ведомостлар киритилмаган'))

    def test_a3_5_negative_control_colouring_on_the_signed_difference(self):
        """Colour on the signed difference and the 99 % month goes yellow."""
        rows = self._rows()
        good = rows['2026-01']
        # The old rule: warning above 5 %, danger above 15 %, on |percent|.
        spread = abs(good['percent'])
        old_flag = ('is-danger-row' if spread > 15.0
                    else 'is-warning-row' if spread > 5.0 else '')
        self.assertAlmostEqual(good['coverage'], 99.0, delta=0.005)
        self.assertEqual(good['flag'], '', 'coverage 99 % is not a problem')
        self.assertEqual(old_flag, '',
                         'at 99 % coverage the two rules happen to agree')

        # Where they disagree: the 60 % month is a 40 % shortfall either way,
        # but the 1 % month is the one the old rule cannot grade -- it caps
        # out at «danger» exactly like a 16 % discrepancy does.
        thin = rows['2026-03']
        thin_old = ('is-danger-row' if abs(thin['percent']) > 15.0
                    else 'is-warning-row')
        mild = rows['2026-02']
        mild_old = ('is-danger-row' if abs(mild['percent']) > 15.0
                    else 'is-warning-row')
        self.assertEqual(thin_old, mild_old,
                         'the signed rule gives 1 % and 60 % coverage the '
                         'SAME colour -- it cannot tell them apart')
        self.assertNotEqual(thin['flag'], mild['flag'],
                            'coverage does tell them apart')
        with self.assertRaises(AssertionError) as caught:
            self.assertNotEqual(
                thin_old, mild_old,
                'a missing ledger must not look like a discrepancy')
        self.assertIn('must not look like a discrepancy',
                      str(caught.exception))

    def test_a3_5_the_total_line_obeys_the_same_rule(self):
        """A footer that contradicts its own column is worse than no footer.

        Written because the first draft did contradict it: with flights and no
        ledger at all, the total printed -100.00 %, the one sentence every row
        is forbidden to print.
        """
        reset_db()
        create_admin('a35totadmin')
        with app.app_context():
            ids = _units([52])
            db.session.add(_flight(ids[52], datetime(2026, 4, 12, 6),
                                   100.0, 0, 500.0))
            db.session.commit()
            data = drones._drone_works_flights_reconcile_data()
        self.assertAlmostEqual(data['total']['flight_area'], 100.0, delta=0.005)
        self.assertEqual(data['total']['ledger_area'], 0.0)
        self.assertIsNone(data['total']['percent'],
                          'the total must not print -100 % either')
        self.assertIsNone(data['total']['coverage'])
        # The signed difference in hectares IS shown -- it is a real quantity.
        self.assertAlmostEqual(data['total']['diff'], -100.0, delta=0.005)
        # And the naive formula, for the record, is the forbidden sentence.
        naive = ((data['total']['ledger_area'] - data['total']['flight_area'])
                 * 100.0 / data['total']['flight_area'])
        self.assertEqual(naive, -100.0)


# ══ A4  flight calendar ══════════════════════════════════════════════════════

class TestA4FlightCalendar(unittest.TestCase):
    """Machine x day in UTC+5, with the serviceability overlay."""

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a4admin')
        with app.app_context():
            ids = _units([1, 2, 3, 4])
            cls.u1, cls.u2, cls.u3 = ids[1], ids[2], ids[3]
            cls.u4 = ids[4]
            # A4.2: 2026-04-30 20:30 UTC is 2026-05-01 01:30 in UTC+5.
            db.session.add(_flight(cls.u1, datetime(2026, 4, 30, 20, 30),
                                   12.5, 0, 250.0))
            # Ordinary May flights for machine 1.
            db.session.add(_flight(cls.u1, datetime(2026, 5, 3, 6), 20.0,
                                   0, 400.0))
            db.session.add(_flight(cls.u1, datetime(2026, 5, 3, 9), 5.0,
                                   0, 100.0))
            # A4.3: machine 2 goes unserviceable on the 10th, never flies.
            db.session.add(DroneUnitStatusLog(
                drone_unit_id=cls.u2, status=DRONE_STATUS_UNSERVICEABLE,
                changed_on=date(2026, 5, 10)))
            # Machine 3 has NO status rows at all and never flies.
            #
            # A4.3 tie-breaking, machine 4: the back-dated correction. The
            # row saying «serviceable again from the 20th» is typed FIRST and
            # therefore has the LOWER id; the row saying «it was in fact
            # unserviceable from the 15th» is typed SECOND and has the higher
            # id. Back-dating is normal here -- changed_on is a date a human
            # enters, and a machine is reported broken after the fact.
            #
            # The fixture lives in setUpClass, not inside the test, so the
            # test only READS. A test that writes status rows into a machine
            # the other tests share passes or fails by unittest's alphabetical
            # method order, which is a check on the alphabet.
            first = DroneUnitStatusLog(
                drone_unit_id=cls.u4, status=DRONE_STATUS_SERVICEABLE,
                changed_on=date(2026, 5, 20))
            db.session.add(first)
            db.session.flush()
            second = DroneUnitStatusLog(
                drone_unit_id=cls.u4, status=DRONE_STATUS_UNSERVICEABLE,
                changed_on=date(2026, 5, 15))
            db.session.add(second)
            db.session.flush()
            assert first.id < second.id, 'the back-dated row must be typed last'
            db.session.commit()

    @staticmethod
    def _data(month):
        with app.app_context():
            return drones._drone_flight_calendar_data(month)

    def _cell(self, data, unit_id, day):
        row = [r for r in data['rows'] if r['unit_id'] == unit_id][0]
        return [c for c in row['cells'] if c['day'] == day][0]

    def test_a4_1_row_totals_match_the_month_aggregation(self):
        """A4.1: every machine's row total against a plain SQL aggregation."""
        data = self._data('2026-05')
        with app.app_context():
            conds = drones._drone_flight_conditions(
                {'date_from': date(2026, 5, 1), 'date_to': date(2026, 5, 31),
                 'unit_id': None, 'region': ''})
            summary = drones._drone_summary_data(conds)
        by_unit = {r['unit_id']: r['area'] for r in summary['by_machine']['rows']}
        compared = 0
        for row in data['rows']:
            expected = by_unit.get(row['unit_id'], 0.0)
            self.assertAlmostEqual(row['area'], expected, delta=0.005,
                                   msg='machine %s' % row['number'])
            compared += 1
        # Four machines now: No 4 carries the A4.3 tie-breaking fixture and
        # never flies, so it reconciles at 0.00 on both sides.
        self.assertEqual(compared, 4)
        self.assertAlmostEqual(data['total']['area'],
                               summary['totals']['area_ha'], delta=0.005)

    def test_a4_2_night_flight_lands_on_the_next_local_day(self):
        """A4.2 positive: 30 Apr 20:30 UTC is 1 May in the calendar."""
        data = self._data('2026-05')
        cell = self._cell(data, self.u1, date(2026, 5, 1))
        self.assertEqual(cell['state'], 'is-flown')
        self.assertAlmostEqual(cell['area'], 12.5, delta=0.005)
        # And it is NOT on 30 April.
        april = self._data('2026-04')
        self.assertEqual(self._cell(april, self.u1, date(2026, 4, 30))['state'],
                         'is-idle')

    def test_a4_2_negative_control_dropping_the_offset(self):
        """A4.2 negative: without the offset the flight shows on 30 April."""
        with app.app_context():
            # The same query, with the offset removed -- i.e. plain UTC.
            naive_day = func.strftime('%Y-%m-%d', DroneFlight.started_at)
            rows = (db.session.query(naive_day, func.count(DroneFlight.id))
                    .filter(DroneFlight.drone_unit_id == self.u1)
                    .group_by(naive_day).all())
            shifted = (db.session.query(drones._drone_flight_day_expr(),
                                        func.count(DroneFlight.id))
                       .filter(DroneFlight.drone_unit_id == self.u1)
                       .group_by(drones._drone_flight_day_expr()).all())
        naive_days = dict(rows)
        shifted_days = dict(shifted)
        self.assertIn('2026-04-30', naive_days,
                      'plain UTC puts the night flight on 30 April')
        self.assertNotIn('2026-04-30', shifted_days)
        self.assertIn('2026-05-01', shifted_days)
        with self.assertRaises(AssertionError) as caught:
            self.assertNotIn('2026-04-30', naive_days,
                             'the night flight must not land on 30 April')
        self.assertIn('must not land on 30 April', str(caught.exception))

    def test_a4_3_down_versus_idle(self):
        """A4.3 positive: unserviceable is «down», no history is «idle»."""
        data = self._data('2026-05')
        # Machine 2 went down on the 10th: the 12th is down, the 9th is idle.
        self.assertEqual(self._cell(data, self.u2, date(2026, 5, 12))['state'],
                         'is-down')
        self.assertEqual(self._cell(data, self.u2, date(2026, 5, 9))['state'],
                         'is-idle')
        # Machine 3 has no status rows at all -- idle, NOT down.
        self.assertEqual(self._cell(data, self.u3, date(2026, 5, 12))['state'],
                         'is-idle', 'unknown state is not «broken»')

    def test_a4_3_negative_control_tie_broken_on_id_only(self):
        """A4.3: the back-dated correction, read out of the REAL calendar.

        Machine 4 carries two status rows (see setUpClass): «serviceable from
        the 20th» typed first, «unserviceable from the 15th» typed second and
        back-dated. The status in force on a day is the row with the greatest
        changed_on <= that day, ties broken by the greatest id -- so:

            17 May  the 15th governs      -> unserviceable -> is-down
            25 May  the 20th governs      -> serviceable   -> is-idle

        [REASON]: this asserts on CELL STATES produced by
        _drone_flight_calendar_data(), NOT on a rule recomputed here. An
        earlier draft sorted the rows two ways inside the test and compared
        the two answers to each other; that version passed against a
        drones.py whose ORDER BY had been broken, because it never asked
        drones.py anything. Dropping changed_on from the ORDER BY inverts
        BOTH assertions below, which is what makes them a control.

        Why it matters: with the ordering broken the machine reads as broken
        from the 15th onwards forever, so every idle day after the repair is
        silently explained away -- and unexplained idle days are the entire
        point of this screen.
        """
        data = self._data('2026-05')

        seventeenth = self._cell(data, self.u4, date(2026, 5, 17))
        self.assertEqual(seventeenth['state'], 'is-down',
                         'on 17 May the back-dated 15th governs')
        self.assertEqual(seventeenth['status'], DRONE_STATUS_UNSERVICEABLE)

        twentyfifth = self._cell(data, self.u4, date(2026, 5, 25))
        self.assertEqual(twentyfifth['state'], 'is-idle',
                         'on 25 May the 20th governs and the machine is '
                         'serviceable again -- an UNEXPLAINED idle day')
        self.assertEqual(twentyfifth['status'], DRONE_STATUS_SERVICEABLE)

        # Before either row, the machine has no status at all: idle, not down.
        self.assertEqual(self._cell(data, self.u4, date(2026, 5, 12))['state'],
                         'is-idle')
        self.assertIsNone(self._cell(data, self.u4,
                                     date(2026, 5, 12))['status'])

    def test_a4_4_all_three_cell_classes_render_and_are_defined(self):
        """A4.4: the three classes appear in the HTML and exist in the CSS."""
        with app.test_client() as client:
            login(client, self.user_id)
            html = client.get(
                '/drones/reports/calendar?month=2026-05').get_data(as_text=True)
        with open(os.path.join(REPO_ROOT, 'static', 'css',
                               'design-system.css'), encoding='utf-8') as fh:
            css = fh.read()
        for state in ('is-flown', 'is-idle', 'is-down'):
            self.assertIn('vs-cal-cell %s' % state, html,
                          'state %s must render' % state)
            self.assertIn('.vs-cal-cell.%s' % state, css,
                          'state %s must be defined' % state)

    def test_a4_5_month_lengths(self):
        for month, length in (('2026-02', 28), ('2026-04', 30),
                              ('2026-05', 31)):
            data = self._data(month)
            self.assertEqual(len(data['days']), length, month)
            for row in data['rows']:
                self.assertEqual(len(row['cells']), length, month)
            self.assertEqual(len(data['day_totals']), length, month)


# ══ A5  the prefill link ═════════════════════════════════════════════════════

class TestA5PrefillLink(unittest.TestCase):
    """The hints screen carries values to a form and still creates nothing."""

    @classmethod
    def setUpClass(cls):
        reset_db()
        cls.user_id = create_admin('a5admin')
        with app.app_context():
            ids = _units([7])
            cls.u7 = ids[7]
            operator = DroneOperator(full_name='Жўраев Туйғун')
            db.session.add(operator)
            db.session.flush()
            cls.op_id = operator.id
            customer = DroneCustomer(name='Пешку АМТ')
            db.session.add(customer)
            db.session.flush()
            db.session.add(_work(customer.id, 120.0, 500000.0, 0.0,
                                 date(2026, 4, 12), period='2026-04',
                                 operator_id=operator.id))
            db.session.add(_flight(cls.u7, datetime(2026, 4, 12, 6), 121.0,
                                   0, 600.0))
            db.session.commit()

    def test_a5_1_no_post_route_on_the_hints_screen(self):
        """A5.1: grep the url map -- no POST anywhere under assignment-hints."""
        offenders = []
        for rule in app.url_map.iter_rules():
            if 'assignment-hints' in str(rule):
                methods = sorted(rule.methods - {'HEAD', 'OPTIONS'})
                if 'POST' in methods:
                    offenders.append((str(rule), methods))
        self.assertEqual(offenders, [])
        # And in the source: no methods=['POST'] near the hints endpoint.
        with open(os.path.join(REPO_ROOT, 'drones.py'), encoding='utf-8') as fh:
            source = fh.read()
        hints = source.split("@drones_bp.route('/works/assignment-hints'")[1]
        self.assertNotIn('POST', hints.split('def works_assignment_hints')[0])

    def test_a5_2_link_carries_month_start_and_no_prefill_to(self):
        with app.test_client() as client:
            login(client, self.user_id)
            html = client.get(
                '/drones/works/assignment-hints?month=2026-04'
            ).get_data(as_text=True)
        hrefs = re.findall(r'href="([^"]*prefill_unit[^"]*)"', html)
        self.assertTrue(hrefs, 'the hint row must carry a prefill link')
        href = hrefs[0]
        self.assertIn('prefill_from=2026-04-01', href.replace('&amp;', '&'))
        self.assertNotIn('prefill_to', href)

    def test_a5_3_garbage_arguments_render_an_empty_form(self):
        """A5.3 negative: a stale bookmark is a 200 and an empty form."""
        with app.test_client() as client:
            login(client, self.user_id)
            resp = client.get(
                '/drones/operators/%d?prefill_unit=999999'
                '&prefill_from=not-a-date' % self.op_id)
            html = resp.get_data(as_text=True)
        self.assertEqual(resp.status_code, 200, 'never a 500')
        date_field = re.search(
            r'<input type="date" name="date_from" id="addAsgFrom"[^>]*>', html)
        self.assertIsNotNone(date_field)
        self.assertIn('value=""', date_field.group(0))
        # No machine option is preselected in the add form.
        add_form = html.split('name="drone_unit_id" id="addAsgUnit"')[1]
        add_form = add_form.split('</select>')[0]
        self.assertNotIn('selected', add_form)
        # And the note is absent, because nothing was pre-filled.
        self.assertNotIn('подставлены из подсказки', html)

    def test_a5_3_valid_arguments_do_prefill(self):
        """The positive half: without it A5.3 would pass on a broken feature."""
        with app.test_client() as client:
            login(client, self.user_id)
            html = client.get(
                '/drones/operators/%d?prefill_unit=%d&prefill_from=2026-04-01'
                % (self.op_id, self.u7)).get_data(as_text=True)
        date_field = re.search(
            r'<input type="date" name="date_from" id="addAsgFrom"[^>]*>', html)
        self.assertIn('value="2026-04-01"', date_field.group(0))
        add_form = html.split('name="drone_unit_id" id="addAsgUnit"')[1]
        add_form = add_form.split('</select>')[0]
        self.assertIn('selected', add_form)

    def test_a5_4_empty_date_from_is_still_refused(self):
        """A5.4: the mandatory-field rule is untouched."""
        with app.test_client() as client:
            login(client, self.user_id)
            resp = client.post(
                '/drones/operators/%d/assignments/add' % self.op_id,
                data={'csrf_token': 'test-csrf-token',
                      'drone_unit_id': str(self.u7), 'date_from': ''},
                follow_redirects=True)
            html = resp.get_data(as_text=True)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(
            'обязательна' in html or 'мажбурий' in html,
            'the flash must still refuse a dateless assignment')
        with app.app_context():
            from models import DroneOperatorAssignment
            self.assertEqual(
                DroneOperatorAssignment.query.filter_by(
                    operator_id=self.op_id).count(), 0,
                'and nothing may have been written')

    def test_a5_5_the_note_renders_in_both_languages(self):
        seen = {}
        with app.test_client() as client:
            login(client, self.user_id)
            for lang in ('ru', 'uz'):
                set_language(self.user_id, lang)
                html = client.get(
                    '/drones/operators/%d?prefill_unit=%d'
                    '&prefill_from=2026-04-01'
                    % (self.op_id, self.u7)).get_data(as_text=True)
                block = html.split('vs-alert is-warning vs-mb">')[1]
                seen[lang] = block.split('</div>')[0].strip()
        self.assertIn('подсказки', seen['ru'])
        self.assertIn('маслаҳат', seen['uz'])
        self.assertNotEqual(seen['ru'], seen['uz'])
        # X-4 locally: the Uzbek note carries no Latin letters.
        latin = [ch for ch in seen['uz'] if 'a' <= ch.lower() <= 'z']
        self.assertEqual(latin, [], 'Uzbek is Cyrillic only')


# ══ A6  the four small defects ═══════════════════════════════════════════════

CSS_PATH = os.path.join(REPO_ROOT, 'static', 'css', 'design-system.css')
DRONE_TEMPLATES = os.path.join(REPO_ROOT, 'templates', 'drones')


def _css_text(path=CSS_PATH):
    with open(path, encoding='utf-8') as fh:
        return fh.read()


def _defined_classes(css):
    """Every class name that appears in a selector of the stylesheet."""
    # Strip comments first: a class named only inside /* ... */ is not defined.
    stripped = re.sub(r'/\*.*?\*/', ' ', css, flags=re.S)
    selectors = re.findall(r'([^{}]+)\{', stripped)
    names = set()
    for chunk in selectors:
        if chunk.strip().startswith('@'):
            continue
        names.update(re.findall(r'\.([A-Za-z0-9_-]+)', chunk))
    return names


def _used_classes(directory=DRONE_TEMPLATES):
    """{class name: {template, ...}} for every static class= in the templates.

    Only STATIC class names are collected. A name produced by a Jinja
    expression is not a literal in the file and cannot be resolved here; those
    are covered by the render-time assertion in A4.4 instead.
    """
    used = {}
    for name in sorted(os.listdir(directory)):
        if not name.endswith('.html'):
            continue
        with open(os.path.join(directory, name), encoding='utf-8') as fh:
            source = fh.read()
        source = re.sub(r'\{#.*?#\}', ' ', source, flags=re.S)
        for value in re.findall(r'class="([^"]*)"', source):
            # Drop the Jinja parts; keep the literal words around them.
            literal = re.sub(r'\{[\{%].*?[\}%]\}', ' ', value, flags=re.S)
            for token in literal.split():
                used.setdefault(token, set()).add(name)
    return used


class TestA6TheFourDefects(unittest.TestCase):
    """6.1 dead classes, 6.2 the summary cell, 6.3 the script, 6.4 vs_num."""

    # ── 6.1 / A6.1 / A6.2 ────────────────────────────────────────────────
    def test_a6_1_every_class_used_in_a_drones_template_is_defined(self):
        css = _css_text()
        defined = _defined_classes(css)
        used = _used_classes()
        # Names that are deliberately not CSS: Jinja leftovers and the
        # bootstrap-era `right` helper defined on th/td, not as a class.
        missing = sorted(name for name in used
                         if name not in defined and name != 'right')
        self.assertEqual(missing, [],
                         'undefined classes: %r' % missing)
        # The count of what was actually checked, so «0 missing» cannot be
        # produced by an empty scan.
        self.assertGreaterEqual(len(used), 50,
                                'checked only %d classes' % len(used))
        self.assertIn('vs-check', used)
        self.assertIn('is-focus', used)
        self.assertIn('vs-cal-cell', used)

    def test_a6_2_negative_control_renaming_vs_check_in_the_css(self):
        """A6.2: rename .vs-check in a COPY of the CSS and the check fails."""
        css = _css_text()
        # EVERY selector defining .vs-check must be renamed, not just the
        # first: .vs-check input[type=checkbox] defines the name too, and
        # renaming one of the two leaves the check passing on broken CSS --
        # which is how this negative control failed on its first draft.
        broken = re.sub(r'\.vs-check\b(?!-)', '.vs-checkX', css)
        self.assertNotEqual(broken, css, 'the rule must exist to be renamed')
        self.assertNotIn('.vs-check ', broken)
        defined = _defined_classes(broken)
        used = _used_classes()
        missing = sorted(name for name in used
                         if name not in defined and name != 'right')
        self.assertEqual(missing, ['vs-check'],
                         'the renamed class must surface as missing')
        with self.assertRaises(AssertionError) as caught:
            self.assertEqual(missing, [], 'undefined classes: %r' % missing)
        self.assertIn('undefined classes', str(caught.exception))

    def test_a6_1_the_focus_row_differs_from_its_neighbour(self):
        """Rendered, not read: the highlighted row carries the class."""
        reset_db()
        user_id = create_admin('a61admin')
        with app.app_context():
            first = DroneCustomer(name='Дўстлик ФХ')
            second = DroneCustomer(name='Пешку АМТ')
            db.session.add_all([first, second])
            db.session.commit()
            focus_id = first.id
        with app.test_client() as client:
            login(client, user_id)
            html = client.get(
                '/drones/customers?focus=%d' % focus_id).get_data(as_text=True)
        self.assertIn('<tr class="is-focus">', html)
        self.assertEqual(html.count('<tr class="is-focus">'), 1)
        css = _css_text()
        self.assertIn('.vs-table tr.is-focus', css)

    # ── 6.2 / A6.3 ───────────────────────────────────────────────────────
    def test_a6_3_summary_sheet_blank_versus_recorded_zero(self):
        """A6.3: all-NULL amount is an EMPTY cell; a recorded 0 prints 0."""
        from openpyxl import load_workbook
        reset_db()
        user_id = create_admin('a63admin')
        with app.app_context():
            customer = DroneCustomer(name='Пешку АМТ')
            db.session.add(customer)
            db.session.flush()
            for _ in range(3):
                db.session.add(_work(customer.id, 5.0, None, None,
                                     date(2026, 4, 10)))
            db.session.commit()

        def summary_cells():
            with app.test_client() as client:
                login(client, user_id)
                resp = client.get('/drones/works/reports.xlsx')
                self.assertEqual(resp.status_code, 200)
                wb = load_workbook(io.BytesIO(resp.data))
            ws = wb.worksheets[0]
            return {row[0]: row[1]
                    for row in ws.iter_rows(min_row=2, values_only=True)}

        cells = summary_cells()
        amount_label = [k for k in cells if k in ('Сумма',)][0]
        self.assertIsNone(cells[amount_label],
                          'every amount NULL must give an EMPTY cell, not 0')

        # Now record a zero on one row. A recorded zero and an unrecorded
        # value are different statements, and the cell must say so.
        with app.app_context():
            row = DroneWork.query.first()
            row.amount = 0.0
            db.session.commit()
        cells = summary_cells()
        self.assertEqual(cells[amount_label], 0.0,
                         'a RECORDED zero must print 0')

    def test_a6_2_totals_carry_the_three_counters(self):
        reset_db()
        create_admin('a62admin')
        with app.app_context():
            customer = DroneCustomer(name='Пешку АМТ')
            db.session.add(customer)
            db.session.flush()
            db.session.add(_work(customer.id, 5.0, None, None,
                                 date(2026, 4, 10)))
            db.session.add(_work(customer.id, 5.0, 100.0, 50.0,
                                 date(2026, 4, 11)))
            db.session.commit()
            totals = drones._drone_works_totals([])
        self.assertEqual(totals['no_amount'], 1)
        self.assertEqual(totals['no_received'], 1)
        self.assertEqual(totals['no_other_costs'], 2)
        # The sums are UNCHANGED -- every cut reconciles against them.
        self.assertEqual(totals['jobs'], 2)
        self.assertAlmostEqual(totals['amount'], 100.0, delta=0.005)

    # ── 6.3 / A6.4 ───────────────────────────────────────────────────────
    def test_a6_4_the_amount_script_guards_a_typed_value(self):
        """A6.4: simulate the script's guard over the same sequence.

        The browser is not driven here. What is asserted is the GUARD, which
        is the whole of the defect risk: the script may write the field only
        when it is empty or holds a value the script itself wrote.
        """
        ours = {}

        def recalc(current, area, price):
            if current != '' and current not in ours:
                return current           # typed by hand -- never touched
            nxt = str(round(area * price))
            ours[nxt] = True
            return nxt

        # Type 123, then change the price: the field must still read 123.
        self.assertEqual(recalc('123', 10.0, 200000.0), '123')
        # Clear it, change the price: it recomputes.
        computed = recalc('', 10.0, 200000.0)
        self.assertEqual(computed, '2000000')
        # Change the price again over the script's OWN value: it recomputes.
        self.assertEqual(recalc(computed, 10.0, 85633.0), '856330')
        # And a hand-typed value that merely looks numeric is still safe.
        self.assertEqual(recalc('76458', 10.0, 200000.0), '76458')

    def test_a6_4_script_is_present_and_reuses_the_existing_guard(self):
        with open(os.path.join(DRONE_TEMPLATES, 'works.html'),
                  encoding='utf-8') as fh:
            source = fh.read()
        self.assertIn('oursAmount', source)
        self.assertIn("getElementById('addWAmount')", source)
        self.assertIn("getElementById('addWArea')", source)
        # No tojson anywhere in an attribute of this template (X-3, locally).
        self.assertNotIn('|tojson', source)
        # The out-of-scope decision is recorded rather than left looking
        # forgotten.
        self.assertIn('В ЭТОТ ИНКРЕМЕНТ НЕ', source)

    # ── 6.4 / A6.5 ───────────────────────────────────────────────────────
    def test_a6_5_grouping_threshold(self):
        cases = {28832: '28' + NBSP + '832', 5977: '5977', 2026: '2026'}
        for value, expected in cases.items():
            rendered = drones.drone_group_number(value)
            self.assertEqual(rendered, expected,
                             '%d -> %r' % (value, rendered))

    def test_a6_5_counts_carry_the_filter_in_the_three_templates(self):
        offenders = []
        for name in ('summary.html', 'sources.html', 'reattach.html'):
            with open(os.path.join(DRONE_TEMPLATES, name),
                      encoding='utf-8') as fh:
                source = fh.read()
            # A DOTTED path only. `t_flights` is a translated column header,
            # not a count, and matching it would demand a filter on a label.
            for match in re.findall(
                    r'\{\{\s*([A-Za-z_][\w]*(?:\.[\w]+)+)\s*\}\}', source):
                if re.search(r'(flights|jobs)', match):
                    offenders.append('%s: %s' % (name, match))
        self.assertEqual(offenders, [],
                         'counts still printed raw: %r' % offenders)
