# -*- coding: utf-8 -*-
"""DRONE-FLIGHTS-VS-BOOKS-DAILY-001: посуточная сверка вылетов и ведомостей.

Синтетическая база, где каждая условность посажена нарочно:

  * вылет 2026-03-31 22:30 UTC -- это 2026-04-01 03:30 по UTC+5 и обязан
    попасть в 1 апреля (и в апрель), а вылет 2026-02-28 19:30 UTC -- в 1 марта;
  * работа с датой 23.04 в мартовской книге считается апрельской
    (месяц даты, не месяц книги -- как _drone_work_month_expr);
  * многодневная работа лежит целиком в дне начала и помечена;
  * работа без даты не входит ни в один день -- только в блок и в месяц;
  * работа с датой июля в июньской книге в окно марта--июня НЕ входит;
  * нераспознанный ник виден отдельной строкой «ник: …» в листе машин.

Run:
  python -m unittest tests.test_drone_flights_vs_books_daily_001 -v
"""
import contextlib
import hashlib
import io
import os
import sqlite3
import sys
import tempfile
import unittest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO_ROOT, 'tools'))

import drone_flights_vs_books_daily as tool  # noqa: E402

BOOK_APRIL = 'Апрель книга.xlsx'
BOOK_MARCH = 'Март книга.xlsx'


def build_db(path):
    con = sqlite3.connect(path)
    con.execute('CREATE TABLE drone_units ('
                'id INTEGER PRIMARY KEY, number INTEGER)')
    con.execute('CREATE TABLE drone_flights ('
                'id INTEGER PRIMARY KEY, drone_unit_id INTEGER, '
                'nickname_raw TEXT, started_at TEXT, area_ha REAL)')
    con.execute('CREATE TABLE drone_works ('
                'id INTEGER PRIMARY KEY, period_month TEXT, '
                'work_date_from TEXT, work_date_to TEXT, date_raw TEXT, '
                'customer_raw TEXT, area_ha REAL, subdivision_name TEXT, '
                'operator_raw TEXT, source_file TEXT, source_sheet TEXT, '
                'source_row INTEGER)')
    con.executemany(
        'INSERT INTO drone_units (id, number) VALUES (?, ?)',
        [(1, 2), (2, 10)])
    con.executemany(
        'INSERT INTO drone_flights '
        '(drone_unit_id, nickname_raw, started_at, area_ha) '
        'VALUES (?, ?, ?, ?)',
        [
            (1, None, '2026-04-10 06:00:00.000000', 2.5),
            (2, None, '2026-04-10 07:00:00.000000', 1.5),
            # 22:30 UTC = 03:30 следующего дня по UTC+5.
            (1, None, '2026-03-31 22:30:00.000000', 3.0),
            (None, '8 GardenU', '2026-05-05 05:00:00.000000', 1.2),
            # 19:30 UTC февраля = 00:30 первого марта по UTC+5.
            (2, None, '2026-02-28 19:30:00.000000', 0.8),
            # Целиком вне окна: февраль и июль по локальному времени.
            (1, None, '2026-02-20 10:00:00.000000', 9.9),
            (1, None, '2026-07-01 02:00:00.000000', 9.9),
        ])
    con.executemany(
        'INSERT INTO drone_works '
        '(period_month, work_date_from, work_date_to, date_raw, '
        ' customer_raw, area_ha, subdivision_name, operator_raw, '
        ' source_file, source_sheet, source_row) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        [
            ('2026-04', '2026-04-10', '2026-04-10', '10.04.2026',
             'Тонг фх', 4.0, 'Гарден', 'Нурали', BOOK_APRIL,
             'свод ичи', 4),
            # Апрельская дата в мартовской книге: месяц -- апрель.
            ('2026-03', '2026-04-23', '2026-04-23', '23.04.2026',
             'Сиддик Саид фх', 5.0, 'Пешку', 'Беҳзод', BOOK_MARCH,
             'свод ичи', 10),
            # Многодневная: целиком в дне начала.
            ('2026-04', '2026-04-01', '2026-04-03', '1-3.04.2026',
             'Баркамол фх', 6.0, 'Гарден', 'Нурали', BOOK_APRIL,
             'свод ичи', 5),
            # Без даты: только месяц книги.
            ('2026-05', None, None, None,
             'Оазис', 7.5, 'Сервис', 'Шахзод', 'Май книга.xlsx',
             'свод ичи', 7),
            # Вне окна: февральская книга и июльская дата в июньской книге.
            ('2026-02', None, None, None,
             'Феврал фх', 9.9, None, None, 'Февраль.xlsx', 'свод ичи', 3),
            ('2026-06', '2026-07-02', '2026-07-02', '02.07.2026',
             'Июл фх', 9.9, None, None, 'Июнь книга.xlsx', 'свод ичи', 3),
        ])
    con.commit()
    con.close()


class DailyReconcileTests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix='fl_vs_books_')
        cls.db = os.path.join(cls.tmp, 'transport.db')
        cls.out = os.path.join(cls.tmp, 'report.xlsx')
        build_db(cls.db)
        with open(cls.db, 'rb') as fh:
            cls.db_sha_before = hashlib.sha256(fh.read()).hexdigest()
        cls.stdout = io.StringIO()
        with contextlib.redirect_stdout(cls.stdout):
            cls.code = tool.main(['--db', cls.db, '--out', cls.out])
        from openpyxl import load_workbook
        cls.wb = load_workbook(cls.out)

    def summary_row(self, month):
        ws = self.wb['Сводка по месяцам']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == month:
                return row
        self.fail('month %s not found in the summary' % month)

    def daily_row(self, day):
        ws = self.wb['По дням']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == day:
                return row
        self.fail('day %s not found in the daily sheet' % day)

    def test_exit_code_zero_and_database_untouched(self):
        self.assertEqual(self.code, 0)
        with open(self.db, 'rb') as fh:
            self.assertEqual(hashlib.sha256(fh.read()).hexdigest(),
                             self.db_sha_before,
                             'read-only инструмент изменил файл базы')

    def test_console_is_ascii_only(self):
        text = self.stdout.getvalue()
        self.assertTrue(text.strip())
        self.assertTrue(all(ord(ch) < 128 for ch in text),
                        'в консоли не-ASCII: %r' % text[:120])

    def test_offset_matches_the_application(self):
        """Дубликат константы UTC+5 обязан совпадать с drones.py."""
        with open(os.path.join(REPO_ROOT, 'drones.py'),
                  encoding='utf-8') as fh:
            source = fh.read()
        self.assertIn('DRONE_DISPLAY_UTC_OFFSET = timedelta(hours=5)', source)
        self.assertEqual(tool.UTC_OFFSET_MINUTES, 300)

    def test_summary_numbers_and_notes(self):
        march = self.summary_row('2026-03')
        # Февральский по UTC вылет 19:30 -- первое марта по UTC+5.
        self.assertEqual((march[1], march[2]), (1, 0.8))
        self.assertEqual((march[3], march[4]), (0, 0.0))
        self.assertEqual(march[9], tool.NOTE_NO_BOOKS)

        april = self.summary_row('2026-04')
        self.assertEqual((april[1], april[2]), (3, 7.0))
        self.assertEqual((april[3], april[4]), (3, 15.0))
        self.assertAlmostEqual(april[7], 8.0)      # 15 - 7
        self.assertAlmostEqual(april[8], 15.0 * 100.0 / 7.0, places=2)

        may = self.summary_row('2026-05')
        self.assertEqual((may[1], may[2]), (1, 1.2))
        self.assertEqual((may[3], may[4]), (1, 7.5))
        self.assertEqual((may[5], may[6]), (1, 7.5))

        june = self.summary_row('2026-06')
        self.assertEqual((june[1], june[3]), (0, 0))
        self.assertEqual(june[9], tool.NOTE_NO_FLIGHTS)

        total = self.summary_row('ИТОГО')
        self.assertEqual(total[1], 5)
        self.assertAlmostEqual(total[2], 9.0)
        self.assertEqual(total[3], 4)
        self.assertAlmostEqual(total[4], 22.5)

    def test_boundary_flights_fall_into_local_days(self):
        self.assertEqual(self.daily_row('2026-04-01')[1:3], (1, 3.0))
        self.assertEqual(self.daily_row('2026-03-01')[1:3], (1, 0.8))
        # А в исходных днях UTC их нет.
        self.assertEqual(self.daily_row('2026-03-31')[1:3], (0, 0.0))

    def test_daily_diff_and_files_where_sides_disagree(self):
        matched = self.daily_row('2026-04-10')
        self.assertEqual(matched[3:5], (1, 4.0))
        self.assertAlmostEqual(matched[5], 0.0)
        self.assertIn(BOOK_APRIL, matched[8])

        planted = self.daily_row('2026-04-23')
        self.assertEqual(planted[1:3], (0, 0.0))
        self.assertEqual(planted[3:5], (1, 5.0))
        self.assertAlmostEqual(planted[5], 5.0)
        self.assertIn(BOOK_MARCH, planted[8],
                      'день расхождения обязан называть файл-источник')

    def test_daily_sheet_covers_every_calendar_day(self):
        ws = self.wb['По дням']
        days = [row[0] for row in ws.iter_rows(min_row=3, values_only=True)
                if row[0] and str(row[0]).count('-') == 2]
        self.assertEqual(len(days), 31 + 30 + 31 + 30)
        self.assertEqual(days[0], '2026-03-01')
        self.assertEqual(days[-1], '2026-06-30')

    def test_undated_work_is_a_block_not_a_day(self):
        ws = self.wb['По дням']
        rows = list(ws.iter_rows(values_only=True))
        for row in rows:
            # Только строки-дни (полная дата): строка «2026-05» из блока
            # внизу — месяц, а не день, и сюда не относится.
            if (row[0] and str(row[0]).startswith('2026-05')
                    and str(row[0]).count('-') == 2):
                self.assertEqual(row[3], 0,
                                 'работа без даты попала в день %s' % row[0])
        block = [i for i, row in enumerate(rows)
                 if row[0] and str(row[0]).startswith('Работы без даты')]
        self.assertEqual(len(block), 1)
        tail = rows[block[0]:]
        month_rows = [r for r in tail if r[0] == '2026-05']
        self.assertEqual(len(month_rows), 1)
        self.assertEqual(month_rows[0][1], 1)
        self.assertAlmostEqual(month_rows[0][2], 7.5)

    def test_multiday_work_lands_whole_on_its_start_day(self):
        start = self.daily_row('2026-04-01')
        self.assertEqual(start[3:5], (1, 6.0))
        self.assertEqual(start[6:8], (1, 6.0))
        self.assertEqual(self.daily_row('2026-04-02')[3], 0)
        self.assertEqual(self.daily_row('2026-04-03')[3], 0)

    def test_window_excludes_by_the_screen_month_rule(self):
        ws = self.wb['Работы']
        customers = [row[5] for row in ws.iter_rows(min_row=2,
                                                    values_only=True)]
        self.assertNotIn('Феврал фх', customers)
        self.assertNotIn('Июл фх', customers,
                         'июльская дата в июньской книге -- вне окна')
        self.assertIn('Сиддик Саид фх', customers)

    def test_provenance_sheet_names_file_sheet_row(self):
        ws = self.wb['Работы']
        by_customer = dict((row[5], row) for row in ws.iter_rows(
            min_row=2, values_only=True))
        moved = by_customer['Сиддик Саид фх']
        self.assertEqual(moved[1], '2026-04')
        self.assertEqual((moved[9], moved[10], moved[11]),
                         (BOOK_MARCH, 'свод ичи', 10))
        multiday = by_customer['Баркамол фх']
        self.assertEqual(multiday[12], 'да')

    def test_day_by_machine_sheet(self):
        ws = self.wb['Вылеты по машинам']
        cells = dict(((row[0], row[1]), (row[2], row[3]))
                     for row in ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(cells[('2026-04-10', '№2')], (1, 2.5))
        self.assertEqual(cells[('2026-04-10', '№10')], (1, 1.5))
        self.assertEqual(cells[('2026-04-01', '№2')], (1, 3.0))
        self.assertEqual(cells[('2026-03-01', '№10')], (1, 0.8))
        self.assertEqual(cells[('2026-05-05', 'ник: 8 GardenU')], (1, 1.2))
        self.assertNotIn(('2026-02-20', '№2'), cells)

    def test_missing_database_exits_two(self):
        code = tool.main(['--db', os.path.join(self.tmp, 'absent.db'),
                          '--out', os.path.join(self.tmp, 'x.xlsx')])
        self.assertEqual(code, 2)

    def test_bad_month_arguments_exit_one(self):
        self.assertEqual(tool.main(['--db', self.db, '--month-from', '2026',
                                    '--out', self.out]), 1)
        self.assertEqual(tool.main(['--db', self.db,
                                    '--month-from', '2026-05',
                                    '--month-to', '2026-03',
                                    '--out', self.out]), 1)


if __name__ == '__main__':
    unittest.main()
