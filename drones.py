# -*- coding: utf-8 -*-
"""drones.py -- drones module blueprint.

Route map:
  GET  /drones/                 -- flight list: server-side pagination
                                   (50/page), filters by date range, machine
                                   and region. Correct on an empty table.
  GET  /drones/units            -- the 15 machines with their nickname
                                   aliases grouped.
  POST /drones/api/flight_sync  -- ingest endpoint for raw DJI flight
                                   payloads (DRONE-002). Token-authenticated
                                   via DRONE_API_TOKEN (deny-by-default),
                                   CSRF-exempt in app.py:is_csrf_exempt().

Browser routes are decorated with @module_required('drones'): the admin-UI
permission toggles are enforced at the route, not only at the sidebar link.
The ingest endpoint is a machine API authenticated by token in the request
body (same convention as the Topaz fuel sync), not by a browser session.

Out of scope here: the Playwright collector (DRONE-003), reporting screens
and Excel (DRONE-004).
"""

import io
import json
import re

from datetime import datetime, timedelta

from flask import (Blueprint, abort, current_app, flash, jsonify, redirect,
                   render_template, request, send_file, url_for, g)
from flask_login import current_user
from sqlalchemy import and_, case, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import aliased, joinedload

from ingest_common import verify_api_token, extract_token
from models import (
    db,
    DRONE_BATTERY_CONDITIONS,
    DRONE_CUSTOMER_EXTERNAL,
    DRONE_CUSTOMER_INTERNAL,
    DRONE_CUSTOMER_TYPES,
    DRONE_STATUSES,
    DRONE_STATUS_SERVICEABLE,
    DRONE_STATUS_UNSERVICEABLE,
    DRONE_WORK_PAYMENT_CASH,
    DRONE_WORK_PAYMENT_INTERNAL,
    DRONE_WORK_PAYMENT_TRANSFER,
    DRONE_WORK_PAYMENT_TYPES,
    DRONE_WORK_PAYMENT_UNKNOWN,
    DRONE_WORK_PRICE_SUGGESTIONS,
    DRONE_RECEIVED_KIND_OPERATOR_DUE,
    DRONE_RECEIVED_KIND_RECEIVED,
    DroneBattery,
    DroneCustomer,
    DroneCustomerAlias,
    DroneOperator,
    DroneOperatorAssignment,
    DroneUnit,
    DroneUnitStatusLog,
    DroneNickname,
    DroneFlight,
    DroneReattachRun,
    DroneSyncLog,
    DroneWork,
    module_required,
)

drones_bp = Blueprint('drones', __name__, url_prefix='/drones')

# [REASON]: 150 flights a day in season, peak 476 (drones track) --
# server-side pagination is mandatory from the first line; 50 matches the
# SmartFarm page size the operators already know.
DRONE_PAGE_SIZE = 50

# [REASON]: flight timestamps are stored in UTC (unix seconds in the DJI
# payload); the business reads them in UTC+5. The offset is applied at
# render time only -- storage and filtering stay consistent in the DB.
DRONE_DISPLAY_UTC_OFFSET = timedelta(hours=5)


# ─── UI-NUMBER-FORMAT-001: thousands grouping ────────────────────────────────

# [REASON]: money and counts rendered unseparated -- 2491814505. At that length
# nobody reads the magnitude, and these are reports people decide from.
#
# U+00A0, not a plain space: a plain space lets a browser break the line
# between «2 491» and «814 505», and half a number on each line is worse than
# no grouping at all.
DRONE_GROUP_SEPARATOR = '\u00a0'

# [REASON]: THE YEAR RULE, and it is structural rather than a convention every
# template author has to remember. A number is grouped only when its integer
# part is FIVE digits or longer, so 2026 can never become «2 026» no matter
# where the filter is applied -- and neither can 906, 0 or any other short
# count. Grouping a year is the classic failure of exactly this change; making
# it impossible beats forbidding it in a comment. The cost is that a genuine
# 5 000 so'm renders ungrouped, which is readable either way.
DRONE_GROUP_MIN_DIGITS = 5

_DRONE_NUMERIC_RE = re.compile(r'^-?\d+(\.\d+)?$')


@drones_bp.app_template_filter('vs_num')
def drone_group_number(value):
    """Group an integer part by three with a non-breaking space.

    2491814505 -> '2 491 814 505'   (with U+00A0)
    '15893.64' -> '15 893.64'       decimal separator stays the POINT
    906        -> '906'
    2026       -> '2026'            a year is never grouped -- see above
    '2026-04'  -> '2026-04'         not numeric, passes through untouched
    None       -> None              the template renders its own dash

    Published from the blueprint with @app_template_filter, so app.py is not
    touched -- the same route DRONE-FLEET-001 used for the dashboard tile.
    """
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = ('%d' % value) if value.is_integer() else repr(value)
    elif isinstance(value, str):
        text = value.strip()
        # [REASON]: anything that is not purely numeric passes through
        # UNCHANGED. '2026-04' is a period, not a number; a nickname, a serial
        # and a phone number are identifiers. The filter must be safe to apply
        # to a value that turns out not to be a number, or every call site
        # becomes a decision.
        if not _DRONE_NUMERIC_RE.match(text):
            return value
    else:
        return value

    sign = '-' if text.startswith('-') else ''
    body = text[1:] if sign else text
    whole, _, fraction = body.partition('.')
    if len(whole) < DRONE_GROUP_MIN_DIGITS:
        return value if isinstance(value, str) else text
    groups = []
    while len(whole) > 3:
        groups.insert(0, whole[-3:])
        whole = whole[:-3]
    groups.insert(0, whole)
    grouped = sign + DRONE_GROUP_SEPARATOR.join(groups)
    return grouped + ('.' + fraction if fraction else '')


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _drone_lang():
    lang = getattr(g, 'lang', None)
    if not lang and getattr(current_user, 'is_authenticated', False):
        lang = getattr(current_user, 'language', None)
    return 'ru' if lang == 'ru' else 'uz'


def _drone_t(uz_text, ru_text):
    # Same pattern as _spare_t (spare_parts.py): route-level bilingual helper
    # for strings that are built outside templates.
    return ru_text if _drone_lang() == 'ru' else uz_text


def _drone_parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), '%Y-%m-%d').date()
    except (TypeError, ValueError, AttributeError):
        return None


def _drone_fmt_dt(dt):
    """UTC-stored datetime -> display string in UTC+5, em dash when empty."""
    if not dt:
        return '—'
    return (dt + DRONE_DISPLAY_UTC_OFFSET).strftime('%d.%m.%Y %H:%M')


def _drone_today_local():
    """Today's date as the operators see it (UTC+5), not as the server sees it."""
    return (datetime.utcnow() + DRONE_DISPLAY_UTC_OFFSET).date()


def _drone_fmt_date(value):
    return value.strftime('%d.%m.%Y') if value else '—'


def _drone_assignment_is_current(row, today):
    return (row.date_from is not None and row.date_from <= today
            and (row.date_to is None or row.date_to >= today))


def _drone_usage_labels():
    return {
        0: _drone_t('Пуркаш', 'Опрыскивание'),
        1: _drone_t('Экиш', 'Сев'),
    }


# ─── Routes (read-only) ───────────────────────────────────────────────────────

@drones_bp.route('/')
@module_required('drones')
def index():
    page = request.args.get('page', 1, type=int) or 1
    if page < 1:
        page = 1
    date_from_s = (request.args.get('date_from') or '').strip()
    date_to_s = (request.args.get('date_to') or '').strip()
    unit_id = request.args.get('unit_id', type=int)
    region = (request.args.get('region') or '').strip()

    date_from = _drone_parse_date(date_from_s)
    date_to = _drone_parse_date(date_to_s)

    q = DroneFlight.query
    # [REASON]: the operator picks dates in local time (UTC+5) while
    # started_at is stored in UTC, so the day boundaries are shifted by the
    # display offset -- a flight at 02:00 local on the 20th (21:00 UTC on
    # the 19th) belongs to the 20th for the person filtering.
    if date_from:
        q = q.filter(DroneFlight.started_at >=
                     datetime.combine(date_from, datetime.min.time())
                     - DRONE_DISPLAY_UTC_OFFSET)
    if date_to:
        q = q.filter(DroneFlight.started_at <=
                     datetime.combine(date_to, datetime.max.time())
                     - DRONE_DISPLAY_UTC_OFFSET)
    if unit_id:
        q = q.filter(DroneFlight.drone_unit_id == unit_id)
    if region:
        q = q.filter(DroneFlight.region == region)

    total = q.count()
    pages = max(1, (total + DRONE_PAGE_SIZE - 1) // DRONE_PAGE_SIZE)
    if page > pages:
        page = pages
    flights = (q.options(joinedload(DroneFlight.drone_unit))
                .order_by(DroneFlight.started_at.desc())
                .offset((page - 1) * DRONE_PAGE_SIZE)
                .limit(DRONE_PAGE_SIZE)
                .all())

    units = DroneUnit.query.order_by(DroneUnit.number).all()
    regions = [row[0] for row in
               db.session.query(DroneFlight.region)
               .filter(DroneFlight.region.isnot(None))
               .distinct().order_by(DroneFlight.region).all()]

    # Filter args echoed into pagination links, only the ones actually set.
    filter_args = {}
    if date_from_s:
        filter_args['date_from'] = date_from_s
    if date_to_s:
        filter_args['date_to'] = date_to_s
    if unit_id:
        filter_args['unit_id'] = unit_id
    if region:
        filter_args['region'] = region

    return render_template(
        'drones/list.html',
        flights=flights,
        total=total,
        page=page,
        pages=pages,
        units=units,
        regions=regions,
        # [REASON]: labels are for display only. The <option> value and the
        # filter comparison stay the RAW stored string, so
        # DroneFlight.region == filters['region'] keeps working unchanged and
        # links already in circulation keep resolving.
        region_labels=_drone_region_label_map(regions),
        filters={'date_from': date_from_s, 'date_to': date_to_s,
                 'unit_id': unit_id, 'region': region},
        filter_args=filter_args,
        fmt_dt=_drone_fmt_dt,
        usage_labels=_drone_usage_labels(),
    )


# ─── Ingest endpoint (DRONE-002) ──────────────────────────────────────────────

# [REASON]: the collector chunks its uploads; a hard cap keeps one request
# from ballooning memory and the SQLite write transaction. 1000 > the
# collector's default chunk of 500 with headroom.
DRONE_SYNC_MAX_BATCH = 1000

DRONE_SYNC_KINDS = ('backfill', 'incremental', 'replay')

# Error text on the sync log keeps at most this many per-row messages.
DRONE_SYNC_MAX_ERRORS_LOGGED = 50


def _drone_normalize_nickname(nickname):
    """Lowercase, ALL whitespace removed -- must match the migration seed."""
    return ''.join(nickname.lower().split())


def _drone_region_from_location(location):
    """Region from the reverse-geocoded address, rule verified on 10 385 rows.

    [REASON]: split on commas and take the LAST segment containing the word
    'Region'; when none does, use 'Tashkent' if a segment equals 'Tashkent';
    otherwise NULL. Naive 'second-to-last segment' splitting is wrong: in
    roughly a third of the rows that segment is a postal code. Measured
    coverage on the sample is 100 % (10 381 by the Region rule + 4 by the
    Tashkent fallback). See docs/DRONES_DOMAIN.md section 4.
    """
    if not location or not isinstance(location, str):
        return None
    segments = [seg.strip() for seg in location.split(',')]
    for seg in reversed(segments):
        if 'Region' in seg:
            return seg
    for seg in segments:
        if seg == 'Tashkent':
            return 'Tashkent'
    return None


# [REASON]: DRONE-007 -- stems of the stored region strings, i.e. the value
# lowercased with a trailing ' region' removed. DJI sends the region in Latin
# ('Bukhara Region') and it landed unchanged on an Uzbek Cyrillic screen and
# in Excel. This map is DISPLAY ONLY: nothing here is ever written to
# drone_flights.region, and every filter keeps comparing the raw stored value,
# so an existing bookmark or export link keeps working.
#
# Confirmed present in production data: bukhara, navoiy, qashqadaryo,
# samarqand, jizzakh, tashkent. The remaining Uzbek regions are listed too, so
# the first flight from a new region does not appear raw among translated
# neighbours. Uzbek is Cyrillic, as everywhere in this module.
DRONE_REGION_LABELS = {
    'bukhara':        ('Бухоро вилояти', 'Бухарская область'),
    'navoiy':         ('Навоий вилояти', 'Навоийская область'),
    'qashqadaryo':    ('Қашқадарё вилояти', 'Кашкадарьинская область'),
    'samarqand':      ('Самарқанд вилояти', 'Самаркандская область'),
    'jizzakh':        ('Жиззах вилояти', 'Джизакская область'),
    'tashkent':       ('Тошкент', 'Ташкент'),
    'andijan':        ('Андижон вилояти', 'Андижанская область'),
    'fergana':        ('Фарғона вилояти', 'Ферганская область'),
    'namangan':       ('Наманган вилояти', 'Наманганская область'),
    'sirdaryo':       ('Сирдарё вилояти', 'Сырдарьинская область'),
    'surxondaryo':    ('Сурхондарё вилояти', 'Сурхандарьинская область'),
    'xorazm':         ('Хоразм вилояти', 'Хорезмская область'),
    'karakalpakstan': ('Қорақалпоғистон Республикаси',
                       'Республика Каракалпакстан'),
}

DRONE_REGION_SUFFIX = ' region'


def _drone_region_label(region):
    """Display label for a stored region value, in the current language.

    [REASON]: an UNKNOWN region is returned VERBATIM, never blank and never a
    placeholder. A region that fell out of this map must still be readable and
    still be recognisable as the string the filter works on; rendering it as
    an empty cell would look exactly like the NULL-region case, which has its
    own explicit line and its own wording.
    """
    if not region:
        return region
    stem = region.strip().lower()
    if stem.endswith(DRONE_REGION_SUFFIX):
        stem = stem[:-len(DRONE_REGION_SUFFIX)].strip()
    labels = DRONE_REGION_LABELS.get(stem)
    if labels is None:
        return region
    return _drone_t(labels[0], labels[1])


def _drone_region_label_map(regions):
    """{raw value: label} for a list of raw region strings.

    Handed to the templates instead of registering a global Jinja filter --
    a global would have to be registered in app.py, which this change does
    not touch.
    """
    return {region: _drone_region_label(region) for region in regions}


def _drone_num(value):
    """float(value) or None -- missing and non-numeric both become None."""
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _drone_int(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _drone_utc(value):
    """Unix seconds -> naive UTC datetime, None when missing or non-numeric."""
    ts = _drone_num(value)
    if ts is None:
        return None
    try:
        return datetime.utcfromtimestamp(ts)
    except (OverflowError, OSError, ValueError):
        return None


def _drone_text(value, limit):
    if value is None:
        return None
    text = str(value)
    return text[:limit] if text else None


def _drone_nickname_maps():
    """Build the two-step resolution maps in one query.

    Resolution order (drones track): exact match on nickname; on a miss,
    match on normalized; if the normalized match points at more than one
    drone_unit_id, the nickname is UNRESOLVED rather than guessed.
    """
    exact = {}
    normalized = {}
    # [REASON]: DRONE-007 -- the units screen offers a per-alias disable
    # toggle. Without this filter the toggle would change nothing at ingest
    # and the UI would be lying about what it does. isnot(False), never
    # == True: in SQLite the column is BOOLEAN DEFAULT 1 and a legacy or
    # hand-inserted row can hold NULL, which == True would silently drop,
    # turning known spellings into unresolved flights. NULL therefore counts
    # as active, which is the behaviour that was there before. All twenty
    # rows seeded by migrate_drones_foundation_001.py carry is_active = 1, so
    # nothing changes on existing data.
    for nick in DroneNickname.query.filter(
            DroneNickname.is_active.isnot(False)).all():
        exact[nick.nickname] = nick.drone_unit_id
        normalized.setdefault(nick.normalized, set()).add(nick.drone_unit_id)
    return exact, normalized


def _drone_resolve_unit(nickname, exact, normalized):
    if not nickname:
        return None
    unit_id = exact.get(nickname)
    if unit_id is not None:
        return unit_id
    candidates = normalized.get(_drone_normalize_nickname(nickname))
    if candidates and len(candidates) == 1:
        return next(iter(candidates))
    return None


@drones_bp.route('/api/flight_sync', methods=['POST'])
def api_flight_sync():
    """Ingest a batch of raw DJI flight payloads.

    Counter semantics (also stated on the DroneSyncLog model): every seen
    flight lands in exactly one of new / duplicates / errors, so

        seen = new + duplicates + errors

    `new` counts every row that entered the table -- it is the answer to
    "did data arrive". `unresolved` is a SUBSET of `new`, not a sibling
    bucket (unresolved <= new): rows stored with an unrecognised nickname
    and drone_unit_id NULL. A batch-level failure rolls everything back
    and reports new = duplicates = unresolved = 0, errors = seen.
    """
    payload = request.get_json(force=True, silent=True)
    token = extract_token(payload)
    # [REASON]: a missing DRONE_API_TOKEN must DENY, never accept --
    # verify_api_token already implements that contract; no fallback here.
    if not verify_api_token(token, current_app.config.get('DRONE_API_TOKEN')):
        return jsonify(error='unauthorized'), 401

    flights = payload.get('flights')
    if not isinstance(flights, list):
        return jsonify(error='flights must be a list'), 400
    if len(flights) > DRONE_SYNC_MAX_BATCH:
        return jsonify(error='batch too large: %d flights, the cap is %d '
                             'per request -- send chunks'
                             % (len(flights), DRONE_SYNC_MAX_BATCH)), 413
    kind = payload.get('kind')
    if kind not in DRONE_SYNC_KINDS:
        return jsonify(error='kind must be one of %s'
                             % ', '.join(DRONE_SYNC_KINDS)), 400

    log = DroneSyncLog(
        kind=kind,
        period_from=_drone_parse_date(payload.get('period_from')),
        period_to=_drone_parse_date(payload.get('period_to')),
        started_at=datetime.utcnow(),
        status='running',
        records_seen=len(flights),
        source_ip=request.remote_addr,
    )
    db.session.add(log)
    # Commit the running row first: a crash mid-batch must leave a visible
    # 'running' log, not nothing.
    db.session.commit()

    new = duplicates = unresolved = errors = 0
    # [REASON]: counted for the batch-level warning below only. It is NOT a
    # sixth DroneSyncLog counter and NOT part of the JSON response: the log's
    # five counters carry a documented invariant and the response shape is
    # what the collector reads, so the signal channel here is the app log.
    new_without_region = 0
    error_lines = []

    try:
        exact, normalized = _drone_nickname_maps()

        # [REASON]: dedup on dji_flight_id -- the existing ids for the whole
        # batch are read in ONE query; the per-row IntegrityError guard below
        # stays as well, because two collector runs may overlap.
        candidate_ids = []
        for flight in flights:
            if isinstance(flight, dict):
                fid = _drone_int(flight.get('id'))
                if fid is not None:
                    candidate_ids.append(fid)
        existing = set()
        if candidate_ids:
            rows = (db.session.query(DroneFlight.dji_flight_id)
                    .filter(DroneFlight.dji_flight_id.in_(candidate_ids))
                    .all())
            existing = {row[0] for row in rows}

        for position, flight in enumerate(flights):
            try:
                if not isinstance(flight, dict):
                    raise ValueError('flight #%d is not an object' % position)
                fid = _drone_int(flight.get('id'))
                if fid is None:
                    raise ValueError('flight #%d has no numeric id'
                                     % position)
                if fid in existing:
                    duplicates += 1
                    continue
                started_at = _drone_utc(flight.get('start_timestamp'))
                if started_at is None:
                    raise ValueError('flight id=%d has no valid '
                                     'start_timestamp' % fid)

                nickname = _drone_text(flight.get('nickname'), 100)
                unit_id = _drone_resolve_unit(nickname, exact, normalized)

                area = _drone_num(flight.get('new_work_area'))
                spray = _drone_num(flight.get('spray_usage'))
                sow = _drone_num(flight.get('sow_usage'))
                manual = flight.get('manual_mode')
                region = _drone_region_from_location(flight.get('location'))

                row = DroneFlight(
                    dji_flight_id=fid,
                    # [REASON]: an unknown nickname never rejects a flight --
                    # a hardcoded nickname dictionary cost the previous
                    # system 2 036 flights, and a rejected flight is
                    # unrecoverable once the DJI history window rolls past
                    # it. Unresolved rows are stored with drone_unit_id NULL
                    # and nickname_raw filled, and can be re-attributed after
                    # the alias map learns the spelling.
                    drone_unit_id=unit_id,
                    nickname_raw=nickname,
                    serial_number=_drone_text(flight.get('serial_number'), 50),
                    flyer_name=_drone_text(flight.get('flyer_name'), 100),
                    team_name=_drone_text(flight.get('team_name'), 100),
                    started_at=started_at,
                    finished_at=_drone_utc(flight.get('end_timestamp')),
                    work_seconds=_drone_int(flight.get('work_time_seconds')),
                    # Units, verified on the sample: new_work_area is m2,
                    # spray_usage is ml, sow_usage is GRAMS.
                    area_ha=(area / 10000.0) if area is not None else 0,
                    spray_liters=(spray / 1000.0) if spray is not None else None,
                    sow_kg=(sow / 1000.0) if sow is not None else None,
                    usage_type=_drone_int(flight.get('usage_type')),
                    mode_name=_drone_int(flight.get('mode_name')),
                    manual_mode=bool(manual) if manual is not None else None,
                    work_speed=_drone_num(flight.get('work_speed')),
                    spray_width=_drone_num(flight.get('spray_width')),
                    radar_height=_drone_num(flight.get('radar_height')),
                    lat=_drone_num(flight.get('lat')),
                    lng=_drone_num(flight.get('lng')),
                    location_text=_drone_text(flight.get('location'), 500),
                    # [REASON]: region is computed at ingest and stored, so
                    # reports never re-parse tens of thousands of address
                    # strings; location_text keeps the source verbatim.
                    region=region,
                    raw_json=json.dumps(flight, ensure_ascii=False),
                    sync_log_id=log.id,
                    ingested_at=datetime.utcnow(),
                )
                try:
                    with db.session.begin_nested():
                        db.session.add(row)
                except IntegrityError:
                    # Same dji_flight_id landed from an overlapping run (or
                    # twice in one batch) between the dedup read and here.
                    duplicates += 1
                    continue
                existing.add(fid)
                # [REASON]: `new` counts every row that entered the table,
                # whether or not its nickname resolved -- it is the answer
                # to "did data arrive". `unresolved` is a SUBSET of `new`,
                # not a sibling bucket: an unattributed flight is still a
                # stored flight, and it becomes attributed later without
                # the log changing.
                new += 1
                if unit_id is None:
                    unresolved += 1
                if region is None:
                    new_without_region += 1
            except Exception as exc:
                # [REASON]: a row that fails to parse increments errors and
                # is recorded, but never aborts the batch -- the rest of the
                # chunk must still land.
                errors += 1
                if len(error_lines) < DRONE_SYNC_MAX_ERRORS_LOGGED:
                    error_lines.append(str(exc))

        # [REASON]: DRONE-007 -- the region parser looks for the ENGLISH word
        # 'Region' in the reverse-geocoded address. If the DJI account
        # language ever changes, the token disappears, the column goes NULL
        # for every new row and nothing else about the run looks different:
        # the flights arrive, the counters are healthy, and the by-region
        # report quietly empties out. A whole batch with no region at all is
        # the shape of that failure, so it is said out loud. The threshold is
        # ALL of them, not some: a handful of addresses without the token is
        # ordinary, and a warning that fires on ordinary data stops being
        # read. The parsing rule itself is NOT changed -- it is verified at
        # 100 % coverage on 10 385 rows, and changing it on a guess would
        # corrupt a working column.
        if new > 0 and new_without_region == new:
            current_app.logger.warning(
                'DRONE INGEST: all %d newly stored flight(s) in this batch '
                'have region NULL. That is what a change of the DJI account '
                'language looks like -- the address no longer contains the '
                'token the region parser matches on. Check the SmartFarm '
                'account language; the parsing rule was not changed and is '
                'verified at 100%% coverage on 10385 rows. Sync log id %s.',
                new, log.id)

        log.records_new = new
        log.records_duplicate = duplicates
        log.records_unresolved = unresolved
        log.records_error = errors
        log.error_text = '\n'.join(error_lines) if error_lines else None
        log.status = 'ok'
        log.finished_at = datetime.utcnow()
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        # [REASON]: the rollback above discarded every row of this batch,
        # so the counters must say so. Recording the in-flight values here
        # would leave a log claiming N new rows that do not exist -- worse
        # than no log at all. records_seen stays: what was submitted is
        # still true.
        log.records_new = 0
        log.records_duplicate = 0
        log.records_unresolved = 0
        log.records_error = len(flights)
        log.status = 'error'
        log.error_text = ('batch rolled back, nothing was stored: %s' % exc)
        log.finished_at = datetime.utcnow()
        db.session.commit()
        return jsonify(status='error', log_id=log.id, error=str(exc)), 500

    return jsonify(status='ok', log_id=log.id, seen=len(flights), new=new,
                   duplicates=duplicates, unresolved=unresolved,
                   errors=errors)


@drones_bp.route('/units')
@module_required('drones')
def units():
    unit_rows = (DroneUnit.query
                 .options(joinedload(DroneUnit.organization))
                 .order_by(DroneUnit.number).all())
    # [REASON]: one query for all aliases, grouped in Python -- no per-unit
    # N+1 through the lazy='dynamic' relationship.
    nicknames_by_unit = {}
    for nick in DroneNickname.query.order_by(DroneNickname.id).all():
        nicknames_by_unit.setdefault(nick.drone_unit_id, []).append(nick)

    # DRONE-FLEET-001: the same one-query-then-group-in-Python rule for the
    # battery count as for the aliases above.
    battery_counts = dict(
        db.session.query(DroneBattery.drone_unit_id,
                         func.count(DroneBattery.id))
        .group_by(DroneBattery.drone_unit_id).all())

    return render_template(
        'drones/units.html',
        units=unit_rows,
        nicknames_by_unit=nicknames_by_unit,
        current_status=_drone_current_status_rows(),
        status_labels=_drone_status_labels(),
        battery_counts=battery_counts,
    )


# ─── Alias management (DRONE-007) ─────────────────────────────────────────────
#
# [REASON]: DroneNickname was writable only by migration until now, so the
# forty historical 2025 spellings sitting in drone_flights.nickname_raw could
# not be mapped without a code change. Two POST routes -- add and update --
# cover the whole surface the owner needs: a new spelling, a rename, a move to
# another machine, and a disable.
#
# There is deliberately NO delete route. Deactivation is the reversible
# operation and it preserves the record that a spelling ever existed, which is
# what makes an old attribution explainable years later; deleting the row
# loses that permanently and gains nothing, because an inactive alias already
# stops resolving at ingest (_drone_nickname_maps filters on is_active).


@drones_bp.app_context_processor
def _drone_dashboard_context():
    """Expose a LAZY fleet-status callable to every template (DRONE-FLEET-001).

    [REASON]: registered from the BLUEPRINT rather than added to the dashboard
    route. The main dashboard belongs to another track and app.py is a shared
    file; this keeps the cross-track change down to one additive tile in
    templates/index.html and nothing else.

    [REASON]: the value is a FUNCTION, not the numbers. An app-wide context
    processor runs on every render in every module, and computing the counts
    eagerly would put two queries on every page of the whole application to
    feed one tile on one page. Nothing runs until a template calls it.
    """
    def drone_fleet_status():
        if not getattr(current_user, 'is_authenticated', False):
            return None
        if not current_user.has_module_access('drones'):
            return None
        try:
            serviceable, unserviceable, unknown = _drone_fleet_status_counts()
        except Exception:
            # [REASON]: a drones table that is missing or mid-migration must
            # not take the whole dashboard down with it -- but it must not be
            # reported as "0 broken machines" either, which is a false and
            # comfortable answer. The rollback is required: a failed statement
            # leaves the session unusable for the rest of the page.
            db.session.rollback()
            current_app.logger.exception('drone fleet status tile failed')
            return {'error': True}
        return {
            'error': False,
            'serviceable': serviceable,
            'unserviceable': unserviceable,
            'unknown': unknown,
            'total': serviceable + unserviceable + unknown,
        }

    return {'drone_fleet_status': drone_fleet_status}


def _drone_units_redirect():
    return redirect(url_for('drones.units'))


def _drone_nickname_ambiguous_with(normalized, drone_unit_id,
                                   exclude_id=None):
    """An ACTIVE alias normalising the same way but pointing elsewhere.

    [REASON]: `normalized` is deliberately not unique ('8 Garden U' and
    '8 GardenU' both normalise to '8gardenu'), and that is fine while every
    row of a normalized group points at one machine. When a group straddles
    two machines, _drone_resolve_unit() stops using the normalized fallback
    for the whole group and returns None rather than guessing -- so the save
    still happens, but the operator has to be told that resolution for these
    spellings is now exact-match only.
    """
    query = DroneNickname.query.filter(
        DroneNickname.normalized == normalized,
        DroneNickname.drone_unit_id != drone_unit_id,
        DroneNickname.is_active.isnot(False))
    if exclude_id is not None:
        query = query.filter(DroneNickname.id != exclude_id)
    return query.first()


def _drone_flash_ambiguity(other):
    unit = DroneUnit.query.get(other.drone_unit_id)
    number = unit.number if unit else other.drone_unit_id
    flash(_drone_t(
        'Сақланди, аммо огоҳлантириш: «%s» ёзилиши № %s машинага '
        'ишора қилади ва нормаллаштирилгандан кейин худди шундай бўлади. '
        'Бу ёзилишлар гуруҳи учун нормаллаштирилган мослик энди ноаниқ — '
        'аниқлаш фақат аниқ мослик бўйича ишлайди.',
        'Сохранено, но внимание: написание «%s» указывает на машину № %s и '
        'после нормализации совпадает с этим. Для этой группы написаний '
        'нормализованное сопоставление стало неоднозначным — распознавание '
        'будет работать только по точному совпадению.')
        % (other.nickname, number), 'warning')


def _drone_flash_duplicate(nickname):
    """Name the machine the existing spelling already points at, not a 500."""
    existing = DroneNickname.query.filter(
        DroneNickname.nickname == nickname).first()
    if existing is None:
        flash(_drone_t('«%s» ёзилиши сақланмади: у аллақачон мавжуд.',
                       'Написание «%s» не сохранено: оно уже существует.')
              % nickname, 'danger')
        return
    unit = DroneUnit.query.get(existing.drone_unit_id)
    number = unit.number if unit else existing.drone_unit_id
    flash(_drone_t(
        '«%s» ёзилиши аллақачон мавжуд ва № %s машинага бириктирилган. '
        'Ҳеч нарса ўзгартирилмади.',
        'Написание «%s» уже существует и привязано к машине № %s. '
        'Ничего не изменено.') % (nickname, number), 'danger')


@drones_bp.route('/units/nicknames/add', methods=['POST'])
@module_required('drones')
def units_nickname_add():
    """Add one raw DJI spelling and attach it to a machine."""
    # The admin-only decorators live inside create_app() and are not
    # importable from a blueprint; spare_parts.py and fuel_routes.py check the
    # flag inline for the same reason.
    if not current_user.can_edit:
        abort(403)

    nickname = (request.form.get('nickname') or '').strip()
    unit_id = request.form.get('drone_unit_id', type=int)

    if not nickname:
        flash(_drone_t('Ник бўш бўлиши мумкин эмас.',
                       'Ник не может быть пустым.'), 'warning')
        return _drone_units_redirect()

    unit = DroneUnit.query.get(unit_id) if unit_id else None
    if unit is None:
        flash(_drone_t('Машина танланмаган ёки топилмади.',
                       'Машина не выбрана или не найдена.'), 'warning')
        return _drone_units_redirect()

    # [REASON]: normalized is recomputed here with the ingest's own helper on
    # every write, so the column can never drift from the rule the resolution
    # uses -- the migration seeds it with the identical rule.
    normalized = _drone_normalize_nickname(nickname)
    ambiguous = _drone_nickname_ambiguous_with(normalized, unit.id)

    row = DroneNickname(drone_unit_id=unit.id, nickname=nickname,
                        normalized=normalized, is_active=True,
                        created_at=datetime.utcnow())
    db.session.add(row)
    try:
        db.session.commit()
    except IntegrityError:
        # nickname is UNIQUE in the database; a readable message beats a 500.
        db.session.rollback()
        _drone_flash_duplicate(nickname)
        return _drone_units_redirect()

    flash(_drone_t('«%s» ники № %s машинага қўшилди.',
                   'Ник «%s» добавлен машине № %s.')
          % (nickname, unit.number), 'success')
    if ambiguous is not None:
        _drone_flash_ambiguity(ambiguous)
    return _drone_units_redirect()


@drones_bp.route('/units/nicknames/<int:nick_id>/update', methods=['POST'])
@module_required('drones')
def units_nickname_update(nick_id):
    """Rename a spelling, move it to another machine, activate or disable it.

    One route covers all three edits of a row: they are one form on the
    screen, and splitting them would make "rename and re-attach in one go"
    two round trips that can half-fail.
    """
    if not current_user.can_edit:
        abort(403)

    row = DroneNickname.query.get(nick_id)
    if row is None:
        flash(_drone_t('Ник топилмади.', 'Ник не найден.'), 'warning')
        return _drone_units_redirect()

    nickname = (request.form.get('nickname') or '').strip()
    unit_id = request.form.get('drone_unit_id', type=int)
    # An unchecked checkbox is simply absent from the form body.
    is_active = request.form.get('is_active') is not None

    if not nickname:
        flash(_drone_t('Ник бўш бўлиши мумкин эмас.',
                       'Ник не может быть пустым.'), 'warning')
        return _drone_units_redirect()

    unit = DroneUnit.query.get(unit_id) if unit_id else None
    if unit is None:
        flash(_drone_t('Машина танланмаган ёки топилмади.',
                       'Машина не выбрана или не найдена.'), 'warning')
        return _drone_units_redirect()

    normalized = _drone_normalize_nickname(nickname)
    # Only an alias that will BE active can make a group ambiguous, and the
    # row being edited is excluded from its own check.
    ambiguous = None
    if is_active:
        ambiguous = _drone_nickname_ambiguous_with(normalized, unit.id,
                                                   exclude_id=row.id)

    row.nickname = nickname
    row.normalized = normalized
    row.drone_unit_id = unit.id
    row.is_active = is_active
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        _drone_flash_duplicate(nickname)
        return _drone_units_redirect()

    if is_active:
        flash(_drone_t('«%s» ники янгиланди: № %s машина, фаол.',
                       'Ник «%s» обновлён: машина № %s, активен.')
              % (nickname, unit.number), 'success')
    else:
        # [REASON]: say plainly what a disabled alias does, because it is the
        # only setting on this screen that changes how new data is ingested:
        # _drone_nickname_maps() skips it, so flights carrying this spelling
        # are stored unattributed from now on. Nothing already stored moves.
        flash(_drone_t(
            '«%s» ники ўчирилди: у энди янги парвозларда танилмайди, '
            'аллақачон сақланган парвозлар ўзгармайди.',
            'Ник «%s» отключён: он больше не распознаётся при загрузке новых '
            'вылетов, уже сохранённые вылеты не меняются.')
            % nickname, 'success')
    if ambiguous is not None:
        _drone_flash_ambiguity(ambiguous)
    return _drone_units_redirect()


# ─── Re-attribution pass (DRONE-007) ──────────────────────────────────────────
#
# [REASON]: THE INVARIANT OF THIS WHOLE SECTION -- a flight whose
# drone_unit_id IS NOT NULL is never modified by any route here. Re-attribution
# only fills blanks. Correcting an attribution that already points at a machine
# is a different problem with a different risk profile (it silently moves
# hectares between machines in reports that have already been sent) and is not
# in this change. Every write below therefore carries
# DroneFlight.drone_unit_id.is_(None) IN THE SQL WHERE CLAUSE, never a filter
# applied in Python after loading, so the guarantee does not depend on the
# rows the request happened to read.

# [REASON]: SQLite caps the number of bound parameters in one statement
# (999 before 3.32, 32766 after). The worst case here is 5 977 ids in one
# undo, which lands on the wrong side of that line on an older SQLite and
# fails the whole statement. Chunking is unconditional so behaviour does not
# depend on the library version that happens to be linked into the Python on
# the server.
DRONE_ID_CHUNK = 400


def _drone_id_chunks(ids):
    for start in range(0, len(ids), DRONE_ID_CHUNK):
        yield ids[start:start + DRONE_ID_CHUNK]


def _drone_unattributed_groups():
    """Unattributed flights grouped by raw spelling, with their resolution.

    Returns (resolvable, unresolvable, totals). Reads only -- this is shared
    by the preview and by apply, and apply recomputes it inside its own
    request rather than trusting anything the client posted back.
    """
    rows = (db.session.query(
        DroneFlight.nickname_raw,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0))
        .filter(DroneFlight.drone_unit_id.is_(None))
        .group_by(DroneFlight.nickname_raw).all())

    exact, normalized = _drone_nickname_maps()
    unit_numbers = {u.id: u.number for u in DroneUnit.query.all()}

    resolvable, unresolvable = [], []
    for nickname, flights, area in rows:
        unit_id = _drone_resolve_unit(nickname, exact, normalized)
        entry = {
            'nickname': nickname,
            'flights': flights,
            'area': float(area or 0.0),
            'drone_unit_id': unit_id,
            'unit_number': unit_numbers.get(unit_id),
        }
        (resolvable if unit_id is not None else unresolvable).append(entry)

    resolvable.sort(key=lambda r: (-r['flights'], r['nickname'] or ''))
    unresolvable.sort(key=lambda r: (-r['flights'], r['nickname'] or ''))
    totals = {
        'resolvable_flights': sum(r['flights'] for r in resolvable),
        'resolvable_area': sum(r['area'] for r in resolvable),
        'unresolvable_flights': sum(r['flights'] for r in unresolvable),
        'unresolvable_area': sum(r['area'] for r in unresolvable),
    }
    return resolvable, unresolvable, totals


def _drone_run_detail(run):
    """detail_json parsed, never raising on a malformed or truncated row."""
    try:
        detail = json.loads(run.detail_json or '{}')
    except (TypeError, ValueError):
        return {'by_nickname': {}, 'flight_ids': []}
    if not isinstance(detail, dict):
        return {'by_nickname': {}, 'flight_ids': []}
    by_nickname = detail.get('by_nickname')
    flight_ids = detail.get('flight_ids')
    return {
        'by_nickname': by_nickname if isinstance(by_nickname, dict) else {},
        'flight_ids': [i for i in flight_ids
                       if isinstance(i, int)] if isinstance(flight_ids,
                                                            list) else [],
    }


@drones_bp.route('/units/reattach')
@module_required('drones')
def reattach():
    """Preview of the pass. WRITES NOTHING."""
    if not current_user.can_edit:
        abort(403)

    resolvable, unresolvable, totals = _drone_unattributed_groups()
    runs = (DroneReattachRun.query
            .options(joinedload(DroneReattachRun.performer),
                     joinedload(DroneReattachRun.undoer))
            .order_by(DroneReattachRun.id.desc())
            .limit(50).all())
    run_rows = []
    for run in runs:
        detail = _drone_run_detail(run)
        run_rows.append({
            'id': run.id,
            'performed_at': run.performed_at,
            'performed_by': (run.performer.full_name or
                             run.performer.username) if run.performer else None,
            'rows_matched': run.rows_matched,
            'rows_updated': run.rows_updated,
            'undone_at': run.undone_at,
            'undone_by': (run.undoer.full_name or run.undoer.username)
                         if run.undoer else None,
            'nicknames': sorted(detail['by_nickname'].keys()),
        })

    return render_template(
        'drones/reattach.html',
        resolvable=resolvable,
        unresolvable=unresolvable,
        totals=totals,
        runs=run_rows,
        fmt_dt=_drone_fmt_dt,
    )


@drones_bp.route('/units/reattach/apply', methods=['POST'])
@module_required('drones')
def reattach_apply():
    """Fill drone_unit_id on unattributed flights whose spelling now resolves.

    The grouping is recomputed HERE. Nothing the client posted about which
    rows to change is read, because a form field naming rows to update is a
    form field an attacker can rewrite.
    """
    if not current_user.can_edit:
        abort(403)

    resolvable, _unresolvable, totals = _drone_unattributed_groups()
    if not resolvable:
        # [REASON]: a run row recording zero changes is noise in an audit
        # ledger -- it makes "what has been done to this data" longer to read
        # without adding a fact. Nothing resolvable means nothing happened.
        flash(_drone_t(
            'Қайта бириктириш учун ҳеч нарса йўқ: бириктирилмаган '
            'парвозларнинг ҳеч бир ёзилиши жорий никлар харитасида '
            'танилмади.',
            'Переназначать нечего: ни одно написание среди непривязанных '
            'вылетов не распознаётся текущей картой ников.'), 'info')
        return redirect(url_for('drones.reattach'))

    by_nickname = {}
    flight_ids = []
    rows_updated = 0
    try:
        for group in resolvable:
            nickname = group['nickname']
            unit_id = group['drone_unit_id']
            # The id list and the UPDATE carry the SAME conditions, and
            # drone_unit_id IS NULL is one of them on both sides.
            conds = (DroneFlight.drone_unit_id.is_(None),
                     DroneFlight.nickname_raw == nickname)
            ids = [row[0] for row in
                   db.session.query(DroneFlight.id).filter(*conds).all()]
            if not ids:
                continue
            updated = (db.session.query(DroneFlight).filter(*conds)
                       .update({DroneFlight.drone_unit_id: unit_id},
                               synchronize_session=False))
            rows_updated += updated or 0
            flight_ids.extend(ids)
            by_nickname[nickname] = {
                'drone_unit_id': unit_id,
                'unit_number': group['unit_number'],
                'count': len(ids),
            }

        run = DroneReattachRun(
            performed_by=current_user.id,
            performed_at=datetime.utcnow(),
            rows_matched=totals['resolvable_flights'],
            rows_updated=rows_updated,
            detail_json=json.dumps({'by_nickname': by_nickname,
                                    'flight_ids': flight_ids},
                                   ensure_ascii=False),
        )
        db.session.add(run)
        db.session.commit()
    except Exception as exc:
        # One transaction: either the whole pass and its ledger row land, or
        # neither does. A ledger row describing rows that were rolled back
        # would be worse than no ledger at all.
        db.session.rollback()
        current_app.logger.exception('Drone re-attribution failed')
        flash(_drone_t('Қайта бириктириш бажарилмади, ҳеч нарса '
                       'ўзгартирилмади: %s',
                       'Переназначение не выполнено, ничего не изменено: %s')
              % exc, 'danger')
        return redirect(url_for('drones.reattach'))

    flash(_drone_t(
        'Қайта бириктирилди: %d та ёзилиш, %d та парвоз машиналарга '
        'бириктирилди. Прогон №%d — уни бекор қилиш мумкин.',
        'Переназначено: %d написаний, %d вылетов привязано к машинам. '
        'Прогон №%d — его можно откатить.')
        % (len(by_nickname), rows_updated, run.id), 'success')
    return redirect(url_for('drones.reattach'))


@drones_bp.route('/units/reattach/<int:run_id>/undo', methods=['POST'])
@module_required('drones')
def reattach_undo(run_id):
    """Put the blanks back for exactly the rows this run filled."""
    if not current_user.can_edit:
        abort(403)

    run = DroneReattachRun.query.get(run_id)
    if run is None:
        flash(_drone_t('Прогон топилмади.', 'Прогон не найден.'), 'warning')
        return redirect(url_for('drones.reattach'))
    if run.undone_at is not None:
        flash(_drone_t('Прогон №%d аллақачон бекор қилинган (%s). Ҳеч нарса '
                       'ўзгартирилмади.',
                       'Прогон №%d уже откачен (%s). Ничего не изменено.')
              % (run.id, _drone_fmt_dt(run.undone_at)), 'warning')
        return redirect(url_for('drones.reattach'))

    detail = _drone_run_detail(run)
    recorded_ids = detail['flight_ids']
    expected_unit = {}
    for nickname, info in detail['by_nickname'].items():
        if isinstance(info, dict) and info.get('drone_unit_id') is not None:
            expected_unit[nickname] = info['drone_unit_id']

    reverted = 0
    try:
        # [REASON]: only rows that STILL point at the machine this run
        # recorded for their spelling are reverted. A row re-attributed since
        # -- by a later pass after this one was undone and redone, say -- is
        # skipped rather than blanked, because undoing this run must not undo
        # somebody else's work. The id restriction is equally load-bearing: it
        # separates the rows this run filled from flights of the same spelling
        # that arrived afterwards and were attributed at ingest, which this
        # run never touched and must not blank.
        for chunk in _drone_id_chunks(recorded_ids):
            rows = (db.session.query(DroneFlight.id, DroneFlight.nickname_raw,
                                     DroneFlight.drone_unit_id)
                    .filter(DroneFlight.id.in_(chunk)).all())
            by_unit = {}
            for flight_id, nickname_raw, unit_id in rows:
                if (unit_id is not None
                        and expected_unit.get(nickname_raw) == unit_id):
                    by_unit.setdefault(unit_id, []).append(flight_id)
            # The machine condition is repeated in the UPDATE itself, so the
            # guarantee rests on SQL and not only on the read above.
            for unit_id, ids in by_unit.items():
                reverted += (db.session.query(DroneFlight).filter(
                    DroneFlight.id.in_(ids),
                    DroneFlight.drone_unit_id == unit_id)
                    .update({DroneFlight.drone_unit_id: None},
                            synchronize_session=False) or 0)
        run.undone_at = datetime.utcnow()
        run.undone_by = current_user.id
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('Drone re-attribution undo failed')
        flash(_drone_t('Бекор қилиш бажарилмади, ҳеч нарса ўзгартирилмади: %s',
                       'Откат не выполнен, ничего не изменено: %s')
              % exc, 'danger')
        return redirect(url_for('drones.reattach'))

    skipped = len(recorded_ids) - reverted
    flash(_drone_t(
        'Прогон №%d бекор қилинди: %d та парвоз яна бириктирилмаган ҳолатга '
        'қайтарилди, %d таси ўтказиб юборилди (улар ўшандан бери '
        'ўзгартирилган).',
        'Прогон №%d откачен: %d вылетов возвращено в непривязанное '
        'состояние, %d пропущено (они были изменены с тех пор).')
        % (run.id, reverted, skipped), 'success')
    return redirect(url_for('drones.reattach'))


# ─── Machine card: identifiers, status ledger, batteries (DRONE-FLEET-001) ────
#
# [REASON]: EXACTLY TWO STATUSES, and no third one may be added. The source
# uses `Соз` (serviceable) and `Носоз` (unserviceable); everything else it
# records -- "in Tashkent", "hit a 380 kV line", "5 propellers to replace" --
# is free text and belongs in `comment`. The moment a third value exists, half
# the rows drift into it, the two real states stop being countable, and the
# question the owner actually asks ("how many machines can fly today?") stops
# having an answer.
#
# [REASON]: the ledger is APPEND-ONLY. A status change writes a new row; no
# route here ever UPDATEs one. The reason a machine was down in May has to
# stay readable in August, and an in-place edit erases it silently.


def _drone_status_labels():
    return {
        DRONE_STATUS_SERVICEABLE: _drone_t('Соз', 'Исправна'),
        DRONE_STATUS_UNSERVICEABLE: _drone_t('Носоз', 'Неисправна'),
    }


def _drone_battery_condition_labels():
    return {
        'working': _drone_t('Соз', 'Рабочая'),
        'faulty': _drone_t('Носоз', 'Неисправная'),
        'new_sealed': _drone_t('Янги, очилмаган', 'Новая, невскрытая'),
    }


def _drone_current_status_rows():
    """{drone_unit_id: DroneUnitStatusLog} -- the CURRENT status of each machine.

    Current = the greatest changed_on, ties broken by the greatest id. Both
    halves of that rule are resolved in SQL over two aggregations rather than
    by loading the ledger and sorting in Python: the ledger only grows, and a
    Python max() over a page of loaded rows would quietly start answering with
    whatever happened to be loaded.
    """
    newest = (db.session.query(
        DroneUnitStatusLog.drone_unit_id.label('unit_id'),
        func.max(DroneUnitStatusLog.changed_on).label('changed_on'))
        .group_by(DroneUnitStatusLog.drone_unit_id).subquery())
    winners = (db.session.query(func.max(DroneUnitStatusLog.id))
               .join(newest,
                     and_(DroneUnitStatusLog.drone_unit_id == newest.c.unit_id,
                          DroneUnitStatusLog.changed_on == newest.c.changed_on))
               .group_by(DroneUnitStatusLog.drone_unit_id))
    ids = [row[0] for row in winners.all()]
    if not ids:
        return {}
    rows = DroneUnitStatusLog.query.filter(DroneUnitStatusLog.id.in_(ids)).all()
    return {row.drone_unit_id: row for row in rows}


def _drone_fleet_status_counts():
    """(serviceable, unserviceable, unknown) over all ACTIVE machines.

    A machine with no status row at all is counted as UNKNOWN and never as
    serviceable: "nobody has said" and "somebody said it flies" are different
    facts, and folding them together is how a broken machine gets planned into
    a working day.
    """
    current = _drone_current_status_rows()
    serviceable = unserviceable = unknown = 0
    for unit in DroneUnit.query.all():
        row = current.get(unit.id)
        if row is None:
            unknown += 1
        elif row.status == DRONE_STATUS_SERVICEABLE:
            serviceable += 1
        elif row.status == DRONE_STATUS_UNSERVICEABLE:
            unserviceable += 1
        else:
            unknown += 1
    return serviceable, unserviceable, unknown


def _drone_duplicate_battery_serials():
    """Serial numbers carried by more than one battery row.

    [REASON]: this is a WARNING, never a constraint. Batteries No 8 and No 10
    genuinely share 65VPN19DA50MSU in the source. Whether that is one
    transcription error or two identical labels can only be settled at the
    battery itself, so the register shows what the sheet says and names the
    contradiction.
    """
    rows = (db.session.query(DroneBattery.serial_number)
            .group_by(DroneBattery.serial_number)
            .having(func.count(DroneBattery.id) > 1).all())
    return set(row[0] for row in rows)


def _drone_unit_redirect(unit_id):
    return redirect(url_for('drones.unit_card', unit_id=unit_id))


@drones_bp.route('/units/<int:unit_id>')
@module_required('drones')
def unit_card(unit_id):
    """One machine: identifiers, status history, batteries, nicknames."""
    unit = DroneUnit.query.get(unit_id)
    if unit is None:
        abort(404)
    history = (DroneUnitStatusLog.query
               .options(joinedload(DroneUnitStatusLog.author))
               .filter(DroneUnitStatusLog.drone_unit_id == unit_id)
               .order_by(DroneUnitStatusLog.changed_on.desc(),
                         DroneUnitStatusLog.id.desc())
               .all())
    batteries = (DroneBattery.query
                 .filter(DroneBattery.drone_unit_id == unit_id)
                 .order_by(DroneBattery.label_number, DroneBattery.id).all())
    assignments = (DroneOperatorAssignment.query
                   .options(joinedload(DroneOperatorAssignment.operator))
                   .filter(DroneOperatorAssignment.drone_unit_id == unit_id)
                   .order_by(DroneOperatorAssignment.date_from.desc(),
                             DroneOperatorAssignment.id.desc()).all())
    today = _drone_today_local()
    return render_template(
        'drones/unit_card.html',
        unit=unit,
        history=history,
        current_status=(history[0] if history else None),
        batteries=batteries,
        duplicate_serials=_drone_duplicate_battery_serials(),
        condition_labels=_drone_battery_condition_labels(),
        status_labels=_drone_status_labels(),
        nicknames=(DroneNickname.query
                   .filter(DroneNickname.drone_unit_id == unit_id)
                   .order_by(DroneNickname.id).all()),
        assignments=assignments,
        current_flags={a.id: _drone_assignment_is_current(a, today)
                       for a in assignments},
        today=today,
    )


@drones_bp.route('/units/<int:unit_id>/fields', methods=['POST'])
@module_required('drones')
def unit_fields_update(unit_id):
    """Edit hardware_id, generator_id and subdivision_name of one machine.

    [REASON]: the field is hardware_id and never serial_number. In this module
    serial_number means the per-flight record id of the DJI payload -- 22 855
    distinct values on 22 855 flights -- and reusing the name for the airframe
    would make the whole module ambiguous to read.
    """
    if not current_user.can_edit:
        abort(403)

    unit = DroneUnit.query.get(unit_id)
    if unit is None:
        abort(404)

    unit.hardware_id = (request.form.get('hardware_id') or '').strip() or None
    unit.generator_id = (request.form.get('generator_id') or '').strip() or None
    unit.subdivision_name = (request.form.get('subdivision_name') or '').strip() or None
    db.session.commit()
    flash(_drone_t('№ %s машина маълумотлари сақланди.',
                   'Данные машины № %s сохранены.') % unit.number, 'success')
    return _drone_unit_redirect(unit_id)


@drones_bp.route('/units/<int:unit_id>/status', methods=['POST'])
@module_required('drones')
def unit_status_add(unit_id):
    """Append ONE row to the status ledger. Nothing is ever updated in place."""
    if not current_user.can_edit:
        abort(403)

    unit = DroneUnit.query.get(unit_id)
    if unit is None:
        abort(404)

    status = (request.form.get('status') or '').strip()
    if status not in DRONE_STATUSES:
        # [REASON]: the closed set is enforced here, in the only place that
        # writes the column. A machine "in repair" or "written off" is still
        # either flyable or not, and its situation belongs in the comment.
        flash(_drone_t('Ҳолат фақат «Соз» ёки «Носоз» бўлиши мумкин.',
                       'Статус может быть только «Исправна» или '
                       '«Неисправна».'), 'warning')
        return _drone_unit_redirect(unit_id)

    changed_on = _drone_parse_date(request.form.get('changed_on'))
    if changed_on is None:
        flash(_drone_t('Ўзгариш санаси кўрсатилмаган.',
                       'Дата изменения не указана.'), 'warning')
        return _drone_unit_redirect(unit_id)
    if changed_on > _drone_today_local():
        # [REASON]: a past date is normal and expected -- a machine is
        # reported broken after the fact. Only the future is refused, because
        # a future row would become the "current" status of the machine
        # immediately and hide the real one.
        flash(_drone_t('Ўзгариш санаси келажакда бўлиши мумкин эмас.',
                       'Дата изменения не может быть в будущем.'), 'warning')
        return _drone_unit_redirect(unit_id)

    db.session.add(DroneUnitStatusLog(
        drone_unit_id=unit.id,
        status=status,
        comment=(request.form.get('comment') or '').strip() or None,
        changed_on=changed_on,
        changed_by=(current_user.id
                    if getattr(current_user, 'is_authenticated', False)
                    else None)))
    db.session.commit()
    flash(_drone_t('№ %s машина ҳолати ёзилди: %s, %s.',
                   'Статус машины № %s записан: %s, %s.')
          % (unit.number, _drone_status_labels()[status],
             _drone_fmt_date(changed_on)), 'success')
    return _drone_unit_redirect(unit_id)


# ─── Fleet directories: operators and assignments (DRONE-FLEET-001) ───────────
#
# [REASON]: THE RULE THIS WHOLE SECTION EXISTS FOR. An assignment of an
# operator to a machine WITHOUT a validity period is worthless. The
# 2026-08-03 reconciliation against the dispatchers' April ledgers matched the
# month total to 0.15 % while the per-subdivision breakdown carried two errors
# of +460 ha and -544 ha that happened to cancel: one operator flew machine
# No 4 until 13 April and No 8 from 14 April, and the dispatcher kept booking
# his work to his own subdivision. Our data confirms the boundary
# independently -- No 4 has 198 flights and none after 13 April, No 8 jumps
# from 107 to 541 flights on that date. With dates, all twelve operators fell
# within +-9 %. So date_from is mandatory everywhere below, and closing an
# assignment means SETTING date_to -- there is no delete route, for an
# operator or for an assignment, exactly as there is none for a nickname.
#
# [REASON]: OVERLAPS ARE SAVED, NOT REFUSED. Two operators genuinely appear on
# one machine in the source, and a substitution is written into the sheet as
# "Anvarov Usmon (Qodirov Nurali)". A screen that refuses the real data is a
# screen nobody can use, and it would push somebody to invent dates that make
# the form happy. The save succeeds and a warning names the other assignment;
# the report gives a multi-covered flight its own visible row.


def _drone_operators_redirect(op_id=None):
    if op_id:
        return redirect(url_for('drones.operator_card', op_id=op_id))
    return redirect(url_for('drones.operators'))


def _drone_assignment_overlaps(drone_unit_id, date_from, date_to,
                               exclude_id=None):
    """Assignments on the SAME machine whose period intersects the given one.

    Two half-open intervals overlap when A.from <= B.to and B.from <= A.to,
    with a NULL end meaning "still current", i.e. +infinity. The NULL branch
    is spelled out rather than folded into a COALESCE with a far-future date:
    a sentinel date is a value somebody eventually filters on by accident.

    This is a WARNING source, never a veto -- see the section header.
    """
    q = DroneOperatorAssignment.query.filter(
        DroneOperatorAssignment.drone_unit_id == drone_unit_id)
    if exclude_id:
        q = q.filter(DroneOperatorAssignment.id != exclude_id)
    if date_to is not None:
        q = q.filter(DroneOperatorAssignment.date_from <= date_to)
    q = q.filter(or_(DroneOperatorAssignment.date_to.is_(None),
                     DroneOperatorAssignment.date_to >= date_from))
    return (q.options(joinedload(DroneOperatorAssignment.operator),
                      joinedload(DroneOperatorAssignment.drone_unit))
            .order_by(DroneOperatorAssignment.date_from).all())


def _drone_flash_overlap(rows):
    """Name every other assignment the saved one overlaps with."""
    if not rows:
        return
    names = ', '.join(
        '%s (№ %s, %s — %s)' % (
            r.operator.full_name if r.operator else '?',
            r.drone_unit.number if r.drone_unit else '?',
            _drone_fmt_date(r.date_from),
            _drone_fmt_date(r.date_to))
        for r in rows)
    flash(_drone_t(
        'Сақланди, аммо давр бошқа бириктириш билан кесишади: %s. '
        'Ҳисоботда бундай парвозлар «Бир нечта оператор» қаторига тушади — '
        'улар йўқолмайди, аммо бирор операторга ёзилмайди ҳам.',
        'Сохранено, но период пересекается с другим назначением: %s. '
        'В отчёте такие вылеты попадут в строку «Несколько операторов» — '
        'они не теряются, но и не приписываются никому одному.')
        % names, 'warning')


def _drone_assignments_by_operator(operator_ids=None):
    """All assignments, grouped by operator id, machines joined in one query."""
    q = (DroneOperatorAssignment.query
         .options(joinedload(DroneOperatorAssignment.drone_unit)))
    if operator_ids is not None:
        q = q.filter(DroneOperatorAssignment.operator_id.in_(operator_ids))
    grouped = {}
    for row in q.order_by(DroneOperatorAssignment.date_from.desc(),
                          DroneOperatorAssignment.id.desc()).all():
        grouped.setdefault(row.operator_id, []).append(row)
    return grouped


@drones_bp.route('/operators')
@module_required('drones')
def operators():
    """Directory of pilots with the machines they currently fly."""
    rows = (DroneOperator.query
            .order_by(DroneOperator.is_active.desc(), DroneOperator.full_name)
            .all())
    # One query for every assignment, grouped in Python -- the lazy='dynamic'
    # relationship would give one query per operator.
    assignments = _drone_assignments_by_operator()
    today = _drone_today_local()
    current_units = {}
    for op in rows:
        current_units[op.id] = [
            a for a in assignments.get(op.id, [])
            if _drone_assignment_is_current(a, today)]
    return render_template(
        'drones/operators.html',
        operators=rows,
        current_units=current_units,
        assignment_counts={op.id: len(assignments.get(op.id, []))
                           for op in rows},
        today=today,
    )


@drones_bp.route('/operators/add', methods=['POST'])
@module_required('drones')
def operators_add():
    if not current_user.can_edit:
        abort(403)

    full_name = (request.form.get('full_name') or '').strip()
    if not full_name:
        flash(_drone_t('Оператор Ф.И.Ш. бўш бўлиши мумкин эмас.',
                       'Ф.И.О. оператора не может быть пустым.'), 'warning')
        return _drone_operators_redirect()

    row = DroneOperator(
        full_name=full_name,
        phone_primary=(request.form.get('phone_primary') or '').strip() or None,
        phone_secondary=(request.form.get('phone_secondary') or '').strip() or None,
        subdivision_name=(request.form.get('subdivision_name') or '').strip() or None,
        note=(request.form.get('note') or '').strip() or None,
        is_active=True,
    )
    db.session.add(row)
    db.session.commit()
    flash(_drone_t('«%s» оператори қўшилди.', 'Оператор «%s» добавлен.')
          % full_name, 'success')
    return _drone_operators_redirect(row.id)


@drones_bp.route('/operators/<int:op_id>')
@module_required('drones')
def operator_card(op_id):
    """One pilot: his details and the full history of his assignments."""
    operator = DroneOperator.query.get(op_id)
    if operator is None:
        abort(404)
    assignments = _drone_assignments_by_operator([op_id]).get(op_id, [])
    today = _drone_today_local()
    return render_template(
        'drones/operator_card.html',
        operator=operator,
        assignments=assignments,
        current_flags={a.id: _drone_assignment_is_current(a, today)
                       for a in assignments},
        units=DroneUnit.query.order_by(DroneUnit.number).all(),
        today=today,
    )


@drones_bp.route('/operators/<int:op_id>/update', methods=['POST'])
@module_required('drones')
def operator_update(op_id):
    """Edit a pilot, or deactivate him. There is deliberately no delete.

    [REASON]: an operator who left still owns the hectares he flew. Deleting
    the row would orphan every assignment that explains an already-sent
    report; deactivating keeps the history readable and takes him out of the
    pickers.
    """
    if not current_user.can_edit:
        abort(403)

    operator = DroneOperator.query.get(op_id)
    if operator is None:
        abort(404)

    full_name = (request.form.get('full_name') or '').strip()
    if not full_name:
        flash(_drone_t('Оператор Ф.И.Ш. бўш бўлиши мумкин эмас.',
                       'Ф.И.О. оператора не может быть пустым.'), 'warning')
        return _drone_operators_redirect(op_id)

    operator.full_name = full_name
    operator.phone_primary = (request.form.get('phone_primary') or '').strip() or None
    operator.phone_secondary = (request.form.get('phone_secondary') or '').strip() or None
    operator.subdivision_name = (request.form.get('subdivision_name') or '').strip() or None
    operator.note = (request.form.get('note') or '').strip() or None
    # An unchecked checkbox is simply absent from the form body.
    operator.is_active = request.form.get('is_active') is not None
    db.session.commit()

    if operator.is_active:
        flash(_drone_t('«%s» оператори янгиланди.',
                       'Оператор «%s» обновлён.') % full_name, 'success')
    else:
        flash(_drone_t(
            '«%s» оператори фаолсизлантирилди: у янги бириктиришлар '
            'рўйхатида кўринмайди, аммо бириктиришлари ва гектарлари '
            'сақланиб қолади.',
            'Оператор «%s» деактивирован: он не появляется в списках выбора, '
            'но его назначения и гектары сохраняются.')
            % full_name, 'success')
    return _drone_operators_redirect(op_id)


def _drone_assignment_form_values():
    """(drone_unit, date_from, date_to, note, error_message) from the form."""
    unit_id = request.form.get('drone_unit_id', type=int)
    unit = DroneUnit.query.get(unit_id) if unit_id else None
    date_from = _drone_parse_date(request.form.get('date_from'))
    date_to = _drone_parse_date(request.form.get('date_to'))
    note = (request.form.get('note') or '').strip() or None

    if unit is None:
        return None, None, None, None, _drone_t(
            'Машина танланмаган ёки топилмади.',
            'Машина не выбрана или не найдена.')
    if date_from is None:
        # [REASON]: the mandatory field of this task. A dateless assignment
        # cannot reproduce April and will be wrong again at the next move.
        return None, None, None, None, _drone_t(
            '«Санадан» мажбурий: даврсиз бириктириш ҳисоботни тиклай олмайди.',
            '«Дата с» обязательна: назначение без периода не может '
            'воспроизвести отчёт.')
    if date_to is not None and date_to < date_from:
        return None, None, None, None, _drone_t(
            '«Санагача» «Санадан»дан олдин бўлиши мумкин эмас.',
            '«Дата по» не может быть раньше «Даты с».')
    return unit, date_from, date_to, note, None


@drones_bp.route('/operators/<int:op_id>/assignments/add', methods=['POST'])
@module_required('drones')
def operator_assignment_add(op_id):
    if not current_user.can_edit:
        abort(403)

    operator = DroneOperator.query.get(op_id)
    if operator is None:
        abort(404)

    unit, date_from, date_to, note, error = _drone_assignment_form_values()
    if error:
        flash(error, 'warning')
        return _drone_operators_redirect(op_id)

    overlaps = _drone_assignment_overlaps(unit.id, date_from, date_to)
    row = DroneOperatorAssignment(
        operator_id=operator.id, drone_unit_id=unit.id, date_from=date_from,
        date_to=date_to, note=note,
        created_by=(current_user.id
                    if getattr(current_user, 'is_authenticated', False)
                    else None))
    db.session.add(row)
    db.session.commit()

    flash(_drone_t('Бириктириш қўшилди: № %s машина, %s дан.',
                   'Назначение добавлено: машина № %s, с %s.')
          % (unit.number, _drone_fmt_date(date_from)), 'success')
    _drone_flash_overlap(overlaps)
    return _drone_operators_redirect(op_id)


@drones_bp.route('/operators/assignments/<int:assign_id>/update',
                 methods=['POST'])
@module_required('drones')
def operator_assignment_update(assign_id):
    """Edit one assignment, including closing it by setting date_to.

    [REASON]: closing is SETTING date_to, never deleting the row. The row is
    the only record of who flew that machine in April; a report printed in May
    stops being explainable the moment it is gone.
    """
    if not current_user.can_edit:
        abort(403)

    row = DroneOperatorAssignment.query.get(assign_id)
    if row is None:
        abort(404)

    unit, date_from, date_to, note, error = _drone_assignment_form_values()
    if error:
        flash(error, 'warning')
        return _drone_operators_redirect(row.operator_id)

    overlaps = _drone_assignment_overlaps(unit.id, date_from, date_to,
                                          exclude_id=row.id)
    row.drone_unit_id = unit.id
    row.date_from = date_from
    row.date_to = date_to
    row.note = note
    db.session.commit()

    if date_to is None:
        flash(_drone_t('Бириктириш янгиланди: № %s машина, %s дан, амалда.',
                       'Назначение обновлено: машина № %s, с %s, действует.')
              % (unit.number, _drone_fmt_date(date_from)), 'success')
    else:
        flash(_drone_t('Бириктириш ёпилди: № %s машина, %s — %s.',
                       'Назначение закрыто: машина № %s, %s — %s.')
              % (unit.number, _drone_fmt_date(date_from),
                 _drone_fmt_date(date_to)), 'success')
    _drone_flash_overlap(overlaps)
    return _drone_operators_redirect(row.operator_id)


# ─── Summary and Excel exports (DRONE-004) ────────────────────────────────────

# [REASON]: a silently truncated export reads as complete data on the other
# side of an e-mail; above the cap the route refuses with a clear message
# asking to narrow the period instead.
DRONE_FLIGHTS_XLSX_CAP = 50000


def _drone_filters_from_args(args, default_current_month):
    """Parse the shared filter set (dates, machine, region) from a query
    string, mirroring index(): the same _drone_parse_date and the same UTC+5
    day-boundary shift -- one convention, not two.

    When default_current_month is true and NEITHER date parameter is present
    in the query string at all, the current calendar month (in UTC+5) is
    preselected. Parameters that are present but empty mean "no bound" --
    that is how the operator asks for all time by clearing the inputs.
    """
    has_date_args = ('date_from' in args) or ('date_to' in args)
    date_from_s = (args.get('date_from') or '').strip()
    date_to_s = (args.get('date_to') or '').strip()
    date_from = _drone_parse_date(date_from_s)
    date_to = _drone_parse_date(date_to_s)
    if default_current_month and not has_date_args:
        today_local = (datetime.utcnow() + DRONE_DISPLAY_UTC_OFFSET).date()
        date_from = today_local.replace(day=1)
        if today_local.month == 12:
            date_to = today_local.replace(day=31)
        else:
            date_to = (today_local.replace(month=today_local.month + 1, day=1)
                       - timedelta(days=1))
        date_from_s = date_from.isoformat()
        date_to_s = date_to.isoformat()
    return {
        'date_from': date_from,
        'date_to': date_to,
        'date_from_s': date_from_s,
        'date_to_s': date_to_s,
        # [REASON]: the flag travels with the parsed filters so a caller can
        # tell "no date filter was ever specified" from "the date filter was
        # explicitly cleared". Both look like an empty date_from_s, but only
        # the second one must survive into an export link -- see
        # _drone_link_args.
        'has_date_args': has_date_args,
        'unit_id': args.get('unit_id', type=int),
        'region': (args.get('region') or '').strip(),
    }


def _drone_flight_conditions(filters):
    """Filter conditions over DroneFlight for the parsed filter set."""
    conds = []
    if filters['date_from']:
        conds.append(DroneFlight.started_at >=
                     datetime.combine(filters['date_from'],
                                      datetime.min.time())
                     - DRONE_DISPLAY_UTC_OFFSET)
    if filters['date_to']:
        conds.append(DroneFlight.started_at <=
                     datetime.combine(filters['date_to'],
                                      datetime.max.time())
                     - DRONE_DISPLAY_UTC_OFFSET)
    if filters['unit_id']:
        conds.append(DroneFlight.drone_unit_id == filters['unit_id'])
    if filters['region']:
        conds.append(DroneFlight.region == filters['region'])
    return conds


def _drone_link_args(filters):
    """Query args for drill-down links and exports.

    A set date is passed through. A date that was explicitly cleared is passed
    through as an EMPTY value rather than dropped: _drone_filters_from_args
    decides whether to apply its current-month default by the PRESENCE of the
    key, so dropping a cleared date makes the target silently fall back to the
    current month while the page that produced the link shows all time.
    """
    link = {}
    if filters['date_from_s']:
        link['date_from'] = filters['date_from_s']
    elif filters.get('has_date_args'):
        link['date_from'] = ''
    if filters['date_to_s']:
        link['date_to'] = filters['date_to_s']
    elif filters.get('has_date_args'):
        link['date_to'] = ''
    if filters['unit_id']:
        link['unit_id'] = filters['unit_id']
    if filters['region']:
        link['region'] = filters['region']
    return link


def _drone_share(area, total_area):
    if not total_area:
        return 0.0
    return round(area * 100.0 / total_area, 1)


def _drone_rate(numerator, denominator):
    """numerator/denominator, or None when the denominator is zero.

    [REASON]: DRONE-FLEET-001 section 6.2. 899 April-to-May flights have zero
    area and about 3 % of all flights do -- manual and technical take-offs,
    normal and not an error. None travels to the template and the workbook as
    an EM DASH: a zero would read as "this machine sprays nothing per
    hectare", which is a different and false statement, and an unguarded
    division would 500 the whole page over data that is fine.
    """
    if not denominator:
        return None
    return numerator / denominator


# Bucket keys of the operator cut. Both are rows of the table and both count
# towards the grand total, exactly like «Не распознано» and «Область не
# определена» already do.
DRONE_OPERATOR_NONE = 'none'
DRONE_OPERATOR_MANY = 'many'


def _drone_flight_operator_subquery(conds):
    """One row per flight: how many operators cover it, and which one.

    THE ATTRIBUTION RULE, implemented exactly as specified and nowhere else:
    a flight belongs to the operator whose assignment covers ITS MACHINE on
    ITS LOCAL DATE (UTC+5). No cover -> «Оператор не определён». More than one
    -> «Несколько операторов». A flight is never silently handed to one of
    several candidates.

    [REASON]: the count is over DISTINCT operator_id, not over assignment
    rows. Two rows for the SAME man on the same machine (a period he re-opened
    after a repair, say) are one candidate, not an ambiguity; counting rows
    would push his own hectares into «Несколько операторов» and make the cut
    read as chaos where the data is merely verbose.

    [REASON]: both sides of the date comparison are wrapped in date(). The
    DATE columns have NUMERIC affinity in SQLite while date(started_at, ...)
    yields TEXT, and leaning on affinity coercion to make that comparison work
    is a silent dependency on storage format. date() on both sides makes it a
    text-to-text comparison of two ISO strings, which is what it must be.

    [REASON]: the whole cut is derived from ONE row per flight (GROUP BY the
    primary key), which is what makes it reconcile with every other cut on the
    page. A flight cannot be counted twice by an operator holding two
    overlapping assignments.
    """
    offset_minutes = int(DRONE_DISPLAY_UTC_OFFSET.total_seconds() // 60)
    local_date = func.date(DroneFlight.started_at,
                           '%+d minutes' % offset_minutes)
    assign = aliased(DroneOperatorAssignment)
    return (db.session.query(
        DroneFlight.id.label('flight_id'),
        DroneFlight.area_ha.label('area_ha'),
        func.count(func.distinct(assign.operator_id)).label('covers'),
        func.min(assign.operator_id).label('operator_id'),
    ).outerjoin(assign, and_(
        assign.drone_unit_id == DroneFlight.drone_unit_id,
        func.date(assign.date_from) <= local_date,
        or_(assign.date_to.is_(None),
            func.date(assign.date_to) >= local_date),
    )).filter(*conds).group_by(DroneFlight.id).subquery())


def _drone_operator_cut(conds, total_area):
    """Flights, hectares and share per operator, plus the two special rows."""
    per_flight = _drone_flight_operator_subquery(conds)
    groups = (db.session.query(
        per_flight.c.covers,
        per_flight.c.operator_id,
        func.count(per_flight.c.flight_id),
        func.coalesce(func.sum(per_flight.c.area_ha), 0.0),
    ).group_by(per_flight.c.covers, per_flight.c.operator_id).all())

    names = {op.id: op.full_name for op in DroneOperator.query.all()}
    by_operator = {}
    undetermined = {'flights': 0, 'area': 0.0}
    multiple = {'flights': 0, 'area': 0.0}
    for covers, operator_id, flights, area in groups:
        area = float(area or 0.0)
        if not covers:
            undetermined['flights'] += flights
            undetermined['area'] += area
        elif covers == 1:
            row = by_operator.setdefault(
                operator_id, {'operator_id': operator_id,
                              'name': names.get(operator_id, '?'),
                              'flights': 0, 'area': 0.0})
            row['flights'] += flights
            row['area'] += area
        else:
            multiple['flights'] += flights
            multiple['area'] += area

    rows = sorted(by_operator.values(), key=lambda r: r['area'], reverse=True)
    for row in rows:
        row['share'] = _drone_share(row['area'], total_area)
    for special in (undetermined, multiple):
        special['share'] = _drone_share(special['area'], total_area)

    o_flights = (sum(r['flights'] for r in rows) + undetermined['flights']
                 + multiple['flights'])
    o_area = (sum(r['area'] for r in rows) + undetermined['area']
              + multiple['area'])
    return {
        'rows': rows,
        # Both special rows are ALWAYS present in the structure; the template
        # hides an empty one. They are never folded into a named operator.
        'undetermined': undetermined,
        'multiple': multiple,
        'total': {'flights': o_flights, 'area': o_area},
    }


def _drone_operator_by_flight(conds):
    """{flight id: operator label} under the same rule as the cut above.

    Built from the same subquery, so the flat export and the operator cut
    cannot drift apart: one attribution rule, one place it is written.
    """
    per_flight = _drone_flight_operator_subquery(conds)
    rows = db.session.query(per_flight.c.flight_id, per_flight.c.covers,
                            per_flight.c.operator_id).all()
    names = {op.id: op.full_name for op in DroneOperator.query.all()}
    label_none = _drone_t('Оператор аниқланмаган', 'Оператор не определён')
    label_many = _drone_t('Бир нечта оператор', 'Несколько операторов')
    labels = {}
    for flight_id, covers, operator_id in rows:
        if not covers:
            labels[flight_id] = label_none
        elif covers == 1:
            labels[flight_id] = names.get(operator_id, label_none)
        else:
            labels[flight_id] = label_many
    return labels


def _drone_summary_data(conds):
    """Aggregate the filtered flights for the summary page and the exports.

    [REASON]: every breakdown must reconcile with the grand total, and
    unattributed flights must appear as their own visible line.
    drone_flights.drone_unit_id is nullable BY DESIGN -- a flight whose DJI
    nickname is not yet in the alias map is stored with NULL rather than
    rejected. A per-machine table that simply groups by machine silently
    omits those flights and their hectares; that is precisely how the
    previous system lost 2 036 flights and 2 082 hectares into an invisible
    bucket. So: NULL drone_unit_id gets an explicit line, NULL region gets
    an explicit line, every table carries a total row, and each table's
    total is compared against the grand total -- a mismatch is surfaced on
    the page (reconciled=False), never hidden.

    Aggregation happens in SQL (func.count / func.sum / group_by), never by
    looping over loaded rows: the table holds ~29 000 flights after the
    historical backfill and keeps growing.
    """
    totals_row = (db.session.query(
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
        func.coalesce(func.sum(DroneFlight.work_seconds), 0),
        func.coalesce(func.sum(DroneFlight.spray_liters), 0.0),
        func.coalesce(func.sum(DroneFlight.sow_kg), 0.0),
    ).filter(*conds).one())
    zero_area = (db.session.query(func.count(DroneFlight.id))
                 .filter(*conds).filter(DroneFlight.area_ha == 0).scalar())

    totals = {
        'flights': totals_row[0] or 0,
        'area_ha': float(totals_row[1] or 0.0),
        'hours': (totals_row[2] or 0) / 3600.0,
        'spray_liters': float(totals_row[3] or 0.0),
        'sow_kg': float(totals_row[4] or 0.0),
        'zero_area': zero_area or 0,
    }
    total_area = totals['area_ha']

    def reconciled(flights_sum, area_sum):
        return (flights_sum == totals['flights']
                and abs(area_sum - totals['area_ha']) < 0.005)

    # По машинам -- ordered by machine number, NULL as its own line.
    machine_groups = (db.session.query(
        DroneFlight.drone_unit_id,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
        # DRONE-FLEET-001: the two productivity columns are derived from
        # fields already stored -- no new column, no new table.
        func.coalesce(func.sum(DroneFlight.work_seconds), 0),
        func.coalesce(func.sum(DroneFlight.spray_liters), 0.0),
    ).filter(*conds).group_by(DroneFlight.drone_unit_id).all())
    unit_numbers = {u.id: u.number for u in DroneUnit.query.all()}
    machine_rows = []
    machine_unattributed = None
    m_seconds = 0
    m_liters = 0.0
    for unit_id, flights, area, seconds, liters in machine_groups:
        area = float(area or 0.0)
        seconds = int(seconds or 0)
        liters = float(liters or 0.0)
        m_seconds += seconds
        m_liters += liters
        cell = {
            'flights': flights,
            'area': area,
            'share': _drone_share(area, total_area),
            'hours': seconds / 3600.0,
            'spray_liters': liters,
            # Guarded: a machine with no recorded flight time and a machine
            # with no sprayed area both render as an em dash, never as zero.
            'ha_per_hour': _drone_rate(area, seconds / 3600.0),
            'liters_per_ha': _drone_rate(liters, area),
        }
        if unit_id is None:
            machine_unattributed = cell
        else:
            cell['unit_id'] = unit_id
            cell['number'] = unit_numbers.get(unit_id)
            machine_rows.append(cell)
    machine_rows.sort(key=lambda r: (r['number'] is None, r['number']))
    m_flights = (sum(r['flights'] for r in machine_rows)
                 + (machine_unattributed['flights']
                    if machine_unattributed else 0))
    m_area = (sum(r['area'] for r in machine_rows)
              + (machine_unattributed['area']
                 if machine_unattributed else 0.0))
    by_machine = {
        'rows': machine_rows,
        'unattributed': machine_unattributed,
        'total': {
            'flights': m_flights,
            'area': m_area,
            'hours': m_seconds / 3600.0,
            'spray_liters': m_liters,
            'ha_per_hour': _drone_rate(m_area, m_seconds / 3600.0),
            'liters_per_ha': _drone_rate(m_liters, m_area),
        },
        'reconciled': reconciled(m_flights, m_area),
    }

    # По областям -- ordered by hectares descending, NULL as its own line.
    region_groups = (db.session.query(
        DroneFlight.region,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
    ).filter(*conds).group_by(DroneFlight.region).all())
    region_rows = []
    region_undetermined = None
    for region, flights, area in region_groups:
        area = float(area or 0.0)
        if region is None:
            region_undetermined = {'flights': flights, 'area': area,
                                   'share': _drone_share(area, total_area)}
        else:
            region_rows.append({
                'region': region,
                # The label rides alongside the raw value, never replaces it:
                # the drill-down link and the Excel row both need the raw
                # string to keep filtering, and both need the label to be
                # readable. Computed here so the page and summary_xlsx use
                # one helper rather than two.
                'label': _drone_region_label(region),
                'flights': flights,
                'area': area,
                'share': _drone_share(area, total_area),
            })
    region_rows.sort(key=lambda r: r['area'], reverse=True)
    r_flights = (sum(r['flights'] for r in region_rows)
                 + (region_undetermined['flights']
                    if region_undetermined else 0))
    r_area = (sum(r['area'] for r in region_rows)
              + (region_undetermined['area'] if region_undetermined else 0.0))
    by_region = {
        'rows': region_rows,
        'undetermined': region_undetermined,
        'total': {'flights': r_flights, 'area': r_area},
        'reconciled': reconciled(r_flights, r_area),
    }

    # По типам работ -- spraying and sowing separately; litres belong only
    # to the spraying row and kilograms only to the sowing row (different
    # substances in different units are NEVER added together); any other
    # usage_type value gets its own row instead of being folded into either.
    usage_groups = (db.session.query(
        DroneFlight.usage_type,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
        func.coalesce(func.sum(DroneFlight.spray_liters), 0.0),
        func.coalesce(func.sum(DroneFlight.sow_kg), 0.0),
    ).filter(*conds).group_by(DroneFlight.usage_type).all())
    usage_labels = _drone_usage_labels()
    usage_rows = []
    for usage_type, flights, area, liters, kg in usage_groups:
        area = float(area or 0.0)
        label = usage_labels.get(usage_type)
        if label is None:
            label = (_drone_t('Номаълум тур', 'Неизвестный тип')
                     + ': ' + str(usage_type))
        usage_rows.append({
            'usage_type': usage_type,
            'label': label,
            'flights': flights,
            'area': area,
            'share': _drone_share(area, total_area),
            'spray_liters': float(liters or 0.0) if usage_type == 0 else None,
            'sow_kg': float(kg or 0.0) if usage_type == 1 else None,
        })
    usage_rows.sort(key=lambda r: (r['usage_type'] is None,
                                   r['usage_type']
                                   if r['usage_type'] is not None else 0))
    u_flights = sum(r['flights'] for r in usage_rows)
    u_area = sum(r['area'] for r in usage_rows)
    by_usage = {
        'rows': usage_rows,
        'total': {'flights': u_flights, 'area': u_area},
        'reconciled': reconciled(u_flights, u_area),
    }

    # По месяцам -- calendar months in the operator's timezone, ascending.
    # The shift must match the display shift exactly: a flight at 19:30 UTC
    # on the 31st belongs to the next month for the operator.
    # [REASON]: the modifier is derived from DRONE_DISPLAY_UTC_OFFSET, never
    # written as a literal. Eight other sites in this module already derive
    # from the constant; a ninth that hardcodes it would silently keep
    # grouping at the old offset if the constant ever changes, and the month
    # table would stop reconciling with the header cards while still looking
    # plausible. Minutes, not hours, so a half-hour timezone stays correct.
    _offset_minutes = int(DRONE_DISPLAY_UTC_OFFSET.total_seconds() // 60)
    month_expr = func.strftime(
        '%Y-%m',
        func.datetime(DroneFlight.started_at,
                      '%+d minutes' % _offset_minutes))
    month_groups = (db.session.query(
        month_expr,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
    ).filter(*conds).group_by(month_expr).order_by(month_expr).all())
    month_rows = [{
        'month': month,
        'flights': flights,
        'area': float(area or 0.0),
        'share': _drone_share(float(area or 0.0), total_area),
    } for month, flights, area in month_groups]
    mo_flights = sum(r['flights'] for r in month_rows)
    mo_area = sum(r['area'] for r in month_rows)
    by_month = {
        'rows': month_rows,
        'total': {'flights': mo_flights, 'area': mo_area},
        'reconciled': reconciled(mo_flights, mo_area),
    }

    # По операторам (DRONE-FLEET-001) -- derived from the assignment table,
    # never observed. It must reconcile with the grand total exactly like the
    # four cuts above, and it does because it is built from one row per flight.
    by_operator = _drone_operator_cut(conds, total_area)
    by_operator['reconciled'] = reconciled(by_operator['total']['flights'],
                                           by_operator['total']['area'])

    return {
        'totals': totals,
        'by_machine': by_machine,
        'by_region': by_region,
        'by_usage': by_usage,
        'by_month': by_month,
        'by_operator': by_operator,
    }


@drones_bp.route('/summary')
@module_required('drones')
def summary():
    filters = _drone_filters_from_args(request.args,
                                       default_current_month=True)
    conds = _drone_flight_conditions(filters)
    data = _drone_summary_data(conds)

    units = DroneUnit.query.order_by(DroneUnit.number).all()
    regions = [row[0] for row in
               db.session.query(DroneFlight.region)
               .filter(DroneFlight.region.isnot(None))
               .distinct().order_by(DroneFlight.region).all()]

    return render_template(
        'drones/summary.html',
        data=data,
        filters=filters,
        link_args=_drone_link_args(filters),
        units=units,
        regions=regions,
        region_labels=_drone_region_label_map(regions),
    )


@drones_bp.route('/sources')
@module_required('drones')
def sources():
    """Sync observability: who has gone silent, and what the last runs did.

    [REASON]: DRONE-FLEET-001 section 5. Everything on this page is DERIVED --
    no table was added for it. Two facts motivated it and neither is visible
    anywhere else today: machines No 11 and No 13 produced 200.11 ha in May
    that no dispatcher ledger records, and machines No 4, No 9 and No 12
    produced nothing at all in May. Silence from a machine has to be a row
    somebody can see, not something noticed when the numbers stop.

    [REASON]: A MACHINE WITH ZERO FLIGHTS IN THE PERIOD IS SHOWN AS A ROW.
    Omitting it would make it indistinguishable from a machine nobody looked
    at, which is precisely the failure this screen exists to prevent. The
    "last flight" column is therefore computed over ALL time and not inside
    the period: a machine that has been silent for three months must show the
    date it last spoke, not an em dash.
    """
    filters = _drone_filters_from_args(request.args,
                                       default_current_month=True)
    # Only the date bounds apply here: a per-machine silence table filtered to
    # one machine would answer a question nobody asks and hide the rest.
    period_conds = _drone_flight_conditions(dict(filters, unit_id=None,
                                                 region=''))

    period_rows = (db.session.query(
        DroneFlight.drone_unit_id,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
    ).filter(*period_conds).group_by(DroneFlight.drone_unit_id).all())
    period_by_unit = {unit_id: (flights, float(area or 0.0))
                      for unit_id, flights, area in period_rows}

    # Last flight EVER received per machine -- deliberately unfiltered.
    last_rows = (db.session.query(DroneFlight.drone_unit_id,
                                  func.max(DroneFlight.started_at))
                 .group_by(DroneFlight.drone_unit_id).all())
    last_by_unit = {unit_id: last for unit_id, last in last_rows}

    today = _drone_today_local()
    current_status = _drone_current_status_rows()
    rows = []
    for unit in DroneUnit.query.order_by(DroneUnit.number).all():
        flights, area = period_by_unit.get(unit.id, (0, 0.0))
        last_utc = last_by_unit.get(unit.id)
        last_local = ((last_utc + DRONE_DISPLAY_UTC_OFFSET).date()
                      if last_utc else None)
        status_row = current_status.get(unit.id)
        rows.append({
            'unit': unit,
            'flights': flights,
            'area': area,
            'last_local': last_local,
            'silent_days': ((today - last_local).days
                            if last_local is not None else None),
            'status': status_row.status if status_row else None,
        })

    # Flights whose machine is still unresolved are their own visible line for
    # the same reason as on the summary: they are received data, and a page
    # about what arrived must not drop them.
    unattributed = period_by_unit.get(None)

    logs = (DroneSyncLog.query
            .order_by(DroneSyncLog.started_at.desc(), DroneSyncLog.id.desc())
            .limit(20).all())
    log_rows = []
    for log in logs:
        seen = log.records_seen or 0
        new = log.records_new or 0
        duplicate = log.records_duplicate or 0
        error = log.records_error or 0
        unresolved = log.records_unresolved or 0
        # [REASON]: the invariant is only DISPLAYED here. Nothing on this page
        # writes to drone_sync_logs, and the ingest keeps producing exactly
        # what it produced before -- seen = new + duplicate + error, with
        # unresolved a subset of new. Verification scripts depend on it, so
        # this screen reports a violation instead of repairing one.
        log_rows.append({
            'log': log,
            'holds': (seen == new + duplicate + error and unresolved <= new),
        })

    return render_template(
        'drones/sources.html',
        rows=rows,
        unattributed=unattributed,
        totals={
            'flights': sum(r['flights'] for r in rows)
                       + (unattributed[0] if unattributed else 0),
            'area': sum(r['area'] for r in rows)
                    + (unattributed[1] if unattributed else 0.0),
            'silent': sum(1 for r in rows if r['flights'] == 0),
            'never': sum(1 for r in rows if r['last_local'] is None),
        },
        log_rows=log_rows,
        filters=filters,
        link_args=_drone_link_args(dict(filters, unit_id=None, region='')),
        status_labels=_drone_status_labels(),
        today=today,
    )


def _drone_xlsx_safe(value):
    """Neutralize spreadsheet formula injection in a data-driven string.

    [REASON]: same defense as _xlsx_safe in spare_parts.py (SP-F-004) --
    Excel/LibreOffice execute a cell starting with = + - @ as a formula.
    Nicknames and reverse-geocoded addresses come from the DJI cloud, i.e.
    outside this system; the fix is one helper and it also stops Excel from
    mangling ordinary values. Applied only to data-driven strings; fixed
    labels and numeric columns are left untouched.
    """
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped and stripped[0] in ('=', '+', '-', '@'):
            return "'" + value
    return value


def _drone_xlsx_rate(value):
    """None -> em dash in a spreadsheet cell, exactly as on the screen.

    [REASON]: a workbook cell is where a zero does the most damage -- it is
    summed, averaged and charted downstream. An undefined rate (no flight
    time recorded, or no sprayed area) must arrive as text nobody can add up,
    not as a number that means "zero litres per hectare".
    """
    return '—' if value is None else value


def _drone_operator_derived_note():
    return _drone_t(
        'Оператор бириктиришлар жадвалидан келтириб чиқарилган, DJI '
        'маълумотларидан олинган эмас: бутун парк битта аккаунт остида '
        'учади. Бириктириш даври нотўғри бўлса, гектарлар бошқа одамга '
        'ёзилади.',
        'Оператор выведен из таблицы назначений, а не получен из данных DJI: '
        'весь парк летает под одним аккаунтом. Если период назначения задан '
        'неверно, гектары попадут на другого человека.')


def _drone_operator_sheet(wb, st, by_operator):
    """The «По операторам» sheet, identical in both workbooks.

    Carries the derived-attribution note in its first row, because the task
    requires every export showing an operator cut to state it visibly, and a
    workbook that leaves the caveat on the web page arrives without it.
    """
    ws = wb.create_sheet(_drone_t('Операторлар бўйича', 'По операторам'))
    ws.append([_drone_operator_derived_note()])
    ws.append([_drone_t('Оператор', 'Оператор'),
               _drone_t('Парвозлар', 'Вылеты'),
               _drone_t('Гектар', 'Гектары'),
               _drone_t('Улуш, %', 'Доля, %')])
    for r in by_operator['rows']:
        ws.append([_drone_xlsx_safe(r['name']), r['flights'], r['area'],
                   r['share']])
    # Both special rows are written whenever they carry anything, and they are
    # part of the total -- never a remainder outside it.
    if by_operator['undetermined']['flights']:
        u = by_operator['undetermined']
        ws.append([_drone_t('Оператор аниқланмаган', 'Оператор не определён'),
                   u['flights'], u['area'], u['share']])
    if by_operator['multiple']['flights']:
        m = by_operator['multiple']
        ws.append([_drone_t('Бир нечта оператор', 'Несколько операторов'),
                   m['flights'], m['area'], m['share']])
    ws.append([_drone_t('Жами', 'Итого'), by_operator['total']['flights'],
               by_operator['total']['area'], 100.0])
    # The note occupies row 1, so the header the styler freezes is row 2.
    st.style_table(ws, num_formats={3: '0.00', 4: '0.0'},
                   bold_rows=(ws.max_row,), header_row=2)
    return ws


def _drone_xlsx_styler():
    """Shared openpyxl styling for the drones workbooks.

    Modeled on _spare_report_styler (spare_parts.py): header fill and bold
    font, thin borders, frozen header row, auto column widths. Local on
    purpose -- excel_export.py is the eight-sheet daily activity machinery
    and is not a fit here.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from types import SimpleNamespace

    header_fill = PatternFill('solid', fgColor='D9EAD3')
    header_font = Font(bold=True)
    total_font = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_table(ws, num_formats=None, bold_rows=(), datetime_format=None,
                    header_row=1):
        """num_formats: {column index: number format} for numeric cells;
        datetime_format: {column index: format} for datetime cells.

        header_row says which row carries the column names. It is 1 for every
        sheet but the operator cut, which puts the derived-attribution note
        above the header -- the caveat has to be the first thing read, and a
        styler that assumed row 1 would bold the note and freeze the header
        out of view.
        """
        ws.freeze_panes = 'A%d' % (header_row + 1)
        ws.sheet_view.showGridLines = False
        num_formats = num_formats or {}
        datetime_format = datetime_format or {}
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                                max_col=ws.max_column):
            for cell in row:
                if (cell.column in num_formats
                        and isinstance(cell.value, (int, float))
                        and not isinstance(cell.value, bool)):
                    cell.number_format = num_formats[cell.column]
                elif (cell.column in datetime_format
                        and isinstance(cell.value, datetime)):
                    cell.number_format = datetime_format[cell.column]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column):
            for cell in row:
                cell.border = border
        for row_idx in bold_rows:
            for cell in ws[row_idx]:
                cell.font = total_font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = 10
            for cell in ws[letter]:
                val = '' if cell.value is None else str(cell.value)
                width = max(width, min(len(val) + 2, 42))
            ws.column_dimensions[letter].width = width

    return SimpleNamespace(style_table=style_table, total_font=total_font)


def _drone_xlsx_response(wb, base_name, filters):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    if filters['date_from_s'] or filters['date_to_s']:
        fname = '%s_%s_%s.xlsx' % (base_name,
                                   filters['date_from_s'] or 'all',
                                   filters['date_to_s'] or 'all')
    else:
        fname = '%s_all.xlsx' % base_name
    return send_file(
        buffer,
        as_attachment=True,
        download_name=fname,
        mimetype=('application/vnd.openxmlformats-officedocument'
                  '.spreadsheetml.sheet'),
    )


@drones_bp.route('/summary.xlsx')
@module_required('drones')
def summary_xlsx():
    """Five-sheet workbook of the summary page, same filters, same order."""
    from openpyxl import Workbook

    filters = _drone_filters_from_args(request.args,
                                       default_current_month=True)
    data = _drone_summary_data(_drone_flight_conditions(filters))
    st = _drone_xlsx_styler()

    label_total = _drone_t('Жами', 'Итого')
    label_unattr = _drone_t('Аниқланмаган', 'Не распознано')
    label_noregion = _drone_t('Вилоят аниқланмаган', 'Область не определена')
    head_flights = _drone_t('Парвозлар', 'Вылеты')
    head_area = _drone_t('Гектар', 'Гектары')
    head_share = _drone_t('Улуш, %', 'Доля, %')
    head_ha_hour = _drone_t('Га/соат', 'Га/час')
    head_l_ha = _drone_t('Л/га', 'Л/га')
    unbounded = _drone_t('чекланмаган', 'не ограничен')

    wb = Workbook()

    # 1. Сводка / Жамланма
    ws = wb.active
    ws.title = _drone_t('Жамланма', 'Сводка')
    ws.append([_drone_t('Кўрсаткич', 'Показатель'),
               _drone_t('Қиймат', 'Значение')])
    summary_rows = [
        (_drone_t('Давр: бошланиши', 'Период: с'),
         filters['date_from_s'] or unbounded, None),
        (_drone_t('Давр: охири', 'Период: по'),
         filters['date_to_s'] or unbounded, None),
        (_drone_t('Парвозлар', 'Вылетов'), data['totals']['flights'], None),
        (_drone_t('Гектар', 'Гектаров'), data['totals']['area_ha'], '0.00'),
        (_drone_t('Ҳавода соат', 'Часов в воздухе'),
         data['totals']['hours'], '0.00'),
        (_drone_t('Литр эритма', 'Литров раствора'),
         data['totals']['spray_liters'], '0.00'),
        (_drone_t('Кг уруғ', 'Кг посева'), data['totals']['sow_kg'], '0.000'),
        (_drone_t('Ноль майдонли парвозлар', 'Вылетов с нулевой площадью'),
         data['totals']['zero_area'], None),
    ]
    for label, value, fmt in summary_rows:
        ws.append([label, value])
        if fmt:
            ws.cell(row=ws.max_row, column=2).number_format = fmt
    st.style_table(ws)

    # 2. По машинам / Машиналар бўйича
    # DRONE-FLEET-001: two productivity columns appended at the END. Existing
    # columns keep their position and their business meaning -- the file goes
    # to operators and to accounting.
    ws = wb.create_sheet(_drone_t('Машиналар бўйича', 'По машинам'))
    ws.append([_drone_t('Машина (№)', 'Машина (№)'), head_flights,
               head_area, head_share, head_ha_hour, head_l_ha])
    for r in data['by_machine']['rows']:
        ws.append([r['number'], r['flights'], r['area'], r['share'],
                   _drone_xlsx_rate(r['ha_per_hour']),
                   _drone_xlsx_rate(r['liters_per_ha'])])
    if data['by_machine']['unattributed']:
        u = data['by_machine']['unattributed']
        ws.append([label_unattr, u['flights'], u['area'], u['share'],
                   _drone_xlsx_rate(u['ha_per_hour']),
                   _drone_xlsx_rate(u['liters_per_ha'])])
    ws.append([label_total, data['by_machine']['total']['flights'],
               data['by_machine']['total']['area'], 100.0,
               _drone_xlsx_rate(data['by_machine']['total']['ha_per_hour']),
               _drone_xlsx_rate(data['by_machine']['total']['liters_per_ha'])])
    st.style_table(ws, num_formats={3: '0.00', 4: '0.0', 5: '0.00',
                                    6: '0.00'},
                   bold_rows=(ws.max_row,))

    # 3. По областям / Вилоятлар бўйича
    ws = wb.create_sheet(_drone_t('Вилоятлар бўйича', 'По областям'))
    ws.append([_drone_t('Вилоят', 'Область'), head_flights,
               head_area, head_share])
    for r in data['by_region']['rows']:
        # The same label helper as the screen -- an export that disagrees with
        # the page it was taken from is a support ticket waiting to happen.
        # _drone_xlsx_safe still wraps it: an unknown region passes through as
        # its raw string, which came from outside this system.
        ws.append([_drone_xlsx_safe(r['label']), r['flights'], r['area'],
                   r['share']])
    if data['by_region']['undetermined']:
        u = data['by_region']['undetermined']
        ws.append([label_noregion, u['flights'], u['area'], u['share']])
    ws.append([label_total, data['by_region']['total']['flights'],
               data['by_region']['total']['area'], 100.0])
    st.style_table(ws, num_formats={3: '0.00', 4: '0.0'},
                   bold_rows=(ws.max_row,))

    # 4. По типам работ / Иш турлари бўйича
    ws = wb.create_sheet(_drone_t('Иш турлари бўйича', 'По типам работ'))
    ws.append([_drone_t('Иш тури', 'Тип работы'), head_flights, head_area,
               head_share, _drone_t('Литр', 'Литров'),
               _drone_t('Килограмм', 'Килограммов')])
    for r in data['by_usage']['rows']:
        ws.append([r['label'], r['flights'], r['area'], r['share'],
                   r['spray_liters'], r['sow_kg']])
    ws.append([label_total, data['by_usage']['total']['flights'],
               data['by_usage']['total']['area'], 100.0, None, None])
    st.style_table(ws, num_formats={3: '0.00', 4: '0.0', 5: '0.00',
                                    6: '0.000'},
                   bold_rows=(ws.max_row,))

    # 5. По месяцам / Ойлар бўйича
    ws = wb.create_sheet(_drone_t('Ойлар бўйича', 'По месяцам'))
    ws.append([_drone_t('Ой', 'Месяц'), head_flights, head_area, head_share])
    for r in data['by_month']['rows']:
        ws.append([r['month'], r['flights'], r['area'], r['share']])
    ws.append([label_total, data['by_month']['total']['flights'],
               data['by_month']['total']['area'], 100.0])
    st.style_table(ws, num_formats={3: '0.00', 4: '0.0'},
                   bold_rows=(ws.max_row,))

    # 6. По операторам / Операторлар бўйича (DRONE-FLEET-001)
    _drone_operator_sheet(wb, st, data['by_operator'])

    return _drone_xlsx_response(wb, 'drones_summary', filters)


@drones_bp.route('/flights.xlsx')
@module_required('drones')
def flights_xlsx():
    """Flat one-sheet export of the flight list, honouring its filters."""
    from openpyxl import Workbook

    filters = _drone_filters_from_args(request.args,
                                       default_current_month=False)
    conds = _drone_flight_conditions(filters)

    total = (db.session.query(func.count(DroneFlight.id))
             .filter(*conds).scalar() or 0)
    if total > DRONE_FLIGHTS_XLSX_CAP:
        # [REASON]: a silently truncated file reads as complete data on the
        # other side of an e-mail -- refuse with a clear message instead.
        flash(_drone_t(
            'Экспорт жуда катта: %d та парвоз, чегара — %d. '
            'Даврни торайтиринг.',
            'Экспорт слишком велик: %d вылетов при пределе %d. '
            'Сузьте период.') % (total, DRONE_FLIGHTS_XLSX_CAP), 'warning')
        return redirect(url_for('drones.index', **_drone_link_args(filters)))

    flights = (DroneFlight.query.filter(*conds)
               .options(joinedload(DroneFlight.drone_unit))
               .order_by(DroneFlight.started_at.desc())
               .all())
    usage_labels = _drone_usage_labels()
    # DRONE-FLEET-001: the per-flight operator, resolved by the SAME rule and
    # the SAME query the summary cut uses -- one implementation, so the flat
    # export and the cut can never disagree about who flew a flight.
    operator_by_flight = _drone_operator_by_flight(conds)
    total_area = float(db.session.query(
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0))
        .filter(*conds).scalar() or 0.0)

    wb = Workbook()
    ws = wb.active
    ws.title = _drone_t('Парвозлар', 'Вылеты')
    ws.append([
        _drone_t('Сана ва вақт (Тошкент)', 'Дата и время (UTC+5)'),
        _drone_t('Машина (№)', 'Машина (№)'),
        _drone_t('Ник (DJI)', 'Ник (DJI)'),
        _drone_t('Вилоят', 'Область'),
        _drone_t('Манзил', 'Адрес'),
        _drone_t('Гектар', 'Гектары'),
        _drone_t('Дақиқа', 'Минуты'),
        _drone_t('Литр', 'Литры'),
        _drone_t('Килограмм', 'Килограммы'),
        _drone_t('Иш тури', 'Тип работы'),
        'DJI id',
        # The caveat rides in the column NAME: a flat sheet has nowhere else
        # to put it, and this column must never be read as observed fact.
        _drone_t('Оператор (бириктиришдан келтирилган)',
                 'Оператор (выведен по назначениям)'),
        _drone_t('Га/соат', 'Га/час'),
        _drone_t('Л/га', 'Л/га'),
    ])
    for f in flights:
        usage_label = usage_labels.get(f.usage_type)
        if usage_label is None and f.usage_type is not None:
            usage_label = str(f.usage_type)
        hours = (f.work_seconds or 0) / 3600.0
        ws.append([
            (f.started_at + DRONE_DISPLAY_UTC_OFFSET)
            if f.started_at else None,
            f.drone_unit.number if f.drone_unit else None,
            _drone_xlsx_safe(f.nickname_raw),
            _drone_xlsx_safe(_drone_region_label(f.region)),
            _drone_xlsx_safe(f.location_text),
            f.area_ha,
            (f.work_seconds / 60.0) if f.work_seconds is not None else None,
            f.spray_liters,
            f.sow_kg,
            usage_label,
            f.dji_flight_id,
            _drone_xlsx_safe(operator_by_flight.get(f.id, '')),
            _drone_xlsx_rate(_drone_rate(f.area_ha or 0.0, hours)),
            _drone_xlsx_rate(_drone_rate(f.spray_liters or 0.0, f.area_ha)),
        ])
    st = _drone_xlsx_styler()
    st.style_table(ws,
                   num_formats={6: '0.00', 7: '0.0', 8: '0.00', 9: '0.000',
                                13: '0.00', 14: '0.00'},
                   datetime_format={1: 'DD.MM.YYYY HH:MM'})

    # The same operator cut sheet as summary.xlsx, over the same filters --
    # the flat list answers "which flight", the cut answers "how much whose",
    # and a reader who has only this file needs both.
    _drone_operator_sheet(wb, st, _drone_operator_cut(conds, total_area))
    return _drone_xlsx_response(wb, 'drones_flights', filters)


# ─── DRONE-WORKS-001: the work ledger ────────────────────────────────────────

# [REASON]: the commercial side of the drone operation has lived in the
# dispatchers' Excel books since September 2025 -- about 700 rows, 526 farms,
# 12 767 ha. Everything below reads drone_works, which stores the OPERATOR and
# never the machine: the sheets name the man, and the machine is derived from
# operator + date through drone_operator_assignments. Nothing here reads or
# writes drone_flights, and no route in this section touches the ingest.

DRONE_WORKS_PAGE_SIZE = 50

# [REASON]: DRONE-CUSTOMERS-PAGE-SLOW-001 -- 713 customers after the
# 2026-08-05 import. Same 50 as everywhere else in this project.
DRONE_CUSTOMERS_PAGE_SIZE = 50

# Same reason as DRONE_FLIGHTS_XLSX_CAP: a silently truncated workbook reads
# as complete data on the other side of an e-mail.
DRONE_WORKS_XLSX_CAP = 20000

# Query-string sentinel for «not resolved». -1 rather than an empty string,
# because an empty value already means «no filter» in every other picker on
# this module and reusing it would make the two indistinguishable.
DRONE_WORK_UNRESOLVED = -1

# [REASON]: DRONE-UI-CUSTOMER-LABEL-001. A second sentinel, because a NULL
# drone_customer_id means two different things and one filter option cannot
# ask for both. -1 NARROWS here to «a spelling exists and matched no alias»,
# which is the case somebody has to act on by adding an alias; -2 is «the
# sheet carried no farm name at all», which is how the holding's own land is
# written down and needs nothing done. The report cut has always split them --
# see _drone_works_report_data() -- and the ledger filter now agrees with it.
DRONE_WORK_CUSTOMER_UNSTATED_FILTER = -2

# Group keys of the two customer service rows. Strings, because the group
# expression has to produce one column type and the ordinary keys are customer
# ids rendered with printf.
DRONE_WORK_CUSTOMER_UNSTATED = 'unstated'
DRONE_WORK_CUSTOMER_UNRESOLVED = 'unresolved'


def _drone_customer_service_labels():
    """The two labels a NULL drone_customer_id can carry -- and they are two.

    [REASON]: DRONE-UI-CUSTOMER-LABEL-001. These strings existed inline in
    _drone_works_report_data() and the ledger cell carried a THIRD spelling of
    its own, «Заказчик не определён», applied to both cases at once. Every job
    with no customer was therefore reported as an alias somebody must add, and
    on the production data all 8 such jobs were the other kind -- a person
    reading the ledger went looking for eight aliases that must not be created.
    One function owns the wording now, so the cell, the filter option and the
    report cut cannot drift apart again.
    """
    return {
        DRONE_WORK_CUSTOMER_UNSTATED: _drone_t('Буюртмачи кўрсатилмаган',
                                               'Заказчик не указан'),
        DRONE_WORK_CUSTOMER_UNRESOLVED: _drone_t('Буюртмачи аниқланмаган',
                                                 'Заказчик не определён'),
    }


def _drone_work_customer_unstated_cond():
    """«The sheet carried no farm name at all.»

    [REASON]: byte-for-byte the same expression the customer_group CASE uses
    in _drone_works_report_data(). If the two ever diverge, a job could be
    counted in one service row by the report and returned by the other filter
    on the ledger, and nothing would report the disagreement.
    """
    return func.trim(func.coalesce(DroneWork.customer_raw, '')) == ''


def _drone_payment_labels():
    return {
        DRONE_WORK_PAYMENT_CASH: _drone_t('Нақд', 'Наличные'),
        DRONE_WORK_PAYMENT_TRANSFER: _drone_t('Справка', 'Справка'),
        DRONE_WORK_PAYMENT_INTERNAL: _drone_t('Тизим корхонаси',
                                              'Своё предприятие'),
        # [REASON]: not «прочее» and not an empty label. The sheet did not say,
        # and the row is waiting for somebody to say -- which is a different
        # statement from «this job was paid in some fourth way».
        DRONE_WORK_PAYMENT_UNKNOWN: _drone_t('Тўлов тури кўрсатилмаган',
                                             'Тип оплаты не указан'),
    }


def _drone_received_kind_labels():
    return {
        DRONE_RECEIVED_KIND_RECEIVED: _drone_t('Кирим қилинган',
                                               'Кирим қилинган — получено'),
        DRONE_RECEIVED_KIND_OPERATOR_DUE: _drone_t(
            'Оператор топшириши керак',
            'Оператор топшириши керак — оператор ещё должен'),
    }


def _drone_customer_type_labels():
    return {
        DRONE_CUSTOMER_EXTERNAL: _drone_t('Ташқи', 'Внешний'),
        DRONE_CUSTOMER_INTERNAL: _drone_t('Ички', 'Внутренний'),
    }


def _drone_work_month_expr():
    """'YYYY-MM' of a job: its own date when it has one, the month otherwise.

    [REASON]: period_month is NOT NULL precisely so a job with no date is
    still attributable to a month -- 34 rows of ~708 carry no date at all and
    would otherwise fall out of every monthly report. A job that DOES carry a
    date is grouped by that date, never by the manifest month: one book is
    filed as 2026-03 while its rows run 17.03--25.04, and filing April work in
    March is the error this expression exists to avoid.
    """
    return func.coalesce(func.strftime('%Y-%m', DroneWork.work_date_from),
                         DroneWork.period_month)


def _drone_work_outstanding_expr():
    """amount - received, both defaulting to zero. «олинмаган пуллар»."""
    return (func.coalesce(DroneWork.amount, 0.0)
            - func.coalesce(DroneWork.received_amount, 0.0))


def _drone_num_arg(value):
    """A money/area field from a form: float, or None when the field is blank.

    [REASON]: blank must become NULL and never 0.0. A zero price means «this
    farm was charged nothing per hectare», which is a different and false
    statement from «the price was not written down», and the two must stay
    distinguishable in a ledger somebody bills from.
    """
    text = (value or '').strip()
    if not text:
        return None
    text = (text.replace('\xa0', '').replace(' ', '').replace(' ', '')
            .replace(',', '.'))
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _drone_works_filters_from_args(args):
    """Parse the ledger filter set. Every value round-trips into link_args."""
    period = (args.get('period') or '').strip()
    if not re.match(r'^\d{4}-\d{2}$', period or ''):
        period = ''
    payment = (args.get('payment') or '').strip()
    if payment not in DRONE_WORK_PAYMENT_TYPES:
        payment = ''
    resolved = (args.get('resolved') or '').strip()
    if resolved not in ('resolved', 'unresolved'):
        resolved = ''
    return {
        'period': period,
        'operator_id': args.get('operator_id', type=int),
        'customer_id': args.get('customer_id', type=int),
        'payment': payment,
        'subdivision': (args.get('subdivision') or '').strip(),
        'resolved': resolved,
        'q': (args.get('q') or '').strip(),
    }


def _drone_work_conditions(filters):
    """SQLAlchemy conditions for the parsed ledger filters."""
    conds = []
    if filters['period']:
        conds.append(_drone_work_month_expr() == filters['period'])
    if filters['operator_id'] == DRONE_WORK_UNRESOLVED:
        conds.append(DroneWork.drone_operator_id.is_(None))
    elif filters['operator_id']:
        conds.append(DroneWork.drone_operator_id == filters['operator_id'])
    # [REASON]: DRONE-UI-CUSTOMER-LABEL-001. -1 NARROWED here: it used to mean
    # every NULL customer and now means only the ones with a spelling that
    # matched no alias. On the production data of 2026-08-06 that turns 8 rows
    # into 0, and the 8 move to -2. That is the fix, not a regression: the 8
    # were never aliases waiting to be created.
    if filters['customer_id'] == DRONE_WORK_UNRESOLVED:
        conds.append(and_(DroneWork.drone_customer_id.is_(None),
                          ~_drone_work_customer_unstated_cond()))
    elif filters['customer_id'] == DRONE_WORK_CUSTOMER_UNSTATED_FILTER:
        conds.append(and_(DroneWork.drone_customer_id.is_(None),
                          _drone_work_customer_unstated_cond()))
    elif filters['customer_id']:
        conds.append(DroneWork.drone_customer_id == filters['customer_id'])
    if filters['payment']:
        conds.append(DroneWork.payment_type == filters['payment'])
    if filters['subdivision']:
        conds.append(DroneWork.subdivision_name == filters['subdivision'])
    if filters['resolved'] == 'resolved':
        conds.append(and_(DroneWork.drone_customer_id.isnot(None),
                          DroneWork.drone_operator_id.isnot(None)))
    elif filters['resolved'] == 'unresolved':
        conds.append(or_(DroneWork.drone_customer_id.is_(None),
                         DroneWork.drone_operator_id.is_(None)))
    if filters['q']:
        # The verbatim spelling is what people remember and search for; the
        # canonical name lives on the customer row and is searched too.
        like = '%%%s%%' % filters['q']
        conds.append(or_(DroneWork.customer_raw.ilike(like),
                         DroneWork.operator_raw.ilike(like),
                         DroneWork.note.ilike(like)))
    return conds


def _drone_works_link_args(filters):
    """Filters as URL parameters, empty ones dropped -- shared by every link."""
    args = {}
    if filters['period']:
        args['period'] = filters['period']
    if filters['operator_id']:
        args['operator_id'] = filters['operator_id']
    if filters['customer_id']:
        args['customer_id'] = filters['customer_id']
    if filters['payment']:
        args['payment'] = filters['payment']
    if filters['subdivision']:
        args['subdivision'] = filters['subdivision']
    if filters['resolved']:
        args['resolved'] = filters['resolved']
    if filters['q']:
        args['q'] = filters['q']
    return args


def _drone_work_periods():
    """Every month present in the ledger, newest first."""
    month = _drone_work_month_expr()
    return [row[0] for row in
            db.session.query(month).distinct().order_by(month.desc()).all()
            if row[0]]


def _drone_work_subdivisions():
    return [row[0] for row in
            db.session.query(DroneWork.subdivision_name)
            .filter(DroneWork.subdivision_name.isnot(None))
            .distinct().order_by(DroneWork.subdivision_name).all()]


def _drone_works_totals(conds):
    row = db.session.query(
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.area_ha), 0.0),
        func.coalesce(func.sum(DroneWork.amount), 0.0),
        func.coalesce(func.sum(DroneWork.received_amount), 0.0),
        func.coalesce(func.sum(DroneWork.other_costs), 0.0),
    ).filter(*conds).one()
    return {
        'jobs': row[0] or 0,
        'area': float(row[1] or 0.0),
        'amount': float(row[2] or 0.0),
        'received': float(row[3] or 0.0),
        'other_costs': float(row[4] or 0.0),
        'outstanding': float(row[2] or 0.0) - float(row[3] or 0.0),
    }


def _drone_works_pickers():
    """The three pickers every works screen shares."""
    return {
        'operators': DroneOperator.query.order_by(
            DroneOperator.is_active.desc(), DroneOperator.full_name).all(),
        'customers': DroneCustomer.query.order_by(
            DroneCustomer.is_active.desc(), DroneCustomer.name).all(),
        'periods': _drone_work_periods(),
        'subdivisions': _drone_work_subdivisions(),
        'payment_labels': _drone_payment_labels(),
    }


@drones_bp.route('/works')
@module_required('drones')
def works():
    """The work ledger: one row per job, with the filters the owner asked for."""
    page = request.args.get('page', 1, type=int) or 1
    if page < 1:
        page = 1
    filters = _drone_works_filters_from_args(request.args)
    conds = _drone_work_conditions(filters)

    query = (DroneWork.query.filter(*conds)
             .options(joinedload(DroneWork.drone_operator),
                      joinedload(DroneWork.drone_customer))
             # [REASON]: ordered by the EFFECTIVE month, which is never NULL,
             # then by the date inside it. Sorting on work_date_from alone
             # would need NULLS LAST -- unsupported on SQLite below 3.30 --
             # and would scatter the 34 undated jobs to one end of the whole
             # ledger instead of keeping them inside their own month.
             .order_by(_drone_work_month_expr().desc(),
                       DroneWork.work_date_from.desc(),
                       DroneWork.id.desc()))
    total = db.session.query(func.count(DroneWork.id)).filter(*conds).scalar() or 0
    pages = max(1, (total + DRONE_WORKS_PAGE_SIZE - 1) // DRONE_WORKS_PAGE_SIZE)
    if page > pages:
        page = pages
    rows = (query.offset((page - 1) * DRONE_WORKS_PAGE_SIZE)
            .limit(DRONE_WORKS_PAGE_SIZE).all())

    context = _drone_works_pickers()
    return render_template(
        'drones/works.html',
        rows=rows,
        filters=filters,
        link_args=_drone_works_link_args(filters),
        # The totals are over the WHOLE filtered set, not the page: a page
        # total on a paginated ledger is a number people quote by accident.
        totals=_drone_works_totals(conds),
        page=page,
        pages=pages,
        total=total,
        unresolved_key=DRONE_WORK_UNRESOLVED,
        # DRONE-UI-CUSTOMER-LABEL-001: the ledger renders the two customer
        # labels from the same function the report cut uses, so that the two
        # screens cannot disagree about what a NULL customer means.
        unstated_key=DRONE_WORK_CUSTOMER_UNSTATED_FILTER,
        customer_labels=_drone_customer_service_labels(),
        customer_unstated_group=DRONE_WORK_CUSTOMER_UNSTATED,
        customer_unresolved_group=DRONE_WORK_CUSTOMER_UNRESOLVED,
        price_suggestions=DRONE_WORK_PRICE_SUGGESTIONS,
        payment_types=DRONE_WORK_PAYMENT_TYPES,
        **context)


def _drone_work_redirect(filters=None):
    args = _drone_works_link_args(filters) if filters else {}
    return redirect(url_for('drones.works', **args))


def _drone_work_form_values(row=None):
    """(values dict, error message). Shared by add and edit.

    [REASON]: the manual form is the one place a price may be SUGGESTED --
    200 000 so'm/ha for cash and certificate work, 85 633 for the holding's
    own land. Both are prefilled and both are overridable, because nine other
    prices occur in the real books and one internal sheet quotes 76 458. The
    import must never do this; a price it invented would be
    indistinguishable from a price the farm agreed.
    """
    period = (request.form.get('period_month') or '').strip()
    date_from = _drone_parse_date(request.form.get('work_date_from'))
    date_to = _drone_parse_date(request.form.get('work_date_to'))
    customer_raw = (request.form.get('customer_raw') or '').strip()
    area = _drone_num_arg(request.form.get('area_ha'))
    payment = (request.form.get('payment_type') or '').strip()

    if not customer_raw:
        return None, _drone_t('Буюртмачи номи бўш бўлиши мумкин эмас.',
                              'Название заказчика не может быть пустым.')
    if area is None or area <= 0:
        return None, _drone_t('Майдон нолдан катта бўлиши керак.',
                              'Площадь должна быть больше нуля.')
    if payment not in DRONE_WORK_PAYMENT_TYPES:
        return None, _drone_t('Тўлов тури танланмаган.',
                              'Тип оплаты не выбран.')
    if date_to is not None and date_from is not None and date_to < date_from:
        return None, _drone_t('«Санагача» «Санадан»дан олдин бўлиши мумкин эмас.',
                              '«Дата по» не может быть раньше «Даты с».')
    if date_to is not None and date_from is None:
        return None, _drone_t('«Санагача» бор, «Санадан» йўқ — даврнинг '
                              'боши кўрсатилмаган.',
                              'Указана «Дата по» без «Даты с» — начало '
                              'периода не задано.')
    if not re.match(r'^\d{4}-\d{2}$', period or ''):
        # [REASON]: period_month is NOT NULL and is what keeps an undated job
        # inside its month's report. It is derived from the date when there
        # is one, so the operator does not have to type it twice.
        if date_from is not None:
            period = date_from.strftime('%Y-%m')
        else:
            return None, _drone_t(
                'Давр (ЙЙЙЙ-ОО) кўрсатилиши шарт: санаси йўқ иш ҳам ўз '
                'ойининг ҳисоботида қолиши керак.',
                'Период (ГГГГ-ММ) обязателен: работа без даты всё равно '
                'должна остаться в отчёте своего месяца.')

    customer_id = request.form.get('drone_customer_id', type=int) or None
    operator_id = request.form.get('drone_operator_id', type=int) or None
    if customer_id and DroneCustomer.query.get(customer_id) is None:
        return None, _drone_t('Буюртмачи топилмади.', 'Заказчик не найден.')
    if operator_id and DroneOperator.query.get(operator_id) is None:
        return None, _drone_t('Оператор топилмади.', 'Оператор не найден.')

    return {
        'period_month': period,
        'work_date_from': date_from,
        'work_date_to': date_to if date_to is not None else date_from,
        'date_raw': (request.form.get('date_raw') or '').strip() or None,
        'drone_operator_id': operator_id,
        'operator_raw': (request.form.get('operator_raw') or '').strip() or None,
        'drone_customer_id': customer_id,
        'customer_raw': customer_raw,
        'area_ha': area,
        'price_per_ha': _drone_num_arg(request.form.get('price_per_ha')),
        'amount': _drone_num_arg(request.form.get('amount')),
        'other_costs': _drone_num_arg(request.form.get('other_costs')),
        'received_amount': _drone_num_arg(request.form.get('received_amount')),
        'payment_type': payment,
        'subdivision_name': (request.form.get('subdivision_name') or '').strip()
                            or None,
        'note': (request.form.get('note') or '').strip() or None,
    }, None


@drones_bp.route('/works/add', methods=['POST'])
@module_required('drones')
def works_add():
    """A job typed by hand. Not everything comes out of a file."""
    if not current_user.can_edit:
        abort(403)
    filters = _drone_works_filters_from_args(request.args)
    values, error = _drone_work_form_values()
    if error:
        flash(error, 'warning')
        return _drone_work_redirect(filters)

    row = DroneWork(created_by=(current_user.id
                                if getattr(current_user, 'is_authenticated',
                                           False) else None),
                    **values)
    db.session.add(row)
    db.session.commit()
    flash(_drone_t('Иш қўшилди: «%s», %.2f га.', 'Работа добавлена: «%s», %.2f га.')
          % (values['customer_raw'], values['area_ha']), 'success')
    return _drone_work_redirect(filters)


@drones_bp.route('/works/<int:work_id>/update', methods=['POST'])
@module_required('drones')
def works_update(work_id):
    """Edit one job.

    [REASON]: customer_raw stays editable but its ORIGINAL value is the
    evidence behind the resolution, so the import never rewrites it and a
    hand edit of an imported row is a deliberate act by a person who can see
    the source file name in the same table.
    """
    if not current_user.can_edit:
        abort(403)
    row = DroneWork.query.get(work_id)
    if row is None:
        abort(404)
    filters = _drone_works_filters_from_args(request.args)
    values, error = _drone_work_form_values(row)
    if error:
        flash(error, 'warning')
        return _drone_work_redirect(filters)
    for key, value in values.items():
        setattr(row, key, value)
    db.session.commit()
    flash(_drone_t('Иш янгиланди: «%s».', 'Работа обновлена: «%s».')
          % values['customer_raw'], 'success')
    return _drone_work_redirect(filters)


# ─── DRONE-WORKS-001: the customer directory and its aliases ─────────────────

def _drone_customer_key(value):
    """normalized_key of a farm spelling: case-folded, whitespace-collapsed.

    [REASON]: the SAME rule tools/import_drone_works.py writes, spelled out
    here rather than imported, because tools/ is not a package and drones.py
    must not depend on it at request time. The two are pinned together by a
    test that compares them on the same inputs -- if one drifts, the alias the
    screen creates stops matching the alias the import looks up, and every
    re-import quietly creates a second customer.
    """
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().casefold()


# [REASON]: DRONE-CUSTOMERS-PAGE-SLOW-001. The SAME fold
# tools/import_drone_works.py applies to operator names, spelled out here
# rather than imported: tools/ is not a package and drones.py must not depend
# on it at request time. A test compares the two on the same inputs, so a
# drift shows up as a failing test rather than as a search that quietly stops
# finding «Фурқат» when somebody typed «Фуркат».
#
# This is the SEARCH key, not the alias key. _drone_customer_key() -- case and
# whitespace only -- is what gets STORED in drone_customer_aliases and must
# never change: the screen has to be able to reproduce it by eye. Search may
# be more forgiving than storage; storage may not be more forgiving than the
# import.
_DRONE_UZ_FOLD = {
    'Қ': 'К', 'Ғ': 'Г', 'Ҳ': 'Х', 'Ў': 'У', 'Ҷ': 'Ж', 'Ҝ': 'К',
    'қ': 'к', 'ғ': 'г', 'ҳ': 'х', 'ў': 'у', 'ҷ': 'ж',
    'Ъ': '', 'ъ': '', 'Ь': '', 'ь': '', 'Ё': 'Е', 'ё': 'е',
    '\u0451': 'е',
}
_DRONE_UZ_TABLE = {ord(k): (v or None) for k, v in _DRONE_UZ_FOLD.items()}


def _drone_search_key(value):
    """Uzbek fold, lower-cased, whitespace collapsed. Search only.

    Collapsing whitespace is what lets «Гурганжи» find «Гурганжи Чорва
    (Завкиддин)» -- the real row has two spaces inside it.
    """
    if value is None:
        return ''
    return re.sub(r'\s+', ' ',
                  str(value).translate(_DRONE_UZ_TABLE).lower()).strip()


def _drone_customers_redirect(customer_id=None):
    if customer_id:
        return redirect(url_for('drones.customers', focus=customer_id))
    return redirect(url_for('drones.customers'))


def _drone_alias_maps():
    """{normalized_key: customer id} for resolution. isnot(False), never == True.

    [REASON]: is_active is BOOLEAN DEFAULT 1 and a hand-inserted or legacy row
    can hold NULL. `== True` silently drops those rows, turning a known
    spelling into an unresolved one and, on the next import, into a duplicate
    customer. The identical bug was already caught once in
    _drone_nickname_maps(); it is not being reintroduced here.
    """
    mapping = {}
    for alias in DroneCustomerAlias.query.filter(
            DroneCustomerAlias.is_active.isnot(False)).all():
        mapping[alias.normalized_key] = alias.drone_customer_id
    return mapping


@drones_bp.route('/customers')
@module_required('drones')
def customers():
    """The drone service's own customer directory and every spelling of it.

    [REASON]: a directory SEPARATE from `customers`, by the owner's decision
    of 2026-08-03. Nothing on this screen joins to `customers` and no row is
    moved between them; merging the two by INN is a separate task, and a
    foreign key added now would presume its outcome.
    """
    rows = (DroneCustomer.query
            .order_by(DroneCustomer.is_active.desc(), DroneCustomer.name)
            .all())
    aliases = {}
    for alias in (DroneCustomerAlias.query
                  .order_by(DroneCustomerAlias.raw_name).all()):
        aliases.setdefault(alias.drone_customer_id, []).append(alias)

    # [REASON]: DRONE-CUSTOMERS-PAGE-SLOW-001. The 2026-08-05 import created
    # 713 customers and the page rendered all of them, each with its alias
    # list and its own save form; the browser QA agent failed to load it four
    # times and gave up. Filtering happens in PYTHON, not in SQL, because the
    # match has to run over the FOLDED name AND every folded alias spelling,
    # and neither fold exists inside SQLite. 713 rows and their aliases are
    # already in memory by this point -- the filter costs one pass over a list
    # that was loaded either way.
    query_text = (request.args.get('q') or '').strip()
    matched = rows
    if query_text:
        needle = _drone_search_key(query_text)
        matched = [c for c in rows
                   if needle in _drone_search_key(c.name)
                   or any(needle in _drone_search_key(a.raw_name)
                          for a in aliases.get(c.id, ()))]

    page = request.args.get('page', 1, type=int) or 1
    if page < 1:
        page = 1
    pages = max(1, ((len(matched) + DRONE_CUSTOMERS_PAGE_SIZE - 1)
                    // DRONE_CUSTOMERS_PAGE_SIZE))
    if page > pages:
        page = pages
    start = (page - 1) * DRONE_CUSTOMERS_PAGE_SIZE
    page_rows = matched[start:start + DRONE_CUSTOMERS_PAGE_SIZE]

    # Jobs and hectares per customer, in one query -- the lazy='dynamic'
    # relationship would give one query per customer and there are ~526.
    usage = {}
    for customer_id, jobs, area in db.session.query(
            DroneWork.drone_customer_id,
            func.count(DroneWork.id),
            func.coalesce(func.sum(DroneWork.area_ha), 0.0)
    ).group_by(DroneWork.drone_customer_id).all():
        usage[customer_id] = {'jobs': jobs, 'area': float(area or 0.0)}

    unresolved = db.session.query(
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.area_ha), 0.0)
    ).filter(DroneWork.drone_customer_id.is_(None)).one()

    return render_template(
        'drones/customers.html',
        # The page, and separately EVERY customer. The pickers -- «Добавить
        # написание» and the per-alias repoint select -- must list all of
        # them, or an alias could not be moved to a customer on page 7.
        customers=page_rows,
        all_customers=rows,
        aliases=aliases,
        usage=usage,
        # [REASON]: counted over the WHOLE table, never the page. This banner
        # is how somebody notices there is work to do at all; a per-page count
        # would read as «almost nothing to fix» on page 1 of 15.
        unresolved={'jobs': unresolved[0] or 0,
                    'area': float(unresolved[1] or 0.0)},
        type_labels=_drone_customer_type_labels(),
        customer_types=DRONE_CUSTOMER_TYPES,
        focus=request.args.get('focus', type=int),
        query_text=query_text,
        page=page,
        pages=pages,
        total=len(rows),
        matched_total=len(matched),
    )


@drones_bp.route('/customers/add', methods=['POST'])
@module_required('drones')
def customers_add():
    """Add a customer and, in the same transaction, its first alias.

    [REASON]: a customer with no alias resolves nothing -- the import looks
    up spellings, not names. Creating the two together is what makes the
    screen and the import agree from the first row.
    """
    if not current_user.can_edit:
        abort(403)
    name = (request.form.get('name') or '').strip()
    if not name:
        flash(_drone_t('Буюртмачи номи бўш бўлиши мумкин эмас.',
                       'Название заказчика не может быть пустым.'), 'warning')
        return _drone_customers_redirect()

    customer_type = (request.form.get('customer_type') or '').strip()
    if customer_type not in DRONE_CUSTOMER_TYPES:
        customer_type = DRONE_CUSTOMER_EXTERNAL

    key = _drone_customer_key(name)
    existing = _drone_alias_maps().get(key)
    if existing is not None:
        other = DroneCustomer.query.get(existing)
        flash(_drone_t(
            'Бундай ёзилиш аллақачон «%s» буюртмачисига боғланган. '
            'Иккинчи ёзувни яратиш ўрнига мавжуд алиасни кўчиринг.',
            'Такое написание уже привязано к заказчику «%s». Перенесите '
            'существующий алиас, а не создавайте вторую запись.')
            % (other.name if other else '?'), 'warning')
        return _drone_customers_redirect(existing)

    row = DroneCustomer(
        name=name,
        inn=(request.form.get('inn') or '').strip() or None,
        customer_type=customer_type,
        note=(request.form.get('note') or '').strip() or None,
        is_active=True)
    db.session.add(row)
    db.session.flush()
    db.session.add(DroneCustomerAlias(
        raw_name=name, normalized_key=key, drone_customer_id=row.id,
        is_active=True))
    db.session.commit()
    flash(_drone_t('«%s» буюртмачиси ва унинг биринчи ёзилиши қўшилди.',
                   'Заказчик «%s» и его первое написание добавлены.')
          % name, 'success')
    return _drone_customers_redirect(row.id)


@drones_bp.route('/customers/<int:customer_id>/update', methods=['POST'])
@module_required('drones')
def customers_update(customer_id):
    """Edit a customer, or deactivate it. There is deliberately no delete.

    [REASON]: a customer that stops being served still owns the hectares it
    was billed for. Deleting the row would orphan every work row that
    explains an already-sent report; deactivating keeps the history readable
    and takes it out of the pickers.
    """
    if not current_user.can_edit:
        abort(403)
    row = DroneCustomer.query.get(customer_id)
    if row is None:
        abort(404)
    name = (request.form.get('name') or '').strip()
    if not name:
        flash(_drone_t('Буюртмачи номи бўш бўлиши мумкин эмас.',
                       'Название заказчика не может быть пустым.'), 'warning')
        return _drone_customers_redirect(customer_id)
    customer_type = (request.form.get('customer_type') or '').strip()
    if customer_type not in DRONE_CUSTOMER_TYPES:
        customer_type = row.customer_type or DRONE_CUSTOMER_EXTERNAL

    row.name = name
    row.inn = (request.form.get('inn') or '').strip() or None
    row.customer_type = customer_type
    row.note = (request.form.get('note') or '').strip() or None
    row.is_active = request.form.get('is_active') is not None
    db.session.commit()
    if row.is_active:
        flash(_drone_t('«%s» янгиланди.', '«%s» обновлён.') % name, 'success')
    else:
        flash(_drone_t(
            '«%s» фаолсизлантирилди: у танлаш рўйхатларида кўринмайди, аммо '
            'ишлари ва гектарлари сақланиб қолади.',
            '«%s» деактивирован: он не появляется в списках выбора, но его '
            'работы и гектары сохраняются.') % name, 'success')
    return _drone_customers_redirect(customer_id)


@drones_bp.route('/customers/aliases/add', methods=['POST'])
@module_required('drones')
def customer_alias_add():
    if not current_user.can_edit:
        abort(403)
    customer_id = request.form.get('drone_customer_id', type=int)
    customer = DroneCustomer.query.get(customer_id) if customer_id else None
    raw_name = (request.form.get('raw_name') or '').strip()
    if customer is None:
        flash(_drone_t('Буюртмачи танланмаган ёки топилмади.',
                       'Заказчик не выбран или не найден.'), 'warning')
        return _drone_customers_redirect()
    if not raw_name:
        flash(_drone_t('Ёзилиш бўш бўлиши мумкин эмас.',
                       'Написание не может быть пустым.'), 'warning')
        return _drone_customers_redirect(customer_id)

    key = _drone_customer_key(raw_name)
    existing = DroneCustomerAlias.query.filter_by(normalized_key=key).first()
    if existing is not None:
        other = DroneCustomer.query.get(existing.drone_customer_id)
        # [REASON]: name the customer it already points at instead of letting
        # the unique index raise a 500. Which customer holds the spelling is
        # exactly what the person is trying to find out.
        flash(_drone_t(
            'Бундай ёзилиш аллақачон «%s» буюртмачисида. Икки буюртмачини '
            'бирлаштириш — алиасни кўчириш, иш қаторларини таҳрирлаш эмас.',
            'Такое написание уже у заказчика «%s». Слияние двух заказчиков — '
            'это перенос алиаса, а не правка строк работ.')
            % (other.name if other else '?'), 'warning')
        return _drone_customers_redirect(existing.drone_customer_id)

    db.session.add(DroneCustomerAlias(
        raw_name=raw_name, normalized_key=key, drone_customer_id=customer.id,
        is_active=True))
    db.session.commit()
    flash(_drone_t('«%s» ёзилиши «%s»га боғланди.',
                   'Написание «%s» привязано к «%s».')
          % (raw_name, customer.name), 'success')
    return _drone_customers_redirect(customer.id)


@drones_bp.route('/customers/aliases/<int:alias_id>/update', methods=['POST'])
@module_required('drones')
def customer_alias_update(alias_id):
    """Repoint a spelling at another customer, or switch it off.

    [REASON]: THIS IS HOW TWO CUSTOMERS ARE MERGED -- by repointing the
    alias, never by editing work rows. customer_raw on a work row is the
    evidence behind its resolution and rewriting it destroys the only record
    of what the sheet actually said.
    """
    if not current_user.can_edit:
        abort(403)
    alias = DroneCustomerAlias.query.get(alias_id)
    if alias is None:
        abort(404)
    target_id = request.form.get('drone_customer_id', type=int)
    target = DroneCustomer.query.get(target_id) if target_id else None
    if target is None:
        flash(_drone_t('Буюртмачи танланмаган ёки топилмади.',
                       'Заказчик не выбран или не найден.'), 'warning')
        return _drone_customers_redirect(alias.drone_customer_id)

    moved = target.id != alias.drone_customer_id
    alias.drone_customer_id = target.id
    alias.is_active = request.form.get('is_active') is not None
    db.session.commit()

    if moved:
        flash(_drone_t('«%s» ёзилиши «%s»га кўчирилди. Иш қаторлари '
                       'ўзгартирилмади: улар келгуси импортда қайта '
                       'боғланади.',
                       'Написание «%s» перенесено к «%s». Строки работ не '
                       'изменены: они перепривяжутся при следующем импорте.')
              % (alias.raw_name, target.name), 'success')
    elif alias.is_active:
        flash(_drone_t('«%s» ёзилиши ёқилди.', 'Написание «%s» включено.')
              % alias.raw_name, 'success')
    else:
        flash(_drone_t(
            '«%s» ёзилиши ўчирилди: у янги импортда танилмайди, аммо '
            'аллақачон сақланган ишлар ўзгармайди.',
            'Написание «%s» отключено: оно не распознаётся при новом '
            'импорте, а уже сохранённые работы не меняются.')
          % alias.raw_name, 'success')
    return _drone_customers_redirect(target.id)


# ─── DRONE-WORKS-001: reports ────────────────────────────────────────────────

def _drone_work_cut(conds, group_expr, total_row, service_labels,
                    label_map=None, order_by_area=True):
    """One breakdown of the ledger, with its service buckets as visible rows.

    service_labels is an ORDERED sequence of (key, label): every group whose
    key is one of them becomes a service row instead of an ordinary one, in
    that order, and is still part of the total.

    [REASON]: the rule the flight summary already obeys and this must match --
    a group-by that simply drops NULL silently omits exactly the rows nobody
    has resolved yet, which is how the previous system lost 2 036 flights and
    2 082 hectares into an invisible bucket. So the missing bucket gets a row,
    the row is part of the total, and the total is compared against the grand
    total; a mismatch is shown on the page (reconciled=False), never hidden.

    [REASON]: SEVERAL service rows, not one. «Заказчик не указан» -- the sheet
    carried no farm name at all, which is what all nine internal jobs of
    «Агрокластер Дрон маълумот МАЙ.xlsx» look like -- and «Заказчик не
    определён» -- a spelling that exists but matched no alias -- are different
    facts needing different actions. The first is how the holding's own land
    is written down and needs nothing; the second is an alias somebody has to
    add. One combined row would hide the second inside the first.
    """
    groups = db.session.query(
        group_expr,
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.area_ha), 0.0),
        func.coalesce(func.sum(DroneWork.amount), 0.0),
        func.coalesce(func.sum(DroneWork.received_amount), 0.0),
        # [REASON]: DRONE-REPORT-ZERO-AMOUNT-001. SUM over a group whose every
        # amount is NULL is 0.0 after the coalesce above, and the cut then
        # reports «this farm was charged nothing» -- a different and false
        # statement from «the price was never written down». 15 jobs on
        # production carry no price and no amount. This counts them so the
        # screen can tell an empty column from a zero one; the sum itself is
        # NOT changed, because the totals reconcile against it.
        func.sum(case((DroneWork.amount.is_(None), 1), else_=0)),
        # [REASON]: DRONE-ZERO-VS-UNKNOWN-001. received_amount carries the
        # same NULL-versus-zero distinction as amount, and the NULLs are real:
        # of the 904 imported jobs 852 carry a received figure and 52 do not.
        #
        # [REASON]: DRONE-RECEIVED-BLANK-IS-ZERO-001, and read this before
        # deleting the column. no_received is COMPUTED AND DELIBERATELY NOT
        # RENDERED. On 2026-08-07 the owner decided that a blank «олинмаган
        # пуллар» cell in the dispatcher sheets means NOTHING WAS COLLECTED,
        # not «nobody wrote it down», so «Получено» prints 0 and never a dash
        # and never a note. The counter stays because it is a real measurement
        # that costs one SUM over a column already being scanned, and because
        # the decision is the owner's to reverse -- if he does, the number is
        # already here and only the rendering changes back.
        #
        # A counter computed and never shown is normally a smell. This one is
        # the exception, and it says so here so that nobody removes it in six
        # months believing they are cleaning up an oversight -- or, worse,
        # restores the dash believing they are fixing one.
        func.sum(case((DroneWork.received_amount.is_(None), 1), else_=0)),
    ).filter(*conds).group_by(group_expr).all()

    service_map = dict(service_labels)
    rows = []
    found = {}
    for key, jobs, area, amount, received, no_amount, no_received in groups:
        cell = {
            'key': key,
            'jobs': jobs,
            'area': float(area or 0.0),
            'amount': float(amount or 0.0),
            'received': float(received or 0.0),
            'no_amount': int(no_amount or 0),
            'no_received': int(no_received or 0),
        }
        cell['outstanding'] = cell['amount'] - cell['received']
        cell['share'] = _drone_share(cell['area'], total_row['area'])
        if key in service_map:
            cell['label'] = service_map[key]
            found[key] = cell
        else:
            cell['label'] = (label_map or {}).get(key, key)
            rows.append(cell)

    if order_by_area:
        rows.sort(key=lambda r: r['area'], reverse=True)
    else:
        rows.sort(key=lambda r: str(r['key']))

    services = [found[key] for key, _label in service_labels if key in found]
    everything = rows + services
    total = {
        'jobs': sum(r['jobs'] for r in everything),
        'area': sum(r['area'] for r in everything),
        'amount': sum(r['amount'] for r in everything),
        'received': sum(r['received'] for r in everything),
        'no_amount': sum(r['no_amount'] for r in everything),
        'no_received': sum(r['no_received'] for r in everything),
    }
    total['outstanding'] = total['amount'] - total['received']
    reconciled = (total['jobs'] == total_row['jobs']
                  and abs(total['area'] - total_row['area']) < 0.005
                  and abs(total['amount'] - total_row['amount']) < 0.005
                  and abs(total['received'] - total_row['received']) < 0.005)
    return {'rows': rows, 'services': services, 'total': total,
            'reconciled': reconciled}


def _drone_work_bucket_row(label, rows):
    """One collapsed service line over `rows`, or None when there are none.

    Every key an ordinary row carries is named here. A counter that is not
    named does not reach the line, and nothing raises -- see the note on
    no_amount below.
    """
    if not rows:
        return None
    # [REASON]: DRONE-ZERO-VS-UNKNOWN-001. This dict is assembled by naming its
    # keys, so every counter _drone_work_cut() adds to an ordinary row has to
    # be named here too or it silently does not reach the service row.
    # no_amount was added by DRONE-REPORT-ZERO-AMOUNT-001 and never arrived;
    # nothing rendered it, so nothing failed. The moment the debts template
    # started calling money_cell() this row would have printed a sum where the
    # ordinary rows print a dash.
    #
    # [REASON]: DRONE-RECEIVED-BLANK-IS-ZERO-001. no_received is summed here
    # and DELIBERATELY NOT RENDERED anywhere -- the owner decided on
    # 2026-08-07 that a blank «олинмаган пуллар» means nothing was collected,
    # so «Получено» never dashes. It is kept because it is a real measurement
    # and because reversing that decision must not need a new query. Do not
    # drop it as dead weight; see the same note in _drone_work_cut().
    row = {
        'label': label,
        'jobs': sum(r['jobs'] for r in rows),
        'area': sum(r['area'] for r in rows),
        'amount': sum(r['amount'] for r in rows),
        'received': sum(r['received'] for r in rows),
        'no_amount': sum(r['no_amount'] for r in rows),
        'no_received': sum(r['no_received'] for r in rows),
    }
    row['outstanding'] = row['amount'] - row['received']
    return row


def _drone_work_debt_cut(cut, no_debt_label, unknown_label):
    """The same cut, re-read as money: who owes, how much, and who cannot say.

    [REASON]: rows with nothing outstanding are collapsed into ONE visible
    service line rather than dropped, so the column still sums to the grand
    total. A debt table that silently omits the settled rows reads as "this is
    everybody" and there is no way to tell it apart from a filter mistake.

    [REASON]: DRONE-ZERO-VS-UNKNOWN-001 -- THREE buckets, not two. A group
    whose price was never written down has amount = 0, received = 0 and
    therefore outstanding = 0, so two buckets put it under «Долга нет
    (остальные)» and the page stated that the farm owes nothing. That is not a
    wrong number, it is a wrong sentence: the debt is unknown, and unknown is
    not zero. It gets its own line.

    ORDER OF EVALUATION IS LOAD-BEARING: unknown is tested BEFORE settled, or
    every unknown group stays exactly where it was. owing is tested first and
    is unchanged; it cannot collide with unknown, because a group with every
    amount NULL sums to 0 and its outstanding can never exceed 0.005.
    """
    everything = cut['rows'] + list(cut['services'])
    owing, unknown, settled = [], [], []
    # One pass, so «which bucket» is decided in exactly one place and every
    # row lands in exactly one of them by construction rather than by three
    # comprehensions agreeing with each other.
    for row in everything:
        if row['outstanding'] > 0.005:
            owing.append(row)
        elif row['jobs'] and row['jobs'] == row['no_amount']:
            unknown.append(row)
        else:
            settled.append(row)
    owing.sort(key=lambda r: r['outstanding'], reverse=True)
    return {'rows': owing,
            'unknown': _drone_work_bucket_row(unknown_label, unknown),
            'rest': _drone_work_bucket_row(no_debt_label, settled),
            'total': cut['total'],
            'reconciled': cut['reconciled']}


def _drone_works_report_data(conds):
    """Four cuts, a debt view, and the grand total they all reconcile with."""
    totals = _drone_works_totals(conds)

    customer_names = {c.id: c.name for c in DroneCustomer.query.all()}
    operator_names = {o.id: o.full_name for o in DroneOperator.query.all()}

    # [REASON]: the customer cut has TWO service rows, and the key is a CASE
    # expression rather than the raw column, because a NULL drone_customer_id
    # means two different things. «Заказчик не указан» is a sheet that carried
    # no farm name -- how the holding's own land is written down, and nothing
    # needs doing about it. «Заказчик не определён» is a spelling that exists
    # and matched no alias -- somebody has to add the alias. One combined row
    # would hide the second inside the first.
    customer_group = case(
        (DroneWork.drone_customer_id.isnot(None),
         func.printf('%d', DroneWork.drone_customer_id)),
        (_drone_work_customer_unstated_cond(), DRONE_WORK_CUSTOMER_UNSTATED),
        else_=DRONE_WORK_CUSTOMER_UNRESOLVED)
    # The wording now comes from _drone_customer_service_labels(), which the
    # ledger cell and the ledger filter read too -- see DRONE-UI-CUSTOMER-
    # LABEL-001. The strings are unchanged; only their single owner is new.
    service_labels = _drone_customer_service_labels()
    by_customer = _drone_work_cut(
        conds, customer_group, totals,
        ((DRONE_WORK_CUSTOMER_UNSTATED,
          service_labels[DRONE_WORK_CUSTOMER_UNSTATED]),
         (DRONE_WORK_CUSTOMER_UNRESOLVED,
          service_labels[DRONE_WORK_CUSTOMER_UNRESOLVED])),
        label_map={('%d' % c.id): c.name for c in DroneCustomer.query.all()})
    by_operator = _drone_work_cut(
        conds, DroneWork.drone_operator_id, totals,
        ((None, _drone_t('Оператор аниқланмаган', 'Оператор не определён')),),
        label_map=operator_names)
    by_subdivision = _drone_work_cut(
        conds, DroneWork.subdivision_name, totals,
        ((None, _drone_t('Бўлинма кўрсатилмаган',
                         'Подразделение не указано')),))
    # [REASON]: the payment cut exists because 'unknown' does. Roughly two
    # fifths of the corpus has no payment type in the source, and a report
    # that never shows the bucket makes it invisible to the person who is
    # supposed to fill it in on the works screen.
    by_payment = _drone_work_cut(
        conds, DroneWork.payment_type, totals,
        ((DRONE_WORK_PAYMENT_UNKNOWN,
          _drone_t('Тўлов тури кўрсатилмаган', 'Тип оплаты не указан')),),
        label_map=_drone_payment_labels(), order_by_area=False)

    # [REASON]: the month cut groups on the job's OWN date, and jobs with no
    # date get the «Дата не указана» row instead of vanishing. Grouping every
    # row by period_month instead would look tidier and would file the April
    # work of a March-labelled book in March -- the exact error the manifest
    # note about «Имомов Бехзод Пешку ПТЗ.xlsx» warns about. The service row
    # then carries a per-period breakdown, which is what period_month is FOR:
    # an undated job is still attributable to a month, just not to a day.
    by_month = _drone_work_cut(
        conds, func.strftime('%Y-%m', DroneWork.work_date_from), totals,
        ((None, _drone_t('Сана кўрсатилмаган', 'Дата не указана')),),
        order_by_area=False)
    if by_month['services']:
        breakdown = db.session.query(
            DroneWork.period_month,
            func.count(DroneWork.id),
            func.coalesce(func.sum(DroneWork.area_ha), 0.0),
        ).filter(*conds).filter(DroneWork.work_date_from.is_(None)) \
            .group_by(DroneWork.period_month) \
            .order_by(DroneWork.period_month).all()
        by_month['services'][0]['periods'] = [
            {'period': period, 'jobs': jobs, 'area': float(area or 0.0)}
            for period, jobs, area in breakdown]

    no_debt = _drone_t('Қарзи йўқ (қолганлари)', 'Долга нет (остальные)')
    # [REASON]: DRONE-ZERO-VS-UNKNOWN-001. Its own wording, because the row
    # says something «Долга нет» cannot: not «this farm owes nothing» but «we
    # cannot say what this farm owes, because the price was never written
    # down». The two are different sentences and the second one names the
    # action -- somebody has to enter the amount on the works screen.
    debt_unknown = _drone_t('Қарз номаълум — сумма ёзилмаган',
                            'Долг неизвестен — сумма не записана')
    # [REASON]: what the debt column is actually made of. «Кирим қилинган» is
    # money handed in; «Оператор топшириши керак» is what the operator still
    # owes and is NOT a collection. Both live in received_amount, so the debt
    # tables carry a memo saying how much of the «получено» column is the
    # second kind. It changes no total -- it says what the totals contain.
    operator_due = db.session.query(
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.received_amount), 0.0),
    ).filter(*conds).filter(
        DroneWork.received_kind == DRONE_RECEIVED_KIND_OPERATOR_DUE).one()

    return {
        'totals': totals,
        'by_customer': by_customer,
        'by_operator': by_operator,
        'by_subdivision': by_subdivision,
        'by_month': by_month,
        'by_payment': by_payment,
        'operator_due': {'jobs': operator_due[0] or 0,
                         'amount': float(operator_due[1] or 0.0)},
        'debt_by_customer': _drone_work_debt_cut(by_customer, no_debt,
                                                 debt_unknown),
        'debt_by_operator': _drone_work_debt_cut(by_operator, no_debt,
                                                 debt_unknown),
    }


# ─── DRONE-REPORTS-HUB-001: the reports launcher ─────────────────────────────

# [REASON]: five drone reports had accumulated in five places with no index.
# The owner went looking for the debt tables on 2026-08-05 and could not find
# them. Spare parts solved exactly this with /spare-parts/reports -- a grid of
# tiles, one per report -- so this copies that pattern rather than inventing a
# second one. Every class it uses already exists in design-system.css; not one
# rule is added or changed.
#
# The launcher is deliberately DATA-FREE: no filters, no query, no export
# button. Those live inside each report, the same division spare parts makes.
#
# `endpoint` is resolved with url_for at render time, so a renamed route
# breaks the page loudly instead of producing a dead tile.
DRONE_REPORT_TILES = (
    {
        'key': 'flights-summary',
        'endpoint': 'drones.summary',
        'accent': 'is-primary',
        'title_ru': 'Сводка по вылетам',
        'title_uz': 'Парвозлар жамланмаси',
        'subtitle_ru': 'Машины, области, типы работ, месяцы и операторы '
                       'за период',
        'subtitle_uz': 'Давр бўйича машиналар, вилоятлар, иш турлари, ойлар '
                       'ва операторлар',
    },
    {
        'key': 'works',
        'endpoint': 'drones.works_reports',
        'accent': 'is-info',
        'title_ru': 'Работы — заказчики и операторы',
        'title_uz': 'Ишлар — буюртмачилар ва операторлар',
        'subtitle_ru': 'Гектары и суммы по ведомостям диспетчеров, '
                       'пять разрезов',
        'subtitle_uz': 'Диспетчерлар ведомостлари бўйича гектар ва суммалар, '
                       'бешта кесим',
    },
    {
        'key': 'debts',
        'endpoint': 'drones.works_debts',
        'accent': 'is-warning',
        'title_ru': 'Долги',
        'title_uz': 'Қарзлар',
        'subtitle_ru': 'Кто и сколько не заплатил — по заказчикам '
                       'и по операторам',
        'subtitle_uz': 'Ким қанча тўламаган — буюртмачилар ва операторлар '
                       'бўйича',
    },
    {
        'key': 'assignment-hints',
        'endpoint': 'drones.works_assignment_hints',
        'accent': 'is-success',
        'title_ru': 'Подсказка по назначениям',
        'title_uz': 'Бириктиришлар бўйича маслаҳат',
        'subtitle_ru': 'Гектары оператора рядом с гектарами машины '
                       'за тот же месяц',
        'subtitle_uz': 'Операторнинг гектарлари ўша ойдаги машина гектарлари '
                       'ёнида',
    },
    {
        # DRONE-ANALYTICS-001/1. The only tile on a NEW accent: five accents
        # existed and all five were already in use, so a sixth report either
        # repeats a colour or brings one rule. --vs-purple / --vs-purple-bg
        # were already defined (design-system.css line 53) and unused by any
        # tile; the rule that binds them is additive and modifies nothing.
        'key': 'spray',
        'endpoint': 'drones.spray_usage',
        'accent': 'is-purple',
        'title_ru': 'Расход рабочего раствора',
        'title_uz': 'Иш эритмаси сарфи',
        'subtitle_ru': 'Литры на гектар по машинам и месяцам, '
                       'с коридором вокруг медианы',
        'subtitle_uz': 'Машина ва ойлар бўйича гектарига литр, '
                       'медиана атрофида йўлак билан',
    },
    {
        'key': 'sources',
        'endpoint': 'drones.sources',
        'accent': 'is-danger',
        'title_ru': 'Источники данных',
        'title_uz': 'Маълумот манбалари',
        'subtitle_ru': 'Какая машина перестала присылать вылеты и что сделали '
                       'последние загрузки',
        'subtitle_uz': 'Қайси машина парвоз юборишни тўхтатган ва сўнгги '
                       'юклашлар нима қилган',
    },
)


@drones_bp.route('/reports')
@module_required('drones')
def reports():
    """Launcher: one tile per report. No filters, no data, no export.

    [REASON]: the route name drones.reports and the /reports URL are what
    _drones_nav.html links; keep both if this ever moves.
    """
    return render_template('drones/reports.html', tiles=DRONE_REPORT_TILES)


@drones_bp.route('/works/reports')
@module_required('drones')
def works_reports():
    """The five cuts, over the same filters as the ledger.

    The debt view moved to works_debts() in DRONE-REPORTS-HUB-001; this route
    still computes it, because _drone_works_report_data() is one function and
    splitting it would mean rewriting queries this task is not allowed to
    touch. The cost is one unused dict per request, accepted knowingly.
    """
    filters = _drone_works_filters_from_args(request.args)
    conds = _drone_work_conditions(filters)
    data = _drone_works_report_data(conds)
    context = _drone_works_pickers()
    return render_template(
        'drones/works_reports.html',
        data=data,
        filters=filters,
        link_args=_drone_works_link_args(filters),
        unresolved_key=DRONE_WORK_UNRESOLVED,
        payment_types=DRONE_WORK_PAYMENT_TYPES,
        **context)


@drones_bp.route('/works/debts')
@module_required('drones')
def works_debts():
    """The two debt tables, on their own page. «олинмаган пуллар».

    [REASON]: DRONE-REPORTS-HUB-001. These tables used to sit at the bottom of
    /drones/works/reports, below five long cuts; the owner went looking for
    them on 2026-08-05 and could not find them. Nothing is recomputed here --
    the same _drone_works_report_data() the reports page calls, rendered by
    the same markup -- so the figures on this page are the figures that were
    on that one, by construction rather than by coincidence.
    """
    filters = _drone_works_filters_from_args(request.args)
    conds = _drone_work_conditions(filters)
    data = _drone_works_report_data(conds)
    context = _drone_works_pickers()
    return render_template(
        'drones/works_debts.html',
        data=data,
        filters=filters,
        link_args=_drone_works_link_args(filters),
        unresolved_key=DRONE_WORK_UNRESOLVED,
        payment_types=DRONE_WORK_PAYMENT_TYPES,
        **context)


def _drone_works_xlsx_response(wb, base_name, filters):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fname = '%s_%s.xlsx' % (base_name, filters['period'] or 'all')
    return send_file(
        buffer,
        as_attachment=True,
        download_name=fname,
        mimetype=('application/vnd.openxmlformats-officedocument'
                  '.spreadsheetml.sheet'),
    )


def _drone_works_filter_sheet(wb, st, filters, totals):
    """First sheet: which filters produced these numbers, and the grand total.

    [REASON]: a workbook that arrives by e-mail without its filters is a set
    of numbers nobody can reproduce. The filters ride in the file.
    """
    ws = wb.active
    ws.title = _drone_t('Жамланма', 'Сводка')
    ws.append([_drone_t('Кўрсаткич', 'Показатель'),
               _drone_t('Қиймат', 'Значение')])
    unbounded = _drone_t('чекланмаган', 'не ограничен')
    payment_labels = _drone_payment_labels()
    rows = [
        (_drone_t('Давр', 'Период'), filters['period'] or unbounded, None),
        (_drone_t('Тўлов тури', 'Тип оплаты'),
         payment_labels.get(filters['payment'], unbounded), None),
        (_drone_t('Бўлинма', 'Подразделение'),
         filters['subdivision'] or unbounded, None),
        (_drone_t('Ишлар', 'Работ'), totals['jobs'], None),
        (_drone_t('Гектар', 'Гектаров'), totals['area'], '0.00'),
        (_drone_t('Сумма', 'Сумма'), totals['amount'], '0.00'),
        (_drone_t('Кирим қилинган', 'Получено'), totals['received'], '0.00'),
        (_drone_t('Олинмаган', 'Не получено'), totals['outstanding'], '0.00'),
        (_drone_t('Бошқа харажатлар', 'Прочие расходы'),
         totals['other_costs'], '0.00'),
    ]
    for label, value, fmt in rows:
        ws.append([label, _drone_xlsx_safe(value)])
        if fmt:
            ws.cell(row=ws.max_row, column=2).number_format = fmt
    st.style_table(ws)
    return ws


def _drone_money_or_blank(row, key, missing_key):
    """A money figure for a workbook cell, or None when it is not known.

    [REASON]: DRONE-ZERO-VS-UNKNOWN-001. The workbook is what goes to
    management, so a screen that prints «—» beside a workbook that prints 0
    fixes nothing -- the 0 is the copy that gets believed. This is the SAME
    rule money_cell() applies in Jinja, written once here so the two cannot
    drift: the figure is unknown exactly when every job in the group is
    missing the thing it is computed from.

    An EMPTY cell, never 0 and never a string. An empty cell is absent data in
    a numeric column; «—» would turn the column into text and break every SUM
    over it, and money in these workbooks stays a number.
    """
    if row['jobs'] and row[missing_key] == row['jobs']:
        return None
    return row[key]


def _drone_work_money_cells(row):
    """(amount, received, outstanding) for one row of a cut or a debt table.

    «Не получено» keys off no_amount, not off its own counter: the debt is
    unknown when the amount is unknown, whatever was received.

    [REASON]: DRONE-RECEIVED-BLANK-IS-ZERO-001. «Получено» is NOT blanked, and
    that is the OWNER'S DECISION of 2026-08-07, not something derived from the
    data: in the dispatcher sheets an empty «олинмаган пуллар» cell means
    NOTHING WAS COLLECTED, not «nobody wrote it down». DRONE-ZERO-VS-UNKNOWN-
    001 had treated it like a blank «Сумма» and printed a dash, which produced
    rows saying two things at once -- «Бухоро Сервис Агрокластер 3
    55 429 140 — 55 429 140», i.e. «collected: unknown; uncollected: all of
    it». Only «Сумма» and «Не получено» can be unknown now.
    """
    return (_drone_money_or_blank(row, 'amount', 'no_amount'),
            row['received'],
            _drone_money_or_blank(row, 'outstanding', 'no_amount'))


def _drone_work_cut_sheet(wb, st, title, first_column, cut):
    """One cut as one sheet: rows, the service row, and the total."""
    ws = wb.create_sheet(title)
    ws.append([first_column,
               _drone_t('Ишлар', 'Работ'),
               _drone_t('Гектар', 'Гектары'),
               _drone_t('Сумма', 'Сумма'),
               _drone_t('Кирим қилинган', 'Получено'),
               _drone_t('Олинмаган', 'Не получено'),
               _drone_t('Улуш, %', 'Доля, %')])
    for row in cut['rows']:
        ws.append([_drone_xlsx_safe(row['label']), row['jobs'], row['area']]
                  + list(_drone_work_money_cells(row)) + [row['share']])
    # The service rows are part of the table AND part of the total -- never a
    # remainder printed outside it.
    for service in cut['services']:
        ws.append([service['label'], service['jobs'], service['area']]
                  + list(_drone_work_money_cells(service))
                  + [service['share']])
    ws.append([_drone_t('Жами', 'Итого'), cut['total']['jobs'],
               cut['total']['area']]
              + list(_drone_work_money_cells(cut['total'])) + [100.0])
    st.style_table(ws, num_formats={3: '0.00', 4: '0.00', 5: '0.00',
                                    6: '0.00', 7: '0.0'},
                   bold_rows=(ws.max_row,))
    return ws


def _drone_work_debt_sheet(wb, st, title, first_column, debt):
    ws = wb.create_sheet(title)
    ws.append([first_column,
               _drone_t('Ишлар', 'Работ'),
               _drone_t('Сумма', 'Сумма'),
               _drone_t('Кирим қилинган', 'Получено'),
               _drone_t('Олинмаган', 'Не получено')])
    for row in debt['rows']:
        ws.append([_drone_xlsx_safe(row['label']), row['jobs']]
                  + list(_drone_work_money_cells(row)))
    # [REASON]: DRONE-ZERO-VS-UNKNOWN-001. The third bucket is a row of the
    # sheet for the same reason it is a row of the screen: the printed rows
    # have to add up to the printed total. Dropping it here would leave the
    # workbook short by exactly the groups nobody has priced -- the ones the
    # bucket exists to make visible.
    for bucket in (debt.get('unknown'), debt['rest']):
        if bucket is not None:
            ws.append([bucket['label'], bucket['jobs']]
                      + list(_drone_work_money_cells(bucket)))
    ws.append([_drone_t('Жами', 'Итого'), debt['total']['jobs']]
              + list(_drone_work_money_cells(debt['total'])))
    st.style_table(ws, num_formats={3: '0.00', 4: '0.00', 5: '0.00'},
                   bold_rows=(ws.max_row,))
    return ws


@drones_bp.route('/works/reports.xlsx')
@module_required('drones')
def works_reports_xlsx():
    """The four cuts and both debt views, one sheet each, same filters."""
    from openpyxl import Workbook

    filters = _drone_works_filters_from_args(request.args)
    data = _drone_works_report_data(_drone_work_conditions(filters))
    st = _drone_xlsx_styler()
    wb = Workbook()
    _drone_works_filter_sheet(wb, st, filters, data['totals'])
    _drone_work_cut_sheet(wb, st, _drone_t('Буюртмачилар бўйича',
                                           'По заказчикам'),
                          _drone_t('Буюртмачи', 'Заказчик'),
                          data['by_customer'])
    _drone_work_cut_sheet(wb, st, _drone_t('Операторлар бўйича',
                                           'По операторам'),
                          _drone_t('Оператор', 'Оператор'),
                          data['by_operator'])
    _drone_work_cut_sheet(wb, st, _drone_t('Бўлинмалар бўйича',
                                           'По подразделениям'),
                          _drone_t('Бўлинма', 'Подразделение'),
                          data['by_subdivision'])
    _drone_work_cut_sheet(wb, st, _drone_t('Ойлар бўйича', 'По месяцам'),
                          _drone_t('Ой', 'Месяц'), data['by_month'])
    _drone_work_cut_sheet(wb, st, _drone_t('Тўлов турлари бўйича',
                                           'По типам оплаты'),
                          _drone_t('Тўлов тури', 'Тип оплаты'),
                          data['by_payment'])
    _drone_work_debt_sheet(wb, st, _drone_t('Қарз — буюртмачилар',
                                            'Долги — заказчики'),
                           _drone_t('Буюртмачи', 'Заказчик'),
                           data['debt_by_customer'])
    _drone_work_debt_sheet(wb, st, _drone_t('Қарз — операторлар',
                                            'Долги — операторы'),
                           _drone_t('Оператор', 'Оператор'),
                           data['debt_by_operator'])
    return _drone_works_xlsx_response(wb, 'drone_works_reports', filters)


@drones_bp.route('/works/debt.xlsx')
@module_required('drones')
def works_debt_xlsx():
    """«олинмаган пуллар» alone: who owes, by customer and by operator."""
    from openpyxl import Workbook

    filters = _drone_works_filters_from_args(request.args)
    data = _drone_works_report_data(_drone_work_conditions(filters))
    st = _drone_xlsx_styler()
    wb = Workbook()
    _drone_works_filter_sheet(wb, st, filters, data['totals'])
    _drone_work_debt_sheet(wb, st, _drone_t('Қарз — буюртмачилар',
                                            'Долги — заказчики'),
                           _drone_t('Буюртмачи', 'Заказчик'),
                           data['debt_by_customer'])
    _drone_work_debt_sheet(wb, st, _drone_t('Қарз — операторлар',
                                            'Долги — операторы'),
                           _drone_t('Оператор', 'Оператор'),
                           data['debt_by_operator'])
    return _drone_works_xlsx_response(wb, 'drone_works_debt', filters)


@drones_bp.route('/works.xlsx')
@module_required('drones')
def works_xlsx():
    """Flat one-sheet export of the ledger, honouring its filters."""
    from openpyxl import Workbook

    filters = _drone_works_filters_from_args(request.args)
    conds = _drone_work_conditions(filters)
    total = db.session.query(func.count(DroneWork.id)).filter(*conds).scalar() or 0
    if total > DRONE_WORKS_XLSX_CAP:
        flash(_drone_t(
            'Экспорт жуда катта: %d та иш, чегара — %d. Даврни торайтиринг.',
            'Экспорт слишком велик: %d работ при пределе %d. Сузьте период.')
            % (total, DRONE_WORKS_XLSX_CAP), 'warning')
        return redirect(url_for('drones.works', **_drone_works_link_args(filters)))

    rows = (DroneWork.query.filter(*conds)
            .options(joinedload(DroneWork.drone_operator),
                     joinedload(DroneWork.drone_customer))
            .order_by(_drone_work_month_expr().desc(),
                      DroneWork.work_date_from.desc(), DroneWork.id.desc())
            .all())
    payment_labels = _drone_payment_labels()
    label_no_customer = _drone_t('Буюртмачи аниқланмаган',
                                 'Заказчик не определён')
    label_no_operator = _drone_t('Оператор аниқланмаган',
                                 'Оператор не определён')
    received_kind_labels = _drone_received_kind_labels()

    st = _drone_xlsx_styler()
    wb = Workbook()
    _drone_works_filter_sheet(wb, st, filters, _drone_works_totals(conds))
    ws = wb.create_sheet(_drone_t('Ишлар', 'Работы'))
    ws.append([
        _drone_t('Давр', 'Период'),
        _drone_t('Сана дан', 'Дата с'),
        _drone_t('Сана гача', 'Дата по'),
        _drone_t('Сана — матн', 'Дата — текст'),
        _drone_t('Оператор', 'Оператор'),
        _drone_t('Оператор — ёзилганидек', 'Оператор — как написано'),
        _drone_t('Буюртмачи', 'Заказчик'),
        _drone_t('Буюртмачи — ёзилганидек', 'Заказчик — как написано'),
        _drone_t('Гектар', 'Гектары'),
        _drone_t('1 га нархи', 'Цена за га'),
        _drone_t('Сумма', 'Сумма'),
        _drone_t('Бошқа харажатлар', 'Прочие расходы'),
        _drone_t('Кирим қилинган', 'Получено'),
        # [REASON]: without this column «получено» is two different facts in
        # one place -- money handed in, and money the operator still owes.
        _drone_t('«Получено» нимадан олинган', 'Что за «получено»'),
        _drone_t('Олинмаган', 'Не получено'),
        _drone_t('Тўлов тури', 'Тип оплаты'),
        _drone_t('Бўлинма', 'Подразделение'),
        _drone_t('Манба файл', 'Файл-источник'),
        _drone_t('Варақ', 'Лист'),
        _drone_t('Қатор', 'Строка'),
        _drone_t('Изоҳ', 'Примечание'),
    ])
    for w in rows:
        amount = float(w.amount or 0.0)
        received = float(w.received_amount or 0.0)
        ws.append([
            w.period_month,
            w.work_date_from, w.work_date_to,
            _drone_xlsx_safe(w.date_raw),
            _drone_xlsx_safe(w.drone_operator.full_name if w.drone_operator
                             else label_no_operator),
            _drone_xlsx_safe(w.operator_raw),
            _drone_xlsx_safe(w.drone_customer.name if w.drone_customer
                             else label_no_customer),
            _drone_xlsx_safe(w.customer_raw),
            w.area_ha, w.price_per_ha, w.amount, w.other_costs,
            w.received_amount,
            received_kind_labels.get(w.received_kind, ''),
            amount - received,
            payment_labels.get(w.payment_type, w.payment_type),
            _drone_xlsx_safe(w.subdivision_name),
            _drone_xlsx_safe(w.source_file),
            _drone_xlsx_safe(w.source_sheet),
            w.source_row,
            _drone_xlsx_safe(w.note),
        ])
    st.style_table(ws,
                   num_formats={9: '0.00', 10: '0.00', 11: '0.00',
                                12: '0.00', 13: '0.00', 15: '0.00'},
                   datetime_format={2: 'DD.MM.YYYY', 3: 'DD.MM.YYYY'})
    return _drone_works_xlsx_response(wb, 'drone_works', filters)


# ─── DRONE-WORKS-001: assignment hints ───────────────────────────────────────

def _drone_flight_month_expr():
    """'YYYY-MM' of a flight in the operators' timezone (UTC+5).

    Derived from DRONE_DISPLAY_UTC_OFFSET, never written as a literal -- the
    summary page groups months the same way, and a second site hardcoding the
    offset would keep grouping at the old value if the constant ever changed.
    """
    offset_minutes = int(DRONE_DISPLAY_UTC_OFFSET.total_seconds() // 60)
    return func.strftime('%Y-%m',
                         func.datetime(DroneFlight.started_at,
                                       '%+d minutes' % offset_minutes))


def _drone_month_bounds(month):
    """('YYYY-MM-01', last day) for an assignment-overlap test."""
    year, mon = int(month[:4]), int(month[5:7])
    first = datetime(year, mon, 1).date()
    if mon == 12:
        last = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last = datetime(year, mon + 1, 1).date() - timedelta(days=1)
    return first, last


def _drone_assigned_units_in_month(month):
    """{operator id: {unit ids}} for assignments overlapping the month."""
    first, last = _drone_month_bounds(month)
    rows = (DroneOperatorAssignment.query
            .filter(DroneOperatorAssignment.date_from <= last)
            .filter(or_(DroneOperatorAssignment.date_to.is_(None),
                        DroneOperatorAssignment.date_to >= first))
            .all())
    by_operator = {}
    by_unit = {}
    for row in rows:
        by_operator.setdefault(row.operator_id, set()).add(row.drone_unit_id)
        by_unit.setdefault(row.drone_unit_id, set()).add(row.operator_id)
    return by_operator, by_unit


def _drone_assignment_hints(month):
    """Operator hectares from the ledger against machine hectares from DJI.

    [REASON]: THIS SCREEN SUGGESTS AND CREATES NOTHING. It is the one place in
    the module where two independent measurements of the same month can be put
    side by side: what the dispatchers billed, and what the machines flew. On
    April 2026 the comparison put all twelve operators within +-9 %, the best
    being Zhuraev Tuygun on machine No 6 at 0.1 %. It also produced one
    confident wrong answer, for an operator who changed machines mid-month --
    which is precisely why a close match is evidence and not proof, and why
    the human decides. A screen that could write an assignment would turn that
    one wrong answer into a silently wrong report.
    """
    work_month = _drone_work_month_expr()
    ledger = (db.session.query(
        DroneWork.drone_operator_id,
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.area_ha), 0.0),
    ).filter(work_month == month)
        .filter(DroneWork.drone_operator_id.isnot(None))
        .group_by(DroneWork.drone_operator_id).all())

    flight_month = _drone_flight_month_expr()
    machines = (db.session.query(
        DroneFlight.drone_unit_id,
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
    ).filter(flight_month == month)
        .group_by(DroneFlight.drone_unit_id).all())

    unit_numbers = {u.id: u.number for u in DroneUnit.query.all()}
    operator_names = {o.id: o.full_name for o in DroneOperator.query.all()}
    assigned_by_operator, assigned_by_unit = _drone_assigned_units_in_month(
        month)

    pool = []
    unattributed = {'flights': 0, 'area': 0.0}
    for unit_id, flights, area in machines:
        area = float(area or 0.0)
        if unit_id is None:
            # [REASON]: a flight whose DJI nickname is not in the alias map
            # belongs to no machine and therefore cannot be a candidate. It is
            # shown as its own number rather than dropped, because a big
            # unattributed bucket is the reason a hint would look wrong.
            unattributed = {'flights': flights, 'area': area}
            continue
        pool.append({'unit_id': unit_id, 'number': unit_numbers.get(unit_id),
                     'flights': flights, 'area': area,
                     'taken_by': sorted(
                         operator_names.get(op_id, '?')
                         for op_id in assigned_by_unit.get(unit_id, ()))})

    rows = []
    for operator_id, jobs, area in ledger:
        area = float(area or 0.0)
        already = assigned_by_operator.get(operator_id, set())
        candidates = []
        for machine in pool:
            if machine['unit_id'] in already:
                continue
            delta = None if not area else (
                (machine['area'] - area) * 100.0 / area)
            candidates.append(dict(machine, delta=delta))
        candidates.sort(key=lambda c: (c['delta'] is None,
                                       abs(c['delta']) if c['delta'] is not None
                                       else 0.0))
        rows.append({
            'operator_id': operator_id,
            'name': operator_names.get(operator_id, '?'),
            'jobs': jobs,
            'area': area,
            'already': sorted(unit_numbers.get(u) for u in already),
            'best': candidates[0] if candidates else None,
            'second': candidates[1] if len(candidates) > 1 else None,
        })
    rows.sort(key=lambda r: r['area'], reverse=True)

    ledger_area = sum(r['area'] for r in rows)
    machine_area = sum(m['area'] for m in pool) + unattributed['area']
    unresolved = db.session.query(
        func.count(DroneWork.id),
        func.coalesce(func.sum(DroneWork.area_ha), 0.0)
    ).filter(work_month == month) \
        .filter(DroneWork.drone_operator_id.is_(None)).one()
    return {
        'rows': rows,
        'pool': sorted(pool, key=lambda m: (m['number'] is None,
                                            m['number'])),
        'unattributed': unattributed,
        'ledger_area': ledger_area,
        'machine_area': machine_area,
        'unresolved': {'jobs': unresolved[0] or 0,
                       'area': float(unresolved[1] or 0.0)},
    }


@drones_bp.route('/works/assignment-hints')
@module_required('drones')
def works_assignment_hints():
    """Read-only. Suggests a machine for an operator; creates nothing."""
    periods = _drone_work_periods()
    month = (request.args.get('month') or '').strip()
    if not re.match(r'^\d{4}-\d{2}$', month or ''):
        month = periods[0] if periods else ''
    data = _drone_assignment_hints(month) if month else None
    return render_template(
        'drones/works_assignment_hints.html',
        data=data,
        month=month,
        periods=periods,
    )


# ─── DRONE-ANALYTICS-001/1: working-solution consumption ─────────────────────

# [REASON]: the corridor half-width, in per cent of the median. 30 % is a
# starting point the owner can move with ?band=, NOT a measured tolerance --
# nobody has yet established what spread is normal for these treatments. It is
# named and printed on the page in words for exactly that reason: a threshold
# whose value is invisible gets believed.
DRONE_SPRAY_BAND_DEFAULT = 30
DRONE_SPRAY_BAND_MIN = 5
DRONE_SPRAY_BAND_MAX = 100

# [REASON]: a rate computed on a denominator near zero is arithmetic, not
# evidence. Measured on staging 2026-08-07: machine No 9 flew 25 flights
# totalling 0.25 ha and its litres-per-hectare comes out at 141.91 -- that is
# not a machine over-dosing five-fold, it is 35 litres divided by a quarter of
# a hectare. Grouped by machine PER MONTH, as this report is, such cells are
# common rather than exceptional: production 2026-01 holds 12 flights and
# 0.19 ha for the whole fleet. A machine-month below this area still SHOWS its
# rate -- hiding a number because it is inconvenient is the other failure --
# but it does not vote in the median and is never coloured.
DRONE_SPRAY_MIN_AREA_HA = 10.0
DRONE_SPRAY_MIN_AREA_MAX = 1000.0


def _drone_spray_band(args):
    """The corridor half-width in per cent: 5..100, anything else -> 30."""
    raw = args.get('band', type=int)
    if raw is None or raw < DRONE_SPRAY_BAND_MIN or raw > DRONE_SPRAY_BAND_MAX:
        return DRONE_SPRAY_BAND_DEFAULT
    return raw


def _drone_spray_min_area(args):
    """The judging threshold in hectares: 0..1000, anything else -> 10.0.

    0 is LEGAL and means «judge everything». It is how the owner inspects
    exactly the rows this rule keeps out of the median, so it must not be
    folded into the «anything else» branch: a falsy 0 that silently became
    10.0 would make the escape hatch look broken rather than absent.
    """
    raw = args.get('min_area', type=float)
    if raw is None or raw < 0.0 or raw > DRONE_SPRAY_MIN_AREA_MAX:
        return DRONE_SPRAY_MIN_AREA_HA
    return raw


def _drone_median(values):
    """Median of a list of floats, or None when there is nothing to take."""
    ordered = sorted(values)
    count = len(ordered)
    if not count:
        return None
    middle = count // 2
    if count % 2:
        return ordered[middle]
    return (ordered[middle - 1] + ordered[middle]) / 2.0


def _drone_spray_usage_data(conds, band, min_area=DRONE_SPRAY_MIN_AREA_HA):
    """Litres per hectare by machine x month, with the corridor around it.

    [REASON]: THE RATE IS COMPUTED OVER SPRAY FLIGHTS ONLY (usage_type == 0)
    and NOT over every flight. _drone_summary_data's liters_per_ha divides
    total litres by TOTAL area, seeding included, which depresses the figure
    for any machine that also sowed -- the machine looks under-dosed when it
    was simply doing another job. The area of everything that is not a spray
    flight is carried in its own column so the machine's total area still
    reconciles with the grand total and the reader can see where it went.

    [REASON]: usage_type is NULLABLE, and `usage_type == 0` is NULL -- not
    true -- for those rows, so a flight whose type was never recorded lands in
    the OTHER column. That is deliberate: an unknown job type is not evidence
    of spraying, and quietly counting it as spray would move litres and area
    into a rate that is supposed to describe one treatment. MEASURED
    2026-08-07: sowing is 63 flights of 28 832 (17.13 ha of 28 835.34) and
    touches only machines No 10 and No 13, moving their rate by 0.25 and 0.07
    l/ha. So this filter is CORRECTNESS, not a large correction -- it is what
    makes the column mean one thing, and it does not explain the spread.

    [REASON]: DRONE-ZERO-VS-UNKNOWN-001 again, on a third column, and READ
    THIS BEFORE DELETING THE BRANCH AS UNREACHABLE. The NULL-litres branch is
    DEFENSIVE, NOT DESCRIPTIVE. Measured on 2026-08-07: the number of spray
    flights with spray_liters IS NULL is ZERO across 28 832 production
    flights and ZERO across 10 409 on staging. The branch exists because the
    column is NULLABLE and one collector failure is all it takes: without it,
    SUM over an all-NULL group is 0.0 after the coalesce and the cell reads
    0.00 l/ha -- «this machine sprayed nothing over 40 hectares», a false
    sentence, which the corridor then paints red into a false accusation. A
    gap must not become a fabricated zero.
    It is NOT the explanation of the eleven-fold spread. That spread is real
    data: machine No 8 genuinely reports 3.54 l/ha over 2 384.30 ha, with no
    NULLs and no sowing involved. The counter that decides the branch is
    no_liters, in the same shape _drone_work_cut() uses for no_amount, and
    the only thing that exercises it today is the fixture in A1.3 -- which
    makes that negative control the whole of this branch's coverage, and
    therefore more necessary rather than less.
    """
    month_expr = _drone_flight_month_expr()
    # usage_type == 0 is spray; see the second [REASON] above for NULL.
    is_spray = (DroneFlight.usage_type == 0)

    groups = (db.session.query(
        DroneFlight.drone_unit_id,
        month_expr,
        func.count(DroneFlight.id),
        func.coalesce(
            func.sum(case((is_spray, DroneFlight.area_ha), else_=0.0)), 0.0),
        func.coalesce(
            func.sum(case((is_spray, 0.0), else_=DroneFlight.area_ha)), 0.0),
        # SUM skips NULLs by itself, so this is litres over the spray flights
        # that HAVE a figure -- which is why no_liters below is needed to tell
        # «nothing recorded» from «recorded as nothing».
        func.coalesce(
            func.sum(case((is_spray, DroneFlight.spray_liters), else_=None)),
            0.0),
        func.sum(case((is_spray, 1), else_=0)),
        func.sum(case((and_(is_spray, DroneFlight.spray_liters.is_(None)), 1),
                      else_=0)),
    ).filter(*conds).group_by(DroneFlight.drone_unit_id, month_expr).all())

    unit_numbers = {u.id: u.number for u in DroneUnit.query.all()}
    rows = []
    for (unit_id, month, flights, area_spray, area_other, liters,
         spray_flights, no_liters) in groups:
        area_spray = float(area_spray or 0.0)
        area_other = float(area_other or 0.0)
        liters = float(liters or 0.0)
        spray_flights = int(spray_flights or 0)
        no_liters = int(no_liters or 0)
        # The whole point of the counter: all-NULL is unknown, not zero.
        if spray_flights and spray_flights > no_liters:
            rate = _drone_rate(liters, area_spray)
        else:
            rate = None
        rows.append({
            'unit_id': unit_id,
            'number': unit_numbers.get(unit_id),
            'month': month,
            'flights': flights,
            'spray_flights': spray_flights,
            'other_flights': flights - spray_flights,
            'area_spray': area_spray,
            'area_other': area_other,
            'area_total': area_spray + area_other,
            'liters': liters,
            'no_liters': no_liters,
            'rate': rate,
        })

    # Machine number ascending, NULL machine last, newest month first inside
    # each machine: the row a reader looks for is this month's, and the
    # unattributed line must stay visible rather than sort into the middle of
    # the fleet. Two passes, relying on sort stability, so neither key has to
    # encode the other's direction.
    rows.sort(key=lambda r: r['month'] or '', reverse=True)
    rows.sort(key=lambda r: (r['number'] is None, r['number'] or 0))

    # [REASON]: WHICH ROWS ARE JUDGED, and it is not «all of them».
    #
    #   * a rate that could not be computed cannot be compared;
    #   * a machine-month under min_area hectares has a denominator too small
    #     to mean anything -- see DRONE_SPRAY_MIN_AREA_HA;
    #   * THE UNATTRIBUTED LINE NEVER VOTES. It aggregates some forty nickname
    #     spellings across an unknown set of machines, so its rate is a mixture
    #     of an unknown number of treatments. A report that judges machines
    #     must not take its standard from a bucket of unknowns, and must not
    #     accuse that bucket either.
    #
    # Every excluded row STILL SHOWS ITS RATE and carries a label saying why
    # it is uncoloured. Hiding the number would be the opposite failure.
    for row in rows:
        if row['rate'] is None:
            row['judged'] = False
            row['note'] = ''
        elif row['unit_id'] is None:
            row['judged'] = False
            row['note'] = _drone_t('аниқланмаган', 'не распознано')
        elif row['area_spray'] < min_area:
            row['judged'] = False
            row['note'] = _drone_t('майдон кам', 'мало площади')
        else:
            row['judged'] = True
            row['note'] = ''

    median = _drone_median([r['rate'] for r in rows if r['judged']])
    low = high = low2 = high2 = None
    if median is not None:
        low = median * (1.0 - band / 100.0)
        high = median * (1.0 + band / 100.0)
        low2 = median * (1.0 - 2.0 * band / 100.0)
        high2 = median * (1.0 + 2.0 * band / 100.0)
    for row in rows:
        # A cell with no rate is never coloured: an em dash means «nobody
        # wrote it down», and painting it red would accuse the machine of a
        # dose nobody measured. Nor is a row that did not vote in the median.
        row['flag'] = ''
        if median is not None and row['judged']:
            if row['rate'] < low2 or row['rate'] > high2:
                row['flag'] = 'is-danger-row'
            elif row['rate'] < low or row['rate'] > high:
                row['flag'] = 'is-warning-row'

    grand = (db.session.query(
        func.count(DroneFlight.id),
        func.coalesce(func.sum(DroneFlight.area_ha), 0.0),
    ).filter(*conds).one())
    total = {
        'flights': sum(r['flights'] for r in rows),
        'spray_flights': sum(r['spray_flights'] for r in rows),
        'other_flights': sum(r['other_flights'] for r in rows),
        'area_spray': sum(r['area_spray'] for r in rows),
        'area_other': sum(r['area_other'] for r in rows),
        'liters': sum(r['liters'] for r in rows),
        'no_liters': sum(r['no_liters'] for r in rows),
    }
    total['area_total'] = total['area_spray'] + total['area_other']
    total['rate'] = (_drone_rate(total['liters'], total['area_spray'])
                     if total['spray_flights'] > total['no_liters'] else None)
    # A1.2: area_spray + area_other over every row against the grand total of
    # the same filter. Shown on the page as the other cuts show it, never
    # hidden -- a breakdown that quietly loses hectares is the defect this
    # module has already paid for once.
    reconciled = (total['flights'] == (grand[0] or 0)
                  and abs(total['area_total'] - float(grand[1] or 0.0)) < 0.005)
    return {
        'rows': rows,
        'total': total,
        'grand': {'flights': grand[0] or 0, 'area': float(grand[1] or 0.0)},
        'reconciled': reconciled,
        'band': band,
        'min_area': min_area,
        'median': median,
        'low': low,
        'high': high,
        'low2': low2,
        'high2': high2,
        # Cells that VOTED, not cells that have a rate: the sentence on the
        # page says «the median was taken over N pairs» and must name the
        # number that was actually used.
        'rated_cells': sum(1 for r in rows if r['judged']),
        'unjudged_cells': sum(1 for r in rows
                              if r['rate'] is not None and not r['judged']),
    }


@drones_bp.route('/reports/spray')
@module_required('drones')
def spray_usage():
    """Litres per hectare by machine and month, against a median corridor.

    [REASON]: across the fleet the figure runs from 3.54 to 38.53 l/ha against
    a mean of 25.98 -- an eleven-fold spread nobody looks at. Three
    explanations demand three different actions (a different treatment; an
    under-filled tank the customer paid for; a broken flow meter) and nothing
    in the module tells them apart. This screen does not decide which it is;
    it makes the spread visible per machine-month so somebody can go and ask.
    """
    filters = _drone_filters_from_args(request.args,
                                       default_current_month=False)
    band = _drone_spray_band(request.args)
    min_area = _drone_spray_min_area(request.args)
    data = _drone_spray_usage_data(_drone_flight_conditions(filters), band,
                                   min_area)
    link_args = _drone_link_args(filters)
    if band != DRONE_SPRAY_BAND_DEFAULT:
        link_args['band'] = band
    # Compared to the constant, not to a falsy test: min_area=0 is a real
    # choice and has to survive into the Excel link.
    if min_area != DRONE_SPRAY_MIN_AREA_HA:
        link_args['min_area'] = min_area
    return render_template(
        'drones/spray_usage.html',
        data=data,
        filters=filters,
        link_args=link_args,
        units=DroneUnit.query.order_by(DroneUnit.number).all(),
    )


@drones_bp.route('/reports/spray.xlsx')
@module_required('drones')
def spray_usage_xlsx():
    """The same table, two sheets: the numbers and the filters that made them.

    [REASON]: a cell that is an em dash on screen is an EMPTY cell here, never
    0 -- the same rule _drone_money_or_blank() applies to the works workbooks,
    for the same reason: the workbook is the copy that gets believed, and a 0
    in a litres-per-hectare column is an accusation.
    """
    from openpyxl import Workbook

    filters = _drone_filters_from_args(request.args,
                                       default_current_month=False)
    band = _drone_spray_band(request.args)
    min_area = _drone_spray_min_area(request.args)
    data = _drone_spray_usage_data(_drone_flight_conditions(filters), band,
                                   min_area)
    st = _drone_xlsx_styler()
    unbounded = _drone_t('чекланмаган', 'не ограничен')
    label_unattr = _drone_t('Аниқланмаган', 'Не распознано')
    label_total = _drone_t('Жами', 'Итого')

    wb = Workbook()
    ws = wb.active
    ws.title = _drone_t('Жамланма', 'Сводка')
    ws.append([_drone_t('Кўрсаткич', 'Показатель'),
               _drone_t('Қиймат', 'Значение')])
    for label, value in (
        (_drone_t('Давр: бошланиши', 'Период: с'),
         filters['date_from_s'] or unbounded),
        (_drone_t('Давр: тугаши', 'Период: по'),
         filters['date_to_s'] or unbounded),
        (_drone_t('Йўлак, %', 'Коридор, %'), band),
        (_drone_t('Баҳолаш чегараси, га', 'Порог оценки, га'), min_area),
        (_drone_t('Медиана, л/га', 'Медиана, л/га'),
         data['median'] if data['median'] is not None else ''),
        (_drone_t('Медианага кирган жуфтликлар',
                  'Пар вошло в медиану'), data['rated_cells']),
        (_drone_t('Баҳоланмаган жуфтликлар',
                  'Пар не оценивалось'), data['unjudged_cells']),
        (_drone_t('Парвозлар', 'Вылеты'), data['total']['flights']),
        (_drone_t('Пуркаш парвозлари', 'Вылеты с опрыскиванием'),
         data['total']['spray_flights']),
        (_drone_t('Пуркаш гектари', 'Гектары опрыскивания'),
         data['total']['area_spray']),
        (_drone_t('Бошқа гектар', 'Прочие гектары'),
         data['total']['area_other']),
        (_drone_t('Литр', 'Литры'), data['total']['liters']),
        (_drone_t('Литри ёзилмаган парвозлар',
                  'Вылетов без записанных литров'),
         data['total']['no_liters']),
    ):
        ws.append([label, _drone_xlsx_safe(value)])
    st.style_table(ws)

    # [REASON]: no slash in a sheet title -- openpyxl refuses «Л/га» with
    # ValueError and the whole export becomes a 500. Excel forbids / \ ? * [ ]
    # in sheet names; the column header inside the sheet still reads «Л/га».
    ws = wb.create_sheet(_drone_t('Литр гектарига', 'Литров на гектар'))
    ws.append([
        _drone_t('Машина', 'Машина'),
        _drone_t('Ой', 'Месяц'),
        _drone_t('Пуркаш парвозлари', 'Вылеты с опрыскиванием'),
        _drone_t('Литри ёзилмаган', 'Без записанных литров'),
        _drone_t('Пуркаш гектари', 'Гектары опрыскивания'),
        _drone_t('Бошқа гектар', 'Прочие гектары'),
        _drone_t('Жами гектар', 'Всего гектаров'),
        _drone_t('Литр', 'Литры'),
        _drone_t('Л/га', 'Л/га'),
        _drone_t('Изоҳ', 'Примечание'),
    ])
    def _liters_cell(row):
        """Litres for a workbook cell, or None when none were ever recorded.

        Same rule as _drone_money_or_blank(): the figure is unknown exactly
        when every spray flight of the group is missing the thing it is
        computed from. An empty cell, never 0 -- 0 litres over 40 hectares is
        a statement about the machine, and nobody made it.
        """
        return row['liters'] if row['spray_flights'] > row['no_liters'] else None

    for r in data['rows']:
        ws.append([
            (('№ %s' % r['number']) if r['number'] is not None
             else label_unattr),
            r['month'],
            r['spray_flights'],
            r['no_liters'],
            r['area_spray'],
            r['area_other'],
            r['area_total'],
            _liters_cell(r),
            r['rate'],
            r['note'],
        ])
    ws.append([label_total, '',
               data['total']['spray_flights'], data['total']['no_liters'],
               data['total']['area_spray'], data['total']['area_other'],
               data['total']['area_total'], _liters_cell(data['total']),
               data['total']['rate'], ''])
    st.style_table(ws, num_formats={5: '0.00', 6: '0.00', 7: '0.00',
                                    8: '0.00', 9: '0.00'},
                   bold_rows=(ws.max_row,))
    return _drone_xlsx_response(wb, 'drone_spray_usage', filters)
