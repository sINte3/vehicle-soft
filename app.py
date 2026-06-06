"""
Р’СѓС…РѕСЂРѕ РђРіСЂРѕРєР»Р°СЃС‚РµСЂ вЂ” РўСЂР°РЅСЃРїРѕСЂС‚ ТіРёСЃРѕР±РѕС‚Рё v4.5
"""

import os
import hmac
import secrets
import re
from functools import wraps
from datetime import datetime, date, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file, abort, g, session
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user
)
from config import get_config
from models import (
    db, User, Organization, Equipment, WorkType, Customer, DailyRecord,
    Deficiency, VialonMapping, VialonImport, EngineHoursRecord,
    FuelStation, FuelTank, FuelSnapshot, FuelTransaction, FuelSyncLog,
    FuelWarehouse, FuelStation2, FuelInitialBalance, FuelReceipt2, FuelTransaction2,
    FuelSyncLog2, FuelWarningReview,
    SparePartRequest,
    ROLE_ADMIN, ROLE_OPERATOR, ROLE_VIEWER, ROLES,
    CAT_YUKORI, CAT_MTZ, CAT_QATNOV, CAT_MINI, CAT_COMBINE,
    CAT_SPECIAL, CAT_YUK_TRANSPORT, CAT_MOTORCYCLE, CAT_PASSENGER,
    CATEGORIES, REPORT_GROUPS,
    AppModule, UserModulePermission,
    module_required,
)
from excel_export import generate_report
from wialon_import import register_wialon_routes
from fuel_routes import fuel_bp, _perform_fuel_sync, _collect_fuel_report_data  # noqa: F401 (registered below)
from excel_daily_activity import generate_daily_activity
from translations import TRANS
from spare_parts import spare_parts_bp
from sec003a_ext import (
    log_audit, get_user_extra, must_change_password, is_locked,
    record_failed_login, record_successful_login, clear_must_change_password,
    require_temp_password_change, password_is_strong_enough,
    model_snapshot, diff_dict, audit_action_name,
)
from sqlalchemy import text


def create_app():
    app = Flask(__name__)
    app.config.from_object(get_config())
    os.makedirs(app.config['REPORTS_DIR'], exist_ok=True)
    os.makedirs(os.path.join(os.path.dirname(__file__), 'instance'), exist_ok=True)

    db.init_app(app)

    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'РўРёР·РёРјРіР° РєРёСЂРёРЅРі'
    login_manager.login_message_category = 'info'

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    def fmt_sum(val):
        if val is None or val == 0:
            return ''
        return f'{val:,.0f}'.replace(',', ' ')

    app.jinja_env.filters['fmt_sum'] = fmt_sum
    app.jinja_env.globals['ROLES'] = ROLES
    app.jinja_env.globals['CATEGORIES'] = CATEGORIES
    app.jinja_env.globals['REPORT_GROUPS'] = REPORT_GROUPS
    app.jinja_env.globals['CAT_YUKORI'] = CAT_YUKORI
    app.jinja_env.globals['CAT_MTZ'] = CAT_MTZ
    app.jinja_env.globals['CAT_QATNOV'] = CAT_QATNOV
    app.jinja_env.globals['CAT_MINI'] = CAT_MINI
    app.jinja_env.globals['CAT_COMBINE'] = CAT_COMBINE
    app.jinja_env.globals['CAT_SPECIAL'] = CAT_SPECIAL
    app.jinja_env.globals['CAT_YUK_TRANSPORT'] = CAT_YUK_TRANSPORT
    app.jinja_env.globals['CAT_MOTORCYCLE'] = CAT_MOTORCYCLE
    app.jinja_env.globals['CAT_PASSENGER'] = CAT_PASSENGER

    def get_csrf_token():
        token = session.get('_csrf_token')
        if not token:
            token = secrets.token_urlsafe(32)
            session['_csrf_token'] = token
        return token

    app.jinja_env.globals['csrf_token'] = get_csrf_token

    def is_csrf_exempt():
        # Topaz agent API endpoints are protected by FUEL_API_TOKEN and must not
        # require browser-session CSRF tokens.
        if request.path in ('/fuel/api/fuel_sync', '/api/fuel_sync'):
            return True
        return False

    @app.before_request
    def enforce_csrf_protection():
        if request.method != 'POST':
            return None
        if is_csrf_exempt():
            return None
        expected = session.get('_csrf_token')
        submitted = (
            request.form.get('csrf_token')
            or request.headers.get('X-CSRF-Token')
            or request.headers.get('X-CSRFToken')
        )
        if not expected or not submitted or not hmac.compare_digest(expected, submitted):
            abort(400)
        return None

    @app.before_request
    def set_language():
        if current_user.is_authenticated:
            g.lang = getattr(current_user, 'language', 'uz') or 'uz'
        else:
            g.lang = 'uz'

    @app.before_request
    def enforce_temp_password_change():
        if not current_user.is_authenticated:
            return None
        if not must_change_password(db, current_user.id):
            return None
        allowed = {
            'sec003a_change_temp_password', 'logout', 'static',
            'set_language_route'
        }
        if request.endpoint in allowed:
            return None
        if request.path.startswith('/static/'):
            return None
        return redirect(url_for('sec003a_change_temp_password'))

    @app.context_processor
    def inject_translation():
        lang = getattr(g, 'lang', 'uz')
        def t(key):
            return TRANS.get(lang, {}).get(key, key)
        return dict(t=t, lang=lang)

    def ui_t(uz, ru):
        # [REASON]: Small route-level helper for bilingual flash messages in
        # safety guards where templates are not involved.
        return ru if getattr(g, 'lang', 'uz') == 'ru' else uz

    def format_validation_errors(errors, title_uz=None, title_ru=None):
        # [REASON]: Present several validation problems at once so operators
        # can fix the whole form in one pass instead of resubmitting repeatedly.
        unique = []
        seen = set()
        for err in errors or []:
            text = str(err).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            unique.append(text)
        if not unique:
            return ''
        if len(unique) == 1 and not title_uz and not title_ru:
            return unique[0]
        title = ui_t(title_uz or 'Маълумот сақланмади', title_ru or 'Данные не сохранены')
        if len(unique) == 1:
            return title + '\n- ' + unique[0]
        return title + '\n' + '\n'.join('- ' + item for item in unique)

    def flash_validation_errors(errors, title_uz=None, title_ru=None):
        msg = format_validation_errors(errors, title_uz, title_ru)
        if msg:
            flash(msg, 'warning')

    def admin_required(f):
        @wraps(f)
        @login_required
        def decorated(*args, **kwargs):
            if not current_user.is_admin:
                abort(403)
            return f(*args, **kwargs)
        return decorated

    def editor_required(f):
        @wraps(f)
        @login_required
        def decorated(*args, **kwargs):
            if not current_user.can_edit:
                abort(403)
            return f(*args, **kwargs)
        return decorated

    def get_user_orgs():
        if current_user.is_admin:
            return Organization.query.order_by(Organization.sort_order).all()
        return (Organization.query
                .filter(Organization.id.in_([o.id for o in current_user.organizations]))
                .order_by(Organization.sort_order).all())

    def check_org_access(org_id):
        if org_id and not current_user.can_access_org(org_id):
            abort(403)

    def parse_date(s):
        if s:
            try:
                return datetime.strptime(s, '%Y-%m-%d').date()
            except ValueError:
                pass
        return date.today()

    def parse_date_range(args):
        """Parse date range from request args. Returns (mode, date_from, date_to)."""
        mode = args.get('mode', 'day')
        today = date.today()

        if mode == 'week':
            # Monday to Sunday of current week
            d_from = today - timedelta(days=today.weekday())
            d_to = d_from + timedelta(days=6)
        elif mode == 'month':
            d_from = today.replace(day=1)
            # Last day of month
            if today.month == 12:
                d_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                d_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        elif mode == 'range':
            d_from = parse_date(args.get('date_from'))
            d_to = parse_date(args.get('date_to'))
            if d_from > d_to:
                d_from, d_to = d_to, d_from
        else:
            # Single day
            mode = 'day'
            d_from = parse_date(args.get('date'))
            d_to = d_from

        return mode, d_from, d_to

    def parse_form_date_required(value, field_label):
        try:
            if not value:
                raise ValueError()
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            raise ValueError(f'{field_label}: invalid date')

    def parse_non_negative_float(value, field_label, allow_empty=True):
        if value is None or str(value).strip() == '':
            if allow_empty:
                return None
            raise ValueError(f'{field_label}: required')
        try:
            result = float(str(value).replace(',', '.'))
        except (TypeError, ValueError):
            raise ValueError(f'{field_label}: invalid number')
        if result < 0:
            raise ValueError(f'{field_label}: must be non-negative')
        return result

    def parse_positive_float(value, field_label):
        result = parse_non_negative_float(value, field_label, allow_empty=False)
        if result <= 0:
            raise ValueError(f'{field_label}: must be greater than zero')
        return result

    def normalize_ref_text(value):
        return re.sub(r'\s+', ' ', (value or '').strip())

    def normalize_plate(value):
        return normalize_ref_text(value).upper()

    def ref_duplicate_exists(model, field_name, normalized_value, exclude_id=None, org_id=None):
        if not normalized_value:
            return False
        rows = model.query.all()
        target = normalize_ref_text(normalized_value).casefold()
        for row in rows:
            if exclude_id and getattr(row, 'id', None) == exclude_id:
                continue
            if org_id is not None and getattr(row, 'organization_id', None) != org_id:
                continue
            current = normalize_ref_text(getattr(row, field_name, '')).casefold()
            if current == target:
                return True
        return False

    def equipment_plate_duplicate_exists(plate_value, exclude_id=None):
        target = normalize_plate(plate_value)
        if not target:
            return False
        for row in Equipment.query.all():
            if exclude_id and row.id == exclude_id:
                continue
            if normalize_plate(row.plate) == target:
                return True
        return False


    # в”Ђв”Ђв”Ђ AUTH в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            if must_change_password(db, current_user.id):
                return redirect(url_for('sec003a_change_temp_password'))
            return redirect(url_for('index'))
        if request.method == 'POST':
            username = request.form.get('username', '').strip().lower()
            password = request.form.get('password', '')
            user = User.query.filter_by(username=username).first()
            if user and is_locked(db, user.id):
                log_audit(db, 'login_locked', entity_type='user', entity_id=user.id,
                          entity_label=user.username, module='auth', status='blocked',
                          description='Account temporarily locked', actor_user=user)
                db.session.commit()
                flash('Аккаунт вақтинча блокланган. 15 дақиқадан кейин қайта уриниб кўринг.', 'warning')
                return render_template('login.html')
            if user and user.check_password(password) and user.is_active:
                user.last_login = datetime.utcnow()
                record_successful_login(db, user)
                db.session.commit()
                login_user(user, remember=True)
                log_audit(db, 'login_success', entity_type='user', entity_id=user.id,
                          entity_label=user.username, module='auth', description='Successful login', actor_user=user)
                db.session.commit()
                next_page = request.args.get('next')
                flash(f'Хуш келибсиз, {user.full_name or user.username}!', 'success')
                if must_change_password(db, user.id):
                    return redirect(url_for('sec003a_change_temp_password'))
                return redirect(next_page or url_for('index'))
            else:
                if user:
                    record_failed_login(db, user)
                    log_audit(db, 'login_failed', entity_type='user', entity_id=user.id,
                              entity_label=user.username, module='auth', status='failed',
                              description='Invalid password or inactive account', actor_user=user)
                    db.session.commit()
                flash('Логин ёки парол нотўғри', 'warning')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        log_audit(db, 'logout', entity_type='user', entity_id=current_user.id,
                  entity_label=current_user.username, module='auth', description='User logout')
        db.session.commit()
        logout_user()
        return redirect(url_for('login'))

    @app.route('/change-temporary-password', methods=['GET', 'POST'])
    @login_required
    def sec003a_change_temp_password():
        if request.method == 'POST':
            new_pass = request.form.get('new_password', '').strip()
            confirm = request.form.get('confirm_password', '').strip()
            if new_pass != confirm:
                flash('Пароллар мос келмади', 'warning')
                return redirect(url_for('sec003a_change_temp_password'))
            ok, msg = password_is_strong_enough(current_user.username, new_pass)
            if not ok:
                flash(msg, 'warning')
                return redirect(url_for('sec003a_change_temp_password'))
            if current_user.check_password(new_pass):
                flash('Янги парол вақтинчалик паролдан фарқ қилиши керак', 'warning')
                return redirect(url_for('sec003a_change_temp_password'))
            current_user.set_password(new_pass)
            clear_must_change_password(db, current_user.id)
            log_audit(db, 'password_changed', entity_type='user', entity_id=current_user.id,
                      entity_label=current_user.username, module='auth', description='Temporary password changed')
            db.session.commit()
            flash('Парол ўзгартирилди. Энди дастурдан фойдаланишингиз мумкин.', 'success')
            return redirect(url_for('index'))
        return render_template('change_temporary_password.html')

    @app.route('/profile', methods=['GET', 'POST'])
    @login_required
    def profile():
        if request.method == 'POST':
            full_name = request.form.get('full_name', '').strip()
            current_pass = request.form.get('current_password', '')
            new_pass = request.form.get('new_password', '').strip()
            confirm = request.form.get('confirm_password', '').strip()
            current_user.full_name = full_name
            if new_pass:
                if not current_user.check_password(current_pass):
                    flash('Жорий парол нотўғри', 'warning')
                    return redirect(url_for('profile'))
                if new_pass != confirm:
                    flash('Пароллар мос келмади', 'warning')
                    return redirect(url_for('profile'))
                ok, msg = password_is_strong_enough(current_user.username, new_pass)
                if not ok:
                    flash(msg, 'warning')
                    return redirect(url_for('profile'))
                if current_user.check_password(new_pass):
                    flash('Янги парол эски паролдан фарқ қилиши керак', 'warning')
                    return redirect(url_for('profile'))
                current_user.set_password(new_pass)
                clear_must_change_password(db, current_user.id)
                log_audit(db, 'password_changed', entity_type='user', entity_id=current_user.id,
                          entity_label=current_user.username, module='auth', description='Password changed from profile')
                flash('Парол ўзгартирилди', 'success')
            db.session.commit()
            flash('Профиль сақланди', 'success')
            return redirect(url_for('profile'))
        return render_template('profile.html')

    @app.route('/set_language', methods=['POST'])
    @login_required
    def set_language_route():
        lang = request.form.get('lang', 'uz')
        if lang in ('uz', 'ru'):
            current_user.language = lang
            db.session.commit()
        return redirect(request.referrer or url_for('index'))


    def _format_dashboard_dt(value):
        if not value:
            return ''
        try:
            return value.strftime('%d.%m.%Y %H:%M')
        except Exception:
            return str(value)

    def _safe_scalar(query, default=0):
        try:
            value = query.scalar()
            return default if value is None else value
        except Exception:
            return default

    def _latest_backup_info():
        backup_dir = r'D:\transport-report-backups\production\daily'
        info = {'path': '', 'name': '', 'mtime': None, 'status': 'unknown'}
        try:
            if not os.path.isdir(backup_dir):
                return info
            files = [
                os.path.join(backup_dir, name)
                for name in os.listdir(backup_dir)
                if name.lower().endswith('.db')
            ]
            if not files:
                return info
            latest = max(files, key=os.path.getmtime)
            mtime = datetime.fromtimestamp(os.path.getmtime(latest))
            age_hours = (datetime.now() - mtime).total_seconds() / 3600
            info.update({
                'path': latest,
                'name': os.path.basename(latest),
                'mtime': mtime,
                'age_hours': round(age_hours, 1),
                'status': 'ok' if age_hours <= 36 else 'warning',
            })
        except Exception:
            pass
        return info

    def _build_dashboard_context(d_from, d_to, filter_org_ids, records, totals):
        d_from_dt = datetime.combine(d_from, datetime.min.time())
        d_to_dt = datetime.combine(d_to, datetime.max.time())
        now = datetime.now()

        working_rows = [r for r in records if r.status == 'working']
        idle_rows = [r for r in records if r.status != 'working']
        transport = {
            'records': len(records),
            'working_rows': len(working_rows),
            'idle_rows': len(idle_rows),
            'quantity': round(sum((r.quantity or 0) for r in working_rows), 2),
            'amount': totals.get('total', 0),
            'equipment_working': len({r.equipment_id for r in working_rows}),
            'equipment_idle': len({r.equipment_id for r in idle_rows} - {r.equipment_id for r in working_rows}),
        }

        module_access = {
            'transport': True,
            'fuel': current_user.has_module_access('fuel'),
            'spare_parts': current_user.has_module_access('spare_parts'),
            'wialon': current_user.has_module_access('wialon'),
            'admin': current_user.is_admin,
        }

        fuel = {
            'available': module_access['fuel'],
            'issued': 0.0,
            'transactions': 0,
            'warnings': 0,
            'danger_warnings': 0,
            'latest_sync': None,
            'latest_sync_text': '',
            'sync_age_hours': None,
            'review_new': 0,
            'review_in_progress': 0,
            'review_resolved': 0,
        }
        if module_access['fuel']:
            try:
                fuel_report = _collect_fuel_report_data(d_from, d_to)
                fuel_totals = fuel_report.get('totals', {})
                fuel_warning_summary = fuel_report.get('warnings_summary', {})
                fuel['issued'] = fuel_totals.get('issued', 0) or 0
                fuel['transactions'] = fuel_totals.get('transactions', 0) or 0
                fuel['warnings'] = fuel_warning_summary.get('total', 0) or 0
                fuel['danger_warnings'] = fuel_warning_summary.get('danger', 0) or 0
            except Exception:
                pass

            latest_sync = FuelSyncLog2.query.order_by(FuelSyncLog2.synced_at.desc()).first()
            fuel['latest_sync'] = latest_sync
            fuel['latest_sync_text'] = _format_dashboard_dt(getattr(latest_sync, 'synced_at', None))
            if latest_sync and latest_sync.synced_at:
                fuel['sync_age_hours'] = round((now - latest_sync.synced_at).total_seconds() / 3600, 1)

            try:
                review_rows = (db.session.query(FuelWarningReview.status, db.func.count(FuelWarningReview.id))
                               .group_by(FuelWarningReview.status).all())
                review_counts = {status or 'new': int(cnt or 0) for status, cnt in review_rows}
                fuel['review_new'] = review_counts.get('new', 0)
                fuel['review_in_progress'] = review_counts.get('in_progress', 0)
                fuel['review_resolved'] = review_counts.get('resolved', 0)
            except Exception:
                pass

        wialon = {
            'available': module_access['wialon'],
            'total': 0,
            'linked': 0,
            'skipped': 0,
            'unmapped': 0,
        }
        if module_access['wialon']:
            try:
                total_mappings = VialonMapping.query.count()
                linked = VialonMapping.query.filter(VialonMapping.equipment_id.isnot(None), VialonMapping.skip.is_(False)).count()
                skipped = VialonMapping.query.filter(VialonMapping.skip.is_(True)).count()
                wialon.update({
                    'total': total_mappings,
                    'linked': linked,
                    'skipped': skipped,
                    'unmapped': max(0, total_mappings - linked - skipped),
                })
            except Exception:
                pass

        spare = {
            'available': module_access['spare_parts'],
            'draft': 0,
            'submitted': 0,
            'approved': 0,
            'rejected': 0,
            'open': 0,
        }
        if module_access['spare_parts']:
            try:
                q = db.session.query(SparePartRequest.status, db.func.count(SparePartRequest.id))
                if not current_user.is_admin and filter_org_ids:
                    q = q.filter(SparePartRequest.organization_id.in_(filter_org_ids))
                spare_counts = {status or 'draft': int(cnt or 0) for status, cnt in q.group_by(SparePartRequest.status).all()}
                spare.update({
                    'draft': spare_counts.get('draft', 0),
                    'submitted': spare_counts.get('submitted', 0),
                    'approved': spare_counts.get('approved', 0),
                    'rejected': spare_counts.get('rejected', 0),
                })
                spare['open'] = spare['draft'] + spare['submitted']
            except Exception:
                pass

        audit = []
        try:
            audit = db.session.execute(text("""
                SELECT id, created_at, username_snapshot, action, module, status, description
                FROM audit_logs
                ORDER BY id DESC
                LIMIT 8
            """)).mappings().all()
        except Exception:
            audit = []

        return {
            'transport': transport,
            'fuel': fuel,
            'wialon': wialon,
            'spare': spare,
            'audit': audit,
            'backup': _latest_backup_info(),
            'module_access': module_access,
        }


    # в”Ђв”Ђв”Ђ DASHBOARD в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/', methods=['GET', 'POST'])
    @module_required('transport')
    @login_required
    def index():
        if request.method == 'POST':
            source = request.form
            org_ids_form = request.form.getlist('org_ids')
            cat_codes    = request.form.getlist('cat_codes')
            work_types_f = request.form.getlist('work_types')
        else:
            source = request.args
            org_ids_form = []
            cat_codes = []
            work_types_f = []

        mode, d_from, d_to = parse_date_range(source)
        is_range = (d_from != d_to)
        num_days = (d_to - d_from).days + 1

        user_org_ids = current_user.get_org_ids()
        organizations = get_user_orgs()

        if org_ids_form:
            filter_org_ids = [int(x) for x in org_ids_form if x.isdigit() and int(x) in user_org_ids]
        else:
            filter_org_ids = user_org_ids

        q = (DailyRecord.query
             .filter(DailyRecord.work_date >= d_from,
                     DailyRecord.work_date <= d_to)
             .join(Equipment).join(Organization)
             .filter(Organization.id.in_(filter_org_ids)))

        if cat_codes:
            q = q.filter(Equipment.category.in_(cat_codes))

        if work_types_f:
            q = q.filter(DailyRecord.work_type.in_(work_types_f))

        records = q.order_by(Organization.sort_order, Equipment.category,
                             Equipment.name, DailyRecord.work_date, DailyRecord.line_index).all()

        # Work types for multiselect (from actual records in date range)
        all_work_types = sorted(set(
            r[0] for r in DailyRecord.query
            .filter(DailyRecord.work_date >= d_from, DailyRecord.work_date <= d_to)
            .filter(DailyRecord.status == 'working')
            .filter(DailyRecord.work_type != '')
            .with_entities(DailyRecord.work_type).all()
            if r[0]
        ))

        # Totals
        total_cash = sum(r.amount_cash or 0 for r in records)
        total_transfer = sum(r.amount_transfer or 0 for r in records)
        total_internal = sum(r.amount_internal or 0 for r in records)
        total_other = sum(r.amount_other or 0 for r in records)
        total = total_cash + total_transfer + total_internal + total_other

        # Unique equipment counts
        working_eq_ids = set(r.equipment_id for r in records if r.status == 'working')
        idle_eq_ids = set(r.equipment_id for r in records if r.status != 'working')
        working_count = len(working_eq_ids)
        idle_count = len(idle_eq_ids - working_eq_ids)

        # Build hierarchical data: org -> equipment -> records
        orgs_map = {}
        for r in records:
            org = r.equipment.organization
            oid = org.id
            if oid not in orgs_map:
                orgs_map[oid] = {
                    'org': org,
                    'total': 0, 'cash': 0, 'transfer': 0, 'internal': 0, 'other': 0,
                    'working_count': 0, 'idle_count': 0,
                    'equipment': {},
                }
            od = orgs_map[oid]
            od['total'] += r.total_amount
            od['cash'] += r.amount_cash or 0
            od['transfer'] += r.amount_transfer or 0
            od['internal'] += r.amount_internal or 0
            od['other'] += r.amount_other or 0

            eqid = r.equipment_id
            if eqid not in od['equipment']:
                od['equipment'][eqid] = {
                    'eq': r.equipment,
                    'total': 0, 'cash': 0, 'transfer': 0, 'internal': 0, 'other': 0,
                    'working_days': 0, 'idle_days': 0,
                    'records': [],
                }
            ed = od['equipment'][eqid]
            ed['total'] += r.total_amount
            ed['cash'] += r.amount_cash or 0
            ed['transfer'] += r.amount_transfer or 0
            ed['internal'] += r.amount_internal or 0
            ed['other'] += r.amount_other or 0
            ed['records'].append(r)

        # Count working/idle per equipment
        for od in orgs_map.values():
            w_set = set()
            i_set = set()
            for eqid, ed in od['equipment'].items():
                dates_working = set(r.work_date for r in ed['records'] if r.status == 'working')
                dates_idle = set(r.work_date for r in ed['records'] if r.status != 'working')
                ed['working_days'] = len(dates_working)
                ed['idle_days'] = len(dates_idle - dates_working)
                if dates_working:
                    w_set.add(eqid)
                else:
                    i_set.add(eqid)
            od['working_count'] = len(w_set)
            od['idle_count'] = len(i_set)

        # Sort orgs by sort_order
        orgs_data = sorted(orgs_map.values(), key=lambda x: x['org'].sort_order)

        deficiencies_count = (Deficiency.query
                              .filter(Deficiency.work_date >= d_from,
                                      Deficiency.work_date <= d_to).count())

        dashboard = _build_dashboard_context(d_from, d_to, filter_org_ids, records, {
            'total': total,
            'cash': total_cash,
            'transfer': total_transfer,
            'internal': total_internal,
            'other': total_other,
        })

        return render_template('index.html',
                               dashboard=dashboard,
                               mode=mode, d_from=d_from, d_to=d_to,
                               is_range=is_range, num_days=num_days,
                               orgs_data=orgs_data,
                               total_cash=total_cash, total_transfer=total_transfer,
                               total_internal=total_internal, total_other=total_other,
                               total=total,
                               organizations=organizations,
                               categories=CATEGORIES,
                               all_work_types=all_work_types,
                               working_count=working_count, idle_count=idle_count,
                               deficiencies_count=deficiencies_count,
                               selected_org_ids=[int(x) for x in org_ids_form if x.isdigit()] if org_ids_form else [],
                               selected_cats=cat_codes,
                               selected_work_types=work_types_f)

    # в”Ђв”Ђв”Ђ DAILY ENTRY в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/entry')
    @module_required('transport')
    @login_required
    def daily_entry():
        if not current_user.can_edit:
            flash('Сизда маълумот киритиш ҳуқуқи йўқ', 'warning')
            return redirect(url_for('index'))

        _da = request.args.get('date')
        sel = parse_date(_da) if _da else date.today() - timedelta(days=1)
        org_id = request.args.get('org_id', type=int)
        organizations = get_user_orgs()

        equipment_list = []
        existing = {}
        if org_id:
            check_org_access(org_id)
            equipment_list = (Equipment.query
                              .filter_by(organization_id=org_id, is_active=True)
                              .order_by(Equipment.category, Equipment.name).all())
            recs = (DailyRecord.query.filter_by(work_date=sel)
                    .join(Equipment).filter(Equipment.organization_id == org_id).all())
            for r in recs:
                existing.setdefault(r.equipment_id, []).append(r)
            # Sort lines by line_index
            for k in existing:
                existing[k].sort(key=lambda r: r.line_index)

        work_types = WorkType.query.order_by(WorkType.name).all()
        customers = Customer.query.order_by(Customer.name).all()

        return render_template('daily_entry.html', selected_date=sel,
                               organizations=organizations, selected_org_id=org_id,
                               equipment_list=equipment_list, existing=existing,
                               work_types=work_types, customers=customers)

    @app.route('/entry/save', methods=['POST'])
    @module_required('transport')
    @editor_required
    def save_entry():
        data = request.form
        org_id = request.form.get('org_id', type=int)
        if not org_id:
            flash(ui_t('Ташкилотни танланг', 'Выберите организацию'), 'warning')
            return redirect(url_for('daily_entry'))
        check_org_access(org_id)

        try:
            sel = parse_form_date_required(request.form.get('work_date'), 'work_date')
        except ValueError:
            flash(ui_t('Нотўғри сана', 'Некорректная дата'), 'warning')
            return redirect(url_for('daily_entry', org_id=org_id))

        eq_ids = set()
        for key in data.keys():
            if key.startswith('eq_') and '_status' in key:
                try:
                    eq_ids.add(int(key.split('_')[1]))
                except (ValueError, IndexError):
                    pass

        valid_payments = {'internal', 'cash', 'transfer', 'other'}
        validation_errors = []
        prepared = {}

        for eq_id in eq_ids:
            eq = Equipment.query.get(eq_id)
            if not eq or eq.organization_id != org_id or not eq.is_active:
                validation_errors.append(ui_t('Техника нотўғри танланган', 'Некорректно выбрана техника') + f' ID={eq_id}')
                continue

            p = f'eq_{eq_id}_'
            status = data.get(f'{p}status', 'idle')
            if status not in {'idle', 'working'}:
                validation_errors.append(f'{eq.name} {eq.plate}: ' + ui_t('техника ҳолати нотўғри', 'некорректный статус техники'))
                continue

            if status == 'idle':
                idle_reason = data.get(f'{p}idle_reason', '').strip() or ui_t('Вақтинча бўш', 'Временно свободен')
                prepared[eq_id] = {'status': 'idle', 'idle_reason': idle_reason, 'lines': []}
                continue

            valid_lines_str = data.get(f'{p}valid_lines', '')
            if valid_lines_str:
                valid_lines = [int(x) for x in valid_lines_str.split(',') if x.strip().isdigit()]
            else:
                valid_lines = [0]

            lines_to_save = []
            for li in valid_lines:
                lp = f'{p}line_{li}_'
                work_type = data.get(f'{lp}work_type', '').strip()
                customer = data.get(f'{lp}customer', '').strip()
                unit = data.get(f'{lp}unit', '').strip()
                qty_s = data.get(f'{lp}quantity', '').strip()
                price_s = data.get(f'{lp}price', '').strip()
                payment = data.get(f'{lp}payment_type', 'internal')
                note = data.get(f'{lp}note', '').strip()

                if not work_type and not qty_s and not price_s and not customer and not unit and not note:
                    continue

                if not work_type:
                    validation_errors.append(f'{eq.name} {eq.plate}: ' + ui_t('иш турини танланг', 'выберите вид работ'))
                    continue
                if payment not in valid_payments:
                    validation_errors.append(f'{eq.name} {eq.plate}: ' + ui_t('тўлов тури нотўғри', 'некорректный тип оплаты'))
                    continue

                try:
                    qty = parse_positive_float(qty_s, 'quantity')
                    price = parse_non_negative_float(price_s, 'price', allow_empty=True)
                except ValueError:
                    validation_errors.append(f'{eq.name} {eq.plate}: ' + ui_t('миқдор мусбат, нарх эса манфий бўлмаслиги керак',
                                                  'количество должно быть больше нуля, цена не может быть отрицательной'))
                    continue

                price = price if price is not None else 0
                lines_to_save.append({
                    'work_type': work_type,
                    'customer': customer,
                    'unit': unit,
                    'quantity': qty,
                    'price': price,
                    'payment_type': payment,
                    'note': note,
                })

            if not lines_to_save:
                validation_errors.append(f'{eq.name} {eq.plate}: ' + ui_t('ишлайдиган техника учун камида битта иш сатри керак',
                                              'для работающей техники нужна хотя бы одна строка работы'))
                continue

            prepared[eq_id] = {'status': 'working', 'lines': lines_to_save}

        if validation_errors:
            flash_validation_errors(
                validation_errors,
                'Ҳисобот сақланмади. Қуйидаги хатоларни тузатинг:',
                'Отчёт не сохранён. Исправьте следующие ошибки:'
            )
            return redirect(url_for('daily_entry', date=sel.isoformat(), org_id=org_id))

        daily_fields = ['id', 'work_date', 'equipment_id', 'line_index', 'status', 'work_type', 'customer', 'unit', 'quantity', 'price', 'payment_type', 'idle_reason', 'note', 'amount_cash', 'amount_transfer', 'amount_internal', 'amount_other']
        before_records = []
        if prepared:
            before_records = [model_snapshot(r, daily_fields) for r in DailyRecord.query.filter(
                DailyRecord.work_date == sel, DailyRecord.equipment_id.in_(list(prepared.keys()))
            ).order_by(DailyRecord.equipment_id, DailyRecord.line_index).all()]

        saved = 0
        for eq_id, item in prepared.items():
            DailyRecord.query.filter_by(work_date=sel, equipment_id=eq_id).delete()

            if item['status'] == 'idle':
                rec = DailyRecord(work_date=sel, equipment_id=eq_id, status='idle',
                                  idle_reason=item['idle_reason'], line_index=0,
                                  created_by=current_user.id)
                db.session.add(rec)
                saved += 1
                continue

            for line_pos, line in enumerate(item['lines']):
                amount = round((line['quantity'] or 0) * (line['price'] or 0), 2)
                rec = DailyRecord(
                    work_date=sel, equipment_id=eq_id, status='working',
                    work_type=line['work_type'], customer=line['customer'], unit=line['unit'],
                    quantity=line['quantity'], price=line['price'],
                    amount_cash=amount if line['payment_type'] == 'cash' else 0,
                    amount_transfer=amount if line['payment_type'] == 'transfer' else 0,
                    amount_internal=amount if line['payment_type'] == 'internal' else 0,
                    amount_other=amount if line['payment_type'] == 'other' else 0,
                    payment_type=line['payment_type'], note=line['note'], line_index=line_pos,
                    created_by=current_user.id,
                )
                db.session.add(rec)
            saved += 1

        db.session.flush()
        after_records = []
        if prepared:
            after_records = [model_snapshot(r, daily_fields) for r in DailyRecord.query.filter(
                DailyRecord.work_date == sel, DailyRecord.equipment_id.in_(list(prepared.keys()))
            ).order_by(DailyRecord.equipment_id, DailyRecord.line_index).all()]
        log_audit(
            db, 'daily_records_saved', entity_type='daily_records', entity_label=f'{sel.isoformat()} org_id={org_id}',
            module='daily_entry', before=before_records, after=after_records,
            changes={'saved_equipment_count': saved, 'equipment_ids': sorted(prepared.keys())},
            description=f'Daily entry saved for {saved} equipment items on {sel.isoformat()}'
        )
        db.session.commit()
        flash(f'{saved} та техника маълумотлари сақланди', 'success')
        return redirect(url_for('daily_entry', date=sel.isoformat(), org_id=org_id))

    @app.route('/entry/copy_prev', methods=['POST'])
    @module_required('transport')
    @editor_required
    def copy_previous_day():
        sel = parse_date(request.form.get('work_date'))
        org_id = request.form.get('org_id', type=int)
        check_org_access(org_id)
        prev = sel - timedelta(days=1)

        prev_recs = (DailyRecord.query.filter_by(work_date=prev)
                     .join(Equipment).filter(Equipment.organization_id == org_id).all())
        copied = 0
        for pr in prev_recs:
            if not DailyRecord.query.filter_by(work_date=sel, equipment_id=pr.equipment_id).first():
                new = DailyRecord(work_date=sel, equipment_id=pr.equipment_id,
                                  status='idle', idle_reason='Р’Р°Т›С‚РёРЅС‡Р° Р±СћС€',
                                  unit=pr.unit, price=pr.price,
                                  payment_type=pr.payment_type, line_index=0,
                                  created_by=current_user.id)
                db.session.add(new)
                copied += 1
        log_audit(
            db, 'daily_records_copied_previous_day', entity_type='daily_records',
            entity_label=f'{prev.isoformat()} -> {sel.isoformat()} org_id={org_id}', module='daily_entry',
            after={'copied': copied, 'from_date': prev.isoformat(), 'to_date': sel.isoformat(), 'org_id': org_id},
            description=f'Copied {copied} idle daily records from previous day'
        )
        db.session.commit()
        flash(f'{copied} та техника олдинги кундан кўчирилди', 'info')
        return redirect(url_for('daily_entry', date=sel.isoformat(), org_id=org_id))

    # в”Ђв”Ђв”Ђ DEFICIENCIES (РљР°РјС‡РёР»РёРєР»Р°СЂ) в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/deficiencies')
    @module_required('deficiencies')
    def deficiencies_list():
        if not current_user.can_edit:
            flash('Сизда камчиликларни кўришга ҳуқуқ йўқ', 'warning')
            return redirect(url_for('index'))
        # Input date (for adding new deficiencies)
        sel = parse_date(request.args.get('date'))
        # View range
        mode, d_from, d_to = parse_date_range(request.args)
        user_org_ids = current_user.get_org_ids()
        q = Deficiency.query.filter(
            Deficiency.work_date >= d_from,
            Deficiency.work_date <= d_to
        )
        if not current_user.is_admin:
            q = q.filter(
                (Deficiency.organization_id.in_(user_org_ids)) | (Deficiency.organization_id == None)
            )
        items = q.order_by(Deficiency.work_date, Deficiency.sort_order, Deficiency.id).all()
        organizations = get_user_orgs()
        return render_template('deficiencies.html', selected_date=sel,
                               mode=mode, d_from=d_from, d_to=d_to,
                               items=items, organizations=organizations)

    @app.route('/deficiencies/save', methods=['POST'])
    @module_required('deficiencies')
    @editor_required
    def save_deficiency():
        sel = parse_date(request.form.get('work_date'))
        did = request.form.get('id', type=int)
        text = request.form.get('text', '').strip()
        org_id = request.form.get('organization_id', type=int)
        sort_order = request.form.get('sort_order', 0, type=int)

        if not text:
            flash('Матн киритинг', 'warning')
            return redirect(url_for('deficiencies_list', date=sel.isoformat()))

        if not current_user.is_admin:
            if not org_id:
                abort(403)
            check_org_access(org_id)
        elif org_id:
            check_org_access(org_id)

        deficiency_fields = ['id', 'work_date', 'text', 'organization_id', 'sort_order']
        before = None
        if did:
            d = Deficiency.query.get(did)
            if not d:
                abort(404)
            if d.organization_id:
                check_org_access(d.organization_id)
            elif not current_user.is_admin:
                abort(403)
            before = model_snapshot(d, deficiency_fields)
            d.text = text
            d.organization_id = org_id if org_id else None
            d.sort_order = sort_order
            d.work_date = sel
        else:
            d = Deficiency(
                work_date=sel, text=text,
                organization_id=org_id if org_id else None,
                sort_order=sort_order, created_by=current_user.id
            )
            db.session.add(d)
        db.session.flush()
        after = model_snapshot(d, deficiency_fields)
        log_audit(db, audit_action_name('deficiency', did), entity_type='deficiency', entity_id=d.id,
                  entity_label=(text[:120] if text else str(d.id)), module='deficiencies',
                  before=before, after=after, changes=diff_dict(before, after),
                  description='Deficiency saved')
        db.session.commit()
        flash('Камчилик сақланди', 'success')
        return redirect(url_for('deficiencies_list', date=sel.isoformat()))

    @app.route('/deficiencies/delete/<int:did>', methods=['POST'])
    @module_required('deficiencies')
    @editor_required
    def delete_deficiency(did):
        d = Deficiency.query.get_or_404(did)
        if d.organization_id:
            check_org_access(d.organization_id)
        elif not current_user.is_admin:
            abort(403)
        sel = d.work_date
        before = model_snapshot(d, ['id', 'work_date', 'text', 'organization_id', 'sort_order'])
        db.session.delete(d)
        log_audit(db, 'deficiency_deleted', entity_type='deficiency', entity_id=did,
                  entity_label=(before.get('text') or '')[:120], module='deficiencies',
                  before=before, description='Deficiency deleted')
        db.session.commit()
        flash('Камчилик ўчирилди', 'warning')
        return redirect(url_for('deficiencies_list', date=sel.isoformat()))

    # в”Ђв”Ђв”Ђ REPORT в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/report', methods=['GET', 'POST'])
    @module_required('transport')
    @login_required
    def report():
        if request.method == 'POST':
            mode = request.form.get('mode', 'day')
            if mode == 'day':
                d_from = parse_date(request.form.get('report_date'))
                d_to = d_from
            elif mode == 'week':
                today = date.today()
                d_from = today - timedelta(days=today.weekday())
                d_to = d_from + timedelta(days=6)
            elif mode == 'month':
                today = date.today()
                d_from = today.replace(day=1)
                if today.month == 12:
                    d_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    d_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            else:
                d_from = parse_date(request.form.get('date_from'))
                d_to = parse_date(request.form.get('date_to'))
                if d_from > d_to:
                    d_from, d_to = d_to, d_from

            org_ids_form = request.form.getlist('org_ids')
            cat_filter_form = request.form.getlist('cat_codes')

            all_org_ids = current_user.get_org_ids()
            if org_ids_form:
                org_ids = [int(x) for x in org_ids_form if x.isdigit() and int(x) in all_org_ids]
            else:
                org_ids = all_org_ids

            cat_filter = [c for c in cat_filter_form if c in CATEGORIES] if cat_filter_form else None

            report_type = request.form.get('report_type', 'main')
            export_lang = getattr(current_user, 'language', 'uz') or 'uz'
            if export_lang not in ('uz', 'ru'):
                export_lang = 'uz'

            if report_type == 'daily_activity':
                filepath = generate_daily_activity(d_from, app.config['REPORTS_DIR'], org_ids, lang=export_lang)
                if export_lang == 'ru':
                    fname = f'Dnevnaya_zanyatost_{d_from.strftime("%d_%m_%Y")}.xlsx'
                else:
                    fname = f'Kunlik_ish_{d_from.strftime("%d_%m_%Y")}.xlsx'
            else:
                filepath = generate_report(d_from, d_to, app.config['REPORTS_DIR'], org_ids, cat_filter, lang=export_lang)
                prefix = 'Otchet' if export_lang == 'ru' else 'Hisobot'
                if d_from == d_to:
                    fname = f'{prefix}_{d_from.strftime("%d_%m_%Y")}.xlsx'
                else:
                    fname = f'{prefix}_{d_from.strftime("%d_%m_%Y")}_{d_to.strftime("%d_%m_%Y")}.xlsx'
            return send_file(filepath, as_attachment=True, download_name=fname)

        # GET: show summary and a limited preview table for selected range.
        mode, d_from, d_to = parse_date_range(request.args)
        user_org_ids = current_user.get_org_ids()
        selected_org_ids_r = [
            oid for oid in request.args.getlist('org_ids', type=int)
            if oid in user_org_ids
        ]
        legacy_org_id = request.args.get('org_id', type=int)
        if not selected_org_ids_r and legacy_org_id and legacy_org_id in user_org_ids:
            selected_org_ids_r = [legacy_org_id]
        active_org_ids_r = selected_org_ids_r or user_org_ids

        selected_cat_codes = [
            code for code in request.args.getlist('cat_codes')
            if code in CATEGORIES
        ]

        organizations = get_user_orgs()
        q = (DailyRecord.query
             .filter(DailyRecord.work_date >= d_from,
                     DailyRecord.work_date <= d_to)
             .join(Equipment).join(Organization)
             .filter(Organization.id.in_(active_org_ids_r)))
        if selected_cat_codes:
            q = q.filter(Equipment.category.in_(selected_cat_codes))

        records = (q.order_by(DailyRecord.work_date.desc(),
                              Organization.sort_order,
                              Equipment.name,
                              DailyRecord.line_index)
                   .all())

        preview = {
            'cash': sum(r.amount_cash or 0 for r in records),
            'transfer': sum(r.amount_transfer or 0 for r in records),
            'internal': sum(r.amount_internal or 0 for r in records),
            'other': sum(r.amount_other or 0 for r in records),
            'total_quantity': sum((r.quantity or 0) for r in records if r.status == 'working'),
            'records_count': len(records),
            'worked_count': sum(1 for r in records if r.status == 'working'),
            'idle_count': sum(1 for r in records if r.status != 'working'),
            'equipment_count': len({r.equipment_id for r in records}),
        }
        preview['total'] = preview['cash'] + preview['transfer'] + preview['internal'] + preview['other']

        org_summary_map = {}
        work_type_summary_map = {}
        for r in records:
            eq = r.equipment
            org = eq.organization if eq else None
            amount = r.total_amount or 0
            qty = r.quantity or 0

            if org:
                item = org_summary_map.setdefault(org.id, {
                    'name': org.name,
                    'amount': 0,
                    'quantity': 0,
                    'records_count': 0,
                    'equipment_ids': set(),
                })
                item['amount'] += amount
                if r.status == 'working':
                    item['quantity'] += qty
                item['records_count'] += 1
                item['equipment_ids'].add(r.equipment_id)

            if r.status == 'working':
                wt_name = (r.work_type or '').strip() or '—'
                wt = work_type_summary_map.setdefault(wt_name, {
                    'name': wt_name,
                    'amount': 0,
                    'quantity': 0,
                    'records_count': 0,
                })
                wt['amount'] += amount
                wt['quantity'] += qty
                wt['records_count'] += 1

        org_summary = []
        for item in org_summary_map.values():
            item = dict(item)
            item['equipment_count'] = len(item.pop('equipment_ids'))
            org_summary.append(item)
        org_summary.sort(key=lambda x: x['amount'], reverse=True)

        work_type_summary = sorted(
            work_type_summary_map.values(),
            key=lambda x: x['amount'],
            reverse=True
        )[:12]

        preview_rows = []
        for r in records[:300]:
            eq = r.equipment
            org = eq.organization if eq else None
            preview_rows.append({
                'date': r.work_date,
                'organization': org.name if org else '',
                'equipment': eq.name if eq else '',
                'plate': eq.plate if eq else '',
                'category': CATEGORIES.get(eq.category, eq.category) if eq else '',
                'status': r.status,
                'work_type': r.work_type or '',
                'customer': r.customer or '',
                'quantity': r.quantity,
                'unit': r.unit or '',
                'payment_type': r.payment_type or '',
                'total': r.total_amount or 0,
                'note': r.note or r.idle_reason or '',
            })

        return render_template('report.html',
                               mode=mode, d_from=d_from, d_to=d_to,
                               organizations=organizations,
                               selected_org_ids=selected_org_ids_r or user_org_ids,
                               selected_cat_codes=selected_cat_codes,
                               preview=preview,
                               org_summary=org_summary,
                               work_type_summary=work_type_summary,
                               preview_rows=preview_rows,
                               hidden_rows_count=max(0, len(records) - len(preview_rows)),
                               categories=CATEGORIES)

    # в”Ђв”Ђв”Ђ ADMIN: Users в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/admin/users')
    @admin_required
    def admin_users():
        users = User.query.order_by(User.username).all()
        orgs = Organization.query.order_by(Organization.sort_order).all()
        return render_template('admin_users.html', users=users, organizations=orgs)

    @app.route('/admin/users/save', methods=['POST'])
    @admin_required
    def admin_save_user():
        uid = request.form.get('id', type=int)
        username = request.form.get('username', '').strip().lower()
        full_name = request.form.get('full_name', '').strip()
        role = request.form.get('role', ROLE_OPERATOR)
        password = request.form.get('password', '').strip()
        is_active = request.form.get('is_active') == 'on'
        org_ids = request.form.getlist('org_ids', type=int)

        if uid:
            user = User.query.get(uid)
            if not user:
                abort(404)
            before = {'username': user.username, 'full_name': user.full_name, 'role': user.role, 'is_active': user.is_active_user}
            user.username = username
            user.full_name = full_name
            user.role = role
            user.is_active_user = is_active
            action = 'user_updated'
            if password:
                ok, msg = password_is_strong_enough(username, password)
                if not ok:
                    flash(msg, 'warning')
                    return redirect(url_for('admin_users'))
                user.set_password(password)
                action = 'user_password_reset'
        else:
            if User.query.filter_by(username=username).first():
                flash(f'"{username}" фойдаланувчиси мавжуд', 'warning')
                return redirect(url_for('admin_users'))
            if not password:
                flash('Парол киритинг', 'warning')
                return redirect(url_for('admin_users'))
            ok, msg = password_is_strong_enough(username, password)
            if not ok:
                flash(msg, 'warning')
                return redirect(url_for('admin_users'))
            user = User(username=username, full_name=full_name, role=role,
                        is_active_user=is_active)
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            before = None
            action = 'user_created'

        if role != ROLE_ADMIN:
            user.organizations = Organization.query.filter(Organization.id.in_(org_ids)).all()
        else:
            user.organizations = []
        db.session.flush()
        if password:
            require_temp_password_change(db, user.id)
        after = {'username': user.username, 'full_name': user.full_name, 'role': user.role, 'is_active': user.is_active_user,
                 'org_ids': [o.id for o in user.organizations], 'temporary_password_required': bool(password)}
        log_audit(db, action, entity_type='user', entity_id=user.id, entity_label=user.username,
                  module='admin', before=before, after=after, description='Admin saved user account')
        db.session.commit()
        if password:
            flash(f'Фойдаланувчи "{username}" сақланди. Биринчи киришда паролни алмаштириши шарт.', 'success')
        else:
            flash(f'Фойдаланувчи "{username}" сақланди', 'success')
        return redirect(url_for('admin_users'))

    @app.route('/admin/users/delete/<int:uid>', methods=['POST'])
    @admin_required
    def admin_delete_user(uid):
        user = User.query.get_or_404(uid)
        if user.id == current_user.id:
            flash('Ўзингизни блоклай олмайсиз', 'warning')
            return redirect(url_for('admin_users'))
        before = {'username': user.username, 'is_active': user.is_active_user}
        user.is_active_user = False
        log_audit(db, 'user_blocked', entity_type='user', entity_id=user.id, entity_label=user.username,
                  module='admin', before=before, after={'is_active': False}, description='User blocked instead of deleted')
        db.session.commit()
        flash(f'Фойдаланувчи "{user.username}" блокланди', 'warning')
        return redirect(url_for('admin_users'))

    @app.route('/admin/permissions')
    @admin_required
    def admin_permissions():
        users = User.query.filter(User.role != ROLE_ADMIN).order_by(User.username).all()
        modules = AppModule.query.filter_by(is_active=True).order_by(AppModule.id).all()
        perms = {}
        for perm in UserModulePermission.query.all():
            perms[(perm.user_id, perm.module_code)] = perm.has_access
        return render_template('admin_permissions.html', users=users, modules=modules, perms=perms)

    @app.route('/admin/permissions/save', methods=['POST'])
    @admin_required
    def admin_permissions_save():
        users = User.query.filter(User.role != ROLE_ADMIN).all()
        modules = AppModule.query.filter_by(is_active=True).all()
        for user in users:
            for module in modules:
                key = f'perm_{user.id}_{module.code}'
                has_access = request.form.get(key) is not None
                perm = UserModulePermission.query.filter_by(
                    user_id=user.id, module_code=module.code).first()
                if perm:
                    perm.has_access = has_access
                else:
                    perm = UserModulePermission(
                        user_id=user.id, module_code=module.code, has_access=has_access)
                    db.session.add(perm)
        log_audit(db, 'module_permissions_updated', entity_type='permissions', module='admin',
                  description='Module permissions updated')
        db.session.commit()
        flash('Ҳуқуқлар сақланди', 'success')
        return redirect(url_for('admin_permissions'))


    @app.route('/admin/audit')
    @admin_required
    def admin_audit_logs():
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        user_id = request.args.get('user_id', type=int)
        action = request.args.get('action', '').strip()
        module = request.args.get('module', '').strip()
        q = 'SELECT * FROM audit_logs WHERE 1=1'
        params = {}
        if date_from:
            q += ' AND created_at >= :date_from'
            params['date_from'] = date_from + 'T00:00:00'
        if date_to:
            q += ' AND created_at <= :date_to'
            params['date_to'] = date_to + 'T23:59:59'
        if user_id:
            q += ' AND user_id = :user_id'
            params['user_id'] = user_id
        if action:
            q += ' AND action = :action'
            params['action'] = action
        if module:
            q += ' AND module = :module'
            params['module'] = module
        q += ' ORDER BY created_at DESC LIMIT 300'
        raw_logs = db.session.execute(text(q), params).mappings().all()

        # Audit timestamps are stored in UTC. Display them in Uzbekistan local time (UTC+5).
        logs = []
        for row in raw_logs:
            item = dict(row)
            raw_created_at = item.get('created_at')
            try:
                dt = datetime.fromisoformat(str(raw_created_at).replace('Z', ''))
                item['created_at_local'] = (dt + timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                item['created_at_local'] = raw_created_at
            logs.append(item)

        users = User.query.order_by(User.username).all()
        actions = [r[0] for r in db.session.execute(text('SELECT DISTINCT action FROM audit_logs ORDER BY action')).all()]
        modules = [r[0] for r in db.session.execute(text('SELECT DISTINCT module FROM audit_logs WHERE module IS NOT NULL AND module != "" ORDER BY module')).all()]
        return render_template('audit_logs.html', logs=logs, users=users, actions=actions, modules=modules,
                               date_from=date_from, date_to=date_to, selected_user_id=user_id,
                               selected_action=action, selected_module=module)

    # в”Ђв”Ђв”Ђ REFERENCES в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/ref/organizations')
    @module_required('transport')
    @login_required
    def ref_organizations():
        if current_user.is_admin:
            orgs = Organization.query.order_by(Organization.sort_order).all()
        else:
            orgs = get_user_orgs()
        org_delete_info = {}
        for org in orgs:
            linked = {
                'equipment_count': Equipment.query.filter_by(organization_id=org.id).count(),
                'fuel_warehouse_count': FuelWarehouse.query.filter_by(organization_id=org.id).count(),
                'spare_part_request_count': SparePartRequest.query.filter_by(organization_id=org.id).count(),
                'deficiency_count': Deficiency.query.filter_by(organization_id=org.id).count(),
                'user_count': org.users.count() if hasattr(org, 'users') else 0,
            }
            org_delete_info[org.id] = {
                'can_delete': not any(linked.values()),
                'linked_total': sum(linked.values()),
                'linked': linked,
            }
        return render_template('ref_organizations.html', organizations=orgs,
                               org_delete_info=org_delete_info)

    @app.route('/ref/organizations/save', methods=['POST'])
    @module_required('transport')
    @admin_required
    def save_organization():
        oid = request.form.get('id', type=int)
        name = normalize_ref_text(request.form.get('name', ''))
        short = normalize_ref_text(request.form.get('short_name', ''))

        if not name or len(name) < 2:
            flash(ui_t('Ташкилот номини тўлиқ киритинг', 'Введите корректное название организации'), 'warning')
            return redirect(url_for('ref_organizations'))

        if ref_duplicate_exists(Organization, 'name', name, exclude_id=oid):
            flash(ui_t('Бундай ташкилот мавжуд', 'Такая организация уже существует'), 'warning')
            return redirect(url_for('ref_organizations'))

        org_fields = ['id', 'name', 'short_name', 'sort_order']
        before = None
        if oid:
            o = Organization.query.get_or_404(oid)
            before = model_snapshot(o, org_fields)
            sort = o.sort_order
            o.name, o.short_name, o.sort_order = name, short, sort
        else:
            max_sort = db.session.query(db.func.max(Organization.sort_order)).scalar() or 0
            sort = max_sort + 1
            o = Organization(name=name, short_name=short, sort_order=sort)
            db.session.add(o)
        db.session.flush()
        after = model_snapshot(o, org_fields)
        log_audit(db, audit_action_name('organization', oid), entity_type='organization', entity_id=o.id,
                  entity_label=o.name, module='references', before=before, after=after,
                  changes=diff_dict(before, after), description='Organization saved')
        db.session.commit()
        flash('Сақланди', 'success')
        return redirect(url_for('ref_organizations'))

    @app.route('/ref/organizations/delete/<int:oid>', methods=['POST'])
    @module_required('transport')
    @admin_required
    def delete_organization(oid):
        o = Organization.query.get_or_404(oid)
        before = model_snapshot(o, ['id', 'name', 'short_name', 'sort_order'])
        linked = {
            'equipment_count': Equipment.query.filter_by(organization_id=oid).count(),
            'fuel_warehouse_count': FuelWarehouse.query.filter_by(organization_id=oid).count(),
            'spare_part_request_count': SparePartRequest.query.filter_by(organization_id=oid).count(),
            'deficiency_count': Deficiency.query.filter_by(organization_id=oid).count(),
            'user_count': o.users.count() if hasattr(o, 'users') else 0,
        }
        if any(linked.values()):
            log_audit(db, 'organization_delete_blocked', entity_type='organization', entity_id=oid,
                      entity_label=before.get('name', ''), module='references', before=before,
                      after=linked, description='Organization delete blocked because linked records exist')
            db.session.commit()
            flash(ui_t('Ташкилот ўчирилмади: боғланган маълумотлар мавжуд',
                       'Организация не удалена: есть связанные данные'), 'warning')
            return redirect(url_for('ref_organizations'))
        db.session.delete(o)
        log_audit(db, 'organization_deleted', entity_type='organization', entity_id=oid,
                  entity_label=before.get('name', ''), module='references', before=before,
                  description='Organization deleted')
        db.session.commit()
        flash('Ўчирилди', 'warning')
        return redirect(url_for('ref_organizations'))

    @app.route('/ref/equipment')
    @module_required('transport')
    @login_required
    def ref_equipment():
        org_id    = request.args.get('org_id', type=int)
        cat_codes = request.args.getlist('cat_codes')
        eq_types  = request.args.getlist('eq_types')
        orgs      = get_user_orgs()
        user_org_ids = [o.id for o in orgs]

        q = (Equipment.query.join(Organization)
             .filter(Organization.id.in_(user_org_ids))
             .order_by(Organization.sort_order, Equipment.category, Equipment.name))

        if org_id:
            check_org_access(org_id)
            q = q.filter(Equipment.organization_id == org_id)

        if cat_codes:
            q = q.filter(Equipment.category.in_(cat_codes))

        if eq_types:
            q = q.filter(Equipment.eq_type.in_(eq_types))

        equipment = q.all()
        total_count = len(equipment)

        all_eq_types = sorted(set(
            e[0] for e in Equipment.query
            .filter(Equipment.organization_id.in_(user_org_ids))
            .with_entities(Equipment.eq_type).all()
            if e[0]
        ))

        equipment_delete_info = {}
        for eq in equipment:
            linked = {
                'daily_records_count': DailyRecord.query.filter_by(equipment_id=eq.id).count(),
                'engine_hours_count': EngineHoursRecord.query.filter_by(equipment_id=eq.id).count(),
                'wialon_mapping_count': VialonMapping.query.filter_by(equipment_id=eq.id).count(),
                'spare_part_request_count': SparePartRequest.query.filter_by(equipment_id=eq.id).count(),
            }
            linked_total = sum(linked.values())
            equipment_delete_info[eq.id] = {
                'can_delete': linked_total == 0,
                'can_deactivate': linked_total > 0 and bool(eq.is_active),
                'is_disabled': linked_total > 0 and not bool(eq.is_active),
                'linked_total': linked_total,
                'linked': linked,
            }

        return render_template('ref_equipment.html',
            equipment=equipment, organizations=orgs,
            selected_org_id=org_id, categories=CATEGORIES,
            selected_cats=cat_codes, eq_types=all_eq_types,
            selected_eq_types=eq_types, total_count=total_count,
            equipment_delete_info=equipment_delete_info)

    @app.route('/ref/equipment/export')
    @module_required('transport')
    @login_required
    def export_equipment():
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        org_id    = request.args.get('org_id', type=int)
        cat_codes = request.args.getlist('cat_codes')
        eq_types  = request.args.getlist('eq_types')
        orgs      = get_user_orgs()
        user_org_ids = [o.id for o in orgs]

        q = (Equipment.query.join(Organization)
             .filter(Organization.id.in_(user_org_ids))
             .order_by(Organization.sort_order, Equipment.category, Equipment.name))
        if org_id:
            check_org_access(org_id)
            q = q.filter(Equipment.organization_id == org_id)
        if cat_codes:
            q = q.filter(Equipment.category.in_(cat_codes))
        if eq_types:
            q = q.filter(Equipment.eq_type.in_(eq_types))

        equipment = q.all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Texnika'

        filter_info = []
        if org_id:
            org = Organization.query.get(org_id)
            filter_info.append(f'Tashkilot: {org.name}')
        if cat_codes:
            filter_info.append('Kategoriyalar: ' + ', '.join(CATEGORIES.get(c, c) for c in cat_codes))
        if eq_types:
            filter_info.append('Turlar: ' + ', '.join(eq_types))
        if not filter_info:
            filter_info.append('Barcha texnika')

        ws.merge_cells('A1:G1')
        ws['A1'].value = '"Buxoro Agroklastr" MChJ вЂ” Texnika royxati'
        ws['A1'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        ws['A2'].value = ' | '.join(filter_info) + f' | Sana: {date.today().strftime("%d.%m.%Y")}'
        ws['A2'].font = Font(name='Times New Roman', size=11)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:G3')
        ws['A3'].value = f'Jami: {len(equipment)} ta texnika'
        ws['A3'].font = Font(name='Times New Roman', size=11, bold=True)
        ws['A3'].alignment = Alignment(horizontal='left')

        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        headers = ['N', 'Tashkilot', 'Kategoriya', 'Turi', 'Texnika nomi', 'Davlat raqami', 'Faol']
        widths  = [5,   25,          28,            20,     30,              15,               8]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = Font(name='Times New Roman', size=11, bold=True)
            c.fill = yellow
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[c.column_letter].width = w

        for i, eq in enumerate(equipment, 1):
            row = 4 + i
            vals = [i, eq.organization.name, CATEGORIES.get(eq.category, eq.category),
                    eq.eq_type, eq.name, eq.plate, 'Ha' if eq.is_active else 'Yoq']
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Times New Roman', size=10)
                c.alignment = Alignment(vertical='center', wrap_text=True)
            ws.row_dimensions[row].height = 18

        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'Tehnika_{datetime.now().strftime("%d_%m_%Y")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/ref/equipment/save', methods=['POST'])
    @module_required('transport')
    @editor_required
    def save_equipment():
        eid = request.form.get('id', type=int)
        org_id = request.form.get('organization_id', type=int)
        if not org_id:
            flash(ui_t('Ташкилотни танланг', 'Выберите организацию'), 'warning')
            return redirect(url_for('ref_equipment'))
        check_org_access(org_id)

        name = normalize_ref_text(request.form.get('name',''))
        plate = normalize_plate(request.form.get('plate',''))
        eq_type = normalize_ref_text(request.form.get('eq_type',''))
        default_unit = normalize_ref_text(request.form.get('default_unit',''))
        category = request.form.get('category','mtz')

        if not name:
            flash(ui_t('Техника номини киритинг', 'Введите название техники'), 'warning')
            return redirect(url_for('ref_equipment', org_id=org_id))
        if not plate:
            flash(ui_t('Давлат рақамини киритинг', 'Введите государственный номер техники'), 'warning')
            return redirect(url_for('ref_equipment', org_id=org_id))
        if equipment_plate_duplicate_exists(plate, exclude_id=eid):
            flash(ui_t('Бу давлат рақами бошқа техникага бириктирилган', 'Этот госномер уже привязан к другой технике'), 'warning')
            return redirect(url_for('ref_equipment', org_id=org_id))
        if category not in CATEGORIES:
            flash(ui_t('Категория нотўғри танланган', 'Некорректно выбрана категория'), 'warning')
            return redirect(url_for('ref_equipment', org_id=org_id))
        try:
            default_price = parse_non_negative_float(request.form.get('default_price'), 'default_price', allow_empty=True)
        except ValueError:
            flash(ui_t('Нарх манфий бўлмаслиги керак', 'Цена не может быть отрицательной'), 'warning')
            return redirect(url_for('ref_equipment', org_id=org_id))
        kw = dict(name=name,
                  plate=plate,
                  category=category,
                  eq_type=eq_type,
                  organization_id=org_id,
                  default_price=default_price or 0,
                  default_unit=default_unit)
        eq_fields = ['id', 'name', 'plate', 'category', 'eq_type', 'organization_id', 'default_price', 'default_unit', 'is_active']
        before = None
        if eid:
            eq = Equipment.query.get_or_404(eid)
            check_org_access(eq.organization_id)
            before = model_snapshot(eq, eq_fields)
            for k, v in kw.items():
                setattr(eq, k, v)
        else:
            eq = Equipment(**kw)
            db.session.add(eq)
        db.session.flush()
        after = model_snapshot(eq, eq_fields)
        log_audit(db, audit_action_name('equipment', eid), entity_type='equipment', entity_id=eq.id,
                  entity_label=f'{eq.name} {eq.plate}'.strip(), module='references',
                  before=before, after=after, changes=diff_dict(before, after),
                  description='Equipment saved')
        db.session.commit()
        flash('Сақланди', 'success')
        return redirect(url_for('ref_equipment', org_id=org_id))

    @app.route('/ref/equipment/delete/<int:eid>', methods=['POST'])
    @module_required('transport')
    @editor_required
    def delete_equipment(eid):
        eq = Equipment.query.get_or_404(eid)
        check_org_access(eq.organization_id)
        oid = eq.organization_id
        fields = ['id', 'name', 'plate', 'category', 'eq_type', 'organization_id', 'default_price', 'default_unit', 'is_active']
        before = model_snapshot(eq, fields)
        linked = {
            'daily_records_count': DailyRecord.query.filter_by(equipment_id=eid).count(),
            'engine_hours_count': EngineHoursRecord.query.filter_by(equipment_id=eid).count(),
            'wialon_mapping_count': VialonMapping.query.filter_by(equipment_id=eid).count(),
            'spare_part_request_count': SparePartRequest.query.filter_by(equipment_id=eid).count(),
        }
        label = f"{before.get('name', '')} {before.get('plate', '')}".strip()
        if any(linked.values()):
            if eq.is_active:
                eq.is_active = False
                db.session.flush()
                after = model_snapshot(eq, fields)
                after.update(linked)
                action = 'equipment_delete_blocked_deactivated'
                description = 'Equipment delete blocked; equipment was deactivated because linked records exist'
            else:
                after = dict(linked)
                action = 'equipment_delete_blocked'
                description = 'Equipment delete blocked because linked records exist'
            log_audit(db, action, entity_type='equipment', entity_id=eid,
                      entity_label=label, module='references', before=before, after=after,
                      description=description)
            db.session.commit()
            flash(ui_t('Техника ўчирилмади: боғланган маълумотлар мавжуд. Техника фаол эмас қилинди.',
                       'Техника не удалена: есть связанные данные. Техника отключена.'), 'warning')
            return redirect(url_for('ref_equipment', org_id=oid))
        db.session.delete(eq)
        log_audit(db, 'equipment_deleted', entity_type='equipment', entity_id=eid,
                  entity_label=label, module='references', before=before, description='Equipment deleted')
        db.session.commit()
        flash('Ўчирилди', 'warning')
        return redirect(url_for('ref_equipment', org_id=oid))

    @app.route('/ref/equipment/enable/<int:eid>', methods=['POST'])
    @module_required('transport')
    @editor_required
    def enable_equipment(eid):
        eq = Equipment.query.get_or_404(eid)
        check_org_access(eq.organization_id)
        fields = ['id', 'name', 'plate', 'category', 'eq_type', 'organization_id', 'default_price', 'default_unit', 'is_active']
        before = model_snapshot(eq, fields)
        label = f"{before.get('name', '')} {before.get('plate', '')}".strip()
        if not eq.is_active:
            eq.is_active = True
            db.session.flush()
        linked = {
            'daily_records_count': DailyRecord.query.filter_by(equipment_id=eid).count(),
            'engine_hours_count': EngineHoursRecord.query.filter_by(equipment_id=eid).count(),
            'wialon_mapping_count': VialonMapping.query.filter_by(equipment_id=eid).count(),
            'spare_part_request_count': SparePartRequest.query.filter_by(equipment_id=eid).count(),
        }
        after = model_snapshot(eq, fields)
        after.update(linked)
        log_audit(db, 'equipment_reactivated', entity_type='equipment', entity_id=eid,
                  entity_label=label, module='references', before=before, after=after,
                  changes=diff_dict(before, after), description='Equipment reactivated')
        db.session.commit()
        flash(ui_t('Техника қайта фаол қилинди.', 'Техника включена.'), 'success')
        return redirect(url_for('ref_equipment', org_id=eq.organization_id))

    @app.route('/ref/work_types')
    @module_required('transport')
    @login_required
    def ref_work_types():
        work_types = WorkType.query.order_by(WorkType.name).all()
        work_type_usage = {
            wt.id: DailyRecord.query.filter(DailyRecord.work_type == wt.name).count()
            for wt in work_types
        }
        return render_template('ref_work_types.html',
                               work_types=work_types,
                               work_type_usage=work_type_usage)

    @app.route('/ref/work_types/save', methods=['POST'])
    @module_required('transport')
    @editor_required
    def save_work_type():
        wid = request.form.get('id', type=int)
        name = normalize_ref_text(request.form.get('name',''))
        unit = normalize_ref_text(request.form.get('default_unit',''))
        if not name:
            flash(ui_t('Иш тури номини киритинг', 'Введите название вида работ'), 'warning')
            return redirect(url_for('ref_work_types'))
        if ref_duplicate_exists(WorkType, 'name', name, exclude_id=wid):
            flash(ui_t('Бундай иш тури мавжуд', 'Такой вид работ уже существует'), 'warning')
            return redirect(url_for('ref_work_types'))
        try:
            price = parse_non_negative_float(request.form.get('default_price'), 'default_price', allow_empty=True) or 0
        except ValueError:
            flash(ui_t('Нарх манфий бўлмаслиги керак', 'Цена не может быть отрицательной'), 'warning')
            return redirect(url_for('ref_work_types'))
        wt_fields = ['id', 'name', 'default_unit', 'default_price']
        before = None
        if wid:
            w = WorkType.query.get_or_404(wid)
            before = model_snapshot(w, wt_fields)
            w.name, w.default_unit, w.default_price = name, unit, price
        else:
            w = WorkType(name=name, default_unit=unit, default_price=price)
            db.session.add(w)
        db.session.flush()
        after = model_snapshot(w, wt_fields)
        log_audit(db, audit_action_name('work_type', wid), entity_type='work_type', entity_id=w.id,
                  entity_label=w.name, module='references', before=before, after=after,
                  changes=diff_dict(before, after), description='Work type saved')
        db.session.commit()
        flash('Сақланди', 'success')
        return redirect(url_for('ref_work_types'))

    @app.route('/ref/work_types/delete/<int:wid>', methods=['POST'])
    @module_required('transport')
    @admin_required
    def delete_work_type(wid):
        w = WorkType.query.get_or_404(wid)
        before = model_snapshot(w, ['id', 'name', 'default_unit', 'default_price'])
        used_count = DailyRecord.query.filter(DailyRecord.work_type == w.name).count()
        if used_count:
            after = {'daily_records_count': used_count}
            log_audit(db, 'work_type_delete_blocked', entity_type='work_type', entity_id=wid,
                      entity_label=before.get('name', ''), module='references', before=before,
                      after=after, description='Work type delete blocked because it is used in daily records')
            db.session.commit()
            flash(ui_t('Иш тури ўчирилмади: ҳисоботларда ишлатилган',
                       'Вид работ не удалён: используется в отчётах'), 'warning')
            return redirect(url_for('ref_work_types'))
        db.session.delete(w)
        log_audit(db, 'work_type_deleted', entity_type='work_type', entity_id=wid,
                  entity_label=before.get('name', ''), module='references', before=before,
                  description='Work type deleted')
        db.session.commit()
        flash('Ўчирилди', 'warning')
        return redirect(url_for('ref_work_types'))

    @app.route('/ref/customers')
    @module_required('transport')
    @login_required
    def ref_customers():
        customers = Customer.query.order_by(Customer.name).all()
        customer_usage = {
            customer.id: DailyRecord.query.filter(DailyRecord.customer == customer.name).count()
            for customer in customers
        }
        return render_template('ref_customers.html',
                               customers=customers,
                               customer_usage=customer_usage)

    @app.route('/ref/customers/save', methods=['POST'])
    @module_required('transport')
    @editor_required
    def save_customer():
        cid = request.form.get('id', type=int)
        name = normalize_ref_text(request.form.get('name',''))
        ctype = request.form.get('customer_type','external')
        if not name:
            flash(ui_t('Буюртмачи номини киритинг', 'Введите название заказчика'), 'warning')
            return redirect(url_for('ref_customers'))
        if ref_duplicate_exists(Customer, 'name', name, exclude_id=cid):
            flash(ui_t('Бундай буюртмачи мавжуд', 'Такой заказчик уже существует'), 'warning')
            return redirect(url_for('ref_customers'))
        if ctype not in {'external', 'internal'}:
            ctype = 'external'
        customer_fields = ['id', 'name', 'customer_type']
        before = None
        if cid:
            c = Customer.query.get_or_404(cid)
            before = model_snapshot(c, customer_fields)
            c.name, c.customer_type = name, ctype
        else:
            c = Customer(name=name, customer_type=ctype)
            db.session.add(c)
        db.session.flush()
        after = model_snapshot(c, customer_fields)
        log_audit(db, audit_action_name('customer', cid), entity_type='customer', entity_id=c.id,
                  entity_label=c.name, module='references', before=before, after=after,
                  changes=diff_dict(before, after), description='Customer saved')
        db.session.commit()
        flash('Сақланди', 'success')
        return redirect(url_for('ref_customers'))

    @app.route('/ref/customers/delete/<int:cid>', methods=['POST'])
    @module_required('transport')
    @admin_required
    def delete_customer(cid):
        c = Customer.query.get_or_404(cid)
        before = model_snapshot(c, ['id', 'name', 'customer_type'])
        used_count = DailyRecord.query.filter(DailyRecord.customer == c.name).count()
        if used_count:
            after = {'daily_records_count': used_count}
            log_audit(db, 'customer_delete_blocked', entity_type='customer', entity_id=cid,
                      entity_label=before.get('name', ''), module='references', before=before,
                      after=after, description='Customer delete blocked because it is used in daily records')
            db.session.commit()
            flash(ui_t('Буюртмачи ўчирилмади: ҳисоботларда ишлатилган',
                       'Заказчик не удалён: используется в отчётах'), 'warning')
            return redirect(url_for('ref_customers'))
        db.session.delete(c)
        log_audit(db, 'customer_deleted', entity_type='customer', entity_id=cid,
                  entity_label=before.get('name', ''), module='references', before=before,
                  description='Customer deleted')
        db.session.commit()
        flash('Ўчирилди', 'warning')
        return redirect(url_for('ref_customers'))

    # в”Ђв”Ђв”Ђ API в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.route('/api/work_type_defaults/<int:wid>')
    @module_required('transport')
    @login_required
    def api_work_type_defaults(wid):
        wt = WorkType.query.get_or_404(wid)
        return jsonify(default_unit=wt.default_unit, default_price=wt.default_price)

    # в”Ђв”Ђв”Ђ WIALON IMPORT в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    register_wialon_routes(app, editor_required, admin_required)

    # в”Ђв”Ђв”Ђ FUEL (TOPAZ) в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    app.register_blueprint(fuel_bp)

    # [REASON]: Legacy compatibility alias. Older Topaz agent configs may POST to
    # /api/fuel_sync before being updated to the canonical /fuel/api/fuel_sync.
    # This alias applies the same token validation and sync logic with no duplication.
    # The legacy path should be removed once all agent configs are confirmed updated.
    @app.route('/api/fuel_sync', methods=['POST'])
    def api_fuel_sync_legacy():
        app.logger.warning(
            'Topaz agent used deprecated endpoint /api/fuel_sync. '
            'Update agent configuration to /fuel/api/fuel_sync.'
        )
        return _perform_fuel_sync()

    # в”Ђв”Ђв”Ђ SPARE PARTS в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    app.register_blueprint(spare_parts_bp)

    # в”Ђв”Ђв”Ђ ERRORS в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    @app.errorhandler(400)
    def bad_request(e):
        return render_template('error.html', code=400,
                               message='Сўров хавфсизлик текширувидан ўтмади'), 400

    @app.errorhandler(403)
    def forbidden(e):
        return render_template('error.html', code=403,
                               message='Сизда бу саҳифага кириш ҳуқуқи йўқ'), 403

    @app.errorhandler(404)
    def not_found(e):
        return render_template('error.html', code=404,
                               message='Саҳифа топилмади'), 404

    with app.app_context():
        db.create_all()
        if AppModule.query.count() == 0:
            for code, name_uz, name_ru in [
                ('transport', 'РўСЂР°РЅСЃРїРѕСЂС‚ ТіРёСЃРѕР±РѕС‚Рё', 'РўСЂР°РЅСЃРїРѕСЂС‚РЅС‹Р№ РѕС‚С‡С’С‚'),
                ('wialon', 'Р’РёР°Р»РѕРЅ', 'Р’РёР°Р»РѕРЅ GPS'),
                ('fuel', 'РђР—РЎ РјРѕРґСѓР»Рё', 'РњРѕРґСѓР»СЊ РђР—РЎ'),
                ('deficiencies', 'РљР°РјС‡РёР»РёРєР»Р°СЂ', 'РќРµРґРѕСЃС‚Р°С‚РєРё'),
                ('spare_parts', 'Р­ТіС‚РёС’С‚ Т›РёСЃРјР»Р°СЂ', 'Р—Р°РїС‡Р°СЃС‚Рё'),
            ]:
                db.session.add(AppModule(code=code, name_uz=name_uz, name_ru=name_ru))
            db.session.commit()

    return app


app = create_app()

if __name__ == '__main__':
    app.run(debug=app.config.get('DEBUG', False), host='0.0.0.0', port=5050)
