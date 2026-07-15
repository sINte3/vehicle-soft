"""
fuel_routes.py — АЗС модуль v2
Flask Blueprint: /fuel/* и /api/fuel_*

Логика:
  Склад (FuelWarehouse) = организация.
  У склада — несколько АЗС (FuelStation2), каждая с topaz_id.
  Агент Топаз присылает транзакции по topaz_id → списываем со склада.
  Баланс = НачОстаток + Приходы - Расходы (транзакции).
"""

import hmac
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, jsonify, abort, current_app, g, send_file)
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
import hashlib
from io import BytesIO
from sqlalchemy import text
from sqlalchemy import func, and_, or_
from sqlalchemy.orm import joinedload
from types import SimpleNamespace

from models import (
    db, Organization,
    FuelWarehouse, FuelStation2, FuelInitialBalance,
    FuelReceipt2, FuelTransaction2, FuelSyncLog2, FuelWarningReview,
    FuelCard, FuelCardAlias, FuelCardSyncLog,
    module_required,
)

from sec003a_ext import log_audit

fuel_bp = Blueprint('fuel', __name__, url_prefix='/fuel')


def fuel_t(uz, ru):
    # [REASON]: Local flash-message helper for route-level flash() messages.
    # Prefer the persisted user language because POST redirects can otherwise
    # be confusing during module-level UI tests. Fall back to g.lang.
    lang = getattr(g, 'lang', 'uz') or 'uz'
    try:
        if current_user.is_authenticated:
            lang = getattr(current_user, 'language', lang) or lang
    except Exception:
        pass
    return ru if lang == 'ru' else uz


def fuel_format_errors(errors, title_uz=None, title_ru=None):
    unique = []
    seen = set()
    for err in errors or []:
        text = str(err).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        unique.append(text)
    if not unique:
        return ''
    title = fuel_t(title_uz or 'Маълумот сақланмади. Хатоларни тузатинг:',
                   title_ru or 'Данные не сохранены. Исправьте ошибки:')
    return title + '\n' + '\n'.join('- ' + item for item in unique)


def fuel_flash_errors(errors, title_uz=None, title_ru=None):
    msg = fuel_format_errors(errors, title_uz, title_ru)
    if msg:
        flash(msg, 'warning')


def _parse_fuel_date(value, field_label):
    try:
        if not value:
            raise ValueError()
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        raise ValueError(field_label)


def _parse_non_negative(value, field_label, allow_empty=True):
    if value is None or str(value).strip() == '':
        if allow_empty:
            return None
        raise ValueError(field_label)
    try:
        result = float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        raise ValueError(field_label)
    if result < 0:
        raise ValueError(field_label)
    return result


def _parse_float_required(value, field_label):
    if value is None or str(value).strip() == '':
        raise ValueError(field_label)
    try:
        return float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        raise ValueError(field_label)


def _parse_positive(value, field_label):
    result = _parse_non_negative(value, field_label, allow_empty=False)
    if result <= 0:
        raise ValueError(field_label)
    return result



TOPAZ_AUDIT_ACTOR = SimpleNamespace(
    id=None,
    username='system/topaz_agent',
    full_name='Topaz Fuel Agent',
    role='system',
)

# [FUEL-SYNC-013] Known Topaz technical/service cards that must never be
# counted as vehicle fuel consumption. CardID 30 = "ПЕРЕЛИВ" — a Topaz-
# generated counter-mismatch reconciliation record, not a real dispensing
# event. Add more IDs here if similar technical cards are discovered later.
EXCLUDED_CARD_NUMBERS = {'30'}


def _iso_date(value):
    return value.isoformat() if value else None


def _fuel_warehouse_snapshot(warehouse):
    if not warehouse:
        return None
    return {
        'id': getattr(warehouse, 'id', None),
        'name': getattr(warehouse, 'name', '') or '',
        'organization_id': getattr(warehouse, 'organization_id', None),
        'notes': getattr(warehouse, 'notes', '') or '',
    }


def _fuel_station_snapshot(station):
    if not station:
        return None
    return {
        'id': getattr(station, 'id', None),
        'name': getattr(station, 'name', '') or '',
        'topaz_id': getattr(station, 'topaz_id', None),
        'warehouse_id': getattr(station, 'warehouse_id', None),
        'is_active': bool(getattr(station, 'is_active', False)),
    }


# [REASON]: FUEL-REPORT-012H-C card resolver — resolves FuelTransaction2.card_number
# to FuelCard.display_name. Supports both topaz_card_id and rfid alias types.
# Must not display raw card_number/RFID when no mapping exists — show em-dash instead.
def _resolve_card_names(txns):
    """Given a list of FuelTransaction2 rows, return a dict mapping card_number -> display_name."""
    card_numbers = set()
    for txn in txns:
        val = (txn.card_number or '').strip()
        if val:
            card_numbers.add(val)

    if not card_numbers:
        return {}

    aliases = (FuelCardAlias.query
               .filter(FuelCardAlias.alias_type.in_(['topaz_card_id', 'rfid']),
                       FuelCardAlias.alias_value.in_(list(card_numbers)))
               .all())

    result = {}
    for alias in aliases:
        if alias.card and alias.card.display_name:
            result[alias.alias_value] = alias.card.display_name

    return result


def _card_display_name(card_number, name_map):
    """Return display name from map, or em-dash if not found."""
    if not card_number or not name_map:
        return '—'
    return name_map.get(card_number.strip(), '—')


def _card_name_map_for_station_report(station_id, start_date, end_date):
    """Build card_number -> display_name map for station issues report."""
    from datetime import datetime as dt
    txns = (FuelTransaction2.query
            .filter(FuelTransaction2.station_id == station_id,
                    FuelTransaction2.txn_datetime >= dt.combine(start_date, dt.min.time()),
                    FuelTransaction2.txn_datetime <= dt.combine(end_date, dt.max.time()))
            .all())
    return _resolve_card_names(txns)


def _fuel_initial_balance_snapshot(balance):
    if not balance:
        return None
    return {
        'id': getattr(balance, 'id', None),
        'warehouse_id': getattr(balance, 'warehouse_id', None),
        'fuel_type': getattr(balance, 'fuel_type', '') or '',
        'quantity': getattr(balance, 'quantity', 0) or 0,
        'balance_date': _iso_date(getattr(balance, 'balance_date', None)),
        'note': getattr(balance, 'note', '') or '',
    }


def _fuel_receipt_snapshot(receipt):
    if not receipt:
        return None
    quantity = getattr(receipt, 'quantity', 0) or 0
    price = getattr(receipt, 'price_per_liter', 0) or 0
    return {
        'id': getattr(receipt, 'id', None),
        'warehouse_id': getattr(receipt, 'warehouse_id', None),
        'receipt_date': _iso_date(getattr(receipt, 'receipt_date', None)),
        'fuel_type': getattr(receipt, 'fuel_type', '') or '',
        'quantity': quantity,
        'price_per_liter': price,
        'amount': quantity * price,
        'supplier': getattr(receipt, 'supplier', '') or '',
        'doc_number': getattr(receipt, 'doc_number', '') or '',
        'note': getattr(receipt, 'note', '') or '',
    }


def _audit_fuel(action, entity_type='', entity_id=None, entity_label='', before=None,
                after=None, changes=None, status='ok', description='', actor_user=None):
    log_audit(
        db,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_label=entity_label,
        module='fuel',
        before=before,
        after=after,
        changes=changes,
        status=status,
        description=description,
        actor_user=actor_user,
    )


# [REASON]: Token is no longer hardcoded; read per-request from app config which
# reads FUEL_API_TOKEN from the environment. If not configured, all sync requests
# are rejected (deny-all safe default). See docs/DEPLOYMENT_SECURITY.md.
FUEL_TYPES = ['ДТ']


# ─── FUEL-REPORT-012B/012C: Fuel reports center + station issues ──────


def _fuel_stations_active():
    """Return active fuel stations sorted by name, ordered by name and topaz_id."""
    return (FuelStation2.query
            .filter(FuelStation2.is_active.is_(True))
            .order_by(FuelStation2.name, FuelStation2.topaz_id)
            .all())


def _fuel_station_issues_query(station_id, start_date, end_date):
    """Return station issue transactions for given station and period.
    Respects station validity dates when present."""
    q = (FuelTransaction2.query
         .options(joinedload(FuelTransaction2.station))
         .filter(FuelTransaction2.station_id == station_id,
                 FuelTransaction2.txn_datetime >= datetime.combine(start_date, datetime.min.time()),
                 FuelTransaction2.txn_datetime <= datetime.combine(end_date, datetime.max.time()))
         .order_by(FuelTransaction2.txn_datetime, FuelTransaction2.id))

    # [REASON]: Filter by station validity dates per business rule 5:
    # valid_from <= transaction date when valid_from is not null
    # valid_to >= transaction date when valid_to is not null
    station = FuelStation2.query.get(station_id)
    if station:
        if station.valid_from:
            q = q.filter(FuelTransaction2.txn_datetime >=
                         datetime.combine(station.valid_from, datetime.min.time()))
        if station.valid_to:
            q = q.filter(FuelTransaction2.txn_datetime <=
                         datetime.combine(station.valid_to, datetime.max.time()))

    return q.all()


def _fuel_station_issue_summary(txns):
    """Return summary dict from a list of station issue transactions."""
    return {
        'count': len(txns),
        'total_liters': round(sum(t.quantity or 0 for t in txns), 2),
    }


def _fuel_station_issues_workbook(txns, station_name, topaz_id, start_date, end_date, card_name_map=None):
    """Build Excel workbook for station issues with outline grouping by date."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Выдача АЗС"

    # Title
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1).value = f"Выдача топлива по АЗС: {station_name} (Topaz {topaz_id})"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Period info
    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1).value = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    ws.cell(row=2, column=1).font = Font(size=11, color='666666')
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Headers
    headers = ['№', 'Дата/время', 'АЗС', 'Topaz ID', 'Имя карты', 'Литры']
    header_row = 4
    for col, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col).value = title

    header_fill = PatternFill('solid', fgColor='D9EAD3')
    header_font = Font(bold=True)
    thin = Side(style='thin', color='D1D5DB')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Outline: 0=date header, 1=transaction row
    ws.sheet_properties.outlinePr.summaryBelow = True

    grouped = {}
    for txn in txns:
        date_key = txn.txn_datetime.strftime('%Y-%m-%d') if txn.txn_datetime else 'unknown'
        grouped.setdefault(date_key, []).append(txn)

    row_idx = 0
    for date_key in sorted(grouped.keys()):
        day_txns = grouped[date_key]
        date_display = datetime.strptime(date_key, '%Y-%m-%d').strftime('%d.%m.%Y')
        subtotal = round(sum(t.quantity or 0 for t in day_txns), 2)

        # Date header row (outline level 0)
        r = header_row + 1 + row_idx
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).value = f"{date_display}  —  {len(day_txns)} выдач, всего {subtotal:.2f} л"
        ws.cell(row=r, column=1).font = Font(bold=True, size=11, color='1A6B3C')
        ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='F0F9F0')
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row_idx += 1

        for seq, txn in enumerate(day_txns, start=1):
            r = header_row + 1 + row_idx
            ws.row_dimensions[r].outlineLevel = 1
            st_name = txn.station.name if txn.station else station_name
            card_name_val = _card_display_name(txn.card_number, card_name_map) if card_name_map else '—'
            ws.cell(row=r, column=1).value = seq
            ws.cell(row=r, column=2).value = txn.txn_datetime.strftime('%d.%m.%Y %H:%M') if txn.txn_datetime else ''
            ws.cell(row=r, column=3).value = st_name
            ws.cell(row=r, column=4).value = topaz_id
            ws.cell(row=r, column=5).value = card_name_val
            ws.cell(row=r, column=6).value = round(txn.quantity or 0, 2)
            ws.cell(row=r, column=6).number_format = '#,##0.00'
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            row_idx += 1

        # Subtotal row
        r = header_row + 1 + row_idx
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1).value = f"Итого за {date_display}"
        ws.cell(row=r, column=1).font = Font(bold=True, italic=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=6).value = subtotal
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        ws.cell(row=r, column=6).font = Font(bold=True)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row_idx += 1

    # Grand total row
    grand_total = round(sum(t.quantity or 0 for t in txns), 2)
    r = header_row + 1 + row_idx
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).value = "ВСЕГО"
    ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).value = grand_total
    ws.cell(row=r, column=6).number_format = '#,##0.00'
    ws.cell(row=r, column=6).font = Font(bold=True, size=12)
    total_fill = PatternFill('solid', fgColor='FEF3C7')
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = total_fill
        ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    col_widths = {1: 5, 2: 18, 3: 22, 4: 12, 5: 28, 6: 12}
    for col_n, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_n)].width = width

    # Freeze panes and auto filter
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:F{r}'

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return wb


@fuel_bp.route('/reports')
@module_required('fuel')
def reports_center():
    """FUEL-REPORT-012B: Fuel reports center page."""
    return render_template('fuel/reports.html')


@fuel_bp.route('/reports/station-issues')
@module_required('fuel')
def station_issues_report():
    """FUEL-REPORT-012C: Station issue detail report."""
    today = date.today()
    default_start = today.replace(day=1)
    default_end = today

    start_date = _fuel_report_parse_date(request.args.get('start_date'), default_start)
    end_date = _fuel_report_parse_date(request.args.get('end_date'), default_end)
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    station_id = request.args.get('station_id', type=int)
    stations = _fuel_stations_active()

    txns = []
    station_name = ''
    topaz_id = ''
    summary = {'count': 0, 'total_liters': 0.0}
    card_name_map = {}

    if station_id:
        selected = FuelStation2.query.get(station_id)
        if selected:
            station_name = selected.name
            topaz_id = selected.topaz_id
            txns = _fuel_station_issues_query(station_id, start_date, end_date)
            # [REASON]: FUEL-REPORT-012C grouping fix - precompute a safe date string
            # for each transaction so that Jinja's groupby works without calling
            # txn_datetime.date() (which fails because method references are not comparable).
            for txn in txns:
                if txn.txn_datetime and hasattr(txn.txn_datetime, 'strftime'):
                    txn.txn_date = txn.txn_datetime.strftime('%Y-%m-%d')
                else:
                    txn.txn_date = str(txn.txn_datetime or '')[:10]
            summary = _fuel_station_issue_summary(txns)
            card_name_map = _resolve_card_names(txns)

    return render_template('fuel/station_issues_report.html',
                           stations=stations,
                           station_id=station_id,
                           station_name=station_name,
                           topaz_id=topaz_id,
                           start_date=start_date,
                           end_date=end_date,
                           txns=txns,
                           summary=summary,
                            card_name_map=card_name_map)


@fuel_bp.route('/reports/station-issues/export')
@module_required('fuel')
def station_issues_export():
    """FUEL-REPORT-012C Excel export for station issue detail report."""
    today = date.today()
    default_start = today.replace(day=1)
    default_end = today

    start_date = _fuel_report_parse_date(request.args.get('start_date'), default_start)
    end_date = _fuel_report_parse_date(request.args.get('end_date'), default_end)
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    station_id = request.args.get('station_id', type=int)
    if not station_id:
        return 'station_id is required', 400

    station = FuelStation2.query.get(station_id)
    if not station:
        return 'Station not found', 404

    txns = _fuel_station_issues_query(station_id, start_date, end_date)
    card_name_map = _resolve_card_names(txns)

    wb = _fuel_station_issues_workbook(txns, station.name, station.topaz_id, start_date, end_date, card_name_map=card_name_map)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"station_issues_{station.topaz_id}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def parse_fuel_station_date(value):
    """Parse YYYY-MM-DD date from fuel station forms."""
    value = (value or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def parse_topaz_txn_datetime(value):
    """Parse Topaz transaction datetime with safe fallback."""
    value = (value or '').strip()
    if not value:
        return datetime.utcnow()
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return datetime.utcnow()


def station_matches_txn_datetime(station, txn_dt):
    """Return True if station validity period allows this transaction date."""
    if not station or not txn_dt:
        return False
    txn_date = txn_dt.date() if hasattr(txn_dt, 'date') else txn_dt
    if station.valid_from and txn_date < station.valid_from:
        return False
    if station.valid_to and txn_date > station.valid_to:
        return False
    return True


def resolve_fuel_station_for_topaz(topaz_id, txn_dt=None):
    """Resolve Topaz column to station without active-only filtering.

    Historical/inactive stations must still match old transactions.
    If validity dates exist, transaction date is used to choose the station.
    """
    try:
        topaz_id = int(topaz_id or 0)
    except (TypeError, ValueError):
        return None
    if not topaz_id:
        return None

    stations = FuelStation2.query.filter_by(topaz_id=topaz_id).all()
    if not stations:
        return None

    if txn_dt:
        dated = [st for st in stations if station_matches_txn_datetime(st, txn_dt)]
        if dated:
            active_dated = [st for st in dated if st.is_active]
            return active_dated[0] if active_dated else dated[0]

        has_any_validity = any(st.valid_from or st.valid_to for st in stations)
        if has_any_validity:
            return None

    active = [st for st in stations if st.is_active]
    return active[0] if active else stations[0]



# ─── Helpers ─────────────────────────────────────────────────────────

def admin_required_fuel(f):
    from functools import wraps
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# PERF-DASH-001 V3B: bulk fuel dashboard/report helpers.
# [REASON]: Avoid repeated per-warehouse/per-fuel SELECTs on fuel dashboard/report pages.
def _fuel_balance_map(warehouse_ids, fuel_type=None):
    ids = [int(x) for x in (warehouse_ids or []) if x]
    result = {wh_id: {} for wh_id in ids}

    if not ids:
        return result

    ib_query = (FuelInitialBalance.query
                .filter(FuelInitialBalance.warehouse_id.in_(ids))
                .order_by(FuelInitialBalance.warehouse_id,
                          FuelInitialBalance.fuel_type,
                          FuelInitialBalance.id))

    if fuel_type:
        ib_query = ib_query.filter(FuelInitialBalance.fuel_type == fuel_type)

    initial_rows = ib_query.all()

    initial_map = {}
    fuel_types_by_wh = {}

    for ib in initial_rows:
        wh_id = int(ib.warehouse_id)
        key = (wh_id, ib.fuel_type)
        if key not in initial_map:
            initial_map[key] = ib
            fuel_types_by_wh.setdefault(wh_id, [])
            if ib.fuel_type not in fuel_types_by_wh[wh_id]:
                fuel_types_by_wh[wh_id].append(ib.fuel_type)

    receipt_conditions = []
    expense_conditions = []

    for (wh_id, ft), ib in initial_map.items():
        receipt_conditions.append(and_(
            FuelReceipt2.warehouse_id == wh_id,
            FuelReceipt2.fuel_type == ft,
            FuelReceipt2.receipt_date >= ib.balance_date,
        ))
        expense_conditions.append(and_(
            FuelStation2.warehouse_id == wh_id,
            FuelTransaction2.fuel_type == ft,
            FuelTransaction2.txn_datetime >= datetime.combine(ib.balance_date, datetime.min.time()),
        ))

    receipt_sums = {}
    if receipt_conditions:
        receipt_rows = (db.session.query(
                FuelReceipt2.warehouse_id,
                FuelReceipt2.fuel_type,
                func.coalesce(func.sum(FuelReceipt2.quantity), 0)
            )
            .filter(or_(*receipt_conditions))
            .group_by(FuelReceipt2.warehouse_id, FuelReceipt2.fuel_type)
            .all())

        receipt_sums = {
            (int(wh_id), ft): float(total or 0)
            for wh_id, ft, total in receipt_rows
        }

    expense_sums = {}
    if expense_conditions:
        expense_rows = (db.session.query(
                FuelStation2.warehouse_id,
                FuelTransaction2.fuel_type,
                func.coalesce(func.sum(FuelTransaction2.quantity), 0)
            )
            .join(FuelTransaction2, FuelTransaction2.station_id == FuelStation2.id)
            .filter(or_(*expense_conditions))
            .group_by(FuelStation2.warehouse_id, FuelTransaction2.fuel_type)
            .all())

        expense_sums = {
            (int(wh_id), ft): float(total or 0)
            for wh_id, ft, total in expense_rows
        }

    for wh_id in ids:
        ftypes = fuel_types_by_wh.get(wh_id)

        if not ftypes:
            if fuel_type:
                if fuel_type == 'ДТ':
                    result[wh_id]['ДТ'] = None
            else:
                result[wh_id]['ДТ'] = None
            continue

        for ft in ftypes:
            ib = initial_map.get((wh_id, ft))
            if not ib:
                result[wh_id][ft] = None
                continue

            receipts = receipt_sums.get((wh_id, ft), 0.0)
            expenses = expense_sums.get((wh_id, ft), 0.0)
            result[wh_id][ft] = round(float(ib.quantity or 0) + receipts - expenses, 2)

    return result


def _fuel_station_count_map(warehouse_ids, active_only=False):
    ids = [int(x) for x in (warehouse_ids or []) if x]
    if not ids:
        return {}

    q = (db.session.query(FuelStation2.warehouse_id, func.count(FuelStation2.id))
         .filter(FuelStation2.warehouse_id.in_(ids)))

    if active_only:
        q = q.filter(FuelStation2.is_active.is_(True))

    return {
        int(wh_id): int(cnt or 0)
        for wh_id, cnt in q.group_by(FuelStation2.warehouse_id).all()
    }


def _fuel_today_expense_map(warehouse_ids):
    ids = [int(x) for x in (warehouse_ids or []) if x]
    if not ids:
        return {}

    day_start = datetime.combine(date.today(), datetime.min.time())
    day_end = day_start + timedelta(days=1)

    rows = (db.session.query(
            FuelStation2.warehouse_id,
            func.coalesce(func.sum(FuelTransaction2.quantity), 0)
        )
        .join(FuelTransaction2, FuelTransaction2.station_id == FuelStation2.id)
        .filter(FuelStation2.warehouse_id.in_(ids),
                FuelTransaction2.txn_datetime >= day_start,
                FuelTransaction2.txn_datetime < day_end)
        .group_by(FuelStation2.warehouse_id)
        .all())

    return {int(wh_id): float(total or 0) for wh_id, total in rows}


def _fuel_latest_txn_map(warehouse_ids, start_dt=None, end_dt=None, station_id=None):
    ids = [int(x) for x in (warehouse_ids or []) if x]
    if not ids:
        return {}

    max_query = (db.session.query(
            FuelStation2.warehouse_id,
            func.max(FuelTransaction2.txn_datetime)
        )
        .join(FuelTransaction2, FuelTransaction2.station_id == FuelStation2.id)
        .filter(FuelStation2.warehouse_id.in_(ids)))

    if start_dt is not None:
        max_query = max_query.filter(FuelTransaction2.txn_datetime >= start_dt)
    if end_dt is not None:
        max_query = max_query.filter(FuelTransaction2.txn_datetime <= end_dt)
    if station_id:
        max_query = max_query.filter(FuelTransaction2.station_id == station_id)

    max_rows = max_query.group_by(FuelStation2.warehouse_id).all()

    conditions = []
    for wh_id, max_dt in max_rows:
        if max_dt is None:
            continue
        cond = and_(FuelStation2.warehouse_id == wh_id,
                    FuelTransaction2.txn_datetime == max_dt)
        if station_id:
            cond = and_(cond, FuelTransaction2.station_id == station_id)
        conditions.append(cond)

    if not conditions:
        return {}

    txns = (FuelTransaction2.query
            .options(joinedload(FuelTransaction2.station))
            .join(FuelStation2)
            .filter(or_(*conditions))
            .order_by(FuelTransaction2.txn_datetime.desc(),
                      FuelTransaction2.id.desc())
            .all())

    result = {}
    for tx in txns:
        wh_id = getattr(getattr(tx, 'station', None), 'warehouse_id', None)
        if wh_id is not None and int(wh_id) not in result:
            result[int(wh_id)] = tx

    return result


def get_warehouse_balance(warehouse_id, fuel_type=None):
    """
    Возвращает dict {fuel_type: current_liters}.
    current = initial_balance + receipts - transactions  (всё после initial.balance_date).
    """
    return _fuel_balance_map([warehouse_id], fuel_type=fuel_type).get(int(warehouse_id), {})


def get_all_balances():
    """Балансы всех складов для дашборда. Returns list of dicts."""
    warehouses = fuel_warehouse_query_for_ui().order_by(FuelWarehouse.name).all()
    warehouse_ids = [wh.id for wh in warehouses]

    balances_by_wh = _fuel_balance_map(warehouse_ids)
    latest_txn_by_wh = _fuel_latest_txn_map(warehouse_ids)
    today_expense_by_wh = _fuel_today_expense_map(warehouse_ids)
    active_station_count_by_wh = _fuel_station_count_map(warehouse_ids, active_only=True)

    rows = []
    for wh in warehouses:
        rows.append({
            'warehouse': wh,
            'balances': balances_by_wh.get(wh.id, {}),
            'last_txn': latest_txn_by_wh.get(wh.id),
            'today_expense': today_expense_by_wh.get(wh.id, 0),
            'stations': active_station_count_by_wh.get(wh.id, 0),
        })

    return rows


# Fuel management report

def _fuel_report_lang():
    lang = getattr(g, 'lang', 'uz') or 'uz'
    try:
        if current_user.is_authenticated:
            lang = getattr(current_user, 'language', lang) or lang
    except Exception:
        pass
    return 'ru' if lang == 'ru' else 'uz'


def _fuel_report_label(ru, uz):
    return ru if _fuel_report_lang() == 'ru' else uz


def _parse_report_date(value, default):
    try:
        return datetime.strptime(value or '', '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return default


def _sum_receipts_for_period(warehouse_id, start_date, end_date):
    return float(db.session.query(func.coalesce(func.sum(FuelReceipt2.quantity), 0))
                 .filter(FuelReceipt2.warehouse_id == warehouse_id,
                         FuelReceipt2.receipt_date >= start_date,
                         FuelReceipt2.receipt_date <= end_date)
                 .scalar() or 0)


def _sum_issues_for_period(warehouse_id, start_dt, end_dt, station_id=None):
    q = (db.session.query(func.coalesce(func.sum(FuelTransaction2.quantity), 0))
         .join(FuelStation2)
         .filter(FuelStation2.warehouse_id == warehouse_id,
                 FuelTransaction2.txn_datetime >= start_dt,
                 FuelTransaction2.txn_datetime <= end_dt))
    if station_id:
        q = q.filter(FuelTransaction2.station_id == station_id)
    return float(q.scalar() or 0)


def _fuel_opening_balance(warehouse_id, d_from):
    ib = (FuelInitialBalance.query
          .filter_by(warehouse_id=warehouse_id, fuel_type='ДТ')
          .first())
    if not ib:
        return None, None
    start_date = ib.balance_date
    if d_from <= start_date:
        return float(ib.quantity or 0), ib
    before_date = d_from - timedelta(days=1)
    receipts_before = _sum_receipts_for_period(warehouse_id, start_date, before_date)
    issues_before = _sum_issues_for_period(
        warehouse_id,
        datetime.combine(start_date, datetime.min.time()),
        datetime.combine(before_date, datetime.max.time()),
    )
    return round(float(ib.quantity or 0) + receipts_before - issues_before, 2), ib


FUEL_LARGE_TXN_THRESHOLD = 500.0
FUEL_SYNC_STALE_HOURS = 12


def _fuel_warning_key(code, entity_type='', entity_id=None, value=None):
    raw = '|'.join([
        str(code or ''),
        str(entity_type or ''),
        str(entity_id or ''),
        str(value if value is not None else ''),
    ])
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _fuel_warning(code, severity, title_ru, title_uz, details_ru='', details_uz='', value=None,
                  entity_type='', entity_id=None, key_value=None):
    key_seed = value if key_value is None else key_value
    warning_key = _fuel_warning_key(code, entity_type, entity_id, key_seed)
    return {
        'key': warning_key,
        'code': code,
        'severity': severity,
        'title_ru': title_ru,
        'title_uz': title_uz,
        'details_ru': details_ru,
        'details_uz': details_uz,
        'value': value,
        'entity_type': entity_type or '',
        'entity_id': entity_id,
        'review_status': 'new',
        'review_comment': '',
        'review_updated_at': None,
        'reviewer_name': '',
        'is_current': True,
    }


def _fuel_report_warning_text(warning, lang='uz'):
    title = warning.get('title_ru') if lang == 'ru' else warning.get('title_uz')
    details = warning.get('details_ru') if lang == 'ru' else warning.get('details_uz')
    return title or '', details or ''


def _fuel_report_warning_summary(warnings):
    warnings = warnings or []
    return {
        'total': len(warnings),
        'danger': sum(1 for w in warnings if w.get('severity') == 'danger'),
        'warning': sum(1 for w in warnings if w.get('severity') == 'warning'),
        'info': sum(1 for w in warnings if w.get('severity') == 'info'),
    }


FUEL_WARNING_STATUSES = {
    'new': ('Новое', 'Янги'),
    'in_progress': ('В работе', 'Ишда'),
    'resolved': ('Проверено', 'Текширилди'),
    'rejected': ('Отклонено', 'Рад этилди'),
}


def _fuel_warning_status_label(status, lang='uz'):
    ru, uz = FUEL_WARNING_STATUSES.get(status or 'new', FUEL_WARNING_STATUSES['new'])
    return ru if lang == 'ru' else uz


def _fuel_warning_severity_label(severity, lang='uz'):
    labels = {
        'danger': ('Критично', 'Критик'),
        'warning': ('Предупреждение', 'Огоҳлантириш'),
        'info': ('Информация', 'Маълумот'),
    }
    ru, uz = labels.get(severity or 'warning', labels['warning'])
    return ru if lang == 'ru' else uz


def _fuel_warning_reviews_by_key(keys):
    keys = [k for k in (keys or []) if k]
    if not keys:
        return {}
    rows = FuelWarningReview.query.filter(FuelWarningReview.warning_key.in_(keys)).all()
    return {r.warning_key: r for r in rows}


def _apply_fuel_warning_reviews(warnings):
    reviews = _fuel_warning_reviews_by_key([w.get('key') for w in warnings])
    for warning in warnings:
        review = reviews.get(warning.get('key'))
        if not review:
            continue
        warning['review_id'] = review.id
        warning['review_status'] = review.status or 'new'
        warning['review_comment'] = review.comment or ''
        warning['review_updated_at'] = review.updated_at
        warning['reviewer_name'] = (review.reviewer.full_name or review.reviewer.username) if review.reviewer else ''
    return warnings


def _fuel_warning_snapshot(warning, lang='uz'):
    title, details = _fuel_report_warning_text(warning, lang='ru')
    return {
        'warning_key': warning.get('key') or '',
        'warning_code': warning.get('code') or '',
        'severity': warning.get('severity') or 'warning',
        'entity_type': warning.get('entity_type') or '',
        'entity_id': warning.get('entity_id'),
        'title_snapshot': title or warning.get('title_ru') or warning.get('title_uz') or '',
        'details_snapshot': details or warning.get('details_ru') or warning.get('details_uz') or '',
        'value_snapshot': str(warning.get('value') if warning.get('value') is not None else ''),
    }


def _fuel_warning_from_review(review):
    return {
        'key': review.warning_key,
        'code': review.warning_code,
        'severity': review.severity or 'warning',
        'title_ru': review.title_snapshot or review.warning_code,
        'title_uz': review.title_snapshot or review.warning_code,
        'details_ru': review.details_snapshot or '',
        'details_uz': review.details_snapshot or '',
        'value': review.value_snapshot or '',
        'entity_type': review.entity_type or '',
        'entity_id': review.entity_id,
        'review_id': review.id,
        'review_status': review.status or 'new',
        'review_comment': review.comment or '',
        'review_updated_at': review.updated_at,
        'reviewer_name': (review.reviewer.full_name or review.reviewer.username) if review.reviewer else '',
        'is_current': False,
    }


def _collect_fuel_report_data(d_from, d_to, warehouse_id=None, station_id=None):
    d_from_dt = datetime.combine(d_from, datetime.min.time())
    d_to_dt = datetime.combine(d_to, datetime.max.time())

    selected_station = None
    if station_id:
        selected_station = FuelStation2.query.get(station_id)
        if selected_station:
            warehouse_id = selected_station.warehouse_id
        else:
            station_id = None

    wh_query = FuelWarehouse.query.order_by(FuelWarehouse.name)
    if warehouse_id:
        wh_query = wh_query.filter(FuelWarehouse.id == warehouse_id)
    warehouses = wh_query.all()

    warehouse_rows = []
    totals = {
        'opening': 0.0,
        'receipts': 0.0,
        'issued': 0.0,
        'ending': 0.0,
        'warehouses': len(warehouses),
        'stations': 0,
        'transactions': 0,
        'negative_balances': 0,
        'missing_initial': 0,
    }

    warehouse_ids = [wh.id for wh in warehouses]

    initial_rows = (FuelInitialBalance.query
                    .filter(FuelInitialBalance.warehouse_id.in_(warehouse_ids),
                            FuelInitialBalance.fuel_type == 'ДТ')
                    .order_by(FuelInitialBalance.warehouse_id, FuelInitialBalance.id)
                    .all())

    initial_by_wh = {}
    for ib in initial_rows:
        wh_id = int(ib.warehouse_id)
        if wh_id not in initial_by_wh:
            initial_by_wh[wh_id] = ib

    before_receipt_conditions = []
    before_issue_conditions = []

    for wh_id, ib in initial_by_wh.items():
        if d_from <= ib.balance_date:
            continue

        before_date = d_from - timedelta(days=1)
        before_receipt_conditions.append(and_(
            FuelReceipt2.warehouse_id == wh_id,
            FuelReceipt2.receipt_date >= ib.balance_date,
            FuelReceipt2.receipt_date <= before_date,
        ))

        before_issue_conditions.append(and_(
            FuelStation2.warehouse_id == wh_id,
            FuelTransaction2.txn_datetime >= datetime.combine(ib.balance_date, datetime.min.time()),
            FuelTransaction2.txn_datetime <= datetime.combine(before_date, datetime.max.time()),
        ))

    receipts_before_by_wh = {}
    if before_receipt_conditions:
        rows = (db.session.query(
                FuelReceipt2.warehouse_id,
                func.coalesce(func.sum(FuelReceipt2.quantity), 0)
            )
            .filter(or_(*before_receipt_conditions))
            .group_by(FuelReceipt2.warehouse_id)
            .all())
        receipts_before_by_wh = {int(wh_id): float(total or 0) for wh_id, total in rows}

    issues_before_by_wh = {}
    if before_issue_conditions:
        rows = (db.session.query(
                FuelStation2.warehouse_id,
                func.coalesce(func.sum(FuelTransaction2.quantity), 0)
            )
            .join(FuelTransaction2, FuelTransaction2.station_id == FuelStation2.id)
            .filter(or_(*before_issue_conditions))
            .group_by(FuelStation2.warehouse_id)
            .all())
        issues_before_by_wh = {int(wh_id): float(total or 0) for wh_id, total in rows}

    period_receipts_by_wh = {}
    if warehouse_ids:
        rows = (db.session.query(
                FuelReceipt2.warehouse_id,
                func.coalesce(func.sum(FuelReceipt2.quantity), 0)
            )
            .filter(FuelReceipt2.warehouse_id.in_(warehouse_ids),
                    FuelReceipt2.receipt_date >= d_from,
                    FuelReceipt2.receipt_date <= d_to)
            .group_by(FuelReceipt2.warehouse_id)
            .all())
        period_receipts_by_wh = {int(wh_id): float(total or 0) for wh_id, total in rows}

    tx_stats_by_wh = {}
    if warehouse_ids:
        tx_stats_q = (db.session.query(
                FuelStation2.warehouse_id,
                func.coalesce(func.sum(FuelTransaction2.quantity), 0),
                func.count(FuelTransaction2.id),
            )
            .join(FuelTransaction2, FuelTransaction2.station_id == FuelStation2.id)
            .filter(FuelStation2.warehouse_id.in_(warehouse_ids),
                    FuelTransaction2.txn_datetime >= d_from_dt,
                    FuelTransaction2.txn_datetime <= d_to_dt))

        if station_id:
            tx_stats_q = tx_stats_q.filter(FuelTransaction2.station_id == station_id)

        for wh_id, issued_total, tx_count in tx_stats_q.group_by(FuelStation2.warehouse_id).all():
            tx_stats_by_wh[int(wh_id)] = {
                'issued': float(issued_total or 0),
                'tx_count': int(tx_count or 0),
            }

    stations_count_by_wh = _fuel_station_count_map(warehouse_ids, active_only=False)
    last_txn_by_wh = _fuel_latest_txn_map(warehouse_ids, start_dt=d_from_dt, end_dt=d_to_dt, station_id=station_id)

    for wh in warehouses:
        initial_balance = initial_by_wh.get(wh.id)

        if not initial_balance:
            opening = None
        elif d_from <= initial_balance.balance_date:
            opening = float(initial_balance.quantity or 0)
        else:
            opening = round(
                float(initial_balance.quantity or 0)
                + receipts_before_by_wh.get(wh.id, 0.0)
                - issues_before_by_wh.get(wh.id, 0.0),
                2,
            )

        receipts = period_receipts_by_wh.get(wh.id, 0.0)
        tx_stats = tx_stats_by_wh.get(wh.id, {})
        issued = float(tx_stats.get('issued') or 0)
        tx_count = int(tx_stats.get('tx_count') or 0)
        stations_count = int(stations_count_by_wh.get(wh.id, 0))
        last_txn = last_txn_by_wh.get(wh.id)

        ending = None if opening is None else round(opening + receipts - issued, 2)

        if opening is None:
            totals['missing_initial'] += 1
        else:
            totals['opening'] += opening
            totals['ending'] += ending or 0
            if ending is not None and ending < 0:
                totals['negative_balances'] += 1
        totals['receipts'] += receipts
        totals['issued'] += issued
        totals['transactions'] += tx_count
        totals['stations'] += stations_count

        warehouse_rows.append({
            'warehouse': wh,
            'opening': opening,
            'receipts': receipts,
            'issued': issued,
            'ending': ending,
            'initial_balance': initial_balance,
            'stations_count': stations_count,
            'tx_count': tx_count,
            'last_txn': last_txn,
        })

    station_query = (db.session.query(
            FuelStation2.id,
            FuelStation2.name,
            FuelStation2.topaz_id,
            FuelStation2.is_active,
            FuelWarehouse.name.label('warehouse_name'),
            func.coalesce(func.sum(FuelTransaction2.quantity), 0).label('issued'),
            func.count(FuelTransaction2.id).label('tx_count'),
            func.max(FuelTransaction2.txn_datetime).label('last_txn'),
        )
        .join(FuelWarehouse, FuelWarehouse.id == FuelStation2.warehouse_id)
        .outerjoin(FuelTransaction2,
                   (FuelTransaction2.station_id == FuelStation2.id) &
                   (FuelTransaction2.txn_datetime >= d_from_dt) &
                   (FuelTransaction2.txn_datetime <= d_to_dt)))
    if warehouse_id:
        station_query = station_query.filter(FuelStation2.warehouse_id == warehouse_id)
    if station_id:
        station_query = station_query.filter(FuelStation2.id == station_id)
    station_rows = station_query.group_by(
        FuelStation2.id, FuelStation2.name, FuelStation2.topaz_id,
        FuelStation2.is_active, FuelWarehouse.name,
    ).order_by(func.coalesce(func.sum(FuelTransaction2.quantity), 0).desc(), FuelStation2.name).all()

    recent_txns_q = (FuelTransaction2.query
                     .options(joinedload(FuelTransaction2.station))
                     .join(FuelStation2)
                     .filter(FuelTransaction2.txn_datetime >= d_from_dt,
                             FuelTransaction2.txn_datetime <= d_to_dt))
    if warehouse_id:
        recent_txns_q = recent_txns_q.filter(FuelStation2.warehouse_id == warehouse_id)
    if station_id:
        recent_txns_q = recent_txns_q.filter(FuelTransaction2.station_id == station_id)
    recent_txns = recent_txns_q.order_by(FuelTransaction2.txn_datetime.desc()).limit(200).all()

    sync_logs = (FuelSyncLog2.query
                 .filter(FuelSyncLog2.synced_at >= d_from_dt,
                         FuelSyncLog2.synced_at <= d_to_dt)
                 .order_by(FuelSyncLog2.synced_at.desc())
                 .limit(100).all())
    sync_summary = {
        'logs': len(sync_logs),
        'received': sum((l.transactions_received or 0) for l in sync_logs),
        'new': sum((l.transactions_new or 0) for l in sync_logs),
        'dup': sum((l.transactions_dup or 0) for l in sync_logs),
        'unknown': sum((l.unknown_stations or 0) for l in sync_logs),
        'last': sync_logs[0] if sync_logs else None,
    }

    warnings = []
    for row in warehouse_rows:
        wh_name = row['warehouse'].name
        if row['opening'] is None:
            warnings.append(_fuel_warning(
                'missing_initial', 'warning',
                f'Склад «{wh_name}» без начального остатка',
                f'«{wh_name}» омборида бошланғич қолдиқ йўқ',
                'Расчётный остаток по этому складу нельзя считать надёжным.',
                'Бу омбор бўйича ҳисобий қолдиқни ишончли ҳисоблаб бўлмайди.',
                wh_name,
                entity_type='fuel_warehouse', entity_id=wh.id, key_value=wh.id,
            ))
        elif row['ending'] is not None and row['ending'] < 0:
            warnings.append(_fuel_warning(
                'negative_balance', 'danger',
                f'Отрицательный расчётный остаток: {wh_name}',
                f'Манфий ҳисобий қолдиқ: {wh_name}',
                f'Расчётный остаток {row["ending"]:.2f} л. Проверьте начальный остаток, приходы и выдачи.',
                f'Ҳисобий қолдиқ {row["ending"]:.2f} л. Бошланғич қолдиқ, кирим ва беришларни текширинг.',
                row['ending'],
                entity_type='fuel_warehouse', entity_id=wh.id, key_value=wh.id,
            ))

    stations_without_warehouse = FuelStation2.query.filter(FuelStation2.warehouse_id.is_(None)).order_by(FuelStation2.name).all()
    for st in stations_without_warehouse:
        warnings.append(_fuel_warning(
            'station_without_warehouse', 'danger',
            f'АЗС «{st.name}» не привязана к складу',
            f'«{st.name}» АЗС омборга боғланмаган',
            'Выдачи по такой АЗС не могут корректно списываться со склада.',
            'Бундай АЗС бўйича берилган ёқилғи омбордан тўғри чиқим қилинмайди.',
            st.topaz_id,
            entity_type='fuel_station', entity_id=st.id, key_value=st.id,
        ))

    for st in station_rows:
        if not st.is_active and (st.tx_count or 0) > 0:
            warnings.append(_fuel_warning(
                'inactive_station_with_tx', 'warning',
                f'Отключённая АЗС имеет выдачи: {st.name}',
                f'Ўчирилган АЗСда беришлар бор: {st.name}',
                f'За период: {float(st.issued or 0):.2f} л, транзакций: {int(st.tx_count or 0)}.',
                f'Давр бўйича: {float(st.issued or 0):.2f} л, транзакциялар: {int(st.tx_count or 0)}.',
                float(st.issued or 0),
                entity_type='fuel_station', entity_id=st.id, key_value=st.id,
            ))

    sync_issue_logs = [l for l in sync_logs if (l.unknown_stations or 0) > 0 or (l.status or '') != 'ok']
    for log in sync_issue_logs[:10]:
        warnings.append(_fuel_warning(
            'sync_issue', 'danger' if (log.unknown_stations or 0) else 'warning',
            f'Проблема синхронизации Topaz: {log.synced_at.strftime("%d.%m.%Y %H:%M")}',
            f'Topaz синхронизациясида муаммо: {log.synced_at.strftime("%d.%m.%Y %H:%M")}',
            f'Статус: {log.status or ""}; неизвестных АЗС: {log.unknown_stations or 0}; ошибка: {log.error_msg or "—"}.',
            f'Ҳолат: {log.status or ""}; номаълум АЗС: {log.unknown_stations or 0}; хато: {log.error_msg or "—"}.',
            log.unknown_stations or 0,
            entity_type='fuel_sync_log', entity_id=log.id, key_value=log.id,
        ))

    latest_sync = FuelSyncLog2.query.order_by(FuelSyncLog2.synced_at.desc()).first()
    if latest_sync and latest_sync.synced_at:
        age_hours = (datetime.utcnow() - latest_sync.synced_at).total_seconds() / 3600
        if age_hours > FUEL_SYNC_STALE_HOURS:
            warnings.append(_fuel_warning(
                'stale_sync', 'danger',
                'Давно не было синхронизации Topaz',
                'Topaz синхронизацияси узоқ вақт бўлмади',
                f'Последняя синхронизация: {latest_sync.synced_at.strftime("%d.%m.%Y %H:%M")}; прошло примерно {age_hours:.1f} ч.',
                f'Охирги синхронизация: {latest_sync.synced_at.strftime("%d.%m.%Y %H:%M")}; тахминан {age_hours:.1f} соат ўтди.',
                round(age_hours, 1),
                entity_type='fuel_sync_log', entity_id=latest_sync.id, key_value=latest_sync.id,
            ))
    else:
        warnings.append(_fuel_warning(
            'no_sync', 'danger',
            'Нет журналов синхронизации Topaz',
            'Topaz синхронизация журналлари йўқ',
            'Система не видит ни одной записи синхронизации.',
            'Тизимда синхронизация бўйича бирорта ёзув йўқ.',
            entity_type='fuel_sync_log', entity_id=None, key_value='no_sync',
        ))

    large_txn_q = (FuelTransaction2.query
                   .options(joinedload(FuelTransaction2.station))
                   .join(FuelStation2)
                   .filter(FuelTransaction2.txn_datetime >= d_from_dt,
                           FuelTransaction2.txn_datetime <= d_to_dt,
                           FuelTransaction2.quantity >= FUEL_LARGE_TXN_THRESHOLD))
    if warehouse_id:
        large_txn_q = large_txn_q.filter(FuelStation2.warehouse_id == warehouse_id)
    if station_id:
        large_txn_q = large_txn_q.filter(FuelTransaction2.station_id == station_id)
    large_txns = large_txn_q.order_by(FuelTransaction2.quantity.desc()).limit(20).all()
    for txn in large_txns:
        station_name = txn.station.name if txn.station else ''
        warnings.append(_fuel_warning(
            'large_transaction', 'warning',
            f'Крупная выдача топлива: {txn.quantity:.2f} л',
            f'Йирик ёқилғи бериш: {txn.quantity:.2f} л',
            f'{txn.txn_datetime.strftime("%d.%m.%Y %H:%M")}; АЗС: {station_name}; карта: {txn.card_number or "—"}; Topaz ID: {txn.topaz_txn_id or "—"}.',
            f'{txn.txn_datetime.strftime("%d.%m.%Y %H:%M")}; АЗС: {station_name}; карта: {txn.card_number or "—"}; Topaz ID: {txn.topaz_txn_id or "—"}.',
            txn.quantity,
            entity_type='fuel_transaction', entity_id=txn.id, key_value=txn.id,
        ))

    bad_qty_q = (FuelTransaction2.query
                 .options(joinedload(FuelTransaction2.station))
                 .join(FuelStation2)
                 .filter(FuelTransaction2.txn_datetime >= d_from_dt,
                         FuelTransaction2.txn_datetime <= d_to_dt)
                 .filter((FuelTransaction2.quantity == None) | (FuelTransaction2.quantity <= 0)))
    if warehouse_id:
        bad_qty_q = bad_qty_q.filter(FuelStation2.warehouse_id == warehouse_id)
    if station_id:
        bad_qty_q = bad_qty_q.filter(FuelTransaction2.station_id == station_id)
    bad_qty_txns = bad_qty_q.order_by(FuelTransaction2.txn_datetime.desc()).limit(20).all()
    for txn in bad_qty_txns:
        warnings.append(_fuel_warning(
            'bad_quantity', 'danger',
            'Транзакция с нулевым или отрицательным количеством',
            'Ноль ёки манфий миқдорли транзакция',
            f'{txn.txn_datetime.strftime("%d.%m.%Y %H:%M")}; АЗС: {txn.station.name if txn.station else ""}; количество: {txn.quantity}.',
            f'{txn.txn_datetime.strftime("%d.%m.%Y %H:%M")}; АЗС: {txn.station.name if txn.station else ""}; миқдор: {txn.quantity}.',
            txn.quantity,
            entity_type='fuel_transaction', entity_id=txn.id, key_value=txn.id,
        ))

    warnings = _apply_fuel_warning_reviews(warnings)
    warnings_summary = _fuel_report_warning_summary(warnings)
    totals = {k: round(v, 2) if isinstance(v, float) else v for k, v in totals.items()}
    totals['warnings'] = warnings_summary['total']
    totals['danger_warnings'] = warnings_summary['danger']
    totals['large_txn_threshold'] = FUEL_LARGE_TXN_THRESHOLD
    return {
        'd_from': d_from,
        'd_to': d_to,
        'd_from_dt': d_from_dt,
        'd_to_dt': d_to_dt,
        'warehouses': warehouses,
        'warehouse_rows': warehouse_rows,
        'station_rows': station_rows,
        'recent_txns': recent_txns,
        'sync_logs': sync_logs,
        'sync_summary': sync_summary,
        # perf-index-fuel-sync-dup-001_marker: expose already-loaded latest sync to dashboard context.
        'latest_sync': latest_sync,
        'warnings': warnings,
        'warnings_summary': warnings_summary,
        'large_txn_threshold': FUEL_LARGE_TXN_THRESHOLD,
        'totals': totals,
        'selected_warehouse_id': warehouse_id,
        'selected_station_id': station_id,
        'selected_station': selected_station,
    }


def _safe_ws_title(title, used):
    bad = '[]:*?/\\'
    cleaned = ''.join('_' if c in bad else c for c in str(title or 'Sheet')).strip()[:31]
    cleaned = cleaned or 'Sheet'
    base = cleaned
    idx = 2
    while cleaned in used:
        suffix = f' {idx}'
        cleaned = (base[:31-len(suffix)] + suffix).strip()
        idx += 1
    used.add(cleaned)
    return cleaned


def _fuel_report_workbook(data, lang='uz'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def L(ru, uz):
        return ru if lang == 'ru' else uz

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    header_fill = PatternFill('solid', fgColor='D9EAD3')
    title_fill = PatternFill('solid', fgColor='1A6B3C')
    title_font = Font(bold=True, color='FFFFFF', size=12)
    header_font = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_table(ws, header_row=1):
        ws.freeze_panes = f'A{header_row + 1}'
        ws.sheet_view.showGridLines = False
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row >= header_row and max_col:
            ws.auto_filter.ref = f'A{header_row}:{get_column_letter(max_col)}{max_row}'
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            width = 10
            for cell in ws[letter]:
                val = '' if cell.value is None else str(cell.value)
                width = max(width, min(len(val) + 2, 38))
            ws.column_dimensions[letter].width = width
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    ws = wb.create_sheet(_safe_ws_title(L('Сводка', 'Сводка'), used))
    ws.append([L('Отчёт по топливу', 'Ёқилғи ҳисоботи')])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.append([L('Период', 'Давр'), data['d_from'].strftime('%d.%m.%Y'), data['d_to'].strftime('%d.%m.%Y'), 'ДТ'])
    ws.append([])
    ws.append([L('Показатель', 'Кўрсаткич'), L('Значение', 'Қиймат')])
    summary_rows = [
        (L('Начальный остаток, л', 'Бошланғич қолдиқ, л'), data['totals']['opening']),
        (L('Приход, л', 'Кирим, л'), data['totals']['receipts']),
        (L('Выдача Topaz, л', 'Topaz бериш, л'), data['totals']['issued']),
        (L('Расчётный остаток, л', 'Ҳисобий қолдиқ, л'), data['totals']['ending']),
        (L('Склады', 'Омборлар'), data['totals']['warehouses']),
        (L('АЗС', 'АЗС'), data['totals']['stations']),
        (L('Транзакции', 'Транзакциялар'), data['totals']['transactions']),
        (L('Складов без начального остатка', 'Бошланғич қолдиқсиз омборлар'), data['totals']['missing_initial']),
        (L('Отрицательных остатков', 'Манфий қолдиқлар'), data['totals']['negative_balances']),
        (L('Неизвестных АЗС в sync logs', 'Sync logларда номаълум АЗС'), data['sync_summary']['unknown']),
        (L('Предупреждений', 'Огоҳлантиришлар'), data.get('warnings_summary', {}).get('total', 0)),
        (L('Критических предупреждений', 'Критик огоҳлантиришлар'), data.get('warnings_summary', {}).get('danger', 0)),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
    style_table(ws, header_row=4)

    ws = wb.create_sheet(_safe_ws_title(L('Предупреждения', 'Огоҳлантиришлар'), used))
    ws.append([L('Уровень', 'Даража'), L('Проблема', 'Муаммо'), L('Описание', 'Изоҳ'), L('Значение', 'Қиймат')])
    for warning in data.get('warnings', []):
        title, details = _fuel_report_warning_text(warning, lang=lang)
        severity_label = {
            'danger': L('Критично', 'Критик'),
            'warning': L('Предупреждение', 'Огоҳлантириш'),
            'info': L('Информация', 'Маълумот'),
        }.get(warning.get('severity'), warning.get('severity') or '')
        ws.append([severity_label, title, details, warning.get('value')])
    if not data.get('warnings'):
        ws.append([L('OK', 'OK'), L('Критических проблем не найдено', 'Критик муаммолар топилмади'), '', ''])
    style_table(ws)

    ws = wb.create_sheet(_safe_ws_title(L('Склады', 'Омборлар'), used))
    ws.append([L('Склад', 'Омбор'), L('АЗС', 'АЗС'), L('Начальный остаток', 'Бошланғич қолдиқ'), L('Приход', 'Кирим'), L('Выдача', 'Бериш'), L('Расчётный остаток', 'Ҳисобий қолдиқ'), L('Транзакций', 'Транзакциялар'), L('Последняя выдача', 'Охирги бериш')])
    for r in data['warehouse_rows']:
        last = r['last_txn'].txn_datetime.strftime('%d.%m.%Y %H:%M') if r['last_txn'] else ''
        ws.append([r['warehouse'].name, r['stations_count'], r['opening'], r['receipts'], r['issued'], r['ending'], r['tx_count'], last])
    style_table(ws)

    ws = wb.create_sheet(_safe_ws_title(L('АЗС', 'АЗС'), used))
    ws.append([L('АЗС', 'АЗС'), 'Topaz ID', L('Склад', 'Омбор'), L('Активна', 'Фаол'), L('Выдано, л', 'Берилди, л'), L('Транзакций', 'Транзакциялар'), L('Последняя выдача', 'Охирги бериш')])
    for r in data['station_rows']:
        last = r.last_txn.strftime('%d.%m.%Y %H:%M') if r.last_txn else ''
        ws.append([r.name, r.topaz_id, r.warehouse_name, L('Да', 'Ҳа') if r.is_active else L('Нет', 'Йўқ'), float(r.issued or 0), int(r.tx_count or 0), last])
    style_table(ws)

    ws = wb.create_sheet(_safe_ws_title(L('Транзакции', 'Транзакциялар'), used))
    ws.append([L('Дата/время', 'Сана/вақт'), L('АЗС', 'АЗС'), L('Склад', 'Омбор'), L('Карта', 'Карта'), L('Топливо', 'Ёқилғи'), L('Литры', 'Литр'), 'Topaz ID'])
    for t in data['recent_txns']:
        ws.append([t.txn_datetime.strftime('%d.%m.%Y %H:%M'), t.station.name, t.station.warehouse_name, t.card_number or '', t.fuel_type or 'ДТ', t.quantity or 0, t.topaz_txn_id or ''])
    style_table(ws)

    ws = wb.create_sheet(_safe_ws_title(L('Синхронизация', 'Синхронизация'), used))
    ws.append([L('Время', 'Вақт'), L('Агент IP', 'Агент IP'), L('Получено', 'Қабул қилинди'), L('Новых', 'Янги'), L('Дублей', 'Такрор'), L('Неизвестных АЗС', 'Номаълум АЗС'), L('Статус', 'Ҳолат'), L('Ошибка', 'Хато')])
    for l in data['sync_logs']:
        ws.append([l.synced_at.strftime('%d.%m.%Y %H:%M:%S'), l.agent_ip or '', l.transactions_received or 0, l.transactions_new or 0, l.transactions_dup or 0, l.unknown_stations or 0, l.status or '', l.error_msg or ''])
    style_table(ws)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    return wb


@fuel_bp.route('/warnings')
@module_required('fuel')
def fuel_warnings():
    today = date.today()
    default_from = today.replace(day=1)
    d_from = _parse_report_date(request.args.get('date_from'), default_from)
    d_to = _parse_report_date(request.args.get('date_to'), today)
    if d_from > d_to:
        d_from, d_to = d_to, d_from

    status_filter = (request.args.get('status') or '').strip()
    severity_filter = (request.args.get('severity') or '').strip()
    code_filter = (request.args.get('code') or '').strip()
    q = (request.args.get('q') or '').strip().lower()

    data = _collect_fuel_report_data(d_from, d_to)
    warnings = list(data.get('warnings') or [])
    current_keys = {w.get('key') for w in warnings if w.get('key')}

    saved_reviews = (FuelWarningReview.query
                     .order_by(FuelWarningReview.last_seen_at.desc(), FuelWarningReview.updated_at.desc())
                     .limit(300).all())
    for review in saved_reviews:
        if review.warning_key not in current_keys:
            warnings.append(_fuel_warning_from_review(review))

    status_options = ['new', 'in_progress', 'resolved', 'rejected']
    severity_options = ['danger', 'warning', 'info']
    code_options = sorted({w.get('code') for w in warnings if w.get('code')})

    def matches(w):
        if status_filter and (w.get('review_status') or 'new') != status_filter:
            return False
        if severity_filter and (w.get('severity') or '') != severity_filter:
            return False
        if code_filter and (w.get('code') or '') != code_filter:
            return False
        if q:
            hay = ' '.join(str(x or '') for x in [
                w.get('code'), w.get('severity'), w.get('title_ru'), w.get('title_uz'),
                w.get('details_ru'), w.get('details_uz'), w.get('value'),
                w.get('review_comment'), w.get('review_status'),
            ]).lower()
            if q not in hay:
                return False
        return True

    filtered_warnings = [w for w in warnings if matches(w)]
    counts = {
        'total': len(warnings),
        'filtered': len(filtered_warnings),
        'current': sum(1 for w in warnings if w.get('is_current')),
        'danger': sum(1 for w in warnings if w.get('severity') == 'danger'),
        'in_progress': sum(1 for w in warnings if w.get('review_status') == 'in_progress'),
        'resolved': sum(1 for w in warnings if w.get('review_status') == 'resolved'),
    }
    return render_template(
        'fuel/warnings.html',
        warnings=filtered_warnings,
        counts=counts,
        d_from=d_from,
        d_to=d_to,
        status_filter=status_filter,
        severity_filter=severity_filter,
        code_filter=code_filter,
        q=request.args.get('q') or '',
        status_options=status_options,
        severity_options=severity_options,
        code_options=code_options,
        status_label=_fuel_warning_status_label,
        severity_label=_fuel_warning_severity_label,
        warning_text=_fuel_report_warning_text,
    )


@fuel_bp.route('/warnings/<warning_key>/update', methods=['POST'])
@module_required('fuel')
def fuel_warning_update(warning_key):
    if not current_user.can_edit:
        abort(403)
    allowed_statuses = {'new', 'in_progress', 'resolved', 'rejected'}
    status = (request.form.get('status') or 'new').strip()
    if status not in allowed_statuses:
        status = 'new'
    comment = (request.form.get('comment') or '').strip()[:4000]

    review = FuelWarningReview.query.filter_by(warning_key=warning_key).first()
    before = None
    created = False
    if review:
        before = {
            'status': review.status,
            'comment': review.comment,
            'updated_by': review.updated_by,
            'resolved_at': review.resolved_at,
        }
    else:
        review = FuelWarningReview(warning_key=warning_key)
        db.session.add(review)
        created = True

    review.warning_code = (request.form.get('warning_code') or review.warning_code or '').strip()[:80]
    review.severity = (request.form.get('severity') or review.severity or 'warning').strip()[:20]
    review.entity_type = (request.form.get('entity_type') or review.entity_type or '').strip()[:80]
    try:
        review.entity_id = int(request.form.get('entity_id') or review.entity_id or 0) or None
    except (TypeError, ValueError):
        review.entity_id = review.entity_id or None
    review.title_snapshot = (request.form.get('title_snapshot') or review.title_snapshot or '').strip()[:500]
    review.details_snapshot = (request.form.get('details_snapshot') or review.details_snapshot or '').strip()
    review.value_snapshot = (request.form.get('value_snapshot') or review.value_snapshot or '').strip()[:200]
    review.status = status
    review.comment = comment
    review.last_seen_at = datetime.utcnow()
    review.updated_by = current_user.id
    review.updated_at = datetime.utcnow()
    review.resolved_at = datetime.utcnow() if status in ('resolved', 'rejected') else None

    db.session.flush()
    after = {
        'status': review.status,
        'comment': review.comment,
        'updated_by': review.updated_by,
        'resolved_at': review.resolved_at,
    }
    _audit_fuel(
        'fuel_warning_review_created' if created else 'fuel_warning_review_updated',
        entity_type='fuel_warning_review',
        entity_id=review.id,
        entity_label=review.warning_code or warning_key,
        before=before,
        after=after,
        status='ok',
        description='Fuel warning review status updated',
    )
    db.session.commit()
    flash(fuel_t('Огоҳлантириш ҳолати сақланди', 'Статус предупреждения сохранён'), 'success')
    return redirect(request.form.get('return_url') or url_for('fuel.fuel_warnings'))


@fuel_bp.route('/report')
@module_required('fuel')
def fuel_report():
    today = date.today()
    default_from = today.replace(day=1)
    d_from = _parse_report_date(request.args.get('date_from'), default_from)
    d_to = _parse_report_date(request.args.get('date_to'), today)
    if d_from > d_to:
        d_from, d_to = d_to, d_from

    warehouse_id = request.args.get('warehouse_id', type=int)
    station_id = request.args.get('station_id', type=int)
    data = _collect_fuel_report_data(d_from, d_to, warehouse_id=warehouse_id, station_id=station_id)

    if request.args.get('export') == '1':
        wb = _fuel_report_workbook(data, lang=_fuel_report_lang())
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        prefix = 'Fuel_report' if _fuel_report_lang() == 'ru' else 'Yoqilgi_hisoboti'
        fname = f"{prefix}_{d_from.strftime('%d_%m_%Y')}_{d_to.strftime('%d_%m_%Y')}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    # perf-fuel-report-warehouse-query-001_marker: reuse warehouses loaded by report collector.
    data_for_template = dict(data)
    warehouses = data_for_template.pop('warehouses', None) or FuelWarehouse.query.order_by(FuelWarehouse.name).all()
    stations = (FuelStation2.query
                .join(FuelWarehouse)
                .order_by(FuelWarehouse.name, FuelStation2.name).all())
    return render_template('fuel/report.html',
                           warehouses=warehouses,
                           stations=stations,
                           fuel_types=FUEL_TYPES,
                           **data_for_template)


# ─── Dashboard ────────────────────────────────────────────────────────

@fuel_bp.route('/')
@module_required('fuel')
def dashboard():
    balance_rows = get_all_balances()

    # Последняя синхронизация
    last_sync = (FuelSyncLog2.query
                 .filter_by(status='ok')
                 .order_by(FuelSyncLog2.synced_at.desc())
                 .first())

    # Последние 30 транзакций
    recent_txns = (FuelTransaction2.query
                   .options(joinedload(FuelTransaction2.station))
                   .join(FuelStation2)
                   .order_by(FuelTransaction2.txn_datetime.desc())
                   .limit(30).all())

    # Статистика за сегодня
    today_total = (db.session.query(func.coalesce(func.sum(FuelTransaction2.quantity), 0))
                   .filter(FuelTransaction2.txn_datetime >= datetime.combine(date.today(), datetime.min.time()),
                           FuelTransaction2.txn_datetime < datetime.combine(date.today() + timedelta(days=1), datetime.min.time()))
                   .scalar())

    return render_template('fuel/dashboard.html',
                           balance_rows=balance_rows,
                           last_sync=last_sync,
                           recent_txns=recent_txns,
                           today_total=today_total,
                           fuel_types=FUEL_TYPES)


# ─── Warehouses ───────────────────────────────────────────────────────

@fuel_bp.route('/warehouses')
@module_required('fuel')
@admin_required_fuel
def warehouses():
    whs = (fuel_warehouse_query_for_ui()
           .outerjoin(Organization)
           .order_by(FuelWarehouse.name).all())
    orgs = Organization.query.order_by(Organization.sort_order).all()
    edit_id = request.args.get('edit_id', type=int)
    edit_warehouse = None
    if edit_id:
        edit_warehouse = FuelWarehouse.query.get(edit_id)
    # fuel-batch-perf-001b_marker: bulk warehouse/station counts.
    warehouse_ids = [wh.id for wh in whs]
    station_rows = []
    if warehouse_ids:
        station_rows = (FuelStation2.query
            .filter(FuelStation2.warehouse_id.in_(warehouse_ids))
            .order_by(FuelStation2.warehouse_id, FuelStation2.name)
            .all())

    stations_by_warehouse = {wid: [] for wid in warehouse_ids}
    station_ids = []
    for st in station_rows:
        stations_by_warehouse.setdefault(st.warehouse_id, []).append(st)
        station_ids.append(st.id)

    warehouse_station_counts = {wid: len(stations_by_warehouse.get(wid, [])) for wid in warehouse_ids}
    warehouse_receipt_counts = {}
    warehouse_initial_balance_counts = {}
    station_tx_counts = {}

    if warehouse_ids:
        warehouse_receipt_counts = dict(
            db.session.query(FuelReceipt2.warehouse_id, func.count(FuelReceipt2.id))
            .filter(FuelReceipt2.warehouse_id.in_(warehouse_ids))
            .group_by(FuelReceipt2.warehouse_id)
            .all()
        )
        warehouse_initial_balance_counts = dict(
            db.session.query(FuelInitialBalance.warehouse_id, func.count(FuelInitialBalance.id))
            .filter(FuelInitialBalance.warehouse_id.in_(warehouse_ids))
            .group_by(FuelInitialBalance.warehouse_id)
            .all()
        )

    if station_ids:
        station_tx_counts = dict(
            db.session.query(FuelTransaction2.station_id, func.count(FuelTransaction2.id))
            .filter(FuelTransaction2.station_id.in_(station_ids))
            .group_by(FuelTransaction2.station_id)
            .all()
        )

    fuel_batch_total_stations_count = sum(warehouse_station_counts.values())

    warehouse_delete_info = {}
    for wh in whs:
        linked = {
            'stations_count': int(warehouse_station_counts.get(wh.id, 0) or 0),
            'receipts_count': int(warehouse_receipt_counts.get(wh.id, 0) or 0),
            'initial_balances_count': int(warehouse_initial_balance_counts.get(wh.id, 0) or 0),
        }
        warehouse_delete_info[wh.id] = {
            'can_delete': not any(linked.values()),
            'linked_total': sum(linked.values()),
            'linked': linked,
        }

    return render_template('fuel/warehouses.html', warehouses=whs,
                           organizations=orgs, fuel_types=FUEL_TYPES,
                           edit_warehouse=edit_warehouse,
                           warehouse_delete_info=warehouse_delete_info,
                           fuel_batch_stations_by_warehouse=stations_by_warehouse,
                           fuel_batch_station_tx_counts=station_tx_counts,
                           fuel_batch_total_stations_count=fuel_batch_total_stations_count)


@fuel_bp.route('/warehouses/save', methods=['POST'])
@module_required('fuel')
@admin_required_fuel
def save_warehouse():
    wid = request.form.get('id', type=int)
    name = request.form.get('name', '').strip()
    org_id = request.form.get('organization_id', type=int)
    notes = request.form.get('notes', '').strip()

    if not name:
        flash(fuel_t('Омбор номини киритинг', 'Введите название склада'), 'warning')
        return redirect(url_for('fuel.warehouses'))

    created = False
    before = None
    if wid:
        wh = FuelWarehouse.query.get_or_404(wid)
        before = _fuel_warehouse_snapshot(wh)
        wh.name = name
        wh.organization_id = org_id or None
        wh.notes = notes
    else:
        wh = FuelWarehouse(name=name, organization_id=org_id or None, notes=notes)
        db.session.add(wh)
        created = True

    db.session.flush()
    after = _fuel_warehouse_snapshot(wh)
    _audit_fuel(
        'fuel_warehouse_created' if created else 'fuel_warehouse_updated',
        entity_type='fuel_warehouse',
        entity_id=wh.id,
        entity_label=wh.name,
        before=before,
        after=after,
        description='Fuel warehouse saved',
    )
    db.session.commit()
    flash(fuel_t('Омбор сақланди', 'Склад сохранён'), 'success')
    return redirect(url_for('fuel.warehouses'))


@fuel_bp.route('/warehouses/delete/<int:wid>', methods=['POST'])
@admin_required_fuel
def delete_warehouse(wid):
    wh = FuelWarehouse.query.get_or_404(wid)

    station_count = FuelStation2.query.filter_by(warehouse_id=wh.id).count()
    receipt_count = FuelReceipt2.query.filter_by(warehouse_id=wh.id).count()
    balance_count = FuelInitialBalance.query.filter_by(warehouse_id=wh.id).count()
    tx_count = (FuelTransaction2.query
                .join(FuelStation2)
                .filter(FuelStation2.warehouse_id == wh.id)
                .count())

    if station_count or receipt_count or balance_count or tx_count:
        flash('Склад не удалён: есть АЗС, остатки, приходы или исторические выдачи. Удаление заблокировано для сохранения учёта.', 'warning')
        return redirect(url_for('fuel.warehouses'))

    db.session.delete(wh)
    db.session.commit()
    flash('Склад удалён', 'warning')
    return redirect(url_for('fuel.warehouses'))


# ─── Initial Balance ──────────────────────────────────────────────────

@fuel_bp.route('/initial-balance', methods=['GET', 'POST'])
@module_required('fuel')
@admin_required_fuel
def initial_balance():
    warehouses = fuel_warehouse_query_for_ui().order_by(FuelWarehouse.name).all()
    # fuel-batch-perf-001b_marker: bulk initial balance rows.
    warehouse_ids = [wh.id for wh in warehouses]
    initial_balance_rows = []
    if warehouse_ids:
        initial_balance_rows = (FuelInitialBalance.query
            .filter(FuelInitialBalance.warehouse_id.in_(warehouse_ids))
            .order_by(FuelInitialBalance.warehouse_id, FuelInitialBalance.fuel_type)
            .all())

    existing = {wh.id: {} for wh in warehouses}
    for ib in initial_balance_rows:
        existing.setdefault(ib.warehouse_id, {})[ib.fuel_type] = ib
    return render_template('fuel/initial_balance.html',
                           warehouses=warehouses, existing=existing,
                           fuel_types=FUEL_TYPES, today=date.today())


@fuel_bp.route('/initial-balance/save', methods=['POST'])
@admin_required_fuel
def save_initial_balance():
    ib_id = request.form.get('id', type=int)
    warehouse_id = request.form.get('warehouse_id', type=int)
    fuel_type = request.form.get('fuel_type', 'ДТ').strip()
    quantity = request.form.get('quantity', type=float)
    balance_date_s = request.form.get('balance_date', '')
    note = request.form.get('note', '').strip()

    if not warehouse_id or quantity is None or quantity < 0:
        flash('Заполните все поля. Остаток не может быть отрицательным.', 'warning')
        return redirect(url_for('fuel.initial_balance'))

    try:
        balance_date = datetime.strptime(balance_date_s, '%Y-%m-%d').date()
    except ValueError:
        balance_date = date.today()

    existing = None
    if ib_id:
        existing = FuelInitialBalance.query.get_or_404(ib_id)

    duplicate = (FuelInitialBalance.query
                 .filter_by(warehouse_id=warehouse_id, fuel_type=fuel_type)
                 .first())

    if duplicate and (not existing or duplicate.id != existing.id):
        flash('Для этого склада и вида топлива уже есть начальный остаток. Используйте кнопку "Изм." у существующей строки.', 'warning')
        return redirect(url_for('fuel.initial_balance'))

    if existing:
        existing.warehouse_id = warehouse_id
        existing.fuel_type = fuel_type
        existing.quantity = quantity
        existing.balance_date = balance_date
        existing.note = note
        flash('Начальный остаток изменён. Баланс пересчитается автоматически.', 'success')
    else:
        ib = FuelInitialBalance(
            warehouse_id=warehouse_id,
            fuel_type=fuel_type,
            quantity=quantity,
            balance_date=balance_date,
            note=note,
            created_by=current_user.id,
        )
        db.session.add(ib)
        flash('Начальный остаток сохранён. Баланс пересчитается автоматически.', 'success')

    db.session.commit()
    return redirect(url_for('fuel.initial_balance'))


@fuel_bp.route('/initial-balance/delete/<int:ib_id>', methods=['POST'])
@admin_required_fuel
def delete_initial_balance(ib_id):
    ib = FuelInitialBalance.query.get_or_404(ib_id)
    db.session.delete(ib)
    db.session.commit()
    flash('Начальный остаток удалён. Баланс пересчитается автоматически.', 'warning')
    return redirect(url_for('fuel.initial_balance'))


# ─── Receipts (Приходы) ───────────────────────────────────────────────

@fuel_bp.route('/receipts')
@module_required('fuel')
def receipts():
    d_from_s = request.args.get('date_from', '')
    d_to_s   = request.args.get('date_to', '')
    wh_id    = request.args.get('warehouse_id', type=int)

    today = date.today()
    try:
        d_from = datetime.strptime(d_from_s, '%Y-%m-%d').date()
    except ValueError:
        d_from = today.replace(day=1)
    try:
        d_to = datetime.strptime(d_to_s, '%Y-%m-%d').date()
    except ValueError:
        d_to = today

    q = FuelReceipt2.query.filter(
        FuelReceipt2.receipt_date >= d_from,
        FuelReceipt2.receipt_date <= d_to,
    )
    if wh_id:
        q = q.filter(FuelReceipt2.warehouse_id == wh_id)
    else:
        q = fuel_apply_warehouse_filter_for_ui(q, FuelReceipt2.warehouse_id)
    items = q.order_by(FuelReceipt2.receipt_date.desc(), FuelReceipt2.id.desc()).all()

    warehouses = fuel_warehouse_query_for_ui().order_by(FuelWarehouse.name).all()
    total_qty = sum(r.quantity for r in items)
    return render_template('fuel/receipts.html',
                           items=items, warehouses=warehouses,
                           d_from=d_from, d_to=d_to,
                           selected_wh_id=wh_id,
                           fuel_types=FUEL_TYPES, total_qty=total_qty,
                           today=today)


@fuel_bp.route('/receipts/save', methods=['POST'])
@module_required('fuel')
def save_receipt():
    if not current_user.can_edit:
        abort(403)
    rid = request.form.get('id', type=int)
    warehouse_id = request.form.get('warehouse_id', type=int)
    fuel_type = 'ДТ'
    supplier = request.form.get('supplier', '').strip()
    doc_number = request.form.get('doc_number', '').strip()
    note = request.form.get('note', '').strip()
    date_s = request.form.get('receipt_date', '')

    errors = []
    if not warehouse_id or not FuelWarehouse.query.get(warehouse_id):
        errors.append(fuel_t('Омборни танланг', 'Выберите склад'))
    if fuel_type not in FUEL_TYPES:
        errors.append(fuel_t('Ёқилғи тури нотўғри', 'Некорректный тип топлива'))
    try:
        quantity = _parse_positive(request.form.get('quantity'), 'quantity')
    except ValueError:
        quantity = None
        errors.append(fuel_t('Миқдор мусбат бўлиши керак', 'Количество должно быть больше нуля'))
    price = 0
    try:
        receipt_date = _parse_fuel_date(date_s, 'receipt_date')
    except ValueError:
        receipt_date = None
        errors.append(fuel_t('Сана тўғри бўлиши керак', 'Дата должна быть корректной'))
    if errors:
        fuel_flash_errors(errors)
        return redirect(url_for('fuel.receipts'))

    created = False
    before = None
    if rid:
        r = FuelReceipt2.query.get_or_404(rid)
        before = _fuel_receipt_snapshot(r)
        r.warehouse_id = warehouse_id
        r.fuel_type = fuel_type
        r.quantity = quantity
        r.price_per_liter = price
        r.supplier = supplier
        r.doc_number = doc_number
        r.note = note
        r.receipt_date = receipt_date
    else:
        r = FuelReceipt2(
            warehouse_id=warehouse_id, fuel_type=fuel_type, quantity=quantity,
            price_per_liter=price, supplier=supplier, doc_number=doc_number,
            note=note, receipt_date=receipt_date, created_by=current_user.id,
        )
        db.session.add(r)
        created = True

    db.session.flush()
    after = _fuel_receipt_snapshot(r)
    _audit_fuel(
        'fuel_receipt_created' if created else 'fuel_receipt_updated',
        entity_type='fuel_receipt',
        entity_id=r.id,
        entity_label=r.doc_number or f'{r.fuel_type} {r.quantity}',
        before=before,
        after=after,
        description='Fuel receipt saved',
    )
    db.session.commit()
    flash(fuel_t('Кирим сақланди', 'Приход сохранён'), 'success')
    return redirect(url_for('fuel.receipts'))


@fuel_bp.route('/receipts/delete/<int:rid>', methods=['POST'])
@module_required('fuel')
def delete_receipt(rid):
    if not current_user.can_edit:
        abort(403)
    r = FuelReceipt2.query.get_or_404(rid)
    before = _fuel_receipt_snapshot(r)
    label = r.doc_number or f'{r.fuel_type} {r.quantity}'
    _audit_fuel(
        'fuel_receipt_deleted',
        entity_type='fuel_receipt',
        entity_id=r.id,
        entity_label=label,
        before=before,
        description='Fuel receipt deleted',
    )
    db.session.delete(r)
    db.session.commit()
    flash(fuel_t('Кирим ўчирилди', 'Приход удалён'), 'warning')
    return redirect(url_for('fuel.receipts'))


# ─── Transactions (Расходы из Топаз) ──────────────────────────────────

@fuel_bp.route('/transactions')
@module_required('fuel')
def transactions():
    d_from_s = request.args.get('date_from', '')
    d_to_s   = request.args.get('date_to', '')
    wh_id    = request.args.get('warehouse_id', type=int)

    today = date.today()
    try:
        d_from = datetime.strptime(d_from_s, '%Y-%m-%d').date()
    except ValueError:
        d_from = today

    try:
        d_to = datetime.strptime(d_to_s, '%Y-%m-%d').date()
    except ValueError:
        d_to = today

    d_from_dt = datetime.combine(d_from, datetime.min.time())
    d_to_dt   = datetime.combine(d_to, datetime.max.time())

    q = (FuelTransaction2.query
         # perf-fuel-transactions-nplus1-001_marker: eager load station and warehouse for transaction list.
         .options(joinedload(FuelTransaction2.station).joinedload(FuelStation2.warehouse))
         .join(FuelStation2)
         .filter(FuelTransaction2.txn_datetime >= d_from_dt,
                 FuelTransaction2.txn_datetime <= d_to_dt))
    if wh_id:
        q = q.filter(FuelStation2.warehouse_id == wh_id)

    items = q.order_by(FuelTransaction2.txn_datetime.desc()).limit(500).all()

    warehouses = fuel_warehouse_query_for_ui().order_by(FuelWarehouse.name).all()
    total_qty = sum(t.quantity for t in items)
    total_amount = sum(t.amount for t in items)

    sync_logs = (FuelSyncLog2.query
                 .order_by(FuelSyncLog2.synced_at.desc()).limit(10).all())

    return render_template('fuel/transactions.html',
                           items=items, warehouses=warehouses,
                           d_from=d_from, d_to=d_to,
                           selected_wh_id=wh_id,
                           total_qty=total_qty, total_amount=total_amount,
                           sync_logs=sync_logs, today=today)


# ─── Stations (АЗС — справочник) ─────────────────────────────────────

@fuel_bp.route('/stations')
@module_required('fuel')
@admin_required_fuel
def stations():
    all_stations = (FuelStation2.query
                    .join(FuelWarehouse)
                    .order_by(FuelWarehouse.name, FuelStation2.name).all())
    warehouses = fuel_warehouse_query_for_ui().order_by(FuelWarehouse.name).all()
    # PERF-FUEL-STATIONS-NPLUS1-001B_MARKER: bulk transaction counts for fuel stations.
    station_ids = [st.id for st in all_stations]
    station_tx_counts = {}
    if station_ids:
        station_tx_counts = dict(
            db.session.query(FuelTransaction2.station_id, func.count(FuelTransaction2.id))
            .filter(FuelTransaction2.station_id.in_(station_ids))
            .group_by(FuelTransaction2.station_id)
            .all()
        )

    station_delete_info = {}
    for st in all_stations:
        tx_count = int(station_tx_counts.get(st.id, 0) or 0)
        station_delete_info[st.id] = {
            'can_delete': tx_count == 0,
            'can_deactivate': tx_count > 0 and bool(st.is_active),
            'is_disabled': tx_count > 0 and not bool(st.is_active),
            'transactions_count': tx_count,
        }
    return render_template('fuel/stations.html',
                           stations=all_stations, warehouses=warehouses,
                           station_delete_info=station_delete_info)


@fuel_bp.route('/stations/save', methods=['POST'])
@admin_required_fuel
def save_station():
    sid = request.form.get('id', type=int)
    name = request.form.get('name', '').strip()
    topaz_id = request.form.get('topaz_id', type=int)
    warehouse_id = request.form.get('warehouse_id', type=int)
    is_active = request.form.get('is_active') == 'on'
    valid_from = parse_fuel_station_date(request.form.get('valid_from', ''))
    valid_to = parse_fuel_station_date(request.form.get('valid_to', ''))
    replacement_of_id = request.form.get('replacement_of_id', type=int)
    notes = request.form.get('notes', '').strip()

    if not name or not topaz_id or not warehouse_id:
        flash('Заполните все обязательные поля', 'warning')
        return redirect(url_for('fuel.stations'))

    if valid_from and valid_to and valid_from > valid_to:
        flash('Дата начала действия АЗС не может быть позже даты окончания', 'warning')
        return redirect(url_for('fuel.stations'))

    if sid and replacement_of_id == sid:
        flash('АЗС не может заменять саму себя', 'warning')
        return redirect(url_for('fuel.stations'))

    if sid:
        st = FuelStation2.query.get_or_404(sid)
        duplicate = (FuelStation2.query
                     .filter(FuelStation2.topaz_id == topaz_id,
                             FuelStation2.id != sid)
                     .first())
        if duplicate:
            flash(f'АЗС с Topaz ID {topaz_id} уже существует', 'warning')
            return redirect(url_for('fuel.stations'))

        st.name = name
        st.topaz_id = topaz_id
        st.warehouse_id = warehouse_id
        st.is_active = is_active
        st.valid_from = valid_from
        st.valid_to = valid_to
        st.replacement_of_id = replacement_of_id or None
        st.notes = notes
    else:
        existing = FuelStation2.query.filter_by(topaz_id=topaz_id).first()
        if existing:
            flash(f'АЗС с Topaz ID {topaz_id} уже существует', 'warning')
            return redirect(url_for('fuel.stations'))
        st = FuelStation2(name=name, topaz_id=topaz_id,
                          warehouse_id=warehouse_id, is_active=is_active,
                          valid_from=valid_from, valid_to=valid_to,
                          replacement_of_id=replacement_of_id or None,
                          notes=notes)
        db.session.add(st)

    db.session.commit()
    flash('АЗС сохранена', 'success')
    return redirect(url_for('fuel.stations'))





# Target fuel hierarchy Topaz IDs from Иерархия АЗС.xlsx.
# Legacy warehouses remain in DB for history, but ordinary UI hides them by default.
FUEL_TARGET_TOPAZ_IDS = {
    812011, 935501, 935541,
    935531, 898131, 898141, 935471, 935511,
    825261, 898151, 898111,
    935491, 812301, 934451, 935521,
    812021, 895101, 812001,
    811971, 825241,
    898121, 935551, 825271, 946231,
}


def fuel_show_legacy_warehouses():
    """Return True only when user explicitly asks to show legacy warehouses."""
    return request.args.get('show_legacy') == '1'


def fuel_target_warehouse_ids():
    """Warehouse IDs that contain target fuel stations/topaz IDs."""
    rows = (db.session.query(FuelStation2.warehouse_id)
            .filter(FuelStation2.topaz_id.in_(FUEL_TARGET_TOPAZ_IDS))
            .filter(FuelStation2.warehouse_id.isnot(None))
            .distinct()
            .all())
    return [r[0] for r in rows if r[0] is not None]


def fuel_warehouse_query_for_ui():
    """
    Default UI query for fuel warehouses.

    Normal mode: show only target/current warehouses.
    Historical mode: add ?show_legacy=1 to show all warehouses.
    """
    q = FuelWarehouse.query
    if fuel_show_legacy_warehouses():
        return q

    target_ids = fuel_target_warehouse_ids()
    if not target_ids:
        return q.filter(FuelWarehouse.id == -1)

    return q.filter(FuelWarehouse.id.in_(target_ids))


def fuel_apply_warehouse_filter_for_ui(query, warehouse_column):
    """Apply default warehouse filter to receipt/transaction style queries."""
    if fuel_show_legacy_warehouses():
        return query

    target_ids = fuel_target_warehouse_ids()
    if not target_ids:
        return query.filter(warehouse_column == -1)

    return query.filter(warehouse_column.in_(target_ids))

@fuel_bp.route('/stations/enable/<int:sid>', methods=['POST'])
@admin_required_fuel
def enable_station(sid):
    st = FuelStation2.query.get_or_404(sid)

    if st.is_active:
        flash('АЗС уже активна', 'info')
        return redirect(url_for('fuel.stations'))

    if st.valid_to:
        flash('АЗС не включена автоматически: задана дата окончания действия. Для исторической или заменённой АЗС измените даты вручную.', 'warning')
        return redirect(url_for('fuel.stations'))

    st.is_active = True
    db.session.commit()
    flash('АЗС включена', 'success')
    return redirect(url_for('fuel.stations'))


@fuel_bp.route('/stations/delete/<int:sid>', methods=['POST'])
@admin_required_fuel
def delete_station(sid):
    st = FuelStation2.query.get_or_404(sid)
    tx_count = FuelTransaction2.query.filter_by(station_id=st.id).count()

    if tx_count:
        st.is_active = False
        if not st.valid_to:
            st.valid_to = datetime.utcnow().date()
        db.session.commit()
        flash('АЗС имеет исторические выдачи и не удалена. Она переведена в неактивный статус.', 'warning')
        return redirect(url_for('fuel.stations'))

    db.session.delete(st)
    db.session.commit()
    flash('АЗС удалена', 'warning')
    return redirect(url_for('fuel.stations'))


# ─── API: Ping ────────────────────────────────────────────────────────

@fuel_bp.route('/api/fuel_ping')
def api_fuel_ping():
    return jsonify(status='ok', server_time=datetime.utcnow().isoformat())


# ─── API: Fuel Sync (принимаем данные от агента Топаз) ────────────────

def _perform_fuel_sync():
    # [REASON]: Shared sync logic called by both the canonical /fuel/api/fuel_sync
    # and the legacy /api/fuel_sync compatibility alias. Keeping logic in one place
    # prevents drift between the two paths.
    try:
        payload = request.get_json(force=True)
    except Exception:
        return jsonify(error='invalid JSON'), 400

    api_token = current_app.config.get('FUEL_API_TOKEN')
    submitted_token = str(payload.get('token') or '') if payload else ''
    if not api_token or not payload or not hmac.compare_digest(submitted_token, str(api_token)):
        return jsonify(error='unauthorized'), 401

    transactions = payload.get('transactions', [])
    agent_ip = request.remote_addr

    new_count = 0
    dup_count = 0
    unknown_count = 0
    excluded_count = 0
    possible_dup_count = 0
    errors = []
    # Station is resolved per transaction by Topaz ID and transaction datetime.

    for txn in transactions:
        try:
            # [FUEL-SYNC-013]: Technical/service cards (ПЕРЕЛИВ) are Topaz
            # counter-reconciliation records, not real fuel dispensing. They
            # must never enter fuel_transactions2 for any station, so this
            # check runs before station resolution and dedup/insert logic.
            card_number = str(txn.get('card_number', '') or '')
            if card_number in EXCLUDED_CARD_NUMBERS:
                excluded_count += 1
                continue

            col_id = int(txn.get('topaz_col_id') or 0)
            txn_dt = parse_topaz_txn_datetime(txn.get('txn_datetime', ''))
            station = resolve_fuel_station_for_topaz(col_id, txn_dt)

            if not station:
                unknown_count += 1
                continue

            topaz_txn_id = str(txn.get('topaz_txn_id', ''))
            txn_dt_s = txn.get('txn_datetime', '')
            try:
                txn_dt = datetime.fromisoformat(txn_dt_s)
            except Exception:
                txn_dt = datetime.utcnow()

            # Дедупликация
            if topaz_txn_id:
                existing = (FuelTransaction2.query
                            .filter_by(station_id=station.id,
                                       topaz_txn_id=topaz_txn_id)
                            .first())
                if existing:
                    dup_count += 1
                    continue

            quantity = float(txn.get('quantity', 0) or 0)

            # [FUEL-SYNC-013]: Soft duplicate-content check. Same station/card/
            # quantity/datetime under a *different* topaz_txn_id usually means
            # the terminal re-issued a transaction ID after an offline resync.
            # A genuine coincidence is possible, so the row is still inserted —
            # we only warn (greppable tag) and surface a response counter.
            content_dup = (FuelTransaction2.query
                           .filter(FuelTransaction2.station_id == station.id,
                                   FuelTransaction2.card_number == card_number,
                                   FuelTransaction2.quantity == quantity,
                                   FuelTransaction2.txn_datetime == txn_dt,
                                   FuelTransaction2.topaz_txn_id != topaz_txn_id)
                           .first())
            if content_dup:
                current_app.logger.warning(
                    "FUEL_SYNC_POSSIBLE_DUP station_id=%s card_number=%s "
                    "txn_datetime=%s quantity=%s existing_txn_id=%s new_txn_id=%s",
                    station.id, card_number, txn_dt, quantity,
                    content_dup.topaz_txn_id, topaz_txn_id,
                )
                possible_dup_count += 1

            t = FuelTransaction2(
                station_id=station.id,
                topaz_txn_id=topaz_txn_id,
                topaz_col_id=col_id,
                txn_datetime=txn_dt,
                card_number=card_number,
                fuel_type='ДТ',
                quantity=quantity,
                price_per_liter=0,
                amount=float(txn.get('amount', 0) or 0),
            )
            db.session.add(t)
            new_count += 1

        except Exception as e:
            errors.append(str(e))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        summary = {
            'agent_ip': agent_ip,
            'transactions_received': len(transactions),
            'transactions_new': new_count,
            'transactions_dup': dup_count,
            'unknown_stations': unknown_count,
            'transactions_excluded': excluded_count,
            'possible_duplicates': possible_dup_count,
            'errors_count': len(errors),
            'db_error': str(e),
        }
        try:
            _audit_fuel(
                'fuel_topaz_sync_failed',
                entity_type='fuel_sync',
                entity_label='topaz_agent',
                after=summary,
                status='error',
                description='Topaz fuel sync failed during transaction commit',
                actor_user=TOPAZ_AUDIT_ACTOR,
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify(error=str(e)), 500

    # Лог
    log = FuelSyncLog2(
        agent_ip=agent_ip,
        transactions_received=len(transactions),
        transactions_new=new_count,
        transactions_dup=dup_count,
        unknown_stations=unknown_count,
        status='ok' if not errors else 'partial',
        error_msg='; '.join(errors[:5]),
    )
    db.session.add(log)
    db.session.flush()
    summary = {
        'sync_log_id': log.id,
        'agent_ip': agent_ip,
        'transactions_received': len(transactions),
        'transactions_new': new_count,
        'transactions_dup': dup_count,
        'unknown_stations': unknown_count,
        'transactions_excluded': excluded_count,
        'possible_duplicates': possible_dup_count,
        'errors_count': len(errors),
        'status': log.status,
    }
    _audit_fuel(
        'fuel_topaz_sync_completed',
        entity_type='fuel_sync',
        entity_id=log.id,
        entity_label='topaz_agent',
        after=summary,
        status='ok' if not errors else 'partial',
        description='Topaz fuel sync completed',
        actor_user=TOPAZ_AUDIT_ACTOR,
    )
    db.session.commit()

    return jsonify(
        status='ok',
        received=len(transactions),
        new=new_count,
        duplicates=dup_count,
        unknown=unknown_count,
        excluded=excluded_count,
        possible_duplicates=possible_dup_count,
    )


@fuel_bp.route('/api/fuel_sync', methods=['POST'])
def api_fuel_sync():
    return _perform_fuel_sync()


# ─── FUEL-REPORT-012H-C: Card Directory API ─────────────────────────

@fuel_bp.route('/api/card_sync', methods=['POST'])
def api_card_sync():
    """FUEL-REPORT-012H-C: Receive card directory data from Topaz agent."""
    try:
        payload = request.get_json(force=True)
    except Exception:
        return jsonify(error='invalid JSON'), 400

    api_token = current_app.config.get('FUEL_API_TOKEN')
    submitted_token = str(payload.get('token') or '') if payload else ''
    if not api_token or not payload or not hmac.compare_digest(submitted_token, str(api_token)):
        return jsonify(error='unauthorized'), 401

    source = str(payload.get('source') or 'topaz_dcCards').strip()
    cards_data = payload.get('cards', [])

    rows_received = len(cards_data)
    cards_created = 0
    cards_updated = 0
    aliases_created = 0
    aliases_updated = 0
    aliases_conflicted = 0
    rows_skipped = 0
    now = datetime.utcnow()

    for card_data in cards_data:
        try:
            card_id = str(card_data.get('card_id') or '').strip()
            if not card_id:
                rows_skipped += 1
                continue

            name = str(card_data.get('name') or '').strip()
            code = str(card_data.get('code') or '').strip()
            has_valid_name = bool(name) and name not in ('< Не выбран >', '<Не выбранно>', '<не выбрана>')
            has_valid_code = bool(code) and code not in ('< Не выбран >', '<Не выбранно>', '<не выбрана>')

            if not has_valid_name and not has_valid_code:
                rows_skipped += 1
                continue

            display_name = name if has_valid_name else code
            rfid_code = code if code else None
            partner_id = str(card_data.get('partner_id') or '').strip() or None
            enabled_val = card_data.get('enabled')
            enabled = bool(enabled_val) if enabled_val is not None else True
            car_number = str(card_data.get('car_number') or '').strip() or None
            car_model = str(card_data.get('car_model') or '').strip() or None
            transaction_id = str(card_data.get('transaction_id') or '').strip() or None

            # Upsert FuelCard
            card = FuelCard.query.filter_by(topaz_card_id=card_id).first()
            if card:
                card.display_name = display_name
                card.rfid_code = rfid_code
                card.partner_id = partner_id
                card.enabled = enabled
                card.car_number = car_number
                card.car_model = car_model
                card.topaz_transaction_id = transaction_id
                card.source = source
                card.last_seen = now
                card.updated_at = now
                cards_updated += 1
            else:
                card = FuelCard(
                    topaz_card_id=card_id,
                    display_name=display_name,
                    rfid_code=rfid_code,
                    partner_id=partner_id,
                    enabled=enabled,
                    car_number=car_number,
                    car_model=car_model,
                    topaz_transaction_id=transaction_id,
                    source=source,
                    first_seen=now,
                    last_seen=now,
                    created_at=now,
                    updated_at=now,
                )
                db.session.add(card)
                cards_created += 1

            db.session.flush()

            # Alias: topaz_card_id
            alias_rec = FuelCardAlias.query.filter_by(
                alias_type='topaz_card_id', alias_value=card_id).first()
            if alias_rec:
                if alias_rec.card_id != card.id:
                    aliases_conflicted += 1
                else:
                    alias_rec.last_seen = now
                    alias_rec.updated_at = now
                    aliases_updated += 1
            else:
                db.session.add(FuelCardAlias(
                    card_id=card.id, alias_type='topaz_card_id',
                    alias_value=card_id, source=source,
                    first_seen=now, last_seen=now, created_at=now, updated_at=now,
                ))
                aliases_created += 1

            # Alias: rfid
            if rfid_code:
                alias_rec = FuelCardAlias.query.filter_by(
                    alias_type='rfid', alias_value=rfid_code).first()
                if alias_rec:
                    if alias_rec.card_id != card.id:
                        aliases_conflicted += 1
                    else:
                        alias_rec.last_seen = now
                        alias_rec.updated_at = now
                        aliases_updated += 1
                else:
                    db.session.add(FuelCardAlias(
                        card_id=card.id, alias_type='rfid',
                        alias_value=rfid_code, source=source,
                        first_seen=now, last_seen=now, created_at=now, updated_at=now,
                    ))
                    aliases_created += 1

        except Exception:
            db.session.rollback()
            rows_skipped += 1
            continue

    sync_log = FuelCardSyncLog(
        synced_at=now, source=source,
        rows_received=rows_received, cards_created=cards_created,
        cards_updated=cards_updated, aliases_created=aliases_created,
        aliases_updated=aliases_updated, aliases_conflicted=aliases_conflicted,
        rows_skipped=rows_skipped, status='ok',
    )
    db.session.add(sync_log)
    db.session.commit()

    return jsonify(ok=True, rows_received=rows_received,
                   cards_created=cards_created, cards_updated=cards_updated,
                   aliases_created=aliases_created, aliases_updated=aliases_updated,
                   aliases_conflicted=aliases_conflicted, rows_skipped=rows_skipped)


@fuel_bp.route('/cards')
@module_required('fuel')
def card_directory():
    """FUEL-REPORT-012H-C: Card directory page with Unicode-safe search."""
    q_raw = request.args.get('q', '').strip()

    def norm(value):
        value = str(value or '')
        value = value.replace('\u0401', '\u0415').replace('\u0451', '\u0435')
        value = ' '.join(value.split())
        return value.casefold()

    all_cards = FuelCard.query.order_by(FuelCard.display_name).all()

    if q_raw:
        q_norm = norm(q_raw)
        cards = []
        for card in all_cards:
            hay = ' '.join([
                str(card.display_name or ''),
                str(card.topaz_card_id or ''),
                str(card.rfid_code or ''),
                str(card.car_number or ''),
                str(card.car_model or ''),
            ])
            if q_norm in norm(hay):
                cards.append(card)
    else:
        cards = all_cards

    total_cards = len(cards)
    active_cards = sum(1 for c in cards if c.enabled)
    total_aliases = FuelCardAlias.query.count()
    last_sync = FuelCardSyncLog.query.order_by(FuelCardSyncLog.synced_at.desc()).first()
    sync_logs = FuelCardSyncLog.query.order_by(FuelCardSyncLog.synced_at.desc()).limit(20).all()

    return render_template('fuel/cards.html',
        cards=cards, total_cards=total_cards, active_cards=active_cards,
        total_aliases=total_aliases, last_sync=last_sync,
        sync_logs=sync_logs, q=request.args.get('q', ''))


# ─── FUEL-REPORT-011A: Fuel balance period report ─────────────────────

def _fuel_report_parse_date(value, default_value):
    if not value:
        return default_value
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return default_value


def _fuel_report_date_items(start_date, end_date):
    days = []
    cur = start_date
    while cur <= end_date:
        days.append({
            "date": cur,
            "key": cur.isoformat(),
            "label": cur.strftime("%d.%m"),
        })
        cur = cur + timedelta(days=1)
    return days


def _fuel_report_sum_receipts(warehouse_id, fuel_type, date_from=None, date_to=None):
    q = db.session.query(func.coalesce(func.sum(FuelReceipt2.quantity), 0)).filter(
        FuelReceipt2.warehouse_id == warehouse_id,
        FuelReceipt2.fuel_type == fuel_type,
    )
    if date_from:
        q = q.filter(FuelReceipt2.receipt_date >= date_from)
    if date_to:
        q = q.filter(FuelReceipt2.receipt_date <= date_to)
    return float(q.scalar() or 0)


def _fuel_report_sum_transactions(warehouse_id, fuel_type, dt_from=None, dt_to_exclusive=None):
    q = db.session.query(func.coalesce(func.sum(FuelTransaction2.quantity), 0)).join(FuelStation2).filter(
        FuelStation2.warehouse_id == warehouse_id,
        FuelTransaction2.fuel_type == fuel_type,
    )
    if dt_from:
        q = q.filter(FuelTransaction2.txn_datetime >= dt_from)
    if dt_to_exclusive:
        q = q.filter(FuelTransaction2.txn_datetime < dt_to_exclusive)
    return float(q.scalar() or 0)




# ─── FUEL-REPORT-011G: manual fuel expenses included in balance report ──

def _fuel_report_sum_manual_expenses(warehouse_id, fuel_type, date_from=None, date_to=None):
    conditions = [
        "warehouse_id = :warehouse_id",
        "fuel_type = :fuel_type",
        "coalesce(is_deleted, 0) = 0",
    ]
    params = {
        "warehouse_id": warehouse_id,
        "fuel_type": fuel_type,
    }

    if date_from is not None:
        conditions.append("expense_date >= :date_from")
        params["date_from"] = date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from)

    if date_to is not None:
        conditions.append("expense_date <= :date_to")
        params["date_to"] = date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to)

    sql = "select coalesce(sum(quantity), 0) from fuel_manual_expenses where " + " and ".join(conditions)

    try:
        value = db.session.execute(text(sql), params).scalar()
        return float(value or 0)
    except Exception:
        db.session.rollback()
        return 0.0


def _fuel_report_daily_manual_expenses(warehouse_id, fuel_type, start_date, end_date):
    params = {
        "warehouse_id": warehouse_id,
        "fuel_type": fuel_type,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
    }

    try:
        rows = db.session.execute(
            text("""
                select expense_date, coalesce(sum(quantity), 0) as quantity
                from fuel_manual_expenses
                where warehouse_id = :warehouse_id
                  and fuel_type = :fuel_type
                  and coalesce(is_deleted, 0) = 0
                  and expense_date >= :start_date
                  and expense_date <= :end_date
                group by expense_date
                order by expense_date
            """),
            params,
        ).fetchall()
    except Exception:
        db.session.rollback()
        return {}

    return {str(row[0]): float(row[1] or 0) for row in rows}


def _fuel_report_daily_receipts(warehouse_id, fuel_type, start_date, end_date):
    rows = (
        db.session.query(
            FuelReceipt2.receipt_date,
            func.coalesce(func.sum(FuelReceipt2.quantity), 0),
        )
        .filter(
            FuelReceipt2.warehouse_id == warehouse_id,
            FuelReceipt2.fuel_type == fuel_type,
            FuelReceipt2.receipt_date >= start_date,
            FuelReceipt2.receipt_date <= end_date,
        )
        .group_by(FuelReceipt2.receipt_date)
        .all()
    )
    return {
        d.isoformat(): float(qty or 0)
        for d, qty in rows
        if d is not None
    }


def _fuel_report_daily_expenses(warehouse_id, fuel_type, start_date, end_date):
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    rows = (
        db.session.query(
            func.date(FuelTransaction2.txn_datetime),
            func.coalesce(func.sum(FuelTransaction2.quantity), 0),
        )
        .join(FuelStation2)
        .filter(
            FuelStation2.warehouse_id == warehouse_id,
            FuelTransaction2.fuel_type == fuel_type,
            FuelTransaction2.txn_datetime >= start_dt,
            FuelTransaction2.txn_datetime < end_dt,
        )
        .group_by(func.date(FuelTransaction2.txn_datetime))
        .all()
    )

    result = {}
    for d, qty in rows:
        if d is None:
            continue
        result[str(d)] = float(qty or 0)
    return result


def _fuel_report_build_rows(start_date, end_date, show_zero=True, fuel_type="ДТ"):
    date_items = _fuel_report_date_items(start_date, end_date)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt_exclusive = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    warehouses = (
        fuel_warehouse_query_for_ui()
        .outerjoin(Organization)
        .order_by(Organization.sort_order, FuelWarehouse.name)
        .all()
    )

    rows = []

    totals = {
        "opening": 0.0,
        "receipts": 0.0,
        "expenses": 0.0,
        "closing": 0.0,
    }

    for wh in warehouses:
        initial = (
            FuelInitialBalance.query
            .filter_by(warehouse_id=wh.id, fuel_type=fuel_type)
            .order_by(FuelInitialBalance.balance_date.asc(), FuelInitialBalance.id.asc())
            .first()
        )

        opening = 0.0
        initial_date = None
        initial_qty = 0.0

        if initial:
            initial_date = initial.balance_date
            initial_qty = float(initial.quantity or 0)

            if initial_date and initial_date <= start_date:
                before_receipts = _fuel_report_sum_receipts(
                    wh.id,
                    fuel_type,
                    date_from=initial_date,
                    date_to=start_date - timedelta(days=1),
                )
                before_expenses = _fuel_report_sum_transactions(
                    wh.id,
                    fuel_type,
                    dt_from=datetime.combine(initial_date, datetime.min.time()),
                    dt_to_exclusive=start_dt,
                )
                before_manual_expenses = _fuel_report_sum_manual_expenses(
                    wh.id,
                    fuel_type,
                    date_from=initial_date,
                    date_to=start_date - timedelta(days=1),
                )
                opening = initial_qty + before_receipts - before_expenses - before_manual_expenses
            elif initial_date and initial_date > start_date:
                opening = 0.0
            else:
                opening = initial_qty

        period_receipts = _fuel_report_sum_receipts(
            wh.id,
            fuel_type,
            date_from=start_date,
            date_to=end_date,
        )

        period_topaz_expenses = _fuel_report_sum_transactions(
            wh.id,
            fuel_type,
            dt_from=start_dt,
            dt_to_exclusive=end_dt_exclusive,
        )
        period_manual_expenses = _fuel_report_sum_manual_expenses(
            wh.id,
            fuel_type,
            date_from=start_date,
            date_to=end_date,
        )
        period_expenses = period_topaz_expenses + period_manual_expenses

        closing = opening + period_receipts - period_expenses

        daily_receipts = _fuel_report_daily_receipts(wh.id, fuel_type, start_date, end_date)
        daily_expenses = _fuel_report_daily_expenses(wh.id, fuel_type, start_date, end_date)
        daily_manual_expenses = _fuel_report_daily_manual_expenses(wh.id, fuel_type, start_date, end_date)

        daily = {}
        for item in date_items:
            key = item["key"]
            daily[key] = {
                "receipts": round(daily_receipts.get(key, 0.0), 2),
                "expenses": round(daily_expenses.get(key, 0.0) + daily_manual_expenses.get(key, 0.0), 2),
            }

        if not show_zero:
            if round(opening, 2) == 0 and round(period_receipts, 2) == 0 and round(period_expenses, 2) == 0 and round(closing, 2) == 0:
                continue

        row = {
            "warehouse_id": wh.id,
            "organization": wh.organization.name if wh.organization else "",
            "warehouse": wh.name,
            "fuel_type": fuel_type,
            "initial_date": initial_date,
            "opening": round(opening, 2),
            "receipts": round(period_receipts, 2),
            "expenses": round(period_expenses, 2),
            "closing": round(closing, 2),
            "daily": daily,
        }

        rows.append(row)

        totals["opening"] += row["opening"]
        totals["receipts"] += row["receipts"]
        totals["expenses"] += row["expenses"]
        totals["closing"] += row["closing"]

    for k in totals:
        totals[k] = round(totals[k], 2)

    return rows, date_items, totals


@fuel_bp.route('/balance-report')
@login_required
def balance_report():
    today = date.today()
    default_start = today.replace(day=1)
    default_end = today

    start_date = _fuel_report_parse_date(request.args.get("start_date"), default_start)
    end_date = _fuel_report_parse_date(request.args.get("end_date"), default_end)

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    max_days = 120
    if (end_date - start_date).days + 1 > max_days:
        flash(f"Период отчёта ограничен {max_days} днями.", "warning")
        end_date = start_date + timedelta(days=max_days - 1)

    show_zero = request.args.get("hide_zero") != "1"

    rows, date_items, totals = _fuel_report_build_rows(
        start_date=start_date,
        end_date=end_date,
        show_zero=show_zero,
        fuel_type="ДТ",
    )

    return render_template(
        "fuel/balance_report.html",
        rows=rows,
        date_items=date_items,
        totals=totals,
        start_date=start_date,
        end_date=end_date,
        show_zero=show_zero,
        fuel_type="ДТ",
    )


@fuel_bp.route('/balance-report/export')
@login_required
def balance_report_export():
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    default_start = today.replace(day=1)
    default_end = today

    start_date = _fuel_report_parse_date(request.args.get("start_date"), default_start)
    end_date = _fuel_report_parse_date(request.args.get("end_date"), default_end)

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    max_days = 120
    if (end_date - start_date).days + 1 > max_days:
        end_date = start_date + timedelta(days=max_days - 1)

    show_zero = request.args.get("hide_zero") != "1"

    rows, date_items, totals = _fuel_report_build_rows(
        start_date=start_date,
        end_date=end_date,
        show_zero=show_zero,
        fuel_type="ДТ",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ФАКТ"

    base_headers = [
        "№",
        "Организация",
        "Склад",
        f"Остаток на {start_date.strftime('%d/%m/%Y')}",
        "Приход",
        "Расход",
        "Текущий остаток",
    ]

    max_col = len(base_headers) + len(date_items) * 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1).value = f"Отчёт по остаткам топлива за период {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for col, title in enumerate(base_headers, start=1):
        ws.cell(row=3, column=col).value = title
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    col = len(base_headers) + 1
    for item in date_items:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.cell(row=3, column=col).value = item["date"]
        ws.cell(row=3, column=col).number_format = "dd.mm.yyyy"
        ws.cell(row=4, column=col).value = "Приход"
        ws.cell(row=4, column=col + 1).value = "Расход"
        col += 2

    data_start_row = 5
    for idx, row in enumerate(rows, start=1):
        r = data_start_row + idx - 1
        ws.cell(row=r, column=1).value = idx
        ws.cell(row=r, column=2).value = row["organization"]
        ws.cell(row=r, column=3).value = row["warehouse"]
        ws.cell(row=r, column=4).value = row["opening"]
        ws.cell(row=r, column=5).value = row["receipts"]
        ws.cell(row=r, column=6).value = row["expenses"]
        ws.cell(row=r, column=7).value = row["closing"]

        col = 8
        for item in date_items:
            daily = row["daily"][item["key"]]
            ws.cell(row=r, column=col).value = daily["receipts"] if daily["receipts"] else None
            ws.cell(row=r, column=col + 1).value = daily["expenses"] if daily["expenses"] else None
            col += 2

    total_row = data_start_row + len(rows)
    ws.cell(row=total_row, column=2).value = "ИТОГО"
    ws.cell(row=total_row, column=4).value = totals["opening"]
    ws.cell(row=total_row, column=5).value = totals["receipts"]
    ws.cell(row=total_row, column=6).value = totals["expenses"]
    ws.cell(row=total_row, column=7).value = totals["closing"]

    thin = Side(style="thin", color="D1D5DB")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    total_fill = PatternFill("solid", fgColor="FEF3C7")

    for row_cells in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=max_col):
        for cell in row_cells:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_cells in ws.iter_rows(min_row=5, max_row=total_row, min_col=1, max_col=max_col):
        for cell in row_cells:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = total_fill

    for col_idx in range(4, max_col + 1):
        for row_idx in range(5, total_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    widths = {
        1: 5,
        2: 28,
        3: 24,
        4: 16,
        5: 14,
        6: 14,
        7: 16,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx in range(8, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    ws.freeze_panes = "H5"
    ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{total_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fuel_balance_report_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

