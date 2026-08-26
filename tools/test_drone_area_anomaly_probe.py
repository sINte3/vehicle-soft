# -*- coding: utf-8 -*-
"""Тесты tools/drone_area_anomaly_probe.py. Без сети и без реальной базы.

    python tools/test_drone_area_anomaly_probe.py

Восемь обязательных проверок режима «только чтение» (задание A2 §6) плюс
отрицательные тесты на каждое значение, которое способно испортить итог тихо:
`NaN`, обе бесконечности, строка вместо числа, отсутствующий контрольный
вылет.

Синтетическая база строится тем же DDL, что стоит в
migrate_drones_foundation_001.py, а не тем, что удобно тесту: проверка на
схеме, отличной от production, проверяет не то.

Замечание про NaN. SQLite хранит `NaN` как NULL, а обе бесконечности -- как
настоящий REAL. Поэтому колонка МОЖЕТ содержать `inf`, и защита нужна на обеих
сторонах: и на `raw_json` (Python пишет и читает литерал `NaN` без возражений),
и на колонке.
"""

import contextlib
import datetime
import io
import json
import math
import os
import sqlite3
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.drone_area_anomaly_probe import (  # noqa: E402
    DEFAULT_LOOKBACK, KNOWN_CASE_FLIGHT_ID, QUALITY_CODES, ProbeError,
    analyse, check_known_case, classify_number, connect_read_only, main,
    sha256_of)

# Точная копия DDL из migrate_drones_foundation_001.py.
DDL_FLIGHTS = """
    CREATE TABLE drone_flights (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        dji_flight_id BIGINT NOT NULL UNIQUE,
        drone_unit_id INTEGER REFERENCES drone_units (id),
        nickname_raw  VARCHAR(100),
        serial_number VARCHAR(50),
        flyer_name    VARCHAR(100),
        team_name     VARCHAR(100),
        started_at    DATETIME NOT NULL,
        finished_at   DATETIME,
        work_seconds  INTEGER,
        area_ha       FLOAT NOT NULL DEFAULT 0,
        spray_liters  FLOAT,
        sow_kg        FLOAT,
        usage_type    INTEGER,
        mode_name     INTEGER,
        manual_mode   BOOLEAN,
        work_speed    FLOAT,
        spray_width   FLOAT,
        radar_height  FLOAT,
        lat           FLOAT,
        lng           FLOAT,
        location_text VARCHAR(500),
        region        VARCHAR(100),
        raw_json      TEXT NOT NULL,
        sync_log_id   INTEGER,
        ingested_at   DATETIME NOT NULL
    )
"""

DDL_UNITS = """
    CREATE TABLE drone_units (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        number      INTEGER NOT NULL UNIQUE,
        hardware_id VARCHAR(50)
    )
"""

BASE = datetime.datetime(2026, 6, 5, 14, 0, 0)

# Ellipsis означает «ключа в payload нет вовсе», None -- «ключ есть, null».
ABSENT = Ellipsis


def flight(dji_id, minutes, area_m2, width=5.95, unit_id=1,
           nickname='8 GardenU', duration=500, broken_json=False,
           area_ha_override=None, width_col_override=ABSENT,
           finished_delta=None):
    """Одна строка вылета в форме, пригодной для INSERT."""
    started = BASE + datetime.timedelta(minutes=minutes)
    finished = started + datetime.timedelta(
        seconds=duration if finished_delta is None else finished_delta)
    payload = {
        'id': dji_id,
        'nickname': nickname,
        'serial_number': 'R%010d' % dji_id,
        'start_timestamp': int(started.timestamp()),
        'end_timestamp': int(finished.timestamp()),
        'work_time_seconds': duration,
    }
    if area_m2 is not ABSENT:
        payload['new_work_area'] = area_m2
    if width is not ABSENT:
        payload['spray_width'] = width
    raw = 'НЕ JSON {{{' if broken_json else json.dumps(payload,
                                                       ensure_ascii=False)
    if area_ha_override is not None:
        area_ha = area_ha_override
    elif isinstance(area_m2, (int, float)) and not isinstance(area_m2, bool) \
            and math.isfinite(area_m2):
        area_ha = area_m2 / 10000.0
    else:
        area_ha = 0.0
    if width_col_override is not ABSENT:
        width_col = width_col_override
    elif isinstance(width, (int, float)) and not isinstance(width, bool) \
            and math.isfinite(width):
        width_col = width
    else:
        width_col = None
    return (dji_id, unit_id, nickname, 'R%010d' % dji_id,
            started.strftime('%Y-%m-%d %H:%M:%S'),
            finished.strftime('%Y-%m-%d %H:%M:%S'),
            duration, area_ha, width_col, raw,
            BASE.strftime('%Y-%m-%d %H:%M:%S'))


def known_case_rows(**overrides):
    """Три вылета 05.06.2026 в настоящем порядке -- контрольный случай.

    Между двумя равными площадями стоит вылет с другой: именно поэтому окно в
    один шаг известный случай не ловит.
    """
    return [
        flight(622715273, 0, 5940.000029700001, width=5.91, duration=362),
        flight(622715274, 10, 3293.3333498, width=ABSENT, duration=90),
        flight(KNOWN_CASE_FLIGHT_ID, 20, 5940.000029700001, width=ABSENT,
               duration=40),
    ]


def build_db(path, rows):
    con = sqlite3.connect(path)
    con.execute(DDL_FLIGHTS)
    con.execute(DDL_UNITS)
    con.execute('INSERT INTO drone_units (id, number, hardware_id) '
                "VALUES (1, 8, 'FIXTURE0000000000000')")
    con.executemany(
        'INSERT INTO drone_flights (dji_flight_id, drone_unit_id, '
        ' nickname_raw, serial_number, started_at, finished_at, work_seconds, '
        ' area_ha, spray_width, raw_json, ingested_at) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?)', rows)
    con.commit()
    con.close()


def open_workbook(path):
    """Открыть xlsx, не оставляя файл открытым.

    [REASON]: openpyxl.load_workbook(path) держит zip открытым и в CI даёт
    ResourceWarning на каждый тест. Предупреждение в выводе проверки -- шум,
    который потом мешает заметить настоящее.
    """
    from openpyxl import load_workbook
    with open(path, 'rb') as handle:
        blob = io.BytesIO(handle.read())
    return load_workbook(blob)


class ProbeCase(unittest.TestCase):
    """Общая мастерская: собрать базу, прогнать анализ, вернуть результат."""

    def make_db(self, rows):
        directory = tempfile.mkdtemp()
        path = os.path.join(directory, 'copy.db')
        build_db(path, rows)
        return path

    def run_probe(self, rows, **kwargs):
        path = self.make_db(rows)
        before = sha256_of(path)
        con = connect_read_only(path)
        try:
            result = analyse(con, **kwargs)
        finally:
            con.close()
        after = sha256_of(path)
        return path, result, before, after

    def states(self, rows, **kwargs):
        _p, result, _b, _a = self.run_probe(rows, **kwargs)
        return result['widths']['states']


# ─── 1, 2, 7: режим только чтения и повторяемость ────────────────────────────

class TestReadOnlyGuarantee(ProbeCase):

    def test_1_sha256_before_and_after_match(self):
        _path, _result, before, after = self.run_probe(known_case_rows())
        self.assertEqual(before, after)

    def test_1b_the_file_is_byte_identical_before_and_after(self):
        """Побайтовое сравнение, а не только хеш."""
        path = self.make_db(known_case_rows())
        with open(path, 'rb') as handle:
            before = handle.read()
        con = connect_read_only(path)
        try:
            analyse(con)
        finally:
            con.close()
        with open(path, 'rb') as handle:
            self.assertEqual(handle.read(), before)

    def test_2_the_connection_really_is_read_only(self):
        """Не «мы открыли с mode=ro», а «запись через него отвергается»."""
        path = self.make_db(known_case_rows())
        con = connect_read_only(path)
        try:
            with self.assertRaises(sqlite3.OperationalError):
                con.execute('DELETE FROM drone_flights')
        finally:
            con.close()

    def test_2b_a_missing_database_is_refused_not_created(self):
        directory = tempfile.mkdtemp()
        missing = os.path.join(directory, 'nope.db')
        with self.assertRaises(ProbeError):
            connect_read_only(missing)
        self.assertFalse(os.path.exists(missing),
                         'the probe must never create a database')

    def test_7_a_second_run_gives_the_same_answer(self):
        rows = known_case_rows() + [flight(1, 60, 9193.0)]
        path, first, _b, _a = self.run_probe(rows)
        con = connect_read_only(path)
        try:
            second = analyse(con)
        finally:
            con.close()
        self.assertEqual(first['repeats']['count'], second['repeats']['count'])
        self.assertEqual(first['widths']['states'], second['widths']['states'])
        self.assertAlmostEqual(first['total_area_ha'], second['total_area_ha'])


class TestSourceHasNoWrites(unittest.TestCase):

    def test_3_no_writing_sql_in_the_probe_source(self):
        here = os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(here, 'drone_area_anomaly_probe.py'),
                  encoding='utf-8') as handle:
            source = handle.read()
        for statement in ('INSERT INTO', 'UPDATE ', 'DELETE FROM',
                          'DROP TABLE', 'ALTER TABLE', 'CREATE TABLE',
                          'REPLACE INTO'):
            self.assertNotIn(statement, source,
                             'writing SQL found in the probe: %s' % statement)


# ─── 4 и 5: контроли на повтор площади ───────────────────────────────────────

class TestRepeatDetection(ProbeCase):

    def test_4_negative_control_no_anomalies_means_zero_candidates(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0), flight(2, 10, 14906.0),
            flight(3, 20, 14713.0), flight(4, 30, 15366.0)])
        self.assertEqual(result['repeats']['count'], 0)
        self.assertEqual(result['repeats']['area_ha'], 0.0)

    def test_5_positive_control_a_repeat_is_found(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0), flight(2, 10, 5940.0),
            flight(3, 20, 5940.0)])
        self.assertEqual(result['repeats']['count'], 1)
        candidate = result['repeats']['candidates'][0]
        self.assertEqual(candidate['dji_flight_id'], 3)
        self.assertAlmostEqual(candidate['area_ha'], 0.594)

    def test_5b_the_known_case_is_recognised_at_lag_two(self):
        _p, result, _b, _a = self.run_probe(known_case_rows())
        known = check_known_case(result)
        self.assertTrue(known['passed'])
        self.assertEqual(known['detail']['lag'], 2)
        self.assertTrue(known['detail']['short_flight'])

    def test_5b2_a_one_step_window_would_miss_the_known_case(self):
        """Отрицательный контроль: расхождение с формулировкой задания реально."""
        _p, narrow, _b, _a = self.run_probe(known_case_rows(), lookback=1)
        self.assertEqual(narrow['repeats']['count'], 0)
        self.assertFalse(check_known_case(narrow)['passed'])

    def test_5c_a_zero_area_repeat_is_not_a_candidate(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 0.0), flight(2, 10, 0.0)])
        self.assertEqual(result['repeats']['count'], 0)

    def test_5d_a_repeat_across_different_machines_is_not_a_candidate(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 5940.0, unit_id=1, nickname='8 GardenU'),
            flight(2, 10, 5940.0, unit_id=2, nickname='9 Garden')])
        self.assertEqual(result['repeats']['count'], 0)

    def test_5e_a_run_of_three_is_counted_as_a_run(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 5940.0), flight(2, 10, 5940.0),
            flight(3, 20, 5940.0)])
        self.assertEqual(result['repeats']['count'], 2)
        self.assertEqual(result['repeats']['runs_3plus_count'], 1)

    def test_5f_lag_one_and_lag_two_are_counted_apart(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 5940.0), flight(2, 10, 5940.0),
            flight(3, 20, 1000.0), flight(4, 30, 1000.0),
            flight(5, 40, 7000.0), flight(6, 50, 3000.0),
            flight(7, 60, 7000.0)])
        self.assertEqual(result['repeats']['lag1_count'], 2)
        self.assertEqual(result['repeats']['lag2plus_count'], 1)


# ─── Контрольный случай как предусловие ──────────────────────────────────────

class TestKnownCasePrecondition(ProbeCase):

    def test_a_missing_flight_fails_the_precondition(self):
        _p, result, _b, _a = self.run_probe([flight(1, 0, 14520.0)])
        known = check_known_case(result)
        self.assertFalse(known['passed'])
        self.assertFalse(known['present_in_database'])
        self.assertIn('отсутствует', known['reason'])

    def test_a_present_but_unrecognised_flight_fails_differently(self):
        """Разные причины -- разные действия, и их нельзя смешивать."""
        _p, result, _b, _a = self.run_probe([
            flight(KNOWN_CASE_FLIGHT_ID, 0, 14520.0)])
        known = check_known_case(result)
        self.assertFalse(known['passed'])
        self.assertTrue(known['present_in_database'])
        self.assertIn('кандидатом не опознан', known['reason'])

    def test_main_refuses_to_write_a_report_without_the_known_case(self):
        path = self.make_db([flight(1, 0, 14520.0), flight(2, 10, 14520.0)])
        directory = os.path.dirname(path)
        out = os.path.join(directory, 'A2.xlsx')
        with contextlib.redirect_stdout(io.StringIO()) as buffer:
            code = main(['--db', path, '--out', out])
        self.assertEqual(code, 4)
        self.assertFalse(os.path.exists(out),
                         'no xlsx may be written when the precondition fails')
        self.assertIn('PRECONDITION FAILED', buffer.getvalue())

    def test_the_escape_hatch_writes_a_stamped_report_and_still_fails(self):
        path = self.make_db([flight(1, 0, 14520.0), flight(2, 10, 14520.0)])
        out = os.path.join(os.path.dirname(path), 'A2.xlsx')
        with contextlib.redirect_stdout(io.StringIO()):
            code = main(['--db', path, '--out', out,
                         '--allow-missing-known-case'])
        self.assertEqual(code, 4, 'the exit code stays non-zero')
        self.assertTrue(os.path.exists(out))
        book = open_workbook(out)
        first_cell = book['Сводка'].cell(row=1, column=1).value
        self.assertIn('НЕ ПРОШЁЛ КОНТРОЛЬНЫЙ СЛУЧАЙ', first_cell)

    def test_a_passing_run_writes_an_unstamped_report(self):
        """Отрицательный контроль к отметке: на годном прогоне её нет."""
        path = self.make_db(known_case_rows())
        out = os.path.join(os.path.dirname(path), 'A2.xlsx')
        with contextlib.redirect_stdout(io.StringIO()):
            code = main(['--db', path, '--out', out])
        self.assertEqual(code, 0)
        book = open_workbook(out)
        first_cell = book['Сводка'].cell(row=1, column=1).value
        self.assertNotIn('НЕ ПРОШЁЛ', first_cell)


# ─── Ширина: одиннадцать состояний ───────────────────────────────────────────

class TestWidthStates(ProbeCase):

    def test_null_missing_minus_one_and_zero_are_told_apart(self):
        states = self.states([
            flight(1, 0, 14520.0, width=5.95),      # POSITIVE_UNVALIDATED
            flight(2, 10, 14520.0, width=None),     # JSON_NULL
            flight(3, 20, 14520.0, width=ABSENT),   # MISSING_KEY
            flight(4, 30, 14520.0, width=-1.0),     # MINUS_ONE
            flight(5, 40, 14520.0, width=0.0),      # ZERO
            flight(6, 50, 14520.0, width=-7.0),     # NEGATIVE
        ])
        self.assertEqual(states.get('POSITIVE_UNVALIDATED'), 1)
        self.assertEqual(states.get('JSON_NULL'), 1)
        self.assertEqual(states.get('MISSING_KEY'), 1)
        self.assertEqual(states.get('MINUS_ONE'), 1)
        self.assertEqual(states.get('ZERO'), 1)
        self.assertEqual(states.get('NEGATIVE'), 1)

    def test_a_string_width_is_invalid_type_not_json_null(self):
        """Сломанный формат и «не записано» -- разные вещи."""
        states = self.states([flight(1, 0, 14520.0, width='5.95')])
        self.assertEqual(states.get('INVALID_TYPE'), 1)
        self.assertIsNone(states.get('JSON_NULL'))

    def test_a_boolean_width_is_invalid_type(self):
        """isinstance(True, int) истинно -- без явной проверки True стало бы 1 м."""
        states = self.states([flight(1, 0, 14520.0, width=True)])
        self.assertEqual(states.get('INVALID_TYPE'), 1)

    def test_a_list_or_object_width_is_invalid_type(self):
        states = self.states([flight(1, 0, 14520.0, width=[5.95]),
                              flight(2, 10, 14520.0, width={'v': 5.95})])
        self.assertEqual(states.get('INVALID_TYPE'), 2)

    def test_nan_and_both_infinities_are_non_finite(self):
        states = self.states([
            flight(1, 0, 14520.0, width=float('nan')),
            flight(2, 10, 14520.0, width=float('inf')),
            flight(3, 20, 14520.0, width=float('-inf')),
        ])
        self.assertEqual(states.get('NON_FINITE'), 3)

    def test_a_non_finite_width_never_becomes_usable(self):
        _p, result, _b, _a = self.run_probe(
            [flight(1, 0, 14520.0, width=float('inf'))],
            min_width=1.0, max_width=20.0)
        self.assertEqual(result['width_usable_flights'], 0)
        self.assertEqual(result['width_positive_flights'], 0)

    def test_without_a_configured_range_nothing_is_usable(self):
        _p, result, _b, _a = self.run_probe([flight(1, 0, 14520.0, width=5.95)])
        self.assertEqual(result['width_usable_flights'], 0)
        self.assertEqual(result['width_positive_flights'], 1)

    def test_with_a_configured_range_a_good_value_becomes_usable(self):
        """Отрицательный контроль к предыдущему."""
        _p, result, _b, _a = self.run_probe([flight(1, 0, 14520.0, width=5.95)],
                                            min_width=1.0, max_width=20.0)
        self.assertEqual(result['width_usable_flights'], 1)

    def test_an_absurdly_large_width_falls_out_of_the_configured_range(self):
        states = self.states([flight(1, 0, 14520.0, width=5.95),
                              flight(2, 10, 14520.0, width=5000.0)],
                             min_width=1.0, max_width=20.0)
        self.assertEqual(states.get('USABLE'), 1)
        self.assertEqual(states.get('OUT_OF_CONFIGURED_RANGE'), 1)

    def test_a_large_width_without_a_range_is_only_unvalidated(self):
        """Без границ инструмент не имеет права звать значение невозможным."""
        states = self.states([flight(1, 0, 14520.0, width=5000.0)])
        self.assertEqual(states.get('POSITIVE_UNVALIDATED'), 1)
        self.assertIsNone(states.get('OUT_OF_CONFIGURED_RANGE'))

    def test_no_width_is_never_substituted(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 10000.0, width=5.95),
            flight(2, 10, 20000.0, width=-1.0)],
            min_width=1.0, max_width=20.0)
        self.assertAlmostEqual(result['width_usable_area_ha'], 1.0)
        self.assertAlmostEqual(
            result['total_area_ha'] - result['width_usable_area_ha'], 2.0)

    def test_an_infinite_column_width_does_not_leak_in(self):
        """SQLite хранит inf настоящим REAL, поэтому колонку тоже проверяем."""
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0, width=ABSENT,
                   width_col_override=float('inf'))],
            min_width=1.0, max_width=20.0)
        self.assertEqual(result['width_usable_flights'], 0)


class TestWidthDistribution(ProbeCase):

    def test_the_distribution_reports_min_median_and_max(self):
        rows = [flight(index, index * 10, 10000.0, width=width)
                for index, width in enumerate([5.0, 6.0, 7.0, 8.0, 9.0], 1)]
        _p, result, _b, _a = self.run_probe(rows)
        distribution = result['widths']['distribution']
        self.assertEqual(distribution['count'], 5)
        self.assertAlmostEqual(distribution['min'], 5.0)
        self.assertAlmostEqual(distribution['median'], 7.0)
        self.assertAlmostEqual(distribution['max'], 9.0)

    def test_an_outlier_is_counted_but_stays_usable(self):
        """Выброс -- редкое значение, а не ошибка DJI."""
        widths = [6.0] * 20 + [60.0]
        rows = [flight(index, index * 10, 10000.0, width=width)
                for index, width in enumerate(widths, 1)]
        _p, result, _b, _a = self.run_probe(rows, min_width=1.0,
                                            max_width=100.0)
        self.assertEqual(result['widths']['distribution']['outliers'], 1)
        self.assertEqual(result['width_usable_flights'], 21,
                         'an outlier must not be demoted to unusable')

    def test_a_flat_distribution_has_no_outliers(self):
        """Отрицательный контроль к предыдущему."""
        rows = [flight(index, index * 10, 10000.0, width=6.0)
                for index in range(1, 21)]
        _p, result, _b, _a = self.run_probe(rows)
        self.assertEqual(result['widths']['distribution']['outliers'], 0)


# ─── Площадь: типы и конечность ──────────────────────────────────────────────

class TestAreaValues(ProbeCase):

    def test_a_string_area_is_reported_and_excluded_from_sums(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 10000.0),
            flight(2, 10, '14520.0', area_ha_override=0.0)])
        self.assertEqual(
            result['quality']['issues'].get('new_work_area: значение не число'),
            1)
        self.assertAlmostEqual(result['total_area_ha'], 1.0)

    def test_a_nan_area_never_enters_the_total(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 10000.0),
            flight(2, 10, float('nan'), area_ha_override=0.0)])
        self.assertTrue(math.isfinite(result['total_area_ha']))
        self.assertAlmostEqual(result['total_area_ha'], 1.0)
        self.assertEqual(
            result['quality']['issues'].get(
                'new_work_area: значение NaN или бесконечность'), 1)

    def test_an_infinite_column_area_never_enters_the_total(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 10000.0),
            flight(2, 10, ABSENT, area_ha_override=float('inf'))])
        self.assertTrue(math.isfinite(result['total_area_ha']))
        self.assertAlmostEqual(result['total_area_ha'], 1.0)
        self.assertEqual(result['rows_without_usable_area'], 1)

    def test_a_non_finite_area_cannot_create_a_repeat(self):
        """NaN == NaN ложно, но полагаться на это нельзя -- проверяем явно."""
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, float('nan'), area_ha_override=0.0),
            flight(2, 10, float('nan'), area_ha_override=0.0)])
        self.assertEqual(result['repeats']['count'], 0)

    def test_classify_number_tells_the_four_problems_apart(self):
        self.assertEqual(classify_number({}, 'x'), (None, 'MISSING_KEY'))
        self.assertEqual(classify_number({'x': None}, 'x'), (None, 'JSON_NULL'))
        self.assertEqual(classify_number({'x': 'a'}, 'x'),
                         (None, 'INVALID_TYPE'))
        self.assertEqual(classify_number({'x': True}, 'x'),
                         (None, 'INVALID_TYPE'))
        self.assertEqual(classify_number({'x': float('nan')}, 'x'),
                         (None, 'NON_FINITE'))
        self.assertEqual(classify_number({'x': 5.5}, 'x'), (5.5, None))


# ─── Битая строка и запасной источник ────────────────────────────────────────

class TestBrokenRowSurvives(ProbeCase):

    def test_a_broken_raw_json_is_counted_not_fatal(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0),
            flight(2, 10, 14906.0, broken_json=True),
            flight(3, 20, 14713.0)])
        self.assertEqual(result['flights_total'], 3)
        self.assertEqual(
            result['quality']['issues'].get('raw_json не разобрался'), 1)

    def test_the_fallback_to_the_column_is_recorded_as_an_issue(self):
        """Молча подставить колонку нельзя -- это скрыло бы расхождение."""
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0, broken_json=True)])
        self.assertAlmostEqual(result['total_area_ha'], 1.452, places=3)
        self.assertEqual(
            result['quality']['issues'].get(
                'площадь взята из колонки: raw_json непригоден'), 1)

    def test_a_bad_value_in_a_readable_json_does_not_fall_back(self):
        """Отрицательный контроль: запасной источник только для битого JSON."""
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 'nonsense', area_ha_override=9.99)])
        self.assertEqual(result['rows_without_usable_area'], 1)
        self.assertAlmostEqual(result['total_area_ha'], 0.0)
        self.assertIsNone(result['quality']['issues'].get(
            'площадь взята из колонки: raw_json непригоден'))

    def test_a_column_json_area_mismatch_is_reported(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0, area_ha_override=9.99)])
        self.assertEqual(
            result['quality']['issues'].get(
                'площадь в колонке и в raw_json расходятся'), 1)

    def test_a_clean_database_reports_no_quality_issues(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0), flight(2, 10, 14906.0)])
        self.assertEqual(result['quality']['issues'], {})


# ─── Схема и идентификатор борта ─────────────────────────────────────────────

class TestSchemaAndIdentity(ProbeCase):

    def test_a_missing_required_column_stops_the_run(self):
        directory = tempfile.mkdtemp()
        path = os.path.join(directory, 'thin.db')
        con = sqlite3.connect(path)
        con.execute('CREATE TABLE drone_flights (id INTEGER PRIMARY KEY, '
                    'dji_flight_id BIGINT)')
        con.commit()
        con.close()
        con = connect_read_only(path)
        try:
            with self.assertRaises(ProbeError):
                analyse(con)
        finally:
            con.close()

    def test_serial_number_is_measured_as_a_flight_id(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0), flight(2, 10, 14906.0),
            flight(3, 20, 14713.0)])
        identity = result['identity']
        self.assertEqual(identity['distinct_serial_number'],
                         identity['flights_total'])

    def test_hardware_id_is_absent_from_the_flight_payload(self):
        _p, result, _b, _a = self.run_probe([flight(1, 0, 14520.0)])
        self.assertEqual(result['identity']['hardware_id_in_raw_json'], 0)
        self.assertFalse(result['identity']['hardware_id_on_flight_row'])

    def test_unresolved_flights_are_not_lumped_together(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 5940.0, unit_id=None, nickname='неизвестный A'),
            flight(2, 10, 5940.0, unit_id=None, nickname='неизвестный Б')])
        self.assertEqual(result['repeats']['count'], 0)
        self.assertEqual(result['identity']['drone_unit_id_null'], 2)


# ─── Отчёт целиком ───────────────────────────────────────────────────────────

class TestReportIsWritable(ProbeCase):

    def test_main_writes_all_seven_sheets(self):
        path = self.make_db(known_case_rows() + [
            flight(700000001, 60, 14520.0, width=5.95),
            flight(700000002, 80, 9193.0, width=0.0),
            flight(700000003, 100, 12000.0, width=None),
            flight(700000004, 120, 11000.0, broken_json=True),
            flight(700000005, 140, 1000.0, width='wide'),
            flight(700000006, 160, 1000.0, unit_id=None,
                   nickname='неизвестный'),
        ])
        out = os.path.join(os.path.dirname(path), 'A2.xlsx')
        before = sha256_of(path)
        with contextlib.redirect_stdout(io.StringIO()):
            code = main(['--db', path, '--out', out,
                         '--min-width-m', '1', '--max-width-m', '20'])
        self.assertEqual(code, 0)
        self.assertEqual(sha256_of(path), before)
        book = open_workbook(out)
        self.assertEqual(book.sheetnames,
                         ['Сводка', 'По месяцам', 'По дронам',
                          'Повторы площади', 'Нет ширины', 'Качество данных',
                          'Методика'])
        self.assertGreater(book['Повторы площади'].max_row, 1)
        self.assertGreater(book['Нет ширины'].max_row, 1)
        self.assertGreater(book['Качество данных'].max_row, 1)

    def test_main_refuses_a_missing_database(self):
        directory = tempfile.mkdtemp()
        missing = os.path.join(directory, 'nope.db')
        with contextlib.redirect_stdout(io.StringIO()):
            self.assertEqual(main(['--db', missing]), 2)
        self.assertFalse(os.path.exists(missing))

    def test_main_refuses_an_inverted_width_range(self):
        path = self.make_db(known_case_rows())
        with contextlib.redirect_stdout(io.StringIO()):
            code = main(['--db', path, '--min-width-m', '20',
                         '--max-width-m', '1'])
        self.assertEqual(code, 1)


class TestQualityCodesAreComplete(ProbeCase):

    def test_every_issue_kind_has_an_ascii_code(self):
        _p, result, _b, _a = self.run_probe([
            flight(1, 0, 14520.0),
            flight(2, 10, 14906.0, broken_json=True),
            flight(3, 20, 14713.0, area_ha_override=9.99),
            flight(4, 30, 1000.0, unit_id=None, nickname='X'),
            flight(5, 40, -5.0),
            flight(6, 50, 1000.0, duration=0, finished_delta=0),
            flight(7, 60, 'nope'),
            flight(8, 70, float('nan')),
            flight(9, 80, 1000.0, width='nope'),
            flight(10, 90, 1000.0, width=float('inf')),
            flight(11, 100, ABSENT, area_ha_override=float('inf')),
        ])
        self.assertTrue(result['quality']['issues'])
        for kind in result['quality']['issues']:
            self.assertIn(kind, QUALITY_CODES,
                          'issue kind without an ASCII code: %s' % kind)


if __name__ == '__main__':
    unittest.main()
