# -*- coding: utf-8 -*-
"""Мост «заказчик книги <-> контур DJI» за месяц. Только чтение.

    python tools/drone_books_vs_contours.py --db instance/transport.db \\
        --month 2025-09 --out books_vs_contours_2025_09.xlsx

ЗАЧЕМ. У вылета нет заказчика: `plot_name` пуст на всех строках. Привязать
вылет к полю по координатам НЕ ВЫШЛО, и причина измерена, а не предположена:
координата вылета -- это место взлёта, а не поле. В четверти групп «день x
машина» все вылеты стоят в одной точке с разбросом 0 метров, а контур площадью
1.68 га получил 45 га работы. Полигон этого не чинит: точка стоит не там.

Что осталось рабочим -- ИМЯ. Оператор вводит его с пульта, диспетчер пишет
своё в книге, и это один и тот же заказчик:

    книга «Аваз Исматов фх»  <->  контур «avaz ismatov 1»

Правило сравнения имён -- в tools/drone_name_match.py, вместе с причиной
каждого решения. Здесь -- только сведение месяца.

ЧТО ЭТОТ ОТЧЁТ НЕ ДЕЛАЕТ. Не считает гектары заново и не переписывает учёт.
Совпадение имени -- повод показать строку владельцу, а не основание что-то
поправить. Дата контура тоже не является датой работы: контур рисуют в один
день, а летают на нём в другой, и по дням площадь контуров к телеметрии
скачет от 0 % до 204 % при 70 % за месяц. Поэтому дата здесь -- столбец для
глаза, а не ключ.

Соединение открывается только на чтение, stdlib sqlite3, без Flask:
`from app import app` здесь недопустим -- create_app() вызывает db.create_all()
на импорте и превращает читателя в писателя.

В консоль -- только ASCII (кодировка консоли Windows), кириллица уходит в xlsx.
"""

import argparse
import collections
import datetime
import os
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.drone_name_match import (  # noqa: E402
    SCORE_CANDIDATE,
    SCORE_CONFIDENT,
    score,
)

# Смещение показа: контуры и вылеты приводятся к одному дню, как в отчётах.
UTC_OFFSET_HOURS = 5

SOURCE_DJI = 'dji'

SHEET_SUMMARY = 'Сводка'
SHEET_MATCHED = 'Совпадения'
SHEET_CANDIDATES = 'Кандидаты'
SHEET_BOOK_ONLY = 'Книга без контура'
SHEET_CONTOUR_ONLY = 'Контуры без книги'


def open_readonly(path):
    if not os.path.exists(path):
        print('ERROR: database not found at %s' % path)
        sys.exit(2)
    uri = 'file:%s?mode=ro' % path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def local_day(value):
    """DATETIME из базы (UTC) -> день по UTC+5, или None."""
    if not value:
        return None
    try:
        moment = datetime.datetime.fromisoformat(str(value)[:19])
    except ValueError:
        return None
    return (moment + datetime.timedelta(hours=UTC_OFFSET_HOURS)).date()


def load_contours(conn, month):
    """Контуры DJI, созданные в этом месяце (по дню UTC+5)."""
    rows = conn.execute(
        'SELECT id, name, serial_number, area_ha, source_created_at '
        'FROM field_contours WHERE source = ?', (SOURCE_DJI,)).fetchall()
    out = []
    for cid, name, serial, area, created in rows:
        day = local_day(created)
        if day is None or day.strftime('%Y-%m') != month:
            continue
        out.append({'id': cid, 'name': name or '', 'serial': serial,
                    'area_ha': float(area or 0), 'day': day})
    return out, len(rows)


def load_books(conn, month):
    """Строки книг этого месяца с написанием заказчика и оператора."""
    rows = conn.execute(
        'SELECT w.id, w.customer_raw, c.name, w.area_ha, w.work_date_from, '
        '       w.work_date_to, w.date_raw, o.full_name, w.operator_raw, '
        '       w.source_file, w.source_sheet '
        'FROM drone_works w '
        'LEFT JOIN drone_customers c ON c.id = w.drone_customer_id '
        'LEFT JOIN drone_operators o ON o.id = w.drone_operator_id '
        'WHERE w.period_month = ? ORDER BY w.id', (month,)).fetchall()
    out = []
    for (wid, raw, canonical, area, dfrom, dto, draw, op, op_raw,
         src_file, src_sheet) in rows:
        out.append({
            'id': wid,
            # [REASON]: сравниваем ПО ОБОИМ написаниям -- verbatim из строки и
            # каноническое из справочника. Диспетчер пишет «Чинор фх», а
            # справочник мог собрать «Чинор фермер хужалиги»; выбрать одно
            # значит потерять половину совпадений.
            'customer_raw': raw or '',
            'customer': canonical or raw or '',
            'area_ha': float(area or 0),
            'date_from': dfrom, 'date_to': dto, 'date_raw': draw,
            'operator': op or op_raw or '',
            'source': '%s / %s' % (src_file or '', src_sheet or ''),
        })
    return out


def best_for_book(book, contours):
    """Лучшая оценка книги против каждого контура: [(оценка, причина, контур)].

    Возвращает ВСЕ контуры с оценкой не ниже SCORE_CANDIDATE, отсортированные
    по убыванию. Отбор «лучший один» здесь не делается: у одного заказчика
    законно бывает несколько контуров -- у Аваза Исматова их пять.
    """
    found = []
    for contour in contours:
        best, why = 0.0, ''
        for spelling in (book['customer'], book['customer_raw']):
            if not spelling:
                continue
            value, reason = score(spelling, contour['name'])
            if value > best:
                best, why = value, reason
        if best >= SCORE_CANDIDATE:
            found.append((best, why, contour))
    found.sort(key=lambda item: (-item[0], -item[2]['area_ha']))
    return found


def build(conn, month):
    contours, contours_total = load_contours(conn, month)
    books = load_books(conn, month)

    matched, candidates, book_only = [], [], []
    used_contour_ids = set()

    for book in books:
        found = best_for_book(book, contours)
        confident = [item for item in found if item[0] >= SCORE_CONFIDENT]
        if confident:
            area = sum(item[2]['area_ha'] for item in confident)
            matched.append((book, confident, area))
            used_contour_ids.update(item[2]['id'] for item in confident)
            continue
        if found:
            candidates.append((book, found[:5]))
            continue
        book_only.append(book)

    contour_only = [c for c in contours if c['id'] not in used_contour_ids]

    stats = {
        'contours_all': contours_total,
        'contours_month': len(contours),
        'contours_area': sum(c['area_ha'] for c in contours),
        'books': len(books),
        'books_area': sum(b['area_ha'] for b in books),
        'matched': len(matched),
        'matched_book_area': sum(m[0]['area_ha'] for m in matched),
        'matched_contour_area': sum(m[2] for m in matched),
        'candidates': len(candidates),
        'candidates_area': sum(c[0]['area_ha'] for c in candidates),
        'book_only': len(book_only),
        'book_only_area': sum(b['area_ha'] for b in book_only),
        'contour_only': len(contour_only),
        'contour_only_area': sum(c['area_ha'] for c in contour_only),
    }
    return stats, matched, candidates, book_only, contour_only


def _dates(book):
    if book['date_from'] and book['date_to'] and book['date_from'] != book['date_to']:
        return '%s .. %s' % (book['date_from'], book['date_to'])
    return str(book['date_from'] or book['date_raw'] or '')


def write_xlsx(path, month, stats, matched, candidates, book_only,
               contour_only):
    from openpyxl import Workbook

    book = Workbook()
    pct = lambda part, whole: (part / whole * 100.0) if whole else 0.0

    sheet = book.active
    sheet.title = SHEET_SUMMARY
    for row in [
        ('Месяц', month),
        ('', ''),
        ('Контуров DJI всего в базе', stats['contours_all']),
        ('  создано в этом месяце', stats['contours_month']),
        ('  их площадь, га', round(stats['contours_area'], 2)),
        ('', ''),
        ('Строк книг за месяц', stats['books']),
        ('  их гектары', round(stats['books_area'], 2)),
        ('', ''),
        ('Строк книг с уверенным совпадением', stats['matched']),
        ('  гектары этих строк по книге', round(stats['matched_book_area'], 2)),
        ('  площадь найденных контуров', round(stats['matched_contour_area'], 2)),
        ('Строк книг с кандидатом (нужен глаз)', stats['candidates']),
        ('  их гектары', round(stats['candidates_area'], 2)),
        ('Строк книг без контура', stats['book_only']),
        ('  их гектары', round(stats['book_only_area'], 2)),
        ('', ''),
        ('Контуров, не отданных ни одной строке книги', stats['contour_only']),
        ('  их площадь, га', round(stats['contour_only_area'], 2)),
        ('', ''),
        ('Доля строк книг, нашедших контур, %',
         round(pct(stats['matched'], stats['books']), 1)),
        ('Доля гектаров книг, нашедших контур, %',
         round(pct(stats['matched_book_area'], stats['books_area']), 1)),
    ]:
        sheet.append(list(row))
    sheet.column_dimensions['A'].width = 44
    sheet.column_dimensions['B'].width = 18

    sheet = book.create_sheet(SHEET_MATCHED)
    sheet.append(['Заказчик (книга)', 'Оператор', 'Га по книге', 'Даты книги',
                  'Контуров', 'Площадь контуров, га', 'Имена контуров',
                  'Номера DJI', 'Дни контуров', 'Оценка', 'Почему совпало',
                  'Книга: файл / лист'])
    for bk, found, area in sorted(matched, key=lambda m: -m[0]['area_ha']):
        sheet.append([
            bk['customer_raw'], bk['operator'], round(bk['area_ha'], 2),
            _dates(bk), len(found), round(area, 2),
            ' | '.join(item[2]['name'] for item in found),
            ' | '.join(str(item[2]['serial'] or '') for item in found),
            ', '.join(sorted({str(item[2]['day']) for item in found})),
            round(found[0][0], 2), found[0][1], bk['source'],
        ])
    for column, width in (('A', 30), ('B', 22), ('G', 46), ('I', 24),
                          ('K', 34), ('L', 34)):
        sheet.column_dimensions[column].width = width

    sheet = book.create_sheet(SHEET_CANDIDATES)
    sheet.append(['Заказчик (книга)', 'Оператор', 'Га по книге', 'Даты книги',
                  'Кандидат-контур', 'Площадь контура, га', 'День контура',
                  'Оценка', 'Почему показан'])
    for bk, found in sorted(candidates, key=lambda c: -c[0]['area_ha']):
        for value, why, contour in found:
            sheet.append([bk['customer_raw'], bk['operator'],
                          round(bk['area_ha'], 2), _dates(bk),
                          contour['name'], round(contour['area_ha'], 2),
                          str(contour['day']), round(value, 2), why])
    for column, width in (('A', 30), ('B', 22), ('E', 30), ('I', 40)):
        sheet.column_dimensions[column].width = width

    sheet = book.create_sheet(SHEET_BOOK_ONLY)
    sheet.append(['Заказчик (книга)', 'Оператор', 'Га по книге', 'Даты книги',
                  'Книга: файл / лист'])
    for bk in sorted(book_only, key=lambda b: -b['area_ha']):
        sheet.append([bk['customer_raw'], bk['operator'],
                      round(bk['area_ha'], 2), _dates(bk), bk['source']])
    for column, width in (('A', 32), ('B', 22), ('E', 38)):
        sheet.column_dimensions[column].width = width

    sheet = book.create_sheet(SHEET_CONTOUR_ONLY)
    sheet.append(['Контур DJI', 'Номер DJI', 'Площадь, га', 'День (UTC+5)'])
    for contour in sorted(contour_only, key=lambda c: (c['day'], -c['area_ha'])):
        sheet.append([contour['name'], contour['serial'],
                      round(contour['area_ha'], 2), str(contour['day'])])
    sheet.column_dimensions['A'].width = 34

    book.save(path)
    return path


def print_summary(stats):
    """ASCII only -- Windows console encoding."""
    pct = lambda part, whole: (part / whole * 100.0) if whole else 0.0
    print('contours in the directory   : %d' % stats['contours_all'])
    print('  created in this month     : %d (%.2f ha)'
          % (stats['contours_month'], stats['contours_area']))
    print('book rows in this month     : %d (%.2f ha)'
          % (stats['books'], stats['books_area']))
    print('  matched to a contour      : %d (%.1f%%), %.2f ha of book'
          % (stats['matched'], pct(stats['matched'], stats['books']),
             stats['matched_book_area']))
    print('  candidate, needs a human  : %d (%.2f ha)'
          % (stats['candidates'], stats['candidates_area']))
    print('  no contour at all         : %d (%.2f ha)'
          % (stats['book_only'], stats['book_only_area']))
    print('contours with no book row   : %d (%.2f ha)'
          % (stats['contour_only'], stats['contour_only_area']))


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Bridge dispatcher book customers to DJI field contours '
                    'by name. Read-only.')
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'))
    parser.add_argument('--month', required=True, metavar='YYYY-MM')
    parser.add_argument('--out', help='write an xlsx report to this path')
    args = parser.parse_args(argv)

    try:
        datetime.datetime.strptime(args.month, '%Y-%m')
    except ValueError:
        print('ERROR: --month must be YYYY-MM, got %r' % args.month)
        sys.exit(1)

    conn = open_readonly(args.db)
    try:
        stats, matched, candidates, book_only, contour_only = build(
            conn, args.month)
    finally:
        conn.close()

    print_summary(stats)
    if args.out:
        try:
            write_xlsx(args.out, args.month, stats, matched, candidates,
                       book_only, contour_only)
        except ImportError:
            print('ERROR: openpyxl is not installed; the console summary '
                  'above is complete, the xlsx was not written.')
            return 1
        print('report written: %s' % args.out)
    return 0


if __name__ == '__main__':
    sys.exit(main())
