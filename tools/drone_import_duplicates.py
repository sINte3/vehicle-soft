#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_import_duplicates.py -- DRONE-IMPORT-DUP-001, только чтение.

Одна и та же книга диспетчера, загруженная ДВАЖДЫ, удваивает гектары и
деньги -- и импортёр этого не замечает.

КАК ЭТО ВЫГЛЯДИТ. Уникальность строки в drone_works -- это тройка
(source_file, source_sheet, source_row). Если ту же книгу загрузить вторым
файлом, у которого браузер приписал к имени « (2)», тройка становится другой,
и обе загрузки живут рядом. В сентябре 2025 это дало +118.40 га:

    Сервис Дрон Маълумот (2).xlsx  свод ичи (Мухриддин)  11 строк  205.00
    Сервис Дрон Маълумот.xlsx      свод ичи (Мухриддин)   6 строк   93.60  <-
    Сервис Дрон Маълумот (2).xlsx  свод ичи (Фурқат)     37 строк  398.90
    Сервис Дрон Маълумот.xlsx      свод ичи (Фурқат)      9 строк   24.80  <-

Подписанные ОБЩИЕ СВОДЫ дают Кудратову 205.00 и Жумаеву 398.90 -- ровно то,
что лежит в файле «(2)». Строки из файла без суффикса -- остаток первой,
неполной загрузки.

ПОПРАВКА 2026-08-20, ПОСЛЕ БОЕВОГО ПРОГОНА. Пример выше НЕВЕРЕН, и по нему
на production удалили живые данные. Строки без суффикса « (2)» -- это НАЛИЧНЫЕ
блоки ОКТЯБРЬСКОЙ книги Сервиса: 9 строк Жумаева на 24.80 га и 6 строк
Кудратова на 93.60 га стоят в октябрьском файле ровно этими числами, и
подписанный октябрьский ОБЩИЙ СВОД подтверждает обе (22.3 + 24.8 = 47.1;
3.4 + 93.6 = 97.0). В сентябрь они попали ПОТОМУ ЧТО У НИХ НЕТ ДАТЫ: месяц
строки -- COALESCE(месяц даты, period_month), а period_month берётся из формы
загрузки, и книгу октября загрузили с периодом сентября. Датированные строки
того же файла легли в октябрь и лежат там до сих пор.

Отчёт этого не различал и предложил DELETE. Разбор -- в
docs/DRONES_OCT2025_RECONCILIATION.md.

ЧТО ДЕЛАЕТ ОТЧЁТ. По каждому месяцу ищет ЛИСТЫ С ОДНИМ ИМЕНЕМ, пришедшие из
РАЗНЫХ файлов. Это и есть след двойной загрузки: один и тот же лист книги не
может законно прийти из двух файлов.

[REASON]: сравнение идёт по ЛИСТУ, а не по имени файла. Имена файлов у
подразделений разные и меняются от месяца к месяцу, а имя листа -- «свод ичи
(Мухриддин)» -- это название книги одного человека, и оно устойчиво. Сравнение
по «похожим именам файлов» пропустило бы переименованную копию и подняло бы
ложную тревогу на честно разных книгах.

ЧЕГО ОТЧЁТ НЕ ДЕЛАЕТ. Он НИЧЕГО НЕ УДАЛЯЕТ и не может: база открывается
read-only. Устав проекта запрещает удалять продовые данные автоматически.
Отчёт печатает готовый SELECT, чтобы владелец посмотрел строки глазами, и
готовый DELETE, чтобы он выполнил его сам, своей рукой, после резервной копии.

Запуск (сервер, служба может работать):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_import_duplicates.py --db instance\\transport.db

Выход -- в консоль (ASCII) и, при --out, в xlsx.
Код возврата 1, если найдена хотя бы одна двойная загрузка.
"""

import argparse
import os
import sqlite3
import sys
from collections import defaultdict


def _ascii(text):
    return str(text).encode('ascii', 'replace').decode('ascii')


def connect_ro(db_path):
    uri = 'file:%s?mode=ro' % db_path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def load(conn, month=None):
    """Строки по (месяц, лист, файл). Месяц -- по правилу отчётов.

    Последняя колонка -- сколько строк группы БЕЗ ДАТЫ. Такая строка лежит в
    месяце не по своей дате, а по period_month из формы загрузки; это половина
    признака «книга другого месяца», вторая половина -- dated_months_by_file().
    """
    where = ''
    params = ()
    if month:
        where = ("WHERE COALESCE(strftime('%Y-%m', w.work_date_from), "
                 'w.period_month) = ? ')
        params = (month,)
    return conn.execute(
        "SELECT COALESCE(strftime('%Y-%m', w.work_date_from), "
        "                w.period_month) AS m, "
        "       COALESCE(w.source_sheet, ''), COALESCE(w.source_file, ''), "
        '       COUNT(*), COALESCE(SUM(w.area_ha), 0), '
        "       COALESCE(SUM(w.amount), 0), COALESCE(o.full_name, ''), "
        '       SUM(CASE WHEN w.work_date_from IS NULL THEN 1 ELSE 0 END) '
        'FROM drone_works w '
        'LEFT JOIN drone_operators o ON o.id = w.drone_operator_id ' + where +
        'GROUP BY 1, 2, 3, 7 ORDER BY 1, 2, 3', params).fetchall()


def dated_months_by_file(conn):
    """{файл: множество месяцев его ДАТИРОВАННЫХ строк}.

    [REASON]: считается по ВСЕЙ таблице и никогда не сужается фильтром
    --month. Доказательство «этот файл -- книга другого месяца» лежит именно
    в строках другого месяца, и фильтр отрезал бы их ровно тогда, когда они
    нужны: при прогоне `--month 2025-09` октябрьские строки октябрьского файла
    в выборку не попадают, а без них сентябрьский остаток от него неотличим
    от второй загрузки сентябрьской книги.
    """
    out = defaultdict(set)
    for source_file, dated_month in conn.execute(
            "SELECT source_file, strftime('%Y-%m', work_date_from) "
            'FROM drone_works '
            'WHERE work_date_from IS NOT NULL AND source_file IS NOT NULL '
            'GROUP BY 1, 2'):
        if source_file and dated_month:
            out[source_file].add(dated_month)
    return dict(out)


NAMESAKE = 'namesake'
CANDIDATE = 'candidate'
UNKNOWN = 'unknown'
OTHER_MONTH = 'other_month'


def find_duplicates(rows, file_dated_months):
    """Листы с одним именем из разных файлов -- и КТО за ними стоит.

    Возвращает {(месяц, лист): {'files': [...], 'kind': ...}}, где вид --
    один из OTHER_MONTH / NAMESAKE / CANDIDATE / UNKNOWN.

    file_dated_months -- результат dated_months_by_file(), ОБЯЗАТЕЛЕН.
    [REASON]: параметр со значением по умолчанию молча вернул бы поведение,
    из-за которого на production удалили 118.40 га живых работ: забытый
    аргумент не отличался бы от пустой базы. Пусть лучше падает.
    """
    by_sheet = defaultdict(lambda: defaultdict(
        lambda: {'rows': 0, 'ha': 0.0, 'amount': 0.0, 'operators': set(),
                 'dateless': 0}))
    for row in rows:
        (month, sheet, source_file, count, hectares, amount,
         operator) = row[:7]
        dateless = int(row[7]) if len(row) > 7 and row[7] is not None else 0
        # [REASON]: строка, набранная руками, не имеет ни файла, ни листа.
        # Она законно «ниоткуда», и складывать такие в одну группу значит
        # объявить двойной загрузкой всю ручную правку месяца.
        if not sheet or not source_file:
            continue
        agg = by_sheet[(month, sheet)][source_file]
        agg['rows'] += count
        agg['ha'] += float(hectares)
        agg['amount'] += float(amount)
        agg['dateless'] += dateless
        if operator:
            agg['operators'].add(operator)

    out = {}
    for key, files in by_sheet.items():
        if len(files) < 2:
            continue
        month = key[0]
        # [REASON]: КНИГА ДРУГОГО МЕСЯЦА, А НЕ ВТОРАЯ ЗАГРУЗКА. Месяц строки
        # -- COALESCE(месяц даты, period_month), а period_month приходит из
        # формы загрузки и относится ко всей партии. Загрузили книгу октября,
        # указав в форме сентябрь, -- её ДАТИРОВАННЫЕ строки ушли в октябрь по
        # своим датам, а строки БЕЗ ДАТЫ остались в сентябре и встали рядом с
        # сентябрьской книгой того же человека. Выглядит в точности как
        # двойная загрузка. По этому сходству 2026-08-20 на production удалили
        # 15 живых октябрьских работ на 118.40 га.
        # Признак из двух половин, и обе обязательны: (1) на этой стороне
        # НЕТ НИ ОДНОЙ датированной строки -- значит она стоит в месяце только
        # по period_month; (2) у того же ФАЙЛА есть датированные строки в
        # ДРУГОМ месяце -- значит его собственная книга не здесь. Одной первой
        # мало: у настоящей неполной загрузки даты тоже могут отсутствовать.
        strangers = []
        undated_sides = []
        for source_file, agg in files.items():
            if agg['dateless'] != agg['rows']:
                continue
            undated_sides.append(source_file)
            elsewhere = sorted(set(file_dated_months.get(source_file, ()))
                               - {month})
            if elsewhere:
                strangers.append((source_file, elsewhere))
        if strangers:
            out[key] = {'files': sorted(
                (name, agg['rows'], agg['ha'], agg['amount'],
                 ', '.join(sorted(agg['operators'])) or '(без оператора)')
                for name, agg in files.items()),
                'kind': OTHER_MONTH,
                'strangers': sorted(strangers)}
            continue
        # [REASON]: ЛИСТ-ТЁЗКА, А НЕ ДУБЛЬ. «свод ичи (Шохрух)» есть и в
        # книге Гардена (Файзуллаев Шохрух), и в книге Когона (Хамроев
        # Шохрух) -- это книги ДВУХ РАЗНЫХ ЛЮДЕЙ с одинаковым именем листа.
        # Первая редакция отчёта объявила гарденскую книгу лишней и
        # предложила удалить 154.70 га настоящих работ. Отличает их
        # оператор: у двойной загрузки он один и тот же по обе стороны.
        operators = [tuple(sorted(agg['operators'])) for agg in files.values()]
        # [REASON]: сторона БЕЗ ОПЕРАТОРА не свидетельствует ни за, ни против.
        # В апреле и мае 2026 «свод ичи (Шахзод)» приходит из книги Ғиждувона
        # (Холмуродов) и из книги Сервиса, где оператор не привязан вовсе.
        # Первая редакция считала «нет оператора» за «оператор тот же» и
        # предлагала удалить 492.15 га как дубль. Молчание -- не согласие:
        # такие группы идут в UNKNOWN, и DELETE по ним не печатается.
        reason = None
        if any(not op for op in operators):
            kind = UNKNOWN
            reason = 'operator'
        elif len(set(operators)) <= 1:
            kind = CANDIDATE
        else:
            kind = NAMESAKE
        # [REASON]: сторона, где НЕТ НИ ОДНОЙ даты, стоит в этом месяце не по
        # собственному свидетельству, а по period_month формы загрузки. Даже
        # когда доказать «книга другого месяца» нечем -- датированных строк
        # того файла нет вообще нигде, -- предлагать по ней DELETE значит
        # верить молчанию. Ровно та же логика, по которой сторона без
        # оператора уходит в UNKNOWN. Различить эти два случая -- работа
        # владельца по подписанному своду, а не отчёта.
        if kind == CANDIDATE and undated_sides:
            kind = UNKNOWN
            reason = 'undated'
        out[key] = {'reason': reason, 'files': sorted(
            (name, agg['rows'], agg['ha'], agg['amount'],
             ', '.join(sorted(agg['operators'])) or '(без оператора)')
            for name, agg in files.items()),
            'kind': kind}
    return out


# Один словарь на весь отчёт.
# [REASON]: вердикт жил в двух местах -- в консоли и в write_xlsx, -- и в
# xlsx условие было двоичным: всё, что не CANDIDATE, подписывалось «ТЁЗКИ,
# РАЗНЫЕ ЛЮДИ». UNKNOWN уже подписывался неверно, и OTHER_MONTH подписался бы
# так же. Тот же класс дефекта, что и разъехавшийся порядок столбцов листа
# «По операторам» в drone_money_audit.
VERDICT_CONSOLE = {
    CANDIDATE: 'CANDIDATE double upload',
    NAMESAKE: 'NAMESAKE sheets -- DIFFERENT PEOPLE, not a duplicate',
    (UNKNOWN, 'operator'): 'UNKNOWN -- one side has no operator, cannot tell; '
                           'no SQL offered',
    (UNKNOWN, 'undated'): 'UNKNOWN -- one side has no dated row at all, its '
                          'month is the form period; no SQL offered',
    OTHER_MONTH: 'ANOTHER MONTH\'S BOOK -- undated rows filed by '
                 'period_month; DO NOT DELETE',
}

VERDICT_XLSX = {
    CANDIDATE: 'двойная загрузка?',
    NAMESAKE: 'ТЁЗКИ, РАЗНЫЕ ЛЮДИ -- не дубль',
    (UNKNOWN, 'operator'): 'НЕ ОПОЗНАНО -- у стороны нет оператора',
    (UNKNOWN, 'undated'): 'НЕ ОПОЗНАНО -- у стороны нет ни одной даты',
    OTHER_MONTH: 'КНИГА ДРУГОГО МЕСЯЦА -- не дубль, не удалять',
}


def verdict(group, table):
    """Подпись группы. UNKNOWN различается по причине, остальные -- по виду."""
    if group['kind'] == UNKNOWN:
        return table[(UNKNOWN, group.get('reason') or 'operator')]
    return table[group['kind']]


def report_lines(duplicates):
    lines = []
    for (month, sheet), group in sorted(duplicates.items()):
        lines.append('%s  %s   [%s]'
                     % (month, _ascii(sheet), verdict(group, VERDICT_CONSOLE)))
        for source_file, count, hectares, amount, operator in group['files']:
            lines.append('    %-40s %4d rows %9.2f ha %14.0f  %s'
                         % (_ascii(source_file)[-40:], count, hectares,
                            amount, _ascii(operator)[:26]))
        for source_file, elsewhere in group.get('strangers', ()):
            lines.append('    ^ %s : no dated row here, dated rows in %s'
                         % (_ascii(source_file)[-40:], ', '.join(elsewhere)))
    return lines


def candidates(duplicates):
    """Только группы, где по обе стороны ОДИН И ТОТ ЖЕ оператор.

    [REASON]: отчёт НЕ выбирает, какой из двух файлов лишний, и не пытается.
    Первая редакция считала полной загрузку с наибольшим числом строк -- на
    апреле 2026 это назвало бы лишним файл с 361.30 га в пользу файла с
    492.15, а на мае -- файл с 548.10 га в пользу файла с 226.51. Число
    строк не говорит, какая загрузка верна; говорит подписанный ОБЩИЙ СВОД,
    и смотрит в него человек. Поэтому печатаются ОБА варианта как
    альтернативы, и выбрать надо ровно один.
    """
    out = []
    for (month, sheet), group in sorted(duplicates.items()):
        if group['kind'] != CANDIDATE:
            continue
        for source_file, count, hectares, amount, operator in group['files']:
            out.append((month, sheet, source_file, count, hectares, amount,
                        operator))
    return out


def of_kind(duplicates, kind):
    """Группы одного вида: тёзки или неопознанные."""
    return [(month, sheet, duplicates[(month, sheet)]['files'])
            for (month, sheet) in sorted(duplicates)
            if duplicates[(month, sheet)]['kind'] == kind]


def namesakes(duplicates):
    """Группы, где за одинаковым именем листа стоят РАЗНЫЕ люди."""
    return of_kind(duplicates, NAMESAKE)


def unknowns(duplicates):
    """Группы, где хотя бы у одной стороны оператор не привязан."""
    return of_kind(duplicates, UNKNOWN)


def other_months(duplicates):
    """Группы, где одна сторона -- книга ДРУГОГО месяца.

    Возвращает (месяц, лист, files, strangers): последнее -- список
    (файл, [месяцы его датированных строк]), то есть само доказательство.
    """
    return [(month, sheet, duplicates[(month, sheet)]['files'],
             duplicates[(month, sheet)].get('strangers', []))
            for (month, sheet) in sorted(duplicates)
            if duplicates[(month, sheet)]['kind'] == OTHER_MONTH]


def sql_lines(items):
    lines = []
    for month, sheet, source_file, _count, _ha, _amount, _op in items:
        where = ("WHERE source_file = '%s' AND source_sheet = '%s' "
                 "AND COALESCE(strftime('%%Y-%%m', work_date_from), "
                 "period_month) = '%s'"
                 % (source_file.replace("'", "''"),
                    sheet.replace("'", "''"), month))
        lines.append(('SELECT id, customer_raw, area_ha, amount, source_row '
                      'FROM drone_works %s;' % where,
                      'DELETE FROM drone_works %s;' % where))
    return lines


def write_xlsx(path, rows, duplicates):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    ws = wb.active
    ws.title = 'Двойные загрузки'
    ws.append(['Месяц', 'Лист-источник', 'Файл-источник', 'Оператор',
               'Строк', 'Га', 'Сумма', 'Вердикт', 'Даты этого файла в'])
    for cell in ws[1]:
        cell.font = bold
    for (month, sheet), group in sorted(duplicates.items()):
        label = verdict(group, VERDICT_XLSX)
        elsewhere = dict(group.get('strangers', ()))
        for source_file, count, hectares, amount, operator in group['files']:
            ws.append([month, sheet, source_file, operator, count,
                       round(hectares, 2), round(amount, 2), label,
                       ', '.join(elsewhere.get(source_file, ()))])
    ws = wb.create_sheet('Все источники')
    ws.append(['Месяц', 'Лист-источник', 'Файл-источник', 'Оператор', 'Строк',
               'Га', 'Сумма', 'Из них без даты'])
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        (month, sheet, source_file, count, hectares, amount,
         operator) = row[:7]
        dateless = int(row[7]) if len(row) > 7 and row[7] is not None else 0
        ws.append([month, sheet, source_file, operator, count,
                   round(float(hectares), 2), round(float(amount), 2),
                   dateless])
    for sheet_obj in wb.worksheets:
        for column, width in (('A', 10), ('B', 26), ('C', 46), ('D', 24),
                              ('E', 8), ('F', 11), ('G', 15), ('H', 34),
                              ('I', 20)):
            sheet_obj.column_dimensions[column].width = width
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(
        description='Find dispatcher books imported twice. Read-only.')
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'))
    parser.add_argument('--month', default=None,
                        help='YYYY-MM; omit to scan every month')
    parser.add_argument('--out', default=None)
    args = parser.parse_args()

    if not os.path.isfile(args.db):
        print('ERROR: database not found: %s' % _ascii(args.db))
        return 2

    conn = connect_ro(args.db)
    try:
        rows = load(conn, args.month)
        file_dated_months = dated_months_by_file(conn)
    finally:
        conn.close()

    duplicates = find_duplicates(rows, file_dated_months)
    print('Months scanned      : %s' % (args.month or 'all'))
    print('Source groups       : %d' % len(rows))
    print('Sheets from 2+ files: %d' % len(duplicates))
    if not duplicates:
        print('')
        print('No book looks imported twice.')
        if args.out:
            write_xlsx(args.out, rows, duplicates)
            print('Written: %s' % _ascii(args.out))
        return 0

    items = candidates(duplicates)
    twins = namesakes(duplicates)
    murky = unknowns(duplicates)
    strays = other_months(duplicates)
    print('Candidate duplicates: %d sheet(s), %d file(s)'
          % (len({(i[0], i[1]) for i in items}), len(items)))
    print('Namesake sheets     : %d  (different people -- NOT duplicates)'
          % len(twins))
    print('Undecidable         : %d  (one side has no operator -- no SQL)'
          % len(murky))
    print("Another month's book: %d  (undated rows filed by period_month "
          '-- no SQL)' % len(strays))
    print('')
    for line in report_lines(duplicates):
        print(line)
    # [REASON]: имена листов и файлов -- кириллица, а консоль Windows её
    # калечит в «?????». Скопированный оттуда SQL не нашёл бы НИ ОДНОЙ строки
    # и выглядел бы при этом безобидно: «удалено 0 строк». Поэтому SQL пишется
    # в UTF-8 файл, а в консоль идёт только путь к нему -- ровно то, что
    # предписывает устав про скрипты с кириллическим выводом.
    sql_path = os.path.join(os.path.dirname(os.path.abspath(args.db)),
                            'drone_import_duplicates.sql')
    with open(sql_path, 'w', encoding='utf-8') as handle:
        handle.write('-- Сформировано tools/drone_import_duplicates.py\n')
        # [REASON]: шапка описывает то, что НИЖЕ, а не то, что бывает вообще.
        # Прежняя редакция всегда обещала «ниже ДВА варианта, выполнить надо
        # РОВНО ОДИН» -- даже когда кандидатов нет вовсе и в файле одни
        # предупреждения. Файл, зовущий выполнить то, чего в нём нет, толкает
        # искать DELETE и раскомментировать первое похожее.
        if items:
            handle.write('-- ОТЧЁТ НЕ ЗНАЕТ, какая из двух загрузок верна: '
                         'это\n')
            handle.write('-- говорит подписанный ОБЩИЙ СВОД. По каждому '
                         'листу\n')
            handle.write('-- ниже ДВА варианта -- выполнить надо РОВНО '
                         'ОДИН,\n')
            handle.write('-- и только после копии instance\\transport.db.\n')
        else:
            handle.write('-- ВЫПОЛНЯТЬ НЕЧЕГО: ни одной двойной загрузки не '
                         'найдено.\n')
            handle.write('-- Ниже только предупреждения; DELETE в этом файле '
                         'нет.\n')
        for month, sheet, group in [(m, sh, duplicates[(m, sh)])
                                    for (m, sh) in sorted(duplicates)
                                    if duplicates[(m, sh)]['kind']
                                    == CANDIDATE]:
            handle.write('\n\n-- ================================\n')
            handle.write('-- %s  %s\n' % (month, sheet))
            for source_file, count, hectares, amount, operator in \
                    group['files']:
                handle.write('--   %s : %d строк, %.2f га, %.0f сум, %s\n'
                             % (source_file, count, hectares, amount,
                                operator))
            pairs = sql_lines([(month, sheet, f[0], f[1], f[2], f[3], f[4])
                               for f in group['files']])
            for (source_file, count, hectares, _a, _o), (select, delete) in \
                    zip(group['files'], pairs):
                handle.write('\n-- ВАРИАНТ: убрать %s (%d строк, %.2f га)\n'
                             % (source_file, count, hectares))
                handle.write('-- сначала посмотреть:\n')
                handle.write(select + '\n')
                handle.write('-- и только потом, если это лишняя загрузка:\n')
                handle.write('-- ' + delete + '\n')
        for month, sheet, files in twins:
            handle.write('\n\n-- ================================\n')
            handle.write('-- %s  %s -- ТЁЗКИ, НЕ ДУБЛЬ. Ничего не удалять.\n'
                         % (month, sheet))
            for source_file, count, hectares, _amount, operator in files:
                handle.write('--   %s : %d строк, %.2f га, %s\n'
                             % (source_file, count, hectares, operator))
        for month, sheet, files in murky:
            handle.write('\n\n-- ================================\n')
            if duplicates[(month, sheet)].get('reason') == 'undated':
                handle.write(
                    '-- %s  %s -- НЕ ОПОЗНАНО. У одной стороны нет НИ ОДНОЙ\n'
                    '-- датированной строки: в этом месяце она стоит по '
                    'period_month из формы\n-- загрузки, а не по своей дате. '
                    'SQL не предлагается: какая из двух\n-- загрузок лишняя '
                    'и лишняя ли вообще -- говорит подписанный ОБЩИЙ СВОД.\n'
                    % (month, sheet))
            else:
                handle.write(
                    '-- %s  %s -- НЕ ОПОЗНАНО. У одной стороны оператор '
                    'не привязан,\n-- и отличить двойную загрузку от '
                    'листа-тёзки нечем. SQL не предлагается:\n'
                    '-- сначала привязать оператора, потом запустить '
                    'отчёт заново.\n' % (month, sheet))
            for source_file, count, hectares, _amount, operator in files:
                handle.write('--   %s : %d строк, %.2f га, %s\n'
                             % (source_file, count, hectares, operator))
        for month, sheet, files, strangers in strays:
            handle.write('\n\n-- ================================\n')
            handle.write('-- %s  %s -- КНИГА ДРУГОГО МЕСЯЦА, НЕ ДУБЛЬ.\n'
                         '-- Ничего не удалять.\n' % (month, sheet))
            for source_file, count, hectares, _amount, operator in files:
                handle.write('--   %s : %d строк, %.2f га, %s\n'
                             % (source_file, count, hectares, operator))
            for source_file, elsewhere in strangers:
                handle.write('--   ^ у «%s» здесь нет НИ ОДНОЙ датированной '
                             'строки,\n--     а датированные строки того же '
                             'файла лежат в %s.\n'
                             % (source_file, ', '.join(elsewhere)))
            handle.write('-- Значит месяц этих строк дал period_month из '
                         'формы загрузки, а не их\n-- собственная дата: книгу '
                         'загрузили, указав чужой период. Лечится\n'
                         '-- ПЕРЕНОСОМ строк в свой месяц, а не удалением. '
                         'Порядок --\n-- docs/DRONES_OCT2025_RECONCILIATION.md.'
                         '\n')

    print('')
    print('NOTHING WAS DELETED -- this report cannot write to the database.')
    print('IT ALSO DOES NOT DECIDE which upload is the surplus one: the row '
          'count does not say that, the signed summary does.')
    print('The SQL is in a UTF-8 file (sheet names are Cyrillic and the '
          'console would mangle them):')
    print('  %s' % _ascii(sql_path))
    print('For each candidate it offers TWO alternatives -- run exactly one, '
          'after copying instance\\transport.db aside.')
    print('Every DELETE is written out COMMENTED. Uncomment the one you '
          'chose.')
    if strays:
        print('')
        print("ANOTHER MONTH'S BOOK is NOT a duplicate and has NO DELETE at "
              'all, not even a commented one:')
        print('  those rows carry no date, so the month came from the upload '
              "form's period, while the")
        print('  same file has dated rows in another month. They are moved, '
              'never deleted.')
    if args.out:
        write_xlsx(args.out, rows, duplicates)
        print('')
        print('Written: %s' % _ascii(args.out))
    return 1


if __name__ == '__main__':
    sys.exit(main())
