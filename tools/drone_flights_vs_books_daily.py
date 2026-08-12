#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_flights_vs_books_daily.py -- DRONE-FLIGHTS-VS-BOOKS-DAILY-001:
посуточная сверка телеметрии DJI с ведомостями диспетчеров.

ЗАЧЕМ. Экран «Ведомости против вылетов» сравнивает месяцы; владелец увидел
за март--июнь 2026 разницу в гектарах и попросил доскональную расшифровку:
по дням, с указанием файла, листа и строки каждой работы. Этот инструмент --
та расшифровка. Он НИЧЕГО НЕ РЕШАЕТ и не красит «кто прав»: он кладёт обе
стороны рядом так, чтобы расхождение можно было прочитать пальцем по строке.

ЧТО СЧИТАЕТСЯ. Обе стороны считаются РОВНО как на экранах приложения:

  * день и месяц вылета -- время UTC+5 (константа приложения
    DRONE_DISPLAY_UTC_OFFSET; здесь -- UTC_OFFSET_MINUTES, тест сверяет их);
  * месяц работы -- месяц её собственной даты начала, а без даты -- месяц
    книги (coalesce, как в _drone_work_month_expr);
  * стороны меряют РАЗНОЕ: телеметрия -- всё, что летало, включая свою землю
    и пробные вылеты; ведомости -- только выставленную работу. Совпадать до
    нуля они не обязаны; смотреть надо на дни и книги, где разрыв большой.

УСЛОВНОСТИ ОТОБРАЖЕНИЯ (не правила учёта, в отчёте подписаны):

  * работа с диапазоном дат («23-24.03») показана ЦЕЛИКОМ в дне её начала --
    делить гектары по дням значило бы выдумать правило, которого нет;
  * работа без даты в дневную раскладку не входит вовсе -- такие собраны
    отдельным блоком по месяцам, чтобы их было видно, а не «размазано».

ЛИСТЫ ФАЙЛА:
  «Сводка по месяцам»  -- вылеты/работы/разница/покрытие, строка на месяц;
  «По дням»            -- строка на каждый день периода + блок работ без дат;
  «Работы»             -- каждая работа периода с файлом, листом и строкой;
  «Вылеты по машинам»  -- день х машина (нераспознанные ники -- отдельными
                          строками), чтобы видеть, кто летал в «пустые» дни.

Только чтение: база открывается через file:-URI mode=ro, записи нет нигде.

Запуск (сервер, из каталога установки):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_flights_vs_books_daily.py --db instance\\transport.db

Период по умолчанию 2026-03..2026-06 (запрос владельца 2026-08-11); другой
задаётся --month-from/--month-to. Консоль -- только ASCII; полный отчёт --
xlsx рядом с базой.
"""

import argparse
import calendar
import datetime as dt
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import import_drone_works as importer  # noqa: E402  [REASON]: _connect с
# mode=ro и отказом кодом 2 при отсутствии базы, _require_tables и _ascii --
# те же, что у бэкфиллов и сверки с книгами: один способ открыть базу.

# [REASON]: продублировано из drones.DRONE_DISPLAY_UTC_OFFSET (300 минут).
# Импортировать drones.py отсюда нельзя -- он тянет Flask и модели, а любой
# импорт app-стека из инструмента делает читателя писателем (db.create_all
# на импорте). Тест test_offset_matches_the_application сверяет обе точки:
# разъедутся -- упадёт он, а не отчёт молча.
UTC_OFFSET_MINUTES = 300

DEFAULT_MONTH_FROM = '2026-03'
DEFAULT_MONTH_TO = '2026-06'

NOTE_NO_BOOKS = 'Ведомости не заведены'
NOTE_NO_FLIGHTS = 'Вылетов в этом месяце нет'

DAILY_NOTE = ('Работа с диапазоном дат учтена целиком в дне её начала — это '
              'условность отображения, а не правило учёта. Работы без даты в '
              'дни не входят — они в блоке ниже. День вылета — время UTC+5.')

SUMMARY_NOTE = ('Стороны меряют разное: телеметрия — всё, что летало '
                '(включая свою землю), ведомости — выставленную работу. '
                'Покрытие — гектары ведомостей в процентах от телеметрии; '
                'оно считается только когда обе стороны ненулевые.')


def month_range(month_from, month_to):
    """['2026-03', '2026-04', ...] включительно."""
    y, m = int(month_from[:4]), int(month_from[5:7])
    out = []
    while True:
        cur = '%04d-%02d' % (y, m)
        if cur > month_to:
            break
        out.append(cur)
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return out


def day_range(month_from, month_to):
    """Каждый календарный день периода, объектами date."""
    first = dt.date(int(month_from[:4]), int(month_from[5:7]), 1)
    ly, lm = int(month_to[:4]), int(month_to[5:7])
    last = dt.date(ly, lm, calendar.monthrange(ly, lm)[1])
    days, cur = [], first
    while cur <= last:
        days.append(cur)
        cur += dt.timedelta(days=1)
    return days


def _valid_month(text):
    try:
        dt.datetime.strptime(text, '%Y-%m')
        return True
    except ValueError:
        return False


def collect_flights(con, month_from, month_to):
    """{день: (вылетов, га)} и {(день, ярлык машины): (вылетов, га)}.

    Ярлык распознанной машины -- '№<номер>'; нераспознанной -- 'ник: <ник>'.
    """
    modifier = '+%d minutes' % UTC_OFFSET_MINUTES
    by_day = {}
    by_day_unit = {}
    rows = con.execute(
        "SELECT date(datetime(f.started_at, ?)) AS d, u.number, "
        "       f.nickname_raw, COUNT(*), COALESCE(SUM(f.area_ha), 0.0) "
        "FROM drone_flights f LEFT JOIN drone_units u "
        "     ON u.id = f.drone_unit_id "
        "WHERE strftime('%Y-%m', datetime(f.started_at, ?)) >= ? "
        "  AND strftime('%Y-%m', datetime(f.started_at, ?)) <= ? "
        "GROUP BY d, u.number, "
        "         CASE WHEN u.number IS NULL THEN f.nickname_raw END",
        (modifier, modifier, month_from, modifier, month_to)).fetchall()
    for day, number, nick, count, area in rows:
        if number is not None:
            label = '№%d' % number
        else:
            # [REASON]: нераспознанный ник не сливается в одну кучу -- по
            # написанию видно, ЧТО именно летало без машины в этот день.
            label = 'ник: %s' % (nick or '—')
        c, a = by_day.get(day, (0, 0.0))
        by_day[day] = (c + count, a + float(area))
        by_day_unit[(day, label)] = (count, float(area))
    return by_day, by_day_unit


def collect_works(con, month_from, month_to):
    """Работы периода, месяц -- как в _drone_work_month_expr."""
    rows = con.execute(
        "SELECT id, "
        "       COALESCE(strftime('%Y-%m', work_date_from), period_month), "
        "       work_date_from, work_date_to, date_raw, customer_raw, "
        "       area_ha, subdivision_name, operator_raw, "
        "       source_file, source_sheet, source_row "
        "FROM drone_works "
        "WHERE COALESCE(strftime('%Y-%m', work_date_from), period_month) >= ? "
        "  AND COALESCE(strftime('%Y-%m', work_date_from), period_month) <= ?",
        (month_from, month_to)).fetchall()
    works = []
    for (wid, month, dfrom, dto, draw, customer, area, subdiv, operator,
         sfile, ssheet, srow) in rows:
        dfrom = dfrom[:10] if dfrom else None
        dto = dto[:10] if dto else None
        works.append({
            'id': wid, 'month': month,
            'date_from': dfrom, 'date_to': dto, 'date_raw': draw,
            'customer': customer, 'area': float(area or 0.0),
            'subdivision': subdiv, 'operator': operator,
            'file': sfile, 'sheet': ssheet, 'row': srow,
            'multiday': bool(dfrom and dto and dto != dfrom),
        })
    works.sort(key=lambda w: (w['month'], w['date_from'] or '9999-99-99',
                              w['file'] or '', w['row'] or 0))
    return works


def build_workbook(months, days, fl_day, fl_day_unit, works):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = Workbook()

    def header(ws, row_index, titles, widths):
        for col, (title, width) in enumerate(zip(titles, widths), start=1):
            cell = ws.cell(row=row_index, column=col, value=title)
            cell.font = bold
            ws.column_dimensions[get_column_letter(col)].width = width

    # -- агрегаты работ ----------------------------------------------------
    wk_month = {}          # месяц -> [работ, га]
    wk_month_nodate = {}   # месяц -> [работ, га] (без даты)
    wk_day = {}            # день  -> [работ, га, многодневных, га многодн.]
    wk_day_files = {}      # день  -> набор файлов
    for w in works:
        wm = wk_month.setdefault(w['month'], [0, 0.0])
        wm[0] += 1
        wm[1] += w['area']
        if w['date_from']:
            wd = wk_day.setdefault(w['date_from'], [0, 0.0, 0, 0.0])
            wd[0] += 1
            wd[1] += w['area']
            if w['multiday']:
                wd[2] += 1
                wd[3] += w['area']
            wk_day_files.setdefault(w['date_from'], set()).add(w['file'])
        else:
            nd = wk_month_nodate.setdefault(w['month'], [0, 0.0])
            nd[0] += 1
            nd[1] += w['area']

    fl_month = {}
    for day, (count, area) in fl_day.items():
        fm = fl_month.setdefault(day[:7], [0, 0.0])
        fm[0] += count
        fm[1] += area

    # -- «Сводка по месяцам» ----------------------------------------------
    ws = wb.active
    ws.title = 'Сводка по месяцам'
    ws.cell(row=1, column=1, value=SUMMARY_NOTE)
    header(ws, 2, ('Месяц', 'Вылетов', 'Га по вылетам', 'Работ',
                   'Га по ведомостям', 'Работ без даты', 'Га без даты',
                   'Разница, га (вед. − выл.)', 'Покрытие, %', 'Примечание'),
           (10, 10, 15, 8, 17, 14, 12, 22, 12, 26))
    ws.freeze_panes = 'A3'
    r = 3
    tot = [0, 0.0, 0, 0.0, 0, 0.0]
    for month in months:
        fc, fa = fl_month.get(month, [0, 0.0])
        wc, wa = wk_month.get(month, [0, 0.0])
        nc, na = wk_month_nodate.get(month, [0, 0.0])
        note, coverage = '', None
        # [REASON]: правило процента -- как на экране сверки: месяц без одной
        # из сторон получает слова, а не «-100 %», выдуманные из отсутствия.
        if wa > 0.005 and fa > 0.005:
            coverage = wa * 100.0 / fa
        elif fa > 0.005:
            note = NOTE_NO_BOOKS
        else:
            note = NOTE_NO_FLIGHTS
        for col, value in enumerate(
                (month, fc, fa, wc, wa, nc, na, wa - fa, coverage, note),
                start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (3, 5, 7, 8, 9) and value is not None:
                cell.number_format = '0.00'
        tot = [tot[0] + fc, tot[1] + fa, tot[2] + wc,
               tot[3] + wa, tot[4] + nc, tot[5] + na]
        r += 1
    for col, value in enumerate(
            ('ИТОГО', tot[0], tot[1], tot[2], tot[3], tot[4], tot[5],
             tot[3] - tot[1],
             (tot[3] * 100.0 / tot[1])
             if tot[3] > 0.005 and tot[1] > 0.005 else None, ''),
            start=1):
        cell = ws.cell(row=r, column=col, value=value)
        cell.font = bold
        if col in (3, 5, 7, 8, 9) and value is not None:
            cell.number_format = '0.00'

    # -- «По дням» ---------------------------------------------------------
    ws = wb.create_sheet('По дням')
    ws.cell(row=1, column=1, value=DAILY_NOTE)
    header(ws, 2, ('Дата', 'Вылетов', 'Га по вылетам',
                   'Работ (по дате начала)', 'Га по ведомостям',
                   'Разница, га (вед. − выл.)', 'Из них многодневных',
                   'Га многодневных', 'Файлы работ дня'),
           (12, 10, 15, 20, 17, 22, 18, 15, 60))
    ws.freeze_panes = 'A3'
    r = 3
    for day in days:
        key = day.isoformat()
        fc, fa = fl_day.get(key, (0, 0.0))
        wc, wa, mc, ma = wk_day.get(key, (0, 0.0, 0, 0.0))
        files = '; '.join(sorted(wk_day_files.get(key, ())))
        for col, value in enumerate(
                (key, fc, fa, wc, wa, wa - fa, mc, ma, files), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (3, 5, 6, 8):
                cell.number_format = '0.00'
        r += 1
    r += 1
    cell = ws.cell(row=r, column=1, value='Работы без даты '
                   '(в дни не входят, учтены в месяце книги)')
    cell.font = bold
    r += 1
    header(ws, r, ('Месяц', 'Работ', 'Га'), (12, 10, 12))
    r += 1
    for month in months:
        nc, na = wk_month_nodate.get(month, [0, 0.0])
        if not nc:
            continue
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=nc)
        ws.cell(row=r, column=3, value=na).number_format = '0.00'
        r += 1

    # -- «Работы» ----------------------------------------------------------
    ws = wb.create_sheet('Работы')
    header(ws, 1, ('ID', 'Месяц', 'Дата от', 'Дата до', 'Дата как в книге',
                   'Заказчик', 'Га', 'Подразделение', 'Оператор (в книге)',
                   'Файл', 'Лист', 'Строка', 'Многодневная'),
           (7, 9, 11, 11, 16, 32, 9, 15, 18, 38, 22, 8, 13))
    ws.freeze_panes = 'A2'
    for r, w in enumerate(works, start=2):
        values = (w['id'], w['month'], w['date_from'], w['date_to'],
                  w['date_raw'], w['customer'], w['area'], w['subdivision'],
                  w['operator'], w['file'], w['sheet'], w['row'],
                  'да' if w['multiday'] else '')
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col == 7:
                cell.number_format = '0.00'

    # -- «Вылеты по машинам» ----------------------------------------------
    ws = wb.create_sheet('Вылеты по машинам')
    header(ws, 1, ('Дата', 'Машина', 'Вылетов', 'Га'), (12, 22, 10, 12))
    ws.freeze_panes = 'A2'
    def unit_sort(item):
        (day, label), _ = item
        # «№2» раньше «ник: …», номера -- по числу, не по строке.
        if label.startswith('№'):
            return (day, 0, int(label[1:]))
        return (day, 1, label)
    for r, ((day, label), (count, area)) in enumerate(
            sorted(fl_day_unit.items(), key=unit_sort), start=2):
        ws.cell(row=r, column=1, value=day)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=count)
        ws.cell(row=r, column=4, value=area).number_format = '0.00'

    return wb


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Daily reconciliation of DJI flights against the '
                    'dispatcher books. Read-only, writes an xlsx report.')
    parser.add_argument('--db', default=importer.DEFAULT_DB_PATH,
                        help='SQLite database (default instance/transport.db)')
    parser.add_argument('--month-from', default=DEFAULT_MONTH_FROM,
                        help='first month, YYYY-MM (default %s)'
                             % DEFAULT_MONTH_FROM)
    parser.add_argument('--month-to', default=DEFAULT_MONTH_TO,
                        help='last month, YYYY-MM (default %s)'
                             % DEFAULT_MONTH_TO)
    parser.add_argument('--out', default=None,
                        help='xlsx path (default: next to the db)')
    args = parser.parse_args(argv)

    if not _valid_month(args.month_from) or not _valid_month(args.month_to):
        print('ERROR: months must look like YYYY-MM')
        return 1
    if args.month_from > args.month_to:
        print('ERROR: --month-from is after --month-to')
        return 1
    if args.out is None:
        args.out = os.path.join(
            os.path.dirname(os.path.abspath(args.db)) or '.',
            'drone_flights_vs_books_daily.xlsx')

    try:
        con = importer._connect(args.db, read_only=True)
    except importer.ImportPreconditionError as exc:
        print(str(exc))
        return 2
    try:
        try:
            importer._require_tables(
                con, ('drone_works', 'drone_flights', 'drone_units'))
        except importer.ImportPreconditionError as exc:
            print(str(exc))
            return 2

        months = month_range(args.month_from, args.month_to)
        days = day_range(args.month_from, args.month_to)
        fl_day, fl_day_unit = collect_flights(
            con, args.month_from, args.month_to)
        works = collect_works(con, args.month_from, args.month_to)
    finally:
        con.close()

    wb = build_workbook(months, days, fl_day, fl_day_unit, works)
    wb.save(args.out)

    fl_month = {}
    for day, (count, area) in fl_day.items():
        fm = fl_month.setdefault(day[:7], [0, 0.0])
        fm[0] += count
        fm[1] += area
    wk_month = {}
    for w in works:
        wm = wk_month.setdefault(w['month'], [0, 0.0])
        wm[0] += 1
        wm[1] += w['area']
    print('daily reconciliation %s..%s (read-only)'
          % (args.month_from, args.month_to))
    for month in months:
        fc, fa = fl_month.get(month, [0, 0.0])
        wc, wa = wk_month.get(month, [0, 0.0])
        print('  %s | flights: %10.2f ha /%6d | books: %10.2f ha /%5d works '
              '| diff: %+11.2f'
              % (month, fa, fc, wa, wc, wa - fa))
    print('report: %s' % importer._ascii(os.path.abspath(args.out)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
