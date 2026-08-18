# -*- coding: utf-8 -*-
"""Сшивка вылетов с контурами DJI по рамке. Только чтение.

    python tools/drone_contours_report.py --db instance/transport.db \\
        --from 2025-09-01 --to 2025-09-30 --out "Контуры Сентябрь 2025.xlsx"

Зачем. У вылета нет ни поля, ни заказчика: `plot_name` пуст на всех строках,
а адрес не различает хозяйства -- 19 и 20 сентября 2025 два РАЗНЫХ хозяйства
пришли одной строкой «Bukhara Region, 500200». Имя заказчика есть только в
контурах Field Management, которые оператор вводит с пульта. Снимок контуров
кладёт `--lands`; этот скрипт проверяет, что из снимка получается.

Правило сшивки -- точка вылета внутри рамки контура. Полигон для этого не
нужен и намеренно не качается: ссылка на него живёт шесть часов. Рамка поля в
3 га -- это примерно 210 x 200 м, поэтому точка почти всегда попадает в свой
контур; исключение -- смежные поля, чьи рамки перекрываются. Такие вылеты
считаются ОТДЕЛЬНО и не раздаются молча одному из контуров: при нескольких
кандидатах ближайший центр выбирается только для сводки «по контурам», а сам
факт спора виден на отдельном листе. Оценивать полноту сшивки надо по числу
однозначных попаданий, а не по итоговым гектарам.

Оператор выводится тем же правилом, что и в приложении: назначение
`drone_operator_assignments` по машине и дню в UTC+5. Вылет, покрытый более
чем одним назначением, оператору не приписывается -- он попадает в «оператор
спорный», как и в отчётах.

Скрипт НИЧЕГО не пишет в базу: соединение открывается только на чтение, через
stdlib sqlite3, без Flask. `from app import app` здесь недопустим -- create_app()
вызывает db.create_all() на импорте и превращает читателя в писателя.

В консоль -- только ASCII (кодировка консоли Windows). Все кириллические
названия уходят в xlsx.
"""

import argparse
import datetime
import os
import sqlite3
import sys

# Смещение показа. Дроны опрыскивают ночью, поэтому день считается по UTC+5:
# в UTC вылеты одной смены разъезжаются по двум датам.
UTC_OFFSET_MINUTES = 300

SOURCE_DJI = 'dji'

SHEET_SUMMARY = 'Сводка'
SHEET_CONTOURS = 'По контурам'
SHEET_NO_CONTOUR = 'Вылеты без контура'
SHEET_DISPUTED = 'Спорные рамки'
SHEET_UNUSED = 'Контуры без вылетов'


class Contour(object):
    __slots__ = ('id', 'name', 'serial', 'area_ha', 'min_lat', 'min_lng',
                 'max_lat', 'max_lng', 'center_lat', 'center_lng',
                 'source_created_at')

    def __init__(self, row):
        (self.id, self.name, self.serial, self.area_ha, self.min_lat,
         self.min_lng, self.max_lat, self.max_lng, self.center_lat,
         self.center_lng, self.source_created_at) = row

    def contains(self, lat, lng):
        return (self.min_lat <= lat <= self.max_lat
                and self.min_lng <= lng <= self.max_lng)

    def distance2(self, lat, lng):
        """Квадрат расстояния до центра. Корень не нужен -- сравниваем между
        собой, а не с порогом."""
        clat = self.center_lat if self.center_lat is not None else (
            (self.min_lat + self.max_lat) / 2.0)
        clng = self.center_lng if self.center_lng is not None else (
            (self.min_lng + self.max_lng) / 2.0)
        return (lat - clat) ** 2 + (lng - clng) ** 2


def open_readonly(path):
    """Соединение только на чтение. Отсутствующий файл -- ошибка, не пустая БД."""
    if not os.path.exists(path):
        print('ERROR: database not found at %s' % path)
        sys.exit(2)
    uri = 'file:%s?mode=ro' % path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def load_contours(conn):
    """Контуры DJI с пригодной рамкой, и число отброшенных."""
    rows = conn.execute(
        'SELECT id, name, serial_number, area_ha, bbox_min_lat, bbox_min_lng, '
        '       bbox_max_lat, bbox_max_lng, center_lat, center_lng, '
        '       source_created_at '
        'FROM field_contours WHERE source = ?', (SOURCE_DJI,)).fetchall()
    usable, without_box = [], 0
    for row in rows:
        contour = Contour(row)
        if None in (contour.min_lat, contour.min_lng, contour.max_lat,
                    contour.max_lng):
            without_box += 1
            continue
        usable.append(contour)
    return usable, len(rows), without_box


def load_flights(conn, date_from, date_to):
    modifier = '+%d minutes' % UTC_OFFSET_MINUTES
    return conn.execute(
        'SELECT f.id, date(datetime(f.started_at, ?)) AS day, '
        '       f.drone_unit_id, u.number, f.area_ha, f.lat, f.lng, '
        '       f.location_text '
        'FROM drone_flights f '
        'LEFT JOIN drone_units u ON u.id = f.drone_unit_id '
        'WHERE date(datetime(f.started_at, ?)) BETWEEN ? AND ? '
        'ORDER BY day, u.number',
        (modifier, modifier, date_from, date_to)).fetchall()


def load_assignments(conn):
    return conn.execute(
        'SELECT a.drone_unit_id, o.full_name, a.date_from, a.date_to '
        'FROM drone_operator_assignments a '
        'JOIN drone_operators o ON o.id = a.operator_id').fetchall()


def operator_for(assignments, unit_id, day):
    """Оператор машины на этот день, или None при отсутствии/споре."""
    names = {name for cover_unit, name, date_from, date_to in assignments
             if cover_unit == unit_id and date_from <= day
             and (date_to is None or date_to >= day)}
    if len(names) == 1:
        return names.pop()
    return None


def index_by_latitude(contours):
    """Контуры, разложенные по целому градусу широты.

    [REASON]: 5 489 контуров на 5 661 вылет -- это 31 млн проверок в лоб.
    Все поля холдинга лежат в двух-трёх градусах широты, поэтому ведро по
    целому градусу срезает перебор в разы и не может дать ложный промах:
    контур кладётся во ВСЕ вёдра, которые задевает его рамка.
    """
    buckets = {}
    for contour in contours:
        low = int(contour.min_lat // 1)
        high = int(contour.max_lat // 1)
        for key in range(low, high + 1):
            buckets.setdefault(key, []).append(contour)
    return buckets


def match(contours, lat, lng):
    """Все контуры, чья рамка накрывает точку."""
    return [c for c in contours if c.contains(lat, lng)]


def build(conn, date_from, date_to):
    contours, total_contours, without_box = load_contours(conn)
    buckets = index_by_latitude(contours)
    assignments = load_assignments(conn)
    flights = load_flights(conn, date_from, date_to)

    stats = {
        'contours_total': total_contours,
        'contours_usable': len(contours),
        'contours_without_box': without_box,
        'flights': len(flights),
        'flights_without_coords': 0,
        'matched_one': 0,
        'matched_many': 0,
        'matched_none': 0,
        'ha_total': 0.0,
        'ha_matched_one': 0.0,
        'ha_matched_many': 0.0,
        'ha_none': 0.0,
    }
    per_contour = {}
    no_contour = []
    disputed = []

    for (_fid, day, unit_id, number, area, lat, lng, address) in flights:
        area = float(area or 0)
        stats['ha_total'] += area
        operator = operator_for(assignments, unit_id, day)

        if lat is None or lng is None:
            stats['flights_without_coords'] += 1
            stats['matched_none'] += 1
            stats['ha_none'] += area
            no_contour.append((day, number, operator, area, address,
                               lat, lng, 'нет координат'))
            continue

        candidates = match(buckets.get(int(lat // 1), ()), lat, lng)
        if not candidates:
            stats['matched_none'] += 1
            stats['ha_none'] += area
            no_contour.append((day, number, operator, area, address,
                               lat, lng, 'ни одна рамка не накрывает точку'))
            continue

        if len(candidates) == 1:
            stats['matched_one'] += 1
            stats['ha_matched_one'] += area
        else:
            stats['matched_many'] += 1
            stats['ha_matched_many'] += area
            disputed.append((day, number, operator, area, lat, lng,
                             len(candidates),
                             ' | '.join(sorted(c.name for c in candidates))))
        # Ближайший центр -- только чтобы сводка «по контурам» была одной
        # строкой на контур. Спор при этом уже посчитан выше и виден отдельно.
        best = min(candidates, key=lambda c: c.distance2(lat, lng))
        bucket = per_contour.setdefault(best.id, {
            'contour': best, 'flights': 0, 'ha': 0.0, 'days': set(),
            'machines': set(), 'operators': set(), 'disputed': 0})
        bucket['flights'] += 1
        bucket['ha'] += area
        bucket['days'].add(day)
        if number is not None:
            bucket['machines'].add(number)
        bucket['operators'].add(operator or 'не определён')
        if len(candidates) > 1:
            bucket['disputed'] += 1

    unused = [c for c in contours if c.id not in per_contour]
    return stats, per_contour, no_contour, disputed, unused


def write_xlsx(path, stats, per_contour, no_contour, disputed, unused,
               date_from, date_to):
    from openpyxl import Workbook

    book = Workbook()

    sheet = book.active
    sheet.title = SHEET_SUMMARY
    pct = (lambda part, whole: (part / whole * 100.0) if whole else 0.0)
    for row in [
        ('Период', '%s .. %s' % (date_from, date_to)),
        ('', ''),
        ('Контуров DJI в базе', stats['contours_total']),
        ('  из них с пригодной рамкой', stats['contours_usable']),
        ('  без рамки (в сшивке не участвуют)', stats['contours_without_box']),
        ('', ''),
        ('Вылетов за период', stats['flights']),
        ('  без координат', stats['flights_without_coords']),
        ('  попали ровно в один контур', stats['matched_one']),
        ('  попали в несколько рамок (спорные)', stats['matched_many']),
        ('  не попали никуда', stats['matched_none']),
        ('', ''),
        ('Гектаров за период', round(stats['ha_total'], 2)),
        ('  однозначно на контуре', round(stats['ha_matched_one'], 2)),
        ('  в спорных рамках', round(stats['ha_matched_many'], 2)),
        ('  без контура', round(stats['ha_none'], 2)),
        ('', ''),
        ('Доля однозначных вылетов, %',
         round(pct(stats['matched_one'], stats['flights']), 1)),
        ('Доля однозначных гектаров, %',
         round(pct(stats['ha_matched_one'], stats['ha_total']), 1)),
        ('', ''),
        ('Контуров, на которые лёг хотя бы один вылет', len(per_contour)),
        ('Контуров без вылетов за период', len(unused)),
    ]:
        sheet.append(list(row))
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 22

    sheet = book.create_sheet(SHEET_CONTOURS)
    sheet.append(['Контур', 'Номер DJI', 'Площадь контура, га',
                  'Опрыскано, га', 'Вылетов', 'Дней', 'Даты', 'Машины',
                  'Операторы', 'Из них спорных вылетов', 'Создан (UTC)'])
    for bucket in sorted(per_contour.values(), key=lambda b: -b['ha']):
        contour = bucket['contour']
        sheet.append([
            contour.name,
            contour.serial,
            round(contour.area_ha, 2) if contour.area_ha is not None else None,
            round(bucket['ha'], 2),
            bucket['flights'],
            len(bucket['days']),
            ', '.join(sorted(bucket['days'])),
            ', '.join(str(m) for m in sorted(bucket['machines'])),
            ', '.join(sorted(bucket['operators'])),
            bucket['disputed'],
            contour.source_created_at,
        ])
    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['I'].width = 28

    sheet = book.create_sheet(SHEET_NO_CONTOUR)
    sheet.append(['Дата', 'Машина', 'Оператор', 'Га', 'Адрес DJI',
                  'Широта', 'Долгота', 'Причина'])
    for row in no_contour:
        sheet.append(list(row[:3]) + [round(row[3], 4)] + list(row[4:]))
    sheet.column_dimensions['E'].width = 44

    sheet = book.create_sheet(SHEET_DISPUTED)
    sheet.append(['Дата', 'Машина', 'Оператор', 'Га', 'Широта', 'Долгота',
                  'Рамок', 'Кандидаты'])
    for row in disputed:
        sheet.append(list(row[:3]) + [round(row[3], 4)] + list(row[4:]))
    sheet.column_dimensions['H'].width = 60

    sheet = book.create_sheet(SHEET_UNUSED)
    sheet.append(['Контур', 'Номер DJI', 'Площадь, га', 'Создан (UTC)'])
    for contour in sorted(unused, key=lambda c: (c.name or '')):
        sheet.append([contour.name, contour.serial,
                      round(contour.area_ha, 2)
                      if contour.area_ha is not None else None,
                      contour.source_created_at])
    sheet.column_dimensions['A'].width = 34

    book.save(path)
    return path


def print_summary(stats):
    """ASCII only -- Windows console encoding."""
    pct = (lambda part, whole: (part / whole * 100.0) if whole else 0.0)
    print('contours (dji) total       : %d' % stats['contours_total'])
    print('  with a usable bbox       : %d' % stats['contours_usable'])
    print('  without a bbox           : %d' % stats['contours_without_box'])
    print('flights in period          : %d' % stats['flights'])
    print('  without coordinates      : %d' % stats['flights_without_coords'])
    print('  in exactly one contour   : %d (%.1f%%)'
          % (stats['matched_one'], pct(stats['matched_one'], stats['flights'])))
    print('  in two or more boxes     : %d' % stats['matched_many'])
    print('  in no contour at all     : %d' % stats['matched_none'])
    print('hectares total             : %.2f' % stats['ha_total'])
    print('  unambiguously on a field : %.2f (%.1f%%)'
          % (stats['ha_matched_one'],
             pct(stats['ha_matched_one'], stats['ha_total'])))
    print('  in disputed boxes        : %.2f' % stats['ha_matched_many'])
    print('  with no contour          : %.2f' % stats['ha_none'])


def parse_day(text, flag):
    try:
        datetime.date.fromisoformat(text)
    except ValueError:
        print('ERROR: %s must be YYYY-MM-DD, got %r' % (flag, text))
        sys.exit(1)
    return text


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Match drone flights to DJI field contours by bounding '
                    'box. Read-only.')
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'))
    parser.add_argument('--from', dest='date_from', required=True,
                        metavar='YYYY-MM-DD')
    parser.add_argument('--to', dest='date_to', required=True,
                        metavar='YYYY-MM-DD')
    parser.add_argument('--out', help='write an xlsx report to this path')
    args = parser.parse_args(argv)

    date_from = parse_day(args.date_from, '--from')
    date_to = parse_day(args.date_to, '--to')
    if date_from > date_to:
        print('ERROR: --from is after --to')
        sys.exit(1)

    conn = open_readonly(args.db)
    try:
        stats, per_contour, no_contour, disputed, unused = build(
            conn, date_from, date_to)
    finally:
        conn.close()

    print_summary(stats)
    if args.out:
        try:
            write_xlsx(args.out, stats, per_contour, no_contour, disputed,
                       unused, date_from, date_to)
        except ImportError:
            print('ERROR: openpyxl is not installed; the console summary '
                  'above is complete, the xlsx was not written.')
            return 1
        print('report written: %s' % args.out)
    return 0


if __name__ == '__main__':
    sys.exit(main())
