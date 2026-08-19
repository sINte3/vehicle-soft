#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_money_audit.py -- DRONE-MONEY-001, только чтение.

Деньги против гектаров. Запрошено владельцем 2026-08-19:

    «Самое главное - потом при проверках всех отчетов было все верно. Не
     только по гектарам но и по деньгам. Одно исходило из другого. Если будут
     излишки или недостатки - то обоснованные согласно гектарам и показателям
     справок и налички.»

ЧТО ПРОВЕРЯЕТСЯ. Три тождества, каждое -- построчно, и каждое сопровождается
числом строк, на которых его ВООБЩЕ БЫЛО чем проверить:

  1. сумма = гектары x ставка          (amount = area_ha * price_per_ha)
  2. получено = сумма - расходы        (received_amount = amount - other_costs)
     только для строк с received_kind = 'received'; 'operator_due' -- это
     ДОЛГ оператора, а не приход, и складывать их нельзя;
  3. ничего не потеряно: сумма строк по каждому оператору равна его итогу,
     а сумма итогов -- итогу месяца.

[REASON]: «0 расхождений» ничего не значит, пока не сказано, сколько строк
проверку прошли. Строка без ставки или без суммы проверить тождество 1 не
даёт, и молчаливое её пропускание превращает отчёт в самообман. Поэтому у
каждого тождества три числа: проверено / сошлось / не проверить, и третье
раскрывается отдельным листом поимённо.

ЧЕМ ОБОСНОВЫВАЕТСЯ РАСХОЖДЕНИЕ. Гектары книги и гектары телеметрии законно
расходятся: книга пишет предъявленную площадь, DJI -- опрысканную. Отчёт
переводит эту разницу в деньги ПО СОБСТВЕННОЙ СРЕДНЕЙ СТАВКЕ ОПЕРАТОРА
(сумма / гектары его же строк), а не по константе: ставки в книгах разные,
от 75 040 до 200 000 сум/га, и общая константа соврала бы там, где работа
шла по внутренней цене.

ЧЕГО ОТЧЁТ НЕ УТВЕРЖДАЕТ. Расхождение -- не вердикт и не недостача в кассе.
Инструмент кладёт рядом числа и источник каждого, ничего не выводя сам.

Ничего не пишет: база открывается read-only через file:-URI, служба может
работать.

Запуск (сервер, из каталога установки):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_money_audit.py --db instance\\transport.db --month 2025-09

Выход -- xlsx (--out, по умолчанию drone_money_audit_<месяц>.xlsx рядом с базой):
  1. «Сводка»              -- месяц: гектары, деньги, все три тождества.
  2. «По операторам»       -- книга/телеметрия в га и в сумах, разница.
  3. «Сумма не сходится»   -- строки, где amount != area_ha * price_per_ha.
  4. «Приход не сходится»  -- строки, где received != amount - costs.
  5. «Проверить нечем»     -- строки без ставки, суммы или прихода.

Код возврата 1, если нарушено тождество 3 (что-то потеряно при сборке) --
это ошибка самого отчёта, а не данных, и молчать о ней нельзя.
Консоль -- только ASCII.
"""

import argparse
import os
import sqlite3
import sys
from collections import defaultdict

TOLERANCE_SUM = 1.0        # сум; книги хранят 1739999.9999999998
TOLERANCE_HA = 0.005
UTC_OFFSET_MINUTES = 300   # тот же сдвиг, что и _drone_flight_month_expr()
RECEIVED_KIND_RECEIVED = 'received'


def _ascii(text):
    return str(text).encode('ascii', 'replace').decode('ascii')


def connect_ro(db_path):
    """Только чтение: служба может работать, отчёт ничего не запишет."""
    uri = 'file:%s?mode=ro' % db_path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def load_works(conn, month):
    """Работы месяца по ЕЁ СОБСТВЕННОЙ дате, как их группирует приложение."""
    return conn.execute(
        "SELECT w.id, COALESCE(o.full_name, ''), w.customer_raw, w.area_ha, "
        "       w.price_per_ha, w.amount, w.other_costs, w.received_amount, "
        "       w.received_kind, w.payment_type, w.date_raw, "
        "       COALESCE(w.source_sheet, ''), COALESCE(w.source_row, 0) "
        "FROM drone_works w "
        "LEFT JOIN drone_operators o ON o.id = w.drone_operator_id "
        "WHERE COALESCE(strftime('%Y-%m', w.work_date_from), "
        "               w.period_month) = ?", (month,)).fetchall()


def load_telemetry(conn, month):
    """Гектары телеметрии по оператору -- через назначения, как в приложении.

    [REASON]: вылет, покрытый двумя назначениями, не отдаётся молча одному из
    них. Он идёт в 'спорное' и печатается в сводке отдельной строкой: отчёт,
    который не может доказать, что ничего не потерял, аудитом не является.
    """
    modifier = '+%d minutes' % UTC_OFFSET_MINUTES
    flights = conn.execute(
        "SELECT f.drone_unit_id, date(datetime(f.started_at, ?)) AS day, "
        "       f.area_ha FROM drone_flights f "
        "WHERE strftime('%Y-%m', datetime(f.started_at, ?)) = ?",
        (modifier, modifier, month)).fetchall()
    covers = conn.execute(
        'SELECT a.drone_unit_id, o.full_name, a.date_from, a.date_to '
        'FROM drone_operator_assignments a '
        'JOIN drone_operators o ON o.id = a.operator_id').fetchall()

    by_operator = defaultdict(float)
    unassigned = ambiguous = total = 0.0
    for unit_id, day, area in flights:
        area = float(area or 0)
        total += area
        names = {name for cover_unit, name, date_from, date_to in covers
                 if cover_unit == unit_id and date_from <= day
                 and (date_to is None or date_to >= day)}
        if not names:
            unassigned += area
        elif len(names) > 1:
            ambiguous += area
        else:
            by_operator[names.pop()] += area
    return by_operator, unassigned, ambiguous, total


def audit(works):
    """Три тождества построчно. Возвращает корзины и итоги по операторам."""
    bad_amount, bad_received, uncheckable = [], [], []
    checked_amount = checked_received = 0
    per_operator = defaultdict(lambda: defaultdict(float))
    rows_total = defaultdict(float)

    for (work_id, operator, customer, area, price, amount, costs, received,
         kind, payment, date_raw, sheet, source_row) in works:
        area = float(area or 0)
        costs = float(costs or 0)
        agg = per_operator[operator]
        agg['ha'] += area
        agg['amount'] += float(amount or 0)
        agg['costs'] += costs
        if kind == RECEIVED_KIND_RECEIVED:
            agg['received'] += float(received or 0)
        elif received is not None:
            agg['operator_due'] += float(received or 0)
        agg['rows'] += 1
        rows_total['ha'] += area
        rows_total['amount'] += float(amount or 0)

        row = (work_id, operator, customer, area, price, amount, costs,
               received, kind, payment, date_raw, sheet, source_row)

        if price is None or amount is None:
            uncheckable.append(row + ('нет ставки' if price is None
                                      else 'нет суммы',))
        else:
            checked_amount += 1
            if abs(area * float(price) - float(amount)) > TOLERANCE_SUM:
                bad_amount.append(row + (area * float(price),))

        if kind != RECEIVED_KIND_RECEIVED or amount is None or received is None:
            reason = ('приход не «кирим»' if kind != RECEIVED_KIND_RECEIVED
                      else 'нет прихода')
            uncheckable.append(row + (reason,))
        else:
            checked_received += 1
            if abs(float(amount) - costs - float(received)) > TOLERANCE_SUM:
                bad_received.append(row + (float(amount) - costs,))

    return dict(bad_amount=bad_amount, bad_received=bad_received,
                uncheckable=uncheckable, checked_amount=checked_amount,
                checked_received=checked_received, per_operator=per_operator,
                rows_total=rows_total, works=len(works))


def conservation(result):
    """Тождество 3: ничего не потеряно при сборке. Ошибка отчёта, не данных."""
    problems = []
    ha = sum(a['ha'] for a in result['per_operator'].values())
    amount = sum(a['amount'] for a in result['per_operator'].values())
    if abs(ha - result['rows_total']['ha']) > TOLERANCE_HA:
        problems.append('hectares lost while grouping: rows %.4f, operators '
                        '%.4f' % (result['rows_total']['ha'], ha))
    if abs(amount - result['rows_total']['amount']) > TOLERANCE_SUM:
        problems.append('money lost while grouping: rows %.2f, operators %.2f'
                        % (result['rows_total']['amount'], amount))
    counted = sum(a['rows'] for a in result['per_operator'].values())
    if counted != result['works']:
        problems.append('rows lost while grouping: read %d, grouped %d'
                        % (result['works'], counted))
    return problems


def summary_rows(month, result, telemetry):
    by_operator, unassigned, ambiguous, total = telemetry
    book_ha = result['rows_total']['ha']
    money = result['per_operator']
    amount = sum(a['amount'] for a in money.values())
    costs = sum(a['costs'] for a in money.values())
    received = sum(a['received'] for a in money.values())
    due = sum(a['operator_due'] for a in money.values())
    checked_amount = result['checked_amount']
    checked_received = result['checked_received']
    return [
        ('Месяц', month),
        ('Работ в книгах, строк', result['works']),
        ('', ''),
        ('Гектары книг', round(book_ha, 2)),
        ('Гектары телеметрии, всего', round(total, 2)),
        ('  из них на операторах', round(sum(by_operator.values()), 2)),
        ('  спорное покрытие (день у двоих)', round(ambiguous, 2)),
        ('  без назначения', round(unassigned, 2)),
        ('Книга минус телеметрия на операторах',
         round(book_ha - sum(by_operator.values()), 2)),
        ('', ''),
        ('Сумма по книгам', round(amount, 2)),
        ('Прочие расходы', round(costs, 2)),
        ('Получено («кирим қилинган»)', round(received, 2)),
        ('Оператор должен сдать («топшириши керак»)', round(due, 2)),
        ('Не получено (сумма - расходы - получено - долг)',
         round(amount - costs - received - due, 2)),
        ('', ''),
        ('Тождество 1: сумма = гектары x ставка', ''),
        ('  Т1: строк, где было чем проверить', checked_amount),
        ('  Т1: из них сошлось', checked_amount - len(result['bad_amount'])),
        ('  Т1: расходится', len(result['bad_amount'])),
        ('Тождество 2: получено = сумма - расходы', ''),
        ('  Т2: строк, где было чем проверить', checked_received),
        ('  Т2: из них сошлось',
         checked_received - len(result['bad_received'])),
        ('  Т2: расходится', len(result['bad_received'])),
        ('Проверить нечем, случаев', len(result['uncheckable'])),
        ('', ''),
        ('Тождество 3: ничего не потеряно',
         'ДА' if not conservation(result) else 'НЕТ'),
    ]


def operator_rows(result, telemetry):
    by_operator, _unassigned, _ambiguous, _total = telemetry
    rows = []
    names = set(result['per_operator']) | set(by_operator)
    for name in sorted(names, key=lambda n: -result['per_operator'][n]['ha']):
        agg = result['per_operator'].get(name, defaultdict(float))
        book_ha = agg['ha']
        tel_ha = by_operator.get(name, 0.0)
        # [REASON]: средняя ставка САМОГО оператора, а не общая константа --
        # в книгах есть и 75 040, и 200 000 сум/га, и общая цифра соврала бы
        # там, где работа шла по внутренней цене.
        # [REASON]: ставка округляется ДО умножения, а не после. Иначе
        # читатель, перемноживший показанные ему столбцы, получит не то
        # число, что стоит в третьем -- проверка, которую нельзя повторить,
        # проверкой не является.
        rate = round(agg['amount'] / book_ha, 2) if book_ha else 0.0
        rows.append((
            name or '(без оператора)', agg['rows'],
            round(book_ha, 2), round(tel_ha, 2),
            round(tel_ha - book_ha, 2),
            round(tel_ha / book_ha * 100, 1) if book_ha else '',
            round(agg['amount'], 2), round(agg['costs'], 2),
            round(agg['received'], 2), round(agg['operator_due'], 2),
            round(agg['amount'] - agg['costs'] - agg['received']
                  - agg['operator_due'], 2),
            rate,
            round((tel_ha - book_ha) * rate, 2),
        ))
    return rows


def write_xlsx(path, month, result, telemetry):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = 'Сводка'
    ws.append(['Показатель', 'Значение'])
    ws['A1'].font = ws['B1'].font = bold
    for label, value in summary_rows(month, result, telemetry):
        ws.append([label, value])
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 20

    ws = wb.create_sheet('По операторам')
    header = ['Оператор', 'Строк', 'Га книги', 'Га телеметрии', 'Разница, га',
              '% телеметрии к книге', 'Сумма', 'Расходы', 'Получено',
              'Долг оператора', 'Не получено', 'Своя средняя ставка, сум/га',
              'Разница в сумах']
    ws.append(header)
    for cell in ws[1]:
        cell.font = bold
    for row in operator_rows(result, telemetry):
        ws.append(list(row))
    for idx, width in enumerate([26, 7, 11, 14, 12, 20, 15, 13, 15, 15, 14,
                                 26, 17], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    detail = ['id', 'Оператор', 'Заказчик', 'Га', 'Ставка', 'Сумма',
              'Расходы', 'Приход', 'Вид прихода', 'Оплата', 'Дата в книге',
              'Лист', 'Строка']
    ws = wb.create_sheet('Сумма не сходится')
    ws.append(detail + ['Должно быть'])
    for cell in ws[1]:
        cell.font = bold
    for row in result['bad_amount']:
        ws.append(list(row))

    ws = wb.create_sheet('Приход не сходится')
    ws.append(detail + ['Должно быть'])
    for cell in ws[1]:
        cell.font = bold
    for row in result['bad_received']:
        ws.append(list(row))

    ws = wb.create_sheet('Проверить нечем')
    ws.append(detail + ['Почему'])
    for cell in ws[1]:
        cell.font = bold
    for row in result['uncheckable']:
        ws.append(list(row))

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(
        description='Money against hectares for one month, read-only.')
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'))
    parser.add_argument('--month', default='2025-09', help='YYYY-MM')
    parser.add_argument('--out', default=None)
    args = parser.parse_args()

    if not os.path.isfile(args.db):
        print('ERROR: database not found: %s' % _ascii(args.db))
        return 2

    conn = connect_ro(args.db)
    try:
        works = load_works(conn, args.month)
        telemetry = load_telemetry(conn, args.month)
    finally:
        conn.close()

    result = audit(works)
    lost = conservation(result)

    out = args.out or os.path.join(
        os.path.dirname(os.path.abspath(args.db)),
        'drone_money_audit_%s.xlsx' % args.month)
    write_xlsx(out, args.month, result, telemetry)

    for label, value in summary_rows(args.month, result, telemetry):
        if label:
            print('%-46s %s' % (_ascii(label), _ascii(value)))
    print('')
    print('Written: %s' % _ascii(out))
    if lost:
        print('')
        print('CONSERVATION BROKEN -- this is a bug in the report itself:')
        for line in lost:
            print('  ' + _ascii(line))
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
