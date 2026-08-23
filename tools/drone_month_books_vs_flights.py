#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_month_books_vs_flights.py -- DRONE-SEPT2025-RECON-001:
помесячный разбор «книга диспетчера против телеметрии DJI» на файлах, а не
на базе.

ЗАЧЕМ. Владелец ведёт разбор месяц за месяцем, начиная с сентября-2025, и
присылает в чат два вида файлов: книги диспетчеров (листы «свод ичи ...») и
выгрузку вылетов приложения. Ни того, ни другого в базе может не оказаться в
нужном виде: книги правятся у диспетчеров, выгрузка снимается на дату.
Инструмент читает ровно эти два источника и отвечает на вопрос владельца:
КАКОЙ ОПЕРАТОР, В КАКИЕ ДАТЫ, НА КАКОЙ МАШИНЕ И У КОГО работал.

ПОЧЕМУ НЕ ЧЕРЕЗ БАЗУ. Разбор сентября показал, что расхождение отчётов
рождается ДО базы: 587.50 га книг датированы октябрём, 7.00 га пропущены
импортёром из-за пустого номера строки, 132.70 га уехали к другому оператору
из-за колонки «Дрон бошқарувчи оператор», заполненной чужим именем. Чтобы это
увидеть, надо читать сам файл книги, а не то, что от него осталось в базе.

ЧТО ИНСТРУМЕНТ ДЕЛАЕТ И ЧЕГО НЕ ДЕЛАЕТ

  * Читает книги: обе таблицы каждого листа -- наличные («Накд») и
    перечисление («Справка»), плюс «Топшириқ», если он есть. Оператор берётся
    из ИМЕНИ ЛИСТА, а не из колонки оператора: разбор сентября доказал, что
    колонка врёт (лист Жураева, 13 строк, 132.70 га подписаны «Қодиров
    Нурали», хотя машина Нурали в эти дни налетала своё).
  * Читает выгрузку вылетов: дата, машина, ник, адрес, гектары.
  * Раскладывает вылеты по операторам ПО ЗАДАННОЙ КАРТЕ назначений
    (--assignments). Карту инструмент НЕ ВЫВОДИТ и не угадывает: она -- ответ
    владельца, проверяемый этим же отчётом. Без карты считаются только итоги
    сторон.
  * Проверяет карту на полноту: каждый вылет обязан попасть ровно к одному
    оператору либо в явную корзину «не определён». Непокрытый вылет -- отказ
    кодом 3, потому что молча потерянные гектары и есть та беда, ради которой
    затеян разбор.

  * НЕ ПРИДУМЫВАЕТ ДАТ строкам без даты. Такие строки идут отдельной
    колонкой «без даты» и в посуточную сверку не входят. Отчёт лишь
    показывает рядом, сколько гектаров телеметрии этого оператора лежит в
    днях, где записи в книге нет: совпадение сумм -- довод для владельца,
    а не проставленная за диспетчера дата.
  * НЕ СУДИТ О ЗАКАЗЧИКЕ вылета: DJI не передаёт ни поля, ни клиента (§3
    трека). Совпадение по контрагенту возможно только в обратную сторону --
    заказчики, записанные в книге на тот же день.
  * НЕ РАВНЯЕТ ГЕКТАРЫ СТОРОН. Книга пишет предъявленную площадь, DJI --
    опрысканную; расхождение в проценты законно. Порог доверия трека
    +-9 %, всё сверх него отчёт помечает как вопрос, а не как вину.

ЛИСТЫ ОТЧЁТА
  «Сводка»            -- итоги месяца по обеим сторонам;
  «Мост к отчёту»     -- почему отчёт приложения даёт другое число: каждая
                         строка расхождения поимённо (нужен --works-report);
  «Назначения»        -- оператор x машина x даты: вылеты, гектары, довод;
  «По операторам»     -- книга против телеметрии, доля, признак выхода за +-9 %;
  «По дням»           -- оператор x день: книга / телеметрия / разница;
  «Машины по дням»    -- машина x день с оператором по карте;
  «Летал-записи нет»  -- дни с вылетами без строки в книге;
  «Запись-полёта нет» -- дни со строкой в книге без вылетов;
  «Строки книг»       -- каждая разобранная строка с файлом, листом и датами.

Только чтение: база не открывается вовсе, входные xlsx не перезаписываются.

Запуск (рабочая машина владельца, Windows):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_month_books_vs_flights.py --books-dir C:\\drones\\sept --flights C:\\drones\\sept\\flights.xlsx --month 2025-09 --assignments tools\\drone_assignments_2025_09.json --out C:\\drones\\sept\\recon.xlsx

Коды возврата: 0 -- разобрано; 1 -- не пройдено предусловие (пустая выгрузка,
имя оператора вне карты, не сошёлся --expect-books-ha); 2 -- не найден каталог
книг, файл вылетов или отчёт; 3 -- карта назначений не покрывает часть вылетов.

Откат: инструмент ничего не пишет, кроме своего xlsx; откат кода git revert,
откат данных не требуется.
"""

import argparse
import datetime
import glob
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - среда без openpyxl
    print('ERROR: openpyxl is required')
    sys.exit(1)


TRUST_BAND = 9.0  # проценты; порог доверия трека, см. docs/tracks/drones.md

# [REASON]: оператор берётся из имени листа. Колонка «Дрон бошқарувчи
# оператор» в книгах заполняется непоследовательно и в сентябре-2025 солгала
# на 132.70 га (лист Жураева подписан именем Нурали во всех строках).
SHEET_OPERATOR_MARKERS = (
    'свод ичи', 'свод ичи ', 'ичи',
)

SECTION_CASH = 'накд'
SECTION_TRANSFER = 'справка'
SECTION_TASK = 'топширик'

FARM_HEADERS = ('фх номи', 'корхона номи')
AREA_HEADERS = ('майдон', 'мадон')
DATE_HEADER_HINT = 'сана'


def _ascii(text):
    """Консоль проекта принимает только ASCII (CLAUDE.md, раздел PowerShell)."""
    return str(text).encode('ascii', 'replace').decode('ascii')


def _norm(text):
    if text is None:
        return ''
    return re.sub(r'\s+', ' ', str(text)).strip()


def _fold(text):
    """Свод написаний: узбекская кириллица к русской, регистр вниз."""
    out = _norm(text).lower()
    for src, dst in (('ў', 'у'), ('ғ', 'г'), ('қ', 'к'), ('ҳ', 'х'),
                     ('ъ', ''), ('ь', ''), ('ё', 'е')):
        out = out.replace(src, dst)
    return out


def parse_date_cell(value, default_year, default_month):
    """Разбирает ячейку даты книги. Возвращает (как записано, [дни ISO]).

    Понимает: настоящую дату Excel; «07.09.2025»; диапазон «06-07.09.2025»;
    перечисление «22,25.09.2025»; смешанное «19-30.10.2025». Всё, чего не
    понял, возвращает с пустым списком дней -- строка попадёт в «без даты»,
    а не получит выдуманное число.
    """
    if value is None:
        return '', []
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d'), [value.strftime('%Y-%m-%d')]
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d'), [value.strftime('%Y-%m-%d')]
    raw = _norm(value)
    compact = raw.replace(' ', '')
    if not compact:
        return raw, []

    # [REASON]: дата, набранная текстом в виде «2025-09-20», встречается в
    # книгах наравне с настоящей датой Excel. Без этой ветки такая строка
    # уходила в «без даты» и выпадала из посуточной сверки молча.
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', compact)
    if match:
        try:
            return raw, [datetime.date(int(match.group(1)),
                                       int(match.group(2)),
                                       int(match.group(3))).isoformat()]
        except ValueError:
            return raw, []

    # «06-07.09.2025», «22,25.09.2025», «11,12,13,26.09.2025»
    match = re.fullmatch(r'(\d{1,2})((?:[-,]\d{1,2})+)\.(\d{1,2})\.?(\d{4})',
                         compact)
    if match:
        first = int(match.group(1))
        month = int(match.group(3))
        year = int(match.group(4))
        days = [first]
        for token in re.finditer(r'([-,])(\d{1,2})', match.group(2)):
            sep, number = token.group(1), int(token.group(2))
            if sep == '-':
                if number < days[-1]:
                    return raw, []
                days.extend(range(days[-1] + 1, number + 1))
            else:
                days.append(number)
        try:
            return raw, [datetime.date(year, month, day).isoformat()
                         for day in days]
        except ValueError:
            return raw, []

    # «07.09.2025» и «7.9.25»
    match = re.fullmatch(r'(\d{1,2})\.(\d{1,2})\.?(\d{2,4})', compact)
    if match:
        year = int(match.group(3))
        if year < 100:
            year += 2000
        try:
            return raw, [datetime.date(year, int(match.group(2)),
                                       int(match.group(1))).isoformat()]
        except ValueError:
            return raw, []

    # «06-07.09» без года -- год и месяц берём у периода отчёта
    match = re.fullmatch(r'(\d{1,2})-(\d{1,2})\.(\d{1,2})', compact)
    if match:
        try:
            return raw, [datetime.date(default_year, int(match.group(3)), day)
                         .isoformat()
                         for day in range(int(match.group(1)),
                                          int(match.group(2)) + 1)]
        except ValueError:
            return raw, []

    return raw, []


def operator_from_sheet(title):
    """Имя оператора из названия листа: «свод ичи Имомов Бехзод» -> имя."""
    name = _norm(title)
    low = _fold(name)
    for marker in SHEET_OPERATOR_MARKERS:
        idx = low.find(_fold(marker))
        if idx >= 0:
            name = name[idx + len(marker):]
            break
    name = name.strip(' ()')
    return name or _norm(title)


def resolve_operator(path, sheet_title, aliases, sheet_map):
    """Приводит имя с листа к справочному.

    [REASON]: имя листа НЕ равно имени в справочнике. «свод ичи (Нурали)»
    даёт «Нурали», а человека зовут «Қодиров Нурали»; «Жумаев Фурқат» и
    «Жумаев Фуркат» -- один человек, «Лист1» книги Достон АКА не содержит
    имени вовсе. Жёстко зашитый словарь таких соответствий уже стоил треку
    20 % данных (docs/tracks/drones.md, грабли), поэтому карта живёт в
    данных, а непокрытое имя объявляется отказом, а не молчаливым нулём.
    """
    key = (os.path.basename(path), _norm(sheet_title))
    if key in sheet_map:
        return sheet_map[key]
    derived = operator_from_sheet(sheet_title)
    return aliases.get(_fold(derived), derived)


def is_detail_sheet(path, sheet_title, sheet_map):
    """Лист расшифровки по оператору -- единственный источник строк работ.

    [REASON]: полная книга месяца содержит ещё и сводные листы («ОБШИЙ
    СВОД»), годовой журнал оператора («Расул ака ... обём», март-октябрь),
    список неполученных денег («олинмаган пуллар») и лист соседнего месяца.
    Все они пересказывают ТЕ ЖЕ работы другими словами. Прочитать их наравне
    с расшифровками -- значит задвоить и затроить месяц: на сентябре-2025
    архив целиком даёт 8 336.03 га против настоящих 5 617.03. Поэтому читаются
    только расшифровки, а всё пропущенное ПЕЧАТАЕТСЯ, чтобы лист не исчез
    молча.
    """
    if (os.path.basename(path), _norm(sheet_title)) in sheet_map:
        return True
    return 'свод ичи' in _fold(sheet_title)


def read_books(books_dir, month, aliases=None, sheet_map=None):
    """Разбирает расшифровки по операторам. -> (строки, пропущенные листы)."""
    year, mon = int(month[:4]), int(month[5:7])
    aliases = aliases or {}
    sheet_map = sheet_map or {}
    paths = sorted(glob.glob(os.path.join(books_dir, '*.xlsx')))
    rows, skipped = [], []
    for path in paths:
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in book.worksheets:
                if is_detail_sheet(path, sheet.title, sheet_map):
                    rows.extend(_read_sheet(path, sheet, year, mon,
                                            aliases, sheet_map))
                else:
                    skipped.append((os.path.basename(path), sheet.title))
        finally:
            book.close()
    return rows, skipped


def read_control(path, sheet_title, aliases):
    """Читает сводный лист диспетчеров как КОНТРОЛЬ, а не как данные.

    [REASON]: у диспетчеров есть собственный подписанный итог месяца по
    операторам -- и гектарами, и деньгами. Он не источник строк (строки в
    расшифровках), но он независимая проверка разбора: если мой итог по
    человеку разошёлся с их сводом, ошибся либо я, либо книга, и знать об
    этом надо до отчёта руководству. На сентябре-2025 свод подтвердил 12
    операторов из 13 в ноль -- и по гектарам, и по разбивке «справка/наличные»,
    и по расходам, -- а по тринадцатому вскрыл расхождение внутри книги.

    Колонки свода идут подряд от колонки оператора и подписаны шапкой из трёх
    строк, которую построчно не разобрать. Поэтому они берутся смещением, а
    правильность смещения проверяется строкой «Жами (хаммаси)»: если сумма
    прочитанных строк с ней не сходится, смещение неверно и контроль об этом
    говорит, а не считает молча.
    """
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[sheet_title]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        book.close()

    fields = ('area_transfer', 'area_cash', 'amount', 'received',
              'not_received', 'expenses', 'wage', 'income')
    col = None
    out = {}
    excluded = {}
    grand = None
    for cells in rows:
        values = [_norm(c) for c in cells]
        if col is None:
            for idx, value in enumerate(values):
                if 'дрон бошкарувчи оператор' in _fold(value):
                    col = idx
                    break
            continue

        def number(offset):
            idx = col + offset
            if idx >= len(cells):
                return 0.0
            value = cells[idx]
            return float(value) if isinstance(value, (int, float)) else 0.0

        name = values[col] if col < len(values) else ''
        joined = _fold(' '.join(v for v in values if v))
        if 'жами (хаммаси)' in joined:
            grand = {key: number(i + 1) for i, key in enumerate(fields)}
            continue
        if not name or _fold(name).startswith('жами'):
            continue
        if not isinstance(cells[col + 1] if col + 1 < len(cells) else None,
                          (int, float)):
            continue
        resolved = aliases.get(_fold(name), name)
        if resolved is None:
            # [REASON]: строка объявлена как чужой месяц и в контроль не идёт,
            # но итоговая строка свода её ВКЛЮЧАЕТ. Чтобы самопроверка на
            # смещение колонок осталась настоящей, исключённое считается
            # отдельно и возвращается в сверку с итогом.
            for offset, key in enumerate(fields, start=1):
                excluded[key] = excluded.get(key, 0.0) + number(offset)
            continue
        row = out.setdefault(resolved,
                             dict({key: 0.0 for key in fields},
                                  spellings=[]))
        for offset, key in enumerate(fields, start=1):
            row[key] += number(offset)
        row['spellings'].append(name)

    if grand:
        for key in ('area_transfer', 'area_cash', 'amount'):
            got = sum(row[key] for row in out.values()) + excluded.get(key, 0.0)
            if abs(got - grand[key]) > 0.05:
                raise ValueError(
                    'control sheet: column %s does not add up to the grand '
                    'total row (%.2f vs %.2f) -- the column offset is wrong'
                    % (key, got, grand[key]))
    return out


def _read_sheet(path, sheet, year, mon, aliases, sheet_map):
    operator = resolve_operator(path, sheet.title, aliases, sheet_map)
    out = []
    section = SECTION_CASH
    marker = None
    cols = None
    for cells in sheet.iter_rows(values_only=True):
        values = [_norm(c) for c in cells]
        head = next((v for v in values if v), '')
        folded = _fold(head)

        # [REASON]: метка «Накд»/«Справка» над таблицей НЕНАДЁЖНА. В книге
        # Агрокластера слово «Справка» стоит подзаголовком денежного блока
        # выше, а метки «Накд» над наличной таблицей нет вовсе -- и все
        # 826.80 га наличных Сайфулло уходили в справку молча. Тип таблицы
        # определяется её СТРУКТУРОЙ (ниже): наличная несёт колонки расхода и
        # прихода, справка -- нет. Метка оставлена только для «Топшириқ»:
        # его структура совпадает со справкой, и отличить их можно лишь так.
        if folded.startswith(SECTION_TASK):
            marker, cols = SECTION_TASK, None
            continue
        if folded.startswith(SECTION_CASH) or folded.startswith(SECTION_TRANSFER):
            marker, cols = None, None
            continue

        if any(h in _fold(v) for v in values for h in FARM_HEADERS):
            # [REASON]: часть книг не подписывает первую таблицу словом
            # «Накд» -- шапка идёт сразу. Считаем такую таблицу наличными:
            # так её и заполняют, а «Справка» ниже всегда подписана.
            if section is None:
                section = SECTION_CASH
            cols = {}
            for idx, value in enumerate(values):
                low = _fold(value)
                # [REASON]: порядок веток значим. «Кирим қилинган сана» -- это
                # ДАТА, а «Кирим қилинган» -- ПОЛУЧЕННЫЕ ДЕНЬГИ, и отличаются
                # они только словом «сана». Проверять дату надо раньше денег,
                # иначе дата попадёт в сумму и баланс оператора станет ложным.
                if any(h in low for h in FARM_HEADERS):
                    cols.setdefault('farm', idx)
                elif any(low.startswith(h) for h in AREA_HEADERS):
                    cols.setdefault('area', idx)
                elif DATE_HEADER_HINT in low:
                    cols.setdefault('date', idx)
                elif 'бошка харажат' in low:
                    cols.setdefault('expenses', idx)
                elif low.startswith('кирим'):
                    cols.setdefault('received', idx)
                elif 'жами сумма' in low:
                    cols.setdefault('amount', idx)
            if 'farm' not in cols or 'area' not in cols:
                cols = None
                continue
            if marker == SECTION_TASK:
                section = SECTION_TASK
            elif 'expenses' in cols and 'received' in cols:
                section = SECTION_CASH
            else:
                section = SECTION_TRANSFER
            marker = None
            continue

        if cols is None:
            continue
        if _fold(values[0]).startswith('жами') or (
                cols['farm'] < len(values)
                and _fold(values[cols['farm']]).startswith('жами')):
            cols = None
            continue

        area = cells[cols['area']] if cols['area'] < len(cells) else None
        if not isinstance(area, (int, float)) or float(area) <= 0:
            continue
        farm = values[cols['farm']] if cols['farm'] < len(values) else ''
        raw, days = ('', [])
        if 'date' in cols and cols['date'] < len(cells):
            raw, days = parse_date_cell(cells[cols['date']], year, mon)

        def money(key):
            idx = cols.get(key)
            if idx is None or idx >= len(cells):
                return 0.0
            value = cells[idx]
            return float(value) if isinstance(value, (int, float)) else 0.0

        out.append({
            'file': os.path.basename(path),
            'sheet': sheet.title,
            'operator': operator,
            'section': section,
            'farm': farm or '(без имени)',
            'ha': round(float(area), 4),
            'amount': round(money('amount'), 2),
            'expenses': round(money('expenses'), 2),
            'received': round(money('received'), 2),
            'date_raw': raw,
            'days': days,
        })
    return out


FLIGHT_COLUMNS = {
    'when': ('дата', 'time'),
    'machine': ('машина',),
    'nick': ('ник',),
    'address': ('адрес',),
    'ha': ('гектар',),
}


def read_flights(path):
    """Разбирает выгрузку вылетов приложения (лист «Вылеты»)."""
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book['Вылеты'] if 'Вылеты' in book.sheetnames \
            else book.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = [_fold(h) for h in next(rows)]
        index = {}
        for key, hints in FLIGHT_COLUMNS.items():
            for idx, name in enumerate(header):
                if any(hint in name for hint in hints):
                    index[key] = idx
                    break
        missing = [k for k in ('when', 'machine', 'ha') if k not in index]
        if missing:
            raise ValueError('flights export lacks columns: %s'
                             % ', '.join(missing))
        out = []
        for cells in rows:
            when = cells[index['when']]
            if when is None:
                continue
            if isinstance(when, (datetime.datetime, datetime.date)):
                day = when.strftime('%Y-%m-%d')
            else:
                day = _norm(when)[:10]
            machine = cells[index['machine']]
            out.append({
                'day': day,
                'machine': None if machine in (None, '') else int(machine),
                'nick': _norm(cells[index['nick']]) if 'nick' in index else '',
                'address': _norm(cells[index['address']])
                           if 'address' in index else '',
                'ha': float(cells[index['ha']] or 0),
            })
        return out
    finally:
        book.close()


def read_works_report(path):
    """Читает отчёт по работам приложения: итог, операторы, заказчики.

    [REASON]: главный вопрос владельца -- почему числа разнятся МЕЖДУ
    отчётами. Ответить на него, не прочитав сам отчёт, нельзя: расхождение
    складывается из строк, которые в него не попали, и строк, которых нет в
    книгах. Поэтому мост строится на данных обеих сторон, а не на рассуждении.
    """
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = {'total': None, 'operators': {}, 'customers': {}}
        if 'Сводка' in book.sheetnames:
            for cells in book['Сводка'].iter_rows(values_only=True):
                if cells and _fold(cells[0]).startswith('гектар'):
                    out['total'] = float(cells[1] or 0)
        for title, key, col in (('По операторам', 'operators', 2),
                                ('По заказчикам', 'customers', 2)):
            if title not in book.sheetnames:
                continue
            rows = book[title].iter_rows(values_only=True)
            next(rows, None)
            for cells in rows:
                name = _norm(cells[0]) if cells else ''
                if not name or _fold(name).startswith('итого'):
                    continue
                out[key][name] = out[key].get(name, 0.0) \
                    + float(cells[col] or 0)
        return out
    finally:
        book.close()


def customer_key(name):
    """Свод написания контрагента: без «фх», «мчж», знаков и регистра."""
    folded = _fold(name)
    folded = re.sub(r'\b(фх|фз|мчж|аж|фермер хужалиги)\b', ' ', folded)
    return re.sub(r'[^а-яa-z0-9]+', ' ', folded).strip()


def build_bridge(books, dropped, report):
    """Мост «книги -> отчёт программы» по контрагентам."""
    parsed_all = sum(r['ha'] for r in books) + sum(r['ha'] for r in dropped)
    # [REASON]: свод написаний нужен для сопоставления, но в отчёт идёт
    # ИСХОДНОЕ имя: владелец ищет строку в книге глазами, а свёрнутый ключ
    # («бухоро агрокластер заминлари») в книге не написан нигде.
    shown = {}
    ours = {}
    for row in list(books) + list(dropped):
        key = customer_key(row['farm'])
        ours[key] = ours.get(key, 0.0) + row['ha']
        shown.setdefault(key, _norm(row['farm']))
    theirs = {}
    for name, area in report['customers'].items():
        key = customer_key(name)
        theirs[key] = theirs.get(key, 0.0) + area
        shown.setdefault(key, _norm(name))

    only_books, only_report, differ = [], [], []
    for key in sorted(set(ours) | set(theirs)):
        a, b = ours.get(key), theirs.get(key)
        name = shown.get(key, key)
        if a is not None and b is not None:
            if abs(a - b) > 0.05:
                differ.append((name, a, b))
        elif a is not None:
            only_books.append((name, a))
        else:
            only_report.append((name, b))
    return {
        'parsed_all': parsed_all,
        'report_total': report['total'],
        'only_books': only_books,
        'only_report': only_report,
        'differ': differ,
        'report_operators': report['operators'],
    }


def load_assignments(path):
    """Читает карту: назначения, алиасы операторов, правки книг."""
    with open(path, encoding='utf-8') as handle:
        data = json.load(handle)
    if isinstance(data, list):
        data = {'assignments': data}
    rules = data['assignments']
    for rule in rules:
        rule['machine'] = int(rule['machine'])
        rule.setdefault('operator', None)
        rule.setdefault('why', '')
    # Значение null означает «эту строку свода не считать» (другой месяц).
    aliases = {_fold(k): v for k, v in data.get('operator_aliases', {}).items()}
    sheet_map = {(item['file'], _norm(item['sheet'])): item['operator']
                 for item in data.get('sheet_operators', [])}
    return {
        'rules': rules,
        'aliases': aliases,
        'sheet_map': sheet_map,
        'redated_rows': data.get('redated_rows', []),
        'control': data.get('control_sheet'),
    }


def operator_for(rules, machine, day):
    for rule in rules:
        if rule['machine'] == machine and rule['from'] <= day <= rule['to']:
            return rule['operator'], True
    return None, False


def month_days(month):
    year, mon = int(month[:4]), int(month[5:7])
    day = datetime.date(year, mon, 1)
    out = []
    while day.month == mon:
        out.append(day.isoformat())
        day += datetime.timedelta(days=1)
    return out


def split_books_by_month(rows, month, redated_rows):
    """Делит строки книг на три корзины по ДАТЕ РАБОТЫ, а не по книге.

    [REASON]: месяц, написанный внутри книги, врёт, а дата работы -- нет
    (docs/tracks/drones.md, грабли трека). Поэтому строка попадает в месяц
    только своей датой. Три корзины различаются намеренно:
      in_month  -- есть хотя бы один день внутри месяца: идёт в посуточную
                   сверку и в итог месяца;
      undated   -- дата не читается вовсе: в итог месяца входит (строка
                   принадлежит книге этого месяца), в посуточную -- нет;
      other     -- дата читается и лежит В ДРУГОМ МЕСЯЦЕ: из итога месяца
                   ИСКЛЮЧАЕТСЯ и называется отдельным листом.
    Смешивать «без даты» с «датирована другим месяцем» нельзя: первое --
    незаполненная клетка, второе -- работа другого периода, и попав в один
    котёл они дают правдоподобный, но неверный итог месяца.

    redated_rows -- объявленная владельцем правка: работа месяца, подписанная
    в книге другим месяцем. Применяется ДО деления, поимённо.
    """
    redate = {(_fold(r['operator']), _fold(r['farm']), round(float(r['ha']), 2)): r
              for r in redated_rows}
    in_month, undated, other = [], [], []
    for row in rows:
        marker = (_fold(row['operator']), _fold(row['farm']),
                  round(row['ha'], 2))
        fix = redate.get(marker)
        if fix:
            row['redated_from'] = row['date_raw']
            row['days'] = list(fix['days'])
        if any(day.startswith(month) for day in row['days']):
            in_month.append(row)
        elif not row['days']:
            undated.append(row)
        else:
            row['excluded'] = 'дата работы вне месяца %s' % month
            other.append(row)
    return in_month, undated, other


def compute(books, flights, rules, month):
    days = month_days(month)
    day_set = set(days)

    tele_by_op = {}
    tele_by_op_day = {}
    tele_by_machine_day = {}
    seg = {}
    uncovered = []
    # [REASON]: ТЕЛЕМЕТРИЯ, УВИДЕННАЯ В МЕСЯЦЕ, считается ДО проверки покрытия
    # и независимо от неё. Без карты назначений ни один вылет не покрыт, и
    # раньше отчёт печатал «flights: 0.00 ha in 186 flights» с кодом 0 --
    # число, неотличимое от пустой выгрузки. А первый прогон месяца всегда
    # идёт БЕЗ карты: карты ещё нет, её и предстоит вывести. Это тот же класс
    # дефекта, что «records_written всегда равнялся records_seen»: признак,
    # одинаковый в двух разных случаях, признаком не является.
    seen_ha = 0.0
    seen_flights = 0
    seen_by_machine = {}
    for flight in flights:
        if flight['day'] not in day_set:
            continue
        seen_ha += flight['ha']
        seen_flights += 1
        cell = seen_by_machine.setdefault(
            flight['machine'], {'flights': 0, 'ha': 0.0, 'first': None,
                                'last': None, 'nicks': set()})
        cell['flights'] += 1
        cell['ha'] += flight['ha']
        cell['nicks'].add(flight['nick'])
        if cell['first'] is None or flight['day'] < cell['first']:
            cell['first'] = flight['day']
        if cell['last'] is None or flight['day'] > cell['last']:
            cell['last'] = flight['day']
        operator, covered = operator_for(rules, flight['machine'],
                                         flight['day'])
        if not covered:
            uncovered.append(flight)
            continue
        label = operator or '(не определён)'
        tele_by_op[label] = tele_by_op.get(label, 0.0) + flight['ha']
        tele_by_op_day.setdefault(label, {})
        tele_by_op_day[label][flight['day']] = \
            tele_by_op_day[label].get(flight['day'], 0.0) + flight['ha']
        tele_by_machine_day.setdefault(flight['machine'], {})
        tele_by_machine_day[flight['machine']][flight['day']] = \
            tele_by_machine_day[flight['machine']].get(flight['day'], 0.0) \
            + flight['ha']
        cell = seg.setdefault((label, flight['machine']),
                              {'flights': 0, 'ha': 0.0,
                               'first': None, 'last': None, 'nicks': set()})
        cell['flights'] += 1
        cell['ha'] += flight['ha']
        cell['nicks'].add(flight['nick'])
        if cell['first'] is None or flight['day'] < cell['first']:
            cell['first'] = flight['day']
        if cell['last'] is None or flight['day'] > cell['last']:
            cell['last'] = flight['day']

    book_by_op = {}
    book_by_op_day = {}
    book_undated = {}
    money = {}
    for row in books:
        operator = row['operator']
        book_by_op[operator] = book_by_op.get(operator, 0.0) + row['ha']
        purse = money.setdefault(operator, {
            'ha_cash': 0.0, 'ha_transfer': 0.0, 'amount_cash': 0.0,
            'amount_transfer': 0.0, 'expenses': 0.0, 'received': 0.0})
        if row['section'] == SECTION_TRANSFER:
            purse['ha_transfer'] += row['ha']
            purse['amount_transfer'] += row['amount']
        else:
            purse['ha_cash'] += row['ha']
            purse['amount_cash'] += row['amount']
        purse['expenses'] += row['expenses']
        purse['received'] += row['received']
        days_in = [d for d in row['days'] if d in day_set]
        if not days_in:
            book_undated[operator] = book_undated.get(operator, 0.0) + row['ha']
            continue
        share = row['ha'] / len(days_in)
        book_by_op_day.setdefault(operator, {})
        for day in days_in:
            book_by_op_day[operator][day] = \
                book_by_op_day[operator].get(day, 0.0) + share

    return {
        'days': days,
        'tele_by_op': tele_by_op,
        'tele_by_op_day': tele_by_op_day,
        'tele_by_machine_day': tele_by_machine_day,
        'segments': seg,
        'uncovered': uncovered,
        'seen_ha': seen_ha,
        'seen_flights': seen_flights,
        'seen_by_machine': seen_by_machine,
        'book_by_op': book_by_op,
        'book_by_op_day': book_by_op_day,
        'book_undated': book_undated,
        'money': money,
    }


THIN = Side(style='thin', color='D0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
WARN_FILL = PatternFill('solid', fgColor='FCE4D6')
NOTE_FILL = PatternFill('solid', fgColor='FFF2CC')


def _head(sheet, row, titles, widths=None):
    for col, title in enumerate(titles, 1):
        cell = sheet.cell(row=row, column=col, value=title)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    if widths:
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = width
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def _put(sheet, row, col, value, bold=False, fmt=None, fill=None):
    # [REASON]: openpyxl решает тип клетки по первому символу: строка,
    # начинающаяся с «=», записывается ФОРМУЛОЙ и открывается ошибкой. Подписи
    # инструмента пишутся человеком, формулы -- только явными шаблонами ниже,
    # поэтому подпись со «=» в начале это всегда описка, а не намерение.
    if isinstance(value, str) and value.startswith('=') \
            and not _FORMULA.fullmatch(value):
        raise ValueError('подпись начинается со знака равенства и уйдёт в файл '
                         'формулой: %r (лист %s, %d:%d)'
                         % (value, sheet.title, row, col))
    cell = sheet.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', size=10, bold=bold)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


NUM = '#,##0.00'
PCT = '0.0"%"'

# Формулы, которые инструмент пишет намеренно. Всё прочее со «=» -- описка.
_FORMULA = re.compile(r'=(SUM\([A-Z]+\d+:[A-Z]+\d+\)'
                      r'|IF\([A-Z]+\d+=0,"",[A-Z]+\d+/[A-Z]+\d+\*100\)'
                      r'|[A-Z]+\d+(?:[-+][A-Z]+\d+)+)')


def _sheet_bridge(sheet, bridge, model, operators):
    _put(sheet, 1, 1, 'Почему отчёт по работам и книги дают разные числа',
         bold=True).font = Font(name='Arial', size=12, bold=True)
    sheet.column_dimensions['A'].width = 56
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 66

    row = 3
    _head(sheet, row, ['Шаг', 'Гектары', 'Что это'], [56, 14, 66])
    row += 1
    start = row
    _put(sheet, row, 1, 'Книги диспетчеров, все разобранные строки')
    _put(sheet, row, 2, round(bridge['parsed_all'], 2), fmt=NUM)
    _put(sheet, row, 3, 'сумма листов «свод ичи» всех книг каталога')
    row += 1
    for key, area in sorted(bridge['only_books'], key=lambda kv: -kv[1]):
        _put(sheet, row, 1, '  − нет в отчёте: %s' % key)
        _put(sheet, row, 2, round(-area, 2), fmt=NUM)
        _put(sheet, row, 3, 'строка книги в отчёт не попала')
        row += 1
    for key, ours, theirs in sorted(bridge['differ'],
                                    key=lambda kv: -abs(kv[1] - kv[2])):
        _put(sheet, row, 1, '  − расходится: %s' % key)
        _put(sheet, row, 2, round(theirs - ours, 2), fmt=NUM)
        _put(sheet, row, 3, 'в книгах %.2f, в отчёте %.2f' % (ours, theirs))
        row += 1
    for key, area in sorted(bridge['only_report'], key=lambda kv: -kv[1]):
        _put(sheet, row, 1, '  + есть только в отчёте: %s' % key)
        _put(sheet, row, 2, round(area, 2), fmt=NUM)
        _put(sheet, row, 3, 'в присланных книгах такой строки нет')
        row += 1
    _put(sheet, row, 1, 'ИТОГО: отчёт по работам (расчёт)', bold=True)
    _put(sheet, row, 2, '=SUM(B%d:B%d)' % (start, row - 1), bold=True, fmt=NUM)
    _put(sheet, row, 3, 'должно совпасть со строкой ниже')
    calc_row = row
    row += 1
    _put(sheet, row, 1, 'Отчёт по работам (как в файле)', bold=True)
    _put(sheet, row, 2, bridge['report_total'], bold=True, fmt=NUM)
    _put(sheet, row, 3, 'лист «Сводка», строка «Гектаров»')
    row += 1
    _put(sheet, row, 1, 'Невязка моста', bold=True)
    _put(sheet, row, 2, '=B%d-B%d' % (calc_row, row - 1), bold=True, fmt=NUM)
    _put(sheet, row, 3, 'должна быть 0.00')

    row += 2
    _put(sheet, row, 1, 'Гектары по операторам: отчёт против раскладки',
         bold=True)
    row += 1
    _head(sheet, row, ['Оператор', 'В отчёте', 'Телеметрия по раскладке'],
          [56, 14, 66])
    row += 1
    names = sorted(set(bridge['report_operators']) | set(operators))
    for name in names:
        their = bridge['report_operators'].get(name)
        ours = model['tele_by_op'].get(name)
        if their is None and ours is None:
            continue
        _put(sheet, row, 1, name)
        _put(sheet, row, 2, round(their, 2) if their is not None else None,
             fmt=NUM)
        _put(sheet, row, 3, round(ours, 2) if ours is not None else None,
             fmt=NUM)
        row += 1


def build_workbook(model, books, dropped, rules, month, sources, bridge=None):
    book = openpyxl.Workbook()
    # [REASON]: openpyxl пишет формулы БЕЗ кэша значения, поэтому до первого
    # пересчёта итоги и проценты читаются как пустые -- и Excel, и любой
    # просмотрщик. Этот флаг заставляет пересчитать при открытии, иначе
    # владелец увидит пустые клетки там, где стоят суммы.
    book.calculation.fullCalcOnLoad = True

    operators = sorted(
        set(model['book_by_op']) | set(model['tele_by_op']) - {'(не определён)'},
        key=lambda name: -model['tele_by_op'].get(name, 0.0))
    if '(не определён)' in model['tele_by_op']:
        operators = [o for o in operators if o != '(не определён)']
        operators.append('(не определён)')

    _sheet_summary(book.active, model, books, dropped, month, sources,
                   operators)
    if bridge:
        _sheet_bridge(book.create_sheet('Мост к отчёту'), bridge, model,
                      operators)
    _sheet_assignments(book.create_sheet('Назначения'), model, rules)
    _sheet_operators(book.create_sheet('По операторам'), model, operators)
    _sheet_balance(book.create_sheet('Баланс по операторам'), model, operators)
    if model.get('control'):
        _sheet_control(book.create_sheet('Контроль по своду'), model, operators)
    _sheet_daily(book.create_sheet('По дням'), model, operators)
    _sheet_machines(book.create_sheet('Машины по дням'), model, rules)
    _sheet_gaps(book.create_sheet('Летал-записи нет'), model, operators, True)
    _sheet_gaps(book.create_sheet('Запись-полёта нет'), model, operators, False)
    _sheet_rows(book.create_sheet('Строки книг'), books, dropped)
    return book


def _sheet_summary(sheet, model, books, dropped, month, sources, operators):
    sheet.title = 'Сводка'
    sheet.column_dimensions['A'].width = 58
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 70

    title = _put(sheet, 1, 1, 'Сверка книг диспетчеров с телеметрией DJI, %s'
                 % month, bold=True)
    title.font = Font(name='Arial', size=13, bold=True)

    row = 3
    _head(sheet, row, ['Показатель', 'Гектары', 'Чем получен'],
          [58, 16, 70])
    row += 1
    book_total = sum(model['book_by_op'].values())
    tele_total = sum(model['tele_by_op'].values())
    dated = sum(sum(d.values()) for d in model['book_by_op_day'].values())
    undated = sum(model['book_undated'].values())

    facts = [
        ('Книги диспетчеров: разобрано строк всего',
         book_total + sum(r['ha'] for r in dropped),
         'все строки листов «свод ичи», без отбора по месяцу'),
        ('  минус строки с датой работы в другом месяце',
         -sum(r['ha'] for r in dropped),
         'дата читается и лежит вне месяца -- лист «Строки книг»'),
        ('КНИГИ ЗА МЕСЯЦ', book_total,
         'то, что сравнивается с телеметрией'),
        ('  из них строки с читаемой датой', dated,
         'участвуют в посуточной сверке'),
        ('  из них строки без даты', undated,
         'в итог месяца входят, в посуточную сверку -- нет'),
        ('ТЕЛЕМЕТРИЯ DJI ЗА МЕСЯЦ', tele_total,
         'выгрузка вылетов, колонка «Гектары»'),
        ('Разница «книги - телеметрия»', book_total - tele_total,
         'книга пишет предъявленную площадь, DJI -- опрысканную'),
    ]
    first_num = row
    for name, value, why in facts:
        _put(sheet, row, 1, name, bold=name.startswith(('Книги', 'Телеметрия')))
        _put(sheet, row, 2, round(value, 2), fmt=NUM)
        _put(sheet, row, 3, why)
        row += 1

    row += 1
    _put(sheet, row, 1, 'Контроль полноты карты назначений', bold=True)
    _put(sheet, row, 2, len(model['uncovered']),
         fill=WARN_FILL if model['uncovered'] else None)
    _put(sheet, row, 3, 'вылетов, не покрытых ни одним правилом; должно быть 0')
    row += 1
    _put(sheet, row, 1, 'Операторов вне порога +-%.0f %%' % TRUST_BAND, bold=True)
    outs = sum(1 for name in operators
               if model['tele_by_op'].get(name, 0.0) > 0
               and abs(100.0 * model['book_by_op'].get(name, 0.0)
                       / model['tele_by_op'][name] - 100.0) > TRUST_BAND)
    _put(sheet, row, 2, outs, fill=WARN_FILL if outs else None)
    _put(sheet, row, 3, 'каждый такой оператор -- вопрос, а не вина')
    row += 2

    _put(sheet, row, 1, 'Источники', bold=True)
    row += 1
    for name in sources:
        _put(sheet, row, 1, name)
        row += 1
    row += 1
    _put(sheet, row, 1,
         'Порог доверия трека: +-%.0f %%. Он не означает, что вне порога '
         'кто-то виноват: он означает, что расхождение надо объяснить.'
         % TRUST_BAND).alignment = Alignment(wrap_text=True)
    return first_num


def _sheet_assignments(sheet, model, rules):
    _head(sheet, 1, ['Оператор', 'Машина', 'С', 'По', 'Вылетов', 'Гектары',
                     'Ники DJI в окне', 'Довод'],
          [24, 9, 12, 12, 10, 12, 30, 74])
    row = 2
    ordered = sorted(model['segments'].items(),
                     key=lambda kv: (-kv[1]['ha'], kv[0][1]))
    for (operator, machine), cell in ordered:
        why = ''
        for rule in rules:
            if rule['machine'] == machine and \
                    (rule['operator'] or '(не определён)') == operator:
                why = rule.get('why', '')
                break
        _put(sheet, row, 1, operator)
        _put(sheet, row, 2, '№%d' % machine)
        _put(sheet, row, 3, cell['first'])
        _put(sheet, row, 4, cell['last'])
        _put(sheet, row, 5, cell['flights'])
        _put(sheet, row, 6, round(cell['ha'], 2), fmt=NUM)
        _put(sheet, row, 7, ', '.join(sorted(n for n in cell['nicks'] if n)))
        _put(sheet, row, 8, why).alignment = Alignment(wrap_text=True,
                                                       vertical='top')
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    _put(sheet, row, 5, '=SUM(E2:E%d)' % (row - 1), bold=True)
    _put(sheet, row, 6, '=SUM(F2:F%d)' % (row - 1), bold=True, fmt=NUM)


def _sheet_operators(sheet, model, operators):
    _head(sheet, 1, ['Оператор', 'Машины и даты', 'Книга, га',
                     'из них без даты', 'Телеметрия, га', 'Книга/телеметрия',
                     'Разница, га', 'Вне порога'],
          [24, 34, 13, 15, 15, 17, 13, 12])
    row = 2
    for operator in operators:
        machines = sorted(
            ((m, c) for (o, m), c in model['segments'].items() if o == operator),
            key=lambda item: item[1]['first'] or '')
        label = ', '.join('№%d (%s-%s)' % (m, c['first'][8:], c['last'][8:])
                          for m, c in machines)
        book = model['book_by_op'].get(operator, 0.0)
        tele = model['tele_by_op'].get(operator, 0.0)
        _put(sheet, row, 1, operator)
        _put(sheet, row, 2, label)
        _put(sheet, row, 3, round(book, 2), fmt=NUM)
        _put(sheet, row, 4, round(model['book_undated'].get(operator, 0.0), 2),
             fmt=NUM)
        _put(sheet, row, 5, round(tele, 2), fmt=NUM)
        if tele > 0:
            _put(sheet, row, 6, '=IF(E%d=0,"",C%d/E%d*100)' % (row, row, row),
                 fmt=PCT)
        else:
            _put(sheet, row, 6, '')
        _put(sheet, row, 7, '=C%d-E%d' % (row, row), fmt=NUM)
        share = (100.0 * book / tele) if tele else 0.0
        out = tele > 0 and abs(share - 100.0) > TRUST_BAND
        _put(sheet, row, 8, 'да' if out else '',
             fill=WARN_FILL if out else None)
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    for col in (3, 4, 5):
        letter = openpyxl.utils.get_column_letter(col)
        _put(sheet, row, col, '=SUM(%s2:%s%d)' % (letter, letter, row - 1),
             bold=True, fmt=NUM)
    _put(sheet, row, 6, '=IF(E%d=0,"",C%d/E%d*100)' % (row, row, row),
         bold=True, fmt=PCT)
    _put(sheet, row, 7, '=C%d-E%d' % (row, row), bold=True, fmt=NUM)


def _sheet_balance(sheet, model, operators):
    """Деньги оператора: сколько выставлено, потрачено, сдано и что осталось.

    [REASON]: «не сдано» считается как СУММА НАЛИЧНЫХ минус расходы минус
    сданное -- по строкам самой книги, а не берётся из её итога. Итог книги
    пишет человек, а невязка ищется как раз в человеческой записи; взять итог
    значило бы проверять запись ею же. Справка в «не сдано» не входит: она
    приходит перечислением на счёт, а не наличными оператору.
    """
    _put(sheet, 1, 1, 'Баланс наличных по операторам (из строк книг)',
         bold=True).font = Font(name='Arial', size=12, bold=True)
    _head(sheet, 2, ['Оператор', 'Га наличные', 'Га справка', 'Га всего',
                     'Сумма наличные', 'Сумма справка', 'Расходы (харажат)',
                     'Сдано наличными', 'НЕ СДАНО',
                     'Свод: не получено', 'Расхождение со сводом'],
          [24, 12, 12, 11, 16, 16, 17, 16, 15, 16, 20])
    row = 3
    first = row
    for name in operators:
        purse = model['money'].get(name)
        if not purse:
            continue
        entry = (model.get('control') or {}).get(name)
        _put(sheet, row, 1, name)
        _put(sheet, row, 2, round(purse['ha_cash'], 2), fmt=NUM)
        _put(sheet, row, 3, round(purse['ha_transfer'], 2), fmt=NUM)
        _put(sheet, row, 4, '=B%d+C%d' % (row, row), fmt=NUM)
        _put(sheet, row, 5, round(purse['amount_cash'], 2), fmt=NUM)
        _put(sheet, row, 6, round(purse['amount_transfer'], 2), fmt=NUM)
        _put(sheet, row, 7, round(purse['expenses'], 2), fmt=NUM)
        _put(sheet, row, 8, round(purse['received'], 2), fmt=NUM)
        unpaid = purse['amount_cash'] - purse['expenses'] - purse['received']
        _put(sheet, row, 9, '=E%d-G%d-H%d' % (row, row, row), fmt=NUM,
             fill=WARN_FILL if abs(unpaid) > 0.5 else None)
        if entry:
            _put(sheet, row, 10, round(entry['not_received'], 2), fmt=NUM)
            _put(sheet, row, 11, '=I%d-J%d' % (row, row), fmt=NUM,
                 fill=NOTE_FILL
                 if abs(unpaid - entry['not_received']) > 0.5 else None)
        else:
            _put(sheet, row, 10, None, fmt=NUM)
            _put(sheet, row, 11, None, fmt=NUM)
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    for col in range(2, 12):
        letter = openpyxl.utils.get_column_letter(col)
        _put(sheet, row, col,
             '=SUM(%s%d:%s%d)' % (letter, first, letter, row - 1),
             bold=True, fmt=NUM)
    row += 2
    for line in (
            '«НЕ СДАНО» = сумма наличными − расходы − сдано, посчитано по '
            'строкам книги. Ноль означает, что деньги за наличную работу '
            'сошлись.',
            '«Свод: не получено» — колонка «Олинмаган сумма» сводного листа '
            'самих диспетчеров: сколько они САМИ признают несобранным.',
            'Расхождение между этими двумя колонками — то, о чём стоит '
            'спрашивать: книга и свод расходятся в деньгах.',
            'Справка в «не сдано» не входит: она приходит перечислением, '
            'а не наличными на руки оператору.'):
        _put(sheet, row, 1, line).alignment = Alignment(wrap_text=True)
        sheet.merge_cells(start_row=row, start_column=1,
                          end_row=row, end_column=11)
        row += 1


def _sheet_control(sheet, model, operators):
    _put(sheet, 1, 1, 'Свод самих диспетчеров против моего разбора книг',
         bold=True).font = Font(name='Arial', size=12, bold=True)
    _head(sheet, 2, ['Оператор', 'Свод: справка', 'Свод: наличные',
                     'Свод: итого', 'Мой разбор книги (все строки)',
                     'Разница', 'Телеметрия', 'Как записан в своде'],
          [24, 14, 15, 13, 20, 11, 13, 34])
    row = 3
    def control_area(name):
        entry = model['control'].get(name)
        return (entry['area_transfer'] + entry['area_cash']) if entry else 0.0

    names = sorted(set(model['control']) | set(operators),
                   key=lambda n: -control_area(n))
    for name in names:
        entry = model['control'].get(name)
        mine = model['book_all_by_op'].get(name)
        if entry is None and mine is None:
            continue
        transfer = entry['area_transfer'] if entry else None
        cash = entry['area_cash'] if entry else None
        _put(sheet, row, 1, name)
        _put(sheet, row, 2, round(transfer, 2) if entry else None, fmt=NUM)
        _put(sheet, row, 3, round(cash, 2) if entry else None, fmt=NUM)
        _put(sheet, row, 4, '=B%d+C%d' % (row, row) if entry else None,
             fmt=NUM)
        _put(sheet, row, 5, round(mine, 2) if mine is not None else None,
             fmt=NUM)
        gap = None if (entry is None or mine is None) \
            else mine - (transfer + cash)
        _put(sheet, row, 6, '=E%d-D%d' % (row, row) if gap is not None else None,
             fmt=NUM, fill=WARN_FILL if gap and abs(gap) > 0.05 else None)
        _put(sheet, row, 7, round(model['tele_by_op'].get(name, 0.0), 2),
             fmt=NUM)
        _put(sheet, row, 8, ', '.join(sorted(set(entry['spellings'])))
             if entry else '')
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    for col in (2, 3, 4, 5, 7):
        letter = openpyxl.utils.get_column_letter(col)
        _put(sheet, row, col, '=SUM(%s3:%s%d)' % (letter, letter, row - 1),
             bold=True, fmt=NUM)
    _put(sheet, row, 6, '=E%d-D%d' % (row, row), bold=True, fmt=NUM)


def _sheet_daily(sheet, model, operators):
    _head(sheet, 1, ['День'] + [o for o in operators for _ in (0, 1)],
          [12] + [11] * (2 * len(operators)))
    col = 2
    for operator in operators:
        sheet.cell(row=1, column=col).value = operator + ' (книга)'
        sheet.cell(row=1, column=col + 1).value = operator + ' (телем.)'
        for shift in (0, 1):
            cell = sheet.cell(row=1, column=col + shift)
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.fill = HEAD_FILL
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        col += 2
    row = 2
    for day in model['days']:
        _put(sheet, row, 1, day)
        col = 2
        for operator in operators:
            book = model['book_by_op_day'].get(operator, {}).get(day, 0.0)
            tele = model['tele_by_op_day'].get(operator, {}).get(day, 0.0)
            _put(sheet, row, col, round(book, 2) if book else None, fmt=NUM)
            _put(sheet, row, col + 1, round(tele, 2) if tele else None, fmt=NUM)
            col += 2
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    for idx in range(2, 2 + 2 * len(operators)):
        letter = openpyxl.utils.get_column_letter(idx)
        _put(sheet, row, idx, '=SUM(%s2:%s%d)' % (letter, letter, row - 1),
             bold=True, fmt=NUM)


def _sheet_machines(sheet, model, rules):
    machines = sorted(model['tele_by_machine_day'])
    _head(sheet, 1, ['День'] + ['№%d' % m for m in machines] + ['Всего'],
          [12] + [10] * len(machines) + [11])
    row = 2
    for day in model['days']:
        _put(sheet, row, 1, day)
        for idx, machine in enumerate(machines, 2):
            value = model['tele_by_machine_day'][machine].get(day, 0.0)
            _put(sheet, row, idx, round(value, 2) if value else None, fmt=NUM)
        last = openpyxl.utils.get_column_letter(1 + len(machines))
        _put(sheet, row, 2 + len(machines),
             '=SUM(B%d:%s%d)' % (row, last, row), fmt=NUM)
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    for idx in range(2, 3 + len(machines)):
        letter = openpyxl.utils.get_column_letter(idx)
        _put(sheet, row, idx, '=SUM(%s2:%s%d)' % (letter, letter, row - 1),
             bold=True, fmt=NUM)
    row += 1
    _put(sheet, row, 1, 'Оператор по карте', bold=True)
    for idx, machine in enumerate(machines, 2):
        parts = ['%s %s-%s' % (r['operator'] or 'не определён',
                               r['from'][8:], r['to'][8:])
                 for r in rules if r['machine'] == machine]
        _put(sheet, row, idx, '; '.join(parts)).alignment = \
            Alignment(wrap_text=True, vertical='top')


def _sheet_gaps(sheet, model, operators, flights_side):
    if flights_side:
        _head(sheet, 1, ['Оператор', 'День', 'Телеметрия, га',
                         'В книге на этот день', 'Замечание'],
              [24, 13, 15, 20, 60])
    else:
        _head(sheet, 1, ['Оператор', 'День', 'Книга, га',
                         'Телеметрия этого дня', 'Замечание'],
              [24, 13, 15, 20, 60])
    row = 2
    for operator in operators:
        undated = model['book_undated'].get(operator, 0.0)
        for day in model['days']:
            book = model['book_by_op_day'].get(operator, {}).get(day, 0.0)
            tele = model['tele_by_op_day'].get(operator, {}).get(day, 0.0)
            if flights_side and (tele < 0.5 or book >= 0.01):
                continue
            if not flights_side and (book < 0.5 or tele >= 0.01):
                continue
            note = ''
            if flights_side and undated > 0:
                note = ('у оператора %.2f га книг без даты -- часть этих '
                        'дней может быть ими' % undated)
            _put(sheet, row, 1, operator)
            _put(sheet, row, 2, day)
            _put(sheet, row, 3, round(tele if flights_side else book, 2),
                 fmt=NUM)
            _put(sheet, row, 4, round(book if flights_side else tele, 2),
                 fmt=NUM)
            _put(sheet, row, 5, note, fill=NOTE_FILL if note else None)
            row += 1
    if row == 2:
        _put(sheet, 2, 1, 'нет строк')
        row = 3
    else:
        _put(sheet, row, 1, 'ИТОГО', bold=True)
        _put(sheet, row, 3, '=SUM(C2:C%d)' % (row - 1), bold=True, fmt=NUM)
    if not flights_side:
        return

    # [REASON]: без этого блока лист врёт по смыслу. У оператора со строками
    # без даты дни «летал -- записи нет» и есть те самые строки: у Нурали
    # 324.30 га справки без даты против 322.54 га таких дней. Показывать
    # только левую половину -- значит выставлять записанную работу
    # незаписанной. Дат при этом инструмент не проставляет: он ставит два
    # числа рядом, решение остаётся за владельцем.
    row += 2
    _put(sheet, row, 1, 'Сопоставление строк книг без даты', bold=True)
    row += 1
    _head(sheet, row, ['Оператор', 'Книг без даты, га',
                       'Телеметрия в днях без записи, га', 'Совпадение',
                       'Дни'], [24, 18, 30, 13, 60])
    row += 1
    for operator in operators:
        undated = model['book_undated'].get(operator, 0.0)
        if undated < 0.01:
            continue
        free = [(day, model['tele_by_op_day'][operator][day])
                for day in model['days']
                if model['tele_by_op_day'].get(operator, {}).get(day, 0.0) >= 0.3
                and model['book_by_op_day'].get(operator, {}).get(day, 0.0) < 0.01]
        total = sum(value for _, value in free)
        _put(sheet, row, 1, operator)
        _put(sheet, row, 2, round(undated, 2), fmt=NUM)
        _put(sheet, row, 3, round(total, 2), fmt=NUM)
        _put(sheet, row, 4, '=IF(C%d=0,"",B%d/C%d*100)' % (row, row, row),
             fmt=PCT)
        _put(sheet, row, 5, ', '.join('%s (%.1f)' % (day[8:], value)
                                      for day, value in free))
        row += 1


def _sheet_rows(sheet, books, dropped):
    _head(sheet, 1, ['Файл', 'Лист', 'Оператор', 'Таблица', 'Контрагент',
                     'Гектары', 'Дата как записана', 'Дни разобраны',
                     'Исключена'],
          [34, 28, 22, 11, 40, 11, 20, 34, 26])
    row = 2
    for item in list(books) + list(dropped):
        _put(sheet, row, 1, item['file'])
        _put(sheet, row, 2, item['sheet'])
        _put(sheet, row, 3, item['operator'])
        _put(sheet, row, 4, item['section'])
        _put(sheet, row, 5, item['farm'])
        _put(sheet, row, 6, item['ha'], fmt=NUM)
        _put(sheet, row, 7, item['date_raw'])
        _put(sheet, row, 8, ', '.join(item['days']))
        _put(sheet, row, 9, item.get('excluded', ''),
             fill=WARN_FILL if item.get('excluded') else None)
        row += 1
    _put(sheet, row, 1, 'ИТОГО', bold=True)
    _put(sheet, row, 6, '=SUM(F2:F%d)' % (row - 1), bold=True, fmt=NUM)


def main():
    parser = argparse.ArgumentParser(
        description='Books of dispatchers versus DJI telemetry, one month')
    parser.add_argument('--books-dir', required=True)
    parser.add_argument('--flights', required=True)
    parser.add_argument('--month', required=True, help='YYYY-MM')
    parser.add_argument('--assignments', default=None,
                        help='JSON map operator x machine x dates')
    parser.add_argument('--out', required=True)
    parser.add_argument('--works-report', default=None,
                        help='xlsx report of the application, to bridge to')
    parser.add_argument('--expect-books-ha', type=float, default=None,
                        help='self-check: fail if parsed books differ by >0.05')
    args = parser.parse_args()

    if not os.path.isdir(args.books_dir):
        print('ERROR: books directory not found: %s' % _ascii(args.books_dir))
        return 2
    if not os.path.isfile(args.flights):
        print('ERROR: flights export not found: %s' % _ascii(args.flights))
        return 2
    if not re.fullmatch(r'\d{4}-\d{2}', args.month):
        print('ERROR: --month must be YYYY-MM')
        return 1

    spec = {'rules': [], 'aliases': {}, 'sheet_map': {}, 'redated_rows': [],
            'control': None}
    if args.assignments:
        spec = load_assignments(args.assignments)
    rules = spec['rules']

    books, skipped = read_books(args.books_dir, args.month, spec['aliases'],
                                spec['sheet_map'])
    flights = read_flights(args.flights)
    if not books:
        print('ERROR: no book rows parsed from %s' % _ascii(args.books_dir))
        return 1
    if not flights:
        print('ERROR: no flights parsed from %s' % _ascii(args.flights))
        return 1

    in_month, undated, dropped = split_books_by_month(books, args.month,
                                                      spec['redated_rows'])
    books = in_month + undated

    # [REASON]: имя из книги, которого нет в карте назначений, означает
    # непокрытый алиас. Без этой проверки такой человек получил бы в отчёте
    # честный ноль телеметрии против сотен гектаров книги -- то самое молчаливое
    # расхождение, из-за которого числа сентября и разошлись между отчётами.
    known = {r['operator'] for r in rules if r['operator']}
    unknown = sorted({r['operator'] for r in books} - known) if rules else []
    if unknown:
        print('ERROR: book operators absent from the assignment map:')
        for name in unknown:
            total = sum(r['ha'] for r in books if r['operator'] == name)
            print('  %-30s %8.2f ha' % (_ascii(name), total))
        print('Add them to operator_aliases / sheet_operators / assignments.')
        return 1

    parsed_ha = sum(r['ha'] for r in books) + sum(r['ha'] for r in dropped)
    if args.expect_books_ha is not None and \
            abs(parsed_ha - args.expect_books_ha) > 0.05:
        print('ERROR: parsed books %.2f ha, expected %.2f ha'
              % (parsed_ha, args.expect_books_ha))
        return 1

    control = {}
    if spec['control']:
        control_path = os.path.join(args.books_dir, spec['control']['file'])
        if not os.path.isfile(control_path):
            print('ERROR: control sheet file not found: %s'
                  % _ascii(spec['control']['file']))
            return 2
        control = read_control(control_path, spec['control']['sheet'],
                               spec['aliases'])

    model = compute(books, flights, rules, args.month)
    model['control'] = control
    # [REASON]: свод диспетчеров подписан на ВСЕ строки книги, включая
    # датированные соседним месяцем. Сравнивать его с итогом месяца значит
    # получить ложное расхождение ровно на эти строки (в сентябре -- 76.00 га
    # Холмуродова). Контроль сверяется с полным разбором книги.
    parsed_all_by_op = {}
    for row in list(books) + list(dropped):
        parsed_all_by_op[row['operator']] = \
            parsed_all_by_op.get(row['operator'], 0.0) + row['ha']
    model['book_all_by_op'] = parsed_all_by_op

    bridge = None
    if args.works_report:
        if not os.path.isfile(args.works_report):
            print('ERROR: works report not found: %s'
                  % _ascii(args.works_report))
            return 2
        bridge = build_bridge(books, dropped,
                              read_works_report(args.works_report))

    sources = [os.path.basename(args.flights)]
    if args.works_report:
        sources.append(os.path.basename(args.works_report))
    sources += sorted({r['file'] for r in books})
    workbook = build_workbook(model, books, dropped, rules, args.month,
                              sources, bridge)
    workbook.save(args.out)

    tele_total = sum(model['tele_by_op'].values())
    book_total = sum(model['book_by_op'].values())
    print('month %s (read-only)' % args.month)
    print('  books   : %8.2f ha in %d rows (%d rows for other months)'
          % (book_total, len(books), len(dropped)))
    if skipped:
        print('  skipped %d sheet(s) that are not per-operator breakdowns:'
              % len(skipped))
        for name, title in skipped:
            print('    %s | %s' % (_ascii(name), _ascii(title)))
    print('  flights : %8.2f ha in %d flights (inside %s)'
          % (model['seen_ha'], model['seen_flights'], args.month))
    if rules:
        print('  attributed to operators: %8.2f ha' % tele_total)
        print('  coverage: %d flights not covered by the map'
              % len(model['uncovered']))
        for operator in sorted(model['tele_by_op'],
                               key=lambda o: -model['tele_by_op'][o]):
            book = model['book_by_op'].get(operator, 0.0)
            tele = model['tele_by_op'][operator]
            share = (100.0 * book / tele) if tele else 0.0
            flag = '' if abs(share - 100.0) <= TRUST_BAND else '  <-- out of band'
            print('    %-24s book %8.2f  tele %8.2f  %6.1f%%%s'
                  % (_ascii(operator), book, tele, share, flag))
    else:
        # [REASON]: карты нет -- значит НИ ОДИН вылет не отнесён к оператору, и
        # листы отчёта по операторам пусты. Сказать это словами обязательно:
        # молча напечатанный ноль читается как «телеметрии за месяц нет», а
        # это разные вещи. Заодно печатается раскладка ПО МАШИНАМ -- ровно то
        # сырьё, из которого карту и выводят.
        print('  NO ASSIGNMENT MAP (--assignments not given): not a single '
              'flight is attributed')
        print('  to an operator, and every per-operator sheet of the report '
              'is empty by construction.')
        print('  Telemetry by machine -- this is the raw material for the '
              'map:')
        for machine in sorted(model['seen_by_machine'],
                              key=lambda m: (m is None, m)):
            cell = model['seen_by_machine'][machine]
            nicks = ', '.join(sorted(n for n in cell['nicks'] if n))
            print('    machine %-6s %5d flights %9.2f ha  %s..%s  %s'
                  % ('-' if machine is None else '#%d' % machine,
                     cell['flights'], cell['ha'], cell['first'], cell['last'],
                     _ascii(nicks)[:48]))
    if control:
        worst = 0.0
        for name, entry in sorted(control.items()):
            theirs = entry['area_transfer'] + entry['area_cash']
            mine = model['book_all_by_op'].get(name, 0.0)
            worst = max(worst, abs(mine - theirs))
            print('    control %-24s book-summary %8.2f  parsed %8.2f  %+7.2f'
                  % (_ascii(name), theirs, mine, mine - theirs))
        print('  control : worst gap %.2f ha' % worst)
    if bridge:
        computed = bridge['parsed_all'] \
            - sum(a for _, a in bridge['only_books']) \
            + sum(t - o for _, o, t in bridge['differ']) \
            + sum(a for _, a in bridge['only_report'])
        print('  bridge  : books %.2f -> report %.2f (file says %.2f, '
              'residual %+.2f)'
              % (bridge['parsed_all'], computed,
                 bridge['report_total'] or 0.0,
                 computed - (bridge['report_total'] or 0.0)))
    print('report: %s' % _ascii(os.path.abspath(args.out)))
    if rules and model['uncovered']:
        return 3
    return 0


if __name__ == '__main__':
    sys.exit(main())
