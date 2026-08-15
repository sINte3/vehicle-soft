#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_operator_audit.py -- DRONE-OPERATOR-AUDIT-001, только чтение.

Помесячный аудит «оператор x месяц»: книга против телеметрии его машин.
Запрошен владельцем 2026-08-15 («поставить работу с дронами на прозрачные
рельсы») и стал возможен после загрузки назначений: до 2026-08-15 в базе
лежало два назначения, теперь 39, и телеметрию можно положить на человека.

ЧТО СЧИТАЕТСЯ. Для каждого оператора и месяца:
  * га телеметрии -- вылеты машин, закреплённых за ним назначениями
    (drone_operator_assignments) на день вылета; день считается в UTC+5;
  * га книг с датой -- его работы по месяцу work_date_from (урок
    DRONE-BOOKS-DRIFT-001: месяц книги врёт, дата работы -- нет);
  * га книг без даты -- его работы без читаемой даты, по месяцу книги;
  * разница и три класса расхождений:
      А) «летал -- записи нет»: день, когда его машина летала, а в его
         книгах этот день не покрыт ни одной работой;
      Б) «записал больше, чем летал»: видно из матрицы (книга >> телеметрия);
      В) «запись -- полёта нет»: работа с датами, ни один день которых не
         был лётным для его машин.

ЧЕСТНОСТЬ АРИФМЕТИКИ. Каждый гектар телеметрии попадает ровно в одну из
четырёх корзин: оператор / спорное (день покрыт ДВУМЯ назначениями --
модель это разрешает, аудит не выбирает между ними, а называет обоих) /
без назначения / без машины. Инвариант «сумма корзин = вся телеметрия
месяца» проверяется в самом отчёте и печатается строкой в «Сводке»; отчёт,
который не может доказать, что ничего не потерял, аудитом не является.

ЧЕГО ОТЧЁТ НЕ УТВЕРЖДАЕТ. Расхождение -- не вердикт. Книга пишет
предъявленную площадь, DJI -- опрысканную; они законно расходятся на
проценты. Смотреть надо на большое и систематическое, и решает человек:
инструмент кладёт рядом числа и источники, ничего не выводя сам.

Ничего не пишет: база открывается read-only через file:-URI, служба может
работать. Ни одного пишущего слова в запросах нет (проверено тестом через
ast, как у drone_books_month_drift).

Запуск (сервер, из каталога установки):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_operator_audit.py --db instance\\transport.db

Выход -- xlsx (--out, по умолчанию drone_operator_audit.xlsx рядом с базой):
  1. «Сводка»          -- месяц: корзины телеметрии, книги, инвариант.
  2. «Оператор x месяц» -- матрица с разницей и процентом.
  3. «Летал -- записи нет»   -- класс А по дням.
  4. «Запись -- полёта нет»  -- класс В по работам, со ссылкой в книгу.
  5. «Без назначения»  -- машино-дни, не покрытые никем.
  6. «Спорное покрытие» -- дни под двумя назначениями (в норме пусто).
  7. «Назначения»      -- использованные назначения с меткой источника.

Числа сверять разбором openpyxl, а не глазами по скриншоту.
Консоль -- только ASCII.
"""

import argparse
import datetime
import os
import sqlite3
import sys
from collections import defaultdict

DAY_EXPR = "strftime('%Y-%m-%d', datetime(started_at, '+300 minutes'))"
INF = '9999-12-31'
UNRESOLVED = '-- не распознан --'
# [REASON]: работы длиннее 60 дней -- мусор дат, а не работа (читатели книг
# режут так же); их дни не считаются покрытыми, чтобы одна кривая строка не
# «закрыла» оператору полгода.
MAX_SPAN_DAYS = 60


def connect_ro(db_path):
    if not os.path.exists(db_path):
        # [REASON]: sqlite3.connect СОЗДАЁТ пустую базу, если файла нет.
        print('ERROR: database not found at %s - refusing to run.' % db_path)
        sys.exit(2)
    uri = 'file:%s?mode=ro' % db_path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def month_range(month_from, month_to):
    out = []
    year, month = int(month_from[:4]), int(month_from[5:7])
    while '%04d-%02d' % (year, month) <= month_to:
        out.append('%04d-%02d' % (year, month))
        month += 1
        if month == 13:
            year, month = year + 1, 1
    return out


def iter_days(date_from, date_to):
    a = datetime.date.fromisoformat(date_from)
    b = datetime.date.fromisoformat(date_to)
    for i in range((b - a).days + 1):
        yield (a + datetime.timedelta(days=i)).isoformat()


def source_of(note):
    note = note or ''
    if note.startswith('[') and ']' in note:
        return note[:note.index(']') + 1]
    return '--'


def load(conn, month_from, month_to):
    """Читает базу в структуры. Ничего не пишет."""
    assignments = conn.execute(
        'SELECT a.id, u.number, o.full_name, a.date_from, a.date_to, a.note '
        'FROM drone_operator_assignments a '
        'JOIN drone_units u ON u.id = a.drone_unit_id '
        'JOIN drone_operators o ON o.id = a.operator_id '
        'ORDER BY u.number, a.date_from').fetchall()

    flight_days = conn.execute(
        'SELECT u.number, %s AS day, COUNT(*), '
        '       COALESCE(SUM(f.area_ha), 0) '
        'FROM drone_flights f '
        'LEFT JOIN drone_units u ON u.id = f.drone_unit_id '
        "WHERE %s BETWEEN ? || '-01' AND ? || '-31' "
        'GROUP BY u.number, day' % (DAY_EXPR, DAY_EXPR),
        (month_from, month_to)).fetchall()

    works = conn.execute(
        'SELECT w.id, o.full_name, w.work_date_from, w.work_date_to, '
        '       w.period_month, COALESCE(w.area_ha, 0), '
        '       w.source_file, w.source_sheet, w.source_row '
        'FROM drone_works w '
        'LEFT JOIN drone_operators o ON o.id = w.drone_operator_id'
    ).fetchall()
    return assignments, flight_days, works


def covering(assignments, number, day):
    return [a for a in assignments
            if a[1] == number and a[3] <= day and (a[4] or INF) >= day]


def audit(assignments, flight_days, works, month_from, month_to):
    """Раскладывает гектары по корзинам. Чистая функция, легко проверяемая."""
    tele = defaultdict(float)             # (op, month) -> ha
    disputed = []                         # (day, number, [ops], flights, ha)
    unassigned = defaultdict(lambda: [0, 0.0])   # (month, number) -> [f, ha]
    no_machine = defaultdict(lambda: [0, 0.0])   # month -> [f, ha]
    total = defaultdict(float)            # month -> ha телеметрии всего
    op_flight_days = defaultdict(dict)    # op -> day -> [(number, f, ha)]

    for number, day, flights, ha in flight_days:
        month = day[:7]
        total[month] += ha
        if number is None:
            no_machine[month][0] += flights
            no_machine[month][1] += ha
            continue
        owners = covering(assignments, number, day)
        if not owners:
            unassigned[(month, number)][0] += flights
            unassigned[(month, number)][1] += ha
        elif len(owners) == 1:
            op = owners[0][2]
            tele[(op, month)] += ha
            op_flight_days[op].setdefault(day, []).append(
                (number, flights, ha))
        else:
            names = sorted({a[2] for a in owners})
            disputed.append((day, number, names, flights, ha))
            # День остаётся лётным для КАЖДОГО из претендентов: класс А
            # не должен обвинять человека в «незаписанном» дне, который
            # аудит сам же не смог отдать ему в матрице.
            for op in names:
                op_flight_days[op].setdefault(day, []).append(
                    (number, flights, ha))

    book_dated = defaultdict(float)       # (op, month) -> ha
    book_undated = defaultdict(float)     # (op, month) -> ha
    op_book_days = defaultdict(set)       # op -> {day}
    no_flight_works = []                  # класс В

    months = set(month_range(month_from, month_to))
    for (_id, name, date_from, date_to, period, ha, src, sheet,
         row) in works:
        op = name or UNRESOLVED
        if not date_from:
            if period in months:
                book_undated[(op, period)] += ha
            continue
        month = date_from[:7]
        if month not in months:
            continue
        book_dated[(op, month)] += ha
        date_to = date_to or date_from
        try:
            span = (datetime.date.fromisoformat(date_to)
                    - datetime.date.fromisoformat(date_from)).days
        except ValueError:
            span = MAX_SPAN_DAYS + 1
        if 0 <= span <= MAX_SPAN_DAYS:
            days = list(iter_days(date_from, date_to))
            for day in days:
                op_book_days[op].add(day)
            if name and not any(day in op_flight_days[name]
                                for day in days):
                no_flight_works.append(
                    (op, month, src, sheet, row, date_from, date_to, ha))

    unrecorded = []                       # класс А
    for op, days in op_flight_days.items():
        for day, items in sorted(days.items()):
            if day not in op_book_days[op]:
                flights = sum(i[1] for i in items)
                ha = sum(i[2] for i in items)
                numbers = ', '.join(str(i[0]) for i in items)
                unrecorded.append((day[:7], day, op, numbers, flights, ha))

    return {
        'tele': tele, 'disputed': disputed, 'unassigned': unassigned,
        'no_machine': no_machine, 'total': total,
        'book_dated': book_dated, 'book_undated': book_undated,
        'unrecorded': unrecorded, 'no_flight_works': no_flight_works,
    }


def invariant_rows(result, months):
    """Месяц -> корзины и проверка «ничего не потеряно»."""
    out = []
    for month in months:
        by_ops = sum(ha for (_op, m), ha in result['tele'].items()
                     if m == month)
        disp = sum(ha for day, _n, _o, _f, ha in result['disputed']
                   if day[:7] == month)
        unas = sum(v[1] for (m, _n), v in result['unassigned'].items()
                   if m == month)
        nomach = result['no_machine'].get(month, [0, 0.0])[1]
        total = result['total'].get(month, 0.0)
        ok = abs(by_ops + disp + unas + nomach - total) < 0.01
        out.append((month, total, by_ops, disp, unas, nomach, ok))
    return out


def write_xlsx(path, result, months, assignments):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = openpyxl.Workbook()

    def sheet(title, header, widths):
        ws = wb.create_sheet(title)
        for i, name in enumerate(header, 1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.font = bold
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
        ws.freeze_panes = 'A2'
        return ws

    wb.remove(wb.active)

    ws = sheet('Сводка',
               ['Месяц', 'Га телеметрии', 'По операторам', 'Спорное',
                'Без назначения', 'Без машины', 'Га книг (дата)',
                'Га книг (без даты)', 'Инвариант'],
               [10, 14, 14, 10, 15, 12, 14, 16, 34])
    for (month, total, by_ops, disp, unas, nomach,
         ok) in invariant_rows(result, months):
        bd = sum(ha for (_o, m), ha in result['book_dated'].items()
                 if m == month)
        bu = sum(ha for (_o, m), ha in result['book_undated'].items()
                 if m == month)
        ws.append([month, round(total, 2), round(by_ops, 2), round(disp, 2),
                   round(unas, 2), round(nomach, 2), round(bd, 2),
                   round(bu, 2),
                   'сходится' if ok else 'НЕ СХОДИТСЯ -- отчёту не верить'])

    ws = sheet('Оператор x месяц',
               ['Оператор', 'Месяц', 'Га телеметрии', 'Га книг (дата)',
                'Га книг (без даты)', 'Га книг всего',
                'Разница (книга - телеметрия)', '% книги от телеметрии'],
               [26, 10, 14, 14, 16, 13, 24, 20])
    keys = sorted(set(list(result['tele'])
                      + list(result['book_dated'])
                      + list(result['book_undated'])),
                  key=lambda k: (k[1], k[0]))
    for op, month in keys:
        t = result['tele'].get((op, month), 0.0)
        bd = result['book_dated'].get((op, month), 0.0)
        bu = result['book_undated'].get((op, month), 0.0)
        ws.append([op, month, round(t, 2), round(bd, 2), round(bu, 2),
                   round(bd + bu, 2), round(bd + bu - t, 2),
                   round(100 * (bd + bu) / t, 1) if t else None])

    ws = sheet('Летал -- записи нет',
               ['Месяц', 'День', 'Оператор', 'Машины', 'Вылетов', 'Га'],
               [10, 12, 26, 10, 9, 10])
    for row in sorted(result['unrecorded']):
        month, day, op, numbers, flights, ha = row
        ws.append([month, day, op, numbers, flights, round(ha, 2)])

    ws = sheet('Запись -- полёта нет',
               ['Оператор', 'Месяц', 'Книга', 'Лист', 'Строка',
                'Дата с', 'Дата по', 'Га'],
               [26, 10, 40, 22, 8, 12, 12, 10])
    for row in sorted(result['no_flight_works']):
        op, month, src, sh, srow, date_from, date_to, ha = row
        ws.append([op, month, src, sh, srow, date_from, date_to,
                   round(ha, 2)])

    ws = sheet('Без назначения',
               ['Месяц', 'Машина', 'Вылетов', 'Га'], [10, 9, 9, 10])
    for (month, number), (flights, ha) in sorted(
            result['unassigned'].items()):
        ws.append([month, number, flights, round(ha, 2)])
    for month, (flights, ha) in sorted(result['no_machine'].items()):
        ws.append([month, 'без машины', flights, round(ha, 2)])

    ws = sheet('Спорное покрытие',
               ['День', 'Машина', 'Операторы', 'Вылетов', 'Га'],
               [12, 9, 40, 9, 10])
    for day, number, names, flights, ha in sorted(result['disputed']):
        ws.append([day, number, '; '.join(names), flights, round(ha, 2)])

    ws = sheet('Назначения',
               ['id', 'Машина', 'Оператор', 'С', 'По', 'Источник'],
               [6, 9, 26, 12, 12, 13])
    for aid, number, name, date_from, date_to, note in assignments:
        ws.append([aid, number, name, date_from, date_to or 'по сей день',
                   source_of(note)])

    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'))
    parser.add_argument('--month-from', default='2025-08')
    parser.add_argument('--month-to', default='2026-08')
    parser.add_argument('--out', default=None)
    args = parser.parse_args()

    conn = connect_ro(args.db)
    try:
        assignments, flight_days, works = load(conn, args.month_from,
                                               args.month_to)
    finally:
        conn.close()

    months = month_range(args.month_from, args.month_to)
    result = audit(assignments, flight_days, works, args.month_from,
                   args.month_to)

    bad = 0
    for (month, total, by_ops, disp, unas, nomach,
         ok) in invariant_rows(result, months):
        if not (total or by_ops):
            continue
        print('%s  tele %9.2f = ops %9.2f + disputed %7.2f + '
              'unassigned %8.2f + no-machine %7.2f  [%s]'
              % (month, total, by_ops, disp, unas, nomach,
                 'ok' if ok else 'MISMATCH'))
        if not ok:
            bad += 1
    print('Class A (flew, not booked): %d day-rows, %.2f ha'
          % (len(result['unrecorded']),
             sum(r[5] for r in result['unrecorded'])))
    print('Class C (booked, no flight): %d work-rows, %.2f ha'
          % (len(result['no_flight_works']),
             sum(r[7] for r in result['no_flight_works'])))
    print('Disputed machine-days: %d' % len(result['disputed']))

    out = args.out or os.path.join(
        os.path.dirname(os.path.abspath(args.db)) or '.',
        'drone_operator_audit.xlsx')
    write_xlsx(out, result, months, assignments)
    print('Report written to %s' % out)
    if bad:
        print('ERROR: %d month(s) failed the invariant - do not trust '
              'this report.' % bad)
        sys.exit(1)


if __name__ == '__main__':
    main()
