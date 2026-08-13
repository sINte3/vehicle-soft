#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tools/drone_body_code_probe.py -- DRONE-BODYCODE-001: можно ли привязать
вылет к БОРТУ вместо строки ника. ЧИТАЕТ И НЕ ПИШЕТ НИЧЕГО.

Зачем. Привязка вылета к машине идёт по строке ника (`drone_nicknames`), а
ник в DJI уникален глобально и не имеет срока действия: одно написание = одна
машина навсегда. Но ярлыки в DJI переезжали между бортами. Доказано снимками
владельца от 2026-08-13: борт, который СЕГОДНЯ называется «14 Servis»,
27.09.2025 подписан «PeshkShodi», а 29.09.2025 -- уже «15 Servis». Итог
месяца из-за этого верен всегда, а распределение по бортам уезжает:
октябрь-2025 -- перестановка по кругу на 205.98 га, март-2026 -- 42.16 га
чужих на машине №12.

Настоящий идентификатор борта существует: в выгрузке djiag.com это колонка
`Body Code` (например 64TBL8H00200BA), она же `drone_units.hardware_id`. Её
НЕ надо путать с колонкой `Serial Number` (например R1790694323) -- это
идентификатор ВЫЛЕТА, он же `drone_flights.serial_number`, 22 855 различных
значений на 22 855 вылетов (ловушка записана в CLAUDE.md).

Скрипт отвечает на три вопроса и ничего не решает за человека:

  A. Есть ли серийник планера в САМОМ payload DJI (`drone_flights.raw_json`)?
     Если да -- починка разовая и вечная: пересчитать `drone_unit_id` по
     этому полю. Если нет -- остаётся мост через выгрузку (вопрос C).
  B. Сходятся ли `hardware_id` в карточках парка с теми Body Code, которые
     реально встречаются в выгрузке?
  C. Если подать выгрузку (`--export`): сколько её строк находится в базе по
     `serial_number`, и сколько вылетов сменили бы машину при привязке по
     Body Code. Разбивка по месяцам и парам «откуда -> куда».

ЧТО ДЕЛАТЬ С ВЫВОДОМ. Ответ «A: YES» разрешает миграцию, пересчитывающую
привязку по полю payload. Ответ «A: NO» означает, что источник истины --
выгрузка, и чинить надо ею. Раздел C -- предварительный просмотр будущей
починки: он показывает ровно те вылеты, которые переедут, ДО того как
что-либо будет записано.

ВАЖНО О ПОРЯДКЕ. Если раздел B показал перепутанные `hardware_id`, чинить
карточки надо ПЕРВЫМ шагом (migrate_drones_hardware_id_fix_001.py), иначе
привязка по Body Code разложит вылеты по тем же перепутанным машинам --
уверенно и неправильно.

Единицы. В выгрузке колонка `Sprayed area` -- в МУ, а не в гектарах: 1 га =
15 му. Проверено на 84 зимних вылетах, 386.15 му / 15 = 25.74 га, и это
ровно итог нашей базы за тот же период. Скрипт делит на 15 и называет
результат гектарами; без деления сверка завышена в пятнадцать раз.

Запуск (служба останавливаться НЕ должна -- открывается только чтение):

  cd C:\\transport-report
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_body_code_probe.py --db instance\\transport.db
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_body_code_probe.py --db instance\\transport.db --export C:\\qa\\dji_export_sentyabr.xlsx --report C:\\qa\\bodycode_probe.txt

ОТКАТ НЕ ТРЕБУЕТСЯ: база открывается через file:-URI в режиме `mode=ro`,
транзакций не открывается, ни одного INSERT/UPDATE/DELETE в файле нет.
Откат кода -- git revert коммита.

Консоль -- только ASCII (службы на сервере пишут журнал в cp1251). Полный
читаемый вывод с кириллицей -- в отчёте UTF-8 через --report.
"""

import argparse
import collections
import datetime
import json
import os
import re
import sqlite3
import sys

# [REASON]: выгрузка djiag.com отдаёт площадь в МУ, а не в гектарах. 1 га =
# 15 му. Константа проверена на полной зимней выгрузке (84 вылета, четыре
# борта): 386.15 му / 15 = 25.7433 га против 25.74 га в нашей базе, и так же
# сходится по каждому борту отдельно. Без этого деления любая сверка с
# выгрузкой завышена ровно в пятнадцать раз и выглядит правдоподобно.
MU_PER_HECTARE = 15.0

# Смещение показа времени: вылеты хранятся в UTC, холдинг живёт в UTC+5.
DISPLAY_TZ_HOURS = 5

# Серийник планера DJI Agras T40: «64TBL» + буквы и цифры. Шаблон нарочно
# широкий -- он ищет КАНДИДАТОВ в payload, а не проверяет корректность.
BODY_CODE_RE = re.compile(r'^[0-9A-Z]{12,20}$')

# Сколько строк payload разбирать в поиске поля с серийником планера.
RAW_JSON_SAMPLE = 800

# Названия колонок выгрузки. DJI пишет их по-английски и с опечаткой в
# `Pliot Name`; здесь нужны только четыре.
COL_FLIGHT_TIME = 'Flight time'
COL_AIRCRAFT = 'Aircraft name'
COL_SERIAL = 'Serial Number'
COL_BODY = 'Body Code'
COL_AREA = 'Sprayed area'


def open_readonly(db_path):
    """Только чтение, через file:-URI. Отсутствие файла -- код 2."""
    if not os.path.exists(db_path):
        sys.stderr.write('ERROR: database not found: %s\n' % db_path)
        sys.exit(2)
    uri = 'file:%s?mode=ro' % db_path.replace('?', '%3f').replace('#', '%23')
    return sqlite3.connect(uri, uri=True)


def table_exists(conn, name):
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (name,)).fetchone()
    return row is not None


def load_units(conn):
    """Парк: номер -> (id, hardware_id). Пустой hardware_id -> None."""
    units = {}
    for uid, number, hw in conn.execute(
            'SELECT id, number, hardware_id FROM drone_units ORDER BY number'):
        units[number] = (uid, (hw or '').strip() or None)
    return units


def probe_raw_json(conn):
    """Раздел A: ищет в payload поле, похожее на серийник планера.

    Возвращает (ключи payload, кандидаты, сколько строк разобрано).
    Кандидат -- ключ, у которого ВСЕ непустые значения в выборке подходят под
    шаблон серийника И различных значений заметно меньше, чем строк: борт
    повторяется из вылета в вылет, а идентификатор вылета -- нет.
    """
    rows = conn.execute(
        'SELECT raw_json FROM drone_flights '
        'WHERE raw_json IS NOT NULL AND raw_json != "" '
        'ORDER BY id LIMIT ?', (RAW_JSON_SAMPLE,)).fetchall()
    keys = collections.Counter()
    values = collections.defaultdict(set)
    parsed = 0
    for (blob,) in rows:
        try:
            obj = json.loads(blob)
        except (ValueError, TypeError):
            continue
        if not isinstance(obj, dict):
            continue
        parsed += 1
        for key, value in obj.items():
            keys[key] += 1
            if isinstance(value, str) and value:
                # Множество ограничено: интересна МОЩНОСТЬ, а не содержимое.
                if len(values[key]) < 200:
                    values[key].add(value.strip())
    candidates = []
    for key, seen in values.items():
        if not seen:
            continue
        if all(BODY_CODE_RE.match(v.upper()) for v in seen):
            # Идентификатор вылета уникален на строку; борт -- нет.
            candidates.append((key, len(seen), sorted(seen)[:3]))
    candidates.sort(key=lambda item: item[1])
    return keys, candidates, parsed


def probe_serial_shape(conn):
    """Сколько различных значений serial_number на сколько вылетов."""
    total = conn.execute('SELECT COUNT(*) FROM drone_flights').fetchone()[0]
    distinct = conn.execute(
        'SELECT COUNT(DISTINCT serial_number) FROM drone_flights '
        'WHERE serial_number IS NOT NULL AND serial_number != ""').fetchone()[0]
    filled = conn.execute(
        'SELECT COUNT(*) FROM drone_flights '
        'WHERE serial_number IS NOT NULL AND serial_number != ""').fetchone()[0]
    return total, filled, distinct


def read_export(path):
    """Строки выгрузки djiag.com. Возвращает список словарей.

    Читается openpyxl -- он уже зависимость проекта (excel_export.py), новых
    в requirements.txt не добавляется.
    """
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write('ERROR: openpyxl is required to read %s\n' % path)
        sys.exit(3)
    if not os.path.exists(path):
        sys.stderr.write('ERROR: export not found: %s\n' % path)
        sys.exit(2)
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    header = None
    out = []
    for raw in sheet.iter_rows(values_only=True):
        if header is None:
            header = [(str(c).strip() if c is not None else '') for c in raw]
            missing = [c for c in (COL_SERIAL, COL_BODY)
                       if c not in header]
            if missing:
                sys.stderr.write(
                    'ERROR: %s has no column(s): %s\n'
                    % (os.path.basename(path), ', '.join(missing)))
                sys.exit(3)
            continue
        row = dict(zip(header, raw))
        serial = (str(row.get(COL_SERIAL) or '')).strip()
        body = (str(row.get(COL_BODY) or '')).strip()
        if not serial and not body:
            continue
        area_mu = row.get(COL_AREA)
        try:
            area_mu = float(area_mu) if area_mu is not None else 0.0
        except (TypeError, ValueError):
            area_mu = 0.0
        out.append({
            'serial': serial,
            'body': body,
            'aircraft': (str(row.get(COL_AIRCRAFT) or '')).strip(),
            'flight_time': (str(row.get(COL_FLIGHT_TIME) or '')).strip(),
            'area_ha': area_mu / MU_PER_HECTARE,
        })
    book.close()
    return out


def month_of(flight_time):
    """'2026-02-25 13:58:08-13:58:59' -> '2026-02'. Иначе '(без даты)'."""
    text = (flight_time or '').strip()
    if len(text) >= 7 and text[4] == '-':
        return text[:7]
    return '(no date)'


def join_export(conn, rows, units):
    """Раздел C: сопоставление выгрузки с базой по serial_number.

    Ничего не пишет. Возвращает словарь с итогами и списком переносов.
    """
    by_body = {}
    for number, (uid, hw) in units.items():
        if hw:
            by_body.setdefault(hw.upper(), []).append((number, uid))

    serials = [r['serial'] for r in rows if r['serial']]
    found = {}
    # IN (...) режется на порции: SQLite ограничивает число параметров.
    chunk = 400
    for start in range(0, len(serials), chunk):
        part = serials[start:start + chunk]
        marks = ','.join('?' * len(part))
        sql = ('SELECT serial_number, id, drone_unit_id, started_at, area_ha '
               'FROM drone_flights WHERE serial_number IN (%s)' % marks)
        for serial, fid, unit_id, started, area in conn.execute(sql, part):
            found.setdefault(serial, []).append(
                {'id': fid, 'unit_id': unit_id, 'started_at': started,
                 'area_ha': area or 0.0})

    unit_number_by_id = {uid: number for number, (uid, _hw) in units.items()}
    stats = {
        'export_rows': len(rows),
        'export_ha': sum(r['area_ha'] for r in rows),
        'matched_rows': 0,
        'matched_ha': 0.0,
        'missing_rows': 0,
        'ambiguous_serials': 0,
        'body_unknown': collections.Counter(),
        'moves': [],
        'same': 0,
        'was_null': 0,
    }
    for row in rows:
        hits = found.get(row['serial'])
        if not hits:
            stats['missing_rows'] += 1
            continue
        if len(hits) > 1:
            stats['ambiguous_serials'] += 1
            continue
        hit = hits[0]
        stats['matched_rows'] += 1
        stats['matched_ha'] += hit['area_ha']
        target = by_body.get(row['body'].upper())
        if not target:
            stats['body_unknown'][row['body']] += 1
            continue
        if len(target) > 1:
            stats['body_unknown']['DUPLICATE hardware_id: %s' % row['body']] += 1
            continue
        want_number, want_uid = target[0]
        have_uid = hit['unit_id']
        if have_uid == want_uid:
            stats['same'] += 1
            continue
        if have_uid is None:
            stats['was_null'] += 1
        stats['moves'].append({
            'flight_id': hit['id'],
            'month': month_of(row['flight_time']),
            'from_number': unit_number_by_id.get(have_uid),
            'to_number': want_number,
            'area_ha': hit['area_ha'],
            'aircraft': row['aircraft'],
            'body': row['body'],
        })
    return stats


def format_report(db_path, units, keys, candidates, parsed, shape,
                  export_paths, joined):
    """Полный отчёт (UTF-8, кириллица разрешена)."""
    lines = []
    add = lines.append
    now = (datetime.datetime.utcnow()
           + datetime.timedelta(hours=DISPLAY_TZ_HOURS))
    add('DRONE-BODYCODE-001 — разведка привязки по серийнику планера')
    add('База: %s' % db_path)
    add('Снято: %s (UTC+%d)' % (now.strftime('%Y-%m-%d %H:%M'),
                                DISPLAY_TZ_HOURS))
    add('')
    add('A. ЕСТЬ ЛИ СЕРИЙНИК ПЛАНЕРА В PAYLOAD DJI')
    add('-' * 68)
    total, filled, distinct = shape
    add('Вылетов в базе: %d' % total)
    add('serial_number заполнен у %d, различных значений %d'
        % (filled, distinct))
    if filled and distinct >= filled * 0.9:
        add('  => serial_number — идентификатор ВЫЛЕТА, а не борта '
            '(подтверждает CLAUDE.md).')
    add('Разобрано строк raw_json: %d' % parsed)
    if not parsed:
        add('  raw_json пуст или не разбирается — вывод по разделу A сделать '
            'нельзя.')
    else:
        add('Ключи payload (%d): %s'
            % (len(keys), ', '.join(sorted(keys))))
        if candidates:
            add('')
            add('КАНДИДАТЫ на серийник планера (все значения похожи на код '
                'борта):')
            for key, uniq, sample in candidates:
                add('  %-24s различных значений: %-5d пример: %s'
                    % (key, uniq, ', '.join(sample)))
            add('')
            add('ВЫВОД A: поле-кандидат есть. Если различных значений около '
                'числа бортов (15), это он — починка делается миграцией по '
                'этому полю, выгрузка не нужна.')
        else:
            add('')
            add('ВЫВОД A: поля с серийником планера в payload НЕТ. Источник '
                'истины — выгрузка djiag.com (колонка Body Code); чинить '
                'через неё.')
    add('')
    add('B. КАРТОЧКИ ПАРКА: hardware_id')
    add('-' * 68)
    for number in sorted(units):
        uid, hw = units[number]
        add('  №%-3d id=%-4d hardware_id=%s' % (number, uid, hw or '(пусто)'))
    empty = [n for n in units if not units[n][1]]
    if empty:
        add('  ПУСТО у машин: %s — привязка по борту для них невозможна, '
            'пока не заполнены.' % ', '.join('№%d' % n for n in sorted(empty)))
    dupes = collections.Counter(
        units[n][1].upper() for n in units if units[n][1])
    for hw, cnt in dupes.items():
        if cnt > 1:
            add('  ДУБЛЬ hardware_id %s стоит у %d машин — исправить до '
                'привязки по борту.' % (hw, cnt))
    add('')
    add('C. СВЕРКА С ВЫГРУЗКОЙ')
    add('-' * 68)
    if not export_paths:
        add('  Выгрузка не подана (--export). Раздел пропущен.')
        return '\n'.join(lines) + '\n'
    add('Файлы: %s' % ', '.join(os.path.basename(p) for p in export_paths))
    add('Строк в выгрузке: %d, гектаров (му/15): %.2f'
        % (joined['export_rows'], joined['export_ha']))
    add('Нашлось в базе по serial_number: %d (%.2f га)'
        % (joined['matched_rows'], joined['matched_ha']))
    add('Не нашлось: %d' % joined['missing_rows'])
    if joined['ambiguous_serials']:
        add('serial_number встретился в базе больше одного раза: %d — эти '
            'строки пропущены.' % joined['ambiguous_serials'])
    add('Машина уже верна: %d' % joined['same'])
    add('Переехали бы: %d вылетов (%.2f га), из них с пустой машиной: %d'
        % (len(joined['moves']),
           sum(m['area_ha'] for m in joined['moves']),
           joined['was_null']))
    if joined['body_unknown']:
        add('')
        add('Body Code без машины в парке:')
        for body, cnt in joined['body_unknown'].most_common():
            add('  %-24s %d строк' % (body, cnt))
    if joined['moves']:
        add('')
        add('ПРЕДПОЛАГАЕМЫЕ ПЕРЕНОСЫ (месяц: откуда -> куда):')
        grouped = collections.defaultdict(lambda: [0, 0.0])
        for move in joined['moves']:
            key = (move['month'],
                   move['from_number'] if move['from_number'] else 'нет',
                   move['to_number'])
            grouped[key][0] += 1
            grouped[key][1] += move['area_ha']
        for key in sorted(grouped, key=lambda k: (k[0], str(k[1]), k[2])):
            month, src, dst = key
            cnt, area = grouped[key]
            add('  %s  №%s -> №%s : %4d вылетов, %8.2f га'
                % (month, src, dst, cnt, area))
    return '\n'.join(lines) + '\n'


def print_ascii(units, candidates, parsed, shape, export_paths, joined):
    """Консоль: только ASCII."""
    total, filled, distinct = shape
    print('DRONE-BODYCODE-001 probe')
    print('flights=%d serial_number filled=%d distinct=%d'
          % (total, filled, distinct))
    print('raw_json rows parsed=%d' % parsed)
    if not parsed:
        print('A: UNKNOWN (no raw_json parsed)')
    elif candidates:
        print('A: CANDIDATE FIELD(S) FOUND')
        for key, uniq, sample in candidates:
            print('   key=%s distinct=%d sample=%s'
                  % (key, uniq, ','.join(sample)))
    else:
        print('A: NO airframe serial in payload -- use the djiag.com export')
    empty = [n for n in units if not units[n][1]]
    print('B: units=%d, hardware_id empty for %d' % (len(units), len(empty)))
    if empty:
        print('   empty: %s' % ','.join(str(n) for n in sorted(empty)))
    if not export_paths:
        print('C: skipped (no --export)')
        return
    print('C: export rows=%d matched=%d missing=%d ambiguous=%d'
          % (joined['export_rows'], joined['matched_rows'],
             joined['missing_rows'], joined['ambiguous_serials']))
    print('   already correct=%d, would move=%d (%.2f ha), of them null=%d'
          % (joined['same'], len(joined['moves']),
             sum(m['area_ha'] for m in joined['moves']), joined['was_null']))
    grouped = collections.Counter()
    for move in joined['moves']:
        grouped[(move['month'],
                 move['from_number'] or 0,
                 move['to_number'])] += 1
    for key in sorted(grouped):
        print('   %s unit %s -> %s : %d flights'
              % (key[0], key[1] or 'NULL', key[2], grouped[key]))


def main():
    parser = argparse.ArgumentParser(
        description='DRONE-BODYCODE-001: can flights be bound to the airframe '
                    'serial instead of the nickname? Read-only.')
    parser.add_argument('--db', default=os.path.join('instance',
                                                     'transport.db'),
                        help='path to transport.db (read-only)')
    parser.add_argument('--export', action='append', default=[],
                        help='djiag.com export .xlsx (repeatable)')
    parser.add_argument('--report', default=None,
                        help='write the full UTF-8 report to this file')
    args = parser.parse_args()

    conn = open_readonly(args.db)
    try:
        for name in ('drone_units', 'drone_flights'):
            if not table_exists(conn, name):
                sys.stderr.write('ERROR: table %s is missing -- is this the '
                                 'right database?\n' % name)
                return 1
        units = load_units(conn)
        if not units:
            sys.stderr.write('ERROR: drone_units is empty\n')
            return 1
        keys, candidates, parsed = probe_raw_json(conn)
        shape = probe_serial_shape(conn)
        rows = []
        for path in args.export:
            rows.extend(read_export(path))
        joined = join_export(conn, rows, units) if rows else None
    finally:
        conn.close()

    print_ascii(units, candidates, parsed, shape, args.export, joined)
    if args.report:
        text = format_report(args.db, units, keys, candidates, parsed, shape,
                             args.export, joined)
        with open(args.report, 'w', encoding='utf-8') as handle:
            handle.write(text)
        print('report written: %s' % args.report)
    return 0


if __name__ == '__main__':
    sys.exit(main())
