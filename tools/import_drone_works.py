#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tools/import_drone_works.py -- DRONE-WORKS-001 import of the dispatchers'
Excel books into drone_works / drone_customers / drone_customer_aliases.

Since September 2025 the whole commercial side of the drone operation has
lived in Excel files on somebody's laptop: about 700 rows, 526 farms,
12 767 ha, eleven months. This tool moves them into the database. The flight
tables, the ingest endpoint and the collector are not touched by it.

--dry-run IS THE DEFAULT. Running without --apply parses, reports and writes
nothing, so it is safe to point at a production database. The real first run
is the owner's, and the summary is printed so it can be compared against a
prediction written beforehand.

Rules this tool obeys, each of which cost a wrong number when the real files
were parsed by hand on 2026-08-03:

  a) A DATA ROW CAN LOOK LIKE A HEADER. The «Изоҳ» column contains sentences
     mentioning «майдон», so a header detector that searches for that word
     alone matched every data row and parsed whole sheets as empty --
     understating September by 1 548 ha. A row is a header only when ALL
     THREE hold: the first cell is exactly '№', some cell in the first four
     contains «ФХ номи», and some cell in the first six contains «Майдон» OR
     «Мадон» (the typo is in the real file).

  b) JUNK BELOW THE TOTALS ROW. After «Жами сумма» some sheets continue with
     rows like 'Доллар | 12220'; read as hectares that is 12 220 ha in one
     row. A data row must start with the dispatchers' own row counter (a
     number) and carry 0 < area_ha <= 1000.

  c) TWO PAYMENT BLOCKS PER SHEET, in either order, marked by lines reading
     «Справка (<month> ойи)» and «Нақд (<month> ойи)». The block a row falls
     in determines payment_type. Internal work appears as its own block
     («Тизим корхонаси») or as rows whose customer is a holding subdivision.

  d) DUPLICATE SHEETS AND DUPLICATE FILES. One file carries «свод ичи
     (Беҳзод)» and «свод ичи » with identical content; two other files
     overlap partially. Rows matching on
     (customer_raw, area_ha, work_date_from, amount) across the whole run are
     imported once and every skipped row is listed with BOTH provenances.

  e) DATE FORMATS. A datetime; '23-24.03.2026'; '13,16,17.05.2026';
     '27,30.03.2026'; '08-09.092025' (the dot is missing); empty. Ranges
     become from/to, comma lists become min/max, the original always stays in
     date_raw. A form that cannot be parsed keeps date_raw with both dates
     NULL and is counted and listed -- never dropped.

  f) PRICES ARE NOT CONSTANT AND ARE NEVER SUBSTITUTED. 200 000 so'm/ha
     dominates but nine other values occur, and the internal tariff is quoted
     as 76 458 on one sheet and 85 633 on another. An empty price cell is
     stored as NULL. It is NEVER computed from amount / area and never
     defaulted: a price this tool invented is indistinguishable from a price
     the farm agreed. The suggestion belongs to the manual entry form.

  g) SHEET TITLES. «свод ичи (Фурқат)», «свод ичи Шохрух», «свод ичи » --
     with brackets, without, sometimes empty. operator_raw comes from the
     title; when the title carries no name the file stem is used instead and
     the report says so for every such sheet.

  h) «Иш хаки» -- the operator's wage -- IS NOT IMPORTED. That is operator
     sub-accounting and was explicitly deferred by the owner.

Console output is ASCII only (NSSM / Windows cp1251 log safety). Everything
that has to name a farm, an operator or a sheet goes into the UTF-8 report
file written next to the manifest (or wherever --report points), because the
names are Cyrillic and the console cannot be trusted to carry them.

Safety:
  - stdlib sqlite3 only, never `from app import app`: create_app() calls
    db.create_all() at import time and would turn this into a writer even in
    --dry-run;
  - refuses to run and exits 2 when the database file is missing --
    sqlite3.connect() would otherwise create an empty one;
  - --dry-run opens the database read-only through a file: URI with mode=ro,
    so a mistake cannot write;
  - --apply writes inside ONE transaction and rolls the whole run back on any
    error; nothing is written outside drone_works, drone_customers and
    drone_customer_aliases.

Usage:

  python tools/import_drone_works.py --dir <folder> \\
      --manifest tools/drone_works_manifest.txt

  python tools/import_drone_works.py --dir <folder> \\
      --manifest tools/drone_works_manifest.txt --apply --batch 2026-08-first

On Windows (PowerShell has no &&, one command per line):

  cd C:\\transport-report
  & "C:\\Program Files\\Python314\\python.exe" tools\\import_drone_works.py --dir C:\\drone-books --manifest tools\\drone_works_manifest.txt
"""

import argparse
import datetime
import os
import re
import sqlite3
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from plate_norm import normalize_plate  # noqa: E402

DEFAULT_DB_PATH = os.path.join(REPO_ROOT, 'instance', 'transport.db')

PAYMENT_CASH = 'cash'
PAYMENT_TRANSFER = 'transfer'
PAYMENT_INTERNAL = 'internal'
# [REASON]: owner's decision of 2026-08-04 -- see parse_sheet. Roughly two
# fifths of the real corpus sits in sheets with no payment block at all.
PAYMENT_UNKNOWN = 'unknown'
PAYMENT_TYPES = (PAYMENT_CASH, PAYMENT_TRANSFER, PAYMENT_INTERNAL,
                 PAYMENT_UNKNOWN)
# --default-payment may not be used to declare rows unknown: that is what
# happens anyway, and offering it as a choice would suggest it does something.
PAYMENT_OVERRIDE_TYPES = (PAYMENT_CASH, PAYMENT_TRANSFER, PAYMENT_INTERNAL)

RECEIVED_KIND_RECEIVED = 'received'
RECEIVED_KIND_OPERATOR_DUE = 'operator_due'

# A single job is never larger than this. See trap (b).
MAX_AREA_HA = 1000.0

# What the owner's hand parser measured over all 28 books on 2026-08-04,
# grouped by the MANIFEST month. Printed next to the actual figures so the run
# can be compared against a prediction written beforehand instead of being
# read as "some numbers came out".
#
# [REASON]: April is the anchor and the only line corroborated by anything
# other than the same hand parser -- 6 388.83 ha agrees with our own DJI
# flight data for April 2026 (6 379.24 ha over 6 403 flights) to 0.15 %. The
# monthly split for September and October WILL differ here, because this tool
# groups by the row's own date and one book straddles the two months; the
# total is the firm number and the monthly rows are indicative. These figures
# come from a parser that has been wrong twice already: a disagreement is a
# finding to report, not a number to force.
EXPECTED_BY_MONTH = (
    ('2025-09', 378, 4937.73),
    ('2025-10', 28, 790.90),
    ('2026-03', 46, 1378.48),
    ('2026-04', 331, 6388.83),
    ('2026-05', 106, 2179.72),
    ('2026-07', 6, 62.73),
)
EXPECTED_TOTAL_ROWS = 895
EXPECTED_TOTAL_AREA = 15738.39
EXPECTED_ANCHOR_MONTH = '2026-04'


DATE_KIND_EXACT = 'date'      # one real day, from a datetime cell or dd.mm.yyyy
DATE_KIND_SPAN = 'span'       # a text range or comma list over several days
DATE_KIND_NONE = 'none'       # the cell was empty
DATE_KIND_UNPARSED = 'unparsed'  # a form this tool refuses to guess at


# ─── Normalisation ───────────────────────────────────────────────────────────

# [REASON]: an UZBEK ORTHOGRAPHIC fold applied BEFORE plate_norm's homoglyph
# fold, not instead of it. plate_norm.normalize_plate() folds the twelve
# Cyrillic/Latin homoglyph pairs, upper-cases and strips punctuation, and it
# is called here rather than reimplemented. What it deliberately does not do
# is fold the Uzbek letters, and the dispatchers write the same man as
# «Фурқат» and «Фуркат», «Беҳзод» and «Бехзод». Without this pass those are
# two different people. The target letters are the plain Cyrillic ones, so
# the result still passes through the homoglyph table unchanged in meaning.
_UZ_FOLD = {
    'Қ': 'К', 'Ғ': 'Г', 'Ҳ': 'Х', 'Ў': 'У', 'Ҷ': 'Ж', 'Ҝ': 'К',
    'қ': 'к', 'ғ': 'г', 'ҳ': 'х', 'ў': 'у', 'ҷ': 'ж',
    'Ъ': '', 'ъ': '', 'Ь': '', 'ь': '', 'Ё': 'Е', 'ё': 'е',
    '\u0451': 'е',
}
_UZ_TABLE = {ord(k): (v or None) for k, v in _UZ_FOLD.items()}

_WS_RE = re.compile(r'\s+')


def uz_fold(value):
    """Uzbek-letter fold, lower-cased, whitespace collapsed. Keyword matching."""
    if value is None:
        return ''
    return _WS_RE.sub(' ', str(value).translate(_UZ_TABLE).lower()).strip()


def name_key(value):
    """Comparison key for a person's name: Uzbek fold, then plate_norm."""
    return normalize_plate(str(value).translate(_UZ_TABLE)) if value else ''


def name_tokens(value):
    """Set of normalised word tokens of a name. Empty set for empty input.

    Tokenised BEFORE normalisation, because normalize_plate strips spaces:
    the dispatchers write «Шохрух» where the directory holds «Хамроев
    Шохрух», and a whole-string comparison could never match the two.
    """
    if not value:
        return set()
    tokens = set()
    for part in _WS_RE.split(str(value).strip()):
        key = name_key(part)
        if key:
            tokens.add(key)
    return tokens


def customer_key(value):
    """normalized_key of a farm spelling: case-folded, whitespace-collapsed.

    [REASON]: exactly what the task specifies for drone_customer_aliases, and
    deliberately NOT the Uzbek fold used for operator names. This key is
    stored in the database and the aliases screen has to be able to reproduce
    it by eye; a fold nobody can see would make two visibly different rows
    collide and look like a bug.
    """
    if value is None:
        return ''
    return _WS_RE.sub(' ', str(value)).strip().casefold()


def _text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value):
    """float(value) or None. Accepts '1 234,56', '1 234.56', '200 000 сум'."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, (datetime.date, datetime.datetime)):
        return None
    text = str(value).strip()
    if not text:
        return None
    # Non-breaking spaces and thin spaces are what Excel puts in a thousands
    # separator when the value was typed as text.
    text = text.replace('\xa0', '').replace('\u202f', '').replace(' ', '')
    text = text.replace(',', '.')
    match = re.match(r'^-?\d+(\.\d+)?', text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _int_counter(value):
    """The dispatchers' own row counter: an integer, nothing else."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = _text(value)
    if re.match(r'^\d+$', text):
        return int(text)
    return None


# ─── Trap (a): header detection ──────────────────────────────────────────────

HEADER_CUSTOMER_MARK = uz_fold('ФХ номи')
HEADER_AREA_MARKS = (uz_fold('Майдон'), uz_fold('Мадон'))


def is_header_row(cells):
    """True only when all three conditions hold. See trap (a).

    A header detector that merely searches for «майдон» matches every data
    row whose «Изоҳ» tells a story about a field, and the sheet parses as
    empty. tools/test_import_drone_works.py monkeypatches this function with
    that naive version and asserts the wrong number appears -- a guard with
    no failing negative control is not covered.
    """
    if not cells:
        return False
    if _text(cells[0]) != '№':
        return False
    if not any(HEADER_CUSTOMER_MARK in uz_fold(c) for c in cells[:4]):
        return False
    return any(mark in uz_fold(c)
               for c in cells[:6] for mark in HEADER_AREA_MARKS)


# ─── Sheet selection ─────────────────────────────────────────────────────────

DETAIL_SHEET_PREFIX = uz_fold('свод ичи')


def is_detail_sheet(title):
    """True only for the per-job detail sheets. Everything else is skipped.

    [REASON]: the 2026-08-04 dry run over the real books offered rows from
    «олинмаган пуллар» and «свод фх» as well as from «свод ичи». They were all
    rejected at the time for a different reason -- no payment block -- which
    hid the problem; once that rejection is lifted they would be IMPORTED, and
    they are duplicates of rows already present in the detail sheet
    («Бахром Хайрулло фх, 33.5 га» is in both) plus «ЖАМИ:» total lines. The
    dedup would catch some of them and silently pick whichever file listing
    reached it first, which is not a decision a tool should make.

    The rule is the sheet TITLE, not its shape: «ОБШИЙ СВОД» and its suffixed
    variants, «олинмаган пуллар», «свод фх», «Свод1», «Расход» and
    «карпаратив почта» are summaries and registers, and the payment split
    lives only in «ОБШИЙ СВОД». A detail sheet may name its operator in
    brackets, after a space, or not at all -- «свод ичи (Фурқат)»,
    «свод ичи Шохрух», «свод ичи » -- and all three are imported.
    """
    return uz_fold(title).startswith(DETAIL_SHEET_PREFIX)


# ─── Column mapping ──────────────────────────────────────────────────────────

# EXACT normalised phrases, not fragments.
#
# [REASON]: the first version matched header fragments, and on the real books
# «суммаси» hit «Хизмат кўрсатиш суммаси» -- the PRICE -- and stored it as the
# amount, while «Жами сумма (обьём)», the real amount, was never mapped at all.
# Every one of the twenty-four layouts reported `amount->3 ... НЕ ОПОЗНАНЫ
# КОЛОНКИ: price`, and the run's amount total came out eleven times smaller
# than its received total, which is arithmetically impossible. Two headers also
# differ by two letters: «Хизмат кўрсатиш СУМмаси» is the price and «Хизмат
# кўрсатиш САНаси» is a date, and seven layouts carry the second.
#
# Within a field the phrases are in PREFERENCE order: the first one present in
# the sheet wins, and any other match is reported rather than silently dropped.
# Across fields the phrases are matched longest-first, so a longer phrase can
# never be shadowed by a shorter one contained in it.
FIELD_PHRASES = (
    ('customer', ('ФХ номи',)),
    ('area', ('Майдон (га)', 'Мадон (га)')),
    ('price', ('Хизмат кўрсатиш суммаси',)),
    ('amount', ('Жами сумма (обьём)', 'Жами сумма', 'Сумма (без НДС)')),
    ('received', ('Кирим қилинган', 'Кирим қилинган (Фактура)',
                  'Кирим қилиши керак')),
    ('operator_due', ('Оператор топшириши керак',)),
    ('date', ('Дрон ишлаган сана', 'Кирим қилинган сана', 'Сана',
              'Хизмат кўрсатиш санаси')),
    ('note', ('*Изоҳ', '*Изоҳ / Йуналиш')),
    ('row_operator', ('Дрон бошқарувчи оператор',)),
)

# Expense columns. ALL of them present in a sheet are summed into other_costs;
# six of the twenty-four layouts carry two at once.
EXPENSE_PHRASES = ('Бошка харажатлар', 'Транспорт харажати', 'ЁММ харажати')
# [REASON]: the vehicle plate follows the header in brackets --
# «Транспорт харажати (Мерс 80 L 422 MA)», «(Газель 80 120 FBA)», «Газел» --
# so this one phrase is matched as a PREFIX. It is the only prefix rule there
# is; everything else is exact, because a prefix rule is how «суммаси» ate the
# price column in the first place.
EXPENSE_PREFIXES = ('Транспорт харажати',)

# Headers that exist, are understood, and are deliberately not imported.
#   Иш хаки              -- the operator's wage, sub-accounting, deferred
#   Сумма ( НДС)         -- the VAT part; «Сумма (без НДС)» is the amount
#   Ишлаган контур рақами -- the field-contour number, no consumer yet
#   №                    -- the dispatchers' row counter, read positionally
#                           as cells[0] by the data-row guard, never a field
IGNORED_PHRASES = ('Иш хаки', 'Сумма ( НДС)', 'Ишлаган контур рақами', '№')

# Without these two a sheet cannot be read at all: there is no job without a
# place to read the farm and the area from. Everything else degrades to NULL
# and is reported per sheet.
REQUIRED_FIELDS = ('customer', 'area')
OPTIONAL_FIELDS = ('price', 'amount', 'received', 'date', 'note')

_ZERO_WIDTH = dict.fromkeys(
    [0x00ad, 0x200b, 0x200c, 0x200d, 0xfeff], None)


def header_key(value):
    """Normalised comparison form of a header cell.

    Whitespace of every kind -- including the double space in «Дрон  ишлаган
    сана», embedded newlines and non-breaking spaces -- collapses to one
    space; then trim, lower-case and the same Uzbek-letter fold uz_fold()
    already applies to names. Zero-width characters are removed outright:
    they are invisible in Excel and would make two identical-looking headers
    compare unequal with nothing on screen to explain it.
    """
    if value is None:
        return ''
    text = str(value).translate(_ZERO_WIDTH)
    text = text.replace('\xa0', ' ').replace(' ', ' ')
    return uz_fold(text)


def _build_phrase_index():
    """[(key, field)] over every phrase, longest key first."""
    entries = []
    for field, phrases in FIELD_PHRASES:
        for phrase in phrases:
            entries.append((header_key(phrase), field))
    for phrase in EXPENSE_PHRASES:
        entries.append((header_key(phrase), 'expense'))
    for phrase in IGNORED_PHRASES:
        entries.append((header_key(phrase), 'ignored'))
    entries.sort(key=lambda pair: len(pair[0]), reverse=True)
    return entries


PHRASE_INDEX = _build_phrase_index()
EXPENSE_PREFIX_KEYS = tuple(sorted((header_key(p) for p in EXPENSE_PREFIXES),
                                   key=len, reverse=True))
# Preference order inside a field, as a lookup: phrase key -> its position.
FIELD_PREFERENCE = {}
for _field, _phrases in FIELD_PHRASES:
    for _position, _phrase in enumerate(_phrases):
        FIELD_PREFERENCE[(_field, header_key(_phrase))] = _position


class ColumnMap(object):
    """What one header row resolved to. Everything the report prints."""

    def __init__(self):
        self.fields = {}          # field -> column index
        self.chosen = {}          # field -> the header text that won
        self.expenses = []        # [(index, header)] -- ALL of them, summed
        self.unmatched = []       # [(index, header)] -- named in the report
        self.ignored = []         # [(index, header)] -- known, not imported
        self.shadowed = []        # [(field, index, header)] -- second match

    def get(self, field, default=None):
        return self.fields.get(field, default)

    def __contains__(self, field):
        return field in self.fields

    def __bool__(self):
        return bool(self.fields)

    __nonzero__ = __bool__

    @property
    def received_kind(self):
        """Which header the figure in received_amount came from, or None.

        [REASON]: «Кирим қилинган» is money already handed in; «Оператор
        топшириши керак» is what the operator still owes. They sit in the same
        position and both equal amount - other_costs, so a sheet carrying only
        the second must not be read as collections. When a sheet carries both,
        the actual receipt wins and the other column is reported as shadowed.
        """
        if 'received' in self.fields:
            return 'received'
        if 'operator_due' in self.fields:
            return 'operator_due'
        return None

    @property
    def received_index(self):
        if 'received' in self.fields:
            return self.fields['received']
        return self.fields.get('operator_due')

    def missing_fields(self):
        return [f for f in OPTIONAL_FIELDS
                if f not in self.fields
                and not (f == 'received' and 'operator_due' in self.fields)]


def map_columns(header_cells):
    """Resolve one header row into a ColumnMap. Exact phrases only.

    [REASON]: NO POSITIONAL FALLBACK, and no fragment matching. Guessing that
    the column after the area is the price is exactly what produced a run
    whose amount column held prices. A header that matches nothing keeps its
    field NULL and is NAMED in the report with its sheet -- there were 24
    layouts and there may be a 25th, and the report is how the 25th gets
    found instead of quietly mis-parsed.
    """
    mapping = ColumnMap()
    for index, cell in enumerate(header_cells):
        key = header_key(cell)
        if not key:
            continue
        header = _text(cell)
        field = None
        for phrase_key, candidate in PHRASE_INDEX:
            if key == phrase_key:
                field = candidate
                break
        if field is None:
            for prefix in EXPENSE_PREFIX_KEYS:
                if key.startswith(prefix):
                    field = 'expense'
                    break
        if field is None:
            mapping.unmatched.append((index, header))
            continue
        if field == 'ignored':
            mapping.ignored.append((index, header))
            continue
        if field == 'expense':
            mapping.expenses.append((index, header))
            continue
        if field in mapping.fields:
            # A second column matched the same field. Keep the one whose
            # phrase is earlier in that field's preference list, and report
            # the loser rather than dropping it in silence.
            old_key = header_key(mapping.chosen[field])
            old_rank = FIELD_PREFERENCE.get((field, old_key), 99)
            new_rank = FIELD_PREFERENCE.get((field, key), 99)
            if new_rank < old_rank:
                mapping.shadowed.append(
                    (field, mapping.fields[field], mapping.chosen[field]))
                mapping.fields[field] = index
                mapping.chosen[field] = header
            else:
                mapping.shadowed.append((field, index, header))
            continue
        mapping.fields[field] = index
        mapping.chosen[field] = header
    return mapping


# ─── Trap (c) and (h): block markers ─────────────────────────────────────────

MARK_TRANSFER = uz_fold('Справка')
MARK_CASH = uz_fold('Нақд')
MARK_INTERNAL = uz_fold('Тизим корхонаси')
MARK_WAGE = uz_fold('Иш хаки')


def is_wage_marker(cells):
    """True for «Иш хаки» -- the operator's wage. Never imported, trap (h)."""
    return any(MARK_WAGE in uz_fold(c) for c in cells)


def payment_marker(cells):
    """'cash' / 'transfer' / 'internal' / None for a block-separator row.

    A sheet may start with either block, so there is no default: rows that
    appear before any marker are REJECTED with a reason and listed, never
    assigned a payment type this tool made up. --default-payment is the
    owner's explicit override for that case and is reported loudly.
    """
    folded = ' '.join(uz_fold(c) for c in cells if _text(c))
    if not folded:
        return None
    if MARK_INTERNAL in folded:
        return PAYMENT_INTERNAL
    if MARK_TRANSFER in folded:
        return PAYMENT_TRANSFER
    if MARK_CASH in folded:
        return PAYMENT_CASH
    return None


# ─── Trap (b): what counts as a data row ─────────────────────────────────────

def area_in_range(area):
    """0 < area <= 1000. See trap (b): 'Доллар | 12220' below the totals row."""
    return area is not None and 0 < area <= MAX_AREA_HA


def payment_when_no_block():
    """What a row with no payment block above it becomes. None = reject it.

    [REASON]: a named decision rather than a literal, because it USED to be
    rejection and the change is the whole of section 4 of
    DRONE-WORKS-IMPORT-FIX-001. The negative control in
    tools/test_import_drone_works.py swaps in the version that returns None
    and watches the two rows disappear again -- which is what happened to 360
    real rows on 2026-08-04.
    """
    return PAYMENT_UNKNOWN


# ─── Trap (e): dates ─────────────────────────────────────────────────────────

# '23-24.03.2026', '13,16,17.05.2026', '27,30.03.2026', '08-09.092025'.
# The dot between month and year is optional -- it is genuinely missing in one
# real file. Day part: digits, commas, dashes and spaces.
_DATE_TEXT_RE = re.compile(
    r'^\s*([\d]{1,2}(?:\s*[,\-]\s*[\d]{1,2})*)\s*\.\s*(\d{1,2})\s*\.?\s*'
    r'(\d{4})\s*$')


def parse_date_cell(value):
    """(date_from, date_to, kind). Never raises, never guesses.

    kind is one of DATE_KIND_EXACT / DATE_KIND_SPAN / DATE_KIND_NONE /
    DATE_KIND_UNPARSED. An unparsed form keeps its raw text and both dates
    NULL; the report counts and lists every one of them.
    """
    if value is None:
        return None, None, DATE_KIND_NONE
    if isinstance(value, datetime.datetime):
        return value.date(), value.date(), DATE_KIND_EXACT
    if isinstance(value, datetime.date):
        return value, value, DATE_KIND_EXACT
    text = _text(value)
    if not text:
        return None, None, DATE_KIND_NONE

    match = _DATE_TEXT_RE.match(text)
    if not match:
        return None, None, DATE_KIND_UNPARSED

    day_part, month_s, year_s = match.groups()
    try:
        month = int(month_s)
        year = int(year_s)
    except ValueError:
        return None, None, DATE_KIND_UNPARSED

    days = []
    for chunk in day_part.split(','):
        chunk = chunk.strip()
        if not chunk:
            return None, None, DATE_KIND_UNPARSED
        if '-' in chunk:
            bounds = [p.strip() for p in chunk.split('-')]
            if len(bounds) != 2 or not all(b.isdigit() for b in bounds):
                return None, None, DATE_KIND_UNPARSED
            low, high = int(bounds[0]), int(bounds[1])
            # [REASON]: a descending range means the span crosses a month
            # boundary ('30-01.04'). min/max would read it as the whole month.
            # Refusing is the honest outcome -- the row is listed, not dropped.
            if low > high:
                return None, None, DATE_KIND_UNPARSED
            days.extend([low, high])
        else:
            if not chunk.isdigit():
                return None, None, DATE_KIND_UNPARSED
            days.append(int(chunk))
    if not days:
        return None, None, DATE_KIND_UNPARSED

    try:
        date_from = datetime.date(year, month, min(days))
        date_to = datetime.date(year, month, max(days))
    except ValueError:
        return None, None, DATE_KIND_UNPARSED

    kind = DATE_KIND_EXACT if date_from == date_to else DATE_KIND_SPAN
    return date_from, date_to, kind


# ─── Trap (f): the price is never computed ───────────────────────────────────

def resolve_price(price_cell, amount, area):
    """The price EXACTLY as written, or None. amount and area are ignored.

    They are parameters only so the negative control in
    tools/test_import_drone_works.py can swap in the tempting version that
    returns amount / area and watch invented prices appear.
    """
    return _number(price_cell)


# ─── Trap (g): operator from the sheet title ─────────────────────────────────

_TITLE_PREFIX_RE = re.compile(r'^\s*свод\s*ичи\s*', re.IGNORECASE)
_BRACKETS_RE = re.compile(r'[()\[\]]')


def operator_from_sheet_title(title, file_stem):
    """(operator_raw, took_it_from_the_file_name).

    «свод ичи (Фурқат)» -> 'Фурқат'; «свод ичи Шохрух» -> 'Шохрух';
    «свод ичи » -> the file stem, flagged, and named in the report.
    """
    raw = _BRACKETS_RE.sub(' ', _TITLE_PREFIX_RE.sub('', _text(title)))
    raw = _WS_RE.sub(' ', raw).strip()
    if raw:
        return raw, False
    return _WS_RE.sub(' ', _text(file_stem)).strip(), True


# ─── Trap (d): duplicates ────────────────────────────────────────────────────

def duplicate_key(row):
    """(customer, area, date_from, amount) -- the identity of a job.

    The whole run shares one set of these keys, because the duplication is
    across sheets AND across files: one book carries «свод ичи (Беҳзод)» and
    «свод ичи » with identical content, and two Servis files overlap.
    """
    return (customer_key(row['customer_raw']),
            round(float(row['area_ha']), 4),
            row['work_date_from'],
            None if row['amount'] is None else round(float(row['amount']), 4))


# ─── The manifest ────────────────────────────────────────────────────────────

_MANIFEST_RE = re.compile(r'^(.*\S)\s+(\d{4}-\d{2})\s*$')


def read_manifest(path):
    """{file name: 'YYYY-MM'}. Two columns, '#' starts a comment.

    The file name may contain spaces, so the LAST whitespace-separated field
    is the period and everything before it is the name.
    """
    mapping = {}
    problems = []
    with open(path, encoding='utf-8-sig') as fh:
        for lineno, line in enumerate(fh, 1):
            line = line.split('#', 1)[0].rstrip()
            if not line.strip():
                continue
            match = _MANIFEST_RE.match(line)
            if not match:
                problems.append((lineno, line.strip()))
                continue
            mapping[match.group(1).strip()] = match.group(2)
    return mapping, problems


# ─── Reading one workbook ────────────────────────────────────────────────────

class SheetReport(object):
    """What one sheet turned out to be. Everything the report file prints."""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.operator_raw = ''
        self.operator_from_file_name = False
        self.header_row = None
        self.column_map = None
        self.header_cells = []
        self.missing_fields = []
        self.rows = 0
        self.area = 0.0
        self.rejections = []
        self.wage_rows = 0
        self.blocks = []
        # DRONE-WORKS-IMPORT-FIX-001
        self.skipped_reason = None       # non-detail sheets, see is_detail_sheet
        self.operator_from_row = 0       # rows resolved from the row column
        self.operator_from_sheet = 0     # rows resolved from the sheet title
        self.unknown_payment_rows = 0
        self.empty_customer_rows = 0


def _row_cells(row):
    return list(row)


def parse_sheet(ws, file_name, sheet_name, period_month, subdivision,
                default_payment=None, formula_cells=None):
    """Parse one worksheet. Returns (rows, SheetReport).

    rows are plain dicts, not yet resolved against the database and not yet
    de-duplicated: both of those are decisions about the whole run.
    """
    report = SheetReport(file_name, sheet_name)
    stem = os.path.splitext(file_name)[0]
    report.operator_raw, report.operator_from_file_name = (
        operator_from_sheet_title(sheet_name, stem))

    rows = []
    mapping = None
    current_payment = default_payment
    in_wage_block = False

    for excel_row, raw in enumerate(ws.iter_rows(values_only=True), 1):
        cells = _row_cells(raw)
        if not any(_text(c) for c in cells):
            continue

        if is_header_row(cells):
            mapping = map_columns(cells)
            report.header_row = excel_row
            report.header_cells = [_text(c) for c in cells]
            report.column_map = mapping
            report.missing_fields = mapping.missing_fields()
            missing_required = [f for f in REQUIRED_FIELDS if f not in mapping]
            if missing_required:
                report.rejections.append(
                    (excel_row, 'header without %s column'
                     % '/'.join(missing_required), cells))
                mapping = None
            continue

        counter = _int_counter(cells[0] if cells else None)
        area = (_number(cells[mapping.get('area')])
                if mapping is not None and 'area' in mapping
                and mapping.get('area') < len(cells)
                else None)
        looks_like_data = counter is not None and area_in_range(area)

        if not looks_like_data:
            if is_wage_marker(cells):
                in_wage_block = True
                report.blocks.append((excel_row, 'wage'))
                continue
            marker = payment_marker(cells)
            if marker is not None:
                current_payment = marker
                in_wage_block = False
                report.blocks.append((excel_row, marker))
                continue
            # Junk, totals rows and everything else that is not a job.
            if counter is not None and area is not None:
                # Trap (b) proper: it has a counter and a number where the
                # area lives, but the number cannot be a single job.
                report.rejections.append(
                    (excel_row, 'area out of range (%.2f)' % area, cells))
            continue

        if in_wage_block or is_wage_marker(cells):
            report.wage_rows += 1
            continue

        if mapping is None:
            report.rejections.append(
                (excel_row, 'data row before any recognised header', cells))
            continue

        def cell(field):
            index = mapping.get(field)
            if index is None or index >= len(cells):
                return None
            return cells[index]

        def cell_at(index):
            if index is None or index >= len(cells):
                return None
            return cells[index]

        # [REASON]: an EMPTY customer is data, not an error. All nine internal
        # jobs in «Агрокластер Дрон маълумот МАЙ.xlsx / свод ичи (Сайфулло)»
        # carry no farm name and the internal tariff 85 632.96: on the
        # holding's own land nobody writes one down. Rejecting them was why
        # the first run reported «internal 0 rows». The row is kept with an
        # empty customer_raw, no drone_customers row and no alias are created,
        # and the reports show «Заказчик не указан» separately from «Заказчик
        # не определён» -- missing data and unmatched data are different facts.
        customer_raw = _text(cell('customer'))
        if not customer_raw:
            report.empty_customer_rows += 1

        payment = current_payment
        # Internal work also appears as ordinary rows whose customer IS a
        # holding subdivision -- see trap (c).
        if customer_raw and MARK_INTERNAL in uz_fold(customer_raw):
            payment = PAYMENT_INTERNAL
        if payment is None:
            # [REASON]: owner's decision of 2026-08-04. The September and
            # October detail sheets carry no «Справка (…ойи)» / «Нақд (…ойи)»
            # markers at all -- the split lives only in the «ОБШИЙ СВОД»
            # summary sheets, which are not imported. Rejecting those rows
            # threw away 360 jobs with their hectares, customers and dates
            # over the one field that was missing. 'unknown' keeps the row and
            # names the gap; the owner sets the type on the works screen.
            payment = payment_when_no_block()
            if payment is None:
                report.rejections.append(
                    (excel_row, 'no payment block above this row', cells))
                continue
            report.unknown_payment_rows += 1

        amount = _number(cell('amount'))
        date_cell = cell('date')
        date_from, date_to, date_kind = parse_date_cell(date_cell)

        # [REASON]: other_costs is the SUM of every expense column present.
        # Six of the twenty-four layouts carry two at once, e.g. «ЁММ
        # харажати» together with «Транспорт харажати (Мерс 80 L 422 MA)».
        # Taking only the first would understate the costs of exactly those
        # sheets and nothing would say so.
        other_costs = None
        for index, _header in mapping.expenses:
            value = _number(cell_at(index))
            if value is not None:
                other_costs = value if other_costs is None else other_costs + value

        # [REASON]: a per-row «Дрон бошқарувчи оператор» beats the sheet
        # title. It is the more specific source and it is written in full --
        # «Қодиров Нурали» rather than «Нурали» -- which removes the
        # «Шахзод» / «Шохрух» ambiguity for those rows outright.
        operator_raw = report.operator_raw
        operator_source = 'file' if report.operator_from_file_name else 'sheet'
        row_operator = _text(cell('row_operator'))
        if row_operator:
            operator_raw = row_operator
            operator_source = 'row'
            report.operator_from_row += 1
        else:
            report.operator_from_sheet += 1

        formula_no_cache = False
        if formula_cells:
            checked = [mapping.get('amount'), mapping.get('price'),
                       mapping.received_index]
            checked.extend(index for index, _h in mapping.expenses)
            for index in checked:
                if index is None:
                    continue
                if ((excel_row, index) in formula_cells
                        and cell_at(index) is None):
                    formula_no_cache = True

        rows.append({
            'source_file': file_name,
            'source_sheet': sheet_name,
            'source_row': excel_row,
            'period_month': period_month,
            'customer_raw': customer_raw,
            'area_ha': area,
            'price_per_ha': resolve_price(cell('price'), amount, area),
            'amount': amount,
            'other_costs': other_costs,
            'received_amount': _number(cell_at(mapping.received_index)),
            'received_kind': mapping.received_kind,
            'payment_type': payment,
            'date_raw': _text(date_cell) or None,
            'work_date_from': date_from,
            'work_date_to': date_to,
            'date_kind': date_kind,
            'operator_raw': operator_raw,
            'operator_source': operator_source,
            'subdivision_name': subdivision,
            'note': _text(cell('note')) or None,
            'formula_no_cache': formula_no_cache,
        })
        report.rows += 1
        report.area += area

    return rows, report


def _formula_cells(path):
    """{(row, col0) : formula} for cells that hold a formula.

    [REASON]: the workbook is read with data_only=True, which returns the
    value Excel cached the last time it saved the file. A file produced by a
    tool that never calculated -- or edited and closed without saving --
    hands back None for every formula, and an amount column of NULLs looks
    exactly like a book where nobody filled the amounts in. The second load
    exists only to be able to SAY which it was.
    """
    from openpyxl import load_workbook
    found = {}
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            for row_index, row in enumerate(ws.iter_rows(values_only=True), 1):
                for col_index, value in enumerate(row):
                    if isinstance(value, str) and value.startswith('='):
                        found[(ws.title, row_index, col_index)] = value
    finally:
        wb.close()
    return found


# ─── Resolution against the database ─────────────────────────────────────────

def _connect(db_path, read_only):
    if not os.path.exists(db_path):
        # [REASON]: sqlite3.connect(path) creates an empty database when the
        # file is missing. A tool that silently creates one and reports
        # "0 operators, 0 customers" is worse than one that refuses.
        print('ERROR: database not found at %s - refusing to run.' % db_path)
        sys.exit(2)
    if read_only:
        uri = 'file:%s?mode=ro' % path_to_uri(db_path)
        con = sqlite3.connect(uri, uri=True)
    else:
        con = sqlite3.connect(db_path)
    con.execute('PRAGMA busy_timeout=30000')
    return con


def path_to_uri(path):
    return path.replace('?', '%3f').replace('#', '%23').replace('\\', '/')


def _require_tables(con, tables):
    have = {row[0] for row in con.execute(
        "SELECT name FROM sqlite_master WHERE type='table'")}
    missing = [t for t in tables if t not in have]
    if missing:
        print('ERROR: table(s) missing: %s. Run migrate_drones_works_001.py '
              'first.' % ', '.join(missing))
        sys.exit(2)


def _require_received_kind(con):
    """Refuse to run against a database that predates DRONES_WORKS_002.

    [REASON]: without the column the INSERT raises inside the transaction on
    --apply, after the whole corpus has been parsed and reported. Checking it
    up front means --dry-run says so too, which is where somebody is actually
    looking.
    """
    columns = {row[1] for row in con.execute('PRAGMA table_info(drone_works)')}
    if 'received_kind' not in columns:
        print('ERROR: drone_works.received_kind is missing. Run '
              'migrate_drones_works_002.py first.')
        sys.exit(2)


class Directory(object):
    """Operators, subdivisions and customer aliases, read once."""

    def __init__(self, con):
        self.operators = []
        for row in con.execute('SELECT id, full_name, subdivision_name '
                               'FROM drone_operators'):
            self.operators.append({
                'id': row[0],
                'full_name': row[1],
                'subdivision_name': row[2],
                'tokens': name_tokens(row[1]),
            })
        self.subdivisions = set()
        for table in ('drone_operators', 'drone_units'):
            for row in con.execute('SELECT DISTINCT subdivision_name FROM %s '
                                   'WHERE subdivision_name IS NOT NULL '
                                   "AND subdivision_name <> ''" % table):
                self.subdivisions.add(row[0])
        self.aliases = {}
        # [REASON]: is_active IS NOT 0, never = 1. The column is
        # BOOLEAN DEFAULT 1 and a hand-inserted row can hold NULL, which
        # '= 1' silently drops -- turning a known spelling into a new
        # duplicate customer. The same bug was already caught once in
        # _drone_nickname_maps().
        for row in con.execute(
                'SELECT normalized_key, drone_customer_id '
                'FROM drone_customer_aliases '
                'WHERE is_active IS NOT 0'):
            self.aliases[row[0]] = row[1]
        self.customer_names = {}
        for row in con.execute('SELECT id, name FROM drone_customers'):
            self.customer_names[row[0]] = row[1]
        self.existing_provenance = set()
        for row in con.execute('SELECT source_file, source_sheet, source_row '
                               'FROM drone_works WHERE source_file IS NOT NULL'):
            self.existing_provenance.add((row[0], row[1], row[2]))

    def subdivision_for_file(self, file_name):
        """The subdivision named in the file name, or None. Data-driven.

        The candidate list comes from the two subdivision columns already in
        the database, so nothing is hardcoded here: a subdivision the fleet
        directory has never heard of cannot be invented by this tool.
        """
        folded = uz_fold(file_name)
        best = None
        for name in self.subdivisions:
            key = uz_fold(name)
            if key and key in folded:
                if best is None or len(key) > len(uz_fold(best)):
                    best = name
        return best

    def resolve_operator(self, operator_raw, subdivision):
        """(operator id or None, reason). Never guesses between two people."""
        tokens = name_tokens(operator_raw)
        if not tokens:
            return None, 'empty operator'
        candidates = [op for op in self.operators
                      if op['tokens'] and tokens <= op['tokens']]
        if len(candidates) == 1:
            return candidates[0]['id'], 'matched'
        if not candidates:
            return None, 'no operator matches'
        if subdivision:
            key = uz_fold(subdivision)
            narrowed = [op for op in candidates
                        if uz_fold(op['subdivision_name'] or '') == key]
            if len(narrowed) == 1:
                return narrowed[0]['id'], 'matched by subdivision'
        # [REASON]: «Шахзод» is Kholmurodov Shakhzod in the Ghijduvon files
        # and Boltaev Shakhzod in the Servis files; «Шохрух» is Khamroev in
        # Kogon and Fayzullaev in Garden. When the subdivision cannot separate
        # them the row keeps drone_operator_id NULL and is listed. Never guess.
        return None, ('ambiguous between %d operators'
                      % len(candidates))


# ─── The run ─────────────────────────────────────────────────────────────────

class RunResult(object):
    def __init__(self):
        self.files_seen = []
        self.files_skipped_no_manifest = []
        self.sheets = []
        self.rows = []
        self.duplicates = []
        self.rejections = []
        self.manifest_problems = []
        self.created_customers = []
        self.matched_customers = 0
        self.unresolved_operators = {}
        self.matched_operators = 0
        self.already_in_db = 0
        self.month_outliers = {}
        self.formula_no_cache = 0
        self.wage_rows = 0
        # DRONE-WORKS-IMPORT-FIX-001
        self.skipped_sheets = []          # [(file, sheet title)]
        self.unmatched_headers = {}       # normalised header -> [(file, sheet)]
        self.unknown_payment_rows = 0
        self.empty_customer_rows = 0
        self.operator_from_row = 0
        self.operator_from_sheet = 0
        self.received_kinds = {}          # kind -> row count


def collect(directory_path, manifest, dirs_files, dir_obj, default_payment):
    """Parse every manifest-listed file in the directory. Pure reading."""
    from openpyxl import load_workbook

    result = RunResult()
    seen_keys = {}

    for file_name in dirs_files:
        if file_name not in manifest:
            result.files_skipped_no_manifest.append(file_name)
            continue
        period_month = manifest[file_name]
        full = os.path.join(directory_path, file_name)
        subdivision = dir_obj.subdivision_for_file(file_name)
        formulas = _formula_cells(full)
        wb = load_workbook(full, data_only=True, read_only=True)
        try:
            result.files_seen.append((file_name, period_month, subdivision,
                                      len(wb.worksheets)))
            for ws in wb.worksheets:
                # [REASON]: gate everything else. «олинмаган пуллар» and
                # «свод фх» repeat rows that already live in «свод ичи», and
                # «ОБШИЙ СВОД» is a summary whose «ЖАМИ:» lines are not jobs.
                if not is_detail_sheet(ws.title):
                    skipped = SheetReport(file_name, ws.title)
                    skipped.skipped_reason = (
                        'not a detail sheet (title does not start with '
                        '"свод ичи")')
                    result.sheets.append(skipped)
                    result.skipped_sheets.append((file_name, ws.title))
                    continue
                sheet_formulas = {(r, c) for (title, r, c) in formulas
                                  if title == ws.title}
                rows, report = parse_sheet(
                    ws, file_name, ws.title, period_month, subdivision,
                    default_payment=default_payment,
                    formula_cells=sheet_formulas)
                result.sheets.append(report)
                result.wage_rows += report.wage_rows
                result.unknown_payment_rows += report.unknown_payment_rows
                result.empty_customer_rows += report.empty_customer_rows
                result.operator_from_row += report.operator_from_row
                result.operator_from_sheet += report.operator_from_sheet
                if report.column_map is not None:
                    for _index, header in report.column_map.unmatched:
                        result.unmatched_headers.setdefault(
                            header, []).append((file_name, ws.title))
                for excel_row, reason, cells in report.rejections:
                    result.rejections.append(
                        (file_name, ws.title, excel_row, reason,
                         ' | '.join(_text(c) for c in cells)[:200]))
                for row in rows:
                    if row['formula_no_cache']:
                        result.formula_no_cache += 1
                    key = duplicate_key(row)
                    first = seen_keys.get(key) if key is not None else None
                    if first is not None:
                        result.duplicates.append((first, row))
                        continue
                    if key is not None:
                        seen_keys[key] = row
                    result.rows.append(row)
                    if row['received_kind']:
                        result.received_kinds[row['received_kind']] = (
                            result.received_kinds.get(row['received_kind'], 0)
                            + 1)
                    if (row['work_date_from'] is not None
                            and row['work_date_from'].strftime('%Y-%m')
                            != period_month):
                        result.month_outliers.setdefault(file_name, set()).add(
                            row['work_date_from'].strftime('%Y-%m'))
        finally:
            wb.close()
    return result


def resolve(result, dir_obj):
    """Fill drone_operator_id / drone_customer_id. Mutates result.rows."""
    pending_customers = {}
    for row in result.rows:
        operator_id, reason = dir_obj.resolve_operator(row['operator_raw'],
                                                       row['subdivision_name'])
        row['drone_operator_id'] = operator_id
        if operator_id is None:
            key = (row['operator_raw'], reason)
            result.unresolved_operators[key] = (
                result.unresolved_operators.get(key, 0) + 1)
        else:
            result.matched_operators += 1

        # [REASON]: an EMPTY customer creates NOTHING. A drone_customers row
        # named '' would be a farm called nothing, it would collect every
        # nameless internal job under one meaningless heading, and its alias
        # would then swallow every future nameless row. The work row keeps
        # customer_raw = '' and drone_customer_id NULL, and the reports say
        # «Заказчик не указан» -- which is the fact.
        if not row['customer_raw']:
            row['drone_customer_id'] = None
            row['customer_is_new'] = False
            continue

        key = customer_key(row['customer_raw'])
        customer_id = dir_obj.aliases.get(key)
        if customer_id is not None:
            row['drone_customer_id'] = customer_id
            row['customer_is_new'] = False
            result.matched_customers += 1
        else:
            row['drone_customer_id'] = None
            row['customer_is_new'] = True
            if key not in pending_customers:
                pending_customers[key] = row['customer_raw']
    result.created_customers = sorted(pending_customers.items(),
                                      key=lambda kv: kv[1])

    result.already_in_db = sum(
        1 for row in result.rows
        if (row['source_file'], row['source_sheet'], row['source_row'])
        in dir_obj.existing_provenance)
    return result


# ─── Writing ─────────────────────────────────────────────────────────────────

def apply_rows(con, result, batch):
    """Insert customers, aliases and works inside ONE transaction."""
    now = datetime.datetime.utcnow().isoformat()
    cur = con.cursor()
    cur.execute('BEGIN')
    created = 0
    for key, raw_name in result.created_customers:
        cur.execute('SELECT drone_customer_id FROM drone_customer_aliases '
                    'WHERE normalized_key = ?', (key,))
        found = cur.fetchone()
        if found:
            customer_id = found[0]
        else:
            cur.execute(
                'INSERT INTO drone_customers (name, customer_type, is_active, '
                'note, created_at, updated_at) VALUES (?, ?, 1, ?, ?, ?)',
                (raw_name, 'external',
                 'import %s' % batch, now, now))
            customer_id = cur.lastrowid
            cur.execute(
                'INSERT INTO drone_customer_aliases (raw_name, '
                'normalized_key, drone_customer_id, is_active, created_at) '
                'VALUES (?, ?, ?, 1, ?)',
                (raw_name, key, customer_id, now))
            created += 1
        for row in result.rows:
            # An empty customer_raw normalises to '' and must never be
            # attached to a customer created for some other spelling.
            if (row['customer_raw']
                    and customer_key(row['customer_raw']) == key):
                row['drone_customer_id'] = customer_id

    inserted = 0
    skipped = 0
    for row in result.rows:
        cur.execute('SELECT id FROM drone_works WHERE source_file = ? AND '
                    'source_sheet = ? AND source_row = ?',
                    (row['source_file'], row['source_sheet'],
                     row['source_row']))
        if cur.fetchone():
            skipped += 1
            continue
        cur.execute(
            'INSERT INTO drone_works (period_month, work_date_from, '
            'work_date_to, date_raw, drone_operator_id, operator_raw, '
            'drone_customer_id, customer_raw, area_ha, price_per_ha, amount, '
            'other_costs, received_amount, received_kind, payment_type, '
            'subdivision_name, source_file, source_sheet, source_row, '
            'import_batch, note, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '
            '?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (row['period_month'],
             row['work_date_from'].isoformat() if row['work_date_from'] else None,
             row['work_date_to'].isoformat() if row['work_date_to'] else None,
             row['date_raw'], row['drone_operator_id'], row['operator_raw'],
             row['drone_customer_id'], row['customer_raw'], row['area_ha'],
             row['price_per_ha'], row['amount'], row['other_costs'],
             row['received_amount'], row['received_kind'],
             row['payment_type'],
             row['subdivision_name'], row['source_file'], row['source_sheet'],
             row['source_row'], batch, row['note'], now))
        inserted += 1
    con.commit()
    return created, inserted, skipped


# ─── Reporting ───────────────────────────────────────────────────────────────

def _ascii(text):
    return str(text).encode('ascii', 'backslashreplace').decode('ascii')


def summarize(result):
    """Numbers only -- safe for an ASCII console."""
    rows = result.rows
    kinds = {DATE_KIND_EXACT: 0, DATE_KIND_SPAN: 0, DATE_KIND_NONE: 0,
             DATE_KIND_UNPARSED: 0}
    kind_area = dict.fromkeys(kinds, 0.0)
    payments = dict.fromkeys(PAYMENT_TYPES, 0)
    payment_area = dict.fromkeys(PAYMENT_TYPES, 0.0)
    months = {}
    manifest_months = {}
    customers = set()
    area = 0.0
    amount = 0.0
    received = 0.0
    other = 0.0
    priced = 0
    for row in rows:
        kinds[row['date_kind']] += 1
        kind_area[row['date_kind']] += row['area_ha']
        payments[row['payment_type']] += 1
        payment_area[row['payment_type']] += row['area_ha']
        month = (row['work_date_from'].strftime('%Y-%m')
                 if row['work_date_from'] else row['period_month'])
        bucket = months.setdefault(month, [0, 0.0])
        bucket[0] += 1
        bucket[1] += row['area_ha']
        # The same rows keyed by the MANIFEST month, which is what the
        # prediction in EXPECTED_BY_MONTH is grouped by. Both are printed:
        # comparing a date-grouped actual against a manifest-grouped
        # expectation is how a straddling book looks like a defect.
        bucket = manifest_months.setdefault(row['period_month'], [0, 0.0])
        bucket[0] += 1
        bucket[1] += row['area_ha']
        if row['customer_raw']:
            customers.add(customer_key(row['customer_raw']))
        area += row['area_ha']
        amount += row['amount'] or 0.0
        received += row['received_amount'] or 0.0
        other += row['other_costs'] or 0.0
        if row['price_per_ha'] is not None:
            priced += 1
    return {
        'files': len(result.files_seen),
        'sheets': len(result.sheets),
        'rows': len(rows),
        'area': area,
        'amount': amount,
        'received': received,
        'other_costs': other,
        'customers': len(customers),
        'priced': priced,
        'kinds': kinds,
        'kind_area': kind_area,
        'payments': payments,
        'payment_area': payment_area,
        'months': months,
        'manifest_months': manifest_months,
    }


def print_console(result, stats, args, batch):
    print('=' * 72)
    print('DRONE-WORKS-001 import  --  %s'
          % ('APPLY' if args.apply else 'DRY RUN (nothing written)'))
    print('=' * 72)
    print('directory   : %s' % _ascii(args.dir))
    print('manifest    : %s' % _ascii(args.manifest))
    print('database    : %s' % _ascii(args.db))
    print('batch       : %s' % _ascii(batch))
    print('report file : %s' % _ascii(args.report))
    print('')
    print('files parsed             : %d' % stats['files'])
    print('files skipped (no entry) : %d' % len(result.files_skipped_no_manifest))
    print('sheets                   : %d' % stats['sheets'])
    print('rows accepted            : %d' % stats['rows'])
    print('hectares                 : %.2f' % stats['area'])
    print('amount total             : %.2f' % stats['amount'])
    print('received total           : %.2f' % stats['received'])
    print('other costs total        : %.2f' % stats['other_costs'])
    print('distinct customers       : %d' % stats['customers'])
    print('rows with a price cell   : %d' % stats['priced'])
    print('')
    print('by payment type:')
    for payment in PAYMENT_TYPES:
        print('  %-9s %5d rows  %10.2f ha'
              % (payment, stats['payments'][payment],
                 stats['payment_area'][payment]))
    print('')
    print('by date kind:')
    for kind in (DATE_KIND_EXACT, DATE_KIND_SPAN, DATE_KIND_NONE,
                 DATE_KIND_UNPARSED):
        print('  %-9s %5d rows  %10.2f ha'
              % (kind, stats['kinds'][kind], stats['kind_area'][kind]))
    print('')
    print('by month (real date when present, manifest month otherwise):')
    for month in sorted(stats['months']):
        count, month_area = stats['months'][month]
        print('  %-8s %5d rows  %10.2f ha' % (month, count, month_area))
    print('')
    print('by MANIFEST month, against the prediction of 2026-08-04 for the')
    print('28 REAL books -- a run over any other corpus differs wholesale:')
    print('  %-8s %6s %11s | %6s %11s | %s'
          % ('month', 'rows', 'ha', 'exp.', 'exp. ha', 'delta ha'))
    expected = dict((m, (r, a)) for m, r, a in EXPECTED_BY_MONTH)
    for month in sorted(set(stats['manifest_months']) | set(expected)):
        count, month_area = stats['manifest_months'].get(month, (0, 0.0))
        exp_rows, exp_area = expected.get(month, (0, 0.0))
        delta = month_area - exp_area
        print('  %-8s %6d %11.2f | %6d %11.2f | %+10.2f'
              % (month, count, month_area, exp_rows, exp_area, delta))
    total_delta = stats['area'] - EXPECTED_TOTAL_AREA
    total_pct = (total_delta * 100.0 / EXPECTED_TOTAL_AREA
                 if EXPECTED_TOTAL_AREA else 0.0)
    print('  %-8s %6d %11.2f | %6d %11.2f | %+10.2f (%+.2f%%)'
          % ('TOTAL', stats['rows'], stats['area'], EXPECTED_TOTAL_ROWS,
             EXPECTED_TOTAL_AREA, total_delta, total_pct))
    print('  The monthly rows are INDICATIVE: this tool groups a dated row by')
    print('  its own date and one book straddles September and October. The')
    print('  total is the firm number. %s is the anchor -- it agrees with our'
          % EXPECTED_ANCHOR_MONTH)
    print('  own DJI flight data to 0.15 %. A disagreement is a finding to')
    print('  report, not a number to force.')
    print('')
    print('customers matched to an alias : %d' % result.matched_customers)
    print('customers to be created       : %d' % len(result.created_customers))
    print('operator resolved rows        : %d' % result.matched_operators)
    print('operator UNRESOLVED rows      : %d'
          % sum(result.unresolved_operators.values()))
    print('duplicate rows skipped        : %d' % len(result.duplicates))
    print('rows already in the database  : %d' % result.already_in_db)
    print('rows rejected                 : %d' % len(result.rejections))
    print('wage rows ignored (Ish haki)  : %d' % result.wage_rows)
    print('formula cells with no cache   : %d' % result.formula_no_cache)
    print('')
    print('sheets skipped (not "svod ichi") : %d' % len(result.skipped_sheets))
    print('headers matching no known phrase : %d distinct'
          % len(result.unmatched_headers))
    print('rows with payment type unknown   : %d' % result.unknown_payment_rows)
    print('rows with an empty customer      : %d' % result.empty_customer_rows)
    print('operator taken from the row cell : %d' % result.operator_from_row)
    print('operator taken from sheet/file   : %d' % result.operator_from_sheet)
    for kind in (RECEIVED_KIND_RECEIVED, RECEIVED_KIND_OPERATOR_DUE):
        print('received figure of kind %-13s: %d'
              % (kind, result.received_kinds.get(kind, 0)))
    if result.month_outliers:
        print('')
        print('files whose dated rows fall outside the manifest month: %d'
              % len(result.month_outliers))
        print('  (their dated rows keep their own dates; only undated rows')
        print('   inherit the manifest month -- see the report file)')
    print('')
    print('Cyrillic detail (farms, operators, sheets, every skipped and')
    print('rejected row) is in the UTF-8 report file above.')


def write_report(path, result, stats, args, batch):
    lines = []
    add = lines.append
    add('DRONE-WORKS-001 — отчёт импорта')
    add('Режим: %s' % ('ЗАПИСЬ (--apply)' if args.apply
                       else 'СУХОЙ ПРОГОН (ничего не записано)'))
    add('Каталог: %s' % args.dir)
    add('Манифест: %s' % args.manifest)
    add('База: %s' % args.db)
    add('Партия импорта: %s' % batch)
    add('')
    add('ИТОГИ')
    add('  файлов разобрано: %d' % stats['files'])
    add('  листов: %d' % stats['sheets'])
    add('  строк принято: %d' % stats['rows'])
    add('  гектаров: %.2f' % stats['area'])
    add('  сумма: %.2f' % stats['amount'])
    add('  получено: %.2f' % stats['received'])
    add('  прочие расходы: %.2f' % stats['other_costs'])
    add('  различных написаний заказчика: %d' % stats['customers'])
    add('')
    add('ПО ТИПАМ ОПЛАТЫ')
    for payment in PAYMENT_TYPES:
        add('  %-9s %5d строк  %10.2f га'
            % (payment, stats['payments'][payment],
               stats['payment_area'][payment]))
    add('')
    add('ПО ВИДУ ДАТЫ')
    for kind in (DATE_KIND_EXACT, DATE_KIND_SPAN, DATE_KIND_NONE,
                 DATE_KIND_UNPARSED):
        add('  %-9s %5d строк  %10.2f га'
            % (kind, stats['kinds'][kind], stats['kind_area'][kind]))
    add('')
    add('ПО МЕСЯЦАМ')
    for month in sorted(stats['months']):
        count, month_area = stats['months'][month]
        add('  %-8s %5d строк  %10.2f га' % (month, count, month_area))

    add('')
    add('ФАЙЛЫ')
    for file_name, period, subdivision, sheets in result.files_seen:
        add('  %s — период %s, подразделение %s, листов %d'
            % (file_name, period, subdivision or '(не определено)', sheets))
    if result.files_skipped_no_manifest:
        add('')
        add('ФАЙЛЫ БЕЗ ЗАПИСИ В МАНИФЕСТЕ — НЕ ОБРАБОТАНЫ')
        add('  Период не угадывается: угаданный период поставил бы сентябрьскую')
        add('  работу в март. Добавьте строку в манифест и повторите прогон.')
        for file_name in result.files_skipped_no_manifest:
            add('  - %s' % file_name)
    if result.manifest_problems:
        add('')
        add('СТРОКИ МАНИФЕСТА, КОТОРЫЕ НЕ РАЗОБРАНЫ')
        for lineno, text in result.manifest_problems:
            add('  - строка %d: %s' % (lineno, text))

    add('')
    add('ПО МЕСЯЦАМ МАНИФЕСТА, ПРОТИВ ПРЕДСКАЗАНИЯ ОТ 2026-08-04')
    add('  Помесячные строки ОРИЕНТИРОВОЧНЫ: инструмент группирует')
    add('  датированную строку по её собственной дате, а одна книга лежит')
    add('  между сентябрём и октябрём. Твёрдое число — итог. Апрель —')
    add('  якорь: 6 388.83 га сходятся с нашими же данными DJI за апрель')
    add('  2026 (6 379.24 га на 6 403 вылетах) до 0.15 %.')
    _expected = dict((m, (r, a)) for m, r, a in EXPECTED_BY_MONTH)
    for month in sorted(set(stats['manifest_months']) | set(_expected)):
        count, month_area = stats['manifest_months'].get(month, (0, 0.0))
        exp_rows, exp_area = _expected.get(month, (0, 0.0))
        add('  %-8s %5d строк %11.2f га | ожидалось %5d %11.2f | %+.2f'
            % (month, count, month_area, exp_rows, exp_area,
               month_area - exp_area))
    add('  ИТОГО    %5d строк %11.2f га | ожидалось %5d %11.2f | %+.2f'
        % (stats['rows'], stats['area'], EXPECTED_TOTAL_ROWS,
           EXPECTED_TOTAL_AREA, stats['area'] - EXPECTED_TOTAL_AREA))

    add('')
    add('ПРОПУЩЕННЫЕ ЛИСТЫ: %d' % len(result.skipped_sheets))
    add('  Импортируются только листы, чьё название начинается на «свод ичи».')
    add('  «ОБШИЙ СВОД», «олинмаган пуллар», «свод фх» и подобные — это')
    add('  сводки и реестры: их строки повторяют уже импортированные из')
    add('  «свод ичи», а строки «ЖАМИ:» вообще не являются работами.')
    for file_name, sheet_name in result.skipped_sheets:
        add('  - %s / %s' % (file_name, sheet_name))

    add('')
    add('ЗАГОЛОВКИ, НЕ СОВПАВШИЕ НИ С ОДНОЙ ИЗВЕСТНОЙ ФРАЗОЙ: %d'
        % len(result.unmatched_headers))
    add('  Поле остаётся NULL, строка импортируется. Настоящих раскладок')
    add('  двадцать четыре; может появиться двадцать пятая — вот она.')
    for header, places in sorted(result.unmatched_headers.items()):
        add('  - «%s» — %d раз(а)' % (header, len(places)))
        for file_name, sheet_name in places[:5]:
            add('      %s / %s' % (file_name, sheet_name))
        if len(places) > 5:
            add('      ... и ещё %d' % (len(places) - 5))

    add('')
    add('ЛИСТЫ')
    for sheet in result.sheets:
        add('  %s / %s' % (sheet.file_name, sheet.sheet_name))
        if sheet.skipped_reason:
            add('    ПРОПУЩЕН: %s' % sheet.skipped_reason)
            continue
        add('    оператор: %s%s'
            % (sheet.operator_raw or '(пусто)',
               ' — ВЗЯТ ИЗ ИМЕНИ ФАЙЛА, в заголовке листа имени нет'
               if sheet.operator_from_file_name else ''))
        add('    строк: %d, гектаров: %.2f' % (sheet.rows, sheet.area))
        if sheet.operator_from_row:
            add('    оператор из колонки строки: %d, из названия листа: %d'
                % (sheet.operator_from_row, sheet.operator_from_sheet))
        if sheet.header_row:
            add('    шапка в строке %d: %s'
                % (sheet.header_row, ' | '.join(sheet.header_cells)))
            add('    колонки: %s'
                % ', '.join('%s->%d («%s»)'
                            % (field, index,
                               sheet.column_map.chosen.get(field, ''))
                            for field, index
                            in sorted(sheet.column_map.fields.items())))
            if sheet.column_map.expenses:
                add('    расходы суммируются из: %s'
                    % ', '.join('«%s» (кол. %d)' % (header, index)
                                for index, header
                                in sheet.column_map.expenses))
            if sheet.column_map.received_kind:
                add('    «получено» прочитано как: %s'
                    % sheet.column_map.received_kind)
            if sheet.column_map.shadowed:
                add('    ВТОРОЕ СОВПАДЕНИЕ, не использовано: %s'
                    % ', '.join('%s кол. %d («%s»)' % (f, i, h)
                                for f, i, h in sheet.column_map.shadowed))
            if sheet.column_map.unmatched:
                add('    ЗАГОЛОВКИ БЕЗ СООТВЕТСТВИЯ: %s'
                    % ', '.join('«%s» (кол. %d)' % (h, i)
                                for i, h in sheet.column_map.unmatched))
            if sheet.missing_fields:
                add('    НЕ ОПОЗНАНЫ КОЛОНКИ: %s — эти поля записаны как NULL'
                    % ', '.join(sheet.missing_fields))
        else:
            add('    шапка не найдена — лист не дал ни одной строки')
        if sheet.blocks:
            add('    блоки: %s'
                % ', '.join('%d:%s' % (r, b) for r, b in sheet.blocks))
        if sheet.wage_rows:
            add('    строк «Иш хаки» пропущено: %d' % sheet.wage_rows)
        if sheet.unknown_payment_rows:
            add('    строк без блока оплаты (unknown): %d'
                % sheet.unknown_payment_rows)
        if sheet.empty_customer_rows:
            add('    строк без имени заказчика: %d'
                % sheet.empty_customer_rows)

    add('')
    add('ДУБЛИ, ПРОПУЩЕННЫЕ В ЭТОМ ПРОГОНЕ: %d' % len(result.duplicates))
    add('  Совпадение по (заказчик, площадь, дата, сумма). Обе принадлежности')
    add('  показаны: ни одна строка не склеена молча и не посчитана дважды.')
    for first, second in result.duplicates:
        add('  - «%s» %.2f га %s сумма %s'
            % (second['customer_raw'], second['area_ha'],
               second['work_date_from'] or 'без даты',
               second['amount'] if second['amount'] is not None else '—'))
        add('      принято: %s / %s / строка %d'
            % (first['source_file'], first['source_sheet'],
               first['source_row']))
        add('      пропущено: %s / %s / строка %d'
            % (second['source_file'], second['source_sheet'],
               second['source_row']))

    add('')
    add('СТРОКИ, ОТКЛОНЁННЫЕ ПРИ РАЗБОРЕ: %d' % len(result.rejections))
    for file_name, sheet_name, excel_row, reason, preview in result.rejections:
        add('  - %s / %s / строка %d: %s'
            % (file_name, sheet_name, excel_row, reason))
        add('      %s' % preview)

    unparsed = [r for r in result.rows if r['date_kind'] == DATE_KIND_UNPARSED]
    add('')
    add('ДАТЫ, КОТОРЫЕ НЕ РАЗОБРАНЫ: %d' % len(unparsed))
    add('  Строки СОХРАНЕНЫ: date_raw заполнен, обе даты NULL. Не потеряны.')
    for row in unparsed:
        add('  - %s / %s / строка %d: «%s»'
            % (row['source_file'], row['source_sheet'], row['source_row'],
               row['date_raw']))

    add('')
    add('ОПЕРАТОРЫ, КОТОРЫЕ НЕ ОПРЕДЕЛЕНЫ')
    add('  drone_operator_id остаётся NULL. Строка не отклоняется.')
    for (raw, reason), count in sorted(result.unresolved_operators.items(),
                                       key=lambda kv: -kv[1]):
        add('  - «%s»: %s, строк %d' % (raw, reason, count))

    add('')
    add('ЗАКАЗЧИКИ')
    add('  сопоставлено с алиасом: %d строк' % result.matched_customers)
    add('  будет заведено новых: %d' % len(result.created_customers))
    for key, raw_name in result.created_customers:
        add('  + %s' % raw_name)

    if result.month_outliers:
        add('')
        add('ФАЙЛЫ, ЧЬИ ДАТИРОВАННЫЕ СТРОКИ ВЫХОДЯТ ЗА МЕСЯЦ МАНИФЕСТА')
        add('  Датированные строки сохраняют СВОИ даты; месяц манифеста')
        add('  наследуют только строки без даты.')
        for file_name, months in sorted(result.month_outliers.items()):
            add('  - %s: встречены месяцы %s'
                % (file_name, ', '.join(sorted(months))))

    if result.formula_no_cache:
        add('')
        add('ЯЧЕЙКИ С ФОРМУЛОЙ БЕЗ СОХРАНЁННОГО ЗНАЧЕНИЯ: %d строк'
            % result.formula_no_cache)
        add('  Файл сохранён программой, которая не считала формулы.')
        add('  Откройте книгу в Excel, пересчитайте, сохраните и повторите.')

    with open(path, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(lines) + '\n')


# ─── Entry point ─────────────────────────────────────────────────────────────

def build_parser():
    parser = argparse.ArgumentParser(
        description='Import the dispatchers\' drone books into drone_works.')
    parser.add_argument('--dir', required=True,
                        help='directory of .xlsx files')
    parser.add_argument('--manifest', required=True,
                        help='file -> YYYY-MM mapping')
    parser.add_argument('--db', default=DEFAULT_DB_PATH,
                        help='SQLite database (default instance/transport.db)')
    parser.add_argument('--dry-run', action='store_true', default=True,
                        help='DEFAULT. Parse, report, write nothing.')
    parser.add_argument('--apply', action='store_true',
                        help='actually write')
    parser.add_argument('--batch', default=None, help='import_batch value')
    parser.add_argument('--report', default=None,
                        help='UTF-8 report file (default next to the manifest)')
    parser.add_argument('--default-payment', choices=PAYMENT_OVERRIDE_TYPES,
                        default=None,
                        help='payment type for rows appearing before any '
                             "block marker. Off by default: such rows get "
                             "'unknown', which is reported and is fixable on "
                             'the works screen.')
    return parser


def main(argv=None):
    args = build_parser().parse_args(argv)
    if args.apply:
        args.dry_run = False
    if not os.path.isdir(args.dir):
        print('ERROR: not a directory: %s' % _ascii(args.dir))
        return 2
    if not os.path.isfile(args.manifest):
        print('ERROR: manifest not found: %s' % _ascii(args.manifest))
        return 2
    if args.report is None:
        args.report = os.path.join(os.path.dirname(
            os.path.abspath(args.manifest)), 'drone_works_import_report.txt')
    batch = args.batch or ('import_%s' % datetime.datetime.now()
                           .strftime('%Y%m%d_%H%M%S'))

    manifest, manifest_problems = read_manifest(args.manifest)
    files = sorted(f for f in os.listdir(args.dir)
                   if f.lower().endswith('.xlsx') and not f.startswith('~$'))

    con = _connect(args.db, read_only=not args.apply)
    try:
        _require_tables(con, ('drone_works', 'drone_customers',
                              'drone_customer_aliases', 'drone_operators'))
        _require_received_kind(con)
        dir_obj = Directory(con)
        result = collect(args.dir, manifest, files, dir_obj,
                         args.default_payment)
        result.manifest_problems = manifest_problems
        resolve(result, dir_obj)
        stats = summarize(result)

        if args.apply:
            created, inserted, skipped = apply_rows(con, result, batch)
        else:
            created = inserted = skipped = 0
    except Exception:
        try:
            con.rollback()
        except sqlite3.Error:
            pass
        raise
    finally:
        con.close()

    write_report(args.report, result, stats, args, batch)
    print_console(result, stats, args, batch)
    if args.apply:
        print('')
        print('WRITTEN: %d customers created, %d works inserted, '
              '%d already present and skipped.' % (created, inserted, skipped))
    else:
        print('')
        print('DRY RUN: nothing was written. Re-run with --apply to write.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
