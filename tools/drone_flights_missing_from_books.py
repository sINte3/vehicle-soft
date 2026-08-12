#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""drone_flights_missing_from_books.py -- DRONE-FLIGHTS-MISSING-001:
гектары, которые есть в телеметрии DJI, но не записаны в книги диспетчеров.

ЗАЧЕМ. Посуточная сверка (DRONE-FLIGHTS-VS-BOOKS-DAILY-001, #69) показала
владельцу, ГДЕ расходятся стороны; следом он попросил обратную выборку:
только незаписанное -- по месяцам, дням, гектарам, машинам, с оператором и
заказчиком, насколько их знает телеметрия. Этот инструмент -- та выборка.
Он НИЧЕГО НЕ РЕШАЕТ: он называет гектары, летавшие без строки в книге,
чтобы их можно было записать, а не чтобы кого-то оштрафовать.

ЧЕСТНЫЕ ПРЕДЕЛЫ (подписаны в самом отчёте, а не подразумеваются):

  * ЗАКАЗЧИКА У НЕЗАПИСАННОЙ РАБОТЫ В ДАННЫХ НЕТ. DJI не передаёт ни поле,
    ни клиента: plot_name пуст на всех вылетах (проверенный факт трека, §3),
    есть только координаты и геокодированный адрес. Поэтому вместо заказчика
    отчёт даёт место по DJI (регион и частые адреса) и заказчиков, которые
    ЗАПИСАНЫ в книгах на ту же машину в тот же день. Заказчик незаписанной
    работы появится только когда её запишет диспетчер.
  * ОПЕРАТОР -- ПРОИЗВОДНАЯ, НЕ НАБЛЮДЕНИЕ. Весь парк летает под одним
    аккаунтом DzzDron, оператора в телеметрии нет. Оператор дня выводится
    через drone_operator_assignments ровно по правилу экранов приложения
    (_drone_flight_operator_subquery в drones.py): покрытий нет -> «привязки
    нет»; больше одного оператора -> названы ВСЕ, никому не отдаётся молча.
  * МАШИНА РАБОТЫ выводится тем же правилом в обратную сторону: привязки её
    оператора на ДАТУ НАЧАЛА работы, DISTINCT по машинам. Работа без даты,
    с нераспознанным оператором, без привязки на дату или с двумя машинами
    на дату к машине НЕ привязывается: она лежит в листе «Работы без машины»
    с причиной, и её гектары не вычитаются ни из одной машины. Раскладывать
    их по машинам «правдоподобно» значило бы выдумать атрибуцию.
  * Телеметрия меряет всё, что летало, включая свою землю и пробные вылеты;
    ведомости -- выставленную работу. «Нет в книгах» не равно «недополучены
    деньги»: это список для проверки диспетчерами, а не счёт.

СВЕРКА С ПОСУТОЧНЫМ ОТЧЁТОМ. «Нетто» месяца здесь (вылеты минус ВСЕ работы
месяца) равно колонке «Разница» посуточной сверки с обратным знаком. Сумма
дневных недостач по машинам БОЛЬШЕ нетто: день, где в книге записано больше,
чем налётано (например старт многодневной работы), в недостачу не входит,
а в нетто входит. Оба числа показаны рядом, чтобы это было видно.

ЛИСТЫ:
  «Сводка»              -- месяц: вылеты, книги (привязано / не привязано),
                           нетто и сумма дневных недостач;
  «Нет в книгах по дням» -- день х машина, ТОЛЬКО строки с недостачей:
                           оператор(ы), место по DJI, записанные заказчики
                           дня, пометки о многодневных диапазонах;
  «По месяцам и машинам» -- месяц х машина: вылеты, привязанные книги,
                           нетто; корзины непривязанных работ; итог месяца;
  «Работы без машины»    -- каждая непривязанная работа с причиной и
                           файлом/листом/строкой книги.

Только чтение: база открывается через file:-URI mode=ro, записи нет нигде.

Запуск (сервер, из каталога установки):
  & "C:\\Program Files\\Python314\\python.exe" tools\\drone_flights_missing_from_books.py --db instance\\transport.db

Период по умолчанию 2026-03..2026-06 (запрос владельца 2026-08-12); другой
задаётся --month-from/--month-to. Консоль -- только ASCII; полный отчёт --
xlsx рядом с базой.

Откат: инструмент не пишет в базу, миграций нет -- откат кода git revert,
откат данных не требуется.
"""

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import import_drone_works as importer  # noqa: E402  [REASON]: _connect с
# mode=ro и отказом кодом 2 при отсутствии базы, _require_tables и _ascii --
# те же, что у бэкфиллов и сверок: один способ открыть базу на весь каталог.

# [REASON]: правила окна, календаря и смещения UTC+5 берутся из посуточной
# сверки, а не дублируются третий раз: два отчёта, считающие день по-разному,
# хуже, чем один неправильный. Тест смещения так и остаётся одним на обоих.
from drone_flights_vs_books_daily import (  # noqa: E402
    NOTE_NO_BOOKS, NOTE_NO_FLIGHTS, UTC_OFFSET_MINUTES,
    _valid_month, month_range)

DEFAULT_MONTH_FROM = '2026-03'
DEFAULT_MONTH_TO = '2026-06'

# Меньше этого гектары считаются нулём: и стороны, и их разность.
EPS = 0.005

NO_ASSIGNMENT = 'привязки нет'

REASON_NO_DATE = 'без даты (только месяц книги)'
REASON_NO_OPERATOR = 'оператор не распознан'
REASON_NO_COVER = 'нет привязки на дату'
REASON_MANY_UNITS = 'несколько машин на дату'
REASONS = (REASON_NO_DATE, REASON_NO_OPERATOR,
           REASON_NO_COVER, REASON_MANY_UNITS)
# Консоль инструментов -- только ASCII (правило PowerShell-консоли сервера).
REASON_ASCII = {
    REASON_NO_DATE: 'no date',
    REASON_NO_OPERATOR: 'operator unresolved',
    REASON_NO_COVER: 'no assignment on date',
    REASON_MANY_UNITS: 'several machines on date',
}

SUMMARY_NOTES = (
    'Заказчика у незаписанной работы в данных нет: DJI не передаёт ни поле, '
    'ни клиента (plot_name пуст на всех вылетах) — только координаты и '
    'адрес. Заказчик появится, когда работу запишет диспетчер.',
    'Оператор — производная привязок оператор×машина×даты, как на экранах '
    'приложения: телеметрия оператора не знает (весь парк летает под одним '
    'аккаунтом). Несколько привязок на день — названы все.',
    'Телеметрия меряет всё, что летало, включая свою землю; ведомости — '
    'выставленную работу. «Нет в книгах» — список для проверки, а не счёт. '
    'Сумма дневных недостач больше нетто: день, где записано больше, чем '
    'налётано (старт многодневной работы), в недостачу не входит.',
)

DAILY_SHEET_NOTE = (
    'Только дни и машины, где налётано больше, чем записано. «Га в книгах» — '
    'работы этой машины (через привязку оператора) с датой начала в этот '
    'день. Пометка «внутри диапазона» значит: день попадает в многодневную '
    'работу, записанную целиком на день её начала, — гектары дня могут быть '
    'учтены там.')

MONTHLY_SHEET_NOTE = (
    'Нетто месяца по всем машинам и корзинам обязано сходиться со «Сводкой» '
    'и с посуточной сверкой (знак обратный). Корзины «не привязано» — '
    'работы, которые нельзя честно отдать ни одной машине; их гектары не '
    'вычитаются ни из одной строки.')

UNMAPPED_SHEET_NOTE = (
    'Работы, не привязанные к машине, с причиной. Привязка работы: машины '
    'её оператора на дату начала работы (DISTINCT), ровно как вылет '
    'получает оператора на экранах. Ноль машин, две машины, нет даты или '
    'нет оператора — работа остаётся здесь и не вычитается ниоткуда.')


def collect_units(con):
    """{unit_id: номер машины} для подписи привязанных работ."""
    return dict(con.execute('SELECT id, number FROM drone_units').fetchall())


def collect_flights(con, month_from, month_to):
    """{(день, ярлык): агрегат вылетов} с местами по DJI.

    Ярлык распознанной машины -- '№<номер>' (unit_id запомнен для привязок);
    нераспознанной -- 'ник: <ник>', как в посуточной сверке: по написанию
    видно, ЧТО именно летало без машины.
    """
    modifier = '+%d minutes' % UTC_OFFSET_MINUTES
    cells = {}
    rows = con.execute(
        "SELECT date(datetime(f.started_at, ?)), u.id, u.number, "
        "       f.nickname_raw, f.area_ha, f.region, f.location_text "
        "FROM drone_flights f LEFT JOIN drone_units u "
        "     ON u.id = f.drone_unit_id "
        "WHERE strftime('%Y-%m', datetime(f.started_at, ?)) BETWEEN ? AND ?",
        (modifier, modifier, month_from, month_to)).fetchall()
    for day, unit_id, number, nick, area, region, location in rows:
        if number is not None:
            label = '№%d' % number
        else:
            label = 'ник: %s' % (nick or '—')
            unit_id = None
        cell = cells.setdefault((day, label), {
            'unit_id': unit_id, 'flights': 0, 'area': 0.0,
            'regions': {}, 'locations': {}})
        cell['flights'] += 1
        cell['area'] += float(area or 0.0)
        if region:
            cell['regions'][region] = cell['regions'].get(region, 0) + 1
        if location:
            cell['locations'][location] = cell['locations'].get(location,
                                                                0) + 1
    return cells


def collect_assignments(con):
    """Все привязки оператор-машина, даты -- ISO-строками (сравнение текста).

    [REASON]: date() на обеих сторонах сравнения -- то же правило, что в
    _drone_flight_operator_subquery: DATE-колонки в SQLite несут NUMERIC
    affinity, и опираться на коэрцию форматов -- молчаливая зависимость от
    способа хранения. Здесь сравнение идёт в Python по ISO-строкам.
    """
    rows = con.execute(
        "SELECT a.drone_unit_id, a.operator_id, "
        "       date(a.date_from), date(a.date_to), "
        "       COALESCE(o.full_name, 'id=' || a.operator_id) "
        "FROM drone_operator_assignments a "
        "LEFT JOIN drone_operators o ON o.id = a.operator_id").fetchall()
    return [{'unit_id': u, 'operator_id': op, 'from': dfrom, 'to': dto,
             'name': name} for u, op, dfrom, dto, name in rows]


def units_covering(assignments, operator_id, day):
    """DISTINCT машины, на которые оператор привязан в этот день.

    [REASON]: DISTINCT по машинам, не по строкам привязок -- два периода
    одного человека на одной машине (переоткрытый после ремонта) дают одну
    машину, а не ложную неоднозначность. Зеркало правила экранов.
    """
    return sorted({a['unit_id'] for a in assignments
                   if a['operator_id'] == operator_id
                   and a['from'] is not None and a['from'] <= day
                   and (a['to'] is None or a['to'] >= day)})


def operators_covering(assignments, unit_id, day):
    """Имена операторов машины в этот день, DISTINCT по оператору."""
    seen = {}
    for a in assignments:
        if (a['unit_id'] == unit_id
                and a['from'] is not None and a['from'] <= day
                and (a['to'] is None or a['to'] >= day)):
            seen[a['operator_id']] = a['name']
    return sorted(seen.values())


def collect_works(con, month_from, month_to):
    """Работы окна; месяц -- как в _drone_work_month_expr (дата, иначе книга)."""
    rows = con.execute(
        "SELECT w.id, "
        "       COALESCE(strftime('%Y-%m', w.work_date_from), "
        "                w.period_month), "
        "       date(w.work_date_from), date(w.work_date_to), "
        "       w.customer_raw, w.area_ha, w.operator_raw, "
        "       w.drone_operator_id, o.full_name, "
        "       w.source_file, w.source_sheet, w.source_row "
        "FROM drone_works w "
        "LEFT JOIN drone_operators o ON o.id = w.drone_operator_id "
        "WHERE COALESCE(strftime('%Y-%m', w.work_date_from), "
        "               w.period_month) BETWEEN ? AND ?",
        (month_from, month_to)).fetchall()
    works = []
    for (wid, month, dfrom, dto, customer, area, op_raw, op_id, op_name,
         sfile, ssheet, srow) in rows:
        works.append({
            'id': wid, 'month': month, 'date_from': dfrom, 'date_to': dto,
            'customer': customer, 'area': float(area or 0.0),
            'operator_raw': op_raw, 'operator_id': op_id,
            'operator_name': op_name,
            'file': sfile, 'sheet': ssheet, 'row': srow,
        })
    return works


def attach_works(works, assignments):
    """Каждой работе -- машина по правилу привязок либо причина отказа."""
    for w in works:
        w['unit_id'] = None
        if not w['date_from']:
            w['reason'] = REASON_NO_DATE
        elif w['operator_id'] is None:
            w['reason'] = REASON_NO_OPERATOR
        else:
            units = units_covering(assignments, w['operator_id'],
                                   w['date_from'])
            if not units:
                w['reason'] = REASON_NO_COVER
            elif len(units) > 1:
                w['reason'] = REASON_MANY_UNITS
            else:
                w['reason'] = None
                w['unit_id'] = units[0]


def _label_sort_key(label):
    """'№2' раньше '№10' (по числу), ники -- после машин, по строке."""
    if label.startswith('№'):
        return (0, int(label[1:]), '')
    return (1, 0, label)


def _places(cell):
    """('регионы', 'частые адреса') из агрегата вылетов дня и машины."""
    regions = '; '.join(sorted(cell['regions'],
                               key=lambda k: (-cell['regions'][k], k)))
    top = sorted(cell['locations'],
                 key=lambda k: (-cell['locations'][k], k))
    shown = [loc if len(loc) <= 80 else loc[:77] + '…' for loc in top[:2]]
    rest = len(top) - len(shown)
    if rest > 0:
        shown.append('ещё адресов: %d' % rest)
    return regions, '; '.join(shown)


def compute(months, cells, works, assignments, unit_numbers):
    """Вся арифметика отчёта, без openpyxl -- числа проверяются тестами."""
    label_by_unit = {uid: '№%d' % num for uid, num in unit_numbers.items()}

    mapped = [w for w in works if w['unit_id'] is not None]
    unmapped = [w for w in works if w['unit_id'] is None]

    booked_day = {}     # (день, unit_id) -> [работы, начавшиеся в этот день]
    spans = []          # многодневные привязанные работы, для пометок
    for w in mapped:
        booked_day.setdefault((w['date_from'], w['unit_id']), []).append(w)
        if w['date_to'] and w['date_to'] > w['date_from']:
            spans.append(w)

    # -- день х машина: только недостача --------------------------------
    day_rows = []
    missing_month_label = {}   # (месяц, ярлык) -> сумма дневных недостач
    for (day, label), cell in sorted(
            cells.items(), key=lambda item: (item[0][0],) +
            _label_sort_key(item[0][1])):
        unit_id = cell['unit_id']
        day_works = booked_day.get((day, unit_id), []) if unit_id else []
        booked = sum(w['area'] for w in day_works)
        missing = cell['area'] - booked
        if missing <= EPS:
            continue
        month = day[:7]
        key = (month, label)
        missing_month_label[key] = missing_month_label.get(key, 0.0) + missing
        if unit_id is not None:
            operators = (operators_covering(assignments, unit_id, day)
                         or [NO_ASSIGNMENT])
        else:
            # [REASON]: у вылета без машины оператора вывести не из чего --
            # привязки ссылаются на машину; пустая клетка честнее догадки.
            operators = []
        inside = [w for w in spans
                  if unit_id is not None and w['unit_id'] == unit_id
                  and w['date_from'] < day <= w['date_to']]
        notes = '; '.join(
            'внутри диапазона работы ID %d (%s, стр. %s, %s..%s)'
            % (w['id'], w['file'] or '—', w['row'] or '—',
               w['date_from'], w['date_to'])
            for w in inside)
        customers = '; '.join(
            '%s (%.2f)' % (w['customer'], w['area'])
            for w in sorted(day_works, key=lambda w: -w['area']))
        regions, locations = _places(cell)
        day_rows.append({
            'month': month, 'day': day, 'label': label,
            'operators': '; '.join(operators),
            'flights': cell['flights'], 'flown': cell['area'],
            'booked': booked, 'missing': missing,
            'customers': customers, 'regions': regions,
            'locations': locations, 'notes': notes,
        })

    # -- месяц х машина ---------------------------------------------------
    fl_month_label = {}   # (месяц, ярлык) -> [вылетов, га, unit_id, дни]
    for (day, label), cell in cells.items():
        row = fl_month_label.setdefault(
            (day[:7], label), [0, 0.0, cell['unit_id'], set()])
        row[0] += cell['flights']
        row[1] += cell['area']
        row[3].add(day)
    bk_month_label = {}   # (месяц, ярлык) -> [работ, га, дни начала]
    for w in mapped:
        label = label_by_unit[w['unit_id']]
        row = bk_month_label.setdefault((w['month'], label), [0, 0.0, set()])
        row[0] += 1
        row[1] += w['area']
        row[2].add(w['date_from'])
    buckets = {}          # (месяц, причина) -> [работ, га]
    for w in unmapped:
        row = buckets.setdefault((w['month'], w['reason']), [0, 0.0])
        row[0] += 1
        row[1] += w['area']

    month_rows, month_totals = [], {}
    for month in months:
        labels = sorted(
            {lbl for (m, lbl) in fl_month_label if m == month}
            | {lbl for (m, lbl) in bk_month_label if m == month},
            key=_label_sort_key)
        fl_a = fl_c = bk_all = missing_sum = 0.0
        for label in labels:
            fc, fa, unit_id, fdays = fl_month_label.get(
                (month, label), (0, 0.0, None, set()))
            wc, wa, wdays = bk_month_label.get((month, label),
                                               (0, 0.0, set()))
            names = set()
            if unit_id is None:
                # Машина без вылетов в месяце, но с привязанными работами.
                for uid, lbl in label_by_unit.items():
                    if lbl == label:
                        unit_id = uid
            if unit_id is not None:
                for day in sorted(fdays | wdays):
                    names.update(operators_covering(assignments, unit_id,
                                                    day))
            miss = missing_month_label.get((month, label), 0.0)
            missing_sum += miss
            fl_a += fa
            fl_c += fc
            bk_all += wa
            month_rows.append({
                'month': month, 'label': label,
                'operators': '; '.join(sorted(names)),
                'flights': fc, 'flown': fa, 'works': wc, 'booked': wa,
                'net': fa - wa, 'missing': miss, 'kind': 'machine',
            })
        for reason in REASONS:
            wc, wa = buckets.get((month, reason), (0, 0.0))
            if not wc:
                continue
            bk_all += wa
            month_rows.append({
                'month': month, 'label': '(не привязано: %s)' % reason,
                'operators': '', 'flights': None, 'flown': None,
                'works': wc, 'booked': wa, 'net': None, 'missing': None,
                'kind': 'bucket',
            })
        month_totals[month] = {
            'flights': int(fl_c), 'flown': fl_a, 'booked_all': bk_all,
            'net': fl_a - bk_all, 'missing': missing_sum,
        }
        month_rows.append({
            'month': month, 'label': 'ИТОГО МЕСЯЦА', 'operators': '',
            'flights': int(fl_c), 'flown': fl_a, 'works': None,
            'booked': bk_all, 'net': fl_a - bk_all, 'missing': missing_sum,
            'kind': 'total',
        })

    # -- сводка ------------------------------------------------------------
    summary_rows = []
    for month in months:
        tot = month_totals[month]
        mapped_a = sum(w['area'] for w in mapped if w['month'] == month)
        unmapped_a = sum(w['area'] for w in unmapped if w['month'] == month)
        note = ''
        if tot['flown'] <= EPS:
            note = NOTE_NO_FLIGHTS
        elif tot['booked_all'] <= EPS:
            note = NOTE_NO_BOOKS
        summary_rows.append({
            'month': month, 'flown': tot['flown'],
            'booked_all': tot['booked_all'], 'mapped': mapped_a,
            'unmapped': unmapped_a, 'net': tot['net'],
            'missing': tot['missing'], 'note': note,
        })

    unmapped.sort(key=lambda w: (w['month'], REASONS.index(w['reason']),
                                 w['file'] or '', w['row'] or 0))
    return {'day_rows': day_rows, 'month_rows': month_rows,
            'summary_rows': summary_rows, 'unmapped': unmapped,
            'buckets': buckets}


def build_workbook(model):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = Workbook()

    def header(ws, row_index, titles, widths):
        for col, (title, width) in enumerate(zip(titles, widths), start=1):
            cell = ws.cell(row=row_index, column=col, value=title)
            cell.font = bold
            ws.column_dimensions[get_column_letter(col)].width = width

    # -- «Сводка» -----------------------------------------------------------
    ws = wb.active
    ws.title = 'Сводка'
    for i, note in enumerate(SUMMARY_NOTES, start=1):
        ws.cell(row=i, column=1, value=note)
    header(ws, 4, ('Месяц', 'Га по вылетам', 'Га по ведомостям',
                   'Из них привязано к машинам', 'Не привязано к машинам',
                   'Нет в книгах, нетто (выл − вед)',
                   'Сумма дневных недостач, га', 'Примечание'),
           (10, 14, 15, 24, 20, 26, 22, 26))
    ws.freeze_panes = 'A5'
    r = 5
    tot = [0.0] * 6
    for row in model['summary_rows']:
        values = (row['month'], row['flown'], row['booked_all'],
                  row['mapped'], row['unmapped'], row['net'],
                  row['missing'], row['note'])
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if 2 <= col <= 7:
                cell.number_format = '0.00'
        for i in range(6):
            tot[i] += values[i + 1]
        r += 1
    for col, value in enumerate(('ИТОГО',) + tuple(tot) + ('',), start=1):
        cell = ws.cell(row=r, column=col, value=value)
        cell.font = bold
        if 2 <= col <= 7:
            cell.number_format = '0.00'

    # -- «Нет в книгах по дням» ----------------------------------------------
    ws = wb.create_sheet('Нет в книгах по дням')
    ws.cell(row=1, column=1, value=DAILY_SHEET_NOTE)
    header(ws, 2, ('Месяц', 'Дата', 'Машина', 'Оператор(ы) по привязке',
                   'Га по вылетам', 'Вылетов', 'Га в книгах (день, машина)',
                   'Нет в книгах, га', 'Заказчики в книгах (день, машина)',
                   'Регион(ы) по DJI', 'Адреса по DJI (частые)', 'Пометки'),
           (9, 12, 14, 26, 13, 9, 22, 15, 34, 16, 46, 46))
    ws.freeze_panes = 'A3'
    for r, row in enumerate(model['day_rows'], start=3):
        values = (row['month'], row['day'], row['label'], row['operators'],
                  row['flown'], row['flights'], row['booked'],
                  row['missing'], row['customers'], row['regions'],
                  row['locations'], row['notes'])
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (5, 7, 8):
                cell.number_format = '0.00'

    # -- «По месяцам и машинам» ----------------------------------------------
    ws = wb.create_sheet('По месяцам и машинам')
    ws.cell(row=1, column=1, value=MONTHLY_SHEET_NOTE)
    header(ws, 2, ('Месяц', 'Машина', 'Оператор(ы) месяца', 'Вылетов',
                   'Га по вылетам', 'Работ (привязано)',
                   'Га в книгах (привязано)', 'Нетто (выл − книги), га',
                   'Нет в книгах по дням, га'),
           (9, 30, 30, 9, 14, 15, 20, 20, 20))
    ws.freeze_panes = 'A3'
    for r, row in enumerate(model['month_rows'], start=3):
        values = (row['month'], row['label'], row['operators'],
                  row['flights'], row['flown'], row['works'],
                  row['booked'], row['net'], row['missing'])
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (5, 7, 8, 9) and value is not None:
                cell.number_format = '0.00'
            if row['kind'] == 'total':
                cell.font = bold

    # -- «Работы без машины» ---------------------------------------------------
    ws = wb.create_sheet('Работы без машины')
    ws.cell(row=1, column=1, value=UNMAPPED_SHEET_NOTE)
    header(ws, 2, ('Причина', 'ID', 'Месяц', 'Дата от', 'Дата до',
                   'Оператор (в книге)', 'Оператор (справочник)',
                   'Заказчик', 'Га', 'Файл', 'Лист', 'Строка'),
           (26, 7, 9, 11, 11, 20, 24, 32, 9, 38, 22, 8))
    ws.freeze_panes = 'A3'
    for r, w in enumerate(model['unmapped'], start=3):
        values = (w['reason'], w['id'], w['month'], w['date_from'],
                  w['date_to'], w['operator_raw'], w['operator_name'],
                  w['customer'], w['area'], w['file'], w['sheet'], w['row'])
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col == 9:
                cell.number_format = '0.00'

    return wb


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Hectares present in DJI telemetry but absent from the '
                    'dispatcher books. Read-only, writes an xlsx report.')
    parser.add_argument('--db', default=importer.DEFAULT_DB_PATH,
                        help='SQLite database (default instance/transport.db)')
    parser.add_argument('--month-from', default=DEFAULT_MONTH_FROM,
                        help='first month, YYYY-MM (default %s)'
                             % DEFAULT_MONTH_FROM)
    parser.add_argument('--month-to', default=DEFAULT_MONTH_TO,
                        help='last month, YYYY-MM (default %s)'
                             % DEFAULT_MONTH_TO)
    parser.add_argument('--out', default=None,
                        help='xlsx path (default: next to the db)')
    args = parser.parse_args(argv)

    if not _valid_month(args.month_from) or not _valid_month(args.month_to):
        print('ERROR: months must look like YYYY-MM')
        return 1
    if args.month_from > args.month_to:
        print('ERROR: --month-from is after --month-to')
        return 1
    if args.out is None:
        args.out = os.path.join(
            os.path.dirname(os.path.abspath(args.db)) or '.',
            'drone_flights_missing_from_books.xlsx')

    try:
        con = importer._connect(args.db, read_only=True)
    except importer.ImportPreconditionError as exc:
        print(str(exc))
        return 2
    try:
        try:
            importer._require_tables(
                con, ('drone_works', 'drone_flights', 'drone_units',
                      'drone_operators', 'drone_operator_assignments'))
        except importer.ImportPreconditionError as exc:
            print(str(exc))
            return 2

        months = month_range(args.month_from, args.month_to)
        unit_numbers = collect_units(con)
        cells = collect_flights(con, args.month_from, args.month_to)
        assignments = collect_assignments(con)
        works = collect_works(con, args.month_from, args.month_to)
    finally:
        con.close()

    attach_works(works, assignments)
    model = compute(months, cells, works, assignments, unit_numbers)
    wb = build_workbook(model)
    wb.save(args.out)

    print('missing-from-books %s..%s (read-only)'
          % (args.month_from, args.month_to))
    for row in model['summary_rows']:
        print('  %s | flights: %10.2f ha | books: %10.2f ha '
              '(mapped %10.2f, unmapped %8.2f) | net missing: %+11.2f '
              '| day-machine missing: %10.2f'
              % (row['month'], row['flown'], row['booked_all'],
                 row['mapped'], row['unmapped'], row['net'], row['missing']))
    for reason in REASONS:
        wc = sum(c for (m, rsn), (c, a) in model['buckets'].items()
                 if rsn == reason)
        wa = sum(a for (m, rsn), (c, a) in model['buckets'].items()
                 if rsn == reason)
        if wc:
            print('  unmapped works [%s]: %d works, %.2f ha'
                  % (REASON_ASCII[reason], wc, wa))
    print('report: %s' % importer._ascii(os.path.abspath(args.out)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
