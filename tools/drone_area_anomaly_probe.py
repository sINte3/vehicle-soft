# -*- coding: utf-8 -*-
"""DRONE-COVERAGE-001, этап A2: аудит применимости данных DJI по всему парку.

    python tools/drone_area_anomaly_probe.py --db C:\\backups\\transport_2026-08-26.db ^
        --out "A2 аудит парка 2026-08-26.xlsx"

ТОЛЬКО ЧТЕНИЕ. База открывается через file:-URI с `mode=ro`, ни одной пишущей
команды в файле нет, отсутствующая база не создаётся (код возврата 2). Скрипт
не импортирует ни `app`, ни `models`, ни SQLAlchemy: `app = create_app()`
вызывает `db.create_all()` на импорте и превращает читателя в писателя.

**Запускать на КОПИИ, а не на рабочей базе.** Копия снимается штатным
`backup_production_db.bat`. Причина не в записи -- её здесь нет, -- а в том,
что чтение живой базы под нагрузкой даёт несогласованный срез: часть строк
прочитана до чужого коммита, часть после.

Что измеряется и зачем
----------------------
Три вопроса, от которых зависит, имеет ли смысл этап B (сбор маршрутов).

1. **Повторы площади.** 05.06.2026 у машины №8 нашлась запись `622715275`:
   40 секунд, 13 точек маршрута, 108 метров пути -- и площадь 0.594 га, в
   точности равная площади вылета ЧЕРЕЗ ОДИН (между ними стоит `622715274` с
   другой площадью, ручной режим). Один случай ничего не говорит о парке.
   Здесь считается, сколько таких по всей базе и на сколько гектаров.

   Наличие этой записи -- **обязательное предусловие прогона**: если она не
   находится, отчёт не считается годным, обычный xlsx не пишется и код
   возврата ненулевой. См. `--allow-missing-known-case`.

   **Совпадение площади само по себе не является ошибкой DJI.** Два соседних
   вылета по одному заданию вполне могут дать одну и ту же площадь. Поэтому
   строка помечается `ANOMALY_CANDIDATE`, а не «ошибка», и разбивается на
   признаки, которые случайное совпадение объяснить труднее: нулевая для
   работы длительность, слишком короткий интервал, серия из трёх и более.

2. **Отсутствующая ширина захвата.** Без `spray_width` геометрическое покрытие
   не считается вовсе -- решение владельца от 2026-08-25: `DATA_UNAVAILABLE`,
   подстановка запрещена (ни медиана, ни паспорт, ни соседнее значение). Доля
   таких вылетов -- это доля парка, до которой этап B не дотянется никогда.

   **Положительное число само по себе пригодной шириной не считается.** Пока
   допустимый диапазон не назван источником, такое значение получает статус
   `POSITIVE_UNVALIDATED`, а `USABLE` не получает никто. Диапазон задаётся
   аргументами `--min-width-m` и `--max-width-m`, и использованные границы
   печатаются в отчёте. Придумывать физику захвата T40 этот скрипт не будет.

3. **Качество исходных данных.** Всё, что помешает расчёту: битый `raw_json`,
   невозможные времена, отрицательные величины, расхождение колонки и JSON.

Скрипт НИЧЕГО не исправляет. Только измеряет и показывает.

Что он НЕ делает и почему
-------------------------
Не подставляет ширину. Не объявляет кандидата ошибкой. Не назначает порогов
«допустимого процента» -- порог это решение владельца, а не свойство данных.
Не выводит машину из ника: см. раздел «Идентификатор борта» ниже.

Вывод в консоль -- только ASCII (кодировка консоли Windows). Кириллица уходит
в xlsx.
"""

import argparse
import collections
import datetime
import hashlib
import json
import math
import os
import sqlite3
import sys

# Смещение показа. Дроны опрыскивают ночью, поэтому день считается по UTC+5:
# в UTC вылеты одной смены разъезжаются по двум датам. То же значение, что
# DRONE_DISPLAY_UTC_OFFSET в drones.py; здесь оно повторено, а не
# импортировано, потому что импорт drones.py тянет за собой приложение.
UTC_OFFSET_MINUTES = 300

# Колонки `drone_flights`, без которых аудит невозможен. Сверено с DDL
# migrate_drones_foundation_001.py, а не с models.py: на production схему
# создавала миграция.
REQUIRED_COLUMNS = (
    'id', 'dji_flight_id', 'drone_unit_id', 'nickname_raw', 'serial_number',
    'started_at', 'finished_at', 'work_seconds', 'area_ha', 'spray_width',
    'raw_json',
)

# Границы допустимой ширины захвата НЕ ЗАДАНЫ ПО УМОЛЧАНИЮ, и это осознанно.
#
# [REASON]: физически допустимый диапазон захвата T40 не подтверждён ни одним
# источником, которым располагает проект. Вписать сюда «от 3 до 12 метров»
# значило бы выдумать бизнес-правило -- ровно то, что запрещает устав. Пока
# границы не названы владельцем, положительное значение честно называется
# POSITIVE_UNVALIDATED: оно похоже на ширину, но не проверено. Когда границы
# заданы аргументами, они печатаются в отчёте рядом с числами, чтобы читатель
# видел, ЧЕМ мерили.
DEFAULT_MIN_WIDTH_M = None
DEFAULT_MAX_WIDTH_M = None

# Множитель межквартильного размаха для статистического выброса. Классический
# критерий Тьюки.
#
# [REASON]: выброс -- это «редкое значение», а НЕ «ошибка DJI». Ширина 11 м
# может быть законной настройкой, встречающейся раз в год. Поэтому выбросы
# считаются отдельно от физически невозможных значений и ни при каких
# обстоятельствах не переводят вылет в непригодные.
OUTLIER_IQR_FACTOR = 1.5

# Ключ полной площади в payload DJI. В базе лежит `area_ha` = m2 / 10000, и
# деление уже потеряло знаки: 5940.000029700001 превращается в
# 0.5940000029700001. Для сравнения «равна ли площадь предыдущей» берётся
# сырое значение, а колонка служит запасным источником и предметом сверки.
RAW_AREA_KEY = 'new_work_area'
RAW_WIDTH_KEY = 'spray_width'

# Длительность, ниже которой вылет не может быть настоящей работой. Не порог
# «допустимого», а признак для разбора: 40-секундная запись 622715275 попала
# бы сюда. Названо константой, чтобы число было видно, а не спрятано в коде.
SHORT_FLIGHT_SECONDS = 60

# Доля самых коротких интервалов между вылетами, которая считается «аномально
# коротким интервалом».
#
# [REASON]: порог НЕ выдуман -- он вычисляется из самих данных как 5-й
# процентиль интервалов между соседними вылетами той же машины, и полученное
# значение печатается в отчёте. Фиксированное число секунд здесь было бы
# правилом, взятым с потолка: интервал зависит от того, как быстро заправляют
# бак, а это разное на разных площадках.
SHORT_GAP_PERCENTILE = 5

# Минимальная длина серии одинаковых площадей, которую стоит показать отдельно.
RUN_MIN_LENGTH = 3

# На сколько вылетов назад искать равную площадь.
#
# [REASON]: задание формулирует правило как «совпадает с площадью ПРЕДЫДУЩЕГО
# по времени вылета». Проверено на известном случае -- правило его не ловит.
# 05.06.2026 у машины №8 порядок такой: 622715273 = 5940.0000297,
# 622715274 = 3293.3333 (ручной режим), 622715275 = 5940.0000297. То есть
# 622715275 повторяет площадь вылета ЧЕРЕЗ ОДИН, а непосредственно перед ним
# стоит вылет с другой площадью. Окно в один шаг дало бы ноль кандидатов на
# том самом случае, ради которого аудит затеян.
#
# Поэтому сравнение идёт по окну, а расстояние (`lag`) записывается в каждую
# строку и разделяется в отчёте: lag = 1 и lag >= 2 -- разные по
# правдоподобию события, и смешивать их нельзя. Значение 3 покрывает
# известный случай с запасом; больше не берётся, потому что чем шире окно,
# тем выше вероятность случайного совпадения.
DEFAULT_LOOKBACK = 3

# Русское название проблемы -> ASCII-код для консоли.
#
# [REASON]: консоль Windows кириллицу в этом проекте не печатает (правило
# устава), а «issue #1: 1 rows» не говорит ничего. Код даёт понятную строку
# в консоли, полное название остаётся в xlsx.
QUALITY_CODES = {
    'raw_json не разобрался': 'RAW_JSON_UNPARSABLE',
    'нет dji_flight_id': 'NO_FLIGHT_ID',
    'повторяющийся dji_flight_id': 'DUPLICATE_FLIGHT_ID',
    'нет или не разобрано время начала': 'NO_START_TIME',
    'время начала в будущем': 'START_IN_FUTURE',
    'время начала раньше 2024 года': 'START_TOO_EARLY',
    'конец раньше начала': 'END_BEFORE_START',
    'отрицательная длительность': 'NEGATIVE_DURATION',
    'нулевая длительность': 'ZERO_DURATION',
    'отрицательная площадь': 'NEGATIVE_AREA',
    'вылет не привязан к машине': 'NO_MACHINE',
    'площадь в колонке и в raw_json расходятся': 'AREA_COLUMN_VS_JSON',
    'ширина есть в raw_json, но колонка пуста': 'WIDTH_COLUMN_EMPTY',
    'ширина в колонке и в raw_json расходятся': 'WIDTH_COLUMN_VS_JSON',
    'вылет начался раньше конца предыдущего': 'OVERLAPPING_FLIGHTS',
    'new_work_area: значение не число': 'AREA_INVALID_TYPE',
    'new_work_area: значение NaN или бесконечность': 'AREA_NON_FINITE',
    'spray_width: значение не число': 'WIDTH_INVALID_TYPE',
    'spray_width: значение NaN или бесконечность': 'WIDTH_NON_FINITE',
    'area_ha в колонке не конечное число': 'AREA_COLUMN_NON_FINITE',
    'spray_width в колонке не конечное число': 'WIDTH_COLUMN_NON_FINITE',
    'площадь взята из колонки: raw_json непригоден': 'AREA_FALLBACK_TO_COLUMN',
    'ширина взята из колонки: raw_json непригоден': 'WIDTH_FALLBACK_TO_COLUMN',
    'new_work_area: ключа нет в payload': 'AREA_MISSING_KEY',
    'new_work_area: значение null': 'AREA_JSON_NULL',
}

# Известный случай, служащий предусловием прогона. DISCOVERY §6.2.
KNOWN_CASE_FLIGHT_ID = 622715275
KNOWN_CASE_EXPECTED_LAG = 2

EXIT_OK = 0
EXIT_PRECONDITION = 1
EXIT_NO_DB = 2
EXIT_KNOWN_CASE_FAILED = 4


class ProbeError(Exception):
    """Аудит невозможен: нет базы, нет таблицы, нет обязательной колонки."""


# ─── Подключение ─────────────────────────────────────────────────────────────

def connect_read_only(db_path, immutable=False):
    """Соединение только на чтение. Отсутствующий файл -- ошибка, не пустая база.

    [REASON]: `sqlite3.connect(path)` СОЗДАЁТ файл, если его нет. Читатель,
    который молча создал пустую базу и отчитался «0 аномалий», хуже, чем
    падение: ноль выглядит как хорошая новость.

    `immutable=1` разрешён только для статической копии: он говорит SQLite,
    что файл никто не меняет, и отключает проверку журнала. На живой базе это
    портит чтение, поэтому по умолчанию выключен.
    """
    if not os.path.exists(db_path):
        raise ProbeError('database not found: %s' % db_path)
    quoted = db_path.replace('?', '%3f').replace('#', '%23')
    uri = 'file:%s?mode=ro' % quoted
    if immutable:
        uri += '&immutable=1'
    con = sqlite3.connect(uri, uri=True)
    con.row_factory = sqlite3.Row
    return con


def sha256_of(path):
    digest = hashlib.sha256()
    with open(path, 'rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


# ─── Схема ───────────────────────────────────────────────────────────────────

def describe_schema(con):
    """Фактические колонки `drone_flights`, а не ожидаемые.

    Возвращает (список колонок, список отсутствующих обязательных).
    """
    rows = list(con.execute('PRAGMA table_info(drone_flights)'))
    if not rows:
        raise ProbeError(
            'table drone_flights not found -- is this a Vehicle Soft database?')
    present = [row['name'] for row in rows]
    missing = [name for name in REQUIRED_COLUMNS if name not in present]
    return present, missing


def machine_identity(con):
    """Чем в этой базе опознаётся борт. Измерение, а не пересказ модели.

    Задание требует установить, какой идентификатор действительно обозначает
    один и тот же планер, и не принимать отображаемое имя за технический ключ.
    Ответ по фактической схеме: **на строке вылета такого идентификатора
    нет**.

    * `serial_number` в payload DJI -- идентификатор ВЫЛЕТА, не машины. Здесь
      это перепроверяется на живых данных: доля различных значений к числу
      строк должна быть около единицы.
    * `nickname_raw` -- отображаемое имя с пульта. Оператор его меняет.
    * `drone_unit_id` -- РЕЗУЛЬТАТ разбора ника через `drone_nicknames`,
      единственная связь вылета с машиной. Наследует все слабости ника.
    * `hardware_id` (Body Code) -- настоящий идентификатор планера, но он
      лежит на `drone_units`, а в списке вылетов его нет вовсе. В
      `raw_json` его тоже нет: приёмник кладёт строку списка целиком, а
      список этого поля не содержит.
    """
    facts = {}
    total = con.execute('SELECT count(*) FROM drone_flights').fetchone()[0]
    facts['flights_total'] = total
    facts['distinct_serial_number'] = con.execute(
        'SELECT count(DISTINCT serial_number) FROM drone_flights '
        'WHERE serial_number IS NOT NULL').fetchone()[0]
    facts['serial_number_null'] = con.execute(
        'SELECT count(*) FROM drone_flights '
        'WHERE serial_number IS NULL').fetchone()[0]
    facts['distinct_nickname_raw'] = con.execute(
        'SELECT count(DISTINCT nickname_raw) FROM drone_flights '
        'WHERE nickname_raw IS NOT NULL').fetchone()[0]
    facts['distinct_drone_unit_id'] = con.execute(
        'SELECT count(DISTINCT drone_unit_id) FROM drone_flights '
        'WHERE drone_unit_id IS NOT NULL').fetchone()[0]
    facts['drone_unit_id_null'] = con.execute(
        'SELECT count(*) FROM drone_flights '
        'WHERE drone_unit_id IS NULL').fetchone()[0]

    # Один ник, указывающий на разные машины, -- признак того, что группировка
    # по нику дала бы неверный результат. Проверяется, а не предполагается.
    facts['nicknames_pointing_at_many_units'] = con.execute(
        'SELECT count(*) FROM ('
        '  SELECT nickname_raw FROM drone_flights '
        '  WHERE nickname_raw IS NOT NULL AND drone_unit_id IS NOT NULL '
        '  GROUP BY nickname_raw HAVING count(DISTINCT drone_unit_id) > 1)'
    ).fetchone()[0]

    facts['hardware_id_on_flight_row'] = False
    facts['hardware_id_in_raw_json'] = None  # заполняется при обходе строк

    try:
        facts['units_total'] = con.execute(
            'SELECT count(*) FROM drone_units').fetchone()[0]
        facts['units_with_hardware_id'] = con.execute(
            "SELECT count(*) FROM drone_units WHERE hardware_id IS NOT NULL "
            "AND hardware_id <> ''").fetchone()[0]
    except sqlite3.Error:
        facts['units_total'] = None
        facts['units_with_hardware_id'] = None
    return facts


# ─── Чтение строк ────────────────────────────────────────────────────────────

class FlightRow(object):
    """Одна строка вылета в форме, пригодной для арифметики."""

    __slots__ = ('row_id', 'dji_flight_id', 'unit_id', 'nickname', 'serial',
                 'started_at', 'finished_at', 'work_seconds', 'area_ha_col',
                 'width_col', 'raw_area_m2', 'raw_width', 'width_key_present',
                 'area_key_present', 'raw_ok', 'raw_error',
                 'raw_has_hardware_id', 'local_month', 'local_date',
                 'area_problem', 'width_problem', 'area_from_column',
                 'width_from_column')

    def __init__(self, **kwargs):
        for name in self.__slots__:
            setattr(self, name, kwargs.get(name))

    @property
    def area_m2(self):
        """Площадь в м2 из наиболее точного ПРИГОДНОГО источника, или None.

        Сначала `raw_json.new_work_area`, затем колонка `area_ha` x 10000.
        Непригодное значение (не число, NaN, бесконечность) НЕ используется ни
        из одного источника: возвращается None, а причина уже записана в
        `area_problem` при чтении строки.

        [REASON]: None здесь -- не «ноль», а «нечего складывать». Все суммы и
        доли пропускают такие строки, иначе NaN отравил бы весь итог одним
        значением, а подстановка нуля дала бы тихое занижение гектаров.
        """
        if self.raw_area_m2 is not None:
            return self.raw_area_m2
        if (self.area_from_column and self.area_ha_col is not None
                and _finite_number(self.area_ha_col)):
            return float(self.area_ha_col) * 10000.0
        return None

    @property
    def duration_seconds(self):
        """Длительность: из `work_seconds`, иначе из разности времён."""
        if self.work_seconds is not None and _finite_number(self.work_seconds):
            return self.work_seconds
        if self.started_at is not None and self.finished_at is not None:
            return (self.finished_at - self.started_at).total_seconds()
        return None

    @property
    def width_value(self):
        """Ширина из наиболее надёжного ПРИГОДНОГО источника, или None.

        [REASON]: раздельные `width_state` и `width_value` были дефектом,
        найденным собственным тестом: при неразобранном `raw_json` состояние
        определялось по КОЛОНКЕ, а значение бралось из JSON, где его нет, и
        отчёт падал на `round(None)`. Один источник на оба ответа исключает
        расхождение по построению.
        """
        if self.raw_width is not None:
            return self.raw_width
        if (self.width_from_column and self.width_col is not None
                and _finite_number(self.width_col)):
            return float(self.width_col)
        return None

    def width_state(self, min_width=None, max_width=None):
        """Одно из одиннадцати состояний ширины захвата.

        `MISSING_KEY`            -- ключа spray_width нет в payload вовсе;
        `JSON_NULL`              -- ключ есть, значение ровно null;
        `INVALID_TYPE`           -- ключ есть, значение не число (строка,
                                    массив, объект, boolean);
        `NON_FINITE`             -- число, но NaN или бесконечность;
        `COLUMN_NULL`            -- raw_json непригоден И колонка пуста;
        `MINUS_ONE`              -- ровно -1, наблюдаемый маркер «не записано»;
        `ZERO`                   -- ровно 0, как радиус буфера непригоден;
        `NEGATIVE`               -- иное отрицательное;
        `POSITIVE_UNVALIDATED`   -- положительное, но диапазон не задан;
        `OUT_OF_CONFIGURED_RANGE`-- положительное вне заданных границ;
        `USABLE`                 -- положительное внутри заданных границ.

        [REASON]: одиннадцать состояний, а не «есть/нет», потому что они
        означают разное для диагностики и ведут к разным действиям.
        `MISSING_KEY` -- смена контракта DJI. `INVALID_TYPE` -- дефект на
        стороне DJI или приёмника, и его нельзя молча свалить в одну кучу с
        честным null: null означает «DJI поле знает и для этого вылета не
        записал», а строка вместо числа означает, что сломался формат.
        `USABLE` не выдаётся, пока границы не заданы -- см. DEFAULT_MIN_WIDTH_M.
        """
        if self.width_problem == 'INVALID_TYPE':
            return 'INVALID_TYPE'
        if self.width_problem == 'NON_FINITE':
            return 'NON_FINITE'
        if self.raw_ok and not self.width_key_present:
            return 'MISSING_KEY'
        value = self.width_value
        if value is None:
            if not self.raw_ok:
                return 'COLUMN_NULL'
            return 'JSON_NULL'
        if value == -1:
            return 'MINUS_ONE'
        if value == 0:
            return 'ZERO'
        if value < 0:
            return 'NEGATIVE'
        if min_width is None and max_width is None:
            return 'POSITIVE_UNVALIDATED'
        if min_width is not None and value < min_width:
            return 'OUT_OF_CONFIGURED_RANGE'
        if max_width is not None and value > max_width:
            return 'OUT_OF_CONFIGURED_RANGE'
        return 'USABLE'


# Состояния, при которых ширина В ПРИНЦИПЕ пригодна как радиус буфера.
# `POSITIVE_UNVALIDATED` сюда НЕ входит: пригодность не проверена.
WIDTH_STATES_USABLE = ('USABLE',)

# Состояния, где значение положительное. Нужны для распределения и для
# честного ответа «сколько вылетов имеют хоть какое-то положительное число».
WIDTH_STATES_POSITIVE = ('USABLE', 'POSITIVE_UNVALIDATED',
                         'OUT_OF_CONFIGURED_RANGE')


def _finite_number(value):
    """True только для настоящего конечного числа.

    [REASON]: `isinstance(True, int)` возвращает True, поэтому boolean надо
    отсекать явно -- иначе `spray_width: true` прошло бы как ширина 1.0 метра.
    `math.isfinite` отсекает NaN и обе бесконечности: любое из них, попав в
    сумму, делает весь итог NaN, а в сравнении площадей ведёт себя так, что
    `x == x` ложно, и повтор перестаёт находиться.
    """
    if isinstance(value, bool) or value is None:
        return False
    if not isinstance(value, (int, float)):
        return False
    return math.isfinite(value)


def classify_number(payload, key):
    """(значение, проблема) для числового поля payload.

    Проблема -- одно из None / 'MISSING_KEY' / 'JSON_NULL' / 'INVALID_TYPE' /
    'NON_FINITE'. Значение возвращается только когда проблемы нет.
    """
    if key not in payload:
        return None, 'MISSING_KEY'
    value = payload[key]
    if value is None:
        return None, 'JSON_NULL'
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None, 'INVALID_TYPE'
    if not math.isfinite(value):
        return None, 'NON_FINITE'
    return float(value), None


def parse_datetime(value):
    """SQLite DATETIME -> naive datetime, None когда разобрать нельзя."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace('T', ' ')
    if text.endswith('Z'):
        text = text[:-1]
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
                '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def to_local(moment):
    if moment is None:
        return None
    return moment + datetime.timedelta(minutes=UTC_OFFSET_MINUTES)


def load_rows(con):
    """Все вылеты в память. Ошибка разбора одной строки не прекращает обход.

    [REASON]: 31 тысяча строк -- это десятки мегабайт с `raw_json`, что для
    разового аудита приемлемо и много проще, чем оконные функции в SQL, числа
    которых нечем проверить тестом. Арифметика здесь на Python и покрыта
    тестами; SQL остаётся простым SELECT.
    """
    rows = []
    for record in con.execute(
            'SELECT id, dji_flight_id, drone_unit_id, nickname_raw, '
            '       serial_number, started_at, finished_at, work_seconds, '
            '       area_ha, spray_width, raw_json '
            'FROM drone_flights ORDER BY started_at, id'):
        raw_ok = True
        raw_error = None
        raw_area = None
        raw_width = None
        area_problem = None
        width_problem = None
        width_key_present = False
        area_key_present = False
        raw_has_hw = False
        try:
            payload = json.loads(record['raw_json'])
            if not isinstance(payload, dict):
                raise ValueError('raw_json is not an object')
            area_key_present = RAW_AREA_KEY in payload
            width_key_present = RAW_WIDTH_KEY in payload
            raw_area, area_problem = classify_number(payload, RAW_AREA_KEY)
            raw_width, width_problem = classify_number(payload, RAW_WIDTH_KEY)
            raw_has_hw = 'hardware_id' in payload
        except Exception as exc:          # noqa: BLE001 -- измеряем, не чиним
            raw_ok = False
            raw_error = '%s: %s' % (type(exc).__name__, exc)

        # Запасной источник разрешён ТОЛЬКО когда raw_json целиком непригоден.
        #
        # [REASON]: если raw_json разобрался, а значение в нём оказалось
        # строкой или NaN, подставлять колонку нельзя молча -- это скрыло бы
        # расхождение между источниками. Такая строка остаётся без значения, а
        # причина попадает в отчёт о качестве. Колонка выручает только там,
        # где читать было нечего вовсе.
        area_from_column = (not raw_ok)
        width_from_column = (not raw_ok)

        started = parse_datetime(record['started_at'])
        finished = parse_datetime(record['finished_at'])
        local_started = to_local(started)
        rows.append(FlightRow(
            row_id=record['id'],
            dji_flight_id=record['dji_flight_id'],
            unit_id=record['drone_unit_id'],
            nickname=record['nickname_raw'],
            serial=record['serial_number'],
            started_at=started,
            finished_at=finished,
            work_seconds=record['work_seconds'],
            area_ha_col=record['area_ha'],
            width_col=record['spray_width'],
            raw_area_m2=raw_area,
            raw_width=raw_width,
            area_key_present=area_key_present,
            width_key_present=width_key_present,
            area_problem=area_problem,
            width_problem=width_problem,
            area_from_column=area_from_column,
            width_from_column=width_from_column,
            raw_ok=raw_ok,
            raw_error=raw_error,
            raw_has_hardware_id=raw_has_hw,
            local_month=(local_started.strftime('%Y-%m')
                         if local_started else None),
            local_date=(local_started.date().isoformat()
                        if local_started else None),
        ))
    return rows


# ─── 4.1 Повторы площади ─────────────────────────────────────────────────────

def percentile(values, pct):
    """Процентиль по ближайшему рангу. Пустой вход -> None."""
    if not values:
        return None
    ordered = sorted(values)
    index = int(round((pct / 100.0) * (len(ordered) - 1)))
    return ordered[max(0, min(index, len(ordered) - 1))]


def group_key(row):
    """Ключ группировки «один борт».

    `drone_unit_id`, а не ник: ник это отображаемое имя. Вылеты без машины
    группируются отдельным ключом и НЕ смешиваются в одну кучу с чужими --
    ('unresolved', ник) сохраняет различие между спеллингами.
    """
    if row.unit_id is not None:
        return ('unit', row.unit_id)
    return ('unresolved', row.nickname or '')


def find_repeats(rows, lookback=DEFAULT_LOOKBACK):
    """Кандидаты «площадь равна площади одного из недавних вылетов той же машины».

    Возвращает (список кандидатов, статистика интервалов).

    Кандидат -- это `ANOMALY_CANDIDATE`, а не ошибка: два вылета по одному
    заданию могут дать одну площадь законно. Признаки, сужающие круг,
    считаются отдельно и не складываются в один вердикт.

    `lag` -- на сколько вылетов назад нашлось совпадение. Ближайшее
    совпадение выигрывает: если площадь равна и предыдущему, и позапрошлому,
    записывается lag = 1.
    """
    by_machine = collections.defaultdict(list)
    for row in rows:
        if row.started_at is None:
            continue
        by_machine[group_key(row)].append(row)

    gaps = []
    negative_gap_rows = []
    for machine_rows in by_machine.values():
        machine_rows.sort(key=lambda item: (item.started_at, item.row_id))
        for index in range(1, len(machine_rows)):
            previous, current = machine_rows[index - 1], machine_rows[index]
            end = previous.finished_at or previous.started_at
            if end is None or current.started_at is None:
                continue
            gap = (current.started_at - end).total_seconds()
            if gap < 0:
                # Вылет начался раньше, чем закончился предыдущий. В
                # процентиль такие не идут -- иначе горстка накладок утянет
                # порог ниже нуля и признак «короткий интервал» перестанет
                # означать что-либо. Считаются отдельно, как дефект данных.
                negative_gap_rows.append(current)
                continue
            gaps.append(gap)

    short_gap_threshold = percentile(gaps, SHORT_GAP_PERCENTILE)

    candidates = []
    for key, machine_rows in by_machine.items():
        run_length = 1
        for index in range(1, len(machine_rows)):
            current = machine_rows[index]
            area_now = current.area_m2
            immediate = machine_rows[index - 1].area_m2

            # Серия считается только по непосредственным соседям: три подряд
            # одинаковых -- это серия, а совпадение через один -- нет.
            if (area_now is not None and immediate is not None
                    and area_now == immediate):
                run_length += 1
            else:
                run_length = 1

            if area_now is None or area_now == 0:
                # Нулевая площадь совпадает у соседей сплошь и рядом; это не
                # тот случай, ради которого затеян разбор.
                continue

            match = None
            for lag in range(1, min(lookback, index) + 1):
                candidate_previous = machine_rows[index - lag]
                if candidate_previous.area_m2 == area_now:
                    match = (lag, candidate_previous)
                    break
            if match is None:
                continue
            lag, previous = match

            end = previous.finished_at or previous.started_at
            gap = None
            if end is not None and current.started_at is not None:
                gap = (current.started_at - end).total_seconds()
            duration = current.duration_seconds
            candidates.append({
                'machine': key,
                'dji_flight_id': current.dji_flight_id,
                'previous_dji_flight_id': previous.dji_flight_id,
                'lag': lag,
                'local_date': current.local_date,
                'local_month': current.local_month,
                'nickname': current.nickname,
                'unit_id': current.unit_id,
                'area_m2': area_now,
                'area_ha': area_now / 10000.0,
                'duration_seconds': duration,
                'gap_seconds': gap,
                'run_length': run_length,
                'short_flight': (duration is not None
                                 and duration < SHORT_FLIGHT_SECONDS),
                'short_gap': (gap is not None and gap >= 0
                              and short_gap_threshold is not None
                              and gap <= short_gap_threshold),
            })
    return candidates, {
        'gap_sample_size': len(gaps),
        'short_gap_threshold_seconds': short_gap_threshold,
        'short_gap_percentile': SHORT_GAP_PERCENTILE,
        'lookback': lookback,
        'negative_gap_rows': negative_gap_rows,
    }


# ─── 4.2 Ширина ──────────────────────────────────────────────────────────────

def positive_width_distribution(values):
    """Распределение положительных значений ширины.

    Выбросы считаются критерием Тьюки (за пределами Q1-1.5*IQR и Q3+1.5*IQR)
    и НЕ объявляются ошибкой: редкое значение и неверное значение -- разные
    вещи, и различить их этими данными нельзя.
    """
    if not values:
        return {'count': 0, 'min': None, 'max': None, 'median': None,
                'p05': None, 'p25': None, 'p75': None, 'p95': None,
                'iqr_low': None, 'iqr_high': None, 'outliers': 0,
                'outlier_values': []}
    ordered = sorted(values)
    q1 = percentile(ordered, 25)
    q3 = percentile(ordered, 75)
    iqr = q3 - q1
    low = q1 - OUTLIER_IQR_FACTOR * iqr
    high = q3 + OUTLIER_IQR_FACTOR * iqr
    outliers = [value for value in ordered if value < low or value > high]
    return {
        'count': len(ordered),
        'min': ordered[0],
        'max': ordered[-1],
        'median': percentile(ordered, 50),
        'p05': percentile(ordered, 5),
        'p25': q1,
        'p75': q3,
        'p95': percentile(ordered, 95),
        'iqr_low': low,
        'iqr_high': high,
        'outliers': len(outliers),
        'outlier_values': sorted(set(outliers)),
    }


def width_report(rows, min_width=None, max_width=None):
    states = collections.Counter()
    area_by_state = collections.Counter()
    by_month = collections.defaultdict(
        lambda: {'flights': 0, 'usable': 0, 'positive': 0, 'area_ha': 0.0,
                 'unusable_area_ha': 0.0})
    by_machine = collections.defaultdict(
        lambda: {'flights': 0, 'usable': 0, 'positive': 0, 'area_ha': 0.0,
                 'unusable_area_ha': 0.0, 'nickname': None})
    distinct_widths = collections.Counter()
    positive_values = []

    for row in rows:
        state = row.width_state(min_width, max_width)
        states[state] += 1
        area = row.area_m2
        area_ha = (area / 10000.0) if area is not None else 0.0
        area_by_state[state] += area or 0.0
        positive = state in WIDTH_STATES_POSITIVE
        usable = state in WIDTH_STATES_USABLE
        if positive:
            distinct_widths[round(row.width_value, 4)] += 1
            positive_values.append(row.width_value)

        month = by_month[row.local_month or '(нет даты)']
        month['flights'] += 1
        month['area_ha'] += area_ha
        if usable:
            month['usable'] += 1
        if positive:
            month['positive'] += 1
        if not usable:
            month['unusable_area_ha'] += area_ha

        key = group_key(row)
        machine = by_machine[key]
        machine['flights'] += 1
        machine['area_ha'] += area_ha
        machine['nickname'] = machine['nickname'] or row.nickname
        if usable:
            machine['usable'] += 1
        if positive:
            machine['positive'] += 1
        if not usable:
            machine['unusable_area_ha'] += area_ha

    months_with_positive = sorted(
        month for month, data in by_month.items()
        if data['positive'] > 0 and month != '(нет даты)')
    months_without_positive = sorted(
        month for month, data in by_month.items()
        if data['positive'] == 0 and data['flights'] > 0
        and month != '(нет даты)')

    return {
        'states': dict(states),
        'area_ha_by_state': {name: value / 10000.0
                             for name, value in area_by_state.items()},
        'by_month': dict(by_month),
        'by_machine': dict(by_machine),
        'distinct_widths': dict(distinct_widths),
        'distribution': positive_width_distribution(positive_values),
        'first_month_with_width': (months_with_positive[0]
                                   if months_with_positive else None),
        'last_month_with_width': (months_with_positive[-1]
                                  if months_with_positive else None),
        'months_without_any_width': months_without_positive,
        'min_width_m': min_width,
        'max_width_m': max_width,
    }


# ─── 4.3 Качество данных ─────────────────────────────────────────────────────

def quality_report(rows, con):
    issues = collections.Counter()
    samples = collections.defaultdict(list)

    def note(kind, row, detail=''):
        issues[kind] += 1
        if len(samples[kind]) < 20:
            samples[kind].append({
                'dji_flight_id': row.dji_flight_id,
                'row_id': row.row_id,
                'local_date': row.local_date,
                'nickname': row.nickname,
                'detail': detail,
            })

    now_limit = datetime.datetime.utcnow() + datetime.timedelta(days=2)
    earliest = datetime.datetime(2024, 1, 1)

    problem_labels = {
        'INVALID_TYPE': 'значение не число',
        'NON_FINITE': 'значение NaN или бесконечность',
        'MISSING_KEY': 'ключа нет в payload',
        'JSON_NULL': 'значение null',
    }

    for row in rows:
        if not row.raw_ok:
            note('raw_json не разобрался', row, row.raw_error or '')
            if row.area_from_column and row.area_ha_col is not None:
                note('площадь взята из колонки: raw_json непригоден', row,
                     'колонка %s' % row.area_ha_col)
            if row.width_from_column and row.width_col is not None:
                note('ширина взята из колонки: raw_json непригоден', row,
                     'колонка %s' % row.width_col)
        if row.area_problem in problem_labels:
            note('new_work_area: %s' % problem_labels[row.area_problem], row)
        if row.width_problem in ('INVALID_TYPE', 'NON_FINITE'):
            note('spray_width: %s' % problem_labels[row.width_problem], row)
        if (row.area_ha_col is not None
                and not _finite_number(row.area_ha_col)):
            note('area_ha в колонке не конечное число', row,
                 repr(row.area_ha_col))
        if (row.width_col is not None
                and not _finite_number(row.width_col)):
            note('spray_width в колонке не конечное число', row,
                 repr(row.width_col))
        if row.dji_flight_id is None:
            note('нет dji_flight_id', row)
        if row.started_at is None:
            note('нет или не разобрано время начала', row)
        else:
            if row.started_at > now_limit:
                note('время начала в будущем', row, str(row.started_at))
            if row.started_at < earliest:
                note('время начала раньше 2024 года', row, str(row.started_at))
        if (row.started_at is not None and row.finished_at is not None
                and row.finished_at < row.started_at):
            note('конец раньше начала', row)
        duration = row.duration_seconds
        if duration is not None and duration < 0:
            note('отрицательная длительность', row, str(duration))
        elif duration == 0:
            note('нулевая длительность', row)
        area = row.area_m2
        if area is not None and area < 0:
            note('отрицательная площадь', row, str(area))
        if row.unit_id is None:
            note('вылет не привязан к машине', row, row.nickname or '')
        # Расхождение колонки и JSON. Колонка -- это m2/10000, поэтому
        # сравнение ведётся в гектарах с допуском на двоичное представление.
        if (row.raw_area_m2 is not None and row.area_ha_col is not None
                and _finite_number(row.area_ha_col)):
            expected = row.raw_area_m2 / 10000.0
            if abs(expected - row.area_ha_col) > 1e-9:
                note('площадь в колонке и в raw_json расходятся', row,
                     'колонка %.10f, json %.10f' % (row.area_ha_col, expected))
        # Ширина в колонке против ширины в JSON.
        if (row.raw_ok and row.width_key_present
                and row.raw_width is not None
                and (row.width_col is None or _finite_number(row.width_col))):
            if row.width_col is None:
                note('ширина есть в raw_json, но колонка пуста', row,
                     str(row.raw_width))
            elif abs(row.width_col - row.raw_width) > 1e-9:
                note('ширина в колонке и в raw_json расходятся', row,
                     'колонка %s, json %s' % (row.width_col, row.raw_width))

    duplicates = con.execute(
        'SELECT count(*) FROM (SELECT dji_flight_id FROM drone_flights '
        'GROUP BY dji_flight_id HAVING count(*) > 1)').fetchone()[0]
    if duplicates:
        issues['повторяющийся dji_flight_id'] = duplicates

    return {'issues': dict(issues), 'samples': dict(samples)}


# ─── Сводка ──────────────────────────────────────────────────────────────────

def analyse(con, lookback=DEFAULT_LOOKBACK, min_width=None, max_width=None):
    """Вся арифметика аудита, без openpyxl -- числа проверяются тестами."""
    present, missing = describe_schema(con)
    if missing:
        raise ProbeError('drone_flights is missing required column(s): %s'
                         % ', '.join(missing))

    identity = machine_identity(con)
    rows = load_rows(con)
    candidates, gap_stats = find_repeats(rows, lookback=lookback)
    widths = width_report(rows, min_width=min_width, max_width=max_width)
    quality = quality_report(rows, con)

    # Накладка времён -- дефект данных, а не признак повтора площади.
    if gap_stats['negative_gap_rows']:
        quality['issues']['вылет начался раньше конца предыдущего'] = len(
            gap_stats['negative_gap_rows'])
        quality['samples']['вылет начался раньше конца предыдущего'] = [
            {'dji_flight_id': row.dji_flight_id, 'row_id': row.row_id,
             'local_date': row.local_date, 'nickname': row.nickname,
             'detail': ''}
            for row in gap_stats['negative_gap_rows'][:20]]
    gap_stats = {name: value for name, value in gap_stats.items()
                 if name != 'negative_gap_rows'}

    identity['hardware_id_in_raw_json'] = sum(
        1 for row in rows if row.raw_has_hardware_id)

    # Строки без пригодной площади в сумму не идут вовсе -- ни нулём, ни NaN.
    areas = [row.area_m2 for row in rows if row.area_m2 is not None]
    total_area_ha = sum(areas) / 10000.0
    rows_without_area = len(rows) - len(areas)
    candidate_area_ha = sum(item['area_ha'] for item in candidates)

    by_month = collections.defaultdict(
        lambda: {'flights': 0, 'area_ha': 0.0, 'repeat_flights': 0,
                 'repeat_area_ha': 0.0})
    for row in rows:
        month = by_month[row.local_month or '(нет даты)']
        month['flights'] += 1
        month['area_ha'] += (row.area_m2 or 0.0) / 10000.0  # None -> не в сумму
    for item in candidates:
        month = by_month[item['local_month'] or '(нет даты)']
        month['repeat_flights'] += 1
        month['repeat_area_ha'] += item['area_ha']

    by_machine = collections.defaultdict(
        lambda: {'flights': 0, 'area_ha': 0.0, 'repeat_flights': 0,
                 'repeat_area_ha': 0.0, 'nickname': None})
    for row in rows:
        machine = by_machine[group_key(row)]
        machine['flights'] += 1
        machine['area_ha'] += (row.area_m2 or 0.0) / 10000.0
        machine['nickname'] = machine['nickname'] or row.nickname
    for item in candidates:
        machine = by_machine[item['machine']]
        machine['repeat_flights'] += 1
        machine['repeat_area_ha'] += item['area_ha']

    usable = sum(widths['states'].get(name, 0)
                 for name in WIDTH_STATES_USABLE)
    usable_area_ha = sum(widths['area_ha_by_state'].get(name, 0.0)
                         for name in WIDTH_STATES_USABLE)
    positive = sum(widths['states'].get(name, 0)
                   for name in WIDTH_STATES_POSITIVE)
    positive_area_ha = sum(widths['area_ha_by_state'].get(name, 0.0)
                           for name in WIDTH_STATES_POSITIVE)

    return {
        'schema_columns': present,
        'identity': identity,
        'flights_total': len(rows),
        'total_area_ha': total_area_ha,
        'rows_without_usable_area': rows_without_area,
        'known_case_id': KNOWN_CASE_FLIGHT_ID,
        'flight_ids': {row.dji_flight_id for row in rows},
        'repeats': {
            'candidates': candidates,
            'count': len(candidates),
            'area_ha': candidate_area_ha,
            'share_of_flights': (len(candidates) / len(rows)) if rows else 0.0,
            'share_of_area': (candidate_area_ha / total_area_ha)
                             if total_area_ha else 0.0,
            'short_flight_count': sum(1 for item in candidates
                                      if item['short_flight']),
            'short_gap_count': sum(1 for item in candidates
                                   if item['short_gap']),
            'runs_3plus_count': sum(1 for item in candidates
                                    if item['run_length'] >= RUN_MIN_LENGTH),
            'lag1_count': sum(1 for item in candidates if item['lag'] == 1),
            'lag2plus_count': sum(1 for item in candidates if item['lag'] >= 2),
            'lag1_area_ha': sum(item['area_ha'] for item in candidates
                                if item['lag'] == 1),
            'lag2plus_area_ha': sum(item['area_ha'] for item in candidates
                                    if item['lag'] >= 2),
            'gap_stats': gap_stats,
        },
        'widths': widths,
        'width_usable_flights': usable,
        'width_usable_share': (usable / len(rows)) if rows else 0.0,
        'width_usable_area_ha': usable_area_ha,
        'width_usable_area_share': (usable_area_ha / total_area_ha)
                                   if total_area_ha else 0.0,
        'width_positive_flights': positive,
        'width_positive_share': (positive / len(rows)) if rows else 0.0,
        'width_positive_area_ha': positive_area_ha,
        'width_positive_area_share': (positive_area_ha / total_area_ha)
                                     if total_area_ha else 0.0,
        'quality': quality,
        'by_month': dict(by_month),
        'by_machine': dict(by_machine),
    }


def check_known_case(result, dji_flight_id=None):
    """Предусловие прогона: известный случай из DISCOVERY §6.2 обязан найтись.

    [REASON]: задание требует не продолжать молча, если он не найден, и это
    правильно: отсутствие означает одно из трёх, и все три обесценивают
    отчёт целиком.

    * вылета нет в базе -- значит анализируется не тот срез;
    * вылет есть, но кандидатом не опознан -- значит правило сравнения
      работает не так, как в разборе, и все остальные кандидаты под вопросом;
    * опознан на другом расстоянии -- значит порядок вылетов в базе не тот,
      что в снимке.

    Проверка отличает эти случаи, потому что действия по ним разные.
    """
    if dji_flight_id is None:
        dji_flight_id = result.get('known_case_id', KNOWN_CASE_FLIGHT_ID)
    present = dji_flight_id in result.get('flight_ids', set())
    found = [item for item in result['repeats']['candidates']
             if item['dji_flight_id'] == dji_flight_id]
    detail = found[0] if found else None
    if not present:
        reason = ('вылет %s отсутствует в этой базе' % dji_flight_id)
    elif detail is None:
        reason = ('вылет %s есть в базе, но кандидатом не опознан'
                  % dji_flight_id)
    elif detail['lag'] != KNOWN_CASE_EXPECTED_LAG:
        reason = ('вылет %s опознан на расстоянии %d, ожидалось %d'
                  % (dji_flight_id, detail['lag'], KNOWN_CASE_EXPECTED_LAG))
    else:
        reason = None
    return {
        'dji_flight_id': dji_flight_id,
        'present_in_database': present,
        'found_as_candidate': bool(found),
        'passed': reason is None,
        'reason': reason,
        'detail': detail,
    }


# ─── Вывод ───────────────────────────────────────────────────────────────────

def machine_label(key, nickname):
    kind, value = key
    if kind == 'unit':
        return 'машина id=%s (%s)' % (value, nickname or '?')
    return 'без машины: %s' % (value or '(пустой ник)')


def print_console_summary(result, known, db_path, digest_before):
    """ASCII only -- консоль Windows."""
    widths = result['widths']
    print('=' * 72)
    print('A2 fleet audit  --  READ ONLY')
    print('database : %s' % db_path)
    print('sha256   : %s' % digest_before)
    print('=' * 72)
    print('flights total          : %d' % result['flights_total'])
    print('area total, ha         : %.2f' % result['total_area_ha'])
    if result['rows_without_usable_area']:
        print('rows with NO usable area (excluded from every sum): %d'
              % result['rows_without_usable_area'])
    print('')
    print('-- width of swath ------------------------------------------------')
    bounds = (widths['min_width_m'], widths['max_width_m'])
    if bounds == (None, None):
        print('  configured range: NONE. No flight can be USABLE; a positive')
        print('  value is reported as POSITIVE_UNVALIDATED. Pass --min-width-m')
        print('  and --max-width-m once the range is agreed with the owner.')
    else:
        print('  configured range: min=%s max=%s (metres)'
              % (bounds[0], bounds[1]))
    for state in ('USABLE', 'POSITIVE_UNVALIDATED', 'OUT_OF_CONFIGURED_RANGE',
                  'MINUS_ONE', 'ZERO', 'NEGATIVE', 'JSON_NULL', 'MISSING_KEY',
                  'INVALID_TYPE', 'NON_FINITE', 'COLUMN_NULL'):
        count = widths['states'].get(state, 0)
        if count:
            area = widths['area_ha_by_state'].get(state, 0.0)
            print('  %-24s %8d flights  %12.2f ha' % (state, count, area))
    print('  positive value at all : %d flights (%.1f%%), %.2f ha (%.1f%%)'
          % (result['width_positive_flights'],
             100.0 * result['width_positive_share'],
             result['width_positive_area_ha'],
             100.0 * result['width_positive_area_share']))
    print('  USABLE (validated)    : %d flights (%.1f%%), %.2f ha (%.1f%%)'
          % (result['width_usable_flights'],
             100.0 * result['width_usable_share'],
             result['width_usable_area_ha'],
             100.0 * result['width_usable_area_share']))
    distribution = widths['distribution']
    if distribution['count']:
        print('  positive width distribution, m:')
        print('    min %.3f  p05 %.3f  p25 %.3f  median %.3f  p75 %.3f  '
              'p95 %.3f  max %.3f'
              % (distribution['min'], distribution['p05'], distribution['p25'],
                 distribution['median'], distribution['p75'],
                 distribution['p95'], distribution['max']))
        print('    distinct values: %d,  statistical outliers (Tukey): %d'
              % (len(widths['distinct_widths']), distribution['outliers']))
        print('    an outlier is a RARE value, not an error of DJI')
    print('')
    print('-- repeated area (ANOMALY_CANDIDATE, not an error) ---------------')
    repeats = result['repeats']
    print('  candidates           : %d (%.2f%% of flights)'
          % (repeats['count'], 100.0 * repeats['share_of_flights']))
    print('  area of candidates   : %.2f ha (%.2f%% of all hectares)'
          % (repeats['area_ha'], 100.0 * repeats['share_of_area']))
    print('  of them shorter than %ds : %d'
          % (SHORT_FLIGHT_SECONDS, repeats['short_flight_count']))
    threshold = repeats['gap_stats']['short_gap_threshold_seconds']
    print('  of them after a short gap (<= %s s, p%d of %d gaps) : %d'
          % ('n/a' if threshold is None else '%.0f' % threshold,
             SHORT_GAP_PERCENTILE, repeats['gap_stats']['gap_sample_size'],
             repeats['short_gap_count']))
    print('  of them in a run of %d+ : %d'
          % (RUN_MIN_LENGTH, repeats['runs_3plus_count']))
    print('  match with the immediately previous flight (lag 1) : %d, %.2f ha'
          % (repeats['lag1_count'], repeats['lag1_area_ha']))
    print('  match further back (lag 2..%d)                     : %d, %.2f ha'
          % (repeats['gap_stats']['lookback'], repeats['lag2plus_count'],
             repeats['lag2plus_area_ha']))
    print('')
    print('-- known case (PRECONDITION) -------------------------------------')
    print('  flight %d : %s' % (known['dji_flight_id'],
                                'PASS' if known['passed'] else 'FAIL'))
    if not known['passed']:
        print('  reason: %s' % _ascii(known['reason']))
        print('  The report is NOT valid for a decision. Establish the reason')
        print('  before reading any number above.')
    print('')
    print('-- data quality --------------------------------------------------')
    if not result['quality']['issues']:
        print('  no issues found')
    for kind, count in sorted(result['quality']['issues'].items(),
                              key=lambda pair: -pair[1]):
        print('  %-28s %6d rows' % (QUALITY_CODES.get(kind, 'OTHER'), count))
    print('')
    print('-- machine identity ----------------------------------------------')
    identity = result['identity']
    print('  distinct serial_number / flights : %d / %d  '
          '(serial is a FLIGHT id, not a machine id)'
          % (identity['distinct_serial_number'], identity['flights_total']))
    print('  flights with no machine resolved : %d'
          % identity['drone_unit_id_null'])
    print('  nicknames pointing at >1 machine : %d'
          % identity['nicknames_pointing_at_many_units'])
    print('  hardware_id present in raw_json  : %d rows'
          % (identity['hardware_id_in_raw_json'] or 0))
    print('=' * 72)


def _ascii(text):
    """Кириллица в консоль Windows не идёт -- заменяем, не роняя прогон."""
    if text is None:
        return ''
    return str(text).encode('ascii', 'replace').decode('ascii')


def write_xlsx(result, known, path, db_path, digest_before, digest_after):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    book = Workbook()

    def sheet(title, headers, widths):
        ws = book.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = 'A1:%s1' % get_column_letter(len(headers))
        return ws

    # --- Сводка --------------------------------------------------------------
    ws = book.active
    ws.title = 'Сводка'
    ws.column_dimensions['A'].width = 58
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 60

    # [REASON]: отчёт, не прошедший контрольный случай, обязан кричать об этом
    # с первой строки. Читатель может открыть файл, не видев консоли, и цифры
    # из непроверенного прогона выглядят ровно так же убедительно, как из
    # проверенного.
    if not known['passed']:
        ws.append(['НЕ ПРОШЁЛ КОНТРОЛЬНЫЙ СЛУЧАЙ — НЕ ИСПОЛЬЗОВАТЬ ДЛЯ РЕШЕНИЯ'])
        ws.append(['Причина: %s' % (known['reason'] or '')])
        ws.append(['Числа ниже могут быть неполными или неверными. Сначала '
                   'установите причину.'])
        ws.append([])
        marker = Font(bold=True, size=20, color='FFFFFFFF')
        fill = PatternFill('solid', fgColor='FFC00000')
        for row_index in (1, 2, 3):
            cell = ws.cell(row=row_index, column=1)
            cell.font = marker if row_index == 1 else Font(bold=True, size=12)
            cell.fill = fill
        ws.row_dimensions[1].height = 34

    repeats = result['repeats']
    widths = result['widths']
    threshold = repeats['gap_stats']['short_gap_threshold_seconds']
    lines = [
        ('DRONE-COVERAGE-001, этап A2 — аудит парка', '', ''),
        ('Режим', 'только чтение', 'file:...?mode=ro, пишущих команд нет'),
        ('База (копия)', os.path.basename(db_path), db_path),
        ('SHA-256 до анализа', digest_before, ''),
        ('SHA-256 после анализа', digest_after,
         'совпадение доказывает, что файл не изменён'),
        ('', '', ''),
        ('Вылетов всего', result['flights_total'], ''),
        ('Гектаров всего (по данным DJI)', round(result['total_area_ha'], 2),
         'сумма new_work_area / 10000'),
        ('Строк без пригодной площади', result['rows_without_usable_area'],
         'не число, NaN или бесконечность в обоих источниках; в суммы не идут'),
        ('', '', ''),
        ('ШИРИНА ЗАХВАТА', '', ''),
        ('Заданный диапазон, м',
         ('%s … %s' % (widths['min_width_m'], widths['max_width_m'])
          if (widths['min_width_m'] is not None
              or widths['max_width_m'] is not None) else 'НЕ ЗАДАН'),
         ('Пока диапазон не задан, статус USABLE не получает никто, а '
          'положительное значение называется POSITIVE_UNVALIDATED. '
          'Задаётся аргументами --min-width-m и --max-width-m.')),
        ('Вылетов с положительной шириной', result['width_positive_flights'],
         'значение похоже на ширину; пригодность НЕ проверена'),
        ('Доля вылетов с положительной шириной',
         round(100.0 * result['width_positive_share'], 2), 'процентов'),
        ('Гектаров с положительной шириной',
         round(result['width_positive_area_ha'], 2), ''),
        ('Вылетов USABLE (в заданном диапазоне)',
         result['width_usable_flights'],
         'только для них возможен геометрический расчёт'),
        ('Доля вылетов USABLE',
         round(100.0 * result['width_usable_share'], 2), 'процентов'),
        ('Гектаров USABLE', round(result['width_usable_area_ha'], 2), ''),
        ('Доля гектаров USABLE',
         round(100.0 * result['width_usable_area_share'], 2), 'процентов'),
        ('Гектаров DATA_UNAVAILABLE',
         round(result['total_area_ha'] - result['width_usable_area_ha'], 2),
         'подстановка ширины запрещена решением владельца 2026-08-25'),
        ('', '', ''),
        ('ПОВТОРЫ ПЛОЩАДИ (ANOMALY_CANDIDATE)', '',
         'совпадение само по себе не ошибка DJI'),
        ('Кандидатов', repeats['count'], ''),
        ('Доля от вылетов', round(100.0 * repeats['share_of_flights'], 3),
         'процентов'),
        ('Гектаров у кандидатов', round(repeats['area_ha'], 2), ''),
        ('Доля от гектаров', round(100.0 * repeats['share_of_area'], 3),
         'процентов'),
        ('Из них короче %d с' % SHORT_FLIGHT_SECONDS,
         repeats['short_flight_count'], ''),
        ('Из них после короткого интервала', repeats['short_gap_count'],
         ('порог %.0f с — %d-й процентиль интервалов, вычислен по этим же '
          'данным' % (threshold, SHORT_GAP_PERCENTILE))
         if threshold is not None else 'интервалов не набралось'),
        ('Из них в серии из %d и более' % RUN_MIN_LENGTH,
         repeats['runs_3plus_count'], ''),
        ('Совпадение с непосредственно предыдущим (lag 1)',
         repeats['lag1_count'],
         'гектаров: %.2f' % repeats['lag1_area_ha']),
        ('Совпадение через один и дальше (lag 2..%d)'
         % repeats['gap_stats']['lookback'], repeats['lag2plus_count'],
         'гектаров: %.2f. Известный случай 622715275 именно такой'
         % repeats['lag2plus_area_ha']),
        ('', '', ''),
        ('', '', ''),
        ('КОНТРОЛЬНЫЙ СЛУЧАЙ (предусловие)', '', ''),
        ('Вылет %d' % known['dji_flight_id'],
         'ПРОЙДЕН' if known['passed'] else 'НЕ ПРОЙДЕН',
         known['reason'] or 'опознан кандидатом на ожидаемом расстоянии'),
        ('Есть в базе', 'да' if known['present_in_database'] else 'НЕТ', ''),
        ('Опознан кандидатом',
         'да' if known['found_as_candidate'] else 'НЕТ', ''),
    ]
    for row in lines:
        ws.append(list(row))
    for cell in ws['A']:
        cell.font = bold
    for cell in ws['C']:
        cell.alignment = wrap

    # --- По месяцам ----------------------------------------------------------
    ws = sheet('По месяцам',
               ['Месяц (UTC+5)', 'Вылетов', 'Гектаров',
                'С положительной шириной', 'USABLE',
                'Доля USABLE, %', 'Гектаров не-USABLE',
                'Кандидатов повтора', 'Гектаров у кандидатов'],
               [16, 12, 14, 24, 12, 18, 22, 20, 22])
    months = sorted(set(list(result['by_month'].keys())
                        + list(result['widths']['by_month'].keys())))
    for month in months:
        base = result['by_month'].get(
            month, {'flights': 0, 'area_ha': 0.0, 'repeat_flights': 0,
                    'repeat_area_ha': 0.0})
        width = result['widths']['by_month'].get(
            month, {'flights': 0, 'usable': 0, 'positive': 0,
                    'unusable_area_ha': 0.0})
        share = (100.0 * width['usable'] / width['flights']
                 if width['flights'] else 0.0)
        ws.append([month, base['flights'], round(base['area_ha'], 2),
                   width.get('positive', 0), width['usable'], round(share, 1),
                   round(width['unusable_area_ha'], 2),
                   base['repeat_flights'], round(base['repeat_area_ha'], 2)])

    # --- По дронам -----------------------------------------------------------
    ws = sheet('По дронам',
               ['Машина', 'Ник (первый встреченный)', 'Вылетов', 'Гектаров',
                'С положительной шириной', 'USABLE', 'Доля USABLE, %',
                'Гектаров не-USABLE', 'Кандидатов повтора',
                'Гектаров у кандидатов'],
               [26, 26, 12, 14, 24, 12, 18, 22, 20, 22])
    for key in sorted(result['by_machine'], key=lambda item: (item[0], str(item[1]))):
        base = result['by_machine'][key]
        width = result['widths']['by_machine'].get(
            key, {'flights': 0, 'usable': 0, 'positive': 0,
                  'unusable_area_ha': 0.0})
        share = (100.0 * width['usable'] / width['flights']
                 if width['flights'] else 0.0)
        ws.append([machine_label(key, base['nickname']), base['nickname'] or '',
                   base['flights'], round(base['area_ha'], 2),
                   width.get('positive', 0), width['usable'], round(share, 1),
                   round(width['unusable_area_ha'], 2),
                   base['repeat_flights'], round(base['repeat_area_ha'], 2)])

    # --- Повторы площади -----------------------------------------------------
    ws = sheet('Повторы площади',
               ['Дата (UTC+5)', 'Машина', 'Ник', 'dji_flight_id',
                'Совпало с dji_flight_id', 'Вылетов назад (lag)',
                'Площадь, га', 'Длительность, с',
                'Интервал, с', 'Длина серии',
                'Короткий вылет', 'Короткий интервал'],
               [14, 24, 20, 18, 24, 20, 14, 18, 18, 14, 16, 18])
    for item in sorted(result['repeats']['candidates'],
                       key=lambda entry: (entry['local_date'] or '',
                                          entry['dji_flight_id'] or 0)):
        ws.append([
            item['local_date'] or '',
            machine_label(item['machine'], item['nickname']),
            item['nickname'] or '',
            item['dji_flight_id'], item['previous_dji_flight_id'],
            item['lag'],
            round(item['area_ha'], 4),
            '' if item['duration_seconds'] is None else round(item['duration_seconds']),
            '' if item['gap_seconds'] is None else round(item['gap_seconds']),
            item['run_length'],
            'да' if item['short_flight'] else '',
            'да' if item['short_gap'] else '',
        ])

    # --- Нет ширины ----------------------------------------------------------
    ws = sheet('Нет ширины',
               ['Состояние', 'Что это значит', 'Вылетов', 'Гектаров DJI'],
               [16, 62, 12, 16])
    meanings = {
        'USABLE': 'положительное И внутри заданного диапазона — только здесь '
                  'геометрию посчитать можно',
        'POSITIVE_UNVALIDATED': 'положительное, но диапазон не задан — '
                                'пригодность НЕ проверена',
        'OUT_OF_CONFIGURED_RANGE': 'положительное, но вне заданных границ',
        'MINUS_ONE': 'DJI прислал -1 — наблюдаемый маркер «не записано»',
        'JSON_NULL': 'ключ есть, значение null — DJI поле знает, но не записал',
        'MISSING_KEY': 'ключа spray_width в payload нет вовсе — смена контракта DJI',
        'INVALID_TYPE': 'ключ есть, значение не число (строка, массив, объект, '
                        'boolean) — сломан формат, а не «не записано»',
        'NON_FINITE': 'число, но NaN или бесконечность — в расчёт не идёт',
        'ZERO': 'ровно 0 — как радиус буфера непригоден так же, как -1',
        'NEGATIVE': 'иное отрицательное значение',
        'COLUMN_NULL': 'raw_json непригоден И колонка пуста',
    }
    for state, count in sorted(result['widths']['states'].items(),
                               key=lambda pair: -pair[1]):
        ws.append([state, meanings.get(state, ''), count,
                   round(result['widths']['area_ha_by_state'].get(state, 0.0), 2)])
    ws.append([])
    ws.append(['Заданный диапазон, м',
               ('%s … %s' % (widths['min_width_m'], widths['max_width_m'])
                if (widths['min_width_m'] is not None
                    or widths['max_width_m'] is not None) else 'НЕ ЗАДАН'),
               '', ''])
    ws.append([])
    distribution = widths['distribution']
    ws.append(['Распределение ПОЛОЖИТЕЛЬНЫХ значений ширины', '', '', ''])
    for label, key in (('значений', 'count'), ('минимум', 'min'),
                       ('5-й процентиль', 'p05'), ('25-й процентиль', 'p25'),
                       ('медиана', 'median'), ('75-й процентиль', 'p75'),
                       ('95-й процентиль', 'p95'), ('максимум', 'max')):
        ws.append([label, distribution[key], '', ''])
    ws.append(['статистических выбросов (критерий Тьюки)',
               distribution['outliers'],
               'выброс — РЕДКОЕ значение, а не ошибка DJI; за границами '
               '%s … %s' % (
                   None if distribution['iqr_low'] is None
                   else round(distribution['iqr_low'], 3),
                   None if distribution['iqr_high'] is None
                   else round(distribution['iqr_high'], 3)), ''])
    if distribution['outlier_values']:
        ws.append(['значения выбросов',
                   ', '.join(str(round(value, 3))
                             for value in distribution['outlier_values']),
                   '', ''])
    ws.append([])
    ws.append(['Различные положительные значения ширины', '', '', ''])
    ws.append(['Ширина, м', 'Вылетов', '', ''])
    for value, count in sorted(widths['distinct_widths'].items()):
        ws.append([value, count, '', ''])
    ws.append([])
    ws.append(['Первый месяц с положительной шириной',
               widths['first_month_with_width'] or '—', '', ''])
    ws.append(['Последний месяц с положительной шириной',
               widths['last_month_with_width'] or '—', '', ''])
    ws.append(['Месяцы без единой положительной ширины',
               ', '.join(widths['months_without_any_width']) or '—', '', ''])

    # --- Качество данных -----------------------------------------------------
    ws = sheet('Качество данных',
               ['Проблема', 'Код', 'Строк', 'Примеры dji_flight_id'],
               [56, 24, 12, 60])
    if not result['quality']['issues']:
        ws.append(['проблем не найдено', '', 0, ''])
    for kind, count in sorted(result['quality']['issues'].items(),
                              key=lambda pair: -pair[1]):
        examples = ', '.join(
            str(sample['dji_flight_id'])
            for sample in result['quality']['samples'].get(kind, [])[:10])
        ws.append([kind, QUALITY_CODES.get(kind, 'OTHER'), count, examples])
    for cell in ws['D']:
        cell.alignment = wrap

    # --- Методика ------------------------------------------------------------
    ws = book.create_sheet('Методика')
    ws.column_dimensions['A'].width = 110
    method = [
        'DRONE-COVERAGE-001, этап A2. Методика и её границы.',
        '',
        'РЕЖИМ. База открыта через file:-URI с mode=ro. Пишущих SQL-команд в '
        'скрипте нет. SHA-256 файла до и после анализа приведён на листе '
        '«Сводка» — совпадение и есть доказательство неизменности.',
        '',
        'ИСТОЧНИК ПЛОЩАДИ. Берётся new_work_area из raw_json (полная '
        'точность). Колонка area_ha равна m2/10000 и знаки уже потеряла: '
        '5940.000029700001 превращается в 0.5940000029700001. Когда raw_json '
        'не разобрался, используется колонка, а расхождение между источниками '
        'считается отдельно на листе «Качество данных».',
        '',
        'ГРУППИРОВКА ПО БОРТУ. На строке вылета технического идентификатора '
        'планера НЕТ. serial_number в payload DJI — это идентификатор ВЫЛЕТА '
        '(проверяется на листе «Сводка»: число различных значений близко к '
        'числу строк). nickname_raw — отображаемое имя с пульта, оператор его '
        'меняет. Единственная связь вылета с машиной — drone_unit_id, а он '
        'получен разбором ника через drone_nicknames и наследует все слабости '
        'ника. hardware_id (Body Code) — настоящий идентификатор планера, но '
        'он лежит на drone_units, а в списке вылетов его нет вовсе. Поэтому '
        'группировка ведётся по drone_unit_id, вылеты без машины вынесены '
        'отдельными строками и НЕ ссыпаны в одну кучу.',
        '',
        'ПОВТОР ПЛОЩАДИ. Кандидат — вылет, у которого площадь ТОЧНО равна '
        'площади одного из %d предыдущих вылетов той же машины, и она не '
        'ноль. Это ANOMALY_CANDIDATE, а не ошибка DJI: два вылета по одному '
        'заданию могут дать одну площадь законно. Признаки, сужающие круг, '
        'считаются отдельно и НЕ складываются в один вердикт.' % DEFAULT_LOOKBACK,
        '',
        'ПОЧЕМУ ОКНО, А НЕ ОДИН ШАГ. Задание формулировало правило как '
        '«совпадает с предыдущим по времени вылетом». На известном случае оно '
        'не срабатывает: 05.06.2026 у машины №8 между двумя вылетами по '
        '5940.0000297 м2 стоит вылет 622715274 на 3293.3333 м2 в ручном '
        'режиме, то есть 622715275 повторяет площадь вылета ЧЕРЕЗ ОДИН. Окно '
        'в один шаг дало бы ноль кандидатов на том самом случае, ради '
        'которого аудит затеян. Расстояние записано в колонке «Вылетов '
        'назад» и разделено в сводке: lag 1 и lag 2 и дальше — разные по '
        'правдоподобию события.',
        '',
        'КОРОТКИЙ ИНТЕРВАЛ. Порог не выдуман: это %d-й процентиль интервалов '
        'между соседними вылетами той же машины, вычисленный по этим же '
        'данным. Значение напечатано на листе «Сводка».' % SHORT_GAP_PERCENTILE,
        '',
        'КОРОТКИЙ ВЫЛЕТ. Длительность меньше %d секунд. Это признак для '
        'разбора, а не порог «допустимого».' % SHORT_FLIGHT_SECONDS,
        '',
        'ШИРИНА ЗАХВАТА. Одиннадцать состояний, потому что они означают '
        'разное и ведут к разным действиям: MISSING_KEY — смена контракта '
        'DJI; JSON_NULL — DJI поле знает, но для этого вылета не записал; '
        'INVALID_TYPE — значение не число (строка, массив, объект, boolean), '
        'то есть сломан формат, а НЕ «не записано»; NON_FINITE — NaN или '
        'бесконечность; MINUS_ONE — наблюдаемый маркер «не записано»; ZERO и '
        'NEGATIVE — непригодны как радиус буфера; COLUMN_NULL — raw_json '
        'непригоден и колонка пуста; POSITIVE_UNVALIDATED — положительное, но '
        'диапазон не задан; OUT_OF_CONFIGURED_RANGE — вне заданных границ; '
        'USABLE — прошло все проверки.',
        '',
        'ПОЧЕМУ USABLE НЕ ВЫДАЁТСЯ ПО УМОЛЧАНИЮ. Физически допустимый '
        'диапазон захвата T40 не подтверждён ни одним источником, которым '
        'располагает проект. Вписать «от 3 до 12 метров» значило бы выдумать '
        'правило. Пока границы не названы владельцем, положительное значение '
        'честно называется POSITIVE_UNVALIDATED. Границы задаются аргументами '
        '--min-width-m и --max-width-m и печатаются в этом отчёте рядом с '
        'числами, чтобы читатель видел, ЧЕМ мерили.',
        '',
        'СТАТИСТИЧЕСКИЙ ВЫБРОС — НЕ ОШИБКА. Выбросы считаются критерием Тьюки '
        '(за пределами Q1−1.5·IQR и Q3+1.5·IQR) и приводятся отдельно от '
        'физически непригодных значений. Редкая ширина может быть законной '
        'настройкой, встречающейся раз в год; различить редкое и неверное '
        'этими данными нельзя, поэтому выброс не переводит вылет в '
        'непригодные.',
        '',
        'ЧИСЛА, КОТОРЫЕ НЕ ЧИСЛА. Значения площади и ширины проверяются '
        'math.isfinite и на тип. NaN, Infinity и -Infinity не попадают ни в '
        'суммы, ни в проценты, ни в сравнение площадей, ни в пригодную '
        'ширину: одно такое значение сделало бы весь итог NaN, а в сравнении '
        'вело бы себя так, что x == x ложно и повтор перестал бы находиться. '
        'Строка без пригодной площади не заменяется нулём — она исключается '
        'из сумм и считается отдельно.',
        '',
        'ЗАПАСНОЙ ИСТОЧНИК. Колонка используется вместо raw_json ТОЛЬКО '
        'когда raw_json непригоден целиком. Если raw_json разобрался, а '
        'значение в нём оказалось строкой или NaN, колонка НЕ подставляется: '
        'это скрыло бы расхождение между источниками. Каждое применение '
        'запасного источника попадает в лист «Качество данных» отдельной '
        'строкой.',
        '',
        'ЧЕГО ЗДЕСЬ НЕТ. Подстановки ширины — ни медианной, ни паспортной, ни '
        'соседней (решение владельца 2026-08-25: DATA_UNAVAILABLE). '
        'Назначенных порогов «допустимого процента» — порог это решение '
        'владельца, а не свойство данных. Исправлений: скрипт только измеряет.',
        '',
        'ОШИБКА В ОДНОЙ СТРОКЕ не прекращает анализ: она попадает в лист '
        '«Качество данных» и считается там.',
        '',
        'КОНТРОЛЬНЫЙ СЛУЧАЙ — ПРЕДУСЛОВИЕ, А НЕ СПРАВКА. Вылет %d обязан '
        'найтись кандидатом на расстоянии %d. Если он не найден, обычный '
        'отчёт не пишется вовсе и код возврата ненулевой. Файл, полученный с '
        'ключом --allow-missing-known-case, помечен на листе «Сводка» и для '
        'решения не годится.' % (KNOWN_CASE_FLIGHT_ID, KNOWN_CASE_EXPECTED_LAG),
    ]
    for line in method:
        ws.append([line])
    for cell in ws['A']:
        cell.alignment = wrap

    book.save(path)
    return path


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='DRONE-COVERAGE-001 stage A2: read-only fleet audit of '
                    'DJI flight data. Never writes to the database.')
    parser.add_argument('--db', required=True,
                        help='path to a COPY of transport.db (never the live '
                             'file)')
    parser.add_argument('--out', default=None,
                        help='path of the xlsx report to write')
    parser.add_argument('--lookback', type=int, default=DEFAULT_LOOKBACK,
                        help='how many flights back to look for an equal '
                             'area (default %d). The known case %d repeats '
                             'the area of the flight TWO back, so 1 would '
                             'miss it.' % (DEFAULT_LOOKBACK,
                                           KNOWN_CASE_FLIGHT_ID))
    parser.add_argument('--min-width-m', type=float,
                        default=DEFAULT_MIN_WIDTH_M,
                        help='lower bound of an acceptable swath, metres. '
                             'Without both bounds no flight is USABLE and a '
                             'positive value is POSITIVE_UNVALIDATED.')
    parser.add_argument('--max-width-m', type=float,
                        default=DEFAULT_MAX_WIDTH_M,
                        help='upper bound of an acceptable swath, metres.')
    parser.add_argument('--allow-missing-known-case', action='store_true',
                        help='DIAGNOSTIC ONLY. Write the report even when the '
                             'known case fails. The xlsx is then stamped '
                             'unusable and the exit code stays non-zero.')
    parser.add_argument('--immutable', action='store_true',
                        help='add immutable=1 to the URI. Correct ONLY for a '
                             'static backup copy, never for a live database')
    args = parser.parse_args(argv)

    if (args.min_width_m is not None and args.max_width_m is not None
            and args.min_width_m > args.max_width_m):
        print('ERROR: --min-width-m is greater than --max-width-m')
        return EXIT_PRECONDITION

    try:
        digest_before = sha256_of(args.db)
    except OSError as exc:
        print('ERROR: cannot read %s (%s)' % (args.db, exc))
        return EXIT_NO_DB

    try:
        con = connect_read_only(args.db, immutable=args.immutable)
    except ProbeError as exc:
        print('ERROR: %s' % exc)
        return EXIT_NO_DB

    try:
        result = analyse(con, lookback=args.lookback,
                         min_width=args.min_width_m,
                         max_width=args.max_width_m)
    except ProbeError as exc:
        print('ERROR: %s' % exc)
        return EXIT_PRECONDITION
    finally:
        con.close()

    known = check_known_case(result)
    digest_after = sha256_of(args.db)
    print_console_summary(result, known, args.db, digest_before)

    if digest_before != digest_after:
        print('FATAL: the database file changed during the run. The report is')
        print('       not trustworthy and no xlsx was written.')
        return EXIT_PRECONDITION
    print('sha256 after : %s  (unchanged)' % digest_after)

    # [REASON]: предусловие проверяется ПОСЛЕ печати сводки и ДО записи файла.
    # Сводку видеть полезно -- по ней и устанавливают причину. А файл, который
    # выглядит как обычный отчёт, но получен на непроверенном прогоне, опаснее
    # отсутствия файла: его откроют через неделю и не вспомнят, что он был
    # забракован.
    if not known['passed']:
        print('')
        print('PRECONDITION FAILED: %s' % _ascii(known['reason']))
        if not args.allow_missing_known_case:
            print('No xlsx was written. Establish the reason, or re-run with')
            print('--allow-missing-known-case to get a report STAMPED as')
            print('unusable for a decision.')
            return EXIT_KNOWN_CASE_FAILED
        print('--allow-missing-known-case given: writing a STAMPED report.')

    if args.out:
        written = write_xlsx(result, known, args.out, args.db,
                             digest_before, digest_after)
        print('report written: %s' % written)
        if not known['passed']:
            print('  the file is stamped unusable for a decision')
    else:
        print('no --out given, xlsx not written')

    return EXIT_OK if known['passed'] else EXIT_KNOWN_CASE_FAILED


if __name__ == '__main__':
    sys.exit(main())
