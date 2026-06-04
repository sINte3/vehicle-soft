# -*- coding: utf-8 -*-
"""
wialon_import.py v4.3
- Fully automatic import (existing mappings applied silently)
- Auto-match page: algorithm pre-fills 200+ vehicles, admin reviews once,
  saves ALL with one button — no individual confirmations
- Topaz agent endpoint: receives fuel data from local agent on 10.103.40.140
"""

import csv
import io
import json
import re
import zipfile
from datetime import datetime, date

from flask import render_template, request, redirect, url_for, flash, jsonify
from flask_login import current_user

from models import db, Equipment, Organization, VialonMapping, VialonImport, EngineHoursRecord, module_required
from workload_report import get_workload_data, generate_workload_excel
from sec003a_ext import log_audit, model_snapshot



def _safe_int(value):
    try:
        return int(value) if value not in (None, '') else None
    except (TypeError, ValueError):
        return None


def _normalize_wialon_name(value):
    return re.sub(r'\s+', ' ', (value or '').strip())


def _equipment_label(eq):
    if not eq:
        return ''
    parts = [eq.name or '']
    if eq.plate:
        parts.append(eq.plate)
    if getattr(eq, 'organization', None):
        parts.append(eq.organization.short_name or eq.organization.name or '')
    return ' / '.join([p for p in parts if p])


def _validate_mapping_decision(eq_id, skip, current_mapping_id=None):
    if skip:
        return None, None
    if not eq_id:
        return None, 'Техникани танланг ёки “Тизимда йўқ” белгисини қўйинг'
    eq = Equipment.query.get(eq_id)
    if not eq:
        return None, 'Техника топилмади'
    if not getattr(eq, 'is_active', True):
        return None, 'Отключённую технику нельзя привязать к Wialon'
    q = VialonMapping.query.filter(
        VialonMapping.equipment_id == eq_id,
        VialonMapping.skip == False
    )
    if current_mapping_id:
        q = q.filter(VialonMapping.id != current_mapping_id)
    existing = q.first()
    if existing:
        return None, 'Бу техника бошқа Wialon объектига бириктирилган: {}'.format(existing.vialon_name)
    return eq_id, None


def _mapping_snapshot(mapping):
    if not mapping:
        return None
    return {
        'id': getattr(mapping, 'id', None),
        'vialon_name': getattr(mapping, 'vialon_name', ''),
        'equipment_id': getattr(mapping, 'equipment_id', None),
        'skip': bool(getattr(mapping, 'skip', False)),
    }


def _audit_wialon(action, entity_type='', entity_id=None, entity_label='', after=None,
                  before=None, changes=None, description=''):
    log_audit(
        db,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_label=entity_label,
        module='wialon',
        before=before,
        after=after,
        changes=changes,
        description=description,
    )


# ─── Plate-based auto-matcher ─────────────────────────────────────────────────

def _normalize(s):
    """Normalize string for plate matching: uppercase, remove spaces/dashes,
    substitute Cyrillic lookalikes with Latin equivalents."""
    s = s.upper().strip()
    s = re.sub(r'[\s\-\.\(\)]', '', s)
    for cy, lat in [('А','A'),('В','B'),('Е','E'),('О','O'),
                    ('Р','P'),('С','C'),('Т','T'),('Х','X'),
                    ('К','K'),('М','M'),('Н','H'),('У','Y')]:
        s = s.replace(cy, lat)
    return s


def auto_match_vehicles(vialon_names, all_equipment):
    """
    For each Wialon name, find the best matching Equipment record.
    Strategy (in order of priority):
      1. Full plate appears in Wialon name           → HIGH
      2. Plate suffix (without '80 ' prefix) appears → HIGH
      3. Exact normalized name match                 → HIGH
      4. Equipment name is substring of Wialon name  → MEDIUM
    Returns list of dicts:
      {vialon_name, engine_hours, equipment, confidence, method}
    """
    results = []

    for row in vialon_names:
        vname = row if isinstance(row, str) else row['vialon_name']
        vn = _normalize(vname)
        best_eq = best_conf = best_method = None

        for eq in all_equipment:
            plate = eq.plate or ''
            eq_name = eq.name or ''

            # 1. Full plate in Wialon name
            if plate:
                pl = _normalize(plate)
                if len(pl) >= 4 and pl in vn:
                    best_eq, best_conf, best_method = eq, 'high', 'plate_full'
                    break

            # 2. Plate suffix (skip leading "80")
            if plate:
                parts = plate.strip().split()
                if parts and parts[0] in ('80', '25'):
                    suffix = _normalize(' '.join(parts[1:]))
                else:
                    suffix = _normalize(plate)
                if len(suffix) >= 4 and suffix in vn:
                    if best_conf != 'high':
                        best_eq, best_conf, best_method = eq, 'high', 'plate_suffix'

            # 3. Exact name
            en = _normalize(eq_name)
            if len(en) >= 4 and en == vn:
                if best_conf != 'high':
                    best_eq, best_conf, best_method = eq, 'high', 'name_exact'

            # 4. Equipment name substring in Wialon name
            if len(en) >= 6 and en in vn:
                if best_conf not in ('high',):
                    best_eq, best_conf, best_method = eq, 'medium', 'name_partial'

        item = dict(row) if isinstance(row, dict) else {'vialon_name': vname}
        item['equipment']  = best_eq
        item['confidence'] = best_conf or 'none'
        item['method']     = best_method or 'none'
        results.append(item)

    return results


# ─── Parser (unchanged) ───────────────────────────────────────────────────────

def _parse_hms(s):
    """Parse Wialon duration string to float hours.

    Handles two formats:
      Standard:  'HH:MM:SS'  or  'H:MM'
      Long:      'N день HH:MM:SS' / 'N дня HH:MM:SS' / 'N дней HH:MM:SS'
      (Wialon uses Russian day-words when the total duration >= 24 hours)
    """
    s = s.strip().strip('"')
    if not s or s == '----':
        return 0.0
    # Handle 'N день/дня/дней HH:MM:SS'
    m = re.match(
        u'(\\d+)\\s+(?:\\u0434\\u0435\\u043d\\u044c|'
        u'\\u0434\\u043d\\u044f|\\u0434\\u043d\\u0435\\u0439)'
        u'\\s+(\\d+):(\\d+):(\\d+)', s)
    if m:
        days = int(m.group(1))
        h    = int(m.group(2))
        mn   = int(m.group(3))
        sec  = int(m.group(4))
        return days * 24 + h + mn / 60.0 + sec / 3600.0
    # Standard HH:MM:SS or H:MM
    parts = s.split(':')
    try:
        if len(parts) == 3:
            h, mn, sec = int(parts[0]), int(parts[1]), int(parts[2])
            return h + mn / 60.0 + sec / 3600.0
        if len(parts) == 2:
            h, mn = int(parts[0]), int(parts[1])
            return h + mn / 60.0
    except ValueError:
        pass
    return 0.0


def _is_top_level(num_str):
    return bool(re.match(r'^\d+$', num_str.strip().strip('"')))


def _extract_moto_csv(file_storage):
    filename = file_storage.filename or ''
    data = file_storage.read()
    if filename.lower().endswith('.zip') or data[:2] == b'PK':
        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
        except zipfile.BadZipFile:
            raise ValueError('Файл ZIP архив эмас.')
        moto_files = [n for n in zf.namelist()
                      if '\u041c\u043e\u0442\u043e\u0447\u0430\u0441\u044b' in n
                      or 'moto' in n.lower()]
        if not moto_files:
            raise ValueError('ZIP архивда "Моточасы" файли топилмади.')
        return zf.read(moto_files[0]), moto_files[0]
    return data, filename


def parse_moto_csv(csv_bytes):
    for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
        try:
            text = csv_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('Файл кодировкасини аниқлаб бўлмади.')

    results = []
    reader = csv.reader(io.StringIO(text), delimiter=';')
    header = next(reader, None)
    if not header:
        raise ValueError('Файл бўш.')

    cols = [h.strip().strip('\ufeff').strip('"').lower() for h in header]

    def col(names):
        for n in names:
            for i, c in enumerate(cols):
                if n in c:
                    return i
        return None

    idx_num  = col(['no', '№', 'num', '#'])
    idx_name = col(['группировка', 'grouping', 'name'])
    idx_moto = col(['моточасы', 'engine hours'])
    idx_move = col(['в движении', 'moving'])
    idx_idle = col(['холостой', 'idle'])

    if idx_num is None or idx_name is None or idx_moto is None:
        raise ValueError('Нотўғри формат: №, Группировка, Моточасы устунлари топилмади.')

    for row in reader:
        if not row or not any(row):
            continue
        num_val = row[idx_num].strip().strip('"') if idx_num < len(row) else ''
        if not _is_top_level(num_val):
            continue
        name   = row[idx_name].strip().strip('"') if idx_name < len(row) else ''
        moto   = _parse_hms(row[idx_moto]) if idx_moto < len(row) else 0.0
        moving = _parse_hms(row[idx_move]) if idx_move and idx_move < len(row) else 0.0
        idle_h = _parse_hms(row[idx_idle]) if idx_idle and idx_idle < len(row) else 0.0
        if not name:
            continue
        results.append({'vialon_name': name,
                        'engine_hours': round(moto),
                        'hours_moving': round(moving),
                        'hours_idle':   round(idle_h)})
    return results


# ─── Apply existing mappings ──────────────────────────────────────────────────

def apply_mappings(parsed_rows):
    all_maps = {m.vialon_name: m for m in VialonMapping.query.all()}
    matched = []
    unknown = []
    skipped = []
    for row in parsed_rows:
        name = row['vialon_name']
        m = all_maps.get(name)
        if m is None:
            unknown.append(row)
        elif m.skip:
            skipped.append(row)
        elif m.equipment_id:
            eq = Equipment.query.get(m.equipment_id)
            if eq:
                matched.append(dict(row, equipment=eq))
            else:
                unknown.append(row)
        else:
            unknown.append(row)
    return matched, unknown, skipped


# ─── Save engine hours ────────────────────────────────────────────────────────

def save_engine_hours(import_log, work_date, items):
    saved = 0
    for item in items:
        eq = item['equipment']
        existing = EngineHoursRecord.query.filter_by(
            work_date=work_date, equipment_id=eq.id
        ).first()
        if existing:
            existing.engine_hours = item['engine_hours']
            existing.hours_moving = item['hours_moving']
            existing.hours_idle   = item['hours_idle']
            existing.vialon_name  = item['vialon_name']
            existing.import_id    = import_log.id
        else:
            db.session.add(EngineHoursRecord(
                work_date=work_date, equipment_id=eq.id,
                import_id=import_log.id,
                engine_hours=item['engine_hours'],
                hours_moving=item['hours_moving'],
                hours_idle=item['hours_idle'],
                vialon_name=item['vialon_name'],
            ))
        saved += 1
    return saved


# ─── Routes ──────────────────────────────────────────────────────────────────

def register_wialon_routes(app, editor_required, admin_required):

    # ── Main page ─────────────────────────────────────────────────────────
    @app.route('/wialon')
    @module_required('wialon')
    @editor_required
    def wialon_index():
        from datetime import date as date_cls, timedelta as td
        imports = (VialonImport.query
                   .order_by(VialonImport.created_at.desc()).limit(30).all())
        all_mapped = {m.vialon_name for m in VialonMapping.query.all()}
        pending_count = 0
        for imp in imports:
            try:
                names = json.loads(imp.unknown_vehicles_json or '[]')
            except (ValueError, TypeError):
                names = []
            pending_count += sum(1 for n in names if n not in all_mapped)
        yesterday = (date_cls.today() - td(days=1)).isoformat()
        return render_template(
            'wialon.html',
            imports=imports,
            pending_count=pending_count,
            yesterday=yesterday,
            mapping_count=VialonMapping.query.filter(
                VialonMapping.equipment_id.isnot(None),
                VialonMapping.skip == False
            ).count(),
        )

    # ── Upload: fully automatic, supports date range ──────────────────────
    @app.route('/wialon/upload', methods=['POST'])
    @module_required('wialon')
    @editor_required
    def wialon_upload():
        from datetime import date as date_cls, timedelta as td
        f = request.files.get('wialon_file')
        if not f or not f.filename:
            flash('Файл танланмади', 'warning')
            return redirect(url_for('wialon_index'))

        # Parse date range based on mode
        mode = request.form.get('wialon_mode', 'day')
        today = date_cls.today()
        if mode == 'week':
            d_from = today - td(days=today.weekday())
            d_to   = d_from + td(days=6)
        elif mode == 'month':
            d_from = today.replace(day=1)
            if today.month == 12:
                d_to = today.replace(year=today.year+1, month=1, day=1) - td(days=1)
            else:
                d_to = today.replace(month=today.month+1, day=1) - td(days=1)
        elif mode == 'range':
            try:
                d_from = datetime.strptime(request.form.get('date_from',''), '%Y-%m-%d').date()
                d_to   = datetime.strptime(request.form.get('date_to',''),   '%Y-%m-%d').date()
                if d_from > d_to:
                    d_from, d_to = d_to, d_from
            except ValueError:
                flash('Нотўғри сана диапазони', 'warning')
                return redirect(url_for('wialon_index'))
        else:  # day
            report_date_str = request.form.get('report_date', '')
            try:
                d_from = datetime.strptime(report_date_str, '%Y-%m-%d').date()
                d_to   = d_from
            except ValueError:
                flash('Нотўғри сана', 'warning')
                return redirect(url_for('wialon_index'))

        try:
            csv_bytes, csv_name = _extract_moto_csv(f)
            parsed = parse_moto_csv(csv_bytes)
        except ValueError as e:
            flash(str(e), 'warning')
            return redirect(url_for('wialon_index'))
        if not parsed:
            flash('Файлда маълумот топилмади', 'warning')
            return redirect(url_for('wialon_index'))

        matched, unknown, skipped = apply_mappings(parsed)
        hours_by_eq = {item['equipment'].id: item for item in matched}

        all_mappings_with_eq = (VialonMapping.query
                                .filter(VialonMapping.equipment_id.isnot(None),
                                        VialonMapping.skip == False).all())
        user_org_ids = current_user.get_org_ids()
        all_eq_to_write = []
        for m in all_mappings_with_eq:
            eq = Equipment.query.get(m.equipment_id)
            if eq and eq.organization_id in user_org_ids:
                all_eq_to_write.append((eq, hours_by_eq.get(eq.id)))

        # Number of days in the range (for dividing period totals)
        num_days = (d_to - d_from).days + 1

        # Write for each day in the range.
        # For multi-day imports the CSV contains PERIOD totals, so we divide
        # by num_days to get the average daily value — this ensures the SUM
        # over any date range in the workload report stays correct.
        total_saved = 0
        current_date = d_from
        while current_date <= d_to:
            import_log = VialonImport(
                import_date=current_date, filename=csv_name,
                vehicles_in_file=len(parsed),
                vehicles_matched=len(matched),
                vehicles_saved=0,
                vehicles_skipped=len(skipped),
                vehicles_unknown=len(unknown),
                unknown_vehicles_json=json.dumps(
                    [r['vialon_name'] for r in unknown], ensure_ascii=False),
                created_by=current_user.id,
            )
            db.session.add(import_log)
            db.session.flush()

            saved = 0
            for eq, item in all_eq_to_write:
                # Divide period total by number of days → daily average
                eng_h = (item['engine_hours'] / num_days) if item else 0.0
                mov_h = (item['hours_moving'] / num_days) if item else 0.0
                idl_h = (item['hours_idle']   / num_days) if item else 0.0
                vname = item['vialon_name'] if item else ''

                existing = EngineHoursRecord.query.filter_by(
                    work_date=current_date, equipment_id=eq.id).first()
                if existing:
                    existing.engine_hours = eng_h
                    existing.hours_moving = mov_h
                    existing.hours_idle   = idl_h
                    existing.vialon_name  = vname
                    existing.import_id    = import_log.id
                else:
                    db.session.add(EngineHoursRecord(
                        work_date=current_date, equipment_id=eq.id,
                        import_id=import_log.id,
                        engine_hours=eng_h, hours_moving=mov_h,
                        hours_idle=idl_h, vialon_name=vname,
                    ))
                saved += 1

            import_log.vehicles_saved = saved
            total_saved = saved
            current_date += td(days=1)

        _audit_wialon(
            'wialon_import_uploaded',
            entity_type='vialon_import',
            entity_label=csv_name,
            after={
                'filename': csv_name,
                'date_from': d_from.isoformat(),
                'date_to': d_to.isoformat(),
                'days': num_days,
                'vehicles_in_file': len(parsed),
                'vehicles_matched': len(matched),
                'vehicles_saved': total_saved,
                'vehicles_skipped': len(skipped),
                'vehicles_unknown': len(unknown),
            },
            description='Wialon engine-hours import completed'
        )
        db.session.commit()

        num_days = (d_to - d_from).days + 1
        if num_days == 1:
            flash('{} та техника учун моточасы сақланди ({} та Виалондан, {} та 0 соат)'.format(
                total_saved, len(matched), total_saved - len(matched)), 'success')
        else:
            flash('{} кун учун {} та техника моточасы сақланди'.format(
                num_days, total_saved), 'success')
        if unknown:
            flash('{} та техника маппингда топилмади.'.format(len(unknown)), 'warning')
        return redirect(url_for('wialon_index'))

    # ── Auto-match page ───────────────────────────────────────────────────
    @app.route('/wialon/auto_match')
    @module_required('wialon')
    @admin_required
    def wialon_auto_match():
        """Show all unknown vehicles with algorithm-suggested matches.
        Admin reviews once, saves ALL with one button."""
        # Collect all unresolved unknowns from all imports
        all_mapped = {m.vialon_name: m for m in VialonMapping.query.all()}
        unknown_rows = []
        seen = set()
        for imp in (VialonImport.query
                    .order_by(VialonImport.created_at.desc()).all()):
            try:
                names = json.loads(imp.unknown_vehicles_json or '[]')
            except (ValueError, TypeError):
                names = []
            for name in names:
                if name not in seen:
                    seen.add(name)
                    if name not in all_mapped:
                        unknown_rows.append({'vialon_name': name,
                                             'engine_hours': 0,
                                             'hours_moving': 0,
                                             'hours_idle':   0})

        all_equipment = (Equipment.query
                         .join(Organization)
                         .filter(Equipment.is_active == True)
                         .order_by(Organization.sort_order,
                                   Equipment.category,
                                   Equipment.name).all())

        # Run auto-match
        suggestions = auto_match_vehicles(unknown_rows, all_equipment)

        # Stats
        high   = sum(1 for s in suggestions if s['confidence'] == 'high')
        medium = sum(1 for s in suggestions if s['confidence'] == 'medium')
        no_match = sum(1 for s in suggestions if s['confidence'] == 'none')

        return render_template(
            'wialon_auto_match.html',
            suggestions=suggestions,
            all_equipment=all_equipment,
            stats={'high': high, 'medium': medium, 'no_match': no_match},
        )

    # ── Bulk save mappings (one button) ───────────────────────────────────
    @app.route('/wialon/auto_match/save', methods=['POST'])
    @module_required('wialon')
    @admin_required
    def wialon_auto_match_save():
        """Save ALL mapping decisions at once. No individual confirmations."""
        vnames = request.form.getlist('vialon_name')
        eq_ids = request.form.getlist('equipment_id')
        skips = set(request.form.getlist('skip_name'))

        decisions = []
        seen_vnames = set()
        seen_eq_ids = {}
        validation_errors = []

        for i, raw_vname in enumerate(vnames):
            vname = _normalize_wialon_name(raw_vname)
            if not vname:
                continue
            vkey = vname.casefold()
            if vkey in seen_vnames:
                validation_errors.append('Wialon номлари такрорланган: {}'.format(vname))
                continue
            seen_vnames.add(vkey)

            skip = vname in skips or raw_vname in skips
            if skip:
                eq_id = None
            else:
                try:
                    eq_id = int(eq_ids[i]) if i < len(eq_ids) and eq_ids[i] else None
                except (TypeError, ValueError):
                    eq_id = None

            if eq_id is None and not skip:
                continue  # user left row empty; ignore it

            existing_mapping = VialonMapping.query.filter_by(vialon_name=vname).first()
            current_id = existing_mapping.id if existing_mapping else None
            eq_id, error = _validate_mapping_decision(eq_id, skip, current_mapping_id=current_id)
            if error:
                validation_errors.append('{}: {}'.format(vname, error))
                continue
            if eq_id:
                if eq_id in seen_eq_ids and seen_eq_ids[eq_id] != vname:
                    validation_errors.append('Бир техника бир нечта Wialon объектига танланган: {} / {}'.format(
                        seen_eq_ids[eq_id], vname))
                    continue
                seen_eq_ids[eq_id] = vname
            decisions.append((vname, eq_id, skip))

        if validation_errors:
            flash(validation_errors[0], 'warning')
            return redirect(url_for('wialon_index'))

        saved_mappings = 0
        for vname, eq_id, skip in decisions:
            m = VialonMapping.query.filter_by(vialon_name=vname).first()
            if m:
                m.equipment_id = eq_id
                m.skip = skip
            else:
                m = VialonMapping(vialon_name=vname, equipment_id=eq_id,
                                  skip=skip, created_by=current_user.id)
                db.session.add(m)
            saved_mappings += 1

        _audit_wialon(
            'wialon_auto_match_saved',
            entity_type='vialon_mapping',
            entity_label='auto_match_bulk',
            after={
                'saved_mappings': saved_mappings,
                'total_rows': len(vnames),
                'skip_count': len(skips),
            },
            description='Wialon auto-match mappings saved'
        )
        db.session.commit()
        flash('{} та маппинг сақланди. '
              'Энди файлни қайта юкланг — импорт автоматик ишлайди.'.format(saved_mappings),
              'success')
        return redirect(url_for('wialon_index'))

    # ── Mapping list (admin, simple edit/delete) ──────────────────────────
    @app.route('/wialon/mapping')
    @module_required('wialon')
    @admin_required
    def wialon_mapping_list():
        mappings = VialonMapping.query.order_by(VialonMapping.vialon_name).all()
        all_equipment = (Equipment.query.join(Organization)
                         .filter(Equipment.is_active == True)
                         .order_by(Organization.sort_order,
                                   Equipment.category, Equipment.name).all())
        return render_template('wialon_mapping_list.html',
                               mappings=mappings,
                               all_equipment=all_equipment)

    @app.route('/wialon/mapping/save', methods=['POST'])
    @module_required('wialon')
    @admin_required
    def wialon_mapping_save():
        mid = request.form.get('id', type=int)
        eq_id = request.form.get('equipment_id', type=int)
        skip = request.form.get('skip') == '1'
        vname = _normalize_wialon_name(request.form.get('vialon_name', ''))
        created = False
        before = None

        if mid:
            m = VialonMapping.query.get_or_404(mid)
            before = _mapping_snapshot(m)
            if not vname:
                vname = m.vialon_name
        else:
            if not vname:
                flash('Виалон номини киритинг', 'warning')
                return redirect(url_for('wialon_mapping_list'))
            m = VialonMapping.query.filter_by(vialon_name=vname).first()
            if not m:
                m = VialonMapping(vialon_name=vname, created_by=current_user.id)
                db.session.add(m)
                created = True
            else:
                before = _mapping_snapshot(m)

        if not vname or len(vname) < 2:
            flash('Wialon номи жуда қисқа', 'warning')
            return redirect(url_for('wialon_mapping_list'))

        duplicate_name = VialonMapping.query.filter(
            VialonMapping.vialon_name == vname,
            VialonMapping.id != getattr(m, 'id', 0)
        ).first()
        if duplicate_name:
            flash('Бундай Wialon номи аллақачон мавжуд', 'warning')
            return redirect(url_for('wialon_mapping_list'))

        eq_id, error = _validate_mapping_decision(eq_id, skip, current_mapping_id=getattr(m, 'id', None))
        if error:
            flash(error, 'warning')
            return redirect(url_for('wialon_mapping_list'))

        m.vialon_name = vname
        m.equipment_id = eq_id if not skip else None
        m.skip = skip

        db.session.flush()
        after = _mapping_snapshot(m)
        _audit_wialon(
            'wialon_mapping_created' if created else 'wialon_mapping_updated',
            entity_type='vialon_mapping',
            entity_id=m.id,
            entity_label=m.vialon_name,
            before=before,
            after=after,
            description='Wialon mapping saved'
        )
        db.session.commit()
        flash('Маппинг сақланди', 'success')
        return redirect(url_for('wialon_mapping_list'))

    @app.route('/wialon/mapping/delete/<int:mid>', methods=['POST'])
    @module_required('wialon')
    @admin_required
    def wialon_mapping_delete(mid):
        m = VialonMapping.query.get_or_404(mid)
        before = _mapping_snapshot(m)
        label = m.vialon_name
        _audit_wialon(
            'wialon_mapping_deleted',
            entity_type='vialon_mapping',
            entity_id=m.id,
            entity_label=label,
            before=before,
            description='Wialon mapping deleted'
        )
        db.session.delete(m)
        db.session.commit()
        flash('Маппинг ўчирилди', 'warning')
        return redirect(url_for('wialon_mapping_list'))

    # ── Engine hours report ────────────────────────────────────────────────
    @app.route('/wialon/report')
    @module_required('wialon')
    @editor_required
    def wialon_report():
        from datetime import date as date_cls, timedelta as td
        today = date_cls.today()

        mode = request.args.get('mode', 'day')
        if mode == 'week':
            d_from = today - td(days=today.weekday())
            d_to   = d_from + td(days=6)
        elif mode == 'month':
            d_from = today.replace(day=1)
            if today.month == 12:
                d_to = today.replace(year=today.year+1, month=1, day=1) - td(days=1)
            else:
                d_to = today.replace(month=today.month+1, day=1) - td(days=1)
        elif mode == 'range':
            try:
                d_from = datetime.strptime(request.args.get('date_from',''), '%Y-%m-%d').date()
                d_to   = datetime.strptime(request.args.get('date_to',''),   '%Y-%m-%d').date()
                if d_from > d_to: d_from, d_to = d_to, d_from
            except ValueError:
                d_from = d_to = today
        else:
            mode = 'day'
            date_str = request.args.get('date', (today - td(days=1)).isoformat())
            try:
                d_from = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                d_from = today - td(days=1)
            d_to = d_from

        org_id = request.args.get('org_id', type=int)
        user_org_ids = current_user.get_org_ids()

        # Sum engine hours per equipment over the date range
        q = (EngineHoursRecord.query.join(Equipment).join(Organization)
             .filter(EngineHoursRecord.work_date >= d_from,
                     EngineHoursRecord.work_date <= d_to)
             .filter(Organization.id.in_(user_org_ids))
             .order_by(Organization.sort_order, Equipment.category, Equipment.name))
        if org_id:
            q = q.filter(Equipment.organization_id == org_id)
        records = q.all()

        # Group by org → equipment, summing hours
        from collections import defaultdict
        orgs_map = {}
        eq_hours = defaultdict(float)
        for r in records:
            eq_hours[r.equipment_id] += r.engine_hours
            org = r.equipment.organization
            if org.id not in orgs_map:
                orgs_map[org.id] = {'org': org, 'records': [], 'total_hours': 0.0}

        # Build one record per equipment with summed hours
        eq_seen = {}
        for r in records:
            eid = r.equipment_id
            if eid not in eq_seen:
                eq_seen[eid] = r
                # Attach summed hours
                r._summed_hours = round(eq_hours[eid])
                orgs_map[r.equipment.organization.id]['records'].append(r)
                orgs_map[r.equipment.organization.id]['total_hours'] += eq_hours[eid]

        for od in orgs_map.values():
            od['total_hours'] = round(od['total_hours'])

        orgs_data = sorted(orgs_map.values(), key=lambda x: x['org'].sort_order)
        organizations = (Organization.query.filter(Organization.id.in_(user_org_ids))
                         .order_by(Organization.sort_order).all())

        return render_template('wialon_report.html',
                               mode=mode, d_from=d_from, d_to=d_to,
                               orgs_data=orgs_data,
                               organizations=organizations,
                               selected_org_id=org_id)

    # ── Engine hours Excel export ──────────────────────────────────────────
    @app.route('/wialon/report/export')
    @module_required('wialon')
    @editor_required
    def wialon_report_export():
        from datetime import date as date_cls
        import os, io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.worksheet.properties import PageSetupProperties

        date_str = request.args.get('date', date_cls.today().isoformat())
        try:
            sel_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            sel_date = date_cls.today()
        org_id = request.args.get('org_id', type=int)
        user_org_ids = current_user.get_org_ids()

        q = (EngineHoursRecord.query.join(Equipment).join(Organization)
             .filter(EngineHoursRecord.work_date == sel_date)
             .filter(Organization.id.in_(user_org_ids))
             .order_by(Organization.sort_order, Equipment.category, Equipment.name))
        if org_id:
            q = q.filter(Equipment.organization_id == org_id)
        records = q.all()

        orgs_map = {}
        for r in records:
            org = r.equipment.organization
            if org.id not in orgs_map:
                orgs_map[org.id] = {'org': org, 'records': [], 'total_hours': 0.0}
            orgs_map[org.id]['records'].append(r)
            orgs_map[org.id]['total_hours'] += r.engine_hours
        orgs_data = sorted(orgs_map.values(), key=lambda x: x['org'].sort_order)

        # ── Styles ──
        TNR = 'Times New Roman'
        THIN = Side(style='thin', color='000000')
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        FILL_YELLOW = PatternFill('solid', start_color='FFFF00', end_color='FFFF00')
        FILL_GREEN  = PatternFill('solid', start_color='E2F0D9', end_color='E2F0D9')
        FILL_HEADER = PatternFill('solid', start_color='1F3864', end_color='1F3864')
        FILL_ORG    = PatternFill('solid', start_color='2E4057', end_color='2E4057')
        FILL_GRAY   = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')
        AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
        AL = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        AR = Alignment(horizontal='right',  vertical='center', wrap_text=True)

        def sc(cell, bold=False, size=12, fill=None, align=AC, white=False, num_fmt=None):
            cell.font = Font(name=TNR, size=size, bold=bold,
                             color='FFFFFF' if white else '000000')
            if fill: cell.fill = fill
            cell.alignment = align
            cell.border = BORDER
            if num_fmt: cell.number_format = num_fmt

        wb = Workbook()
        ws = wb.active
        ws.title = sel_date.strftime('%d.%m.%Y')

        # Page setup: portrait A4, fit to 1 page wide
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.7,
                                       header=0.3, footer=0.3)

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14

        # Title row
        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = (u'"Бухоро Агрокластер" МЧЖ тизимидаги техникаларнинг\n'
                   u'МОТОСОАТЛАР БЎЙИЧА МАЪЛУМОТИ')
        c.font = Font(name=TNR, size=16, bold=True)
        c.alignment = AC
        c.border = BORDER
        ws.row_dimensions[1].height = 55

        # Date row
        ws.merge_cells('A2:E2')
        c = ws['A2']
        c.value = u'Сана: {}'.format(sel_date.strftime('%d.%m.%Y'))
        sc(c, bold=True, size=13, fill=FILL_YELLOW)
        ws.row_dimensions[2].height = 28

        # Header row
        ws.row_dimensions[3].height = 40
        headers = [u'№', u'Техника (Давлат рақами)', u'Мотосоат', u'Меъёр (8 соат)', u'Юклама %']
        fills_h  = [FILL_YELLOW]*5
        for i, (h, fh) in enumerate(zip(headers, fills_h), 1):
            c = ws.cell(3, i, h)
            sc(c, bold=True, size=12, fill=fh)

        # Data
        row = 4
        norm = 8.0
        grand_total = 0.0

        for od in orgs_data:
            # Organization header
            ws.merge_cells('A{}:E{}'.format(row, row))
            c = ws.cell(row, 1)
            c.value = u'{}    —    Жами: {:.2f} соат'.format(
                od['org'].name, od['total_hours'])
            c.font = Font(name=TNR, size=13, bold=True, color='FFFFFF')
            c.fill = FILL_ORG
            c.alignment = AL
            c.border = BORDER
            ws.row_dimensions[row].height = 30
            row += 1

            num = 1
            for r in od['records']:
                eq = r.equipment
                # Combined техника + номер
                tech_name = eq.name
                if eq.plate:
                    tech_name = u'{} / {}'.format(eq.name, eq.plate)

                pct = r.engine_hours / norm if norm else 0
                dev = r.engine_hours - norm

                ws.row_dimensions[row].height = 26
                ws.cell(row, 1, num)
                ws.cell(row, 2, tech_name)
                ws.cell(row, 3, round(r.engine_hours, 2))
                ws.cell(row, 4, round(dev, 2))
                ws.cell(row, 5, round(pct, 4))

                row_fill = None
                if pct >= 1.0:
                    row_fill = PatternFill('solid', start_color='E2F0D9', end_color='E2F0D9')
                elif pct >= 0.7:
                    row_fill = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')

                sc(ws.cell(row, 1), size=12, fill=row_fill)
                sc(ws.cell(row, 2), size=12, fill=row_fill, align=AL)
                sc(ws.cell(row, 3), size=13, bold=True, fill=row_fill, num_fmt='0.00')
                sc(ws.cell(row, 4), size=12, fill=row_fill, num_fmt='+0.0;-0.0')
                sc(ws.cell(row, 5), size=12, fill=row_fill, num_fmt='0%')

                grand_total += r.engine_hours
                num += 1
                row += 1

        # Grand total
        ws.merge_cells('A{}:B{}'.format(row, row))
        ws.cell(row, 1, u'ЖАМИ МОТОСОАТ:')
        sc(ws.cell(row, 1), bold=True, size=14, fill=FILL_GREEN, align=AL)
        sc(ws.cell(row, 2), bold=True, size=14, fill=FILL_GREEN)
        ws.cell(row, 3, round(grand_total, 2))
        sc(ws.cell(row, 3), bold=True, size=16, fill=FILL_YELLOW, num_fmt='0.00')
        for col in [4, 5]:
            sc(ws.cell(row, col), fill=FILL_GREEN)
        ws.row_dimensions[row].height = 36

        ws.print_title_rows = '3:3'
        ws.print_area = 'A1:E{}'.format(row)

        # Save to buffer
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file as _sf
        fname = u'Motochaslar_{}.xlsx'.format(sel_date.strftime('%d_%m_%Y'))
        _audit_wialon(
            'wialon_engine_hours_exported',
            entity_type='wialon_report',
            entity_label=fname,
            after={
                'date': sel_date.isoformat(),
                'org_id': org_id,
                'records_count': len(records),
                'filename': fname,
            },
            description='Wialon engine-hours report exported'
        )
        db.session.commit()
        return _sf(buf, as_attachment=True,
                   download_name=fname,
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # ── Workload report (Yuklama Hisobot) ─────────────────────────────────
    @app.route('/wialon/workload')
    @module_required('wialon')
    @editor_required
    def wialon_workload():
        from datetime import date as date_cls, timedelta as td
        def _pd(s):
            if s:
                try:
                    return datetime.strptime(s, '%Y-%m-%d').date()
                except ValueError:
                    pass
            return date_cls.today()
        mode  = request.args.get('mode', 'month')
        today = date_cls.today()
        if mode == 'week':
            d_from = today - td(days=today.weekday())
            d_to   = d_from + td(days=6)
        elif mode == 'month':
            d_from = today.replace(day=1)
            if today.month == 12:
                d_to = today.replace(year=today.year + 1, month=1, day=1) - td(days=1)
            else:
                d_to = today.replace(month=today.month + 1, day=1) - td(days=1)
        elif mode == 'range':
            d_from = _pd(request.args.get('date_from'))
            d_to   = _pd(request.args.get('date_to'))
            if d_from > d_to:
                d_from, d_to = d_to, d_from
        else:
            mode   = 'day'
            d_from = _pd(request.args.get('date'))
            d_to   = d_from
        selected_org_id = request.args.get('org_id', type=int)
        if current_user.is_admin:
            user_orgs      = Organization.query.order_by(Organization.sort_order).all()
            filter_org_ids = None
        else:
            user_orgs      = sorted(current_user.organizations, key=lambda o: o.sort_order)
            filter_org_ids = [o.id for o in user_orgs]
        if selected_org_id:
            if not current_user.can_access_org(selected_org_id):
                from flask import abort
                abort(403)
            filter_org_ids = [selected_org_id]
        orgs_data, calendar_days, norm_hours = get_workload_data(d_from, d_to, filter_org_ids)
        all_rows  = [r for od in orgs_data for r in od['rows']]
        total_eq  = len(all_rows)
        zero_eq   = sum(1 for r in all_rows if r['fact'] == 0)
        avg_pct_v = (sum(r['exec_pct'] for r in all_rows) / total_eq * 100) if total_eq else 0
        avg_pct   = '{:.1f}%'.format(avg_pct_v)
        return render_template('workload.html',
            mode=mode, d_from=d_from, d_to=d_to,
            calendar_days=calendar_days, norm_hours=norm_hours,
            orgs_data=orgs_data,
            user_orgs=user_orgs, selected_org_id=selected_org_id,
            total_eq=total_eq, zero_eq=zero_eq, avg_pct=avg_pct)

    @app.route('/wialon/workload/export')
    @module_required('wialon')
    @editor_required
    def wialon_workload_export():
        import os
        from datetime import date as date_cls, timedelta as td
        from flask import send_file, current_app
        def _pd(s):
            if s:
                try:
                    return datetime.strptime(s, '%Y-%m-%d').date()
                except ValueError:
                    pass
            return date_cls.today()
        mode  = request.args.get('mode', 'month')
        today = date_cls.today()
        if mode == 'week':
            d_from = today - td(days=today.weekday())
            d_to   = d_from + td(days=6)
        elif mode == 'month':
            d_from = today.replace(day=1)
            if today.month == 12:
                d_to = today.replace(year=today.year + 1, month=1, day=1) - td(days=1)
            else:
                d_to = today.replace(month=today.month + 1, day=1) - td(days=1)
        elif mode == 'range':
            d_from = _pd(request.args.get('date_from'))
            d_to   = _pd(request.args.get('date_to'))
            if d_from > d_to:
                d_from, d_to = d_to, d_from
        else:
            d_from = _pd(request.args.get('date'))
            d_to   = d_from
        selected_org_id = request.args.get('org_id', type=int)
        if current_user.is_admin:
            filter_org_ids = None
        else:
            filter_org_ids = [o.id for o in current_user.organizations]
        if selected_org_id:
            if not current_user.can_access_org(selected_org_id):
                from flask import abort
                abort(403)
            filter_org_ids = [selected_org_id]
        output_dir = current_app.config.get('REPORTS_DIR', 'reports')
        fpath = generate_workload_excel(d_from, d_to, output_dir, filter_org_ids)
        _audit_wialon(
            'wialon_workload_exported',
            entity_type='wialon_workload',
            entity_label=os.path.basename(fpath),
            after={
                'date_from': d_from.isoformat(),
                'date_to': d_to.isoformat(),
                'org_id': selected_org_id,
                'filename': os.path.basename(fpath),
            },
            description='Wialon workload report exported'
        )
        db.session.commit()
        return send_file(fpath, as_attachment=True, download_name=os.path.basename(fpath))